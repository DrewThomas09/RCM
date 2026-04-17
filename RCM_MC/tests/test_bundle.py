"""Tests for the diligence deliverable bundle."""
from __future__ import annotations

import os
import tempfile
import unittest

import pandas as pd
import yaml

from rcm_mc.infra._bundle import (
    build_data_requests,
    finalize_bundle,
    organize_detail,
    write_data_requests,
    write_diligence_workbook,
)


BASE_DIR = os.path.dirname(os.path.dirname(__file__))
ACTUAL_PATH = os.path.join(BASE_DIR, "configs", "actual.yaml")


def _load_cfg() -> dict:
    with open(ACTUAL_PATH) as f:
        return yaml.safe_load(f)


def _mock_summary() -> pd.DataFrame:
    return pd.DataFrame(
        {"mean": [5e6, 2e6], "median": [4.9e6, 1.9e6], "p10": [3e6, 1e6], "p90": [7e6, 3e6]},
        index=["ebitda_drag", "economic_drag"],
    )


class TestWorkbook(unittest.TestCase):
    def test_writes_xlsx_with_core_tabs(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            path = write_diligence_workbook(tmp, _mock_summary(), cfg)
            self.assertTrue(os.path.exists(path))
            from openpyxl import load_workbook
            wb = load_workbook(path)
            for sheet in ("Summary", "Payers", "Assumptions"):
                self.assertIn(sheet, wb.sheetnames)
            wb.close()

    def test_value_drivers_tab_included_when_sensitivity_present(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            pd.DataFrame({"driver": ["idr", "fwr"], "corr": [0.5, 0.3]}).to_csv(
                os.path.join(tmp, "sensitivity.csv"), index=False
            )
            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Value Drivers", wb.sheetnames)
            wb.close()

    def test_peer_tabs_included_when_present(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            # Peer comparison CSVs as they'd be written by cli.py
            pd.DataFrame({
                "ccn": ["111111", "222222"],
                "name": ["Peer A", "Peer B"],
                "state": ["OH", "OH"],
                "beds": [900, 1100],
                "net_patient_revenue": [2.5e9, 3.1e9],
                "medicare_day_pct": [0.22, 0.19],
                "similarity_score": [0.12, 0.18],
            }).to_csv(os.path.join(tmp, "peer_comparison.csv"), index=False)
            pd.DataFrame({
                "kpi": ["net_patient_revenue", "beds", "medicare_day_pct"],
                "target": [6.4e9, 1326, 0.225],
                "peer_p10": [2.0e9, 500, 0.15],
                "peer_median": [2.8e9, 800, 0.22],
                "peer_p90": [4.0e9, 1200, 0.30],
                "target_percentile": [95.0, 93.0, 52.0],
            }).to_csv(os.path.join(tmp, "peer_target_percentiles.csv"), index=False)
            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Peer Percentiles", wb.sheetnames)
            self.assertIn("Peer Set", wb.sheetnames)
            wb.close()

    def test_pe_tabs_included_when_artifacts_present(self):
        """Brick 47: PE Bridge / Returns / Hold Grid / Covenant land in workbook."""
        import json as _json
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            # Seed the four PE artifacts as pe_integration.py would
            with open(os.path.join(tmp, "pe_bridge.json"), "w") as f:
                _json.dump({
                    "entry_ev": 450e6, "exit_ev": 659e6,
                    "total_value_created": 209e6,
                    "components": [
                        {"step": "Entry EV", "value": 450e6,
                         "share_of_creation": None, "note": "9.0x × 50M"},
                        {"step": "Organic EBITDA", "value": 71.7e6,
                         "share_of_creation": 0.342, "note": "3%/yr × 5y"},
                        {"step": "RCM uplift", "value": 72e6,
                         "share_of_creation": 0.344, "note": "× 9x entry"},
                        {"step": "Multiple expansion", "value": 66e6,
                         "share_of_creation": 0.314, "note": "Δ+1x"},
                        {"step": "Exit EV", "value": 659e6,
                         "share_of_creation": None, "note": "10.0x × 66M"},
                    ],
                }, f)
            with open(os.path.join(tmp, "pe_returns.json"), "w") as f:
                _json.dump({
                    "entry_equity": 180e6, "exit_proceeds": 459e6,
                    "hold_years": 5.0, "moic": 2.55, "irr": 0.206,
                    "total_distributions": 459e6,
                }, f)
            pd.DataFrame([
                {"hold_years": 5, "exit_multiple": 9.0, "moic": 2.08, "irr": 0.16,
                 "entry_ev": 450e6, "exit_ev": 594e6, "entry_equity": 180e6,
                 "exit_debt": 270e6, "exit_equity": 324e6, "underwater": False,
                 "total_value_created": 144e6, "rcm_uplift_share": 0.5},
            ]).to_csv(os.path.join(tmp, "pe_hold_grid.csv"), index=False)
            with open(os.path.join(tmp, "pe_covenant.json"), "w") as f:
                _json.dump({
                    "ebitda": 50e6, "debt": 270e6,
                    "covenant_max_leverage": 6.5, "actual_leverage": 5.4,
                    "covenant_headroom_turns": 1.1, "ebitda_cushion_pct": 0.17,
                    "covenant_trips_at_ebitda": 41.5e6, "interest_coverage": 2.3,
                }, f)

            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            for name in ("PE Bridge", "PE Returns", "PE Hold Grid", "PE Covenant"):
                self.assertIn(name, wb.sheetnames, f"Missing: {name}")
            # Bridge tab must have 5 component rows (entry + 3 + exit).
            # UI-5 added a note row, so max_row is now 7 (note + header + 5 data).
            ws = wb["PE Bridge"]
            self.assertIn(ws.max_row, (6, 7))
            wb.close()

    def test_trend_signals_tab_included_when_csv_present(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            pd.DataFrame({
                "metric": ["net_patient_revenue", "medicare_day_pct"],
                "start_year": [2020, 2020],
                "end_year": [2022, 2022],
                "start_value": [5.24e9, 0.270],
                "end_value": [6.38e9, 0.225],
                "pct_change": [0.217, None],
                "pts_change": [None, -4.5],
                "direction": ["up", "down"],
            }).to_csv(os.path.join(tmp, "trend_signals.csv"), index=False)
            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Trend Signals", wb.sheetnames)
            ws = wb["Trend Signals"]
            # UI-5 may insert a note on row 1 — try row 1 then row 2
            headers = [c.value for c in ws[1]]
            if not any(h in headers for h in ("metric", "bucket", "kpi")):
                headers = [c.value for c in ws[2]]
            self.assertIn("metric", headers)
            self.assertIn("pct_change", headers)
            wb.close()

    def test_pressure_test_tabs_included_when_present(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            pd.DataFrame({
                "target": ["idr_blended"], "classification": ["stretch"],
                "actual_blended": [0.135], "benchmark_blended": [0.113],
                "target_value": [0.12], "progress_ratio": [0.68],
                "matching_initiatives": ["prior_auth_improvement"],
                "median_ramp_months": [6.0],
            }).to_csv(os.path.join(tmp, "pressure_test_assessments.csv"), index=False)
            pd.DataFrame({
                "achievement": [0.0, 1.0],
                "ebitda_drag_mean": [6e6, 4e6],
                "ebitda_drag_p10": [4e6, 3e6],
                "ebitda_drag_p90": [8e6, 5e6],
            }).to_csv(os.path.join(tmp, "pressure_test_miss_scenarios.csv"), index=False)
            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Plan Pressure Test", wb.sheetnames)
            self.assertIn("Plan Miss Scenarios", wb.sheetnames)
            wb.close()

    def test_value_drivers_prefers_attribution_over_sensitivity(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            pd.DataFrame({"driver": ["x"], "corr": [0.1]}).to_csv(
                os.path.join(tmp, "sensitivity.csv"), index=False
            )
            pd.DataFrame({"bucket": ["Commercial Denials"], "uplift_oat": [1e6]}).to_csv(
                os.path.join(tmp, "attribution_oat.csv"), index=False
            )
            write_diligence_workbook(tmp, _mock_summary(), cfg)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            ws = wb["Value Drivers"]
            # UI-5 may insert a note on row 1 — try row 1 then row 2
            headers = [c.value for c in ws[1]]
            if not any(h in headers for h in ("metric", "bucket", "kpi")):
                headers = [c.value for c in ws[2]]
            self.assertIn("bucket", headers)  # attribution column, not sensitivity column
            wb.close()


class TestDataRequests(unittest.TestCase):
    def test_all_assumed_yields_priority_asks(self):
        cfg = _load_cfg()
        # actual.yaml ships with _default: assumed → every path assumed
        md = build_data_requests(cfg)
        self.assertIn("Priority asks", md)
        # Per-payer section headers present
        self.assertIn("### Medicare", md)

    def test_all_prior_yields_no_priority_asks(self):
        cfg = _load_cfg()
        cfg["_source_map"] = {"_default": "prior"}
        md = build_data_requests(cfg)
        self.assertNotIn("Priority asks", md)
        self.assertIn("industry priors", md)

    def test_observed_entry_lands_in_already_observed_section(self):
        cfg = _load_cfg()
        cfg["_source_map"] = {"_default": "assumed", "payers.Medicare.denials.idr": "observed"}
        md = build_data_requests(cfg)
        self.assertIn("Already observed", md)

    def test_write_creates_file(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            path = write_data_requests(tmp, cfg, hospital_name="Acme Hospital")
            self.assertTrue(os.path.exists(path))
            with open(path) as f:
                content = f.read()
            self.assertIn("Acme Hospital", content)


class TestOrganizeDetail(unittest.TestCase):
    def test_moves_only_non_top_level_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            # Files that should STAY at top
            for name in ("summary.csv", "report.html", "provenance.json"):
                with open(os.path.join(tmp, name), "w") as f:
                    f.write("x")
            # Files that should MOVE
            for name in ("drivers_stage_drag_mean.csv", "waterfall.png", "sensitivity.csv"):
                with open(os.path.join(tmp, name), "w") as f:
                    f.write("x")
            moved = organize_detail(tmp)
            self.assertEqual(len(moved), 3)
            for name in ("summary.csv", "report.html", "provenance.json"):
                self.assertTrue(os.path.exists(os.path.join(tmp, name)))
            for name in ("drivers_stage_drag_mean.csv", "waterfall.png", "sensitivity.csv"):
                self.assertFalse(os.path.exists(os.path.join(tmp, name)))
                self.assertTrue(os.path.exists(os.path.join(tmp, "_detail", name)))

    def test_idempotent(self):
        with tempfile.TemporaryDirectory() as tmp:
            with open(os.path.join(tmp, "x.csv"), "w") as f:
                f.write("x")
            organize_detail(tmp)
            # Second call: nothing to move
            moved = organize_detail(tmp)
            self.assertEqual(moved, [])

    def test_does_not_touch_subdirectories(self):
        with tempfile.TemporaryDirectory() as tmp:
            sub = os.path.join(tmp, "some_subdir")
            os.makedirs(sub)
            with open(os.path.join(sub, "x.csv"), "w") as f:
                f.write("x")
            organize_detail(tmp)
            # The subdir and its contents are untouched.
            self.assertTrue(os.path.isdir(sub))
            self.assertTrue(os.path.exists(os.path.join(sub, "x.csv")))


class TestFinalizeBundle(unittest.TestCase):
    def test_end_to_end(self):
        cfg = _load_cfg()
        with tempfile.TemporaryDirectory() as tmp:
            # Seed with a typical run's outputs
            _mock_summary().to_csv(os.path.join(tmp, "summary.csv"))
            pd.DataFrame({"iteration": [0, 1], "ebitda_drag": [5e6, 5.1e6]}).to_csv(
                os.path.join(tmp, "simulations.csv"), index=False
            )
            pd.DataFrame({"driver": ["idr"], "corr": [0.4]}).to_csv(
                os.path.join(tmp, "sensitivity.csv"), index=False
            )
            with open(os.path.join(tmp, "waterfall.png"), "wb") as f:
                f.write(b"\x89PNG\r\n\x1a\n")  # minimal header

            result = finalize_bundle(tmp, _mock_summary(), cfg, hospital_name="Test Target")
            self.assertTrue(os.path.exists(result["workbook"]))
            self.assertTrue(os.path.exists(result["data_requests"]))
            # summary.csv / simulations.csv stayed
            self.assertTrue(os.path.exists(os.path.join(tmp, "summary.csv")))
            self.assertTrue(os.path.exists(os.path.join(tmp, "simulations.csv")))
            # sensitivity.csv + waterfall.png moved
            self.assertTrue(os.path.exists(os.path.join(tmp, "_detail", "sensitivity.csv")))
            self.assertTrue(os.path.exists(os.path.join(tmp, "_detail", "waterfall.png")))
