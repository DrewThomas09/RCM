"""Tests for the workbook-polish helpers.

Validates: frozen panes, number formats, source-tag coloring, percentile
highlighting, cover sheet creation, and end-to-end presence of the cover
sheet as the first tab of the final diligence_workbook.xlsx.
"""
from __future__ import annotations

import os
import tempfile
import unittest

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook

from rcm_mc.ui._workbook_style import (
    NUMBER_FORMATS,
    add_tab_note,
    apply_color_scale,
    apply_data_bar,
    apply_peer_percentile_row_formats,
    apply_percentile_coloring,
    apply_sheet_polish,
    apply_source_coloring,
    build_cover_sheet,
)
from rcm_mc.infra._bundle import write_diligence_workbook


BASE_DIR = os.path.dirname(os.path.dirname(__file__))


def _mock_summary() -> pd.DataFrame:
    return pd.DataFrame(
        {"mean": [5e6, 2e6], "median": [4.9e6, 1.9e6], "p10": [3e6, 1e6], "p90": [7e6, 3e6]},
        index=["ebitda_drag", "economic_drag"],
    )


def _load_cfg() -> dict:
    with open(os.path.join(BASE_DIR, "configs", "actual.yaml")) as f:
        return yaml.safe_load(f)


class TestApplySheetPolish(unittest.TestCase):
    """Per-sheet polish: header styling, freeze, number formats."""

    def _build_sheet(self, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        return ws

    def test_freezes_header_row(self):
        ws = self._build_sheet(["ccn", "name"], [["360180", "CCF"]])
        apply_sheet_polish(ws)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_styles_header_cells(self):
        ws = self._build_sheet(["ccn", "name", "beds"], [["360180", "CCF", 1326]])
        apply_sheet_polish(ws)
        for cell in ws[1]:
            self.assertTrue(cell.font.bold)
            # White foreground hex
            self.assertEqual(str(cell.font.color.rgb).upper()[-6:], "FFFFFF")

    def test_applies_currency_format_to_money_columns(self):
        ws = self._build_sheet(
            ["ccn", "net_patient_revenue", "operating_expenses"],
            [["360180", 6378833101.0, 7507484998.0]],
        )
        apply_sheet_polish(ws)
        self.assertIn("$", ws["B2"].number_format)
        self.assertIn("$", ws["C2"].number_format)

    def test_applies_integer_format_to_bed_counts(self):
        ws = self._build_sheet(["ccn", "beds"], [["360180", 1326]])
        apply_sheet_polish(ws)
        self.assertEqual(ws["B2"].number_format, "#,##0")

    def test_applies_percent_format_to_payer_mix_columns(self):
        ws = self._build_sheet(
            ["ccn", "medicare_day_pct", "medicaid_day_pct"],
            [["360180", 0.225, 0.052]],
        )
        apply_sheet_polish(ws)
        self.assertIn("%", ws["B2"].number_format)
        self.assertIn("%", ws["C2"].number_format)

    def test_unknown_column_gets_no_format(self):
        """Headers we don't recognize should be left alone (default 'General')."""
        ws = self._build_sheet(["some_random_col"], [[42]])
        apply_sheet_polish(ws)
        # openpyxl's default format is 'General'
        self.assertEqual(ws["A2"].number_format, "General")

    def test_net_income_format_handles_negatives(self):
        """Net income uses the red-negative currency format."""
        self.assertIn("Red", NUMBER_FORMATS["net_income"])

    def test_empty_sheet_is_no_op(self):
        wb = Workbook()
        ws = wb.active
        # No crash on empty sheet
        apply_sheet_polish(ws)


class TestSourceColoring(unittest.TestCase):
    def _build_assumptions_sheet(self, entries):
        """Rows: (path, source, note)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["path", "source", "note"])
        for e in entries:
            ws.append(e)
        return wb, ws

    def test_observed_cell_gets_green_fill(self):
        wb, ws = self._build_assumptions_sheet([("hospital.annual_revenue", "observed", "HCRIS")])
        colored = apply_source_coloring(ws)
        self.assertEqual(colored, 1)
        # Green = C6EFCE
        fill = ws["B2"].fill
        self.assertEqual(str(fill.fgColor.rgb).upper()[-6:], "C6EFCE")

    def test_prior_cell_gets_amber_fill(self):
        wb, ws = self._build_assumptions_sheet([("x.y", "prior", "")])
        apply_source_coloring(ws)
        self.assertEqual(str(ws["B2"].fill.fgColor.rgb).upper()[-6:], "FFEB9C")

    def test_assumed_cell_gets_red_fill(self):
        wb, ws = self._build_assumptions_sheet([("x.y", "assumed", "")])
        apply_source_coloring(ws)
        self.assertEqual(str(ws["B2"].fill.fgColor.rgb).upper()[-6:], "FFC7CE")

    def test_unknown_source_not_colored(self):
        wb, ws = self._build_assumptions_sheet([("x.y", "something_else", "")])
        colored = apply_source_coloring(ws)
        self.assertEqual(colored, 0)

    def test_sheet_without_source_column_is_noop(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["ccn", "beds"])
        ws.append(["360180", 1326])
        # No 'source' column → no cells colored, no crash
        self.assertEqual(apply_source_coloring(ws), 0)


class TestPercentileColoring(unittest.TestCase):
    def _build_sheet(self, values):
        wb = Workbook()
        ws = wb.active
        ws.append(["kpi", "target_percentile"])
        for k, v in values:
            ws.append([k, v])
        return wb, ws

    def test_high_percentile_gets_green(self):
        wb, ws = self._build_sheet([("NPSR", 92.0)])
        apply_percentile_coloring(ws)
        self.assertEqual(str(ws["B2"].fill.fgColor.rgb).upper()[-6:], "C6EFCE")

    def test_low_percentile_gets_amber(self):
        wb, ws = self._build_sheet([("net_income", 12.0)])
        apply_percentile_coloring(ws)
        self.assertEqual(str(ws["B2"].fill.fgColor.rgb).upper()[-6:], "FFC7CE")

    def test_middle_percentile_uncolored(self):
        wb, ws = self._build_sheet([("Medicare %", 50.0)])
        colored = apply_percentile_coloring(ws)
        self.assertEqual(colored, 0)

    def test_respects_custom_thresholds(self):
        wb, ws = self._build_sheet([("KPI", 60.0)])
        # Default thresholds: 60 isn't extreme. Relax cut and it should color.
        apply_percentile_coloring(ws, high_cut=50.0, low_cut=30.0)
        self.assertEqual(str(ws["B2"].fill.fgColor.rgb).upper()[-6:], "C6EFCE")

    def test_non_numeric_cell_skipped(self):
        wb, ws = self._build_sheet([("KPI", "n/a")])
        colored = apply_percentile_coloring(ws)
        self.assertEqual(colored, 0)


class TestPeerPercentileRowFormats(unittest.TestCase):
    """Per-row number formats on the Peer Percentiles tab — derived KPIs need
    different formats (pct, money, int) in the same `target` column."""

    def _build_sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["kpi", "target", "peer_p10", "peer_median", "peer_p90", "target_percentile"])
        for r in rows:
            ws.append(r)
        return wb, ws

    def test_operating_margin_gets_percent_format(self):
        _, ws = self._build_sheet([["operating_margin", -0.177, -0.21, 0.008, 0.073, 20.0]])
        apply_peer_percentile_row_formats(ws)
        self.assertEqual(ws["B2"].number_format, "0.0%")
        self.assertEqual(ws["C2"].number_format, "0.0%")

    def test_cost_per_patient_day_gets_money_format(self):
        _, ws = self._build_sheet([["cost_per_patient_day", 20437, 5987, 7001, 12345, 100.0]])
        apply_peer_percentile_row_formats(ws)
        self.assertIn('"$"', ws["B2"].number_format)

    def test_raw_kpi_uses_shipped_format(self):
        _, ws = self._build_sheet([["net_patient_revenue", 6.38e9, 1.25e9, 1.52e9, 2.13e9, 100.0]])
        apply_peer_percentile_row_formats(ws)
        self.assertIn('"$"', ws["B2"].number_format)

    def test_unknown_kpi_leaves_default_format(self):
        _, ws = self._build_sheet([["unmapped_kpi", 1.0, 1.0, 1.0, 1.0, 50.0]])
        formatted = apply_peer_percentile_row_formats(ws)
        self.assertEqual(formatted, 0)

    def test_missing_kpi_column_returns_zero(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["target", "peer_p10"])
        ws.append([1.0, 2.0])
        self.assertEqual(apply_peer_percentile_row_formats(ws), 0)


class TestUI5Workbook(unittest.TestCase):
    """UI-5: data bars, color scales, tab notes on the workbook."""

    def _sheet(self, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        return wb, ws

    def test_data_bar_applies_conditional_formatting(self):
        _, ws = self._sheet(["metric", "value"], [["A", 100], ["B", 200]])
        ret = apply_data_bar(ws, "value")
        self.assertEqual(ret, 1)
        # Rule persisted in conditional_formatting
        rules = list(ws.conditional_formatting._cf_rules.items())
        self.assertTrue(rules)

    def test_data_bar_missing_column_noop(self):
        _, ws = self._sheet(["a"], [[1], [2]])
        self.assertEqual(apply_data_bar(ws, "nope"), 0)

    def test_data_bar_empty_sheet_noop(self):
        wb = Workbook()
        ws = wb.active
        self.assertEqual(apply_data_bar(ws, "anything"), 0)

    def test_color_scale_higher_is_better(self):
        _, ws = self._sheet(
            ["metric", "moic"],
            [["A", 1.5], ["B", 2.5], ["C", 3.0]],
        )
        ret = apply_color_scale(ws, "moic", higher_is_better=True)
        self.assertEqual(ret, 1)
        rules = list(ws.conditional_formatting._cf_rules.items())
        self.assertTrue(rules)

    def test_color_scale_higher_is_worse(self):
        """For variance columns where high=bad, colors invert."""
        _, ws = self._sheet(
            ["metric", "leverage"],
            [["A", 4.0], ["B", 5.5], ["C", 7.0]],
        )
        ret = apply_color_scale(ws, "leverage", higher_is_better=False)
        self.assertEqual(ret, 1)

    def test_color_scale_missing_column_noop(self):
        _, ws = self._sheet(["a"], [[1]])
        self.assertEqual(apply_color_scale(ws, "nope"), 0)

    def test_add_tab_note_prepends_row_with_merged_cell(self):
        _, ws = self._sheet(
            ["metric", "value", "notes"],
            [["A", 100, "x"], ["B", 200, "y"]],
        )
        add_tab_note(ws, "This tab shows summary metrics")
        # Row 1 now carries the note
        self.assertEqual(ws.cell(row=1, column=1).value,
                         "This tab shows summary metrics")
        # Original header pushed to row 2
        self.assertEqual(ws.cell(row=2, column=1).value, "metric")
        # Freeze pane adjusted to A3
        self.assertEqual(ws.freeze_panes, "A3")

    def test_add_tab_note_single_column_sheet(self):
        _, ws = self._sheet(["solo"], [[1], [2]])
        # Single-column sheets: merge condition guards correctly
        add_tab_note(ws, "Single column note")
        self.assertEqual(ws.cell(row=1, column=1).value, "Single column note")

    def test_add_tab_note_font_styled_italic_muted(self):
        _, ws = self._sheet(["a", "b"], [[1, 2]])
        add_tab_note(ws, "note")
        note_cell = ws.cell(row=1, column=1)
        self.assertTrue(note_cell.font.italic)


class TestCoverSheet(unittest.TestCase):
    def test_cover_sheet_inserted_at_position_zero(self):
        wb = Workbook()
        wb.active.title = "Summary"
        wb.create_sheet("Payers")
        build_cover_sheet(wb, hospital_name="Test Hospital",
                          ccn="360180", outdir="/nonexistent")
        self.assertEqual(wb.sheetnames[0], "Cover")

    def test_cover_contains_hospital_metadata(self):
        wb = Workbook()
        wb.active.title = "Summary"
        build_cover_sheet(wb, hospital_name="Example General",
                          ccn="123456", outdir="/nonexistent")
        cells = [cell.value for row in wb["Cover"].iter_rows() for cell in row if cell.value]
        joined = " ".join(str(c) for c in cells)
        self.assertIn("Example General", joined)
        self.assertIn("123456", joined)
        self.assertIn("RCM Due Diligence Workbook", joined)

    def test_cover_toc_lists_other_tabs(self):
        wb = Workbook()
        wb.active.title = "Summary"
        wb.create_sheet("Payers")
        wb.create_sheet("Peer Set")
        build_cover_sheet(wb, hospital_name=None, ccn=None, outdir="/nonexistent")
        cover_text = " ".join(
            str(cell.value) for row in wb["Cover"].iter_rows() for cell in row if cell.value
        )
        for tab in ("Summary", "Payers", "Peer Set"):
            self.assertIn(tab, cover_text)

    def test_cover_reads_grade_from_provenance(self):
        with tempfile.TemporaryDirectory() as tmp:
            import json
            prov = {
                "sources": {
                    "grade": "B",
                    "counts": {"observed": 15, "prior": 0, "assumed": 17, "total": 32},
                }
            }
            with open(os.path.join(tmp, "provenance.json"), "w") as f:
                json.dump(prov, f)
            wb = Workbook()
            wb.active.title = "Summary"
            build_cover_sheet(wb, hospital_name="X", ccn=None, outdir=tmp)
            text = " ".join(str(c.value) for row in wb["Cover"].iter_rows() for c in row if c.value)
            self.assertIn("grade", text.lower())
            self.assertIn("15 of 32", text)
            # Grade letter B appears in the label value
            self.assertIn("B  —", text)


class TestEndToEndWorkbook(unittest.TestCase):
    """Full write_diligence_workbook produces a polished workbook."""

    def test_cover_sheet_is_first(self):
        with tempfile.TemporaryDirectory() as tmp:
            write_diligence_workbook(tmp, _mock_summary(), _load_cfg())
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertEqual(wb.sheetnames[0], "Cover")
            wb.close()

    def test_every_data_sheet_has_frozen_header(self):
        """Header stays frozen. Pre-UI-5 = A2; tabs with a note = A3."""
        with tempfile.TemporaryDirectory() as tmp:
            write_diligence_workbook(tmp, _mock_summary(), _load_cfg())
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            for name in wb.sheetnames:
                if name == "Cover":
                    continue
                freeze = wb[name].freeze_panes
                self.assertIn(
                    freeze, ("A2", "A3"),
                    msg=f"{name} tab freeze pane unexpected: {freeze}",
                )
            wb.close()

    def test_assumptions_tab_has_colored_source_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            write_diligence_workbook(tmp, _mock_summary(), _load_cfg())
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            ws = wb["Assumptions"]
            # UI-5 inserted a note in row 1; the header is on row 2.
            # Try row 1 first (back-compat for pre-UI-5 workbooks), fall
            # back to row 2 when the note is present.
            def _find_source_col(row_idx: int):
                for idx, cell in enumerate(ws[row_idx], start=1):
                    if str(cell.value) == "source":
                        return idx
                return None

            source_col = _find_source_col(1) or _find_source_col(2)
            self.assertIsNotNone(source_col)
            # Data rows start one below the header row
            header_row = 1 if _find_source_col(1) else 2
            data_start = header_row + 1
            colored = sum(
                1 for row in range(data_start, ws.max_row + 1)
                if ws.cell(row=row, column=source_col).fill.fgColor
                and str(ws.cell(row=row, column=source_col).fill.fgColor.rgb) not in ("00000000", "0")
            )
            self.assertGreater(colored, 0)
            wb.close()
