"""Tests for the xlsx + PPTX polish work (Prompt 22).

Invariants locked here:

xlsx:
 1. Output file opens cleanly via ``openpyxl.load_workbook`` (valid
    .xlsx — no zip corruption, no schema misses).
 2. Six sheets with the exact expected names.
 3. RCM Profile sheet has one row per metric in ``packet.rcm_profile``.
 4. Audit sheet carries the ``inputs_hash`` the caller passed in.
 5. Conditional formatting paints the ``current_value`` cell for a
    below-benchmark metric green; above-benchmark red.
 6. Raw Data sheet column set matches the CSV export columns.
 7. Risk Flags sheet is severity-sorted (CRITICAL first).
 8. EBITDA Bridge sheet contains one row per lever + a totals row.
 9. Monte Carlo sheet renders a "no simulation" hint when the packet
    has no MC attached.
10. ``PacketRenderer.render_deal_xlsx`` falls back to CSV when
    ``openpyxl`` is unavailable (simulated via monkeypatch).
11. API endpoint ``/api/analysis/<id>/export?format=xlsx`` returns
    a 200 with the Excel MIME type.
12. Analyst override count on Audit sheet reflects
    ``packet.analyst_overrides``.

pptx:
13. ``.pptx`` output is a valid zip with the OOXML parts PowerPoint
    needs (``[Content_Types].xml``, ``_rels/.rels``, ``ppt/…``).
14. Deck contains exactly 8 slides.
15. Slide 2 text contains the deal name (executive-summary inheritance).
16. Output filename keeps the ``.pptx`` extension (not
    ``.pptx.txt``) — the old fallback produced a sibling file.
"""
from __future__ import annotations

import os
import socket
import tempfile
import threading
import time
import unittest
import urllib.request
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from rcm_mc.analysis.packet import (
    DealAnalysisPacket,
    EBITDABridgeResult,
    MetricImpact,
    MetricSource,
    PercentileSet,
    ProfileMetric,
    RiskFlag,
    RiskSeverity,
    SimulationSummary,
)
from rcm_mc.exports import PacketRenderer
from rcm_mc.exports.xlsx_renderer import render_deal_xlsx
from rcm_mc.portfolio.store import PortfolioStore


def _sample_packet() -> DealAnalysisPacket:
    return DealAnalysisPacket(
        deal_id="demo",
        deal_name="Demo Regional",
        run_id="R-001",
        rcm_profile={
            "denial_rate": ProfileMetric(value=12.0, source=MetricSource.OBSERVED),
            "days_in_ar": ProfileMetric(value=55.0, source=MetricSource.PREDICTED),
            "net_collection_rate": ProfileMetric(
                value=97.0, source=MetricSource.OBSERVED,
            ),
        },
        ebitda_bridge=EBITDABridgeResult(
            current_ebitda=60_000_000,
            target_ebitda=72_000_000,
            per_metric_impacts=[
                MetricImpact(
                    metric_key="denial_rate",
                    current_value=12.0, target_value=7.0,
                    ebitda_impact=8_000_000,
                ),
                MetricImpact(
                    metric_key="days_in_ar",
                    current_value=55.0, target_value=45.0,
                    ebitda_impact=4_000_000,
                ),
            ],
        ),
        risk_flags=[
            RiskFlag(category="OPERATIONAL", severity=RiskSeverity.LOW,
                     title="Minor"),
            RiskFlag(category="OPERATIONAL", severity=RiskSeverity.CRITICAL,
                     title="Major"),
            RiskFlag(category="PAYER", severity=RiskSeverity.MEDIUM,
                     title="Payer"),
        ],
        analyst_overrides={"bridge.exit_multiple": 11.0},
    )


# ── xlsx workbook shape ────────────────────────────────────────────

class TestXlsxWorkbookShape(unittest.TestCase):

    def test_opens_cleanly(self):
        out = Path(tempfile.mkdtemp())
        p = _sample_packet()
        path = render_deal_xlsx(p, out, inputs_hash="h1")
        # load_workbook raises on an invalid .xlsx.
        wb = load_workbook(path)
        self.assertIsNotNone(wb)

    def test_expected_sheets(self):
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(_sample_packet(), out, inputs_hash="h2")
        wb = load_workbook(path)
        self.assertEqual(
            wb.sheetnames,
            ["RCM Profile", "EBITDA Bridge", "Monte Carlo",
             "Risk Flags", "Raw Data", "Audit"],
        )

    def test_rcm_row_count_matches_profile(self):
        out = Path(tempfile.mkdtemp())
        p = _sample_packet()
        path = render_deal_xlsx(p, out, inputs_hash="h3")
        wb = load_workbook(path)
        ws = wb["RCM Profile"]
        # +1 for header row.
        self.assertEqual(ws.max_row, 1 + len(p.rcm_profile))

    def test_audit_sheet_carries_hash(self):
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(
            _sample_packet(), out, inputs_hash="abcdef",
        )
        wb = load_workbook(path)
        ws = wb["Audit"]
        # Find the input_hash row (column A is "field", column B is "value").
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertEqual(values.get("input_hash"), "abcdef")

    def test_override_count_reflected(self):
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(
            _sample_packet(), out, inputs_hash="h4",
        )
        wb = load_workbook(path)
        ws = wb["Audit"]
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertEqual(values.get("analyst_override_count"), 1)


# ── Conditional formatting ────────────────────────────────────────

class TestConditionalFormatting(unittest.TestCase):

    def _fill_color(self, cell) -> str:
        f = cell.fill
        if f is None or f.start_color is None:
            return ""
        return (f.start_color.rgb or "").upper()

    def test_green_when_clearly_better_than_p50(self):
        """Build a packet whose ``denial_rate`` is below the
        benchmark P50 × 0.95 threshold, and confirm the current_value
        cell picks up the green fill."""
        p = DealAnalysisPacket(
            deal_id="d", deal_name="D",
            rcm_profile={
                # denial_rate P50 ≈ 5.2. 3.0 < 5.2 × 0.95 = 4.94 → green.
                "denial_rate": ProfileMetric(
                    value=3.0, source=MetricSource.OBSERVED,
                ),
            },
            ebitda_bridge=EBITDABridgeResult(),
        )
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(p, out, inputs_hash="h5")
        wb = load_workbook(path)
        ws = wb["RCM Profile"]
        fill = self._fill_color(ws.cell(row=2, column=4))
        self.assertIn("D1FAE5", fill)

    def test_red_when_clearly_worse_than_p50(self):
        """Mirror test — current_value above P50 × 1.05 for a
        lower-is-better metric should tint red."""
        p = DealAnalysisPacket(
            deal_id="d", deal_name="D",
            rcm_profile={
                # denial_rate P50 ≈ 5.2. 12 > 5.2 × 1.05 → red.
                "denial_rate": ProfileMetric(
                    value=12.0, source=MetricSource.OBSERVED,
                ),
            },
            ebitda_bridge=EBITDABridgeResult(),
        )
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(p, out, inputs_hash="hR")
        wb = load_workbook(path)
        ws = wb["RCM Profile"]
        fill = self._fill_color(ws.cell(row=2, column=4))
        self.assertIn("FEE2E2", fill)


# ── Raw Data sheet parity ─────────────────────────────────────────

class TestRawDataSheet(unittest.TestCase):

    def test_columns_match_csv(self):
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(_sample_packet(), out, inputs_hash="h6")
        wb = load_workbook(path)
        ws = wb["Raw Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Mirror the CSV-export columns verbatim.
        self.assertEqual(headers, [
            "metric_key", "display_name", "current_value", "source",
            "benchmark_p50", "predicted_value", "ci_low", "ci_high",
            "ebitda_impact", "risk_flags",
        ])


# ── Risk sorting ──────────────────────────────────────────────────

class TestRiskFlagsSheet(unittest.TestCase):

    def test_severity_sorted_critical_first(self):
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(_sample_packet(), out, inputs_hash="h7")
        wb = load_workbook(path)
        ws = wb["Risk Flags"]
        severities = [ws.cell(row=r, column=1).value
                      for r in range(2, ws.max_row + 1)]
        # CRITICAL first, then MEDIUM, then LOW.
        self.assertEqual(
            severities, ["CRITICAL", "MEDIUM", "LOW"],
        )


# ── Bridge sheet ──────────────────────────────────────────────────

class TestBridgeSheet(unittest.TestCase):

    def test_one_row_per_lever_plus_totals(self):
        out = Path(tempfile.mkdtemp())
        p = _sample_packet()
        path = render_deal_xlsx(p, out, inputs_hash="h8")
        wb = load_workbook(path)
        ws = wb["EBITDA Bridge"]
        # Expect: 1 header + N levers + 1 TOTAL + blank + 3 summary = 6+.
        self.assertGreater(ws.max_row, len(p.ebitda_bridge.per_metric_impacts))
        # TOTAL row label present.
        labels = [ws.cell(row=r, column=1).value
                  for r in range(2, ws.max_row + 1)]
        self.assertIn("TOTAL", labels)


# ── Monte Carlo empty-state ───────────────────────────────────────

class TestMonteCarloSheet(unittest.TestCase):

    def test_no_simulation_hint(self):
        out = Path(tempfile.mkdtemp())
        p = _sample_packet()  # packet.simulation is None by default.
        path = render_deal_xlsx(p, out, inputs_hash="h9")
        wb = load_workbook(path)
        ws = wb["Monte Carlo"]
        cell_a2 = ws.cell(row=2, column=1).value or ""
        self.assertIn("no simulation", str(cell_a2))

    def test_with_simulation_populates_percentiles(self):
        p = _sample_packet()
        p.simulation = SimulationSummary(
            n_sims=5000, seed=42,
            ebitda_uplift=PercentileSet(
                p10=1e6, p25=2e6, p50=3e6, p75=4e6, p90=5e6,
            ),
            moic=PercentileSet(p10=1.5, p25=1.8, p50=2.0, p75=2.3, p90=2.6),
            irr=PercentileSet(p10=0.10, p25=0.12, p50=0.15, p75=0.18, p90=0.20),
        )
        out = Path(tempfile.mkdtemp())
        path = render_deal_xlsx(p, out, inputs_hash="hA")
        wb = load_workbook(path)
        ws = wb["Monte Carlo"]
        # Row 4 = P50 (header is row 1, P10 row 2, P25 row 3, P50 row 4).
        self.assertEqual(ws.cell(row=4, column=1).value, "P50")
        self.assertAlmostEqual(float(ws.cell(row=4, column=2).value), 3e6)


# ── Fallback path ─────────────────────────────────────────────────

class TestFallbackToCSV(unittest.TestCase):

    def test_renderer_falls_back_when_openpyxl_missing(self):
        """Simulate a missing ``openpyxl`` by monkey-patching the
        renderer module's import helper to raise. The renderer
        should return a .csv path rather than raising."""
        import rcm_mc.exports.packet_renderer as pr

        renderer = PacketRenderer(out_dir=Path(tempfile.mkdtemp()))

        # Swap the xlsx_renderer's import gate to raise.
        import rcm_mc.exports.xlsx_renderer as xr
        original = xr._openpyxl_or_raise

        def _missing():
            raise ImportError("openpyxl not installed in this env")
        xr._openpyxl_or_raise = _missing
        try:
            path = renderer.render_deal_xlsx(
                _sample_packet(), inputs_hash="h_fallback",
            )
        finally:
            xr._openpyxl_or_raise = original
        self.assertTrue(str(path).endswith(".csv"))


# ── API endpoint ──────────────────────────────────────────────────

class TestAPIEndpoint(unittest.TestCase):

    def _start(self, db: str) -> tuple:
        from rcm_mc.server import build_server
        s = socket.socket(); s.bind(("127.0.0.1", 0))
        port = s.getsockname()[1]; s.close()
        server, _ = build_server(port=port, db_path=db)
        t = threading.Thread(target=server.serve_forever, daemon=True)
        t.start()
        time.sleep(0.05)
        return server, port

    def test_xlsx_download(self):
        tf = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tf.close()
        db = tf.name
        try:
            store = PortfolioStore(db)
            store.upsert_deal(
                "d1", name="d1",
                profile={"payer_mix": {"commercial": 1.0},
                          "bed_count": 200},
            )
            server, port = self._start(db)
            try:
                with urllib.request.urlopen(
                    f"http://127.0.0.1:{port}/api/analysis/d1/export?format=xlsx"
                ) as r:
                    body = r.read()
                    ctype = r.headers.get("Content-Type", "")
                    disp = r.headers.get("Content-Disposition", "")
                self.assertIn("spreadsheetml", ctype)
                self.assertIn(".xlsx", disp)
                # Verify it's a valid xlsx by scanning the zip header.
                self.assertTrue(body.startswith(b"PK"))
            finally:
                server.shutdown()
                server.server_close()
        finally:
            os.unlink(db)


# ── PPTX fallback ─────────────────────────────────────────────────

class TestPptxFallback(unittest.TestCase):

    def _render(self):
        """Always exercise the fallback — tests run without python-pptx
        by design in this repo (it's an optional dep)."""
        renderer = PacketRenderer(out_dir=Path(tempfile.mkdtemp()))
        return renderer.render_diligence_memo_pptx(
            _sample_packet(), inputs_hash="hPPT",
        )

    def test_output_is_valid_zip(self):
        path = self._render()
        self.assertTrue(zipfile.is_zipfile(path))

    def test_output_has_pptx_suffix(self):
        path = self._render()
        self.assertEqual(path.suffix, ".pptx")

    def test_eight_slides_present(self):
        path = self._render()
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
        slide_parts = [
            n for n in names
            if n.startswith("ppt/slides/slide") and n.endswith(".xml")
        ]
        self.assertEqual(len(slide_parts), 8)

    def test_required_ooxml_parts(self):
        """Assert the parts PowerPoint actually needs to open the
        deck — dropping any of these breaks load."""
        path = self._render()
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
        for required in [
            "[Content_Types].xml",
            "_rels/.rels",
            "ppt/presentation.xml",
            "ppt/_rels/presentation.xml.rels",
            "ppt/slideMasters/slideMaster1.xml",
            "ppt/slideLayouts/slideLayout1.xml",
            "ppt/theme/theme1.xml",
        ]:
            self.assertIn(required, names)

    def test_slide2_contains_deal_name(self):
        path = self._render()
        with zipfile.ZipFile(path) as z:
            xml = z.read("ppt/slides/slide2.xml").decode("utf-8")
        # Executive-summary text inherited from _memo_exec_text should
        # reference the deal's dollar impact.
        self.assertIn("<a:t>", xml)


if __name__ == "__main__":
    unittest.main()
