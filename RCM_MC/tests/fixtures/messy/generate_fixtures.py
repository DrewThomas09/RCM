"""Generators for the 10 adversarial ingester fixtures.

Each generator is deterministic under a seed so regression tests hold
byte-for-byte. Every fixture produces one or more raw files plus an
``expected.json`` alongside. The expected file is the *test contract*:
if the fixture's intent changes, ``expected.json`` changes and the
regression test fails loudly.

Each entry in :data:`FIXTURES` has:

- ``name``: directory name under ``tests/fixtures/messy/``
- ``generate(out_dir)``: writes the raw files
- ``expected()``: returns the ``expected.json`` dict

Running ``regenerate_all()`` (also used in tests when
``REGENERATE_FIXTURES=1`` is set) rewrites every file. Keep seeds
stable — a re-seed requires walking every expected.json.
"""
from __future__ import annotations

import csv
import json
import random
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

# ── Seeds ───────────────────────────────────────────────────────────
#
# One seed per fixture so tweaking one doesn't ripple into the others.
SEEDS = {
    "fixture_01_clean_837":             101,
    "fixture_02_mixed_ehr_rollup":      102,
    "fixture_03_excel_merged_cells":    103,
    "fixture_04_payer_typos":           104,
    "fixture_05_date_format_hell":      105,
    "fixture_06_partial_837":           106,
    "fixture_07_encoding_chaos":        107,
    "fixture_08_cpt_icd_drift":         108,
    "fixture_09_duplicate_claims":      109,
    "fixture_10_zero_balance_writeoffs": 110,
}


# ── Shared small code universes ────────────────────────────────────

SAMPLE_HCPCS = [
    "99213", "99214", "99203", "99204", "36415",
    "85025", "80053", "93000", "71046", "97110",
]
SAMPLE_ICD = ["E11.9", "I10", "M25.561", "J06.9", "Z00.00", "R07.9"]


# ═══════════════════════════════════════════════════════════════════
# fixture_01 — clean 837
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_01(out: Path) -> None:
    seed = SEEDS["fixture_01_clean_837"]
    r = random.Random(seed)
    # Two X12 837P claims, enough to exercise the parser.
    segments = [
        "ISA*00*          *00*          *ZZ*SUBMITTER      *ZZ*RECEIVER       *240115*1200*^*00501*000000001*0*P*:",
        "GS*HC*SUBMITTER*RECEIVER*20240115*1200*1*X*005010X222A1",
        "ST*837*0001*005010X222A1",
        "BHT*0019*00*TESTBATCH*20240115*1200*CH",
    ]
    claim_rows = [
        ("C-00001", "P-1001", "99213", "E11.9", 150.00, "20240115", "Blue Cross"),
        ("C-00002", "P-1002", "99214", "I10",   220.00, "20240118", "Medicare"),
    ]
    for cid, pid, cpt, icd, amt, dos, payer in claim_rows:
        segments += [
            f"NM1*PR*2*{payer}*****PI*12345",
            f"NM1*QC*1*Smith*John****MI*{pid}",
            f"CLM*{cid}*{amt:.2f}***11:B:1*Y*A*Y*Y",
            f"DTP*472*D8*{dos}",
            f"HI*ABK:{icd}",
            f"SV1*HC:{cpt}*{amt:.2f}*UN*1***1",
        ]
    segments += ["SE*20*0001", "GE*1*1", "IEA*1*000000001"]
    text = "~".join(segments) + "~"
    (out / "claims.edi").write_text(text, encoding="latin-1")


def expected_fixture_01() -> Dict[str, Any]:
    return {
        "fixture": "fixture_01_clean_837",
        "description": "Baseline clean 837 with two claims.",
        "canonical_claim_count": 2,
        "distinct_source_systems": ["edi_837"],
        "payer_class_counts": {"COMMERCIAL": 1, "MEDICARE": 1},
        "must_have_transformation_rules": [
            "cpt_validate:ok",
            "icd_validate:icd10_ok",
            "payer_resolve:exact",
            "date_parse:iso",
        ],
        "must_not_have_severity_error": True,
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_02 — mixed-EHR rollup (three clinics, same 10 claims)
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_02(out: Path) -> None:
    r = random.Random(SEEDS["fixture_02_mixed_ehr_rollup"])

    # Canonical 10 logical claims shared across all three EHRs.
    logical: List[Dict[str, Any]] = []
    for i in range(10):
        dos = date(2024, 3, 1) + timedelta(days=i)
        logical.append({
            "patient_mrn": f"P-200{i}",
            "dob": date(1970 + i, 4, 2).isoformat(),
            "dos": dos,
            "cpt": r.choice(SAMPLE_HCPCS),
            "icd": r.choice(SAMPLE_ICD),
            "charge": round(100 + i * 15, 2),
            "paid": round(80 + i * 9, 2),
            "payer": r.choice(["Blue Cross", "Aetna", "Medicare", "UHC"]),
        })

    # Epic-style: snake_case, ISO dates.
    (out / "epic").mkdir(parents=True, exist_ok=True)
    with (out / "epic" / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_id", "patient_id", "dob", "service_date_from",
            "cpt_code", "diagnosis_code_1", "payer",
            "charge_amount", "paid_amount",
        ])
        w.writeheader()
        for i, c in enumerate(logical):
            w.writerow({
                "claim_id": f"EPIC-{i:04d}",
                "patient_id": c["patient_mrn"],
                "dob": c["dob"],
                "service_date_from": c["dos"].isoformat(),
                "cpt_code": c["cpt"],
                "diagnosis_code_1": c["icd"],
                "payer": c["payer"],
                "charge_amount": c["charge"],
                "paid_amount": c["paid"],
            })

    # Cerner-style: MRN column, MM/DD/YYYY dates, "payer_name".
    (out / "cerner").mkdir(parents=True, exist_ok=True)
    with (out / "cerner" / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_number", "mrn", "date_of_birth", "date_of_service",
            "hcpcs_code", "primary_dx", "payer_name",
            "billed_amount", "insurance_paid",
        ])
        w.writeheader()
        for i, c in enumerate(logical):
            dos: date = c["dos"]
            w.writerow({
                "claim_number": f"CER-{i:04d}",
                "mrn": c["patient_mrn"],
                "date_of_birth": date.fromisoformat(c["dob"]).strftime("%m/%d/%Y"),
                "date_of_service": dos.strftime("%m/%d/%Y"),
                "hcpcs_code": c["cpt"],
                "primary_dx": c["icd"],
                "payer_name": c["payer"],
                "billed_amount": c["charge"],
                "insurance_paid": c["paid"],
            })

    # Athena-style: member_number + control_id + ISO dates + Parquet.
    (out / "athena").mkdir(parents=True, exist_ok=True)
    import pyarrow as pa
    import pyarrow.parquet as pq
    tbl_rows: List[Dict[str, Any]] = []
    for i, c in enumerate(logical):
        tbl_rows.append({
            "control_id": f"ATH-{i:04d}",
            "member_number": c["patient_mrn"],
            "birth_date": c["dob"],
            "claim_start_date": c["dos"].isoformat(),
            "procedure_code": c["cpt"],
            "primary_diagnosis": c["icd"],
            "insurance_name": c["payer"],
            "total_charge": c["charge"],
            "payment_amount": c["paid"],
        })
    pq.write_table(pa.Table.from_pylist(tbl_rows), out / "athena" / "claims.parquet")


def expected_fixture_02() -> Dict[str, Any]:
    return {
        "fixture": "fixture_02_mixed_ehr_rollup",
        "description": "Three EHRs (Epic, Cerner, Athena) express the same 10 claims.",
        "canonical_claim_count": 30,            # 10 × 3 source_system rows
        "logical_claim_count_after_rollup": 10,  # collapsed claim_ids
        "distinct_source_systems": ["athena", "cerner", "epic"],
        "must_have_transformation_rules": [
            "multi_ehr_rollup",
            "payer_resolve:exact",
        ],
        "must_not_have_severity_error": True,
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_03 — Excel merged cells + junk header + trailing totals
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_03(out: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Claims"

    # Junk header rows (two banner rows)
    ws.cell(row=1, column=1, value="Q4 2024 Claims Extract")
    ws.cell(row=2, column=1, value="Exported: 2025-01-04")

    # Real header row
    header = [
        "Claim ID", "MRN", "DOS", "CPT", "Payer",
        "Charge", "Paid",
    ]
    for col, val in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=val)

    # Merged cells — merge the Payer across four rows to simulate
    # "same payer for this block" data entry.
    ws.cell(row=5, column=5, value="Aetna")
    ws.merge_cells(start_row=5, end_row=8, start_column=5, end_column=5)

    rows = [
        ("X-001", "P-3001", "2024-11-01", "99213", None, 150.00, 120.00),
        ("X-002", "P-3002", "2024-11-02", "99214", None, 220.00, 180.00),
        ("X-003", "P-3003", "2024-11-03", "85025", None, 45.00,  36.00),
        ("X-004", "P-3004", "2024-11-04", "80053", None, 60.00,  48.00),
        ("X-005", "P-3005", "2024-11-05", "99204", "Cigna", 320.00, 256.00),
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            if c_idx == 5 and val is None:
                continue  # honoured by the merge
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Trailing total row
    ws.cell(row=11, column=1, value="Grand Total")
    ws.cell(row=11, column=6, value=795.00)
    ws.cell(row=11, column=7, value=640.00)

    wb.save(out / "claims.xlsx")


def expected_fixture_03() -> Dict[str, Any]:
    return {
        "fixture": "fixture_03_excel_merged_cells",
        "description": "Excel with junk headers, merged payer cells, trailing total row.",
        "canonical_claim_count": 5,
        "distinct_source_systems": ["claims"],
        # Payer distribution: first 4 rows get Aetna from merged cell,
        # last row gets Cigna.
        "payer_canonical_counts": {"Aetna": 4, "Cigna": 1},
        "must_not_include_source_rows": ["Grand Total"],
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_04 — payer typos
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_04(out: Path) -> None:
    with (out / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_id", "patient_id", "date_of_service", "cpt_code",
            "payer", "charge_amount", "paid_amount",
        ])
        w.writeheader()
        variants = [
            "Blue Cross", "BCBS of IL", "Blue X BS", "bcbs",
            "BLUECROSS", "blue cross blue shield", "BCBS OF TX",
        ]
        for i, v in enumerate(variants):
            w.writerow({
                "claim_id": f"Q-{i:04d}",
                "patient_id": f"P-400{i}",
                "date_of_service": "2024-09-15",
                "cpt_code": "99213",
                "payer": v,
                "charge_amount": 150.00,
                "paid_amount": 120.00,
            })


def expected_fixture_04() -> Dict[str, Any]:
    return {
        "fixture": "fixture_04_payer_typos",
        "description": "Seven Blue Cross spellings must resolve to one canonical + class.",
        "canonical_claim_count": 7,
        "distinct_payer_canonicals": ["Blue Cross Blue Shield"],
        "payer_class_counts": {"COMMERCIAL": 7},
        "must_have_transformation_rules": [
            "payer_resolve:exact",
        ],
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_05 — date format hell
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_05(out: Path) -> None:
    rows = [
        ("D-001", "P-5001", "2024-03-15",    "99213"),   # ISO
        ("D-002", "P-5002", "03/15/2024",    "99213"),   # US
        ("D-003", "P-5003", "03-15-2024",    "99213"),   # US dashed
        ("D-004", "P-5004", "15.03.2024",    "99213"),   # European
        ("D-005", "P-5005", "45366",         "99213"),   # Excel serial (2024-03-15)
        ("D-006", "P-5006", "3/15/24",       "99213"),   # 2-digit year
        ("D-007", "P-5007", "1710460800",    "99213"),   # Unix epoch (2024-03-15)
    ]
    with (out / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_id", "patient_id", "date_of_service", "cpt_code",
            "payer", "charge_amount", "paid_amount",
        ])
        w.writeheader()
        for cid, pid, dos, cpt in rows:
            w.writerow({
                "claim_id": cid, "patient_id": pid,
                "date_of_service": dos, "cpt_code": cpt,
                "payer": "Blue Cross", "charge_amount": 150.00,
                "paid_amount": 120.00,
            })


def expected_fixture_05() -> Dict[str, Any]:
    return {
        "fixture": "fixture_05_date_format_hell",
        "description": "Seven rows, seven date formats, one canonical date each.",
        "canonical_claim_count": 7,
        "all_service_dates_non_null": True,
        "must_have_transformation_rules": [
            "date_parse:iso",
            "date_parse:us_slash",
            "date_parse:us_dash",
            "date_parse:eu_dot",
            "date_parse:excel_serial",
            "date_parse:unix_epoch",
        ],
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_06 — partial/truncated 837
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_06(out: Path) -> None:
    # Two complete CLM segments followed by a truncated third.
    segments = [
        "ISA*00*          *00*          *ZZ*SUBMITTER      *ZZ*RECEIVER       *240115*1200*^*00501*000000001*0*P*:",
        "ST*837*0001*005010X222A1",
        "BHT*0019*00*TESTBATCH*20240115*1200*CH",
        "NM1*PR*2*Aetna*****PI*12345",
        "NM1*QC*1*Smith*John****MI*P-6001",
        "CLM*T-0001*150.00***11:B:1*Y*A*Y*Y",
        "DTP*472*D8*20240115",
        "HI*ABK:E11.9",
        "SV1*HC:99213*150.00*UN*1***1",
        "NM1*PR*2*Blue Cross*****PI*12346",
        "CLM*T-0002*220.00***11:B:1*Y*A*Y*Y",
        "DTP*472*D8*20240120",
        "HI*ABK:I10",
        "SV1*HC:99214*220.00*UN*1***1",
        # Third claim begins but file ends mid-segment.
        "CLM*T-0003*180.00",     # truncated — no terminator.
    ]
    text = "~".join(segments[:-1]) + "~" + segments[-1]   # no trailing ~
    (out / "claims.edi").write_text(text, encoding="latin-1")


def expected_fixture_06() -> Dict[str, Any]:
    return {
        "fixture": "fixture_06_partial_837",
        "description": "Truncated 837 — first two claims parse, third is incomplete.",
        "canonical_claim_count_min": 2,
        "distinct_source_systems": ["edi_837"],
        "reader_reports_malformed": True,
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_07 — encoding chaos
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_07(out: Path) -> None:
    # Mix utf-8 and windows-1252 in the same file.
    rows_utf8 = [
        "claim_id,patient_id,date_of_service,cpt_code,payer,charge_amount,paid_amount",
        "E-001,P-7001,2024-01-15,99213,Aetna,150.00,120.00",
        "E-002,P-7002,2024-01-16,99213,Blue Cross,150.00,120.00",
    ]
    # A row with a cp1252 smart quote and an em dash (0x91, 0x97).
    row_cp1252 = b"E-003,P-7003,2024-01-17,99214,Cigna \x91Plus\x92 \x97 Gold,220.00,180.00\n"
    row_utf8 = "E-004,P-7004,2024-01-18,99214,Humana,220.00,180.00\n"
    data = ("\n".join(rows_utf8) + "\n").encode("utf-8") + row_cp1252 + row_utf8.encode("utf-8")
    (out / "claims.csv").write_bytes(data)


def expected_fixture_07() -> Dict[str, Any]:
    return {
        "fixture": "fixture_07_encoding_chaos",
        "description": "Mixed UTF-8 and windows-1252 in the same file.",
        "canonical_claim_count": 4,
        "reader_encoding_in": ["windows-1252", "mixed"],
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_08 — CPT / ICD drift
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_08(out: Path) -> None:
    rows = [
        # Valid HCPCS + ICD-10
        ("CD-001", "P-8001", "2024-02-01", "99213", "E11.9", "Aetna"),
        # Deprecated CPT (8-char proprietary)
        ("CD-002", "P-8002", "2024-02-02", "LEGACY07", "I10", "Aetna"),
        # ICD-9 legacy code
        ("CD-003", "P-8003", "2024-02-03", "99214", "250.00", "Aetna"),
        # ICD-9 V-code
        ("CD-004", "P-8004", "2024-02-04", "99213", "V70.0", "Aetna"),
    ]
    with (out / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_id", "patient_id", "date_of_service", "cpt_code",
            "diagnosis_code_1", "payer", "charge_amount", "paid_amount",
        ])
        w.writeheader()
        for cid, pid, dos, cpt, icd, payer in rows:
            w.writerow({
                "claim_id": cid, "patient_id": pid,
                "date_of_service": dos, "cpt_code": cpt,
                "diagnosis_code_1": icd, "payer": payer,
                "charge_amount": 150.00, "paid_amount": 120.00,
            })


def expected_fixture_08() -> Dict[str, Any]:
    return {
        "fixture": "fixture_08_cpt_icd_drift",
        "description": "Proprietary CPT + ICD-9 legacy codes. Preserved, marked.",
        "canonical_claim_count": 4,
        "must_have_transformation_rules": [
            "cpt_validate:non_standard",
            "icd_validate:icd9_legacy",
        ],
        "preserves_original_codes": True,
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_09 — duplicate-resubmit
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_09(out: Path) -> None:
    # Same (patient, dos, cpt) with three different claim_ids — the
    # "resubmitted under a new control number" pattern.
    logical = [
        ("P-9001", "2024-05-05", "99213", 150.00, 120.00),
        ("P-9002", "2024-05-06", "99214", 220.00, 180.00),
    ]
    rows = []
    for i, (pid, dos, cpt, ch, pd) in enumerate(logical):
        for k in range(3):                    # three resubmits
            rows.append({
                "claim_id": f"R-{i:03d}-{k}",
                "patient_id": pid,
                "date_of_service": dos,
                "cpt_code": cpt,
                "payer": "Medicare",
                "charge_amount": ch,
                "paid_amount": pd if k == 2 else 0,
            })
    with (out / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def expected_fixture_09() -> Dict[str, Any]:
    return {
        "fixture": "fixture_09_duplicate_claims",
        "description": "Two logical claims each resubmitted three times under different IDs.",
        "canonical_claim_count": 6,
        "must_have_transformation_rules": ["duplicate_resubmit"],
        "duplicate_resubmit_cohorts_min": 2,
    }


# ═══════════════════════════════════════════════════════════════════
# fixture_10 — ZBA write-offs
# ═══════════════════════════════════════════════════════════════════

def gen_fixture_10(out: Path) -> None:
    rows = [
        # Paid in full
        ("Z-001", "P-1001", "2024-07-01", "99213", "Aetna",
         150.00, 150.00, 120.00, 0.0, ""),
        # Partial payment + adjustment
        ("Z-002", "P-1002", "2024-07-02", "99213", "Aetna",
         150.00, 150.00, 90.00,  30.00, "CO-45"),
        # ZERO balance — adjustment code forced paid to 0. Original
        # charge + allowed MUST be preserved.
        ("Z-003", "P-1003", "2024-07-03", "99214", "Aetna",
         220.00, 180.00, 0.00, 180.00, "CO-45;OA-23"),
        # Denial — zero paid, zero adjustment to allowed
        ("Z-004", "P-1004", "2024-07-04", "99214", "Aetna",
         220.00, 0.00,    0.00, 0.00,   "CO-50"),
    ]
    with (out / "claims.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "claim_id", "patient_id", "date_of_service", "cpt_code", "payer",
            "charge_amount", "allowed_amount", "paid_amount",
            "adjustment_amount", "adjustment_reason_codes",
        ])
        w.writeheader()
        for row in rows:
            w.writerow({
                "claim_id": row[0], "patient_id": row[1],
                "date_of_service": row[2], "cpt_code": row[3],
                "payer": row[4], "charge_amount": row[5],
                "allowed_amount": row[6], "paid_amount": row[7],
                "adjustment_amount": row[8], "adjustment_reason_codes": row[9],
            })


def expected_fixture_10() -> Dict[str, Any]:
    return {
        "fixture": "fixture_10_zero_balance_writeoffs",
        "description": "ZBA write-off must preserve original charge + allowed AND record adjustment.",
        "canonical_claim_count": 4,
        "zba_row_must_preserve": {
            "claim_id": "Z-003",
            "charge_amount": 220.00,
            "allowed_amount": 180.00,
            "paid_amount": 0.0,
            "adjustment_amount": 180.00,
            "adjustment_reason_codes_min_length": 2,
        },
        "must_have_transformation_rules": ["zba_writeoff:preserve"],
    }


# ── Registry ────────────────────────────────────────────────────────

FIXTURES: Dict[str, Dict[str, Any]] = {
    "fixture_01_clean_837": {"gen": gen_fixture_01, "expected": expected_fixture_01},
    "fixture_02_mixed_ehr_rollup": {"gen": gen_fixture_02, "expected": expected_fixture_02},
    "fixture_03_excel_merged_cells": {"gen": gen_fixture_03, "expected": expected_fixture_03},
    "fixture_04_payer_typos": {"gen": gen_fixture_04, "expected": expected_fixture_04},
    "fixture_05_date_format_hell": {"gen": gen_fixture_05, "expected": expected_fixture_05},
    "fixture_06_partial_837": {"gen": gen_fixture_06, "expected": expected_fixture_06},
    "fixture_07_encoding_chaos": {"gen": gen_fixture_07, "expected": expected_fixture_07},
    "fixture_08_cpt_icd_drift": {"gen": gen_fixture_08, "expected": expected_fixture_08},
    "fixture_09_duplicate_claims": {"gen": gen_fixture_09, "expected": expected_fixture_09},
    "fixture_10_zero_balance_writeoffs": {"gen": gen_fixture_10, "expected": expected_fixture_10},
}


# ── Regeneration driver ─────────────────────────────────────────────

def regenerate_one(name: str, root: Path) -> Path:
    entry = FIXTURES[name]
    out = root / name
    # Wipe any previous run to keep the regeneration pure.
    if out.exists():
        for p in sorted(out.rglob("*"), reverse=True):
            if p.is_file():
                p.unlink()
            elif p.is_dir():
                p.rmdir()
    out.mkdir(parents=True, exist_ok=True)
    entry["gen"](out)
    (out / "expected.json").write_text(
        json.dumps(entry["expected"](), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return out


def regenerate_all(root: Path | None = None) -> Dict[str, Path]:
    root = root or Path(__file__).resolve().parent
    return {name: regenerate_one(name, root) for name in FIXTURES}


if __name__ == "__main__":
    out = regenerate_all()
    for name, p in out.items():
        print(f"  {name:40s} → {p}")
