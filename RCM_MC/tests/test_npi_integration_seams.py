"""Integration seams between the three NPI-cleaner improvement fronts.

The engine-cases, connectors, and pipeline-report fronts each shipped inside
their own file sets and left the CROSS-FILE joints to the integrator:

  * engine-side CARC group-prefix normalization (pipeline-report P0 handoff),
  * the compliance screen hoisted out of the enrich flag (connectors P0),
  * billed-dollar weights into the bounded NPPES verify cap (connectors P2),
  * dollars_by_npi into the LEIE screen (pipeline-report handoff),
  * rendering-grain coding intensity (pipeline-report handoff),
  * unconditional catalog/plan on the scorecard + streamed runs (connectors P3),
  * autofiled integration gaps (connectors P2),
  * offline pack installs via the CLI (connectors handoff),
  * exec/workbook rendering of the connector coverage payloads (connectors P1).

Every test drives a producer front's payload through a consumer front's REAL
code path (engine.clean_bytes / bigfile.clean_path / cli.main / the report
builders). No mocks of our own modules — only the external network clients
(NPPES lookup, PECOS) and the deep preflight probe are stubbed, exactly the
injection seams those modules expose for tests.
"""
import io
import json
import os
import sqlite3
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from rcm_mc.npi_cleaner import engine

GOOD = "1679576722"     # Luhn-valid NPI
GOOD_B = "1234567893"   # Luhn-valid NPI


def _workdir() -> str:
    return tempfile.mkdtemp(prefix="npi-seams-")


def _tmp_pack_db() -> Path:
    return Path(tempfile.mkdtemp(prefix="npi-seams-packs-")) / "packs.sqlite3"


def _reset_pack_cache():
    """refdata_packs memoizes pack contents per process; tests that swap
    _DB_PATH must drop the memo or they read the previous test's packs."""
    from rcm_mc.npi_cleaner import refdata_packs
    with refdata_packs._CACHE_LOCK:
        refdata_packs._CACHE.clear()


class CarcNormalizationSeamTests(unittest.TestCase):
    """Engine ↔ report seam: group-prefixed CARCs ("CO-45") must grade and
    tally as their bare code so the validity dimension and the playbook join
    survive 835-style exports — the display layers already normalized, the
    engine was the missing half."""

    def _clean(self, body: str):
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            return engine.clean_bytes(body.encode(), "carc.csv")

    def test_group_prefixed_carcs_do_not_flag_invalid(self):
        body = "ClaimID,DenialCode\n" + "".join(
            f"{i},CO-45\n" for i in range(1, 21))
        res = self._clean(body)
        self.assertIsNone(res.sanity.get("carc-invalid"))

    def test_prefixed_and_bare_merge_with_playbook_and_variants(self):
        body = ("ClaimID,DenialCode\n"
                + "".join(f"{i},CO-45\n" for i in range(1, 21))
                + "21,45\n22,PR-1\n23,16\n")
        res = self._clean(body)
        self.assertIsNone(res.sanity.get("carc-invalid"))
        top = {d["code"]: d for d in res.denials["top"]}
        self.assertEqual(top["45"]["count"], 21)          # CO-45 + bare 45
        self.assertIn("CO-45", top["45"].get("as_seen", []))
        self.assertEqual(top["45"]["category"], "contractual")
        self.assertEqual(top["1"]["count"], 1)            # PR-1 → 1
        self.assertIn("preventable_pct", res.denials)

    def test_true_garbage_still_flags(self):
        res = self._clean("ClaimID,DenialCode\n1,PENDING\n2,CO-45\n")
        self.assertEqual(res.sanity.get("carc-invalid"), 1)

    def test_exec_report_renders_merged_denial_mix(self):
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        body = ("ClaimID,DenialCode\n"
                + "".join(f"{i},CO-45\n" for i in range(1, 21)))
        res = self._clean(body)
        html = build_exec_report(res.as_scorecard(), "carc.csv",
                                 "2026-07-07T00:00:00+00:00")
        # One merged row for code 45, carrying the playbook action.
        self.assertIn("45", html)
        self.assertNotIn("CO-45</td><td", html.replace("CO-45 ", ""))
        self.assertIn("contractual", html.lower())


class ComplianceHoistSeamTests(unittest.TestCase):
    """connectors-front P0 handoff: deep-without-enrich must not silently
    drop the compliance screen, and the LEIE match must carry the billed
    exposure the pipeline front's renderers already know how to show."""

    def setUp(self):
        self._env = patch.dict(os.environ,
                               {"RCM_MC_NPI_WORKDIR": _workdir()})
        self._env.start()
        os.environ.pop("RCM_MC_LEIE_CSV", None)
        from rcm_mc.npi_cleaner import compliance, deep_pipeline
        self._stub_cms = patch.object(
            compliance, "screen_cms",
            lambda npis, **kw: {"id": "pecos", "available": False,
                                "checked": 0, "note": "stubbed (test)"})
        self._stub_cms.start()
        self._no_probe = patch.object(
            deep_pipeline, "_default_probe", lambda timeout=4.0: False)
        self._no_probe.start()

    def tearDown(self):
        self._no_probe.stop()
        self._stub_cms.stop()
        self._env.stop()

    def test_deep_without_enrich_runs_leie_with_dollar_exposure(self):
        wd = _workdir()
        leie = Path(wd) / "UPDATED.csv"
        leie.write_text("LASTNAME,FIRSTNAME,NPI,EXCLTYPE,EXCLDATE\n"
                        f"DOE,JOHN,{GOOD_B},1128a1,2023-01-15\n")
        with patch.dict(os.environ, {"RCM_MC_LEIE_CSV": str(leie)}):
            data = ("BillingNPI,BilledAmount\n"
                    f"{GOOD_B},100.50\n{GOOD_B},49.50\n"
                    f"{GOOD},10\n").encode()
            res = engine.clean_bytes(data, "b.csv", deep=True)
        self.assertIsNotNone(res.compliance)     # used to stay None
        leie_s = next(c for c in res.compliance if c.get("id") == "oig_leie")
        self.assertTrue(leie_s["available"])
        self.assertEqual(leie_s["excluded"], 1)
        m = leie_s["matches"][0]
        self.assertEqual(m["npi"], GOOD_B)
        self.assertAlmostEqual(m["billed"], 150.0, places=2)
        self.assertAlmostEqual(leie_s["excluded_billed_total"], 150.0,
                               places=2)
        # Deep itself failed fast on the preflight, deterministic results
        # untouched — the honest offline story.
        self.assertFalse(res.deep.get("ok"))
        self.assertIn("could not reach", str(res.deep.get("error")).lower())

    def test_offline_run_still_skips_compliance(self):
        res = engine.clean_bytes(
            f"BillingNPI\n{GOOD}\n".encode(), "b.csv")
        self.assertIsNone(res.compliance)

    def test_stale_installed_leie_pack_warns_on_the_run(self):
        from rcm_mc.npi_cleaner import refdata_packs
        dbp = _tmp_pack_db()
        with patch.object(refdata_packs, "_DB_PATH", dbp):
            _reset_pack_cache()
            refdata_packs.install_from_bytes(
                "leie", f"NPI\n{GOOD_B}\n".encode(), source="test.csv")
            # Age the install past the 35-day cadence.
            with sqlite3.connect(dbp) as con:
                con.execute("UPDATE pack_meta SET fetched = ?",
                            (time.time() - 90 * 86400,))
            st = next(p for p in refdata_packs.status() if p["id"] == "leie")
            self.assertTrue(st.get("stale"))
            res = engine.clean_bytes(
                f"BillingNPI\n{GOOD_B}\n".encode(), "b.csv", deep=True)
        _reset_pack_cache()
        blob = " ".join(res.warnings)
        self.assertIn("LEIE reference pack is", blob)
        self.assertIn("Refresh the leie pack", blob)


class DollarWeightedVerifySeamTests(unittest.TestCase):
    """connectors-front P2 handoff: the engine sums billed dollars per
    billing NPI and the bridge spends its bounded verify cap on the
    highest-dollar NPIs (proven through the REAL enrich path with only the
    external NPPES lookup stubbed)."""

    def test_engine_passes_weights_and_cap_keeps_big_dollars(self):
        from rcm_mc.data_public import nppes_api_client
        rows = ["BillingNPI,BilledAmount"]
        npis = [f"19{i:08d}" for i in range(1, 42)]   # 41 distinct
        for i, npi in enumerate(npis):
            billed = "99999.00" if i == len(npis) - 1 else "1.00"
            rows.append(f"{npi},{billed}")
        data = ("\n".join(rows) + "\n").encode()
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}), \
                patch.object(nppes_api_client, "fetch_by_npi",
                             lambda npi: None):
            res = engine.clean_bytes(data, "w.csv", enrich=True)
        v = (res.nppes or {}).get("verify") or {}
        self.assertTrue(v.get("ranked_by_dollars"))
        self.assertTrue(v.get("capped"))
        recs = v.get("records") or {}
        # The last-seen NPI carries the dollars: it must be INSIDE the cap;
        # the tie-broken 40th $1 NPI is the one that falls out.
        self.assertIn(npis[-1], recs)
        self.assertNotIn(npis[39], recs)
        self.assertIn("by dollars", str(v.get("note") or "").lower())
        # Coverage counters (connectors front) survive the engine path.
        self.assertEqual(v.get("rows_seen"), 41)
        self.assertEqual(v.get("rows_covered"), 40)


class RenderingGrainSeamTests(unittest.TestCase):
    """pipeline-report handoff: the engine detects a rendering/servicing
    individual-NPI column and analytics grades coding intensity at that
    grain (provider_basis captions the difference)."""

    def _fixture(self, rendering: bool) -> bytes:
        head = ("ClaimID,RenderingNPI,BillingNPI,HCPCS" if rendering
                else "ClaimID,BillingNPI,HCPCS")
        rows = [head]
        for i in range(20):   # hot coder: all level 5
            rows.append((f"{i},{GOOD},{GOOD_B},99215" if rendering
                         else f"{i},{GOOD_B},99215"))
        for i in range(20, 40):   # cold coder: all level 2
            rows.append((f"{i},{GOOD_B},{GOOD_B},99212" if rendering
                         else f"{i},{GOOD_B},99212"))
        return ("\n".join(rows) + "\n").encode()

    def test_rendering_column_switches_the_grain(self):
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            res = engine.clean_bytes(self._fixture(True), "r.csv")
        ci = (res.population or {}).get("coding_intensity") or {}
        self.assertEqual(ci.get("provider_basis"), "rendering")
        hot = [o["npi"] for o in ci.get("outliers") or []]
        self.assertIn(GOOD, hot)

    def test_billing_fallback_keeps_old_basis(self):
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            res = engine.clean_bytes(self._fixture(False), "r.csv")
        ci = (res.population or {}).get("coding_intensity") or {}
        self.assertEqual(ci.get("provider_basis"), "billing")


class ScorecardCanonicalKeysTests(unittest.TestCase):
    """One canonical name per concept across fronts: catalog + connector_plan
    (with the connectors front's `state` field) ride EVERY scorecard now,
    offline included, and the whole payload stays JSON-serializable."""

    def test_offline_scorecard_carries_catalog_and_plan_states(self):
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            res = engine.clean_bytes(
                f"BillingNPI,Dx\n{GOOD},E11.9\n".encode(), "s.csv")
        sc = res.as_scorecard()
        self.assertTrue(sc["catalog"])
        self.assertTrue(sc["connector_plan"])
        for row in sc["connector_plan"]:
            for key in ("id", "name", "applies", "mode", "reason", "state"):
                self.assertIn(key, row)
        json.dumps(sc)   # must round-trip

    def test_streamed_run_carries_plan_catalog_and_merged_carcs(self):
        from rcm_mc.npi_cleaner import bigfile
        d = tempfile.mkdtemp(prefix="npi-seams-big-")
        p = Path(d) / "big.csv"
        lines = ["BillingNPI,BilledAmount,DenialCode"]
        for i in range(200):
            lines.append(f"{GOOD},{100 + i}.25,CO-45")
        p.write_text("\n".join(lines) + "\n")
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}), \
                patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 512), \
                patch.object(bigfile, "CHUNK_TARGET_BYTES", 2048):
            res = bigfile.clean_path(str(p), "big.csv")
        self.assertTrue(res.connector_plan)     # first-chunk plan carried
        self.assertTrue(res.catalog)
        self.assertIsNone(res.sanity.get("carc-invalid"))
        top = {e["code"]: e["count"] for e in res.denials["top"]}
        self.assertEqual(top.get("45"), 200)    # merged bare across chunks
        json.dumps(res.as_scorecard())


class AutofileIntegrationGapSeamTests(unittest.TestCase):
    """connectors-front P2 handoff: missing data / missing egress now land
    on the wishlist beside missing parsers, through the REAL run path."""

    def setUp(self):
        self.wd = _workdir()
        self._env = patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": self.wd})
        self._env.start()
        os.environ.pop("RCM_MC_LEIE_CSV", None)
        from rcm_mc.npi_cleaner import compliance, deep_pipeline, \
            refdata_packs
        self._stub_cms = patch.object(
            compliance, "screen_cms",
            lambda npis, **kw: {"id": "pecos", "available": False,
                                "checked": 0, "note": "stubbed (test)"})
        self._stub_cms.start()
        self._no_probe = patch.object(
            deep_pipeline, "_default_probe", lambda timeout=4.0: False)
        self._no_probe.start()
        self._db = patch.object(refdata_packs, "_DB_PATH", _tmp_pack_db())
        self._db.start()
        _reset_pack_cache()

    def tearDown(self):
        self._db.stop()
        _reset_pack_cache()
        self._no_probe.stop()
        self._stub_cms.stop()
        self._env.stop()

    def _titles(self):
        from rcm_mc.npi_cleaner import wishlist
        return [r["title"] for r in wishlist.list_requests()
                if r["source"] == "auto"]

    def test_leie_without_data_and_uninstalled_pack_file_gaps(self):
        data = (f"BillingNPI,Dx\n{GOOD},E11.9\n{GOOD_B},I10\n").encode()
        engine.clean_bytes(data, "g.csv", deep=True)
        titles = self._titles()
        self.assertTrue(any("LEIE exclusion screen ran without data" in t
                            for t in titles), titles)
        self.assertTrue(any("ICD-10-CM code set" in t
                            and "not installed" in t for t in titles),
                        titles)
        self.assertTrue(any(t.startswith("Deep recovery timed out")
                            for t in titles), titles)

    def test_no_gap_when_leie_data_is_loaded(self):
        leie = Path(self.wd) / "UPDATED.csv"
        leie.write_text(f"LASTNAME,FIRSTNAME,NPI\nDOE,JOHN,{GOOD_B}\n")
        with patch.dict(os.environ, {"RCM_MC_LEIE_CSV": str(leie)}):
            engine.clean_bytes(f"BillingNPI\n{GOOD}\n".encode(),
                               "g.csv", deep=True)
        self.assertFalse(any("LEIE exclusion screen ran without data" in t
                             for t in self._titles()))


class OfflinePackInstallSeamTests(unittest.TestCase):
    """connectors-front CLI handoff: --refdata-install gives a no-egress box
    a working pack path, and the installed pack immediately powers the
    engine's offline exclusion flag and flips the plan row to ready."""

    def setUp(self):
        from rcm_mc.npi_cleaner import refdata_packs
        self._db = patch.object(refdata_packs, "_DB_PATH", _tmp_pack_db())
        self._db.start()
        _reset_pack_cache()
        self._env = patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()})
        self._env.start()
        os.environ.pop("RCM_MC_LEIE_CSV", None)

    def tearDown(self):
        self._env.stop()
        self._db.stop()
        _reset_pack_cache()

    def test_cli_install_leie_lights_the_offline_screen(self):
        from rcm_mc.npi_cleaner import cli, connectors, refdata_packs
        src = Path(tempfile.mkdtemp(prefix="npi-seams-leie-")) / "UPDATED.csv"
        src.write_text(f"LASTNAME,FIRSTNAME,NPI\nDOE,JANE,{GOOD}\n")
        rc = cli.main(["--refdata-install", "leie", str(src)])
        self.assertEqual(rc, 0)
        st = next(p for p in refdata_packs.status() if p["id"] == "leie")
        self.assertTrue(st["installed"])
        self.assertIn("file:", str(st.get("source") or ""))
        # Engine flag fires offline from the installed pack.
        res = engine.clean_bytes(
            f"BillingNPI\n{GOOD}\n{GOOD_B}\n".encode(), "l.csv")
        self.assertEqual(res.sanity.get("leie-excluded-npi"), 1)
        # And the plan row agrees it is ready now.
        plan = connectors.plan({"has_npi": True, "has_billing": True,
                                "blank_npi_pct": 0.0})
        leie_row = next(p for p in plan if p["id"] == "oig_leie")
        self.assertEqual(leie_row["state"], "ready")

    def test_cli_install_rejects_unknown_pack(self):
        from rcm_mc.npi_cleaner import cli
        src = Path(tempfile.mkdtemp(prefix="npi-seams-bad-")) / "x.csv"
        src.write_text("NPI\n1234567893\n")
        self.assertEqual(cli.main(["--refdata-install", "nope", str(src)]), 1)


class _ResetPackCache:
    """Context manager: drop the refdata_packs memo on enter AND exit so a
    swapped _DB_PATH is actually consulted (and no stale memo leaks out)."""

    def __enter__(self):
        _reset_pack_cache()
        return self

    def __exit__(self, *exc):
        _reset_pack_cache()
        return False


class ReportRenderingSeamTests(unittest.TestCase):
    """connectors-front P1 handoff: the exec one-pager and the workbook
    surface the coverage counters, plan states and reference-source health
    the connectors front put on the scorecard."""

    def _enriched_result(self):
        from rcm_mc.data_public import nppes_api_client
        data = (f"BillingNPI,BilledAmount,Dx\n"
                f"{GOOD},100.00,E11.9\n{GOOD_B},50.00,I10\n").encode()
        from rcm_mc.npi_cleaner import compliance, refdata_packs
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}), \
                patch.object(refdata_packs, "_DB_PATH", _tmp_pack_db()), \
                _ResetPackCache(), \
                patch.object(nppes_api_client, "fetch_by_npi",
                             lambda npi: None), \
                patch.object(compliance, "screen_cms",
                             lambda npis, **kw: {"id": "pecos",
                                                 "available": False,
                                                 "checked": 0,
                                                 "note": "stubbed"}):
            os.environ.pop("RCM_MC_LEIE_CSV", None)
            return engine.clean_bytes(data, "e.csv", enrich=True)

    def test_exec_report_online_screens_section(self):
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        res = self._enriched_result()
        html = build_exec_report(res.as_scorecard(), "e.csv",
                                 "2026-07-07T00:00:00+00:00")
        self.assertIn("Online screens", html)
        # LEIE has no data loaded → it must show as waiting on data.
        self.assertIn("Screen waiting on data", html)
        self.assertIn("OIG LEIE exclusions", html)
        # Reference-source badges ride advanced.reference_status when the
        # vendored adapter ran; guarded otherwise.
        if (res.advanced or {}).get("reference_status"):
            self.assertIn("Reference sources", html)

    def test_workbook_gains_connectors_sheet(self):
        import io as _io
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl unavailable")
        from rcm_mc.npi_cleaner.report import build_workbook
        res = self._enriched_result()
        wb_bytes = build_workbook(res, res.headers,
                                  [[GOOD, "100.00", "E11.9"]])
        wb = load_workbook(_io.BytesIO(wb_bytes))
        self.assertIn("Connectors", wb.sheetnames)
        ws = wb["Connectors"]
        text = " ".join(str(c.value) for row in ws.iter_rows()
                        for c in row if c.value is not None)
        self.assertIn("Applicable source", text)
        self.assertIn("needs_data", text)

    def test_offline_workbook_still_gets_plan_sheet(self):
        import io as _io
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl unavailable")
        from rcm_mc.npi_cleaner.report import build_workbook
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            res = engine.clean_bytes(
                f"BillingNPI\n{GOOD}\n".encode(), "o.csv")
        wb = load_workbook(_io.BytesIO(
            build_workbook(res, res.headers, [[GOOD]])))
        self.assertIn("Connectors", wb.sheetnames)


class JcodeCellGrainSeamTests(unittest.TestCase):
    """Verify-connectors note: rows_seen should mean CELLS for J-codes the
    way it does for NDCs — the engine now hands resolve_drugs the per-cell
    J-code list (resolve_drugs dedupes before any lookup)."""

    def test_jcode_stats_returns_cells_not_distinct(self):
        rows = [["J1745"], ["J1745"], ["99213"], ["J9299"], [""]]
        pct, jc = engine._jcode_stats(rows, 0)
        self.assertAlmostEqual(pct, 75.0, places=1)
        self.assertEqual(jc, ["J1745", "J1745", "J9299"])   # cell grain
        self.assertEqual(set(jc), {"J1745", "J9299"})       # same universe

    def test_distinct_cap_still_bounds_the_universe(self):
        rows = [[f"J{1000 + i}"] for i in range(30)]
        _, jc = engine._jcode_stats(rows, 0, distinct_cap=5)
        self.assertEqual(len(set(jc)), 5)


class WishlistShippedTransitionTests(unittest.TestCase):
    """pipeline-report handoff: mark_shipped is the improvement loop's
    'shipped' transition — prove the exact call sites the integrator runs
    (space-grouped EU money + the no-NPI-column sniff) move only their own
    open auto entries."""

    def test_mark_shipped_moves_only_matching_auto_entries(self):
        from rcm_mc.npi_cleaner import wishlist
        with patch.dict(os.environ, {"RCM_MC_NPI_WORKDIR": _workdir()}):
            wishlist.auto_file(
                "format",
                "European amounts with space-grouped thousands (1 234,56)",
                "seen in engine survey")
            wishlist.auto_file(
                "field", "Detector found no NPI column in an upload",
                "seen in engine survey")
            wishlist.auto_file(
                "format", "Clean .xlsx members inside zip batches",
                "still open")
            self.assertEqual(
                wishlist.mark_shipped("space-grouped thousands"), 1)
            self.assertEqual(wishlist.mark_shipped("no NPI column"), 1)
            open_titles = [r["title"] for r in wishlist.list_requests("open")]
            shipped = [r["title"] for r in wishlist.list_requests("shipped")]
            self.assertIn("Clean .xlsx members inside zip batches",
                          open_titles)
            self.assertEqual(len(shipped), 2)


if __name__ == "__main__":   # pragma: no cover
    unittest.main()
