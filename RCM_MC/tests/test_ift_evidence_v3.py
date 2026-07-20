"""Invariants of the committed IFT Sourced Evidence Master v3 deliverable.

Validates the shipped workbook without regenerating it (no network, no
LibreOffice): structure, governance integrity, the no-illustrative rule, and
the deliverable gates (tabs >= 200, printed-page estimate >= 200, >= 29MB).
"""
import json
import os
import re
import unittest

RCM_MC = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DELIV = os.path.join(RCM_MC, 'deliverables', 'IFT_Sourced_Evidence_Master_v4_3.xlsx')
PIPE = os.path.join(RCM_MC, 'scripts', 'ift_evidence_v3')
CACHE = os.path.join(RCM_MC, 'rcm_mc', 'market_reports', 'reference', 'ift_v3_cache')

try:
    import openpyxl  # noqa: F401
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


@unittest.skipUnless(HAVE_OPENPYXL, 'openpyxl not installed')
@unittest.skipUnless(os.path.exists(DELIV), 'deliverable not present')
class TestIFTEvidenceV3(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        from openpyxl import load_workbook
        cls.wb = load_workbook(DELIV, read_only=False)

    def test_deliverable_gates(self):
        self.assertGreaterEqual(len(self.wb.sheetnames), 200)
        self.assertGreaterEqual(os.path.getsize(DELIV), 29_000_000)

    def test_governance_tabs_present(self):
        for tab in ('README', 'Methodology', 'Fact_Ledger', 'Source_Register',
                    'Source_Index', 'Verification_Log', 'V3_Change_Log',
                    'Pull_Manifest', 'Excluded_Not_Sourced'):
            self.assertIn(tab, self.wb.sheetnames)

    def test_fact_ids_contiguous(self):
        fids = []
        for row in self.wb['Fact_Ledger'].iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str) and re.fullmatch(r'F\d+', v):
                fids.append(int(v[1:]))
        self.assertGreaterEqual(max(fids), 460)
        missing = set(range(1, max(fids) + 1)) - set(fids)
        self.assertEqual(missing, set(), f'missing fact IDs: {sorted(missing)[:10]}')

    def test_source_ids_contiguous(self):
        sids = []
        for row in self.wb['Source_Index'].iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str) and re.fullmatch(r'S\d+', v):
                sids.append(int(v[1:]))
        self.assertGreaterEqual(max(sids), 313)
        missing = set(range(1, max(sids) + 1)) - set(sids)
        self.assertEqual(missing, set(), f'missing source IDs: {sorted(missing)[:10]}')

    def test_pull_manifest_matches_cache(self):
        man_path = os.path.join(CACHE, 'manifest.json')
        if not os.path.exists(man_path):
            self.skipTest('cache not present')
        man = json.load(open(man_path))
        cells = set()
        for row in self.wb['Pull_Manifest'].iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str):
                cells.add(v)
        missing = [k for k in man if k not in cells]
        self.assertEqual(missing, [], f'manifest keys missing from tab: {missing[:5]}')

    def test_no_banned_illustrative_figures(self):
        banned = [re.compile(p, re.I) for p in
                  (r'\$6\.5\s*B', r'\$18[-–]22\s*B', r'165\.8')]
        hits = []
        for name in self.wb.sheetnames:
            if name == 'Excluded_Not_Sourced':
                continue
            for row in self.wb[name].iter_rows():
                for c in row:
                    if isinstance(c.value, str):
                        for rx in banned:
                            if rx.search(c.value):
                                hits.append(f'{name}!{c.coordinate}')
        self.assertEqual(hits, [], f'banned figures outside quarantine: {hits[:5]}')

    def test_charts_present(self):
        n = sum(len(self.wb[s]._charts) for s in self.wb.sheetnames)
        self.assertGreaterEqual(n, 150)

    def test_chart_house_style(self):
        """v3.3 chart regression guards: no combo/secondary-axis charts,
        category axes on the bottom, no smoothed line series."""
        import zipfile
        import xml.etree.ElementTree as ET
        C = '{http://schemas.openxmlformats.org/drawingml/2006/chart}'
        z = zipfile.ZipFile(DELIV)
        defects = []
        n_charts = 0
        for name in z.namelist():
            if not re.fullmatch(r'xl/charts/chart\d+\.xml', name):
                continue
            n_charts += 1
            plot = ET.fromstring(z.read(name)).find(f'.//{C}plotArea')
            n_val = len(plot.findall(f'{C}valAx'))
            n_plots = sum(len(plot.findall(f'{C}{t}'))
                          for t in ('lineChart', 'barChart', 'areaChart'))
            if n_val > 1 or n_plots > 1:
                defects.append(f'{name}: combo/secondary axis')
            horiz = any(bd.get('val') == 'bar'
                        for bc in plot.findall(f'{C}barChart')
                        for bd in bc.findall(f'{C}barDir'))
            want_cat = 'l' if horiz else 'b'
            for ax in plot.findall(f'{C}catAx'):
                pos = ax.find(f'{C}axPos')
                if pos is None or pos.get('val') != want_cat:
                    defects.append(f'{name}: catAx not on {want_cat}')
                dele = ax.find(f'{C}delete')
                if dele is None or dele.get('val') not in ('0', 'false'):
                    defects.append(f'{name}: catAx delete not explicit 0')
            for lc in plot.findall(f'{C}lineChart'):
                for ser in lc.findall(f'{C}ser'):
                    sm = ser.find(f'{C}smooth')
                    if sm is None or sm.get('val') not in ('0', 'false'):
                        defects.append(f'{name}: smoothed line series')
        self.assertGreaterEqual(n_charts, 150)
        self.assertEqual(defects, [], f'chart defects: {defects[:8]}')

    def test_methodology_rebuilt(self):
        texts = []
        for row in self.wb['Methodology'].iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    texts.append(c.value)
        blob = '\n'.join(texts)
        self.assertIn('9. How to verify any number in sixty seconds', blob)
        self.assertIn('data-api/v1/', blob)
        self.assertNotIn('data-api/revision 1/', blob)

    def test_granular_families_present(self):
        names = set(self.wb.sheetnames)
        for y in range(2010, 2025):
            self.assertIn(f'PSPS_Detail_{y}', names)
        for y in range(2013, 2025):
            self.assertIn(f'MUP_State_{y}', names)
        for y in range(2020, 2026):
            self.assertIn(f'MS_County_{y}', names)
        for tab in ('PECOS_Registry', 'Hosp_Registry', 'SNF_Registry',
                    'HSA_Hospital_Catchment', 'ED_Timeliness_Registry', 'SP_Index',
                    'Index', 'State_Age_65plus', 'OEWS_EMS_Wages',
                    'MUP_Providers_2024', 'HSA_Corridors', 'County_Age_65plus',
                    'PLACES_County_Chronic', 'QCEW_Quarterly',
                    'HCRIS_Hospital_Panel', 'LEIE_Ambulance_Exclusions'):
            self.assertIn(tab, names)

    def test_v34_analysis_tabs_present(self):
        """v3.4 specificity-and-analysis pass: the facility-pay layer, the
        subject-company book, the A-part analysis tabs, the cohort program,
        the C-part assembly, and the run governance."""
        names = set(self.wb.sheetnames)
        for tab in ('Facility_Pay_Layer', 'MMT_Medicare_Book',
                    'Market_Share_Panels', 'Fragmentation_National',
                    'Insourcing_Bounds', 'HCRIS_Ambulance_CostCenters',
                    'Cohort_Corridors', 'Hub_Spoke_Map', 'Contract_Corpus',
                    'Cohort_990_Contractors', 'System_Research_Cohort',
                    'Footprint_Determination', 'Prospect_Landscape',
                    'County_Whitespace_Screens', 'Growth_Decomposition',
                    'Denial_Economics', 'Transfer_Delay_Burden',
                    'Workforce_Depth', 'Universe_Reconciliation',
                    'LEIE_Read_Panel', 'Throughput_Economics_Public',
                    'GAO_OIG_Shelf', 'REH_Closure_Flow', 'Medicaid_Rate_Card',
                    'RSNAT_Series', 'MA_Book_Calibrator',
                    'Entry_Barrier_Register', 'Balance_Billing_States',
                    'Receiving_Center_Registry', 'Federal_Ambulance_Contracts',
                    'Annual_Market_Structure', 'Realized_Price_Ladders',
                    'Registered_vs_Billing', 'Metro_TAM_Panels',
                    'TAM_Assembly_State', 'Scenario_Matrix',
                    'Growth_Outlook_Shell', 'Vendor_Share_Stack',
                    'Stickiness_Evidence', 'Investor_QA', 'Slide_Feed',
                    'Refresh_Calendar', 'Run_Log',
                    'Claims_Vendor_Recv', 'Commercial_Rate_MRF'):
            self.assertIn(tab, names)

    def test_findings_extended_past_51(self):
        """v3.4 appended findings 52 onward with live references."""
        ws = self.wb['Findings']
        ids = []
        for row in ws.iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str) and v.strip().isdigit():
                ids.append(int(v.strip()))
        self.assertGreaterEqual(max(ids), 80)

    def test_v35_completion_tabs_present(self):
        """v3.5 completion pass closed the deferred v3.4 PENDING items from
        public sources: the footprint-wide 990 sweep, the SNF return-leg
        quality layer, and the CMS Hospital-at-Home participant list."""
        names = set(self.wb.sheetnames)
        for tab in ('Footprint_990_Sweep', 'SNF_Return_Leg_Quality',
                    'Hospital_at_Home_Participants'):
            self.assertIn(tab, names)

    def test_v35_findings_and_ledgers(self):
        """v3.5 added three live-referenced findings and extended the ledgers
        contiguously past the v3.4 maxima (F585 / S409)."""
        ws = self.wb['Findings']
        ids = [int(r[0].value.strip())
               for r in ws.iter_rows(min_col=1, max_col=1)
               if isinstance(r[0].value, str) and r[0].value.strip().isdigit()]
        self.assertGreaterEqual(max(ids), 102)
        fids = [int(r[0].value[1:])
                for r in self.wb['Fact_Ledger'].iter_rows(min_col=1, max_col=1)
                if isinstance(r[0].value, str) and re.fullmatch(r'F\d+', r[0].value)]
        sids = [int(r[0].value[1:])
                for r in self.wb['Source_Index'].iter_rows(min_col=1, max_col=1)
                if isinstance(r[0].value, str) and re.fullmatch(r'S\d+', r[0].value)]
        self.assertGreaterEqual(max(fids), 596)
        self.assertGreaterEqual(max(sids), 417)

    def test_v36_returnleg_structure_tab(self):
        """v3.6 return-leg structure pass: the SNF QRP claims measures joined
        to the CMS Nursing Home provider-info file on the CCN, cross-tabbed by
        ownership, scale and star rating."""
        self.assertIn('SNF_ReturnLeg_Structure', self.wb.sheetnames)
        ids = [int(r[0].value.strip())
               for r in self.wb['Findings'].iter_rows(min_col=1, max_col=1)
               if isinstance(r[0].value, str) and r[0].value.strip().isdigit()]
        self.assertGreaterEqual(max(ids), 103)

    def test_v43_fleet_license_tabs(self):
        """v4.3 (Run 7): the fleet-license identification pass - the license-
        object route map, the 51-jurisdiction route matrix, and the public
        fleet/operator counts (NPPES operator floor, MO open-data services, NJ
        statewide vehicle anchor)."""
        for tab in ('Fleet_License_Route_Map', 'Fleet_License_State_Matrix',
                    'Fleet_Size_Evidence'):
            self.assertIn(tab, self.wb.sheetnames)
        # the state matrix carries all 51 jurisdictions
        col_a = [r[0].value for r in
                 self.wb['Fleet_License_State_Matrix'].iter_rows(min_col=1,
                                                                 max_col=1)]
        juris = [v for v in col_a if isinstance(v, str)
                 and re.search(r'\(([A-Z]{2})\)$', v)]
        self.assertGreaterEqual(len(juris), 51)
        # findings continue past the v3.6 tail
        ids = [int(r[0].value.strip())
               for r in self.wb['Findings'].iter_rows(min_col=1, max_col=1)
               if isinstance(r[0].value, str) and r[0].value.strip().isdigit()]
        self.assertGreaterEqual(max(ids), 118)

    def test_v43_scale_predictor_tabs(self):
        """v4.3 (Run 7, second phase): the corporate-family resolution and
        scale-predictor tabs - why the two national players (GMR/Priority) are
        undercounted and which public signal best predicts real volume. The
        GMR family Medicare volume must be re-derived (nonzero) and the
        findings tail must reach 125."""
        for tab in ('Corporate_Family_Resolution', 'Fleet_Scale_Predictors'):
            self.assertIn(tab, self.wb.sheetnames)
        # the family-resolution tab carries a computed GMR Medicare volume
        vals = [c.value for row in self.wb['Corporate_Family_Resolution']
                .iter_rows() for c in row]
        nums = [v for v in vals if isinstance(v, (int, float)) and v > 100000]
        self.assertTrue(nums, 'expected a six-figure resolved Medicare volume')
        ids = [int(r[0].value.strip())
               for r in self.wb['Findings'].iter_rows(min_col=1, max_col=1)
               if isinstance(r[0].value, str) and r[0].value.strip().isdigit()]
        self.assertGreaterEqual(max(ids), 125)

    def test_leak_check_clean(self):
        """The v3.4 firewall leak check ran and found no violations."""
        p = os.path.join(PIPE, 'leak_check.json')
        if not os.path.exists(p):
            self.skipTest('leak_check.json not present')
        r = json.load(open(p))
        self.assertEqual(r.get('customer_account_violations'), [])
        self.assertEqual(r.get('survey_statistic_violations'), [])
        self.assertEqual(r.get('em_dash_violations'), [])

    def test_state_profiles_are_formula_driven(self):
        ws = self.wb['SP_TX']
        n_formula = sum(1 for row in ws.iter_rows() for c in row
                        if isinstance(c.value, str) and c.value.startswith('='))
        self.assertGreaterEqual(n_formula, 15)


@unittest.skipUnless(os.path.exists(PIPE), 'pipeline not present')
class TestPipelineArtifacts(unittest.TestCase):

    def test_pipeline_files_present(self):
        for f in ('assemble.py', 'pull.py', 'verify.py', 'v3lib.py',
                  'copy_engine.py', 'corrections.py', 'README.md'):
            self.assertTrue(os.path.exists(os.path.join(PIPE, f)), f)

    def test_verify_results_all_green(self):
        p = os.path.join(PIPE, 'verify_results.json')
        self.assertTrue(os.path.exists(p))
        r = json.load(open(p))
        self.assertEqual(r.get('n_errors'), 0)
        self.assertEqual(r.get('copy_diffs'), 0)
        self.assertEqual(r.get('n_mismatch'), 0)
        self.assertEqual(r.get('chart_defects', 0), 0)
        self.assertGreaterEqual(r.get('pages', 0), 200)


if __name__ == '__main__':
    unittest.main()
