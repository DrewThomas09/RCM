"""Tests for the NPI Claims Cleaner (/npi-cleaner).

Covers the stdlib engine (Luhn verdicts, dedup, trimming, missing-billing-NPI
detection) and the full HTTP loop: page render, raw-body upload, status poll,
and cleaned-file download.
"""
from __future__ import annotations

import json
import os
import tempfile
import time
import unittest
import urllib.request as _u
import urllib.parse as _up
import urllib.error as _ue
from unittest.mock import patch

from rcm_mc.npi_cleaner import engine


# Two NPIs that pass the real CMS/NPPES Luhn check (80840 + first 9).
GOOD_A = "1234567893"
GOOD_B = "1679576722"


class TestEngine(unittest.TestCase):
    def test_luhn_matches_cms_rule(self):
        self.assertTrue(engine.luhn_npi_valid(GOOD_A))
        self.assertTrue(engine.luhn_npi_valid(GOOD_B))
        self.assertFalse(engine.luhn_npi_valid("1234567890"))  # bad check digit
        self.assertFalse(engine.luhn_npi_valid("123"))          # too short

    def test_classify(self):
        self.assertEqual(engine.classify_npi(""), "blank")
        self.assertEqual(engine.classify_npi("  "), "blank")
        self.assertEqual(engine.classify_npi("99999"), "malformed")
        self.assertEqual(engine.classify_npi("1234567890"), "checksum")
        self.assertEqual(engine.classify_npi(GOOD_A), "valid")

    def test_clean_bytes_full(self):
        data = (
            "ClaimID,BillingProviderNPI,RenderingNPI\n"
            f"1, {GOOD_A} ,{GOOD_B}\n"       # leading/trailing space → trimmed
            f"2,{GOOD_A},{GOOD_B}\n"          # distinct row (different ClaimID)
            f"2,{GOOD_A},{GOOD_B}\n"          # exact duplicate of row above
            f"3,,{GOOD_B}\n"                  # blank billing NPI
            f"4,99999,{GOOD_B}\n"             # malformed
            "5,1234567890," + GOOD_B + "\n"   # checksum fail
        ).encode()
        res = engine.clean_bytes(data, "sample claims.csv")
        sc = res.as_scorecard()

        self.assertEqual(sc["rows_in"], 6)
        self.assertEqual(sc["duplicates_removed"], 1)
        self.assertEqual(sc["rows_out"], 5)
        self.assertEqual(sc["cells_trimmed"], 1)
        self.assertEqual(sc["billing_column"], "BillingProviderNPI")
        self.assertIn("RenderingNPI", sc["npi_columns"])

        bill = sc["column_stats"]["BillingProviderNPI"]
        # rows 1,2,2(dup counted before dedup)=3 valid, plus blank/malformed/checksum
        self.assertEqual(bill["blank"], 1)
        self.assertEqual(bill["malformed"], 1)
        self.assertEqual(bill["checksum"], 1)
        self.assertEqual(sc["billing_issues"], 3)
        self.assertEqual(sc["out_name"], "sample_claims_cleaned.csv")

        # The cleaned file exists, keeps the header, and trims whitespace.
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("ClaimID,BillingProviderNPI,RenderingNPI", out)
        self.assertNotIn(f" {GOOD_A} ", out)

    def test_no_npi_column_still_cleans(self):
        data = b"a,b\n1, 2 \n1, 2 \n"
        res = engine.clean_bytes(data, "x.csv")
        sc = res.as_scorecard()
        self.assertEqual(sc["npi_columns"], [])
        self.assertEqual(sc["duplicates_removed"], 1)
        self.assertTrue(any("No NPI column" in w for w in sc["warnings"]))

    def test_tsv_delimiter(self):
        data = b"NPI\tAmount\n" + GOOD_A.encode() + b"\t10\n"
        res = engine.clean_bytes(data, "t.tsv")
        self.assertEqual(res.as_scorecard()["delimiter"], "tab")

    def test_normalization_fixes(self):
        # A deliberately messy file exercising each normalizer.
        data = (
            "ClaimID,BillingNPI,ProviderState,AllowedAmt,DateOfService,PatientZip,HCPCS\n"
            "1,'1234567893,Ohio,\"$1,234.50\",2024-01-15,1234,99213\n"
            "2,1234567893.0,TX,\"(50.00)\",45300,08540,g0008\n"
            "3,N/A,texas,\"1,000\",01/15/2024,00501,99214\n"
        ).encode()
        res = engine.clean_bytes(data, "messy.csv")
        rp = res.repairs
        self.assertGreater(res.as_scorecard()["repairs_total"], 5)
        self.assertIn("state-name-to-code", rp)        # Ohio→OH, texas→TX
        self.assertIn("npi-excel-float", rp)           # 1234567893.0
        self.assertIn("leading-apostrophe", rp)        # '1234567893
        self.assertIn("null-token", rp)                # N/A → blank
        self.assertIn("date-excel-serial", rp)         # 45300
        self.assertIn("date-us-to-iso", rp)            # 01/15/2024
        self.assertIn("hcpcs-upper", rp)               # g0008 → G0008
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("OH", out)
        self.assertIn("2024-01-15", out)
        # Negative money is NOT apostrophe-defanged (it is a plain number).
        self.assertIn("-50.00", out)
        self.assertNotIn("'-50.00", out)

    def test_sex_and_icd_normalization(self):
        data = ("ClaimID,PatientSex,Diagnosis\n"
                "1,Male,E1165\n2,F,I10\n3,2,e119\n").encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertIn("sex-normalize", res.repairs)
        self.assertIn("dx-decimal", res.repairs)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("E11.65", out)   # E1165 → decimal inserted
        self.assertIn("I10", out)      # 3-char code unchanged
        self.assertIn("1,M,", out)     # Male → M

    def test_modifier_phone_taxonomy_normalization(self):
        data = ("ClaimID,Modifiers,ProviderPhone,ProviderTaxonomy\n"
                "1,\"26|tc, 59\",5551234567,207q00000x\n"
                "2,25;25,1-555-987-6543,208d00000x\n").encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertIn("modifier-normalize", res.repairs)
        self.assertIn("phone-format", res.repairs)
        self.assertIn("taxonomy-upper", res.repairs)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("26,TC,59", out)          # split/upper/order
        self.assertIn("(555) 123-4567", out)     # 10-digit format
        self.assertIn("(555) 987-6543", out)     # 11-digit w/ leading 1
        self.assertIn("207Q00000X", out)         # taxonomy upper
        self.assertIn("2,25,", out)              # dedup 25;25 → 25

    def test_cross_field_sanity_flags(self):
        data = ("ClaimID,ChargeAmt,AllowedAmt,PaidAmt,Units\n"
                "1,50,90,40,2\n"       # allowed > billed
                "2,200,150,300,1\n"    # paid > allowed
                "3,100,80,-5,0\n"      # negative paid + units<=0
                "4,100,80,60,1.5\n").encode()  # fractional units
        res = engine.clean_bytes(data, "x.csv")
        sf = res.sanity
        self.assertEqual(sf.get("allowed-exceeds-billed"), 1)
        self.assertEqual(sf.get("paid-exceeds-allowed"), 1)
        self.assertEqual(sf.get("negative-paid"), 1)
        self.assertEqual(sf.get("nonpositive-units"), 1)
        self.assertEqual(sf.get("fractional-units"), 1)

    def test_ndc_11digit_normalization(self):
        # Segment-aware padding must match the vendored normalize_ndc11 rule.
        data = ("ClaimID,NDC\n"
                "1,1234-5678-90\n"    # 4-4-2 → pad seg1
                "2,12345-678-90\n"    # 5-3-2 → pad seg2
                "3,00069015001\n"     # already 11 digits → unchanged
                "4,1234567890\n").encode()  # 10-digit unhyphenated → ambiguous
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.repairs.get("ndc-pad-11"), 2)
        self.assertEqual(res.sanity.get("ndc-ambiguous-10digit"), 1)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("01234567890", out)  # 1234-5678-90 padded
        self.assertIn("12345067890", out)  # 12345-678-90 padded

    def test_ndc_matches_package_rule(self):
        try:
            from rcm_mc.npi_cleaner.vendor_v49.npi_recovery import field_validators as FV
        except Exception:
            self.skipTest("vendored field_validators unavailable")
        for c in ("1234-5678-90", "12345-678-90", "12345-6789-0",
                  "0069-0150-01", "00069015001"):
            mine, _ = engine._clean_ndc_cell(c)
            pkg, _st = FV.normalize_ndc11(c)
            self.assertEqual(mine, pkg, f"mismatch on {c}")

    def test_suspected_duplicate_claims(self):
        # Rows 1&2 share provider/patient/DOS/HCPCS/amount (diff ClaimID) → dup.
        # A 3rd row is an exact dup of row 1 (removed by exact dedup, not
        # double-counted). A 4th row differs only by patient → not a dup.
        data = (
            "ClaimID,BillingNPI,PatientID,DateOfService,HCPCS,AllowedAmt\n"
            f"1,{GOOD_B},P100,2024-03-01,99213,80\n"
            f"2,{GOOD_B},P100,2024-03-01,99213,80\n"
            f"1,{GOOD_B},P100,2024-03-01,99213,80\n"
            f"9,{GOOD_B},P200,2024-03-01,99213,80\n"
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.n_dupes_removed, 1)                    # exact dup
        self.assertEqual(res.sanity.get("suspected-duplicate-claim"), 1)

    def test_duplicate_claims_needs_key_columns(self):
        # Without provider/date/code columns the check must not run.
        data = ("ClaimID,Note\n1,a\n2,a\n").encode()
        res = engine.clean_bytes(data, "y.csv")
        self.assertNotIn("suspected-duplicate-claim", res.sanity)

    def test_future_date_flagged(self):
        # A service date or DOB after today is impossible → flagged. Uses a
        # far-future year so the assertion is stable regardless of run date.
        # A US-format future date must flag too (normalized to ISO first), and
        # a legitimately-future column (coverage end) must NOT flag.
        data = (
            "ClaimID,DateOfService,PatientDOB,CoverageEndDate\n"
            "1,2999-01-01,1980-05-14,2999-12-31\n"   # future DOS → flag
            "2,03/04/2999,1975-02-02,2999-12-31\n"   # future DOS (US fmt) → flag
            "3,2020-06-15,2999-05-05,2999-12-31\n"   # future DOB → flag
            "4,2021-01-01,1990-01-01,2999-12-31\n"   # all past svc/dob → clear
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        # Three rows have a future service/birth date; the coverage-end column
        # is in the future on every row but must never trigger the flag.
        self.assertEqual(res.sanity.get("date-in-future"), 3)

    def test_future_date_needs_service_or_dob_column(self):
        # Without a service/birth date column the check must not run even if a
        # generic future date is present.
        data = ("ClaimID,CoverageEndDate\n1,2999-01-01\n").encode()
        res = engine.clean_bytes(data, "y.csv")
        self.assertNotIn("date-in-future", res.sanity)

    def test_zip_state_mismatch_flagged(self):
        # A ZIP whose 3-digit prefix resolves to a different state than the
        # state cell is flagged. 902xx→CA, 331xx→FL, 100xx→NY.
        data = (
            "ClaimID,ProviderState,ProviderZip\n"
            "1,CA,90210\n"        # 902→CA, matches → clear
            "2,NY,90210\n"        # 902→CA but state says NY → mismatch
            "3,FL,33101\n"        # 331→FL, matches → clear
            "4,TX,10001\n"        # 100→NY but state says TX → mismatch
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("zip-state-mismatch"), 2)

    def test_zip_state_pairs_same_entity_only(self):
        # PatientState must not be compared against ProviderZip (different
        # entities) — that pairing would false-positive. Only same-entity
        # provider columns are compared, and here they agree → no flag.
        data = (
            "ClaimID,PatientState,ProviderState,ProviderZip\n"
            "1,NY,CA,90210\n"     # provider CA↔902 agree; patient NY ignored
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertNotIn("zip-state-mismatch", res.sanity)

    def test_zip_state_skips_territories(self):
        # Territory/military prefixes are approximate in the crosswalk, so a
        # PR/VI mismatch is never flagged.
        data = ("ClaimID,ProviderState,ProviderZip\n1,VI,00801\n").encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertNotIn("zip-state-mismatch", res.sanity)

    def test_hcpcs_malformed_flagged(self):
        # Valid: 5-digit CPT, letter+4 HCPCS, 4+letter Cat II; a valid code with
        # an appended modifier must NOT flag. Malformed: 4 digits, 5 letters.
        data = (
            "ClaimID,HCPCS\n"
            "1,99213\n"        # CPT Cat I → valid
            "2,J1885\n"        # HCPCS Level II → valid
            "3,0001F\n"        # CPT Cat II → valid
            "4,99213-25\n"     # valid + modifier → valid (modifier stripped)
            "5,9921\n"         # only 4 digits → malformed
            "6,ABCDE\n"        # 5 letters → malformed
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("hcpcs-malformed"), 2)

    def test_icd10_malformed_flagged(self):
        # Valid: E11.65, I10 (3-char), U07.1 (COVID). Malformed: leading digit,
        # too short.
        data = (
            "ClaimID,DiagnosisCode\n"
            "1,E11.65\n"       # valid
            "2,I10\n"          # valid 3-char
            "3,U07.1\n"        # valid
            "4,123.45\n"       # starts with a digit → malformed
            "5,E1\n"           # too short → malformed
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("icd10-malformed"), 2)

    def test_code_shape_checks_need_columns(self):
        # Without HCPCS/diagnosis columns the checks must not run.
        data = ("ClaimID,Note\n1,hello\n").encode()
        res = engine.clean_bytes(data, "y.csv")
        self.assertNotIn("hcpcs-malformed", res.sanity)
        self.assertNotIn("icd10-malformed", res.sanity)

    def test_money_unparseable_flagged(self):
        # An amount cell that survives cleaning but isn't a number is flagged;
        # normal / accounting-negative / $-and-comma values are not. A null
        # token ("N/A") is blanked by generic cleaning first, so it never flags.
        data = (
            "ClaimID,AllowedAmt\n"
            "1,100.00\n"       # clean number → ok
            "2,\"$1,250.50\"\n"  # $ + comma → parses → ok
            "3,(75.00)\n"      # accounting negative → parses → ok
            "4,N/A\n"          # null token → blanked → not flagged
            "5,pending\n"      # text in $ field → flagged
            "6,\"1,2OO\"\n"    # letter O instead of zero → flagged
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("money-unparseable"), 2)

    def test_money_unparseable_needs_money_column(self):
        data = ("ClaimID,Note\n1,hello world\n").encode()
        res = engine.clean_bytes(data, "y.csv")
        self.assertNotIn("money-unparseable", res.sanity)

    def test_sex_invalid_flagged(self):
        # Values that don't resolve to M/F/U via _clean_sex_cell are flagged;
        # mapped variants (Male, 2, blank) are not.
        data = (
            "ClaimID,PatientSex\n"
            "1,M\n"        # valid
            "2,Male\n"     # → M → valid
            "3,2\n"        # → F → valid
            "4,\n"         # blank → not flagged
            "5,X\n"        # unmapped → flagged
            "6,3\n"        # unmapped digit → flagged
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("sex-invalid"), 2)

    def test_taxonomy_malformed_flagged(self):
        # NUCC taxonomy must be 10 alphanumeric chars; anything else flags.
        data = (
            "ClaimID,ProviderTaxonomy\n"
            "1,207Q00000X\n"   # valid 10-char
            "2,2085R0202X\n"   # valid 10-char
            "3,207Q\n"         # too short → flagged
            "4,207Q00000XX\n"  # 11 chars → flagged
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("taxonomy-malformed"), 2)

    def test_value_domain_checks_need_columns(self):
        data = ("ClaimID,Note\n1,hi\n").encode()
        res = engine.clean_bytes(data, "y.csv")
        self.assertNotIn("sex-invalid", res.sanity)
        self.assertNotIn("taxonomy-malformed", res.sanity)

    def test_quality_report_card(self):
        # A clean file grades high; a dirty file grades lower, and every
        # dimension is a recomputable 0-100 ratio.
        clean = ("ClaimID,BillingNPI,AllowedAmt\n"
                 f"1,{GOOD_B},100.00\n2,{GOOD_B},110.00\n").encode()
        rc = engine.clean_bytes(clean, "clean.csv")
        qc = rc.as_scorecard()["quality"]
        self.assertGreaterEqual(qc["score"], 93)
        self.assertEqual(qc["letter"], "A")
        for dim in ("completeness", "validity", "consistency",
                    "uniqueness", "conformity"):
            self.assertIn(dim, qc["dimensions"])
        dirty = ("ClaimID,BillingNPI,AllowedAmt,PatientSex\n"
                 "1,123,garbage,X\n1,123,garbage,X\n2,,,\n").encode()
        rd = engine.clean_bytes(dirty, "dirty.csv")
        qd = rd.as_scorecard()["quality"]
        self.assertLess(qd["score"], qc["score"])

    def test_changelog_audit_trail(self):
        # Every change lands in the change-log CSV as before → after + rule.
        data = ("NPI,ChargeAmt,ServiceDate\n"
                f"  {GOOD_A}  ,\"$1,250.50\",03/04/2024\n").encode()
        res = engine.clean_bytes(data, "log.csv")
        self.assertGreaterEqual(res.n_changes, 3)
        self.assertIsNotNone(res.changelog_path)
        with open(res.changelog_path, encoding="utf-8") as fh:
            log = fh.read()
        self.assertIn("before", log.splitlines()[0])
        self.assertIn("1250.50", log)          # money after-value
        self.assertIn("2024-03-04", log)       # date after-value
        self.assertIn("date-us-to-iso", log)   # rule provenance

    def test_changelog_never_contains_phi(self):
        # De-id masking must NOT be recorded — originals would leak into the
        # log. Only the pre-deid cleaning changes appear.
        data = ("BillingNPI,PatientName,SSN\n"
                f"  {GOOD_B} ,John Q Public,123-45-6789\n").encode()
        res = engine.clean_bytes(data, "phi.csv", deid=True)
        if res.changelog_path:
            with open(res.changelog_path, encoding="utf-8") as fh:
                log = fh.read()
            self.assertNotIn("John Q Public", log)
            self.assertNotIn("123-45-6789", log)

    def test_payer_variant_clustering(self):
        data = ("ClaimID,Payer\n"
                "1,BCBS of Texas\n2,Blue Cross Blue Shield TX\n"
                "3,B.C.B.S.\n4,Aetna\n5,AETNA INC\n6,Cigna\n").encode()
        res = engine.clean_bytes(data, "p.csv")
        pv = res.payer_variants
        self.assertIsNotNone(pv)
        self.assertEqual(pv["column"], "Payer")
        self.assertEqual(pv["distinct_raw"], 6)
        canon = {c["canonical"]: c for c in pv["multi_spelling"]}
        self.assertIn("BLUE CROSS BLUE SHIELD", canon)
        self.assertEqual(canon["BLUE CROSS BLUE SHIELD"]["n_variants"], 3)
        self.assertIn("AETNA", canon)

    def test_chronology_flags(self):
        data = (
            "ClaimID,PatientDOB,DateOfService,AdmitDate,DischargeDate\n"
            "1,1990-01-01,1989-06-15,2024-03-01,2024-02-27\n"  # both wrong
            "2,1980-05-14,2024-03-01,2024-03-01,2024-03-04\n"  # both fine
        ).encode()
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("service-before-birth"), 1)
        self.assertEqual(res.sanity.get("discharge-before-admit"), 1)

    def test_paid_exceeds_billed_and_stale_date(self):
        data = ("ClaimID,BilledAmt,PaidAmt,DateOfService\n"
                "1,100,120,2024-03-01\n"      # paid > billed
                "2,100,80,2001-05-05\n").encode()  # stale DOS
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.sanity.get("paid-exceeds-billed"), 1)
        self.assertEqual(res.sanity.get("date-stale"), 1)

    def test_pos_and_revenue_code(self):
        data = ("ClaimID,POS,RevenueCode\n"
                "1,11,450\n"      # POS valid; revcode padded to 0450
                "2,1,0450\n"      # POS padded to 01 → valid
                "3,77,ABC\n").encode()  # POS invalid; revcode malformed
        res = engine.clean_bytes(data, "x.csv")
        self.assertEqual(res.repairs.get("revcode-pad"), 1)
        self.assertEqual(res.repairs.get("pos-pad"), 1)
        self.assertEqual(res.sanity.get("pos-invalid"), 1)
        self.assertEqual(res.sanity.get("revenue-code-malformed"), 1)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("0450", out)

    def test_charge_outliers_per_code(self):
        rows = [f"{i},{GOOD_B},99213,{100 + i}" for i in range(12)]
        rows.append(f"99,{GOOD_B},99213,9000")   # far-out charge
        data = ("ClaimID,BillingNPI,HCPCS,BilledAmt\n"
                + "\n".join(rows) + "\n").encode()
        res = engine.clean_bytes(data, "o.csv")
        self.assertEqual(res.sanity.get("charge-outlier"), 1)
        self.assertEqual(res.outliers[0]["code"], "99213")
        self.assertEqual(res.outliers[0]["outliers"], 1)

    def test_structure_findings(self):
        data = ("ClaimID,Note,Note,Empty\n"
                "1,a,b,\n2,c,d,\n").encode()
        res = engine.clean_bytes(data, "s.csv")
        self.assertEqual(res.structure.get("duplicate_headers"), ["Note"])
        self.assertEqual(res.structure.get("empty_columns"), ["Empty"])

    def test_discharge_date_not_money(self):
        # Regression: "DischargeDate" contains the substring "charge" and was
        # routed to the money cleaner — never date-normalized and false-flagged
        # as unparseable money. Date role must win.
        data = ("ClaimID,DischargeDate\n1,03/04/2024\n").encode()
        res = engine.clean_bytes(data, "d.csv")
        self.assertNotIn("money-unparseable", res.sanity)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("2024-03-04", out)   # date cleaner ran on it

    def test_column_fill_rates(self):
        # Per-column completeness profile: Note filled 1 of 4 rows = 25%.
        # Denominator is rows IN, so pct never exceeds 100 even when exact
        # duplicates are removed from the output.
        data = ("ClaimID,Note\n1,a\n2,\n3,\n4,\n").encode()
        res = engine.clean_bytes(data, "f.csv")
        fills = {f["column"]: f for f in res.column_fill}
        self.assertEqual(fills["ClaimID"]["pct"], 100.0)
        self.assertEqual(fills["Note"]["pct"], 25.0)
        dup = ("ClaimID,Note\n1,a\n1,a\n").encode()   # dedupe drops one row
        rd = engine.clean_bytes(dup, "d.csv")
        fd = {f["column"]: f for f in rd.column_fill}
        self.assertLessEqual(fd["Note"]["pct"], 100.0)

    def test_modifier_units_edits(self):
        # JW (discarded drug) with 0 units and bilateral 50 with 2 units are
        # both flagged; correct pairings are not.
        data = ("ClaimID,HCPCS,Modifier,Units\n"
                "1,J1885,JW,0\n"       # JW needs positive units → flag
                "2,J1885,JW,2\n"       # fine
                "3,99213,50,2\n"       # bilateral bills 1 unit → flag
                "4,99213,50,1\n"       # fine
                "5,99213,25,3\n").encode()  # unrelated modifier → fine
        res = engine.clean_bytes(data, "m.csv")
        self.assertEqual(res.sanity.get("jw-zero-units"), 1)
        self.assertEqual(res.sanity.get("bilateral-units"), 1)

    def test_workbook_quality_sheet(self):
        data = ("ClaimID,BillingNPI,AllowedAmt\n"
                f"1,{GOOD_B},100\n").encode()
        res = engine.clean_bytes(data, "q.csv")
        openpyxl = __import__("openpyxl")
        from io import BytesIO
        with open(res.workbook_path, "rb") as fh:
            wb = openpyxl.load_workbook(BytesIO(fh.read()))
        self.assertIn("Quality", wb.sheetnames)
        ws = wb["Quality"]
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        self.assertTrue(any("Overall grade" in c for c in cells))
        self.assertTrue(any("Completeness" in c for c in cells))

    def test_conflicting_amount_claims(self):
        # Same who·when·what key billed at two different amounts → flagged.
        # Equal-amount repeats are the suspected-duplicate signal, not this.
        data = (
            "ClaimID,BillingNPI,PatientID,DateOfService,HCPCS,BilledAmt\n"
            f"1,{GOOD_B},P1,2024-03-01,99213,100\n"
            f"2,{GOOD_B},P1,2024-03-01,99213,250\n"   # conflict with row 1
            f"3,{GOOD_B},P2,2024-03-01,99213,100\n"   # different patient
            f"4,{GOOD_B},P2,2024-03-01,99213,100\n"   # equal repeat, no conflict
        ).encode()
        res = engine.clean_bytes(data, "c.csv")
        self.assertEqual(res.sanity.get("conflicting-amount-claim"), 1)

    def test_carc_domain_and_top_denials(self):
        # Valid CARCs (numeric, letter+digits, multi-code cells) pass; a
        # non-CARC value flags. Top denial reasons are summarized.
        data = ("ClaimID,DenialCode\n"
                "1,16\n2,16\n3,\"16, 97\"\n4,B7\n5,PENDING\n").encode()
        res = engine.clean_bytes(data, "d.csv")
        self.assertEqual(res.sanity.get("carc-invalid"), 1)
        self.assertIsNotNone(res.denials)
        top = {d["code"]: d["count"] for d in res.denials["top"]}
        self.assertEqual(top.get("16"), 3)
        self.assertEqual(top.get("97"), 1)
        self.assertEqual(top.get("B7"), 1)

    def test_refdata_catalogs(self):
        from rcm_mc.npi_cleaner import refdata as rd
        self.assertEqual(rd.pos_name("11"), "Office")
        self.assertEqual(rd.pos_name("1"), "Pharmacy")     # 1-digit keying
        self.assertIn("Blood", rd.revenue_category("0380") or "")
        self.assertIsNotNone(rd.carc_description("197"))
        self.assertIsNotNone(rd.rarc_description("N115"))
        self.assertEqual(rd.modifier_meaning("JW"),
                         "Drug amount discarded / not administered")
        self.assertFalse(rd.tob_invalid("0111"))   # hospital, final claim
        self.assertFalse(rd.tob_invalid("131"))    # 3-digit keying
        self.assertTrue(rd.tob_invalid("999"))     # facility type 9 invalid
        self.assertTrue(rd.tob_invalid("11"))      # too short
        self.assertFalse(rd.discharge_status_invalid("01"))
        self.assertFalse(rd.discharge_status_invalid("1"))  # zero-stripped
        self.assertTrue(rd.discharge_status_invalid("XX"))
        self.assertTrue(rd.modifier_unknown("ZQ"))
        self.assertFalse(rd.modifier_unknown("25"))

    def test_rule_registry(self):
        from rcm_mc.npi_cleaner import rules
        cat = rules.catalog()
        self.assertGreater(len(cat), 50)
        ids = {r["id"] for r in cat}
        # Every engine sanity/repair key that exists today must be described.
        for must in ("hcpcs-malformed", "money-unparseable", "charge-outlier",
                     "revcode-pad", "tob-malformed", "timely-filing-risk",
                     "service-before-birth", "conflicting-amount-claim"):
            self.assertIn(must, ids)
        d = rules.describe("pos-invalid")
        self.assertEqual(d["severity"], "critical")
        self.assertTrue(d["remediation"])
        # Unknown rules degrade gracefully (renderer never breaks).
        self.assertEqual(rules.describe("brand-new-rule")["title"],
                         "brand-new-rule")

    def test_institutional_domain_flags(self):
        data = (
            "ClaimID,TypeOfBill,DischargeStatus,AdmissionType\n"
            "1,0111,01,1\n"       # all valid
            "2,999,XX,7\n"        # all invalid
        ).encode()
        res = engine.clean_bytes(data, "inst.csv")
        self.assertEqual(res.sanity.get("tob-malformed"), 1)
        self.assertEqual(res.sanity.get("discharge-status-invalid"), 1)
        self.assertEqual(res.sanity.get("admission-type-invalid"), 1)

    def test_discharge_status_not_money(self):
        # "DischargeStatus" contains "charge" — must not be routed to the
        # money cleaner ("01" → "1.00") and false-flagged. Sibling of the
        # DischargeDate/date-role regression.
        data = ("ClaimID,DischargeStatus\n1,01\n2,30\n").encode()
        res = engine.clean_bytes(data, "ds.csv")
        self.assertNotIn("discharge-status-invalid", res.sanity)
        self.assertNotIn("money-unparseable", res.sanity)
        with open(res.out_path, encoding="utf-8") as fh:
            self.assertIn("01", fh.read())

    def test_modifier_unknown_and_timely_filing(self):
        data = (
            "ClaimID,Modifier,Units,DateOfService,ReceivedDate\n"
            "1,ZQ,1,2024-01-01,2024-02-01\n"     # unknown modifier
            "2,25,1,2024-01-01,2025-06-01\n"     # 517 days → filing risk
            "3,50,1,2024-01-01,2024-06-01\n"     # fine
        ).encode()
        res = engine.clean_bytes(data, "mt.csv")
        self.assertEqual(res.sanity.get("modifier-unknown"), 1)
        self.assertEqual(res.sanity.get("timely-filing-risk"), 1)

    def test_drg_pad_and_malformed(self):
        # Excel-stripped DRG zeros restored (87 → 087); bad shapes flagged.
        data = ("ClaimID,MSDRG\n"
                "1,87\n"        # → 087 (pad repair)
                "2,470\n"       # valid, untouched
                "3,1234\n"      # 4 digits → malformed
                "4,ABC\n"       # non-numeric → malformed
                "5,000\n").encode()   # 000 is never assigned → malformed
        res = engine.clean_bytes(data, "drg.csv")
        self.assertEqual(res.repairs.get("drg-pad"), 1)
        self.assertEqual(res.sanity.get("drg-malformed"), 3)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("087", out)
        self.assertIn("470", out)
        # A header merely CONTAINING "drg" must not be treated as a DRG
        # column ("DrGroup" was the false-match risk).
        data2 = ("ClaimID,DrGroupNote\n1,87\n").encode()
        res2 = engine.clean_bytes(data2, "drg2.csv")
        self.assertIsNone(res2.repairs.get("drg-pad"))
        self.assertIsNone(res2.sanity.get("drg-malformed"))

    def test_anesthesia_units_implausible(self):
        data = ("ClaimID,HCPCS,Units\n"
                "1,00840,1500\n"     # anesthesia, >1440 → flag
                "2,00840,90\n"       # anesthesia, plausible
                "3,99213,2000\n").encode()   # not anesthesia → no flag
        res = engine.clean_bytes(data, "anes.csv")
        self.assertEqual(res.sanity.get("anesthesia-units-implausible"), 1)

    def test_timely_filing_per_payer_limit(self):
        # 120 days DOS→received: over UnitedHealthcare's 90-day limit,
        # inside Medicare's 365 — the payer on the ROW decides.
        data = ("ClaimID,PayerName,DateOfService,ReceivedDate\n"
                "1,UnitedHealthcare,2024-01-01,2024-04-30\n"
                "2,Medicare,2024-01-01,2024-04-30\n"
                "3,Some Tiny TPA,2024-01-01,2024-04-30\n").encode()
        res = engine.clean_bytes(data, "tf.csv")
        # Only the UHC row breaches: unknown payer falls back to 365.
        self.assertEqual(res.sanity.get("timely-filing-risk"), 1)
        from rcm_mc.npi_cleaner import refdata
        self.assertEqual(refdata.timely_filing_days("MEDICARE"), 365)
        self.assertIsNone(refdata.timely_filing_days("SOME TINY TPA"))

    def test_revenue_tob_mismatch(self):
        # Room & board (0120) on hospital OUTPATIENT 0131 → flag; the same
        # revenue on inpatient 0111 is fine; ancillary 0450 on outpatient
        # is fine. Row 4: clinic TOB (0731) with ICU revenue → flag.
        data = ("ClaimID,TypeOfBill,RevenueCode\n"
                "1,0131,0120\n"
                "2,0111,0120\n"
                "3,0131,0450\n"
                "4,0731,0200\n").encode()
        res = engine.clean_bytes(data, "tobrev.csv")
        self.assertEqual(res.sanity.get("revenue-tob-mismatch"), 2)
        from rcm_mc.npi_cleaner import refdata
        self.assertEqual(refdata.tob_facility_class("0131"), ("1", "3"))
        self.assertEqual(refdata.tob_facility_class("131"), ("1", "3"))
        self.assertIsNone(refdata.tob_facility_class("13"))

    def test_specialty_mix_from_taxonomy(self):
        data = ("ClaimID,TaxonomyCode\n"
                "1,207q00000x\n"          # lower-case → upper repair, counted
                "2,207Q00000X\n"
                "3,363L00000X\n"
                "4,BADCODE\n").encode()   # malformed → not counted
        res = engine.clean_bytes(data, "tax.csv")
        sc = res.as_scorecard()
        specs = {s["code"]: s for s in (sc["specialties"] or [])}
        self.assertEqual(specs["207Q00000X"]["n"], 2)
        self.assertEqual(specs["207Q00000X"]["name"], "Family Medicine")
        self.assertEqual(specs["363L00000X"]["name"], "Nurse Practitioner")
        self.assertNotIn("BADCODE", specs)

    def test_exec_report_credential_and_specialty_sections(self):
        data = ("RenderingProviderName,TaxonomyCode\n"
                "\"SMITH, JOHN, MD\",207Q00000X\n").encode()
        res = engine.clean_bytes(data, "er.csv")
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        html_out = build_exec_report(res.as_scorecard(), "er.csv",
                                     "2026-01-01T00:00:00+00:00")
        self.assertIn("Credential mix", html_out)
        self.assertIn("Doctor of Medicine", html_out)
        self.assertIn("Specialty mix", html_out)
        self.assertIn("Family Medicine", html_out)

    def test_worklist_row_capture(self):
        data = ("ClaimID,HCPCS\n"
                "1,99213\n2,BAD!!\n3,99214\n4,WRONG\n").encode()
        res = engine.clean_bytes(data, "wl.csv")
        # Output rows 2 and 4 carry the malformed codes.
        self.assertEqual(res.flag_rows.get("hcpcs-malformed"), [2, 4])
        sc = res.as_scorecard()
        self.assertEqual(sc["worklists"]["hcpcs-malformed"], 2)
        self.assertGreaterEqual(len(sc["rule_catalog"]), 58)

    def test_run_history_record_and_compare(self):
        from rcm_mc.npi_cleaner import history
        good = ("ClaimID,BillingNPI\n" f"1,{GOOD_B}\n").encode()
        bad = ("ClaimID,BillingNPI,HCPCS\n1,123,BAD!!\n").encode()
        r1 = engine.clean_bytes(good, "hist_good.csv")
        r2 = engine.clean_bytes(bad, "hist_bad.csv")
        runs = history.list_runs(10)
        names = [r["file_name"] for r in runs]
        self.assertIn("hist_good.csv", names)
        self.assertIn("hist_bad.csv", names)
        a = next(r for r in runs if r["file_name"] == "hist_good.csv")
        b = next(r for r in runs if r["file_name"] == "hist_bad.csv")
        cmp_ = history.compare_runs(a["run_id"], b["run_id"])
        self.assertIsNotNone(cmp_)
        self.assertLess(cmp_["score_delta"], 0)   # bad file scores lower
        rules_moved = {d["rule"] for d in cmp_["rule_delta"]}
        self.assertIn("hcpcs-malformed", rules_moved)

    def test_exec_report_builds(self):
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        res = engine.clean_bytes(engine.sample_csv().encode(), "s.csv")
        html_doc = build_exec_report(res.as_scorecard(), "s.csv",
                                     "2026-01-01 00:00 UTC")
        self.assertIn("Claims data-quality report", html_doc)
        self.assertIn("Quality dimensions", html_doc)
        self.assertIn("Findings", html_doc)

    def test_cli_batch_mode(self):
        from rcm_mc.npi_cleaner.cli import main as cli_main
        with tempfile.TemporaryDirectory() as tmp:
            srcp = os.path.join(tmp, "in.csv")
            with open(srcp, "w", encoding="utf-8") as fh:
                fh.write(f"ClaimID,BillingNPI\n1, {GOOD_B} \n")
            rc = cli_main([srcp, "--outdir", tmp])
            self.assertEqual(rc, 0)
            self.assertTrue(os.path.exists(
                os.path.join(tmp, "in_cleaned.csv")))
            # Quality gate: a clean file passes 90, nothing passes 101.
            self.assertEqual(cli_main([srcp, "--outdir", tmp,
                                       "--min-score", "90"]), 0)
            self.assertEqual(cli_main([srcp, "--outdir", tmp,
                                       "--min-score", "101"]), 1)

    def test_profiles_module(self):
        from rcm_mc.npi_cleaner import profiles
        cfg = profiles.save_profile("t-suite", {
            "disabled_rules": ["modifier-unknown", "not-a-real-rule"],
            "accepted_rules": ["date-stale"],
            "thresholds": {"timely_filing_days": 5,      # clamped to 30
                           "stale_years": 2,
                           "outlier_iqr_mult": 99}})     # clamped to 10
        self.assertEqual(cfg["disabled_rules"], ["modifier-unknown"])
        self.assertEqual(cfg["thresholds"]["timely_filing_days"], 30)
        self.assertEqual(cfg["thresholds"]["outlier_iqr_mult"], 10.0)
        loaded = profiles.get_profile("t-suite")
        self.assertEqual(loaded["name"], "t-suite")
        names = [p["name"] for p in profiles.list_profiles()]
        self.assertIn("t-suite", names)
        self.assertTrue(profiles.delete_profile("t-suite"))
        self.assertIsNone(profiles.get_profile("t-suite"))

    def test_profile_applied_to_run(self):
        # Thresholds honored, disabled rule absent, accepted rule reported
        # but excluded from the grade.
        prof = {"name": "px",
                "disabled_rules": ["modifier-unknown"],
                "accepted_rules": ["date-stale"],
                "thresholds": {"timely_filing_days": 30, "stale_years": 2,
                               "outlier_iqr_mult": 3.0}}
        # Clean padding rows keep the consistency ratio off its 0 floor —
        # timely-filing-risk now (correctly) feeds the grade, and with a
        # single row both variants saturated at 0.
        from datetime import date as _d, timedelta as _td
        _dos = (_d.today() - _td(days=10)).isoformat()
        _rcv = (_d.today() - _td(days=5)).isoformat()
        _pad = "".join(f"{i},25,1,{_dos},{_rcv}\n" for i in range(2, 5))
        data = ("ClaimID,Modifier,Units,DateOfService,ReceivedDate\n"
                "1,ZQ,1,2023-01-01,2024-03-01\n" + _pad).encode()
        res = engine.clean_bytes(data, "prof.csv", profile=prof)
        sc = res.as_scorecard()
        self.assertNotIn("modifier-unknown", sc["sanity"])   # disabled
        self.assertIn("date-stale", sc["sanity"])            # 2y horizon hit
        self.assertIn("timely-filing-risk", sc["sanity"])    # 30d limit hit
        self.assertEqual(sc["accepted_rules"], ["date-stale"])
        self.assertEqual(sc["profile"], "px")
        # Without the acceptance the same file grades lower on consistency.
        res2 = engine.clean_bytes(data, "prof2.csv", profile={
            "thresholds": prof["thresholds"]})
        c_acc = res.quality()["dimensions"]["consistency"]
        c_raw = res2.quality()["dimensions"]["consistency"]
        self.assertGreater(c_acc, c_raw)

    def test_workbook_worklist_sheets(self):
        data = ("ClaimID,HCPCS\n1,99213\n2,BAD!!\n").encode()
        res = engine.clean_bytes(data, "wlx.csv")
        openpyxl = __import__("openpyxl")
        from io import BytesIO
        with open(res.workbook_path, "rb") as fh:
            wb = openpyxl.load_workbook(BytesIO(fh.read()))
        wl_sheets = [s for s in wb.sheetnames if s.startswith("WL ")]
        self.assertTrue(wl_sheets)
        ws = wb[wl_sheets[0]]
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        self.assertIn("BAD!!", cells)
        self.assertNotIn("99213", cells)   # unflagged row not in worklist tab

    def test_mbi_validation(self):
        from rcm_mc.npi_cleaner import refdata as rd
        self.assertFalse(rd.mbi_malformed("1EG4-TE5-MK73"))  # valid, hyphens ok
        self.assertTrue(rd.mbi_malformed("1SG4TE5MK73"))     # S excluded
        self.assertTrue(rd.mbi_malformed("0EG4TE5MK73"))     # can't start 0
        self.assertTrue(rd.mbi_malformed("1EG4TE5MK7"))      # 10 chars
        # Engine: fires only on Medicare rows, never under de-id.
        data = ("ClaimID,Payer,MemberID\n"
                "1,Medicare,1EG4-TE5-MK73\n"   # valid MBI
                "2,Medicare,BADMBI\n"          # malformed → flag
                "3,Aetna,BADMBI\n").encode()   # non-Medicare → no check
        res = engine.clean_bytes(data, "mbi.csv")
        self.assertEqual(res.sanity.get("mbi-malformed"), 1)
        res_deid = engine.clean_bytes(data, "mbi2.csv", deid=True)
        self.assertNotIn("mbi-malformed", res_deid.sanity)

    def test_ub_code_catalogs_and_shape(self):
        from rcm_mc.npi_cleaner import refdata as rd
        self.assertEqual(rd.condition_code_meaning("44"),
                         "Inpatient admission changed to outpatient")
        self.assertIsNotNone(rd.occurrence_code_meaning("24"))
        self.assertIsNotNone(rd.value_code_meaning("80"))
        data = ("ClaimID,ConditionCode,OccurrenceCode,ValueCode\n"
                "1,44,24,80\n"          # valid
                "2,\"44, A1\",05,B1\n"  # multi-code cell valid
                "3,4,ABC,8!\n").encode()  # all malformed shapes
        res = engine.clean_bytes(data, "ub.csv")
        self.assertEqual(res.sanity.get("condition-code-malformed"), 1)
        self.assertEqual(res.sanity.get("occurrence-code-malformed"), 1)
        self.assertEqual(res.sanity.get("value-code-malformed"), 1)

    def test_near_duplicate_rows(self):
        # Same content differing only by case/extra spaces → near-dup flag
        # (row kept); an exact repeat is removed by dedupe, not double-flagged.
        data = ("ID,Name\n1,Hello World\n1,HELLO   world\n"
                "1,Hello World\n2,Other\n").encode()
        res = engine.clean_bytes(data, "nd.csv")
        self.assertEqual(res.sanity.get("near-duplicate-row"), 1)
        self.assertEqual(res.n_dupes_removed, 1)
        self.assertEqual(res.n_rows_out, 3)   # near-dup row is KEPT

    def test_provider_name_recase_unit(self):
        # Shouting person names get standard casing with Mc/O'/hyphen
        # handling; credentials stay uppercase; suffixes keep canon form.
        c = engine._clean_provider_name_cell
        for src, want, creds in (
            ("SMITH, JOHN A, MD", "Smith, John A, MD", ["MD"]),
            ("O'BRIEN-SMITH, MARY, NP", "O'Brien-Smith, Mary, NP", ["NP"]),
            ("MCDONALD, RONALD, DO", "McDonald, Ronald, DO", ["DO"]),
            # No Mac- rule on purpose: Macias must NOT become MacIas.
            ("MACIAS, JOSE, M.D.", "Macias, Jose, MD", ["MD"]),
            ("smith, jane, pa-c", "Smith, Jane, PA-C", ["PA-C"]),
            ("SMITH JR, WILLIAM, DDS", "Smith Jr, William, DDS", ["DDS"]),
        ):
            got, hits, seen = c(src)
            self.assertEqual(got, want)
            self.assertEqual(hits, ["provider-name-format"])
            self.assertEqual(seen, creds)
        # Mixed case = already curated → untouched, credentials still parsed.
        got, hits, seen = c("Smith, John, MD")
        self.assertEqual((got, hits, seen), ("Smith, John, MD", [], ["MD"]))
        # An org name in a provider-name column must pass through untouched.
        for org in ("SMITH FAMILY CLINIC LLC", "MERCY HOSPITAL",
                    "ACME MEDICAL GROUP INC"):
            got, hits, seen = c(org)
            self.assertEqual((got, hits, seen), (org, [], []))
        # Digits mean an ID crept in — never re-case those.
        self.assertEqual(c("SMITH 12345"), ("SMITH 12345", [], []))

    def test_provider_name_engine_and_credentials(self):
        # End to end: the provider-name column is re-cased, the org-name
        # column is untouched, and the credential mix lands in the scorecard.
        data = ("RenderingProviderName,OrganizationName,BilledAmt\n"
                "\"SMITH, JOHN A, MD\",MERCY HOSPITAL INC,100\n"
                "\"O'BRIEN, MARY, NP\",MERCY HOSPITAL INC,200\n").encode()
        res = engine.clean_bytes(data, "prov.csv")
        self.assertEqual(res.repairs.get("provider-name-format"), 2)
        sc = res.as_scorecard()
        self.assertEqual(sc["credentials"], {"MD": 1, "NP": 1})
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("Smith, John A, MD", out)
        self.assertIn("O'Brien, Mary, NP", out)
        self.assertIn("MERCY HOSPITAL INC", out)   # org column untouched
        from rcm_mc.npi_cleaner import rules
        self.assertEqual(rules.describe("provider-name-format")["kind"],
                         "repair")

    def test_credential_catalog_in_sync(self):
        # Every credential the engine can parse must have a display meaning
        # in refdata (and vice versa) — the two sets are maintained together.
        from rcm_mc.npi_cleaner import refdata
        self.assertEqual(set(engine._CREDENTIALS), set(refdata.CREDENTIALS))
        self.assertEqual(refdata.credential_meaning("m.d."),
                         "Doctor of Medicine")
        self.assertIsNone(refdata.credential_meaning("XYZ"))

    def test_formula_injection_defanged(self):
        # A cell that would start an Excel formula must be neutralized in CSV.
        data = ("NPI,Note\n" + GOOD_A + ",=SUM(A1:A9)\n").encode()
        res = engine.clean_bytes(data, "x.csv")
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("'=SUM(A1:A9)", out)

    def test_phi_deid_masks_patient_keeps_provider(self):
        # De-id must mask ONLY patient identifiers — provider NPI/name stay
        # intact (NPPES recovery relies on them). Off by default.
        data = (
            "BillingNPI,ProviderName,PatientName,PatientDOB,PatientZip,"
            "MRN,SSN,PatientAcct\n"
            f"{GOOD_B},Dr Jane Smith,John Q Public,1980-05-14,90210,"
            "MR12345,123-45-6789,ACCT-778\n"
            f"{GOOD_B},Dr Jane Smith,Mary Roe,1975-11-02,90210,"
            "MR99999,987-65-4321,ACCT-778\n"
        ).encode()
        res = engine.clean_bytes(data, "phi.csv", deid=True)
        sc = res.as_scorecard()
        self.assertIsNotNone(sc.get("deid"))
        self.assertGreater(sc["deid"]["cells"], 0)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        # Provider identifiers preserved.
        self.assertIn(GOOD_B, out)
        self.assertIn("Dr Jane Smith", out)
        # Patient identifiers masked.
        self.assertNotIn("John Q Public", out)
        self.assertNotIn("Mary Roe", out)
        self.assertNotIn("123-45-6789", out)
        self.assertNotIn("1980-05-14", out)   # DOB reduced to year
        self.assertIn("1980", out)
        self.assertIn("902XX", out)            # ZIP truncated to 3 digits
        # MRN/account tokenized, and the SAME source value → SAME token so
        # rows still link (both rows share PatientAcct ACCT-778).
        self.assertNotIn("MR12345", out)
        self.assertNotIn("ACCT-778", out)
        lines = [ln for ln in out.splitlines() if ln.strip()][1:]
        acct_tokens = [ln.split(",")[-1] for ln in lines]
        self.assertEqual(acct_tokens[0], acct_tokens[1])   # stable token
        self.assertTrue(acct_tokens[0].startswith("PT-"))

    def test_phi_deid_off_by_default(self):
        # Without the flag, patient data passes through untouched.
        data = ("BillingNPI,PatientName,SSN\n"
                f"{GOOD_B},John Q Public,123-45-6789\n").encode()
        res = engine.clean_bytes(data, "phi.csv")
        self.assertIsNone(res.as_scorecard().get("deid"))
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("John Q Public", out)
        self.assertIn("123-45-6789", out)

    def test_xlsx_upload_roundtrip(self):
        openpyxl = __import__("openpyxl")
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ClaimID", "BillingNPI"])
        ws.append([1, GOOD_A])
        ws.append([2, "99999"])
        buf = BytesIO()
        wb.save(buf)
        res = engine.clean_bytes(buf.getvalue(), "claims.xlsx")
        sc = res.as_scorecard()
        self.assertEqual(sc["delimiter"], "xlsx (Excel)")
        self.assertEqual(sc["rows_in"], 2)
        self.assertEqual(sc["billing_column"], "BillingNPI")

    def test_workbook_generated(self):
        data = (f"BillingNPI,AllowedAmt\n{GOOD_A},80\n").encode()
        res = engine.clean_bytes(data, "c.csv")
        self.assertTrue(res.workbook_path)
        openpyxl = __import__("openpyxl")
        wb = openpyxl.load_workbook(res.workbook_path)
        self.assertIn("Scorecard", wb.sheetnames)
        self.assertIn("Cleaned data", wb.sheetnames)
        # Executive summary is the FIRST tab — the page a VP opens to.
        self.assertEqual(wb.sheetnames[0], "Summary")
        cells = [str(c.value) for row in wb["Summary"].iter_rows()
                 for c in row if c.value is not None]
        joined = " | ".join(cells)
        self.assertIn("Grade", joined)
        self.assertIn("Deterministic fixes applied", joined)

    def test_sample_csv_cleans(self):
        res = engine.clean_bytes(engine.sample_csv().encode(), "sample.csv")
        sc = res.as_scorecard()
        self.assertEqual(sc["duplicates_removed"], 1)   # rows 1001/1002
        self.assertGreater(sc["billing_issues"], 0)


class TestVendorAdapter(unittest.TestCase):
    """The real, complete v49 deterministic engine (schema.standardize_any +
    clean_orchestrator.clean_all), driven through vendor_adapter."""

    def setUp(self):
        from rcm_mc.npi_cleaner import vendor_adapter as va
        if not va.available():
            self.skipTest("pandas / vendored v49 engine unavailable")
        self.va = va

    def test_v49_engine_runs_and_sizes_issues(self):
        data = (
            "ClaimID,BillingProviderNPI,ReferringNPI,ChargeAmt,AllowedAmt,"
            "PaidAmt,DateOfService,PaidDate,HCPCS\n"
            f"1,{GOOD_A},{GOOD_B},100,80,60,2024-01-15,2024-02-01,99213\n"
            # allowed>billed + paid>allowed (money), referring==billing (role)
            f"2,{GOOD_A},{GOOD_A},50,90,40,2024-03-10,2024-03-01,99214\n"
            f"3,99999,{GOOD_B},200,150,120,2024-05-01,2024-06-01,99215\n"
        ).encode()
        res = self.va.run(data)
        self.assertIsNotNone(res)
        self.assertIn("clean_all", res["engine"])
        # Real screens fire; issues are sized with a systematic verdict.
        self.assertIn("money_ordering", res["screens"])
        self.assertGreaterEqual(res["screens"]["money_ordering"], 1)
        self.assertGreaterEqual(res["screens"].get("npi_role_coherence", 0), 1)
        issues = {i["issue"] for i in res["issues"]}
        self.assertIn("money_ordering", issues)
        # A corrections companion is produced for the consistency violations.
        self.assertGreater(res["suggestions_n"], 0)
        self.assertTrue(res["suggestions_records"])
        self.assertIn("suggested_value", res["suggestions_records"][0])

    def test_engine_attaches_v49_advanced_and_companion(self):
        data = (
            "BillingNPI,ChargeAmt,AllowedAmt,PaidAmt\n"
            f"{GOOD_A},50,90,40\n"   # allowed>billed + paid>allowed
        ).encode()
        res = engine.clean_bytes(data, "c.csv")
        self.assertIsNotNone(res.advanced)
        self.assertIn("clean_all", res.advanced["engine"])
        # Suggestions companion is written to its own CSV.
        self.assertTrue(res.companion_path)
        with open(res.companion_path, encoding="utf-8") as fh:
            head = fh.readline()
        self.assertIn("suggested_value", head)

    def test_extended_anomaly_screens(self):
        # Enough rows for the Benford screen (needs >=100 amounts).
        lines = ["BillingNPI,Payer,AllowedAmt"]
        for i in range(150):
            amt = 100 + (i * 37) % 900
            lines.append(f"16799999{i%80:02d},PayerA,{amt}")
        data = ("\n".join(lines) + "\n").encode()
        res = self.va.run(data)
        self.assertIsNotNone(res)
        ext = res.get("extended", [])
        keys = {e["key"] for e in ext}
        # Benford + HHI should always be computable on this data.
        self.assertIn("benford", keys)
        self.assertIn("hhi", keys)

    def test_bundled_v49_sample_runs(self):
        from pathlib import Path
        sample = (Path(__import__("rcm_mc").__file__).parent / "npi_cleaner"
                  / "vendor_v49" / "examples" / "sample_claims.xlsx")
        if not sample.exists():
            self.skipTest("bundled sample missing")
        res = engine.clean_bytes(sample.read_bytes(), "sample_claims.xlsx")
        self.assertEqual(res.as_scorecard()["delimiter"], "xlsx (Excel)")
        self.assertIsNotNone(res.advanced)
        # 20 deterministic repairs + the JW/JZ wastage issue on the sample.
        self.assertGreater(res.advanced["repairs"], 0)
        self.assertTrue(res.advanced["issues"])


class TestConnectors(unittest.TestCase):
    """Live drug connectors (RxNorm / openFDA) with an injected opener."""

    def setUp(self):
        from rcm_mc.npi_cleaner import connectors as C
        if not C.available():
            self.skipTest("public_api_clients unavailable")
        self.C = C

    def _opener(self, url, headers, timeout_s):
        if "rxcui.json" in url and "idtype=NDC" in url:
            return json.dumps({"idGroup": {"rxnormId": ["1049502"]}}).encode()
        if "rxcui.json" in url:
            return json.dumps({"idGroup": {"rxnormId": ["860975"]}}).encode()
        if "/properties.json" in url:
            return json.dumps({"properties": {"name": "metformin", "tty": "IN"}}).encode()
        if "/ndcs.json" in url:
            return json.dumps({"ndcGroup": {"ndcList": {"ndc": ["00093-1049"]}}}).encode()
        if "api.fda.gov" in url:
            return json.dumps({"results": [{"brand_name": "GLUCOPHAGE",
                                            "generic_name": "METFORMIN",
                                            "labeler_name": "BMS"}]}).encode()
        return b"{}"

    def test_catalog_lists_sources(self):
        cat = self.C.catalog()
        self.assertGreaterEqual(len(cat), 15)
        ids = {c["id"] for c in cat}
        self.assertIn("rxnorm", ids)
        self.assertIn("nppes", ids)
        self.assertIn("openfda", ids)

    def test_resolve_drugs_rxnorm_and_openfda(self):
        res = self.C.resolve_drugs(
            ["0093-1049-01"], ["metformin"], opener=self._opener)
        by_id = {r["id"]: r for r in res}
        self.assertIn("rxnorm", by_id)
        self.assertEqual(by_id["rxnorm"]["resolved"], 2)  # 1 NDC + 1 name
        self.assertTrue(by_id["rxnorm"]["sample"][0]["rxcui"])
        self.assertIn("openfda", by_id)
        self.assertEqual(by_id["openfda"]["resolved"], 1)
        self.assertEqual(by_id["openfda"]["sample"][0]["brand"], "GLUCOPHAGE")

    def test_engine_online_wires_connectors_and_catalog(self):
        from rcm_mc.npi_cleaner import connectors as C

        def fake_resolve(ndcs, drugs, **kw):
            return [{"id": "rxnorm", "label": "RxNorm / RxNav", "queried": 1,
                     "resolved": 1, "unresolved": 0, "sample": [], "note": "ok"}]

        def fake_fetch(npi, **kw):
            return None

        data = ("BillingNPI,NDC,DrugName\n"
                f"{GOOD_A},0093-1049-01,Metformin\n").encode()
        with patch.object(C, "resolve_drugs", fake_resolve), \
             patch("rcm_mc.data_public.nppes_api_client.fetch_by_npi", fake_fetch):
            res = engine.clean_bytes(data, "c.csv", enrich=True)
        sc = res.as_scorecard()
        self.assertTrue(sc["connectors"])
        self.assertEqual(sc["connectors"][0]["id"], "rxnorm")
        self.assertGreaterEqual(len(sc["catalog"]), 15)


class TestCompliance(unittest.TestCase):
    """OIG LEIE (offline) + Medicare PECOS (mocked client) screening."""

    def test_leie_offline_flags_excluded(self):
        from rcm_mc.npi_cleaner import compliance as C
        p = os.path.join(tempfile.mkdtemp(), "leie.csv")
        with open(p, "w", encoding="utf-8") as fh:
            fh.write("LASTNAME,NPI,EXCLTYPE\n"
                     f"BADDOC,{GOOD_B},1128a1\n")
        r = C.screen_leie([GOOD_B, GOOD_A, "99999"], leie_path=p)
        self.assertTrue(r["available"])
        self.assertEqual(r["excluded"], 1)
        self.assertEqual(r["matches"][0]["npi"], GOOD_B)

    def test_leie_no_dataset_note(self):
        from rcm_mc.npi_cleaner import compliance as C
        r = C.screen_leie([GOOD_A], leie_path="/nonexistent/leie.csv")
        self.assertFalse(r["available"])
        self.assertIn("LEIE", r["note"])

    def test_pecos_with_mock_client(self):
        from rcm_mc.npi_cleaner import compliance as C

        class FakeCMS:
            def enrollment_lookup(self, npi):
                return {"enrolled": npi != GOOD_A}

            def opt_out_lookup(self, npi):
                return {"opted_out": npi == GOOD_B}

        r = C.screen_cms([GOOD_A, GOOD_B], cms_client=FakeCMS())
        self.assertTrue(r["available"])
        self.assertEqual(r["checked"], 2)
        self.assertEqual(r["not_enrolled"], 1)   # GOOD_A
        self.assertEqual(r["opted_out"], 1)      # GOOD_B

    def test_engine_online_attaches_compliance(self):
        from rcm_mc.npi_cleaner import compliance as C, connectors as CN

        def fake_screen(npis, **kw):
            return [{"id": "oig_leie", "label": "OIG LEIE (excluded providers)",
                     "available": True, "checked": len(npis), "excluded": 0,
                     "matches": [], "note": "clean"}]

        def fake_fetch(npi, **kw):
            return None

        data = (f"BillingNPI\n{GOOD_A}\n99999\n").encode()
        with patch.object(C, "screen", fake_screen), \
             patch.object(CN, "resolve_drugs", lambda *a, **k: []), \
             patch("rcm_mc.data_public.nppes_api_client.fetch_by_npi", fake_fetch):
            res = engine.clean_bytes(data, "c.csv", enrich=True)
        sc = res.as_scorecard()
        self.assertTrue(sc["compliance"])
        self.assertEqual(sc["compliance"][0]["id"], "oig_leie")


class TestDeepPipeline(unittest.TestCase):
    """Deep recovery (full v49 run_pipeline) wiring — timeout + graceful fail."""

    def test_available(self):
        from rcm_mc.npi_cleaner import deep_pipeline as D
        # Boolean either way; just exercise the guard.
        self.assertIn(D.available(), (True, False))

    def test_timeout_is_graceful_not_a_hang(self):
        from rcm_mc.npi_cleaner import deep_pipeline as D
        if not D.available():
            self.skipTest("deep pipeline unavailable")
        from rcm_mc.npi_cleaner.vendor_v49.npi_recovery import pipeline as P

        def _hang(*a, **k):
            time.sleep(10)

        with patch.object(P, "run_pipeline", _hang):
            out = D.run(b"BillingNPI\n" + GOOD_A.encode(),
                        "x.csv", timeout_s=1)
        self.assertFalse(out["ok"])
        self.assertIn("timed out", out["error"])

    def test_engine_attaches_deep_result(self):
        from rcm_mc.npi_cleaner import deep_pipeline, engine as eng
        import tempfile
        import os

        wb = os.path.join(tempfile.mkdtemp(), "x_recovered.xlsx")
        with open(wb, "wb") as fh:
            fh.write(b"PK\x03\x04stub")

        def fake_run(data, name, **kw):
            return {"ok": True, "error": None, "stats": {"npis_recovered": 5},
                    "workbook_path": wb, "workbook_name": "x_recovered.xlsx"}

        data = ("BillingNPI\n" + GOOD_A + "\n").encode()
        with patch.object(deep_pipeline, "run", fake_run):
            res = eng.clean_bytes(data, "x.csv", deep=True)
        self.assertIsNotNone(res.deep)
        self.assertTrue(res.deep["ok"])
        self.assertEqual(res.deep_workbook_path, wb)
        self.assertEqual(res.as_scorecard()["deep_workbook_name"],
                         "x_recovered.xlsx")


class TestNppesBridge(unittest.TestCase):
    """Live NPPES verify/recover, cross-using the shared CMS client (mocked)."""

    def _providers(self):
        from rcm_mc.data_public.nppes_api_client import NppesProvider
        return NppesProvider

    def test_verify_active_and_not_found(self):
        from rcm_mc.npi_cleaner import nppes_bridge
        NppesProvider = self._providers()

        def fake_fetch(npi, **kw):
            if npi == GOOD_B:
                return NppesProvider(npi=npi, entity_type=2,
                                     name="MERCY HOSPITAL", state="OH")
            return None

        with patch("rcm_mc.data_public.nppes_api_client.fetch_by_npi", fake_fetch):
            out = nppes_bridge.verify_npis([GOOD_B, GOOD_A, GOOD_B, "123"])
        self.assertEqual(out["checked"], 2)     # GOOD_B + GOOD_A (distinct, 10-digit)
        self.assertEqual(out["active"], 1)      # GOOD_B resolves
        self.assertEqual(out["not_found"], 1)   # GOOD_A returns None
        self.assertEqual(out["records"][GOOD_B]["status"], "active")

    def test_recover_candidates(self):
        from rcm_mc.npi_cleaner import nppes_bridge
        NppesProvider = self._providers()

        def fake_search(name, state="", **kw):
            return [NppesProvider(npi=GOOD_A, entity_type=2,
                                  name="ACME CLINIC LLC", state=state or "TX")]

        with patch("rcm_mc.data_public.nppes_api_client.search_by_organization",
                   fake_search):
            out = nppes_bridge.recover_candidates([
                {"row": "2", "name": "Acme Clinic", "state": "TX"},
                {"row": "3", "name": "Acme Clinic", "state": "TX"},  # deduped
            ])
        self.assertEqual(out["searched"], 1)
        self.assertEqual(out["resolved"], 1)
        self.assertEqual(out["matches"][0]["candidates"][0]["npi"], GOOD_A)

    def test_engine_enrich_end_to_end(self):
        NppesProvider = self._providers()

        def fake_fetch(npi, **kw):
            if npi == GOOD_B:
                return NppesProvider(npi=npi, entity_type=2,
                                     name="MERCY HOSPITAL", state="OH")
            return None

        def fake_search(name, state="", **kw):
            return [NppesProvider(npi=GOOD_A, entity_type=2,
                                  name="ACME CLINIC LLC", state=state or "TX")]

        data = (
            "ClaimID,BillingProviderNPI,OrganizationName,ProviderState\n"
            f"1,{GOOD_B},Mercy Hospital,OH\n"
            "2,,Acme Clinic,TX\n"
        ).encode()
        with patch("rcm_mc.data_public.nppes_api_client.fetch_by_npi", fake_fetch), \
             patch("rcm_mc.data_public.nppes_api_client.search_by_organization",
                   fake_search):
            res = engine.clean_bytes(data, "c.csv", enrich=True)
        self.assertIsNotNone(res.nppes)
        self.assertEqual(res.nppes["verify"]["active"], 1)
        matches = res.nppes["recover"]["matches"]
        self.assertTrue(any(m["candidates"] for m in matches))

    def test_enrich_off_by_default(self):
        data = f"NPI\n{GOOD_A}\n".encode()
        res = engine.clean_bytes(data, "c.csv")
        self.assertIsNone(res.nppes)

    def test_recovery_written_to_cleaned_file(self):
        NppesProvider = self._providers()

        def fake_fetch(npi, **kw):
            return None  # nothing verifies; forces recovery path

        def fake_search(name, state="", **kw):
            return [NppesProvider(npi=GOOD_A, entity_type=2,
                                  name="ACME CLINIC LLC", state=state or "TX")]

        data = (
            "BillingNPI,OrganizationName,ProviderState\n"
            ",Acme Clinic,TX\n"
            "99999,Acme Clinic,TX\n"
        ).encode()
        with patch("rcm_mc.data_public.nppes_api_client.fetch_by_npi", fake_fetch), \
             patch("rcm_mc.data_public.nppes_api_client.search_by_organization",
                   fake_search):
            res = engine.clean_bytes(data, "c.csv", enrich=True)
        # Both Acme rows get the recovered NPI written to a new column.
        self.assertEqual(len(res.recovered_rows), 2)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("recovered_billing_npi", out)
        self.assertEqual(out.count(GOOD_A), 2)


_X12_837P = (
    "ISA*00*          *00*          *ZZ*SUBMITTER      *ZZ*RECEIVER       "
    "*240110*1200*^*00501*000000001*0*P*:~"
    "GS*HC*SUB*REC*20240110*1200*1*X*005010X222A1~"
    "ST*837*0001*005010X222A1~"
    "BHT*0019*00*1*20240110*1200*CH~"
    "NM1*85*2*ACME FAMILY CLINIC*****XX*1497758544~"
    "NM1*IL*1*DOE*JANE****MI*MBR123~"
    "NM1*PR*2*UNITEDHEALTHCARE*****PI*87726~"
    "CLM*ACCT001*225***11:B:1*Y*A*Y*Y~"
    "HI*ABK:E1165~"
    "NM1*82*1*SMITH*JOHN****XX*1234567893~"
    "LX*1~"
    "SV1*HC:99213:25*100*UN*1*11~"
    "DTP*472*D8*20240110~"
    "LX*2~"
    "SV1*HC:93000*125*UN*1~"
    "DTP*472*D8*20240110~"
    "CLM*ACCT002*80***11:B:1*Y*A*Y*Y~"
    "HI*ABK:I10~"
    "LX*1~"
    "SV1*HC:99212*80*UN*1~"
    "DTP*472*D8*20240111~"
    "SE*20*0001~GE*1*1~IEA*1*000000001~").encode()

_X12_837I = (
    "ISA*00*          *00*          *ZZ*S              *ZZ*R              "
    "*240110*1200*^*00501*000000002*0*P*:~"
    "GS*HC*S*R*20240110*1200*2*X*005010X223A2~"
    "ST*837*0002*005010X223A2~"
    "NM1*85*2*MERCY HOSPITAL*****XX*1497758544~"
    "NM1*IL*1*ROE*MARY****MI*M2~"
    "NM1*PR*2*MEDICARE~"
    "CLM*IP001*2500***13:A:1~"
    "DTP*434*RD8*20240101-20240103~"
    "HI*ABK:A419~"
    "NM1*71*1*JONES*SAM****XX*1234567893~"
    "LX*1~"
    "SV2*0450*HC:99284*250*UN*1~"
    "LX*2~"
    "SV2*0120**2250*UN*3~"
    "SE*14*0002~GE*1*2~IEA*1*000000002~").encode()


class TestX12(unittest.TestCase):
    """X12 837 ingestion — the native claims wire format, flattened to one
    row per service line and fed through the normal pipeline."""

    def test_detection(self):
        from rcm_mc.npi_cleaner import x12
        self.assertTrue(x12.looks_like_x12(_X12_837P))
        self.assertFalse(x12.looks_like_x12(b"NPI,Name\n123,ACME\n"))
        # A CSV whose first header starts with ISA must not be mistaken.
        self.assertFalse(x12.looks_like_x12(b"ISAN,Other\n1,2\n"))

    def test_837p_flatten(self):
        from rcm_mc.npi_cleaner import x12
        h, rows = x12.x12_to_table(_X12_837P)
        self.assertEqual(len(rows), 3)          # 2 lines + 1 line
        r0 = dict(zip(h, rows[0]))
        self.assertEqual(r0["BillingNPI"], "1497758544")
        self.assertEqual(r0["RenderingNPI"], "1234567893")
        self.assertEqual(r0["HCPCS"], "99213")
        self.assertEqual(r0["Modifiers"], "25")
        self.assertEqual(r0["DateOfService"], "2024-01-10")
        self.assertEqual(r0["PlaceOfService"], "11")
        self.assertEqual(r0["DiagnosisCode"], "E1165")
        self.assertEqual(r0["PayerName"], "UNITEDHEALTHCARE")
        # Rendering NPI must NOT leak from claim 1 into claim 2.
        r2 = dict(zip(h, rows[2]))
        self.assertEqual(r2["ClaimID"], "ACCT002")
        self.assertEqual(r2["RenderingNPI"], "")

    def test_837i_flatten(self):
        from rcm_mc.npi_cleaner import x12
        h, rows = x12.x12_to_table(_X12_837I)
        i0 = dict(zip(h, rows[0]))
        self.assertEqual(i0["TypeOfBill"], "131")   # facility 13 + freq 1
        self.assertEqual(i0["RevenueCode"], "0450")
        self.assertEqual(i0["AttendingNPI"], "1234567893")
        self.assertEqual(i0["DateOfService"], "2024-01-01")  # RD8 start
        i1 = dict(zip(h, rows[1]))
        self.assertEqual(i1["RevenueCode"], "0120")
        self.assertEqual(i1["Units"], "3")

    def test_837_through_engine(self):
        res = engine.clean_bytes(_X12_837P, "claims.837")
        sc = res.as_scorecard()
        self.assertEqual(sc["delimiter"], "X12 837 (EDI)")
        self.assertEqual(sc["rows_in"], 3)
        self.assertIn("BillingNPI", sc["npi_columns"])
        self.assertIn("RenderingNPI", sc["npi_columns"])
        # 1234567893 fails the NPI Luhn check → flagged, not silently kept.
        self.assertGreater(sc["npi_issues"], 0)
        # ICD-10 decimal repair fires on the flattened dx (E1165 → E11.65).
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("E11.65", out)

    def test_837_deid_masks_patient(self):
        res = engine.clean_bytes(_X12_837P, "claims.837", deid=True)
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertNotIn("DOE, JANE", out)      # patient masked
        self.assertIn("ACME FAMILY CLINIC", out)  # provider kept

    def test_non_837_x12_precise_warning(self):
        ack = (_X12_837P[:106]
               + b"GS*FA*S*R*20240110*1200*3*X*005010X231A1~"
                 b"ST*999*0001~AK1*HC*1~SE*3*0001~")
        res = engine.clean_bytes(ack, "ack.999")
        self.assertTrue(any("no 837 claim" in w for w in res.warnings))
        self.assertEqual(res.n_rows_in, 0)


_X12_835 = (
    "ISA*00*          *00*          *ZZ*PAYER          *ZZ*PROVIDER       "
    "*240215*0800*^*00501*000000003*0*P*:~"
    "GS*HP*PAYER*PROV*20240215*0800*3*X*005010X221A1~"
    "ST*835*0003~"
    "BPR*I*330*C*ACH*CCP~"
    "TRN*1*CHK12345*1512345678~"
    "DTM*405*20240215~"
    "N1*PR*UNITEDHEALTHCARE~"
    "N1*PE*ACME FAMILY CLINIC*XX*1497758544~"
    "CLP*ACCT001*1*225*180*20*12*UHC-ICN-1~"
    "NM1*QC*1*DOE*JANE~"
    "DTM*232*20240110~"
    "SVC*HC:99213:25*100*80~"
    "CAS*CO*45*15~"
    "CAS*PR*3*5~"
    "SVC*HC:93000*125*100~"
    "CAS*CO*45*25~"
    "CLP*ACCT002*4*80*0*0*12*UHC-ICN-2~"
    "NM1*QC*1*ROE*MARY~"
    "DTM*232*20240111~"
    "CAS*CO*29*80~"
    "SE*20*0003~GE*1*3~IEA*1*000000003~").encode()


class TestX835(unittest.TestCase):
    """X12 835 remittance ingestion — CLP/SVC/CAS flattened to one row per
    paid service line, CARCs feeding the existing denial analytics."""

    def test_835_flatten(self):
        from rcm_mc.npi_cleaner import x12
        h, rows = x12.x835_to_table(_X12_835)
        self.assertEqual(len(rows), 3)
        r0 = dict(zip(h, rows[0]))
        self.assertEqual(r0["ClaimID"], "ACCT001")
        self.assertEqual(r0["HCPCS"], "99213")
        self.assertEqual((r0["BilledAmt"], r0["PaidAmt"]), ("100", "80"))
        # Line-level CAS attaches to ITS line only — no cross-line leak.
        self.assertEqual(r0["DenialCodes"], "45, 3")
        self.assertIn("CO-45:15", r0["AdjustmentDetail"])
        r1 = dict(zip(h, rows[1]))
        self.assertEqual(r1["DenialCodes"], "45")
        # A denied claim with no SVC detail still emits a claim-level row.
        r2 = dict(zip(h, rows[2]))
        self.assertEqual(r2["ClaimStatus"], "4")
        self.assertEqual(r2["DenialCodes"], "29")
        self.assertEqual(r0["PayeeNPI"], "1497758544")
        self.assertEqual(r0["PaidDate"], "2024-02-15")
        self.assertEqual(r0["DateOfService"], "2024-01-10")
        # An 837 is not an 835 and vice versa.
        self.assertIsNone(x12.x835_to_table(_X12_837P))
        self.assertIsNone(x12.x12_to_table(_X12_835))

    def test_835_through_engine(self):
        res = engine.clean_bytes(_X12_835, "remit.835")
        sc = res.as_scorecard()
        self.assertEqual(sc["delimiter"], "X12 835 (ERA)")
        # CARCs land in the denial analytics with catalog-known code 45
        # (charges exceed fee schedule) on top.
        self.assertTrue(sc["denials"])
        self.assertEqual(sc["denials"]["top"][0]["code"], "45")
        # Claim rollup rides the ClaimID column automatically.
        self.assertEqual(sc["claims"]["n_claims"], 2)


class TestDupWindowAndRollup(unittest.TestCase):
    def test_possible_duplicate_service_window(self):
        data = ("PatientID,BillingNPI,HCPCS,DateOfService\n"
                f"P1,{GOOD_A},99213,2024-01-01\n"
                f"P1,{GOOD_A},99213,2024-01-03\n"   # 2 days → both flagged
                f"P2,{GOOD_A},99214,2024-01-01\n"
                f"P2,{GOOD_A},99214,2024-01-20\n").encode()  # 19 days → no
        res = engine.clean_bytes(data, "dup.csv")
        self.assertEqual(res.sanity.get("possible-duplicate-service"), 2)
        self.assertEqual(res.flag_rows["possible-duplicate-service"], [1, 2])
        # The window is a profile threshold: at 1 day nothing matches.
        res2 = engine.clean_bytes(
            data, "dup.csv", profile={"thresholds": {"dup_window_days": 1}})
        self.assertIsNone(res2.sanity.get("possible-duplicate-service"))
        # And it participates in the uniqueness dimension + registry.
        from rcm_mc.npi_cleaner import rules, profiles
        self.assertEqual(rules.describe("possible-duplicate-service")
                         ["dimension"], "uniqueness")
        cfg = profiles._sanitize({"thresholds": {"dup_window_days": 99}})
        self.assertEqual(cfg["thresholds"]["dup_window_days"], 30)

    def test_claim_rollup(self):
        data = ("ClaimID,BilledAmt\nC1,100\nC1,50\nC2,300\n").encode()
        res = engine.clean_bytes(data, "roll.csv")
        c = res.as_scorecard()["claims"]
        self.assertEqual(c["n_claims"], 2)
        self.assertEqual(c["avg_lines"], 1.5)
        self.assertEqual(c["max_lines"], 2)
        self.assertEqual(c["charge"]["median"], 225.0)
        self.assertEqual(c["charge"]["max"], 300.0)
        self.assertFalse(c["truncated"])
        # No claim-id column → no rollup, not a crash.
        res2 = engine.clean_bytes(b"NPI,Amt\n123,5\n", "n.csv")
        self.assertIsNone(res2.as_scorecard()["claims"])
        # Rollup + exec report render together.
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        html_out = build_exec_report(res.as_scorecard(), "roll.csv",
                                     "2026-01-01T00:00:00+00:00")
        self.assertIn("Claim rollup", html_out)


class TestReviewFixes(unittest.TestCase):
    """Regression tests for the batch-18 full-diff review findings."""

    def test_money_hint_exclusions(self):
        # "CostCenter" matched the "cost" hint and code 0100 was rewritten
        # to 100.00; description/name columns false-flagged unparseable.
        data = ("CostCenter,ChargeDescription,FeeScheduleName,PaidAmt\n"
                "0100,Office visit,Standard,\"$1,250.50\"\n").encode()
        res = engine.clean_bytes(data, "cc.csv")
        with open(res.out_path, encoding="utf-8") as fh:
            out = fh.read()
        self.assertIn("0100", out)                 # cost center preserved
        self.assertIn("1250.50", out)              # real money still cleaned
        self.assertIsNone(res.sanity.get("money-unparseable"))

    def test_institutional_rules_hit_the_grade(self):
        # tob-malformed (critical) fired but never fed the validity
        # dimension — an all-bad-TOB file still graded A.
        data = ("ClaimID,TypeOfBill\n1,ABC\n2,XYZ\n").encode()
        res = engine.clean_bytes(data, "tob.csv")
        self.assertEqual(res.sanity.get("tob-malformed"), 2)
        self.assertLess(res.quality()["dimensions"]["validity"], 100.0)
        self.assertIn("timely-filing-risk", engine.CleanResult._CONSISTENCY_RULES)

    def test_worklist_captures_late_first_firing_rule(self):
        # Old _wl_open shut off capture globally once every rule seen so
        # far was full; a rule first firing later got zero worklist rows.
        rows = ["ClaimID,DateOfService,HCPCS"]
        for i in range(520):                        # fills date-stale's 500 cap
            rows.append(f"{i},1990-01-01,99213")
        rows.append("bad,2024-01-01,BAD!!")         # first hcpcs hit, row 521
        res = engine.clean_bytes(("\n".join(rows) + "\n").encode(), "wl.csv")
        self.assertEqual(len(res.flag_rows["date-stale"]), 500)
        self.assertEqual(res.flag_rows.get("hcpcs-malformed"), [521])

    def test_835_line_level_service_dates(self):
        from rcm_mc.npi_cleaner import x12
        era = ("ISA*00*          *00*          *ZZ*P              *ZZ*R    "
               "          *240215*0800*^*00501*000000009*0*P*:~"
               "GS*HP*P*R*20240215*0800*9*X*005010X221A1~ST*835*0009~"
               "N1*PR*AETNA~N1*PE*ACME CLINIC*XX*1497758544~"
               "CLP*C1*1*225*180*20*12*ICN1~"
               "SVC*HC:99213*100*80~DTM*472*20250101~"
               "SVC*HC:93000*125*100~DTM*472*20250215~"
               "SE*10*0009~").encode()
        h, rows = x12.x835_to_table(era)
        d0, d1 = dict(zip(h, rows[0])), dict(zip(h, rows[1]))
        self.assertEqual(d0["DateOfService"], "2025-01-01")
        self.assertEqual(d1["DateOfService"], "2025-02-15")  # not line 1's

    def test_835_role_detection(self):
        # name → PayeeName (the provider), never PayerName; ClaimStatus is
        # NOT a state column (the bare "st" substring hint did that).
        from rcm_mc.npi_cleaner import x12
        h = x12.HEADERS_835
        self.assertEqual(h[engine._detect_one(h, engine._NAME_HINTS)],
                         "PayeeName")
        self.assertIsNone(engine._detect_one(h, engine._STATE_HINTS))
        # PaidDate is no longer a never-in-future column: payers stamp
        # forward-dated production dates on ordinary ERAs.
        self.assertNotIn("paiddate", engine._SERVICE_DATE_HINTS)

    def test_credential_surname_not_miscounted(self):
        # "DO, HANH" is the surname Do, not a Doctor of Osteopathy.
        got, hits, creds = engine._clean_provider_name_cell("DO, HANH")
        self.assertEqual(got, "Do, Hanh")
        self.assertEqual(creds, [])
        got2, _, creds2 = engine._clean_provider_name_cell("PA, MINH")
        self.assertEqual(got2, "Pa, Minh")
        self.assertEqual(creds2, [])
        # Trailing credentials still parse.
        _, _, creds3 = engine._clean_provider_name_cell("SMITH, JOHN, DO")
        self.assertEqual(creds3, ["DO"])

    def test_profile_mapping_long_name_roundtrip(self):
        # save truncated names to 64 chars but get/delete didn't — a
        # 70-char name saved fine and was never findable again.
        import uuid as _uuid
        from rcm_mc.npi_cleaner import mappings, profiles
        long_name = "p" * 60 + _uuid.uuid4().hex[:10]      # 70 chars
        profiles.save_profile(long_name, {"thresholds": {"stale_years": 5}})
        try:
            self.assertIsNotNone(profiles.get_profile(long_name))
        finally:
            self.assertTrue(profiles.delete_profile(long_name))
        long_map = "m" * 60 + _uuid.uuid4().hex[:10]
        mappings.save_mapping(long_map, {"billing_npi": "X"})
        try:
            self.assertIsNotNone(mappings.get_mapping(long_map))
        finally:
            self.assertTrue(mappings.delete_mapping(long_map))

    def test_batch_deid_and_probe(self):
        import io as _io
        import zipfile as _zf
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w") as z:
            z.writestr("a.csv", "PatientName,NPI\nJohn Doe,1497758544\n")
        data = buf.getvalue()
        # Metadata-only probe agrees with the extracting detector.
        self.assertTrue(engine.zip_batch_probe(data))
        self.assertFalse(engine.zip_batch_probe(b"NPI\n1\n"))
        # De-id status must survive the batch merge.
        res = engine.clean_bytes(data, "b.zip", deid=True)
        sc = res.as_scorecard()
        self.assertIsNotNone(sc["deid"])
        self.assertGreater(sc["deid"]["cells"], 0)

    def test_near_duplicate_without_dedupe(self):
        # The near-dup screen was silently disabled under --no-dedupe —
        # exactly the mode where the kept variants matter most.
        data = ("ID,Name\n1,Hello World\n1,HELLO   world\n").encode()
        res = engine.clean_bytes(data, "nd.csv", drop_duplicates=False)
        self.assertEqual(res.sanity.get("near-duplicate-row"), 1)

    def test_dup_window_same_date_not_flagged(self):
        # Same-date repeats belong to the duplicate-claim rules; the
        # window scan only flags DIFFERENT dates within the window (and
        # no longer does O(k²) pair parsing on same-date pileups).
        data = ("PatientID,BillingNPI,HCPCS,DateOfService,Units\n"
                f"P1,{GOOD_A},99213,2024-01-01,1\n"
                f"P1,{GOOD_A},99213,2024-01-01,2\n"
                f"P1,{GOOD_A},99213,2024-01-01,3\n").encode()
        res = engine.clean_bytes(data, "sd.csv")
        self.assertIsNone(res.sanity.get("possible-duplicate-service"))

    def test_datetime64_series_keeps_already_date(self):
        # The unique-value memo must not change datetime64 semantics.
        import pandas as pd
        from rcm_mc.npi_cleaner.vendor_v49.npi_recovery import (
            field_validators as fv)
        s = pd.Series(pd.to_datetime(["2024-01-01", None, "2024-02-02"]))
        out = fv.validate_date_series(s, now="2026-01-01")
        self.assertEqual(out["status"].iloc[0], "ALREADY_DATE")
        self.assertFalse(out["unparseable"].iloc[0])

    def test_batch_download_ctype_and_data_route(self):
        # Batch output is a ZIP: the default download must say so, and
        # the pivot /data feed must 404 instead of parsing zip bytes.
        import io as _io
        import socket as _socket
        import threading
        import zipfile as _zf
        from rcm_mc.server import build_server
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w") as z:
            z.writestr("a.csv", "NPI\n1497758544\n")
            z.writestr("b.csv", "ClaimID,HCPCS\n1,99213\n")
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=buf.getvalue(), method="POST",
                    headers={"X-Filename": "sites.zip"})
                with _u.urlopen(req) as r:
                    jid = json.loads(r.read().decode())["job_id"]
                for _ in range(80):
                    with _u.urlopen(f"http://127.0.0.1:{port}"
                                    f"/npi-cleaner/status/{jid}") as r:
                        if json.loads(r.read().decode()).get("done"):
                            break
                    time.sleep(0.05)
                with _u.urlopen(f"http://127.0.0.1:{port}"
                                f"/npi-cleaner/download/{jid}") as r:
                    self.assertEqual(r.headers.get("Content-Type"),
                                     "application/zip")
                try:
                    _u.urlopen(f"http://127.0.0.1:{port}"
                               f"/npi-cleaner/data/{jid}")
                    self.fail("expected 404 for batch pivot data")
                except _u.HTTPError as exc:
                    self.assertEqual(exc.code, 404)
            finally:
                server.shutdown()
                server.server_close()


class TestConsolidation(unittest.TestCase):
    """Merge-confidence checks: registry↔engine consistency, OpenAPI
    coverage of every API route, and the full golden-path walkthrough."""

    def test_registry_engine_consistency(self):
        # Every sanity key the engine can emit exists in the registry, and
        # every registry id appears in the engine source (no dead entries).
        import inspect
        import re as _re
        from rcm_mc.npi_cleaner import rules
        src = inspect.getsource(engine)
        emitted = set(_re.findall(r'res\.sanity\[\s*"([a-z0-9-]+)"', src))
        # These three are emitted through a loop variable, not a literal.
        emitted |= {"condition-code-malformed", "occurrence-code-malformed",
                    "value-code-malformed"}
        known = {r.id for r in rules.all_rules()}
        self.assertEqual(sorted(emitted - known), [],
                         "engine emits sanity keys missing from rules.py")
        dead = [i for i in sorted(known) if f'"{i}"' not in src]
        self.assertEqual(dead, [],
                         "rules.py ids never emitted by the engine")

    def test_openapi_covers_all_api_routes(self):
        # Every /npi-cleaner/api/* route literal in server.py must be a
        # documented path, plus the async upload/status/download pipeline.
        import re as _re
        import rcm_mc.server as _srv
        from rcm_mc.infra.openapi import get_openapi_spec
        src = open(_srv.__file__, encoding="utf-8").read()
        routes = set(_re.findall(r'"(/npi-cleaner/api/[a-z/]+)"', src))
        self.assertTrue(routes)
        spec = set(get_openapi_spec()["paths"])
        missing = sorted(r for r in routes if r not in spec)
        self.assertEqual(missing, [])
        for core in ("/npi-cleaner/upload", "/npi-cleaner/status/{job_id}",
                     "/npi-cleaner/download/{job_id}"):
            self.assertIn(core, spec)

    def test_golden_path_walkthrough(self):
        # The whole platform in one pass against a real server: clean an
        # 837, pull the bundle, reconcile against its 835, see both runs
        # in history, and get a trend alert on a worse re-upload.
        import io as _io
        import socket as _socket
        import threading
        import uuid as _uuid
        import zipfile as _zf
        from rcm_mc.server import build_server
        tag = _uuid.uuid4().hex[:8]
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            base = f"http://127.0.0.1:{port}"

            def run(payload, name):
                req = _u.Request(f"{base}/npi-cleaner/upload", data=payload,
                                 method="POST",
                                 headers={"X-Filename": name})
                with _u.urlopen(req) as r:
                    jid = json.loads(r.read().decode())["job_id"]
                for _ in range(80):
                    with _u.urlopen(f"{base}/npi-cleaner/status/{jid}") as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            return jid, j["scorecard"]
                    time.sleep(0.05)
                self.fail("job never finished")

            try:
                # 1. Clean the 837 — grade, claim rollup, dictionary.
                a, sc_a = run(_X12_837P, f"claims-{tag}.837")
                self.assertEqual(sc_a["delimiter"], "X12 837 (EDI)")
                self.assertIn("quality", sc_a)
                self.assertEqual(sc_a["claims"]["n_claims"], 2)
                self.assertTrue(sc_a["dictionary"])
                # 2. Everything-bundle carries every artifact.
                with _u.urlopen(f"{base}/npi-cleaner/download/{a}"
                                "?fmt=bundle") as r:
                    blob = r.read()
                with _zf.ZipFile(_io.BytesIO(blob)) as z:
                    names = z.namelist()
                for expected in ("exec_report.html", "scorecard.json",
                                 "data_dictionary.csv"):
                    self.assertIn(expected, names)
                # 3. Clean its 835 and reconcile — full match.
                b, sc_b = run(_X12_835, f"remit-{tag}.835")
                self.assertEqual(sc_b["delimiter"], "X12 835 (ERA)")
                req = _u.Request(
                    f"{base}/npi-cleaner/api/reconcile",
                    data=json.dumps({"a": a, "b": b}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    rec = json.loads(r.read().decode())
                self.assertEqual(rec["matched"], 2)
                self.assertEqual(rec["match_rate_pct"], 100.0)
                # 4. History recorded both runs.
                with _u.urlopen(f"{base}/npi-cleaner/api/history") as r:
                    runs = json.loads(r.read().decode())["runs"]
                seen = {x["file_name"] for x in runs}
                self.assertIn(f"claims-{tag}.837", seen)
                self.assertIn(f"remit-{tag}.835", seen)
                # 5. A worse re-upload of the same file trips trend alerts.
                good = ("ClaimID,HCPCS\n"
                        + "".join(f"{i},99213\n" for i in range(40))).encode()
                bad = ("ClaimID,HCPCS\n"
                       + "".join(f"{i},BAD!!\n" for i in range(40))).encode()
                run(good, f"nightly-{tag}.csv")
                _, sc_bad = run(bad, f"nightly-{tag}.csv")
                self.assertTrue(sc_bad["trend_alerts"])
                self.assertTrue(any("hcpcs-malformed" in x
                                    for x in sc_bad["trend_alerts"]))
            finally:
                server.shutdown()
                server.server_close()


class TestHardening(unittest.TestCase):
    """Adversarial inputs must degrade to a warning, never crash or OOM."""

    def test_zip_bomb_rejected_by_declared_size(self):
        # A tiny compressed archive declaring more than the uncompressed
        # cap is rejected BEFORE any decompression. The cap is lowered for
        # the test so the fixture stays small.
        import io as _io
        import zipfile as _zf
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as z:
            z.writestr("bomb.csv", "0,0\n" * 700_000)   # ~2.8 MB declared
        data = buf.getvalue()
        old = engine._BATCH_MAX_UNCOMPRESSED
        engine._BATCH_MAX_UNCOMPRESSED = 1 * 1024 * 1024
        try:
            with self.assertRaises(ValueError):
                engine.zip_batch_members(data)
            res = engine.clean_bytes(data, "bomb.zip")
            self.assertTrue(any("batch limit" in w for w in res.warnings))
            self.assertEqual(res.n_rows_in, 0)
        finally:
            engine._BATCH_MAX_UNCOMPRESSED = old

    def test_adversarial_inputs_never_crash(self):
        import io as _io
        import zipfile as _zf
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w") as z:
            z.writestr("a.csv", "NPI\n1\n" * 100)
        cases = {
            "truncated.zip": buf.getvalue()[:60],
            "quotes.csv": b'NPI,Name\n123,"unclosed\n456,ok\n',
            "one-col.csv": b"NPI\n1497758544\n99999\n",
            "header-only.csv": b"NPI,Name\n",
            "empty.csv": b"",
            "garbage.bin": bytes(range(256)) * 10,
            "nulls.csv": b"NPI,Na\x00me\n123,\x00x\n",
            "long-line.csv": b"NPI\n" + b"9" * 5_000_000 + b"\n",
            "x12-junk.837": b"ISA*junk",
            "ragged.csv": b"A,B,C\n1\n1,2,3,4,5,6\n",
        }
        for name, data in cases.items():
            res = engine.clean_bytes(data, name)   # must not raise
            self.assertIsNotNone(res.out_name, name)

    def test_cli_zip_batch_and_bundle(self):
        import contextlib
        import io as _io
        import zipfile as _zf
        from rcm_mc.npi_cleaner import cli as nc_cli
        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, "sites.zip")
            buf = _io.BytesIO()
            with _zf.ZipFile(buf, "w") as z:
                z.writestr("a.csv", f"NPI,ChargeAmt\n{GOOD_A},\"$1,0\"\n")
                z.writestr("b.csv", "ClaimID,HCPCS\n1,BAD!!\n")
            with open(src, "wb") as fh:
                fh.write(buf.getvalue())
            out = _io.StringIO()
            with contextlib.redirect_stdout(out):
                rc = nc_cli.main([src, "--json", "--outdir", tmp])
            self.assertEqual(rc, 0)
            sc = json.loads(out.getvalue())
            self.assertEqual(len(sc["batch"]), 2)
            self.assertEqual(sc["delimiter"], "zip batch")
            # --bundle on a plain CSV writes the everything-zip locally.
            src2 = os.path.join(tmp, "in.csv")
            with open(src2, "w", encoding="utf-8") as fh:
                fh.write("ClaimID,HCPCS\n1,99213\n2,BAD!!\n")
            out2 = _io.StringIO()
            with contextlib.redirect_stdout(out2):
                rc2 = nc_cli.main([src2, "--bundle", "--json",
                                   "--outdir", tmp])
            self.assertEqual(rc2, 0)
            bundle = os.path.join(tmp, "in_bundle.zip")
            self.assertTrue(os.path.exists(bundle))
            with _zf.ZipFile(bundle) as z:
                names = z.namelist()
            self.assertIn("exec_report.html", names)
            self.assertIn("scorecard.json", names)
            self.assertIn("data_dictionary.csv", names)
            self.assertIn("worklists/hcpcs-malformed.csv", names)


class TestZipBatchAndPayerWorklist(unittest.TestCase):
    """Batch 15: multi-file zip batch upload + per-payer worklists."""

    def _zip(self, entries):
        import io as _io
        import zipfile as _zf
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w") as z:
            for name, text in entries:
                z.writestr(name, text)
        return buf.getvalue()

    def test_zip_batch_detection(self):
        data = self._zip([("a.csv", "NPI\n1\n"),
                          ("notes/readme.md", "skip"),
                          ("__MACOSX/j.csv", "skip"),
                          (".hidden.csv", "skip")])
        members = engine.zip_batch_members(data)
        self.assertEqual([m[0] for m in members], ["a.csv"])
        # An xlsx is also a zip — must NOT be treated as a batch.
        openpyxl = __import__("openpyxl")
        from io import BytesIO
        wb = openpyxl.Workbook()
        wb.active.append(["NPI"])
        xb = BytesIO()
        wb.save(xb)
        self.assertIsNone(engine.zip_batch_members(xb.getvalue()))
        self.assertIsNone(engine.zip_batch_members(b"NPI\n1\n"))
        self.assertIsNone(engine.zip_batch_members(
            self._zip([("only.md", "no claim files")])))

    def test_zip_batch_clean(self):
        data = self._zip([
            ("siteA.csv", f"NPI,ChargeAmt\n{GOOD_A},\"$100.00\"\n99999,50\n"),
            ("siteB.csv", "ClaimID,HCPCS\n1,99213\n2,BAD!!\n")])
        res = engine.clean_bytes(data, "sites.zip")
        sc = res.as_scorecard()
        self.assertEqual(sc["delimiter"], "zip batch")
        self.assertEqual(len(sc["batch"]), 2)
        self.assertEqual(sc["rows_in"], 4)
        by_file = {b["file"]: b for b in sc["batch"]}
        self.assertEqual(by_file["siteB.csv"]["findings"], 1)
        self.assertGreater(by_file["siteA.csv"]["repairs"], 0)
        # Merged counters: the hcpcs flag from siteB is in the parent.
        self.assertEqual(sc["sanity"].get("hcpcs-malformed"), 1)
        # Output is a zip of the cleaned member files.
        import io as _io
        import zipfile as _zf
        self.assertTrue(res.out_name.endswith("_cleaned.zip"))
        with _zf.ZipFile(res.out_path) as z:
            self.assertEqual(len(z.namelist()), 2)
            inner = z.read(z.namelist()[0]).decode()
        self.assertIn(",", inner)

    def test_payer_worklist_route(self):
        import socket as _socket
        import threading
        from rcm_mc.server import build_server
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                csv_b = ("PayerName,HCPCS\nUHC,BAD!!\nUHC,99213\n"
                         "Aetna,99214\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "pw.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(f"http://127.0.0.1:{port}"
                                    f"/npi-cleaner/status/{job_id}") as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                self.assertEqual(j["scorecard"]["payer_worklists"],
                                 {"UNITEDHEALTHCARE": 1})
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=worklist&payer=UNITEDHEALTHCARE"
                ) as r:
                    body = r.read().decode()
                self.assertIn("BAD!!", body)
                self.assertNotIn("99214", body)   # Aetna row excluded
            finally:
                server.shutdown()
                server.server_close()

    def test_date_memo_speedup_still_correct(self):
        # The unique-value memo in the v49 date validator must not change
        # results: repeated + distinct + bad dates all classify the same.
        import pandas as pd
        from rcm_mc.npi_cleaner.vendor_v49.npi_recovery import (
            field_validators as fv)
        s = pd.Series(["2024-01-01", "2024-01-01", "01/15/2024",
                       "20240116", "not-a-date", None, "45000"])
        out = fv.validate_date_series(s, now="2026-01-01")
        self.assertEqual(list(out["status"])[:2], ["ISO", "ISO"])
        self.assertEqual(out["status"].iloc[2], "US_SLASH")
        self.assertEqual(out["status"].iloc[3], "COMPACT")
        self.assertTrue(out["unparseable"].iloc[4])
        self.assertEqual(out["status"].iloc[5], "BLANK")
        self.assertTrue(out["excel_serial"].iloc[6])


class TestPayerQualityAndPortableSuites(unittest.TestCase):
    """Batch 14: per-payer quality split + portable profiles/mappings."""

    def test_payer_quality_split(self):
        data = ("PayerName,HCPCS\n"
                "UnitedHealthcare,99213\n"
                "UHC,BAD!!\n"          # UHC + UnitedHealthcare = one family
                "Aetna,99214\n"
                "Aetna,99215\n").encode()
        res = engine.clean_bytes(data, "pq.csv")
        pq = {p["payer"]: p for p in res.as_scorecard()["payer_quality"]}
        self.assertEqual(pq["UNITEDHEALTHCARE"]["rows"], 2)
        self.assertEqual(pq["UNITEDHEALTHCARE"]["flagged"], 1)
        self.assertEqual(pq["UNITEDHEALTHCARE"]["clean_pct"], 50.0)
        self.assertEqual(pq["UNITEDHEALTHCARE"]["top_rules"][0]["rule"],
                         "hcpcs-malformed")
        self.assertEqual(pq["AETNA"]["clean_pct"], 100.0)
        # The refactored per-row capture still feeds worklists.
        self.assertEqual(res.flag_rows["hcpcs-malformed"], [2])
        # No payer column → no split, not a crash.
        res2 = engine.clean_bytes(b"HCPCS\n99213\n", "n.csv")
        self.assertIsNone(res2.as_scorecard()["payer_quality"])
        # Exec report renders the payer table.
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        html_out = build_exec_report(res.as_scorecard(), "pq.csv",
                                     "2026-01-01T00:00:00+00:00")
        self.assertIn("Quality by payer", html_out)

    def test_profiles_export_import_roundtrip(self):
        import uuid as _uuid
        from rcm_mc.npi_cleaner import profiles
        pn = f"exp-{_uuid.uuid4().hex[:8]}"
        profiles.save_profile(pn, {"accepted_rules": ["charge-outlier"],
                                   "thresholds": {"timely_filing_days": 90}})
        try:
            mine = [p for p in profiles.export_all()["profiles"]
                    if p["name"] == pn]
            self.assertEqual(
                mine[0]["config"]["thresholds"]["timely_filing_days"], 90)
            profiles.delete_profile(pn)
            rep = profiles.import_all({"profiles": mine})
            self.assertEqual(rep["imported"], 1)
            self.assertEqual(profiles.get_profile(pn)["accepted_rules"],
                             ["charge-outlier"])
            with self.assertRaises(ValueError):
                profiles.import_all({"wrong": []})
        finally:
            profiles.delete_profile(pn)

    def test_mappings_export_import_roundtrip(self):
        import uuid as _uuid
        from rcm_mc.npi_cleaner import mappings
        mn = f"map-{_uuid.uuid4().hex[:8]}"
        mappings.save_mapping(mn, {"billing_npi": "Col A"})
        try:
            mine = [m for m in mappings.export_all()["mappings"]
                    if m["name"] == mn]
            mappings.delete_mapping(mn)
            rep = mappings.import_all({"mappings": mine})
            self.assertEqual(rep["imported"], 1)
            self.assertEqual(mappings.get_mapping(mn),
                             {"billing_npi": "Col A"})
        finally:
            mappings.delete_mapping(mn)

    def test_export_import_http_routes(self):
        import socket as _socket
        import threading
        import uuid as _uuid
        from rcm_mc.server import build_server
        from rcm_mc.npi_cleaner import profiles
        pn = f"http-{_uuid.uuid4().hex[:8]}"
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                profiles.save_profile(pn, {"thresholds":
                                           {"stale_years": 5}})
                with _u.urlopen(f"http://127.0.0.1:{port}"
                                "/npi-cleaner/api/profiles/export") as r:
                    self.assertIn("attachment",
                                  r.headers.get("Content-Disposition", ""))
                    exported = json.loads(r.read().decode())
                self.assertTrue(any(p["name"] == pn
                                    for p in exported["profiles"]))
                profiles.delete_profile(pn)
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/profiles/import",
                    data=json.dumps(exported).encode(), method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    rep = json.loads(r.read().decode())
                self.assertGreaterEqual(rep["imported"], 1)
                self.assertIsNotNone(profiles.get_profile(pn))
            finally:
                profiles.delete_profile(pn)
                server.shutdown()
                server.server_close()


class TestReconcile(unittest.TestCase):
    """837↔835 reconciliation: match a claims run against its remittance
    on claim id — unpaid claims, paid-vs-billed variance, denial mix."""

    def test_reconcile_pure(self):
        from rcm_mc.npi_cleaner.reconcile import reconcile
        ha = ["ClaimID", "BilledAmt"]
        ra = [["C1", "100"], ["C1", "50"], ["C2", "300"], ["C3", "75"]]
        hb = ["ClaimID", "PaidAmt", "DenialCodes"]
        rb = [["C1", "120", "45"], ["C2", "0", "29"], ["C9", "40", ""]]
        rep = reconcile(ha, ra, hb, rb)
        self.assertEqual(rep["claims_a"], 3)
        self.assertEqual(rep["matched"], 2)
        self.assertEqual(rep["match_rate_pct"], 66.7)
        self.assertEqual(rep["unpaid_count"], 1)
        self.assertEqual(rep["unpaid"][0]["claim"], "C3")
        self.assertEqual(rep["orphan_remits_count"], 1)
        self.assertEqual(rep["billed_matched"], 450.0)  # C1 150 + C2 300
        self.assertEqual(rep["paid_matched"], 120.0)
        self.assertEqual(rep["variance_total"], 330.0)
        # C2 is the biggest variance (300 billed, 0 paid, CARC 29).
        self.assertEqual(rep["top_variance"][0]["claim"], "C2")
        self.assertIn("29", rep["top_variance"][0]["carcs"])
        den = {d["code"]: d for d in rep["denials"]}
        self.assertEqual(den["29"]["category"], "preventable")
        # Missing claim-id column → clear error, not a crash.
        bad = reconcile(["A"], [["1"]], hb, rb)
        self.assertIn("claim-id", bad["error"])

    def test_reconcile_http_837_vs_835(self):
        # The 837P and 835 fixtures share ACCT001/ACCT002 — a real
        # claims-vs-remit pair, end to end through the server.
        import socket as _socket
        import threading
        from rcm_mc.server import build_server
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                def _run(payload, name):
                    req = _u.Request(
                        f"http://127.0.0.1:{port}/npi-cleaner/upload",
                        data=payload, method="POST",
                        headers={"X-Filename": name})
                    with _u.urlopen(req) as r:
                        jid = json.loads(r.read().decode())["job_id"]
                    for _ in range(50):
                        with _u.urlopen(
                            f"http://127.0.0.1:{port}"
                            f"/npi-cleaner/status/{jid}") as r:
                            if json.loads(r.read().decode()).get("done"):
                                return jid
                        time.sleep(0.05)
                    return jid
                a = _run(_X12_837P, "claims.837")
                b = _run(_X12_835, "remit.835")
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/reconcile",
                    data=json.dumps({"a": a, "b": b}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    rep = json.loads(r.read().decode())
                self.assertEqual(rep["claims_a"], 2)
                self.assertEqual(rep["matched"], 2)
                self.assertEqual(rep["match_rate_pct"], 100.0)
                self.assertEqual(rep["unpaid_count"], 0)
                den_codes = {d["code"] for d in rep["denials"]}
                self.assertIn("45", den_codes)
                self.assertIn("29", den_codes)
                # Unknown job → 404 with a JSON error.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/reconcile",
                    data=json.dumps({"a": "nope", "b": b}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                try:
                    _u.urlopen(req)
                    self.fail("expected 404")
                except _u.HTTPError as exc:
                    self.assertEqual(exc.code, 404)
            finally:
                server.shutdown()
                server.server_close()


class TestDenialPlaybookTrendsDictionary(unittest.TestCase):
    """Batch 13: playbook-enriched denials, trend alerts vs the previous
    run, and the PHI-safe data dictionary."""

    def test_denial_playbook_enrichment(self):
        data = ("ClaimID,DenialCode,BilledAmt,PaidAmt\n"
                "1,45,100,80\n2,45,90,70\n3,29,50,0\n"
                "4,16,60,0\n5,1,70,50\n").encode()
        sc = engine.clean_bytes(data, "den.csv").as_scorecard()
        top = {e["code"]: e for e in sc["denials"]["top"]}
        self.assertEqual(top["45"]["category"], "contractual")
        self.assertEqual(top["29"]["category"], "preventable")
        self.assertEqual(top["29"]["linked_rule"], "timely-filing-risk")
        self.assertEqual(top["1"]["category"], "patient-responsibility")
        # 29 + 16 preventable out of 5 classified mentions → 40%.
        self.assertEqual(sc["denials"]["preventable_pct"], 40.0)
        from rcm_mc.npi_cleaner import refdata
        self.assertIsNone(refdata.carc_playbook("999"))
        # Exec report renders the playbook column + preventable line.
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        html_out = build_exec_report(sc, "den.csv",
                                     "2026-01-01T00:00:00+00:00")
        self.assertIn("preventable", html_out)
        self.assertIn("timely-filing", html_out)

    def test_trend_alerts_vs_previous_run(self):
        import uuid as _uuid
        fname = f"trend-{_uuid.uuid4().hex[:8]}.csv"
        good = ("ClaimID,HCPCS\n"
                + "".join(f"{i},99213\n" for i in range(50))).encode()
        bad = ("ClaimID,HCPCS\n"
               + "".join(f"{i},BAD!!\n" for i in range(50))).encode()
        r1 = engine.clean_bytes(good, fname)
        self.assertEqual(r1.trend_alerts, [])   # first run: no baseline
        r2 = engine.clean_bytes(bad, fname)
        self.assertTrue(any("hcpcs-malformed" in a for a in r2.trend_alerts))
        self.assertTrue(any("score dropped" in a.lower()
                            for a in r2.trend_alerts))
        self.assertIn("trend_alerts", r2.as_scorecard())

    def test_data_dictionary(self):
        data = ("ClaimID,DenialCode,BilledAmt\n1,45,100\n2,29,50\n").encode()
        res = engine.clean_bytes(data, "dict.csv")
        dic = {e["column"]: e for e in res.as_scorecard()["dictionary"]}
        self.assertEqual(dic["DenialCode"]["role"], "carc/denial")
        self.assertEqual(dic["BilledAmt"]["role"], "money")
        self.assertEqual(dic["ClaimID"]["role"], "claim-id")
        self.assertEqual(dic["BilledAmt"]["fill_pct"], 100.0)
        self.assertEqual(dic["ClaimID"]["distinct"], 2)
        text = engine.dictionary_csv(res)
        self.assertIn("detected_role", text)
        # PHI columns never leak raw samples into the dictionary.
        res2 = engine.clean_bytes(
            b"PatientName,NPI\nJohn Doe,1497758544\n", "p.csv")
        dic2 = {e["column"]: e
                for e in res2.as_scorecard()["dictionary"]}
        self.assertEqual(dic2["PatientName"]["samples"],
                         ["(redacted — patient identifier)"])
        self.assertNotIn("John Doe", engine.dictionary_csv(res2))

    def test_dictionary_download_route(self):
        # Exercised through the running server via the download route.
        import socket as _socket
        import threading
        from rcm_mc.server import build_server
        with tempfile.TemporaryDirectory() as tmp:
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=b"ClaimID,HCPCS\n1,99213\n", method="POST",
                    headers={"X-Filename": "d.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(f"http://127.0.0.1:{port}"
                                    f"/npi-cleaner/status/{job_id}") as r:
                        if json.loads(r.read().decode()).get("done"):
                            break
                    time.sleep(0.05)
                with _u.urlopen(f"http://127.0.0.1:{port}/npi-cleaner"
                                f"/download/{job_id}?fmt=dictionary") as r:
                    body = r.read().decode()
                self.assertIn("detected_role", body)
                self.assertIn("hcpcs/cpt", body)
            finally:
                server.shutdown()
                server.server_close()


class TestMappingTemplates(unittest.TestCase):
    """Named column-mapping templates (mappings.py) — map a source system
    once, reuse per upload via X-Mapping."""

    def test_save_sanitize_get_delete(self):
        import uuid as _uuid
        from rcm_mc.npi_cleaner import mappings
        name = f"tpl-{_uuid.uuid4().hex[:8]}"
        stored = mappings.save_mapping(name, {
            "billing_npi": "Prov_Billing_Num",
            "BAD ROLE!": "X",                      # bad role token → dropped
            "state": "",                           # empty header → dropped
            "drug_name": "Y" * 500,                # over-long header → dropped
            "custom_role_9": "SrcCol",             # unknown roles pass through
        })
        self.assertEqual(stored, {"billing_npi": "Prov_Billing_Num",
                                  "custom_role_9": "SrcCol"})
        got = mappings.get_mapping(name)
        self.assertEqual(got["billing_npi"], "Prov_Billing_Num")
        listed = {m["name"]: m for m in mappings.list_mappings()}
        self.assertIn(name, listed)
        self.assertEqual(listed[name]["roles"], 2)
        self.assertTrue(mappings.delete_mapping(name))
        self.assertIsNone(mappings.get_mapping(name))
        # A template that sanitizes to nothing is rejected outright.
        with self.assertRaises(ValueError):
            mappings.save_mapping(f"tpl-{_uuid.uuid4().hex[:8]}",
                                  {"BAD!": "x"})
        with self.assertRaises(ValueError):
            mappings.save_mapping("", {"billing_npi": "A"})


class TestNpiCleanerHttp(unittest.TestCase):
    def _start(self, tmp):
        import socket as _socket
        import threading
        from rcm_mc.server import build_server
        s = _socket.socket()
        s.bind(("127.0.0.1", 0))
        port = s.getsockname()[1]
        s.close()
        server, _ = build_server(port=port, db_path=os.path.join(tmp, "p.db"))
        t = threading.Thread(target=server.serve_forever, daemon=True)
        t.start()
        time.sleep(0.05)
        return server, port

    def test_page_renders(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                with _u.urlopen(f"http://127.0.0.1:{port}/npi-cleaner") as r:
                    body = r.read().decode()
                self.assertIn("NPI Claims Cleaner", body)
                self.assertIn("/npi-cleaner/upload", body)
                self.assertIn("npi-drop", body)
                # Tabbed results + live-connector affordances present.
                self.assertIn("npi-tabs", body)
                self.assertIn('data-panel="connectors"', body)
                self.assertIn("Connections available", body)
                self.assertIn("Go online", body)
            finally:
                server.shutdown()
                server.server_close()

    def test_changelog_download_route(self):
        # ?fmt=changelog streams the audit trail once a job with changes ran.
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv_b = ("NPI,ChargeAmt\n"
                         f"  {GOOD_A}  ,\"$1,250.50\"\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "c.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                self.assertTrue(j["scorecard"]["changelog_name"])
                self.assertIn("quality", j["scorecard"])
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=changelog"
                ) as r:
                    body = r.read().decode()
                self.assertIn("before", body.splitlines()[0])
                self.assertIn("1250.50", body)
            finally:
                server.shutdown()
                server.server_close()

    def test_mapping_template_routes_and_x_mapping_upload(self):
        # Save a template over HTTP, then upload with X-Mapping and confirm
        # the template's billing_npi role is honored — "Prov_Billing_Num"
        # has no "npi" in it, so auto-detection alone would miss it.
        import uuid as _uuid
        tpl = f"epic-{_uuid.uuid4().hex[:8]}"
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/mappings",
                    data=json.dumps({"name": tpl, "mapping": {
                        "billing_npi": "Prov_Billing_Num"}}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    saved = json.loads(r.read().decode())
                self.assertTrue(saved["ok"])
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/mappings") as r:
                    listed = json.loads(r.read().decode())["mappings"]
                self.assertTrue(any(m["name"] == tpl for m in listed))

                csv_b = ("ClaimID,Prov_Billing_Num\n"
                         f"1,{GOOD_A}\n2,99999\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "m.csv", "X-Mapping": tpl})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                sc = j["scorecard"]
                self.assertEqual(sc["billing_column"], "Prov_Billing_Num")
                self.assertIn("Prov_Billing_Num", sc["npi_columns"])

                # Delete route cleans up; the template disappears from list.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/mappings/delete",
                    data=json.dumps({"name": tpl}).encode(), method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    self.assertTrue(json.loads(r.read().decode())["ok"])
            finally:
                server.shutdown()
                server.server_close()

    def test_bundle_download_route(self):
        # ?fmt=bundle → one zip with the cleaned file, workbook, exec
        # report, scorecard JSON, and per-rule worklists.
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv_b = ("ClaimID,HCPCS\n1,99213\n2,BAD!!\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "b.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=bundle"
                ) as r:
                    self.assertEqual(
                        r.headers.get("Content-Type"), "application/zip")
                    blob = r.read()
                import io as _io
                import zipfile as _zf
                with _zf.ZipFile(_io.BytesIO(blob)) as z:
                    names = z.namelist()
                    self.assertIn("exec_report.html", names)
                    self.assertIn("scorecard.json", names)
                    self.assertIn(j["scorecard"]["out_name"], names)
                    self.assertIn("worklists/hcpcs-malformed.csv", names)
                    wl = z.read("worklists/hcpcs-malformed.csv").decode()
                    self.assertIn("BAD!!", wl)
                    sc_j = json.loads(z.read("scorecard.json").decode())
                    self.assertIn("quality", sc_j)
            finally:
                server.shutdown()
                server.server_close()

    def test_cli_profile_and_mapping_flags(self):
        import uuid as _uuid
        from rcm_mc.npi_cleaner import cli as nc_cli
        from rcm_mc.npi_cleaner import mappings
        tpl = f"cli-{_uuid.uuid4().hex[:8]}"
        mappings.save_mapping(tpl, {"billing_npi": "Prov_Billing_Num"})
        try:
            with tempfile.TemporaryDirectory() as tmp:
                src = os.path.join(tmp, "in.csv")
                with open(src, "w", encoding="utf-8") as fh:
                    fh.write(f"ClaimID,Prov_Billing_Num\n1,{GOOD_A}\n")
                import contextlib
                import io as _io
                out = _io.StringIO()
                with contextlib.redirect_stdout(out):
                    rc = nc_cli.main([src, "--mapping", tpl, "--json",
                                      "--outdir", tmp])
                self.assertEqual(rc, 0)
                sc = json.loads(out.getvalue())
                self.assertEqual(sc["billing_column"], "Prov_Billing_Num")
                # Unknown names exit 2 with a clear error, not a traceback.
                self.assertEqual(
                    nc_cli.main([src, "--mapping", "no-such-tpl"]), 2)
                self.assertEqual(
                    nc_cli.main([src, "--profile", "no-such-prof"]), 2)
        finally:
            mappings.delete_mapping(tpl)

    def test_openapi_documents_npi_cleaner(self):
        from rcm_mc.infra.openapi import get_openapi_spec
        paths = get_openapi_spec()["paths"]
        for p in ("/npi-cleaner/api/clean", "/npi-cleaner/api/rules",
                  "/npi-cleaner/api/profiles", "/npi-cleaner/api/mappings",
                  "/npi-cleaner/api/history"):
            self.assertIn(p, paths)

    def test_worklist_exec_history_rules_routes(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv_b = ("ClaimID,HCPCS\n1,99213\n2,BAD!!\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "w.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                # Worklist: just the flagged row, with a _row column.
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=worklist&rule=hcpcs-malformed"
                ) as r:
                    wl = r.read().decode()
                self.assertIn("_row", wl.splitlines()[0])
                self.assertIn("BAD!!", wl)
                self.assertNotIn("99213", wl)      # unflagged row excluded
                # Executive report renders.
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=exec"
                ) as r:
                    self.assertIn("Claims data-quality report",
                                  r.read().decode())
                # History page + API + rules API.
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/history") as r:
                    _hist_html = r.read().decode()
                self.assertIn("Quality-score trend", _hist_html)
                self.assertIn("Per-rule trend", _hist_html)
                self.assertIn("nh-rule-box", _hist_html)
                self.assertIn("Dimension trends", _hist_html)
                self.assertIn("nh-dims-box", _hist_html)
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/history") as r:
                    runs = json.loads(r.read().decode())["runs"]
                self.assertTrue(any(x["file_name"] == "w.csv" for x in runs))
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/rules") as r:
                    self.assertGreater(
                        len(json.loads(r.read().decode())["rules"]), 50)
            finally:
                server.shutdown()
                server.server_close()

    def test_profiles_api_and_sync_clean(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                # Save a profile over the API.
                body = json.dumps({"name": "http-prof", "config": {
                    "accepted_rules": ["date-stale"],
                    "thresholds": {"stale_years": 2}}}).encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/profiles",
                    data=body, method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    self.assertTrue(json.loads(r.read().decode())["ok"])
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/profiles") as r:
                    names = [p["name"] for p in
                             json.loads(r.read().decode())["profiles"]]
                self.assertIn("http-prof", names)
                # Async upload honors X-Profile.
                csv_b = ("ClaimID,DateOfService\n1,2023-01-01\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "p.csv", "X-Profile": "http-prof"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                self.assertEqual(j["scorecard"]["profile"], "http-prof")
                self.assertIn("date-stale", j["scorecard"]["sanity"])
                self.assertIn("date-stale", j["scorecard"]["accepted_rules"])
                # Synchronous API clean with the same profile.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/clean"
                    "?profile=http-prof",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "sync.csv"})
                with _u.urlopen(req) as r:
                    sc = json.loads(r.read().decode())
                self.assertIn("quality", sc)
                self.assertEqual(sc["profile"], "http-prof")
                # Cleanup.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/profiles/delete",
                    data=json.dumps({"name": "http-prof"}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    self.assertTrue(json.loads(r.read().decode())["ok"])
            finally:
                server.shutdown()
                server.server_close()

    def test_large_upload_accepted_others_capped(self):
        # Claims uploads above the 10 MB global POST cap are accepted on the
        # cleaner upload route (200 MB ceiling); every other POST stays capped.
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                header = b"ClaimID,BillingNPI\n"
                row = b"1," + GOOD_B.encode() + b"\n"
                big = header + row * ((11 * 1024 * 1024) // len(row) + 1)
                self.assertGreater(len(big), 10_000_000)  # above the old cap
                # Upload route accepts it → returns a job id.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=big, method="POST", headers={"X-Filename": "big.csv"})
                with _u.urlopen(req) as r:
                    self.assertIn("job_id", json.loads(r.read().decode()))
                # An ordinary POST path still rejects the same oversized body.
                # The guard rejects early on Content-Length and closes the
                # socket, so the client sees either a clean 413 or a broken
                # pipe mid-upload — both prove the body was refused (and the
                # broken pipe is the desired don't-drain-the-body behavior).
                req2 = _u.Request(
                    f"http://127.0.0.1:{port}/__cap_probe__",
                    data=big, method="POST")
                try:
                    _u.urlopen(req2)
                    self.fail("expected rejection on a non-upload POST path")
                except _u.HTTPError as exc:  # type: ignore[attr-defined]
                    self.assertEqual(exc.code, 413)
                except (_ue.URLError, ConnectionError):
                    pass  # server closed the connection early — refused
            finally:
                server.shutdown()
                server.server_close()

    def test_upload_status_download_loop(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv = (
                    "ClaimID,BillingNPI\n"
                    f"1,{GOOD_A}\n"
                    "2,\n"
                    "3,badnpi\n"
                ).encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv, method="POST",
                    headers={"X-Filename": "claims%20file.csv"},
                )
                with _u.urlopen(req) as r:
                    up = json.loads(r.read().decode())
                job_id = up["job_id"]
                self.assertTrue(job_id)

                # Poll to completion.
                sc = None
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        st = json.loads(r.read().decode())
                    if st.get("done"):
                        sc = st.get("scorecard")
                        break
                    time.sleep(0.05)
                self.assertIsNotNone(sc, "job never completed")
                self.assertEqual(sc["rows_in"], 3)
                self.assertEqual(sc["billing_column"], "BillingNPI")
                self.assertEqual(sc["billing_issues"], 2)  # blank + malformed
                self.assertEqual(sc["out_name"], "claims_file_cleaned.csv")

                # Download the cleaned file.
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                ) as r:
                    self.assertIn("attachment", r.headers.get("Content-Disposition", ""))
                    out = r.read().decode()
                self.assertIn("ClaimID,BillingNPI", out)
            finally:
                server.shutdown()
                server.server_close()

    def test_detect_and_override_flow(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv = (
                    "Claim,ProvID,OrgNm,St,AllowedAmt\n"
                    f"1,{GOOD_A},Mercy,OH,80\n"
                    "2,99999,Mercy,OH,40\n"
                ).encode()
                # detect returns headers + a role mapping
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/detect",
                    data=csv, method="POST",
                    headers={"X-Filename": "c.csv"})
                with _u.urlopen(req) as r:
                    det = json.loads(r.read().decode())
                if not det.get("available"):
                    self.skipTest("detector unavailable")
                self.assertIn("ProvID", det["headers"])
                self.assertTrue(any(role["key"] == "billing_npi"
                                    for role in det["roles"]))

                # upload with an explicit override → billing column honored
                ov = json.dumps({"billing_npi": "ProvID", "state": "St"})
                req2 = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv, method="POST",
                    headers={"X-Filename": "c.csv",
                             "X-Overrides": _up.quote(ov)})
                with _u.urlopen(req2) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                sc = None
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        st = json.loads(r.read().decode())
                    if st.get("done"):
                        sc = st["scorecard"]
                        break
                    time.sleep(0.05)
                self.assertIsNotNone(sc)
                self.assertEqual(sc["billing_column"], "ProvID")
                self.assertEqual(sc["billing_issues"], 1)  # the 99999
            finally:
                server.shutdown()
                server.server_close()

    def test_sample_and_xlsx_download(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                # The sample endpoint returns a usable CSV.
                with _u.urlopen(f"http://127.0.0.1:{port}/npi-cleaner/sample") as r:
                    sample = r.read()
                self.assertIn(b"BillingProviderNPI", sample)

                # Upload it and pull the .xlsx workbook back.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=sample, method="POST",
                    headers={"X-Filename": "sample_claims.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                sc = None
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        st = json.loads(r.read().decode())
                    if st.get("done"):
                        sc = st["scorecard"]
                        break
                    time.sleep(0.05)
                self.assertIsNotNone(sc)
                self.assertTrue(sc["workbook_name"].endswith(".xlsx"))
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}?fmt=xlsx"
                ) as r:
                    body = r.read()
                    self.assertEqual(body[:4], b"PK\x03\x04")  # a real zip/xlsx
                    self.assertIn("spreadsheetml",
                                  r.headers.get("Content-Type", ""))
            finally:
                server.shutdown()
                server.server_close()

    def test_pivot_analysis_page_and_data(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                csv = (
                    "ClaimID,ProviderState,AllowedAmt\n"
                    f"1,OH,100\n2,TX,200\n3,OH,50\n"
                ).encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv, method="POST", headers={"X-Filename": "c.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(50):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}/npi-cleaner/status/{job_id}"
                    ) as r:
                        if json.loads(r.read().decode()).get("done"):
                            break
                    time.sleep(0.05)
                # data endpoint returns the cleaned rows as JSON
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/data/{job_id}"
                ) as r:
                    data = json.loads(r.read().decode())
                self.assertIn("ProviderState", data["columns"])
                self.assertEqual(len(data["rows"]), 3)
                # analysis page renders with the pivot builder
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/analyze/{job_id}"
                ) as r:
                    page = r.read().decode()
                self.assertIn("an-fieldlist", page)
                self.assertIn(f"/npi-cleaner/data/", page)
                # enhanced analytics: stat tiles, quick views, chart modes
                self.assertIn("an-tiles", page)
                self.assertIn("Quick views", page)
                self.assertIn('value="heatmap"', page)
                self.assertIn('value="scatter"', page)
                self.assertIn('value="correlation"', page)
                self.assertIn("renderCorrelation", page)
                self.assertIn("Pearson correlation", page)
                self.assertIn('value="box"', page)
                self.assertIn("renderBoxplot", page)
                self.assertIn('value="histogram"', page)
                self.assertIn("renderHistogram", page)
                self.assertIn("% of total", page)
                # unknown job → graceful "expired" page (still 200)
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/analyze/deadbeef"
                ) as r:
                    self.assertIn("expired", r.read().decode())
            finally:
                server.shutdown()
                server.server_close()

    def test_empty_upload_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            server, port = self._start(tmp)
            try:
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=b"", method="POST",
                    headers={"X-Filename": "x.csv", "Content-Length": "0"},
                )
                try:
                    _u.urlopen(req)
                    self.fail("expected 400")
                except _u.HTTPError as e:  # type: ignore[attr-defined]
                    self.assertEqual(e.code, 400)
            finally:
                server.shutdown()
                server.server_close()


class TestBigFileStreaming(unittest.TestCase):
    """Chunked streaming (bigfile.py): a file above the threshold cleans in
    bounded-memory chunks and merges to the SAME numbers the in-memory
    pipeline produces. Thresholds are patched tiny so a kilobyte fixture
    exercises the multi-chunk path."""

    def _fixture(self, n_rows: int = 400) -> bytes:
        # Dirty on purpose: a whitespace NPI (repair), formatted money
        # (repair), one adjacent exact-duplicate pair at the START (same
        # chunk under any boundary), and a future service date on the LAST
        # row so worklist/changelog offsets past chunk 1 are proven global.
        rows = ["ClaimID,BillingProviderNPI,ChargeAmt,DateOfService"]
        rows.append(f"1, {GOOD_A} ,\"$1,250.50\",2024-02-11")
        rows.append(f"1, {GOOD_A} ,\"$1,250.50\",2024-02-11")  # exact dup
        for i in range(2, n_rows - 1):
            rows.append(f"{i},{GOOD_B},420,2024-03-0{1 + i % 9}")
        rows.append(f"{n_rows},{GOOD_A},99,2099-01-01")  # future date, last
        return ("\n".join(rows) + "\n").encode()

    def test_stream_matches_in_memory(self):
        from rcm_mc.npi_cleaner import bigfile
        data = self._fixture()
        ref = engine.clean_bytes(data, "big.csv")
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "big.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 512), \
                    patch.object(bigfile, "CHUNK_TARGET_BYTES", 2048):
                res = bigfile.clean_path(p, "big.csv")

        # Multi-chunk actually happened.
        self.assertTrue(any("Streaming mode" in w and "chunk" in w
                            for w in res.warnings), res.warnings[:2])
        # Core numbers identical to the one-shot pipeline: every counter is
        # a per-row/per-cell sum and the only duplicate pair is adjacent.
        self.assertEqual(res.n_rows_in, ref.n_rows_in)
        self.assertEqual(res.n_rows_out, ref.n_rows_out)
        self.assertEqual(res.n_dupes_removed, 1)
        self.assertEqual(res.repairs, ref.repairs)
        self.assertEqual(res.sanity, ref.sanity)
        self.assertEqual(res.n_changes, ref.n_changes)
        self.assertEqual(res.n_cells_total, ref.n_cells_total)
        self.assertEqual(res.n_cells_filled, ref.n_cells_filled)
        self.assertEqual(res.column_stats, ref.column_stats)
        self.assertEqual(res.quality()["score"], ref.quality()["score"])
        # Worklist indices are OUTPUT-global: the future-date row is the
        # very last output row, far past chunk 1.
        self.assertEqual(res.flag_rows.get("date-in-future"),
                         ref.flag_rows.get("date-in-future"))
        self.assertEqual(res.flag_rows["date-in-future"][-1],
                         res.n_rows_out)

        # The master output: one header + every cleaned row, cells cleaned.
        with open(res.out_path, encoding="utf-8") as fh:
            lines = fh.read().splitlines()
        self.assertEqual(len(lines), 1 + res.n_rows_out)
        self.assertEqual(lines[0].split(",")[1], "BillingProviderNPI")
        self.assertIn(f"1,{GOOD_A},1250.50,2024-02-11", lines[1])

        # The master changelog: global INPUT row indices, including one on
        # the last input row's chunk-2+ territory.
        self.assertTrue(res.changelog_path)
        with open(res.changelog_path, encoding="utf-8") as fh:
            log_lines = fh.read().splitlines()
        self.assertEqual(len(log_lines) - 1, ref.n_changes)
        max_row = max(int(ln.split(",")[0]) for ln in log_lines[1:])
        self.assertLessEqual(max_row, res.n_rows_in)
        self.assertGreater(max_row, 100)  # offsets applied past chunk 1

    def test_quoted_newlines_never_split_a_record(self):
        from rcm_mc.npi_cleaner import bigfile
        rows = ["ClaimID,Note,BillingProviderNPI"]
        for i in range(1, 201):
            rows.append(f'{i},"line one\nline two",{GOOD_A}')
        data = ("\n".join(rows) + "\n").encode()
        ref = engine.clean_bytes(data, "q.csv")
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "q.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 256), \
                    patch.object(bigfile, "CHUNK_TARGET_BYTES", 1024):
                res = bigfile.clean_path(p, "q.csv")
        self.assertEqual(res.n_rows_in, ref.n_rows_in)
        self.assertEqual(res.n_rows_out, ref.n_rows_out)

    def test_small_file_delegates_to_in_memory_pipeline(self):
        from rcm_mc.npi_cleaner import bigfile
        data = engine.sample_csv().encode()
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "s.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            res = bigfile.clean_path(p, "s.csv")
        ref = engine.clean_bytes(data, "s.csv")
        self.assertEqual(res.n_rows_out, ref.n_rows_out)
        # Full artifact set — small files keep the workbook et al.
        self.assertTrue(res.workbook_path)

    def test_huge_non_splittable_format_gets_instructions(self):
        from rcm_mc.npi_cleaner import bigfile
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "w.xlsx")
            with open(p, "wb") as fh:
                fh.write(b"PK\x03\x04" + b"\x00" * 500)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 64), \
                    patch.object(bigfile, "_FORMAT_INMEM_MAX_BYTES", 128):
                res = bigfile.clean_path(p, "w.xlsx")
        self.assertEqual(res.n_rows_out, 0)
        self.assertTrue(any("can't be split" in w for w in res.warnings),
                        res.warnings)

    def test_stream_chunk_mode_skips_run_side_effects(self):
        # A chunk hands the table back in memory: no output files, no
        # history record for the chunk itself.
        from rcm_mc.npi_cleaner import history as _history
        before = len(_history.list_runs(500))
        res = engine.clean_bytes(engine.sample_csv().encode(), "chunk.csv",
                                 _stream_chunk=True)
        self.assertIsNotNone(res.chunk_payload)
        headers, cleaned = res.chunk_payload
        self.assertEqual(len(cleaned), res.n_rows_out)
        self.assertIsNone(res.out_path)
        self.assertIsNone(res.workbook_path)
        self.assertEqual(len(_history.list_runs(500)), before)


class TestWishlist(unittest.TestCase):
    """The "missing something?" backlog: store roundtrip + HTTP routes."""

    def test_store_roundtrip_and_caps(self):
        from rcm_mc.npi_cleaner import wishlist
        import uuid as _uuid
        marker = _uuid.uuid4().hex[:12]
        rec = wishlist.add_request("payer", f"Add Elevance family {marker}",
                                   "BCBS of GA rebranded")
        try:
            self.assertEqual(rec["status"], "open")
            self.assertEqual(rec["category"], "payer")
            listed = wishlist.list_requests()
            mine = [r for r in listed if marker in str(r["title"])]
            self.assertEqual(len(mine), 1)
            # Status lifecycle; unknown statuses refused.
            self.assertTrue(wishlist.set_status(rec["id"], "planned"))
            self.assertFalse(wishlist.set_status(rec["id"], "hacked"))
            planned = wishlist.list_requests("planned")
            self.assertTrue(any(r["id"] == rec["id"] for r in planned))
            # Caps: hostile title/category can't grow the store or invent
            # categories.
            big = wishlist.add_request("<script>", "x" * 5000, "y" * 90000)
            try:
                self.assertEqual(big["category"], "other")
                self.assertEqual(len(big["title"]), 120)
                self.assertEqual(len(big["details"]), 2000)
            finally:
                wishlist.delete_request(big["id"])
            with self.assertRaises(ValueError):
                wishlist.add_request("rule", "   ")
        finally:
            self.assertTrue(wishlist.delete_request(rec["id"]))

    def test_wishlist_http_routes(self):
        import uuid as _uuid
        marker = _uuid.uuid4().hex[:12]
        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            rid = None
            try:
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/wishlist",
                    data=json.dumps({
                        "category": "rule",
                        "title": f"Flag CLIA-required labs {marker}",
                        "details": "HCPCS 8xxxx without a CLIA number",
                    }).encode(), method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    j = json.loads(r.read().decode())
                self.assertTrue(j["ok"])
                rid = j["request"]["id"]
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/wishlist"
                ) as r:
                    listed = json.loads(r.read().decode())
                self.assertIn("categories", listed)
                self.assertTrue(any(marker in q["title"]
                                    for q in listed["requests"]))
                # Bad body → 400, not a 500.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/wishlist",
                    data=json.dumps({"title": ""}).encode(), method="POST",
                    headers={"Content-Type": "application/json"})
                try:
                    _u.urlopen(req)
                    self.fail("expected 400")
                except _ue.HTTPError as e:
                    self.assertEqual(e.code, 400)
                # Move through the backlog, then delete.
                req = _u.Request(
                    f"http://127.0.0.1:{port}"
                    "/npi-cleaner/api/wishlist/status",
                    data=json.dumps({"id": rid,
                                     "status": "shipped"}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    self.assertTrue(json.loads(r.read().decode())["ok"])
                req = _u.Request(
                    f"http://127.0.0.1:{port}"
                    "/npi-cleaner/api/wishlist/delete",
                    data=json.dumps({"id": rid}).encode(), method="POST",
                    headers={"Content-Type": "application/json"})
                with _u.urlopen(req) as r:
                    self.assertTrue(json.loads(r.read().decode())["ok"])
                rid = None
            finally:
                if rid is not None:
                    from rcm_mc.npi_cleaner import wishlist
                    wishlist.delete_request(rid)
                server.shutdown()
                server.server_close()

    def test_page_carries_wishlist_card_and_10gb_copy(self):
        from rcm_mc.ui.npi_cleaner_page import render_npi_cleaner
        body = render_npi_cleaner()
        self.assertIn("npi-wishlist", body)
        self.assertIn("Missing something?", body)
        self.assertIn("/npi-cleaner/api/wishlist", body)
        self.assertIn("10&nbsp;GB", body)


class TestSpooledUploadHttp(unittest.TestCase):
    def test_large_body_spools_to_disk_and_streams(self):
        # Force the spool path (threshold 1 KB) AND the chunked streaming
        # path (tiny bigfile thresholds): the full 10 GB flow, end to end,
        # on a small fixture.
        import rcm_mc.server as _srv
        from rcm_mc.npi_cleaner import bigfile
        rows = ["ClaimID,BillingProviderNPI,ChargeAmt"]
        rows += [f"{i}, {GOOD_A} ,420" for i in range(1, 301)]
        csv_b = ("\n".join(rows) + "\n").encode()
        self.assertGreater(len(csv_b), 1024)
        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                with patch.object(_srv, "_NPI_SPOOL_THRESHOLD_BYTES", 1024), \
                        patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 512), \
                        patch.object(bigfile, "CHUNK_TARGET_BYTES", 2048):
                    req = _u.Request(
                        f"http://127.0.0.1:{port}/npi-cleaner/upload",
                        data=csv_b, method="POST",
                        headers={"X-Filename": "spool.csv"})
                    with _u.urlopen(req) as r:
                        job_id = json.loads(r.read().decode())["job_id"]
                    for _ in range(100):
                        with _u.urlopen(
                            f"http://127.0.0.1:{port}"
                            f"/npi-cleaner/status/{job_id}"
                        ) as r:
                            j = json.loads(r.read().decode())
                            if j.get("done"):
                                break
                        time.sleep(0.05)
                self.assertTrue(j.get("done"), j)
                self.assertNotIn("error", {k: v for k, v in j.items()
                                           if v is not None})
                sc = j["scorecard"]
                self.assertEqual(sc["rows_in"], 300)
                self.assertEqual(sc["rows_out"], 300)
                self.assertTrue(any("Streaming mode" in w
                                    for w in sc["warnings"]))
                # The money-normalize repair (420 → 420.00) ran on every
                # row across every chunk.
                self.assertGreaterEqual(sc["repairs_total"], 300)
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                ) as r:
                    body = r.read().decode()
                self.assertEqual(len(body.splitlines()), 301)
                self.assertIn(f"1,{GOOD_A},420", body)
                # The spool file is removed once the job finishes.
                from rcm_mc.npi_cleaner.engine import WORKDIR
                updir = WORKDIR / "uploads"
                leftovers = (list(updir.glob("upload_*.spool"))
                             if updir.exists() else [])
                self.assertEqual(leftovers, [])
            finally:
                server.shutdown()
                server.server_close()


class TestJobEtaAndCancel(unittest.TestCase):
    """Long-run ergonomics: ETA projection in job status and cooperative
    cancellation — the difference between a 12-hour job being a feature
    and a liability."""

    def test_eta_projection(self):
        job = engine.Job(job_id="x", name="f.csv",
                         created=time.time() - 100, frac=0.5)
        d = job.status_dict()
        self.assertAlmostEqual(d["eta_secs"], 100, delta=15)
        self.assertGreaterEqual(d["elapsed_secs"], 85)
        # No projection before real work starts, none once done.
        early = engine.Job(job_id="y", name="f.csv",
                           created=time.time(), frac=0.01)
        self.assertNotIn("eta_secs", early.status_dict())
        done = engine.Job(job_id="z", name="f.csv",
                          created=time.time() - 100, frac=1.0, done=True)
        self.assertNotIn("eta_secs", done.status_dict())

    def test_cancel_stops_a_running_job(self):
        mgr = engine.JobManager()

        def slow_clean(data, name, **kw):
            cbf = kw.get("progress")
            for i in range(200):
                cbf(f"step {i}", i / 200.0)
                time.sleep(0.01)
            return engine.CleanResult()

        with patch.object(engine, "clean_bytes", slow_clean):
            jid = mgr.submit(b"x", "slow.csv")
            time.sleep(0.15)
            self.assertTrue(mgr.cancel(jid))
            job = mgr.get(jid)
            for _ in range(200):
                if job.done:
                    break
                time.sleep(0.02)
        self.assertTrue(job.done)
        self.assertEqual(job.msg, "Cancelled")
        self.assertIn("Cancelled", job.error)
        self.assertIsNone(job.result)
        # Cancelling a finished job is a refused no-op.
        self.assertFalse(mgr.cancel(jid))
        self.assertFalse(mgr.cancel("no-such-job"))

    def test_cancel_route_over_http(self):
        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)

            def slow_clean(data, name, **kw):
                cbf = kw.get("progress")
                for i in range(300):
                    cbf(f"step {i}", i / 300.0)
                    time.sleep(0.01)
                return engine.CleanResult()

            try:
                with patch.object(engine, "clean_bytes", slow_clean):
                    req = _u.Request(
                        f"http://127.0.0.1:{port}/npi-cleaner/upload",
                        data=b"ClaimID\n1\n", method="POST",
                        headers={"X-Filename": "slow.csv"})
                    with _u.urlopen(req) as r:
                        job_id = json.loads(r.read().decode())["job_id"]
                    time.sleep(0.15)
                    req = _u.Request(
                        f"http://127.0.0.1:{port}"
                        f"/npi-cleaner/cancel/{job_id}",
                        data=b"", method="POST")
                    with _u.urlopen(req) as r:
                        self.assertTrue(json.loads(r.read().decode())["ok"])
                    for _ in range(200):
                        with _u.urlopen(
                            f"http://127.0.0.1:{port}"
                            f"/npi-cleaner/status/{job_id}"
                        ) as r:
                            j = json.loads(r.read().decode())
                            if j.get("done"):
                                break
                        time.sleep(0.02)
                self.assertTrue(j["done"])
                self.assertIn("Cancelled", j["error"])
                # Cancelling again → refused.
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/cancel/{job_id}",
                    data=b"", method="POST")
                with _u.urlopen(req) as r:
                    self.assertFalse(json.loads(r.read().decode())["ok"])
            finally:
                server.shutdown()
                server.server_close()

    def test_pages_carry_cancel_eta_and_backlog_ui(self):
        from rcm_mc.ui.npi_cleaner_page import render_npi_cleaner
        from rcm_mc.ui.npi_history_page import render_npi_history
        cleaner = render_npi_cleaner()
        self.assertIn("npi-cancel", cleaner)
        self.assertIn("npi-bar-eta", cleaner)
        self.assertIn("/npi-cleaner/cancel/", cleaner)
        history = render_npi_history()
        self.assertIn("nh-wish-box", history)
        self.assertIn("/npi-cleaner/api/wishlist/status", history)


class TestPopulationAnalytics(unittest.TestCase):
    """The Tuva-class marts (analytics.py): service mix, encounters,
    chronic conditions, volume integrity, readmissions, coding intensity —
    computed offline from the cleaned table, report-only."""

    def test_classify_line_ladder(self):
        from rcm_mc.npi_cleaner.analytics import classify_line
        # TOB decides first (inpatient acute).
        self.assertEqual(classify_line("0120", "0111", "", "99231"),
                         ("Inpatient", "Acute inpatient"))
        # ED wins inside an outpatient TOB via revenue 045x.
        self.assertEqual(classify_line("0450", "0131", "", ""),
                         ("Outpatient", "Emergency department"))
        # SNF by TOB facility digit 2.
        self.assertEqual(classify_line("", "0211", "", "")[0], "Inpatient")
        # Professional office E&M by POS 11 + code.
        self.assertEqual(classify_line("", "", "11", "99214"),
                         ("Office", "Office visit (E&M)"))
        # HCPCS ranges when nothing else is present.
        self.assertEqual(classify_line("", "", "", "80053"),
                         ("Ancillary", "Laboratory"))
        self.assertEqual(classify_line("", "", "", "72148"),
                         ("Ancillary", "Imaging"))
        self.assertEqual(classify_line("", "", "", "J1100"),
                         ("Pharmacy", "Drugs (J-codes)"))
        self.assertEqual(classify_line("", "", "", ""),
                         ("Unclassified", "Unclassified"))

    def test_encounters_group_and_readmit_window(self):
        csv_rows = ["PatientID,TypeOfBill,RevenueCode,HCPCS,ChargeAmt,"
                    "DateOfService,AdmitDate,DischargeDate"]
        # One stay: two lines inside the admit→discharge span.
        csv_rows.append("P1,0111,0120,99231,5000,2024-01-05,"
                        "2024-01-03,2024-01-08")
        csv_rows.append("P1,0111,0250,J1100,300,2024-01-06,"
                        "2024-01-03,2024-01-08")
        # Readmit 10 days after discharge (counts).
        csv_rows.append("P1,0111,0120,99232,7000,2024-01-18,"
                        "2024-01-18,2024-01-22")
        # Next stay 45 days later (does NOT count).
        csv_rows.append("P1,0111,0120,99232,7000,2024-03-08,"
                        "2024-03-08,2024-03-10")
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "enc.csv")
        enc = (res.population or {}).get("encounters")
        self.assertIsNotNone(enc)
        self.assertEqual(enc["n_encounters"], 3)  # 2 lines merged into stay 1
        r = enc["readmissions"]
        self.assertEqual(r["inpatient_stays"], 3)
        self.assertEqual(r["readmissions_30d"], 1)

    def test_conditions_prevalence_and_multimorbidity(self):
        csv_rows = ["PatientID,DiagnosisCode,HCPCS"]
        csv_rows += ["P1,E11.65,99213", "P1,I10,99213", "P1,N18.3,99213",
                     "P2,J45.909,99213", "P3,,99213"]
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "dx.csv")
        cond = (res.population or {}).get("conditions")
        self.assertIsNotNone(cond)
        by_name = {p["condition"]: p for p in cond["prevalence"]}
        self.assertEqual(by_name["Diabetes"]["patients"], 1)
        self.assertEqual(by_name["Asthma"]["patients"], 1)
        # P1 carries 3 conditions, P2 one. (P3 has no dx → not in the
        # per-patient grouping at all.)
        self.assertEqual(cond["multimorbidity"]["3+"], 1)
        self.assertEqual(cond["multimorbidity"]["1"], 1)

    def test_volume_cliff_detection(self):
        csv_rows = ["ClaimID,DateOfService,ChargeAmt"]
        n = 0
        for m, count in enumerate((100, 100, 100, 10, 100, 100), start=1):
            for i in range(count):
                n += 1
                csv_rows.append(f"{n},2024-0{m}-15,100")
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "vol.csv")
        vol = (res.population or {}).get("volume")
        self.assertIsNotNone(vol)
        self.assertEqual(len(vol["months"]), 6)
        self.assertEqual(len(vol["alerts"]), 1)
        self.assertIn("2024-04", vol["alerts"][0])

    def test_coding_intensity_flags_hot_provider(self):
        csv_rows = ["BillingProviderNPI,HCPCS,DateOfService"]
        csv_rows += [f"{GOOD_A},99215,2024-01-02"] * 25   # all level 5
        csv_rows += [f"{GOOD_B},99213,2024-01-02"] * 100  # all level 3
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "em.csv", drop_duplicates=False)
        ci = (res.population or {}).get("coding_intensity")
        self.assertIsNotNone(ci)
        self.assertEqual(ci["established_visits"], 125)
        outlier_npis = [o["npi"] for o in ci["outliers"]]
        self.assertIn(GOOD_A, outlier_npis)
        self.assertNotIn(GOOD_B, outlier_npis)
        self.assertIn("99214", ci["national_mix"])

    def test_scorecard_excludes_encounter_records(self):
        csv_rows = ["PatientID,HCPCS,DateOfService",
                    "P1,99213,2024-01-05", "P2,99213,2024-01-06"]
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "sc.csv")
        sc = res.as_scorecard()
        self.assertIsNotNone(sc["population"])
        self.assertNotIn("records", sc["population"]["encounters"])
        # The full records stay on the result for the CSV download.
        self.assertTrue(res.population["encounters"]["records"])
        from rcm_mc.npi_cleaner import analytics
        text = analytics.encounters_csv(res.population)
        self.assertTrue(text.startswith("encounter,patient,category"))
        self.assertEqual(len(text.splitlines()), 3)

    def test_streaming_runs_skip_population(self):
        from rcm_mc.npi_cleaner import bigfile
        rows = ["PatientID,HCPCS,DateOfService"]
        rows += [f"P{i},99213,2024-01-0{1 + i % 9}" for i in range(1, 200)]
        data = ("\n".join(rows) + "\n").encode()
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "s.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 512), \
                    patch.object(bigfile, "CHUNK_TARGET_BYTES", 1024):
                res = bigfile.clean_path(p, "s.csv")
        self.assertIsNone(res.population)
        self.assertTrue(any("population marts" in w for w in res.warnings))

    def test_population_tab_and_encounters_route(self):
        from rcm_mc.ui.npi_cleaner_page import render_npi_cleaner
        body = render_npi_cleaner()
        self.assertIn('data-panel="population"', body)
        self.assertIn("fmt=encounters", body)
        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                csv_b = ("PatientID,HCPCS,DateOfService\n"
                         "P1,99213,2024-01-05\nP2,99214,2024-01-06\n"
                         ).encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "e.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(100):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}"
                        f"/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                self.assertIn("population", j["scorecard"])
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=encounters"
                ) as r:
                    text = r.read().decode()
                self.assertIn("encounter,patient,category", text)
            finally:
                server.shutdown()
                server.server_close()


class TestPhase2(unittest.TestCase):
    """Batch 22: cross-chunk dedupe on streamed runs, observed PMPM in the
    volume mart, encounters.csv in bundles, Population in the exec report."""

    def test_streamed_dedupe_crosses_chunk_boundaries(self):
        from rcm_mc.npi_cleaner import bigfile
        rows = ["ClaimID,BillingProviderNPI,ChargeAmt"]
        # The same row appears at the very start and near the very end —
        # guaranteed different chunks under a tiny chunk target.
        dup = f"DUP,{GOOD_A},420"
        rows.append(dup)
        rows += [f"{i},{GOOD_B},100" for i in range(1, 300)]
        rows.append(dup)
        data = ("\n".join(rows) + "\n").encode()
        ref = engine.clean_bytes(data, "x.csv")
        self.assertEqual(ref.n_dupes_removed, 1)
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "x.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 512), \
                    patch.object(bigfile, "CHUNK_TARGET_BYTES", 1024):
                res = bigfile.clean_path(p, "x.csv")
        self.assertGreater(len([w for w in res.warnings
                                if "chunk(s)" in w]), 0)
        self.assertEqual(res.n_dupes_removed, 1)
        self.assertEqual(res.n_rows_out, ref.n_rows_out)
        # And the duplicate row appears exactly once in the output.
        with open(res.out_path, encoding="utf-8") as fh:
            body = fh.read()
        self.assertEqual(body.count(f"DUP,{GOOD_A}"), 1)

    def test_streamed_dedupe_cap_is_surfaced(self):
        from rcm_mc.npi_cleaner import bigfile
        rows = ["ClaimID,BillingProviderNPI"]
        rows += [f"{i},{GOOD_A}" for i in range(1, 200)]
        data = ("\n".join(rows) + "\n").encode()
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "c.csv")
            with open(p, "wb") as fh:
                fh.write(data)
            with patch.object(bigfile, "STREAM_THRESHOLD_BYTES", 256), \
                    patch.object(bigfile, "CHUNK_TARGET_BYTES", 512), \
                    patch.object(engine, "_STREAM_SEEN_CAP", 10):
                res = bigfile.clean_path(p, "c.csv")
        self.assertTrue(any("Duplicate tracking capped" in w
                            for w in res.warnings), res.warnings)

    def test_observed_pmpm_in_volume_mart(self):
        csv_rows = ["PatientID,ChargeAmt,DateOfService"]
        # Jan: 2 patients, $300 total → PMPM 150. Feb: 1 patient, $80.
        csv_rows += ["P1,100,2024-01-05", "P2,200,2024-01-09",
                     "P1,80,2024-02-10"]
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "pmpm.csv")
        vol = (res.population or {}).get("volume")
        self.assertIsNotNone(vol)
        jan, feb = vol["months"][0], vol["months"][1]
        self.assertEqual(jan["observed_pmpm"], 150.0)
        self.assertEqual(feb["observed_pmpm"], 80.0)
        self.assertEqual(vol["median_observed_pmpm"], 150.0)

    def test_exec_report_population_section(self):
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        csv_rows = ["PatientID,TypeOfBill,HCPCS,DiagnosisCode,ChargeAmt,"
                    "DateOfService"]
        csv_rows += [f"P{i},0131,99213,E11.65,100,2024-0{1 + i % 6}-10"
                     for i in range(1, 40)]
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "er.csv")
        html = build_exec_report(res.as_scorecard(), "er.csv", "now")
        self.assertIn("Population profile", html)
        self.assertIn("Care setting", html)
        self.assertIn("Diabetes", html)
        self.assertIn("median observed PMPM", html)

    def test_bundles_carry_encounters_csv(self):
        # CLI --bundle path.
        import contextlib
        import io as _io
        import zipfile as _zf
        from rcm_mc.npi_cleaner import cli as nc_cli
        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, "in.csv")
            with open(src, "w", encoding="utf-8") as fh:
                fh.write("PatientID,HCPCS,DateOfService\n"
                         "P1,99213,2024-01-05\nP2,99214,2024-01-06\n")
            out = _io.StringIO()
            with contextlib.redirect_stdout(out):
                rc = nc_cli.main([src, "--bundle", "--json",
                                  "--outdir", tmp])
            self.assertEqual(rc, 0)
            with _zf.ZipFile(os.path.join(tmp, "in_bundle.zip")) as z:
                self.assertIn("encounters.csv", z.namelist())
                enc = z.read("encounters.csv").decode()
            self.assertIn("encounter,patient,category", enc)

        # HTTP ?fmt=bundle path.
        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                csv_b = ("PatientID,HCPCS,DateOfService\n"
                         "P1,99213,2024-01-05\n").encode()
                req = _u.Request(
                    f"http://127.0.0.1:{port}/npi-cleaner/upload",
                    data=csv_b, method="POST",
                    headers={"X-Filename": "b.csv"})
                with _u.urlopen(req) as r:
                    job_id = json.loads(r.read().decode())["job_id"]
                for _ in range(100):
                    with _u.urlopen(
                        f"http://127.0.0.1:{port}"
                        f"/npi-cleaner/status/{job_id}"
                    ) as r:
                        j = json.loads(r.read().decode())
                        if j.get("done"):
                            break
                    time.sleep(0.05)
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/download/{job_id}"
                    "?fmt=bundle"
                ) as r:
                    blob = r.read()
                import io as _io2
                import zipfile as _zf2
                with _zf2.ZipFile(_io2.BytesIO(blob)) as z:
                    self.assertIn("encounters.csv", z.namelist())
            finally:
                server.shutdown()
                server.server_close()



class TestRefdataPacks(unittest.TestCase):
    """Reference-data packs: pull the real public code sets (NUCC / CMS /
    OIG), store with provenance, and light up pack-gated checks. All
    tests run offline against fixture payloads via a mocked opener."""

    def setUp(self):
        import rcm_mc.npi_cleaner.refdata_packs as packs
        self.packs = packs
        self.tmp = tempfile.TemporaryDirectory()
        self.db = os.path.join(self.tmp.name, "refpacks.sqlite3")
        self.p_db = patch.object(packs, "_DB_PATH",
                                 __import__("pathlib").Path(self.db))
        self.p_db.start()
        packs._CACHE.clear()

    def tearDown(self):
        self.p_db.stop()
        self.packs._CACHE.clear()
        self.tmp.cleanup()

    class _Resp:
        def __init__(self, payload):
            self._buf = __import__("io").BytesIO(payload)

        def read(self, n=-1):
            return self._buf.read(n)

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    def _opener(self, by_url):
        resp_cls = self._Resp

        def open_(req, timeout=0):
            url = req.full_url
            payload = by_url.get(url)
            if payload is None:
                raise OSError(f"blocked: {url}")
            return resp_cls(payload)
        return open_

    def _taxonomy_csv(self):
        return ("Code,Grouping,Classification,Specialization,Definition\n"
                "207X00000X,Allopathic & Osteopathic Physicians,"
                "Orthopaedic Surgery,,d\n"
                "207XS0114X,Allopathic & Osteopathic Physicians,"
                "Orthopaedic Surgery,Adult Reconstructive Orthopaedic "
                "Surgery,d\n").encode()

    def _icd_zip(self):
        import io as _io
        import zipfile as _zf
        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w") as z:
            z.writestr("icd10cm_codes_2026.txt",
                       "E1165   Type 2 diabetes mellitus with hyperglycemia\n"
                       "I10     Essential (primary) hypertension\n")
        return buf.getvalue()

    def _leie_csv(self):
        return ("LASTNAME,FIRSTNAME,EXCLTYPE,NPI,EXCLDATE\n"
                f"DOE,JOHN,1128a1,{GOOD_A},20240101\n"
                "EMPTY,NONE,1128a1,0000000000,20240101\n").encode()

    def test_pull_installs_with_provenance_and_url_fallback(self):
        urls = self.packs._nucc_urls()
        # First candidate 404s, second succeeds — the puller walks on.
        opener = self._opener({urls[1]: self._taxonomy_csv()})
        info = self.packs.pull("taxonomy", opener=opener)
        self.assertEqual(info["rows"], 2)
        self.assertEqual(info["source"], urls[1])
        st = {p["id"]: p for p in self.packs.status()}
        self.assertTrue(st["taxonomy"]["installed"])
        self.assertEqual(st["taxonomy"]["rows"], 2)
        self.assertTrue(st["taxonomy"]["sha256"])
        self.assertFalse(st["leie"]["installed"])
        # The pack now answers lookups the curated subset can't.
        from rcm_mc.npi_cleaner import refdata
        self.assertIn("Adult Reconstructive",
                      refdata.taxonomy_specialty("207XS0114X"))

    def test_unknown_pack_and_dead_urls_raise_readably(self):
        with self.assertRaises(ValueError):
            self.packs.pull("nope", opener=self._opener({}))
        with self.assertRaises(ValueError) as ctx:
            self.packs.pull("leie", opener=self._opener({}))
        self.assertIn("NPI_REFPACK_URL_LEIE", str(ctx.exception))

    def test_icd10_pack_gates_unknown_code_flag(self):
        url = "https://x.test/icd.zip"
        self.packs.pull("icd10cm", opener=self._opener(
            {url: self._icd_zip()}), url=url)
        csv_rows = ["ClaimID,DiagnosisCode",
                    "1,E11.65",   # known (dot stripped)
                    "2,E11.9",    # shaped fine, NOT in the pack
                    "3,NOTACODE"]  # malformed → shape flag, not unknown
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "dx.csv")
        self.assertEqual(res.sanity.get("icd10-unknown-code"), 1)
        self.assertIn("icd10-unknown-code", res.flag_rows)
        # Without the pack the flag never fires.
        self.packs._CACHE.clear()
        with patch.object(self.packs, "icd10_codes", lambda: None):
            res2 = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                      "dx.csv")
        self.assertNotIn("icd10-unknown-code", res2.sanity)

    def test_leie_pack_screens_offline_and_feeds_compliance(self):
        url = "https://x.test/leie.csv"
        self.packs.pull("leie", opener=self._opener(
            {url: self._leie_csv()}), url=url)
        csv_rows = ["ClaimID,BillingProviderNPI",
                    f"1,{GOOD_A}", f"2,{GOOD_B}"]
        res = engine.clean_bytes(("\n".join(csv_rows) + "\n").encode(),
                                 "leie.csv")
        self.assertEqual(res.sanity.get("leie-excluded-npi"), 1)
        # Compliance screen falls back to the pack with no env/path.
        from rcm_mc.npi_cleaner import compliance
        env = dict(os.environ)
        os.environ.pop("RCM_MC_LEIE_CSV", None)
        try:
            out = compliance.screen_leie([GOOD_A, GOOD_B])
        finally:
            os.environ.update(env)
        self.assertTrue(out["available"])
        self.assertEqual(out["excluded"], 1)
        self.assertIn("reference pack", out["source"])

    def test_refdata_routes_and_cli_status(self):
        import contextlib
        import io as _io
        from rcm_mc.npi_cleaner import cli as nc_cli
        out = _io.StringIO()
        with contextlib.redirect_stdout(out):
            rc = nc_cli.main(["--refdata-status"])
        self.assertEqual(rc, 0)
        self.assertIn("taxonomy", out.getvalue())
        self.assertIn("not installed", out.getvalue())

        with tempfile.TemporaryDirectory() as tmp:
            import socket as _socket
            import threading
            from rcm_mc.server import build_server
            s = _socket.socket()
            s.bind(("127.0.0.1", 0))
            port = s.getsockname()[1]
            s.close()
            server, _ = build_server(port=port,
                                     db_path=os.path.join(tmp, "p.db"))
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            time.sleep(0.05)
            try:
                with _u.urlopen(
                    f"http://127.0.0.1:{port}/npi-cleaner/api/refdata"
                ) as r:
                    j = json.loads(r.read().decode())
                ids = {p["id"] for p in j["packs"]}
                self.assertEqual(ids, {"taxonomy", "icd10cm", "hcpcs",
                                       "leie"})
                # Bad pack name → 400, not a silent thread.
                req = _u.Request(
                    f"http://127.0.0.1:{port}"
                    "/npi-cleaner/api/refdata/pull",
                    data=json.dumps({"pack": "bogus"}).encode(),
                    method="POST",
                    headers={"Content-Type": "application/json"})
                try:
                    _u.urlopen(req)
                    self.fail("expected 400")
                except _ue.HTTPError as e:
                    self.assertEqual(e.code, 400)
                # A real pack pulls in the background (pull mocked).
                with patch.object(self.packs, "pull",
                                  lambda pid, **kw: {"rows": 1,
                                                     "source": "t",
                                                     "sha256": "x",
                                                     "pack": pid}):
                    req = _u.Request(
                        f"http://127.0.0.1:{port}"
                        "/npi-cleaner/api/refdata/pull",
                        data=json.dumps({"pack": "taxonomy"}).encode(),
                        method="POST",
                        headers={"Content-Type": "application/json"})
                    with _u.urlopen(req) as r:
                        j = json.loads(r.read().decode())
                    self.assertEqual(j["pulling"], ["taxonomy"])
                    for _ in range(100):
                        state = self.packs._PULLS.get("taxonomy", {})
                        if state.get("state") == "done":
                            break
                        time.sleep(0.02)
                    self.assertEqual(state.get("state"), "done")
            finally:
                self.packs._PULLS.clear()
                server.shutdown()
                server.server_close()

    def test_hcpcs_parser_keeps_level_ii_only(self):
        raw = ("J1100 Injection dexamethasone sodium phosphate 1 mg\n"
               "99213 OFFICE VISIT EST LOW MDM\n"
               "A0428 Ambulance service BLS non-emergency\n").encode()
        rows = dict(self.packs._parse_hcpcs(raw))
        self.assertIn("J1100", rows)
        self.assertIn("A0428", rows)
        self.assertNotIn("99213", rows)  # CPT-4 never stored


class TestMultiSheetWorkbooks(unittest.TestCase):
    """Vendor extracts lead with a cover/'Detail' sheet; the claims table
    sits on a later tab. Reading sheet 0 silently cleaned a 3-column
    cover page while a 13M-cell 'DATA' tab sat ignored — the reader now
    picks the data-bearing sheet everywhere (stdlib + pandas paths) and
    says which sheets it skipped."""

    def _workbook(self):
        import io as _io
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Detail"
        ws1.append(["Extract", "Vendor", "Date"])
        ws1.append(["Infusion TX", "Komodo", "2026-06-18"])
        ws2 = wb.create_sheet("DATA")
        ws2.append(["SERVICE_YEAR", "STATE", "ZIP3", "COUNTY",
                    "BILLING_PROVIDER_NPI", "BILLING_PROVIDER_NAME",
                    "ENTITY_TYPE"])
        for i in range(120):
            ws2.append([2025, "TX", 760 + i % 40, "TARRANT", GOOD_A,
                        f"Hospital {i}", "ORGANIZATION"])
        wb.create_sheet("Sheet2")  # empty trailing tab
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_clean_reads_the_data_sheet_not_the_cover(self):
        data = self._workbook()
        res = engine.clean_bytes(data, "extract.xlsx")
        self.assertEqual(res.n_rows_in, 120)
        self.assertEqual(res.billing_column, "BILLING_PROVIDER_NPI")
        note = [w for w in res.warnings if "Workbook has 3 sheets" in w]
        self.assertEqual(len(note), 1, res.warnings)
        self.assertIn("'DATA'", note[0])
        self.assertIn("'Detail'", note[0])
        self.assertIn("empty", note[0])

    def test_detect_maps_the_data_sheet_and_names_it(self):
        det = engine.detect_columns_preview(self._workbook())
        if det is None:  # pandas unavailable → detector legitimately off
            self.skipTest("v49 detector unavailable")
        self.assertEqual(len(det["headers"]), 7)
        self.assertEqual(det.get("sheet"), "DATA")
        self.assertEqual(det["mapping"].get("billing_npi"),
                         "BILLING_PROVIDER_NPI")

    def test_single_sheet_workbook_gets_no_note(self):
        import io as _io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["NPI"])
        ws.append([GOOD_A])
        buf = _io.BytesIO()
        wb.save(buf)
        res = engine.clean_bytes(buf.getvalue(), "one.xlsx")
        self.assertEqual(res.n_rows_in, 1)
        self.assertFalse([w for w in res.warnings if "Workbook has" in w])


if __name__ == "__main__":
    unittest.main()


