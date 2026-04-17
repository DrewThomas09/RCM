"""Tests for the reverse challenge solver."""
from __future__ import annotations

import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd

from rcm_mc.analysis.challenge import (
    ChallengeResult,
    _binary_search_progress,
    _drag_at_achievement,
    challenge_to_dataframe,
    run_challenge,
    solve_global_progress,
    solve_per_kpi_progress,
)
from rcm_mc.infra.config import load_and_validate


BASE_DIR = Path(__file__).resolve().parents[1]
ACTUAL_PATH = str(BASE_DIR / "configs" / "actual.yaml")
BENCH_PATH = str(BASE_DIR / "configs" / "benchmark.yaml")


class TestBinarySearch(unittest.TestCase):
    """Pure-math tests on the bisection against a synthetic monotonic function."""

    def test_target_above_status_quo_returns_none(self):
        # f(0) = 10 (status quo), f(1) = 2 (full benchmark). Target = 20 is above f(0).
        def f(a): return 10.0 - 8.0 * a
        self.assertIsNone(_binary_search_progress(20.0, f))

    def test_target_below_full_benchmark_returns_one(self):
        def f(a): return 10.0 - 8.0 * a
        self.assertEqual(_binary_search_progress(0.5, f), 1.0)

    def test_finds_linear_midpoint(self):
        def f(a): return 10.0 - 8.0 * a
        result = _binary_search_progress(6.0, f, tol_pct=0.05)
        # f(0.5) = 6 → progress should be near 0.5
        self.assertAlmostEqual(result, 0.5, delta=0.1)

    def test_respects_tolerance(self):
        # Very tight tolerance: should still terminate within max_iter
        def f(a): return 10.0 - 8.0 * a
        result = _binary_search_progress(6.0, f, tol_pct=0.001, max_iter=20)
        self.assertIsNotNone(result)
        self.assertAlmostEqual(f(result), 6.0, delta=0.1)

    def test_nan_target_rejected(self):
        def f(a): return 10.0 - 8.0 * a
        with self.assertRaises(ValueError):
            _binary_search_progress(float("nan"), f)

    def test_invalid_tolerance_rejected(self):
        def f(a): return 10.0 - 8.0 * a
        with self.assertRaises(ValueError):
            _binary_search_progress(6.0, f, tol_pct=0.0)


class TestSolvers(unittest.TestCase):
    """Integration tests against the shipped configs (low n_sims for speed)."""

    @classmethod
    def setUpClass(cls):
        cls.actual = load_and_validate(ACTUAL_PATH)
        cls.bench = load_and_validate(BENCH_PATH)
        # Measure status-quo drag once so tests can pick target values relative to it
        cls.current_drag = _drag_at_achievement(
            cls.actual, cls.bench, 0.0, ["idr_blended"], n_sims=300, seed=42,
        )

    def test_target_near_current_needs_low_progress(self):
        """Target very close to status quo → small progress fraction."""
        target = self.current_drag * 0.95  # 5% drag reduction
        progress = solve_global_progress(
            self.actual, self.bench, target, n_sims=300, seed=42,
        )
        self.assertIsNotNone(progress)
        self.assertGreaterEqual(progress, 0.0)
        self.assertLess(progress, 0.5,
                        msg=f"5% drag reduction shouldn't require >50% progress; got {progress}")

    def test_target_above_current_returns_none(self):
        target = self.current_drag * 2.0
        self.assertIsNone(solve_global_progress(
            self.actual, self.bench, target, n_sims=200, seed=42,
        ))

    def test_impossibly_low_target_returns_one(self):
        """Target well below full-benchmark drag → unreachable marker."""
        progress = solve_global_progress(
            self.actual, self.bench, target_drag=-1e12, n_sims=200, seed=42,
        )
        self.assertEqual(progress, 1.0)

    def test_per_kpi_returns_all_three_levers(self):
        target = self.current_drag * 0.90
        per_kpi = solve_per_kpi_progress(
            self.actual, self.bench, target, n_sims=300, seed=42,
        )
        self.assertEqual(set(per_kpi.keys()), {"idr_blended", "fwr_blended", "dar_blended"})


class TestRunChallenge(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.actual = load_and_validate(ACTUAL_PATH)
        cls.bench = load_and_validate(BENCH_PATH)

    def test_result_fields_populated(self):
        current = _drag_at_achievement(
            self.actual, self.bench, 0.0, ["idr_blended"], n_sims=200, seed=42,
        )
        target = current * 0.80
        result = run_challenge(
            self.actual, self.bench, target, n_sims=200, seed=42,
        )
        self.assertIsInstance(result, ChallengeResult)
        self.assertEqual(result.target_drag, target)
        self.assertGreater(result.current_drag, 0)
        self.assertEqual(
            set(result.blended_values.keys()),
            {"idr_blended", "fwr_blended", "dar_blended"},
        )
        for bd in result.blended_values.values():
            self.assertIn("current", bd)
            self.assertIn("benchmark", bd)
            self.assertIn("target_required", bd)

    def test_challenge_to_dataframe_has_expected_rows(self):
        current = _drag_at_achievement(
            self.actual, self.bench, 0.0, ["idr_blended"], n_sims=200, seed=42,
        )
        result = run_challenge(self.actual, self.bench, current * 0.85, n_sims=200, seed=42)
        df = challenge_to_dataframe(result)
        self.assertEqual(len(df), 4)  # joint + 3 levers
        self.assertEqual(
            list(df.columns),
            ["lever", "current_value", "target_value_required",
             "benchmark_value", "progress_needed"],
        )
        # Joint row has "—" placeholder values
        joint = df.iloc[0]
        self.assertEqual(joint["lever"], "joint (IDR + FWR + DAR)")
        self.assertEqual(joint["current_value"], "—")

    def test_non_finite_target_drag_rejected(self):
        with self.assertRaises(ValueError):
            run_challenge(
                self.actual, self.bench, float("nan"), n_sims=100, seed=42,
            )

    def test_non_positive_n_sims_rejected(self):
        current = _drag_at_achievement(
            self.actual, self.bench, 0.0, ["idr_blended"], n_sims=100, seed=42,
        )
        with self.assertRaises(ValueError):
            run_challenge(
                self.actual, self.bench, current * 0.9, n_sims=0, seed=42,
            )


class TestChallengeCLI(unittest.TestCase):
    """End-to-end: `rcm-mc challenge ... --outdir X` produces the CSV."""

    def test_help_exits_zero(self):
        env = os.environ.copy()
        env.setdefault("MPLCONFIGDIR", "/tmp/mplcache")
        result = subprocess.run(
            [sys.executable, "-m", "rcm_mc", "challenge", "--help"],
            cwd=str(BASE_DIR), env=env,
            capture_output=True, text=True, timeout=30,
        )
        self.assertEqual(result.returncode, 0)
        self.assertIn("rcm-mc challenge", result.stdout)

    def test_end_to_end_writes_csv(self):
        env = os.environ.copy()
        env.setdefault("MPLCONFIGDIR", "/tmp/mplcache")
        with tempfile.TemporaryDirectory() as tmp:
            result = subprocess.run(
                [
                    sys.executable, "-m", "rcm_mc", "challenge",
                    "--actual", ACTUAL_PATH, "--benchmark", BENCH_PATH,
                    "--target-drag", "10000000",
                    "--outdir", tmp,
                    "--n-sims", "200", "--seed", "42",
                ],
                cwd=str(BASE_DIR), env=env,
                capture_output=True, text=True, timeout=240,
            )
            self.assertEqual(
                result.returncode, 0,
                msg=f"STDOUT:{result.stdout}\nSTDERR:{result.stderr}",
            )
            csv_path = os.path.join(tmp, "challenge_analysis.csv")
            self.assertTrue(os.path.exists(csv_path))
            df = pd.read_csv(csv_path)
            self.assertEqual(len(df), 4)

    def test_nan_target_drag_fails_cleanly(self):
        env = os.environ.copy()
        env.setdefault("MPLCONFIGDIR", "/tmp/mplcache")
        result = subprocess.run(
            [
                sys.executable, "-m", "rcm_mc", "challenge",
                "--actual", ACTUAL_PATH, "--benchmark", BENCH_PATH,
                "--target-drag", "nan",
                "--n-sims", "50",
            ],
            cwd=str(BASE_DIR), env=env,
            capture_output=True, text=True, timeout=120,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("challenge failed", result.stderr)


class TestChallengeInWorkbook(unittest.TestCase):
    def test_challenge_tab_added_to_workbook_when_csv_present(self):
        import yaml as _yaml
        from openpyxl import load_workbook

        from rcm_mc.infra._bundle import write_diligence_workbook

        with open(ACTUAL_PATH) as f:
            cfg = _yaml.safe_load(f)
        summary = pd.DataFrame({
            "mean": [5e6], "p10": [3e6], "p90": [7e6],
        }, index=["ebitda_drag"])

        with tempfile.TemporaryDirectory() as tmp:
            # Stub a challenge_analysis.csv
            pd.DataFrame({
                "lever": ["joint"],
                "current_value": ["—"],
                "target_value_required": ["—"],
                "benchmark_value": ["—"],
                "progress_needed": ["60%"],
            }).to_csv(os.path.join(tmp, "challenge_analysis.csv"), index=False)
            write_diligence_workbook(tmp, summary, cfg)
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Challenge", wb.sheetnames)
            wb.close()
