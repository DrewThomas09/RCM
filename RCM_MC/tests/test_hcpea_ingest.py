"""HCPEA deal-tracker ingest → Deal Library.

Covers the pure normalizers (compliance: missing -> None, never 0; canonical
deal-type/quarter) and end-to-end parsing of both workbook layouts
(HCPEA quarterly 'Data' sheet; annual one-sheet-per-quarter) into
deal_library_companies via the audited loader.
"""
from __future__ import annotations

import os
import tempfile
import unittest
from datetime import datetime

import openpyxl

from scripts import ingest_hcpea_trackers as ing
from rcm_mc.portfolio.store import PortfolioStore
from rcm_mc.data import deal_library as dl


class NormalizerTests(unittest.TestCase):
    def test_missing_tokens_become_none_never_zero(self):
        for tok in ("", "-", "NM", "n/a", "  ", "TBD"):
            self.assertIsNone(ing._s(tok))
            self.assertIsNone(ing._money(tok))   # critically: not 0.0

    def test_money_parses_and_keeps_blank_as_none(self):
        self.assertEqual(ing._money("$1,700"), 1700.0)
        self.assertEqual(ing._money("75"), 75.0)
        self.assertIsNone(ing._money(None))

    def test_quarter_from_iso(self):
        self.assertEqual(ing._quarter("2025-09-30"), "2025Q3")
        self.assertEqual(ing._quarter("2026-01-01"), "2026Q1")
        self.assertIsNone(ing._quarter(None))

    def test_iso_date_from_datetime_and_string(self):
        self.assertEqual(ing._iso_date(datetime(2025, 6, 30)), "2025-06-30")
        self.assertEqual(ing._iso_date("2019-03-29 00:00:00"), "2019-03-29")

    def test_deal_type_canonicalizes_variants(self):
        self.assertEqual(ing._norm_deal_type("Add-on"), ing._norm_deal_type("Add-On"))
        self.assertEqual(
            ing._norm_deal_type("Merger / Acquisition"), "Merger/Acquisition")
        self.assertEqual(ing._norm_deal_type("buyout/lbo"), "Buyout/LBO")

    def test_sponsor_extracted_from_acquirer(self):
        self.assertEqual(
            ing._sponsor_from_acquirer("CVC Capital Partners (Jane Doe)"),
            "CVC Capital Partners")


def _hcpea_book(path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    ws["B4"], ws["C4"], ws["D4"] = "Deal Date", "Target", "Acquirer"
    ws["E4"], ws["F4"], ws["G4"], ws["H4"] = (
        "Deal Size\n ($ in mill)", "Post Valuation\n ($ in mill)",
        "Deal Type", "Sector")
    ws["B5"], ws["C5"], ws["D5"] = datetime(2025, 9, 30), "Bahceci", "CVC Capital (x)"
    ws["E5"], ws["G5"], ws["H5"] = 75, "Buyout/LBO", "Clinics/Outpatient"
    wb.create_sheet("Disclaimer")
    wb.save(path)


def _annual_book(path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Q1 2019"
    ws["D4"], ws["F4"] = "Buying Entity", "Selling Entity"
    ws["B7"], ws["C7"], ws["D7"], ws["E7"] = (
        "Date", "Target Company", "Company", "Private Equity Firm")
    ws["F7"], ws["G7"], ws["H7"], ws["I7"] = (
        "Selling Entity", "Size ($ M)", "Sector", "Deal Type")
    ws["B8"], ws["C8"], ws["D8"], ws["E8"] = (
        datetime(2019, 3, 29), "Tangible Difference", "Blue Sprig", "KKR & Co.")
    ws["H8"], ws["I8"] = "Healthcare Services", "Add-on"
    wb.save(path)


class WorkbookParseTests(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()

    def test_hcpea_layout(self):
        from pathlib import Path
        p = Path(self.dir) / "HCPEA_Q325.xlsx"
        _hcpea_book(p)
        rows = ing.parse_workbook(p)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r["company_name"], "Bahceci")
        self.assertEqual(r["buyer_name"], "CVC Capital (x)")
        self.assertEqual(r["sponsor_owner"], "CVC Capital")
        self.assertEqual(r["deal_quarter"], "2025Q3")
        self.assertEqual(r["transaction_value"], "75.0")
        self.assertEqual(r["deal_type"], "Buyout/LBO")

    def test_annual_layout_and_load(self):
        from pathlib import Path
        p = Path(self.dir) / "2019DealTracker.xlsx"
        _annual_book(p)
        rows = ing.parse_workbook(p)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r["company_name"], "Tangible Difference")
        self.assertEqual(r["buyer_name"], "Blue Sprig")        # bare "Company" col
        self.assertEqual(r["sponsor_owner"], "KKR & Co.")
        self.assertEqual(r["deal_quarter"], "2019Q1")
        self.assertEqual(r["deal_type"], "Add-On")

        # End-to-end: the loader upserts into deal_library_companies.
        db = os.path.join(self.dir, "dl.db")
        ing.main(["--files", str(p), "--db", db])
        store = PortfolioStore(db)
        self.assertEqual(dl.count(store), 1)
        # Idempotent: re-running does not duplicate (upsert on company_id).
        ing.main(["--files", str(p), "--db", db])
        self.assertEqual(dl.count(store), 1)


if __name__ == "__main__":
    unittest.main()
