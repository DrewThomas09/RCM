"""Tests for Phase J modules: exit_package, market_intelligence, deal_query.

Covers Prompts 51-53 with 18 tests total.
"""
import json
import os
import sqlite3
import tempfile
import unittest
import zipfile
import zlib
from pathlib import Path

import pandas as pd


# ── Helpers ───────────────────────────────────────────────────────

def _make_store(db_path: str):
    """Create a minimal PortfolioStore-like object for testing."""
    from rcm_mc.portfolio.store import PortfolioStore
    store = PortfolioStore(db_path)
    store.init_db()
    return store


def _seed_plan(store, deal_id="DEAL001"):
    """Insert a value creation plan for the given deal."""
    from rcm_mc.pe.value_creation_plan import (
        Initiative, ValueCreationPlan, save_plan,
    )
    plan = ValueCreationPlan(
        deal_id=deal_id,
        plan_name="Test Deal — Value Creation Plan",
        created_at="2026-01-01T00:00:00+00:00",
        created_by="test",
        initiatives=[
            Initiative(
                initiative_id="init-idr",
                name="Improve initial denial rate",
                lever_key="initial_denial_rate",
                current_value=0.15,
                target_value=0.08,
                target_ebitda_impact=2_500_000,
                status="in_progress",
            ),
            Initiative(
                initiative_id="init-fwr",
                name="Reduce final write off rate",
                lever_key="final_write_off_rate",
                current_value=0.05,
                target_value=0.02,
                target_ebitda_impact=1_000_000,
                status="completed",
            ),
        ],
        total_target_ebitda=3_500_000,
    )
    save_plan(store, plan)
    return plan


def _seed_deal(store, deal_id="DEAL001", name="Test Hospital"):
    """Insert a deal row so FK constraints are satisfied."""
    with store.connect() as con:
        con.execute(
            "INSERT OR IGNORE INTO deals (deal_id, name, created_at) VALUES (?, ?, ?)",
            (deal_id, name, "2026-01-01"),
        )
        con.commit()


def _seed_packet(store, deal_id="DEAL001", deal_name="Test Hospital",
                 state="IL", bed_count=200, idr=12.5, ebitda=5_000_000):
    """Insert a minimal analysis packet into analysis_runs."""
    from rcm_mc.analysis.packet import DealAnalysisPacket, HospitalProfile, ObservedMetric, EBITDABridgeResult
    packet = DealAnalysisPacket(
        deal_id=deal_id,
        deal_name=deal_name,
        run_id=f"run-{deal_id}",
        profile=HospitalProfile(state=state, bed_count=bed_count),
        observed_metrics={
            "initial_denial_rate": ObservedMetric(value=idr),
        },
        ebitda_bridge=EBITDABridgeResult(total_ebitda_impact=ebitda),
    )
    from rcm_mc.analysis.analysis_store import save_packet
    save_packet(store, packet, inputs_hash=f"hash-{deal_id}")
    return packet


def _make_hospitals_df():
    """Create a small in-memory HCRIS-like DataFrame for testing."""
    return pd.DataFrame([
        {"ccn": "140001", "name": "Chicago General", "city": "Chicago", "state": "IL", "beds": 400},
        {"ccn": "140002", "name": "Springfield Memorial", "city": "Springfield", "state": "IL", "beds": 150},
        {"ccn": "140003", "name": "Peoria Health", "city": "Peoria", "state": "IL", "beds": 200},
        {"ccn": "170001", "name": "Kansas City Hospital", "city": "Kansas City", "state": "KS", "beds": 300},
        {"ccn": "050001", "name": "LA Medical Center", "city": "Los Angeles", "state": "CA", "beds": 800},
    ])


# ══════════════════════════════════════════════════════════════════
# Test Suite 1: exports/exit_package.py (Prompt 51)
# ══════════════════════════════════════════════════════════════════

class TestExitPackage(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "test.db")
        self.store = _make_store(self.db_path)
        _seed_deal(self.store, "DEAL001")
        _seed_plan(self.store, "DEAL001")

    def tearDown(self):
        self.tmp.cleanup()

    def test_generate_exit_package_returns_zip_path(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        self.assertIsInstance(result, Path)
        self.assertTrue(result.exists())
        self.assertTrue(str(result).endswith(".zip"))

    def test_zip_contains_required_files(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        with zipfile.ZipFile(result) as z:
            names = z.namelist()
        self.assertIn("01_Exit_Memo.html", names)
        self.assertIn("02_Value_Creation_Summary.xlsx", names)
        self.assertIn("03_Buyer_Data_Room_Checklist.md", names)
        self.assertIn("manifest.json", names)

    def test_exit_memo_html_references_plan(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        with zipfile.ZipFile(result) as z:
            memo = z.read("01_Exit_Memo.html").decode()
        self.assertIn("Value Creation", memo)
        self.assertIn("initial denial rate", memo.lower().replace("_", " ").replace("-", " "))

    def test_manifest_json_valid(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        with zipfile.ZipFile(result) as z:
            manifest = json.loads(z.read("manifest.json"))
        self.assertEqual(manifest["deal_id"], "DEAL001")
        self.assertEqual(manifest["package_type"], "exit")
        self.assertIn("files", manifest)
        self.assertIn("01_Exit_Memo.html", manifest["files"])

    def test_checklist_references_initiatives(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        with zipfile.ZipFile(result) as z:
            checklist = z.read("03_Buyer_Data_Room_Checklist.md").decode()
        self.assertIn("Improve initial denial rate", checklist)
        self.assertIn("Buyer Data Room", checklist)

    def test_xlsx_is_valid_workbook(self):
        from rcm_mc.exports.exit_package import generate_exit_package
        from openpyxl import load_workbook
        from io import BytesIO
        out_dir = Path(self.tmp.name) / "out"
        result = generate_exit_package(self.store, "DEAL001", out_dir=out_dir)
        with zipfile.ZipFile(result) as z:
            xlsx_bytes = z.read("02_Value_Creation_Summary.xlsx")
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        self.assertEqual(ws.title, "Value Creation Summary")
        # Header row + 2 initiative rows.
        self.assertGreaterEqual(ws.max_row, 3)


# ══════════════════════════════════════════════════════════════════
# Test Suite 2: data/market_intelligence.py (Prompt 52)
# ══════════════════════════════════════════════════════════════════

class TestMarketIntelligence(unittest.TestCase):
    def test_haversine_same_point_is_zero(self):
        from rcm_mc.data.market_intelligence import haversine_miles
        d = haversine_miles(41.8781, -87.6298, 41.8781, -87.6298)
        self.assertAlmostEqual(d, 0.0, places=5)

    def test_haversine_known_distance(self):
        """Chicago to Springfield IL is roughly 200 miles."""
        from rcm_mc.data.market_intelligence import haversine_miles
        d = haversine_miles(41.8781, -87.6298, 39.7817, -89.6501)
        self.assertGreater(d, 100)
        self.assertLess(d, 300)

    def test_find_competitors_same_state(self):
        from rcm_mc.data.market_intelligence import find_competitors
        df = _make_hospitals_df()
        # All IL hospitals have the same state centroid, so distance = 0.
        comps = find_competitors("140001", hospitals_df=df, radius_miles=50)
        ccns = {c.ccn for c in comps}
        # Other IL hospitals should be within 0 miles (same centroid).
        self.assertIn("140002", ccns)
        self.assertIn("140003", ccns)
        # CA hospital should be far away.
        self.assertNotIn("050001", ccns)

    def test_find_competitors_excludes_self(self):
        from rcm_mc.data.market_intelligence import find_competitors
        df = _make_hospitals_df()
        comps = find_competitors("140001", hospitals_df=df, radius_miles=5000)
        ccns = {c.ccn for c in comps}
        self.assertNotIn("140001", ccns)

    def test_find_competitors_bed_count(self):
        from rcm_mc.data.market_intelligence import find_competitors
        df = _make_hospitals_df()
        comps = find_competitors("140001", hospitals_df=df, radius_miles=50)
        for c in comps:
            if c.ccn == "140002":
                self.assertEqual(c.bed_count, 150)

    def test_market_summary_hhi_dominant(self):
        """One large hospital dominates -> HHI high -> 'dominant'."""
        from rcm_mc.data.market_intelligence import Competitor, market_summary
        comps = [Competitor(ccn="B", name="Small", distance_miles=5, bed_count=50)]
        ms = market_summary("A", comps, target_bed_count=950)
        self.assertEqual(ms.market_type, "dominant")
        self.assertGreater(ms.market_hhi, 2500)

    def test_market_summary_fragmented(self):
        """Many equal-sized hospitals -> low HHI -> 'fragmented'."""
        from rcm_mc.data.market_intelligence import Competitor, market_summary
        comps = [
            Competitor(ccn=f"H{i}", name=f"Hosp {i}", distance_miles=10, bed_count=100)
            for i in range(9)
        ]
        ms = market_summary("target", comps, target_bed_count=100)
        self.assertEqual(ms.market_type, "fragmented")
        self.assertLess(ms.market_hhi, 1500)

    def test_market_summary_zero_beds(self):
        from rcm_mc.data.market_intelligence import market_summary
        ms = market_summary("A", [], target_bed_count=0)
        self.assertEqual(ms.total_beds, 0)
        self.assertEqual(ms.market_type, "fragmented")


# ══════════════════════════════════════════════════════════════════
# Test Suite 3: analysis/deal_query.py (Prompt 53)
# ══════════════════════════════════════════════════════════════════

class TestDealQuery(unittest.TestCase):
    def test_parse_simple_gt(self):
        from rcm_mc.analysis.deal_query import parse_query
        filters = parse_query("denial rate > 10")
        self.assertEqual(len(filters), 1)
        self.assertEqual(filters[0].field, "denial rate")
        self.assertEqual(filters[0].operator, ">")
        self.assertEqual(filters[0].value, 10)

    def test_parse_state_equals(self):
        from rcm_mc.analysis.deal_query import parse_query
        filters = parse_query("state = IL")
        self.assertEqual(len(filters), 1)
        self.assertEqual(filters[0].field, "state")
        self.assertEqual(filters[0].operator, "=")
        self.assertEqual(filters[0].value, "IL")

    def test_parse_ebitda_with_suffix(self):
        from rcm_mc.analysis.deal_query import parse_query
        filters = parse_query("ebitda > 5M")
        self.assertEqual(len(filters), 1)
        self.assertEqual(filters[0].value, 5_000_000)

    def test_parse_multiple_and(self):
        from rcm_mc.analysis.deal_query import parse_query
        filters = parse_query("state = IL and ebitda > 5M")
        self.assertEqual(len(filters), 2)
        self.assertEqual(filters[0].field, "state")
        self.assertEqual(filters[1].field, "ebitda")

    def test_parse_semicolon_separator(self):
        from rcm_mc.analysis.deal_query import parse_query
        filters = parse_query("beds >= 200; state = TX")
        self.assertEqual(len(filters), 2)

    def test_execute_query_filters_deals(self):
        from rcm_mc.analysis.deal_query import parse_query, execute_query
        with tempfile.TemporaryDirectory() as tmp:
            db_path = os.path.join(tmp, "test.db")
            store = _make_store(db_path)
            _seed_deal(store, "D1", "Hospital Alpha")
            _seed_deal(store, "D2", "Hospital Beta")
            _seed_packet(store, "D1", "Hospital Alpha", state="IL", idr=12.5, ebitda=6_000_000)
            _seed_packet(store, "D2", "Hospital Beta", state="TX", idr=8.0, ebitda=3_000_000)

            # Query: state = IL
            filters = parse_query("state = IL")
            results = execute_query(store, filters)
            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].deal_id, "D1")

    def test_execute_query_numeric_filter(self):
        from rcm_mc.analysis.deal_query import parse_query, execute_query
        with tempfile.TemporaryDirectory() as tmp:
            db_path = os.path.join(tmp, "test.db")
            store = _make_store(db_path)
            _seed_deal(store, "D1", "Hospital Alpha")
            _seed_deal(store, "D2", "Hospital Beta")
            _seed_packet(store, "D1", "Hospital Alpha", state="IL", idr=12.5)
            _seed_packet(store, "D2", "Hospital Beta", state="TX", idr=8.0)

            filters = parse_query("denial rate > 10")
            results = execute_query(store, filters)
            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].deal_id, "D1")


if __name__ == "__main__":
    unittest.main()
