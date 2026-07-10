"""Selectable enrichment (npi_cleaner.enrich) — cleaner AND enricher.

New-file tests (the main test_npi_cleaner.py contract is untouched):

  * registry() lists every enrichment with honest ready / needs_data /
    network statuses; valid_ids() drops unknown ids and duplicates.
  * The zip_cbsa reference pack parses the Census ZCTA↔CBSA relationship
    format (largest-overlap wins) and simple zip,cbsa,name files.
  * Offline appliers add service_category / specialty_name / cbsa columns
    and produce the top-codes, key-players and geography marts.
  * Medicare connectors (data.cms.gov) run over an injected opener only:
    per-HCPCS national benchmark (weighted by services, ratio + gross-up
    rollup) and per-NPI Medicare volumes folded into the key-players
    mart; a total transport failure reports connectivity, not zeros.
  * Engine integration: clean_bytes(enrichments=[...]) appends the new
    columns to the cleaned CSV + workbook input and carries the
    ``enrichment`` block on the scorecard; header collisions dedupe.
  * The executive report is a standalone <!doctype html> document (the
    fragment form was auto-wrapped in the app shell by _send_html) and
    renders the enrichment sections.
  * HTTP: GET /npi-cleaner/api/enrichments serves the registry, and an
    upload carrying ?enrich_ids= produces an enriched scorecard.
"""
import csv
import io
import json
import os
import socket
import tempfile
import threading
import time
import unittest
import urllib.request as _u
from pathlib import Path
from unittest.mock import patch

GOOD_A = "1679576722"   # Luhn-valid synthetic NPIs (same as main suite)
GOOD_B = "1234567893"


def _fresh_packs(testcase):
    """Patch refdata_packs onto a temp SQLite store for one test."""
    import rcm_mc.npi_cleaner.refdata_packs as packs
    tmp = tempfile.TemporaryDirectory()
    testcase.addCleanup(tmp.cleanup)
    p = patch.object(packs, "_DB_PATH", Path(tmp.name) / "packs.sqlite3")
    p.start()
    testcase.addCleanup(p.stop)
    packs._CACHE.clear()
    testcase.addCleanup(packs._CACHE.clear)
    return packs


CENSUS_REL = (
    b"OID_ZCTA5_20|GEOID_ZCTA5_20|AREALAND_PART|OID_CBSA_20|"
    b"GEOID_CBSA_20|NAMELSAD_CBSA_20\n"
    b"1|75001|5000|9|19100|Dallas-Fort Worth-Arlington, TX Metro Area\n"
    b"2|75001|100|9|26420|Houston-The Woodlands-Sugar Land, TX Metro Area\n"
    b"3|77002|900|9|26420|Houston-The Woodlands-Sugar Land, TX Metro Area\n"
    b"4|99999|10||\n")


class FakeResp:
    def __init__(self, body: bytes):
        self.body = body

    def read(self):
        return self.body

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False


def _cms_opener(calls=None, fail_hosts=False):
    """Injected transport for the data.cms.gov connectors."""
    def opener(req, timeout=None):
        url = req.full_url
        if calls is not None:
            calls.append(url)
        if fail_hosts:
            raise OSError("connection refused")
        if "Rndrng_Prvdr_Geo_Lvl" in url and "99213" not in url:
            body = "[]"        # no national rows for the J-code
        elif "Rndrng_Prvdr_Geo_Lvl" in url:
            body = json.dumps([
                {"HCPCS_Cd": "99213", "Tot_Srvcs": "1000",
                 "Avg_Mdcr_Alowd_Amt": "90.00",
                 "Avg_Mdcr_Pymt_Amt": "70.00"},
                {"HCPCS_Cd": "99213", "Tot_Srvcs": "1000",
                 "Avg_Mdcr_Alowd_Amt": "110.00",
                 "Avg_Mdcr_Pymt_Amt": "80.00"},
            ])
        elif GOOD_A in url:
            body = json.dumps([
                {"Rndrng_NPI": GOOD_A,
                 "Rndrng_Prvdr_Last_Org_Name": "Mercy General",
                 "Rndrng_Prvdr_Type": "Family Practice",
                 "Rndrng_Prvdr_State_Abrvtn": "TX",
                 "Tot_Srvcs": "5000", "Tot_Benes": "800",
                 "Tot_Mdcr_Pymt_Amt": "400000.50",
                 "Tot_Sbmtd_Chrg": "900000"}])
        else:
            body = "[]"
        return FakeResp(body.encode())
    return opener


def _table():
    headers = ["ClaimID", "BillingProviderNPI", "OrganizationName", "Zip",
               "ChargeAmt", "DateOfService", "HCPCS", "POS", "Taxonomy"]
    rows = [
        ["1", GOOD_A, "Mercy General", "75001", "420",
         "2024-01-11", "99213", "11", "207Q00000X"],
        ["2", GOOD_A, "Mercy General", "75001", "430",
         "2024-02-11", "99213", "11", "207Q00000X"],
        ["3", GOOD_B, "Riverbend", "77002", "180",
         "2024-03-02", "J1745", "22", "207R00000X"],
        ["4", GOOD_B, "Riverbend", "77002", "190",
         "2024-04-02", "J1745", "22", "207R00000X"],
        ["5", GOOD_A, "Mercy General", "75001", "440",
         "2024-05-11", "99213", "11", "207Q00000X"],
        ["6", GOOD_A, "Mercy General", "75001", "450",
         "2024-06-11", "99213", "11", "207Q00000X"],
    ]
    idx = {"hcpcs_i": 6, "billed_i": 4, "dos_i": 5, "billing_idx": 1,
           "name_idx": 2, "zip_set": {3}, "taxo_set": {8},
           "rev_set": set(), "pos_set": {7}, "tob_i": None}
    return headers, [list(r) for r in rows], idx


class TestRegistry(unittest.TestCase):
    def setUp(self):
        _fresh_packs(self)
        from rcm_mc.npi_cleaner import enrich
        self.E = enrich

    def test_registry_lists_every_enrichment_with_honest_status(self):
        by_id = {e["id"]: e for e in self.E.registry()}
        self.assertEqual(set(by_id), set(self.E.ALL_IDS))
        # No pack installed → geo_msa is needs_data and NOT default-on.
        self.assertEqual(by_id["geo_msa"]["status"], "needs_data")
        self.assertFalse(by_id["geo_msa"]["default"])
        for oid in ("service_category", "taxonomy_specialty",
                    "top_codes_trend", "provider_revenue"):
            self.assertEqual(by_id[oid]["status"], "ready")
        for nid in self.E.NETWORK_IDS:
            self.assertEqual(by_id[nid]["mode"], "network")
            self.assertFalse(by_id[nid]["default"])
        # Column claims are stated up front for the pivot.
        self.assertIn("cbsa_code", by_id["geo_msa"]["adds"])
        self.assertIn("medicare_avg_allowed",
                      by_id["medicare_hcpcs_benchmark"]["adds"])

    def test_registry_marks_geo_ready_once_pack_installed(self):
        packs = _fresh_packs(self)
        packs.install_from_bytes("zip_cbsa", CENSUS_REL, source="census")
        by_id = {e["id"]: e for e in self.E.registry()}
        self.assertEqual(by_id["geo_msa"]["status"], "ready")
        self.assertTrue(by_id["geo_msa"]["default"])

    def test_valid_ids_drops_unknown_and_duplicates(self):
        self.assertEqual(
            self.E.valid_ids(["geo_msa", "bogus", "geo_msa",
                              "top_codes_trend", ""]),
            ["geo_msa", "top_codes_trend"])
        self.assertEqual(self.E.valid_ids(None), [])


class TestZipCbsaPack(unittest.TestCase):
    def test_census_relationship_parse_largest_overlap_wins(self):
        packs = _fresh_packs(self)
        info = packs.install_from_bytes("zip_cbsa", CENSUS_REL,
                                        source="census")
        self.assertEqual(info["rows"], 2)   # 99999 has no CBSA → dropped
        lut = packs.zip_cbsa_lookup()
        self.assertEqual(lut["75001"][0], "19100")   # 5000 beats 100
        self.assertIn("Dallas", lut["75001"][1])
        self.assertEqual(lut["77002"][0], "26420")

    def test_simple_zip_cbsa_csv_also_installs(self):
        packs = _fresh_packs(self)
        raw = (b"zip,cbsa,cbsa_name\n"
               b"10001,35620,New York-Newark-Jersey City NY-NJ-PA\n")
        packs.install_from_bytes("zip_cbsa", raw, source="hud export")
        self.assertEqual(packs.zip_cbsa_lookup()["10001"][0], "35620")

    def test_not_a_crosswalk_raises(self):
        packs = _fresh_packs(self)
        with self.assertRaises(ValueError):
            packs.install_from_bytes("zip_cbsa", b"a,b\n1,2\n",
                                     source="junk")


class TestOfflineAppliers(unittest.TestCase):
    def setUp(self):
        self.packs = _fresh_packs(self)
        from rcm_mc.npi_cleaner import enrich
        self.E = enrich

    def test_offline_columns_and_marts(self):
        self.packs.install_from_bytes("zip_cbsa", CENSUS_REL,
                                      source="census")
        headers, rows, idx = _table()
        out = self.E.apply(headers, rows, idx, list(self.E.OFFLINE_IDS))
        self.assertEqual(out["added_headers"],
                         ["service_category", "service_subcategory",
                          "specialty_name", "cbsa_code", "cbsa_name"])
        for col in out["added_columns"]:
            self.assertEqual(len(col), len(rows))
        # Row 0: office E&M in Dallas by a family-medicine taxonomy.
        cols = dict(zip(out["added_headers"], out["added_columns"]))
        self.assertEqual(cols["service_category"][0], "Office")
        self.assertEqual(cols["cbsa_code"][0], "19100")
        self.assertTrue(cols["specialty_name"][0])
        marts = out["marts"]
        top = {c["code"]: c for c in marts["top_codes"]["codes"]}
        self.assertEqual(top["99213"]["lines"], 4)
        self.assertIsNotNone(top["99213"]["pct_dollars"])
        players = marts["provider_revenue"]["providers"]
        self.assertEqual(players[0]["npi"], GOOD_A)   # highest dollars
        self.assertIsNotNone(marts["provider_revenue"]["hhi"])
        geo = {a["cbsa"]: a for a in marts["geography"]["areas"]}
        self.assertEqual(geo["19100"]["lines"], 4)

    def test_missing_columns_degrade_to_notes_not_errors(self):
        headers = ["A", "B"]
        rows = [["1", "2"]]
        out = self.E.apply(headers, rows, {}, list(self.E.OFFLINE_IDS))
        self.assertEqual(out["added_headers"], [])
        notes = {r["id"]: r["note"] for r in out["results"]}
        self.assertIn("No service columns", notes["service_category"])
        self.assertIn("No billing NPI", notes["provider_revenue"])

    def test_column_name_collision_dedupes(self):
        headers, rows, idx = _table()
        headers = headers + ["service_category"]
        for r in rows:
            r.append("preexisting")
        out = self.E.apply(headers, rows, idx, ["service_category"])
        self.assertEqual(out["added_headers"],
                         ["service_category_2", "service_subcategory"])


class TestMedicareConnectors(unittest.TestCase):
    def setUp(self):
        _fresh_packs(self)
        from rcm_mc.npi_cleaner import enrich
        self.E = enrich

    def test_benchmark_weighted_ratio_and_grossup(self):
        headers, rows, idx = _table()
        calls = []
        out = self.E.apply(headers, rows, idx,
                           ["medicare_hcpcs_benchmark"],
                           opener=_cms_opener(calls))
        mb = out["marts"]["medicare_benchmark"]
        by_code = {c["code"]: c for c in mb["codes"]}
        # Two national place-of-service rows, equal weights → 100.00.
        self.assertAlmostEqual(by_code["99213"]["medicare_avg_allowed"],
                               100.0)
        # File avg for 99213 = (420+430+440+450)/4 = 435 → 4.35x.
        self.assertAlmostEqual(by_code["99213"]["ratio"], 4.35)
        # The J-code got no benchmark rows → matched lines are the 4
        # 99213 lines; Medicare-equivalent = 4 × $100.
        self.assertEqual(mb["matched_lines"], 4)
        self.assertAlmostEqual(mb["medicare_equivalent_dollars"], 400.0)
        self.assertAlmostEqual(mb["matched_dollars"], 1740.0)
        self.assertAlmostEqual(mb["pct_of_medicare"], 435.0)
        # Row column present and aligned; unmatched lines blank.
        cols = dict(zip(out["added_headers"], out["added_columns"]))
        self.assertEqual(cols["medicare_avg_allowed"][0], "100.00")
        self.assertEqual(cols["medicare_avg_allowed"][2], "")
        # Distinct codes only → 2 calls, not 6.
        self.assertEqual(len(calls), 2)

    def test_provider_volumes_fold_into_key_players(self):
        headers, rows, idx = _table()
        out = self.E.apply(headers, rows, idx,
                           ["provider_revenue",
                            "medicare_provider_volume"],
                           opener=_cms_opener())
        players = {p["npi"]: p
                   for p in out["marts"]["provider_revenue"]["providers"]}
        self.assertAlmostEqual(players[GOOD_A]["medicare_payment"],
                               400000.50)
        self.assertEqual(players[GOOD_A]["provider_type"],
                         "Family Practice")
        self.assertIsNone(players[GOOD_B].get("medicare_payment"))

    def test_total_transport_failure_reports_connectivity(self):
        headers, rows, idx = _table()
        out = self.E.apply(headers, rows, idx,
                           ["medicare_hcpcs_benchmark",
                            "medicare_provider_volume"],
                           opener=_cms_opener(fail_hosts=True))
        self.assertEqual(out["added_headers"], [])
        for r in out["results"]:
            self.assertIn("Could not reach data.cms.gov", r["note"])

    def test_env_override_changes_endpoints(self):
        headers, rows, idx = _table()
        calls = []
        with patch.dict(os.environ, {
                "RCM_MC_CMS_GEO_SERVICE_URL": "https://mirror/geo",
                "RCM_MC_CMS_BY_PROVIDER_URL": "https://mirror/prov"}):
            self.E.apply(headers, rows, idx,
                         ["medicare_hcpcs_benchmark",
                          "medicare_provider_volume"],
                         opener=_cms_opener(calls))
        self.assertTrue(all(u.startswith("https://mirror/")
                            for u in calls), calls)


class TestEngineIntegration(unittest.TestCase):
    def setUp(self):
        self.packs = _fresh_packs(self)

    def _csv_bytes(self):
        headers, rows, _idx = _table()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        return buf.getvalue().encode()

    def test_clean_bytes_appends_columns_and_scorecard_block(self):
        self.packs.install_from_bytes("zip_cbsa", CENSUS_REL,
                                      source="census")
        from rcm_mc.npi_cleaner import engine
        res = engine.clean_bytes(
            self._csv_bytes(), "claims.csv",
            enrichments=["service_category", "taxonomy_specialty",
                         "geo_msa", "top_codes_trend",
                         "provider_revenue"])
        sc = res.as_scorecard()
        enr = sc["enrichment"]
        self.assertIn("cbsa_name", enr["columns_added"])
        self.assertIn("top_codes", enr["marts"])
        with open(res.out_path, encoding="utf-8") as fh:
            r = csv.reader(fh)
            hdr = next(r)
            row1 = next(r)
        self.assertIn("service_category", hdr)
        self.assertIn("cbsa_code", hdr)
        self.assertEqual(row1[hdr.index("cbsa_code")], "19100")
        self.assertEqual(row1[hdr.index("service_category")], "Office")

    def test_no_enrichments_leaves_scorecard_null(self):
        from rcm_mc.npi_cleaner import engine
        res = engine.clean_bytes(self._csv_bytes(), "claims.csv")
        self.assertIsNone(res.as_scorecard()["enrichment"])
        with open(res.out_path, encoding="utf-8") as fh:
            hdr = next(csv.reader(fh))
        self.assertNotIn("service_category", hdr)

    def test_unknown_ids_are_ignored_not_fatal(self):
        from rcm_mc.npi_cleaner import engine
        res = engine.clean_bytes(self._csv_bytes(), "claims.csv",
                                 enrichments=["nope", "top_codes_trend"])
        enr = res.as_scorecard()["enrichment"]
        self.assertEqual(enr["requested"], ["top_codes_trend"])


class TestExecReportStandalone(unittest.TestCase):
    def test_exec_report_is_a_full_document_not_a_fragment(self):
        # The fragment form (leading <style>) tripped _send_html's
        # auto-wrap and rendered the one-pager inside the app shell.
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        sc = {"quality": {"score": 92, "letter": "B",
                          "dimensions": {"completeness": 88.8}},
              "rows_in": 6, "rows_out": 6, "duplicates_removed": 0,
              "repairs_total": 0, "changes_logged": 0}
        html = build_exec_report(sc, "claims & co.csv", "2026-07-10")
        self.assertTrue(html.lower().startswith("<!doctype html"))
        self.assertIn("</body></html>", html)
        self.assertIn("<title>", html)
        self.assertIn("claims &amp; co.csv", html)

    def test_exec_report_renders_enrichment_sections(self):
        from rcm_mc.npi_cleaner.exec_report import build_exec_report
        sc = {"quality": {"score": 92, "letter": "B", "dimensions": {}},
              "rows_in": 1, "rows_out": 1, "duplicates_removed": 0,
              "repairs_total": 0, "changes_logged": 0,
              "enrichment": {
                  "columns_added": ["cbsa_code"],
                  "results": [{"id": "geo_msa", "label": "MSA / CBSA",
                               "rows_enriched": 5,
                               "columns_added": ["cbsa_code"],
                               "note": "mapped"}],
                  "marts": {
                      "top_codes": {"codes": [
                          {"code": "99213", "lines": 4, "charges": 100.0,
                           "pct_dollars": 50.0,
                           "trend": {"direction": "rising",
                                     "change_pct": 12.0,
                                     "window": "2024-01–2024-06"}}]},
                      "provider_revenue": {"providers": [
                          {"npi": GOOD_A, "name": "Mercy", "lines": 4,
                           "charges": 100.0, "pct_dollars": 100.0,
                           "medicare_payment": 5.0}], "hhi": 10000},
                      "geography": {"areas": [
                          {"cbsa": "19100", "name": "Dallas", "lines": 4,
                           "charges": 100.0, "pct_dollars": 100.0}],
                          "unmatched_pct": 0.0},
                      "medicare_benchmark": {
                          "codes": [{"code": "99213",
                                     "file_avg_charge": 435.0,
                                     "medicare_avg_allowed": 100.0,
                                     "ratio": 4.35,
                                     "medicare_services": 2000}],
                          "pct_of_medicare": 435.0,
                          "matched_dollars": 1740.0,
                          "medicare_equivalent_dollars": 400.0,
                          "note": "gross-up basis"},
                  }}}
        html = build_exec_report(sc, "x.csv", "2026-07-10")
        self.assertIn("Top procedure codes", html)
        self.assertIn("Key players", html)
        self.assertIn("CBSA / metro mix", html)
        self.assertIn("Medicare benchmark", html)
        self.assertIn("4.35x", html)
        self.assertIn("Enrichment applied", html)


class TestWorkbookSheet(unittest.TestCase):
    def test_enrichment_sheet_present_when_marts_exist(self):
        packs = _fresh_packs(self)
        packs.install_from_bytes("zip_cbsa", CENSUS_REL, source="census")
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        headers, rows, _idx = _table()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        from rcm_mc.npi_cleaner import engine
        res = engine.clean_bytes(
            buf.getvalue().encode(), "claims.csv",
            enrichments=["geo_msa", "top_codes_trend",
                         "provider_revenue"])
        self.assertTrue(res.workbook_path)
        wb = load_workbook(res.workbook_path, read_only=True)
        self.assertIn("Enrichment", wb.sheetnames)
        # The cleaned-data tab carries the appended columns too.
        ws = wb["Cleaned data"]
        first = [c.value for c in next(ws.iter_rows(max_row=1))]
        self.assertIn("cbsa_code", first)


class TestHttpSurface(unittest.TestCase):
    def _start(self, tmp):
        from rcm_mc.server import build_server
        s = socket.socket()
        s.bind(("127.0.0.1", 0))
        port = s.getsockname()[1]
        s.close()
        server, _ = build_server(port=port,
                                 db_path=os.path.join(tmp, "p.db"))
        t = threading.Thread(target=server.serve_forever, daemon=True)
        t.start()
        time.sleep(0.05)
        self.addCleanup(server.shutdown)
        return port

    def test_enrichments_api_and_enriched_upload(self):
        _fresh_packs(self)
        with tempfile.TemporaryDirectory() as tmp:
            port = self._start(tmp)
            with _u.urlopen(f"http://127.0.0.1:{port}"
                            "/npi-cleaner/api/enrichments") as r:
                reg = json.loads(r.read().decode())["enrichments"]
            self.assertIn("service_category", {e["id"] for e in reg})

            headers, rows, _idx = _table()
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(headers)
            w.writerows(rows)
            req = _u.Request(
                f"http://127.0.0.1:{port}/npi-cleaner/upload"
                "?enrich_ids=service_category,top_codes_trend,bogus",
                data=buf.getvalue().encode(), method="POST",
                headers={"X-Filename": "claims.csv"})
            with _u.urlopen(req) as r:
                jid = json.loads(r.read().decode())["job_id"]
            sc = None
            for _ in range(100):
                with _u.urlopen(f"http://127.0.0.1:{port}"
                                f"/npi-cleaner/status/{jid}") as r:
                    j = json.loads(r.read().decode())
                if j.get("done"):
                    sc = j["scorecard"]
                    break
                time.sleep(0.05)
            self.assertIsNotNone(sc, "job never finished")
            enr = sc["enrichment"]
            self.assertEqual(enr["requested"],
                             ["service_category", "top_codes_trend"])
            self.assertIn("service_category", enr["columns_added"])

            # The exec one-pager stays a standalone document over HTTP —
            # no app-shell topbar wrapped around it.
            with _u.urlopen(f"http://127.0.0.1:{port}"
                            f"/npi-cleaner/download/{jid}?fmt=exec") as r:
                doc = r.read().decode()
            self.assertTrue(doc.lower().startswith("<!doctype html"))
            self.assertNotIn("ck-topbar", doc)


if __name__ == "__main__":
    unittest.main()
