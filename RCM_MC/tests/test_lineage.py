"""Tests for the Lineage tab — per-metric formula + source-grade rollup.

Covers the pure-logic helper :func:`rcm_mc._bundle._lineage_tab` and its
integration into the diligence workbook.
"""
from __future__ import annotations

import json
import os
import tempfile
import unittest
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook

from rcm_mc.infra._bundle import _lineage_tab, write_diligence_workbook


BASE_DIR = Path(__file__).resolve().parents[1]


def _write_provenance(
    tmp: str,
    *,
    metrics: list,
    classification: dict,
) -> str:
    path = os.path.join(tmp, "provenance.json")
    with open(path, "w") as f:
        json.dump(
            {
                "run": {"n_sims": 500},
                "metrics": metrics,
                "sources": {
                    "classification": classification,
                    "counts": {"observed": 0, "prior": 0, "assumed": 0, "total": 0},
                    "grade": "D",
                },
            },
            f,
        )
    return path


class TestLineageTab(unittest.TestCase):
    def test_missing_provenance_returns_none(self):
        with tempfile.TemporaryDirectory() as tmp:
            self.assertIsNone(_lineage_tab(tmp))

    def test_empty_metrics_returns_none(self):
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(tmp, metrics=[], classification={})
            self.assertIsNone(_lineage_tab(tmp))

    def test_columns_are_stable(self):
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "ebitda_drag", "formula": "sum of drag components",
                          "config_keys": ["payers.*.denials"], "caveats": []}],
                classification={"payers.Medicare.denials.idr": "observed"},
            )
            df = _lineage_tab(tmp)
            self.assertEqual(
                set(df.columns),
                {"metric", "source_summary", "config_keys", "formula", "caveats"},
            )

    def test_source_summary_counts_matched_paths(self):
        """`payers.*.denials.idr` should expand to all 4 payers."""
        with tempfile.TemporaryDirectory() as tmp:
            classification = {
                "payers.Medicare.denials.idr":   "observed",
                "payers.Medicaid.denials.idr":   "observed",
                "payers.Commercial.denials.idr": "prior",
                "payers.SelfPay.denials.idr":    "assumed",
            }
            _write_provenance(
                tmp,
                metrics=[{"metric": "idr_driven_metric",
                          "formula": "f(idr)",
                          "config_keys": ["payers.*.denials.idr"],
                          "caveats": []}],
                classification=classification,
            )
            df = _lineage_tab(tmp)
            self.assertEqual(len(df), 1)
            summary = df.iloc[0]["source_summary"]
            # 2 observed, 1 prior, 1 assumed
            self.assertIn("2 observed", summary)
            self.assertIn("1 prior", summary)
            self.assertIn("1 assumed", summary)

    def test_no_matching_paths_renders_dash(self):
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "orphan", "formula": "?",
                          "config_keys": ["nonexistent.*.path"], "caveats": []}],
                classification={"payers.Medicare.denials.idr": "observed"},
            )
            df = _lineage_tab(tmp)
            self.assertEqual(df.iloc[0]["source_summary"], "—")

    def test_caveats_joined_with_semicolons(self):
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "x", "formula": "f",
                          "config_keys": [],
                          "caveats": ["caveat A", "caveat B"]}],
                classification={},
            )
            df = _lineage_tab(tmp)
            self.assertIn("caveat A", df.iloc[0]["caveats"])
            self.assertIn(";", df.iloc[0]["caveats"])

    def test_long_formula_truncated(self):
        long = "x" * 2000
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "x", "formula": long, "config_keys": [], "caveats": []}],
                classification={},
            )
            df = _lineage_tab(tmp)
            self.assertLessEqual(len(df.iloc[0]["formula"]), 500)


class TestLineageInWorkbook(unittest.TestCase):
    """Integration: end-to-end run of write_diligence_workbook produces Lineage tab."""

    def test_lineage_tab_written_when_provenance_present(self):
        cfg_path = BASE_DIR / "configs" / "actual.yaml"
        with open(cfg_path) as f:
            cfg = yaml.safe_load(f)
        summary = pd.DataFrame({
            "mean": [5e6, 2e6],
            "p10": [3e6, 1e6],
            "p90": [7e6, 3e6],
        }, index=["ebitda_drag", "economic_drag"])

        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "ebitda_drag",
                          "formula": "drag_rcm_ebitda_impact",
                          "config_keys": ["payers.*.denials"],
                          "caveats": ["scrub applies"]}],
                classification={"payers.Medicare.denials.idr": "observed"},
            )
            write_diligence_workbook(tmp, summary, cfg)
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            self.assertIn("Lineage", wb.sheetnames)
            ws = wb["Lineage"]
            # header row + at least one data row
            self.assertGreaterEqual(ws.max_row, 2)
            # UI-5 may insert a note on row 1; header falls to row 2
            headers = [cell.value for cell in ws[1]]
            if "metric" not in headers:
                headers = [cell.value for cell in ws[2]]
            for col in ("metric", "source_summary", "config_keys", "formula", "caveats"):
                self.assertIn(col, headers)
            wb.close()

    def test_cover_sheet_toc_lists_lineage(self):
        cfg_path = BASE_DIR / "configs" / "actual.yaml"
        with open(cfg_path) as f:
            cfg = yaml.safe_load(f)
        summary = pd.DataFrame({"mean": [1e6], "p10": [5e5], "p90": [2e6]},
                               index=["ebitda_drag"])
        with tempfile.TemporaryDirectory() as tmp:
            _write_provenance(
                tmp,
                metrics=[{"metric": "ebitda_drag", "formula": "f",
                          "config_keys": [], "caveats": []}],
                classification={},
            )
            write_diligence_workbook(tmp, summary, cfg)
            wb = load_workbook(os.path.join(tmp, "diligence_workbook.xlsx"))
            cover_text = " ".join(
                str(c.value) for row in wb["Cover"].iter_rows()
                for c in row if c.value
            )
            self.assertIn("Lineage", cover_text)
            wb.close()
