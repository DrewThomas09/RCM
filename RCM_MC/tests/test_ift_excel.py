"""Real-path tests for the IFT market-study Excel workbook (/api/ift/markets.xlsx).

Pins that the workbook always builds a real, openable .xlsx (a zip of XML parts)
carrying every analytic layer by market, and that it degrades — never raises —
when an analytic is unavailable, because the download must always succeed.
"""
from __future__ import annotations

import io
import unittest
import zipfile

from rcm_mc.market_reports.ift_excel import ift_workbook_xlsx


class IftWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.data = ift_workbook_xlsx()

    def test_returns_a_real_xlsx_zip(self):
        self.assertIsInstance(self.data, bytes)
        self.assertGreater(len(self.data), 5000)
        self.assertEqual(self.data[:2], b"PK")            # zip magic
        z = zipfile.ZipFile(io.BytesIO(self.data))
        names = z.namelist()
        self.assertIn("xl/workbook.xml", names)
        self.assertIn("[Content_Types].xml", names)
        self.assertIsNone(z.testzip())                     # not corrupt

    def _sheet_names(self):
        z = zipfile.ZipFile(io.BytesIO(self.data))
        import re
        from xml.sax.saxutils import unescape
        wb = z.read("xl/workbook.xml").decode()
        return [unescape(n) for n in re.findall(r'name="([^"]+)"', wb)]

    def test_carries_the_full_market_study_spine(self):
        names = self._sheet_names()
        # the funnel + every deepened analytic layer is present as its own sheet
        for expected in ("Overview", "TAM build", "SAM health systems",
                         "SOM footprint", "Markets structure", "Competitive",
                         "Insourcing", "Moat scorecard", "Clinical demand",
                         "Three-lever tracker", "Provenance"):
            self.assertIn(expected, names, f"missing sheet: {expected}")

    def test_provenance_sheet_always_present(self):
        # the honesty legend is the one sheet that must never drop
        self.assertIn("Provenance", self._sheet_names())

    def test_carries_the_qualitative_narrative_sheet_set(self):
        # the "everything IFT" pack: qualitative frames + research + connectors +
        # the market-report narrative all travel in the one download.
        names = self._sheet_names()
        for expected in ("Contents", "Taxonomy", "Ecosystem & journey",
                         "Operating models", "Company profiles",
                         "MMT positioning", "Competitor types",
                         "Industry context", "Connectors", "Fee schedule",
                         "Occupancy demand", "Deal history", "Report narrative",
                         "Report economics", "Report reg & risk"):
            self.assertIn(expected, names, f"missing sheet: {expected}")

    def test_carries_all_nine_research_sections(self):
        names = self._sheet_names()
        research = [n for n in names if n.startswith("R· ")]
        self.assertGreaterEqual(len(research), 9,
                                f"expected 9 research tabs, got {research}")

    def test_carries_the_model_and_routing_depth(self):
        names = self._sheet_names()
        self.assertIn("Assumptions & levers", names)   # the model, low/central/high
        self.assertIn("Clinical routing", names)       # movement of patients

    def test_no_ghost_or_thin_sheets(self):
        # every sheet must carry real content — no title-only "ghost" pages.
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(self.data))
        for ws in wb.worksheets:
            nonempty = sum(1 for row in ws.iter_rows(values_only=True)
                           for v in row if v not in (None, ""))
            self.assertGreaterEqual(nonempty, 6,
                                    f"ghost/thin sheet: {ws.title} ({nonempty})")

    def test_no_dataclass_reprs_leak_generically(self):
        # a value rendered as `str(some_dataclass)` looks like `Name(field=...)`;
        # that must never reach a cell (risks/connections regressed here once).
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        import re
        repr_pat = re.compile(r"^[A-Z][A-Za-z0-9]+\(.*=.*\)")
        wb = openpyxl.load_workbook(io.BytesIO(self.data))
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    self.assertIsNone(repr_pat.match(str(v)),
                                      f"{ws.title}: dataclass repr leaked: {v!r}")

    def test_cells_wrap_and_hyperlinks_resolve(self):
        # formatting contract: text wraps (no clipped prose), and the Contents
        # "where this lives online" links resolve to real https targets.
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        z = zipfile.ZipFile(io.BytesIO(self.data))
        self.assertIn('wrapText="1"', z.read("xl/styles.xml").decode())
        wb = openpyxl.load_workbook(io.BytesIO(self.data))
        links = [c.hyperlink.target for row in wb["Contents"].iter_rows()
                 for c in row if c.hyperlink]
        self.assertGreaterEqual(len(links), 5)
        self.assertTrue(all(u.startswith("https://") for u in links))

    def test_contents_is_the_first_sheet(self):
        # the index must be tab #1 so a partner can navigate 30+ sheets.
        self.assertEqual(self._sheet_names()[0], "Contents")

    def test_no_none_or_dataclass_leaks_in_cells(self):
        # every degraded/optional value must render a real string, never a leaked
        # repr or a bare None — the pages enforce this and so must the workbook.
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(self.data))
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    s = str(v)
                    self.assertNotIn("CompanyProfile(", s)
                    self.assertNotIn("ConnectorResult(", s)
                    self.assertFalse(s.strip() == "None", f"{ws.title}: bare None")

    def test_competitor_types_sheet_has_no_company_names(self):
        # the market-level competitor-TYPE sheet must stay company-name-free;
        # company positioning lives on the Company profiles sheet only.
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = openpyxl.load_workbook(io.BytesIO(self.data))
        ws = wb["Competitor types"]
        blob = " ".join(str(c) for row in ws.iter_rows(values_only=True)
                        for c in row if c).lower()
        for name in ("midwest medical", "amr", "global medical response",
                     "superior air", "ameripro", "ryan brothers", "acadian",
                     "bell ambulance", "cleveland clinic", "allina"):
            self.assertNotIn(name, blob, f"company name leaked: {name}")

    def test_deterministic_and_repeatable(self):
        # cached analytics -> two builds are byte-identical (no Date.now drift)
        self.assertEqual(ift_workbook_xlsx(), ift_workbook_xlsx())

    def test_degrades_when_an_analytic_is_unavailable(self):
        # If a layer raises, its sheet is skipped but the workbook still builds.
        import rcm_mc.market_reports.ift_excel as X

        orig = X._an.health_system_sam
        try:
            X._an.health_system_sam = lambda: (_ for _ in ()).throw(RuntimeError("boom"))
            data = ift_workbook_xlsx()
        finally:
            X._an.health_system_sam = orig
        self.assertEqual(data[:2], b"PK")
        self.assertGreater(len(data), 5000)


if __name__ == "__main__":
    unittest.main()
