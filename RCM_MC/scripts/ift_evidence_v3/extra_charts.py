"""Extra native charts on the load-bearing analytical tabs (v3.11).

A curated set of house-style charts on the key measured series that carried
numbers but no visual - the Medicare interfacility series, the subject-company
book, the NEMSIS type-of-service split, the acute-transfer series, the payer
revenue mix and the RSNAT savings ladder. Each is a live range reference (it
recomputes with the workbook) and is built through v3lib.add_chart, so it
inherits the house style and passes the V9 chart-integrity gate. Runs before
normalize_all_charts, which handles collision geometry.
"""
from openpyxl.utils import get_column_letter

# (tab, anchor_row, title, cat_range, [(series_name, value_range)], kind, y_fmt_attr)
# ranges are A1-style on the same tab; the pass qualifies them with the sheet.
_SPECS = [
    ('Medicare_IFT_Series', 5,
     'Medicare FFS interfacility transports per year, 2010-2024',
     'A5:A19', [('Transports', 'B5:B19')], 'line', 'FMT_INT'),
    ('Medicare_IFT_Series', 21,
     'Medicare FFS interfacility allowed dollars per year (incl mileage), '
     '2010-2024',
     'A5:A19', [('Allowed $', 'F5:F19')], 'line', 'FMT_USD'),
    ('MMT_Medicare_Book', 5,
     'MMT Medicare FFS allowed dollars by vintage',
     'A7:A9', [('Allowed $', 'E7:E9')], 'bar', 'FMT_USD'),
    ('Acute_IFT_Series', 5,
     'Acute-to-acute ED-origin transfer episodes per year, 2007-2023',
     'A5:A21', [('ED-origin transfer episodes', 'B5:B21')], 'line', 'FMT_INT'),
    ('EMS_Transports', 14,
     'NEMSIS activations by type of service, 2024',
     'A15:A21', [('2024 activations', 'E15:E21')], 'bar', 'FMT_INT'),
    ('Facility_Pay_Layer', 14,
     'Ambulance revenue per NPI by payer (GADCS mean)',
     'A15:A19', [('Revenue per NPI $', 'B15:B19')], 'bar', 'FMT_USD'),
    ('RSNAT_Series', 30,
     'RSNAT prior-authorization cumulative Medicare savings by report',
     'A32:A36', [('Cumulative savings $', 'C32:C36')], 'bar', 'FMT_USD'),
]


def add_extra_charts(wb, lib, log=None):
    added = 0
    # anchor column = two past the tab's used width, matching the existing
    # section-chart convention (charts float to the right of the data).
    for tab, row, title, cat, series, kind, fmt_attr in _SPECS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        col = get_column_letter(ws.max_column + 2)
        anchor = f'{col}{row}'
        cat_ref = f"'{tab}'!{_abs(cat)}"
        ser = [(nm, f"'{tab}'!{_abs(rng)}") for nm, rng in series]
        y_fmt = getattr(lib, fmt_attr, None)
        lib.add_chart(ws, anchor, title, cat_ref, ser, kind=kind, y_fmt=y_fmt)
        added += 1
    if log:
        log(f'extra charts: {added} added on analytical tabs')
    return added


def _abs(rng):
    """A1:B2 -> $A$1:$B$2."""
    out = []
    for part in rng.split(':'):
        col = ''.join(ch for ch in part if ch.isalpha())
        num = ''.join(ch for ch in part if ch.isdigit())
        out.append(f'${col}${num}')
    return ':'.join(out)
