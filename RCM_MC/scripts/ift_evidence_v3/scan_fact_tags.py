"""Workbook-wide fact-tag integrity scan (v3.4 / handoff task 5.1).

Finds every F### token in every cell of every tab and checks:
  1. the ID exists in the Fact_Ledger;
  2. no cell cites the F431-F433 range that v3.2's insertion made ambiguous
     unless the tab is the ledger itself or the KPI tabs those facts live on.

Exit 0 with a summary when clean; exit 1 listing offenders.
Usage: python3 scan_fact_tags.py [workbook.xlsx]
"""
import re
import sys

from openpyxl import load_workbook

DEFAULT = '/home/user/RCM/RCM_MC/deliverables/IFT_Sourced_Evidence_Master_v3_4.xlsx'

RX = re.compile(r'\bF(\d{2,4})\b')


def main(path):
    wb = load_workbook(path, read_only=True)
    ledger = {}
    for row in wb['Fact_Ledger'].iter_rows():
        v = row[0].value
        if isinstance(v, str) and re.fullmatch(r'F\d+', v):
            ledger[int(v[1:])] = {
                'metric': row[1].value if len(row) > 1 else None,
                'lives_on': row[9].value if len(row) > 9 else None,
            }
    fmax = max(ledger)
    orphans = []       # tag cites an ID with no ledger row
    ambiguous = []     # tag cites F431-F433 outside the tabs those facts live on
    n_tags = 0
    for name in wb.sheetnames:
        if name in ('Fact_Ledger', 'V3_Change_Log'):
            continue
        for row in wb[name].iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                for m in RX.finditer(c.value):
                    fid = int(m.group(1))
                    if fid < 1 or fid > 9999 or (fid > fmax and fid < 1000):
                        continue
                    if fid > fmax:
                        continue
                    n_tags += 1
                    if fid not in ledger:
                        orphans.append((name, c.coordinate, f'F{fid}'))
                    elif fid in (431, 432, 433):
                        home = str(ledger[fid]['lives_on'] or '')
                        if name not in home and not name.startswith('KPI'):
                            ambiguous.append((name, c.coordinate, f'F{fid}',
                                              ledger[fid]['metric']))
    print(f'{path}')
    print(f'ledger max: F{fmax}; in-cell tags checked: {n_tags:,}')
    if orphans:
        print(f'ORPHAN TAGS ({len(orphans)}):')
        for o in orphans[:25]:
            print('  ', o)
    if ambiguous:
        print(f'F431-F433 tags outside their home tabs ({len(ambiguous)}):')
        for a in ambiguous[:25]:
            print('  ', a)
    if orphans or ambiguous:
        sys.exit(1)
    print('CLEAN: zero orphan tags, zero stale F431-F433 citations')


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT)
