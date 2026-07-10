"""Group G expansion — ONE TAB PER METRO (20 tabs) + Metro_Index.

Per-metro SOURCED facility structure from ift_geo.metro_structure(): counts
from vendored CMS provider rolls (hospital_coords + SNF/IRF/LTCH/hospice/
home-health/dialysis), HCRIS FY2020-22 capacity join by CCN, named hospital
roster, share-of-metro and share-of-footprint formulas, one facility-mix bar
chart per tab. Metro_Index closes the family with cross-tab link counts,
live footprint SUMs and national-universe context.

EXCLUDED by design: modeled SAM/SOM dollars, discharge_base (beds x 53.3),
anchor-system / operator / insource / moat prose (PUBLIC-WEB reads).
"""
import sys

sys.path.insert(0, '/home/user/RCM/RCM_MC')
sys.path.insert(0, '/home/user/RCM')

PURPLE = 'FF7A5195'

# (metro name in ift_geo, tab slug <=31 chars)
METRO_TABS = [
    ('Omaha', 'Metro_Omaha'),
    ('Lincoln', 'Metro_Lincoln'),
    ('North Platte', 'Metro_NorthPlatte'),
    ('Columbus (NE)', 'Metro_ColumbusNE'),
    ('Grand Island / Kearney', 'Metro_GI_Kearney'),
    ('Cleveland', 'Metro_Cleveland'),
    ('Cincinnati', 'Metro_Cincinnati'),
    ('Columbus (OH)', 'Metro_ColumbusOH'),
    ('Dayton', 'Metro_Dayton'),
    ('Kansas City (bi-state)', 'Metro_KansasCity'),
    ('Wichita', 'Metro_Wichita'),
    ('Madison', 'Metro_Madison'),
    ('Milwaukee', 'Metro_Milwaukee'),
    ('Twin Cities', 'Metro_TwinCities'),
    ('Rochester (MN)', 'Metro_RochesterMN'),
    ('Des Moines', 'Metro_DesMoines'),
    ('Crown Point / NW Indiana', 'Metro_NWIndiana'),
    ('Louisville', 'Metro_Louisville'),
    ('Northern Virginia', 'Metro_NoVirginia'),
    ('Cheyenne / Casper (WY)', 'Metro_Cheyenne_Casper'),
]

SHEETS = ([{'name': tab, 'tab_color': PURPLE} for _, tab in METRO_TABS]
          + [{'name': 'Metro_Index', 'tab_color': PURPLE}])

# ── fixed row map for every metro tab (asserted during build) ──────────────
R_HOSP, R_SNF, R_IRF, R_LTCH, R_HOSPICE, R_HH, R_DIAL = 6, 7, 8, 9, 10, 11, 12
R_PA, R_NODES = 13, 14
R_SNFBEDS, R_LTCHBEDS, R_STATIONS = 18, 19, 20
R_HCRISN, R_BEDS, R_PDAYS, R_BDA, R_OCC, R_COV = 21, 22, 23, 24, 25, 26
# ── fixed row map for Metro_Index ──────────────────────────────────────────
IDX_FIRST, IDX_LAST, IDX_TOT, IDX_NAT, IDX_SHARE = 6, 25, 26, 27, 28

# per-roll dataset citation + snapshot date (verbatim source fields in the
# vendored CSVs: rcm_mc/data/*_providers.csv + hospital_coords.csv)
ROLLS = {
    'hosp': ('CMS Care Compare — Hospital General Information, addresses '
             'geocoded via US Census Geocoder (hospital_coords.csv); count '
             'of CCNs in the metro city list', '2026-05-23'),
    'snf': ('CMS Nursing Home Care Compare — Provider Information '
            '(NH_ProviderInfo); CCN count', '2026-04-01'),
    'irf': ('CMS IRF Compare — General Information + Provider Data; '
            'CCN count', '2026-02-13'),
    'ltch': ('CMS LTCH Compare — General Information + Provider Data; '
             'CCN count', '2026-02-13'),
    'hospice': ('CMS Provider Data Catalog — Hospice General Information '
                '(yc9t-dgbk); CCN count', '2026-05-23'),
    'hh': ('CMS Provider Data Catalog — Home Health Care Agencies '
           '(6jpm-sxkc); CCN count', '2026-05-23'),
    'dial': ('CMS Dialysis Facility Compare — Listing by Facility '
             '(DFC_FACILITY); CCN count', '2026-03-25'),
}

PRESENCE_BASIS = {'npi': 'SOURCED', 'unverified': 'SOURCED',
                  'web': 'PUBLIC-WEB', 'adjacent': 'PUBLIC-WEB'}


def _metro_tab(wb, lib, m, tab, presence, tier_labels):
    ws = wb.create_sheet(tab)
    sb = lib.SheetBuilder(
        ws, 7, col_widths=[36, 12, 13, 12, 46, 12, 48], tab_color=PURPLE)
    states = '+'.join(m.states)
    sb.title(f'{m.name} — metro facility structure ({m.region_label}, '
             f'{states})')
    sb.subtitle(
        'The question: how many transfer origins (hospitals) and post-acute '
        f'destinations (SNF / IRF / LTCH, plus hospice, home-health and '
        f'dialysis endpoints) physically exist in {m.name}, counted from CMS '
        'provider rolls city+state-filtered to the metro definition — and '
        'what share of the 20-metro footprint does this metro carry? Counts '
        'are blue SOURCED values; shares, node totals and occupancy are live '
        'formulas.', height=44)
    sb.blank()
    sb.banner('Panel A. Origin / destination facility counts (CMS provider '
              'rolls, city+state filtered)')
    sb.headers(['Facility class', 'Count', 'Share of O/D nodes',
                'Basis', 'Source dataset & locator', 'Snapshot', 'Note'])

    def cls_row(label, val, key, note='', share=True):
        r = sb.r + 1
        doc, snap = ROLLS[key]
        sb.row([(label, 'label'), (int(val), 'src', lib.FMT_INT),
                (f'=IF($B${R_NODES}>0,B{r}/$B${R_NODES},"n/a")', 'fml',
                 lib.FMT_PCT1) if share else
                ('— (endpoint class, not in the O/D node total)', 'note'),
                ('SOURCED', 'text'), (doc, 'text'), (snap, 'text'),
                (note, 'note')], wrap=True)
        return r

    cls_row('Hospitals (transfer origins)', m.n_hospitals, 'hosp',
            'Named roster in Panel E below')
    cls_row('Skilled nursing facilities (SNF)', m.n_snf, 'snf')
    cls_row('Inpatient rehabilitation facilities (IRF)', m.n_irf, 'irf')
    cls_row('Long-term care hospitals (LTCH)', m.n_ltch, 'ltch')
    cls_row('Hospice agencies', m.n_hospice, 'hospice', share=False)
    cls_row('Home health agencies', m.n_home_health, 'hh', share=False)
    cls_row('Dialysis facilities', m.n_dialysis, 'dial', share=False)
    sb.row([('Post-acute destinations (= SNF + IRF + LTCH)', 'label'),
            (f'=B{R_SNF}+B{R_IRF}+B{R_LTCH}', 'fml', lib.FMT_INT),
            (f'=IF($B${R_NODES}>0,B{R_PA}/$B${R_NODES},"n/a")', 'fml',
             lib.FMT_PCT1),
            ('DERIVED', 'text'),
            ('Live formula over the SOURCED counts above', 'text'), '',
            ('Matches module field n_postacute_destinations', 'note')],
           wrap=True)
    sb.row([('Total O/D nodes (= hospitals + post-acute)', 'label'),
            (f'=B{R_HOSP}+B{R_PA}', 'fml', lib.FMT_INT),
            (f'=IF($B${R_NODES}>0,B{R_NODES}/$B${R_NODES},"n/a")', 'fml',
             lib.FMT_PCT1),
            ('DERIVED', 'text'),
            ('Live formula over the SOURCED counts above', 'text'), '',
            ('Matches module field n_nodes; density tier (module '
             f'classification): {m.density_tier}', 'note')], wrap=True)
    assert sb.r == R_NODES, f'{tab}: row drift Panel A (r={sb.r})'

    sb.blank()
    sb.banner('Panel B. Capacity measures — certified beds, stations, '
              'HCRIS cost-report join')
    sb.headers(['Measure', 'Value', '', 'Basis',
                'Source dataset & locator', 'Snapshot', 'Note'])

    def cap_row(label, val, basis, doc, snap, note='', fmt=None, fml=False):
        sb.row([(label, 'label'),
                (val, 'fml' if fml else 'src', fmt or lib.FMT_INT), '',
                (basis, 'text'), (doc, 'text'), (snap, 'text'),
                (note, 'note')], wrap=True)

    cap_row('SNF certified beds', int(m.snf_beds), 'SOURCED',
            'NH_ProviderInfo, sum of certified_beds over metro CCNs',
            '2026-04-01')
    cap_row('LTCH beds', int(m.ltch_beds), 'SOURCED',
            'LTCH Compare, sum of total_beds over metro CCNs', '2026-02-13',
            '0 can mean the bed field is absent in the roll for a CCN')
    cap_row('Dialysis stations', int(m.dialysis_stations), 'SOURCED',
            'DFC_FACILITY, sum of dialysis_stations over metro CCNs',
            '2026-03-25')
    cap_row('Hospitals with an HCRIS cost report', m.n_hospitals_with_hcris,
            'SOURCED', 'CMS HCRIS (Form CMS-2552-10) FY2020-22 panel, '
            'latest filed report per CCN, joined to the hospital roll by CCN',
            'FY2020-22')
    cap_row('HCRIS staffed beds', int(m.hcris_beds), 'SOURCED',
            'HCRIS beds field, summed over joined metro CCNs', 'FY2020-22',
            'Conservative floor where hospitals lack an HCRIS row')
    cap_row('HCRIS inpatient days', int(m.hcris_patient_days), 'SOURCED',
            'HCRIS total_patient_days, summed over joined metro CCNs',
            'FY2020-22')
    cap_row('HCRIS bed-days available', int(m.hcris_bed_days_available),
            'SOURCED', 'HCRIS bed_days_available, summed over joined '
            'metro CCNs', 'FY2020-22')
    cap_row('Metro occupancy (= inpatient days / bed-days available)',
            f'=IF(B{R_BDA}>0,B{R_PDAYS}/B{R_BDA},"n/a")', 'DERIVED',
            'Live formula over the two HCRIS rows above', '',
            fmt=lib.FMT_PCT1, fml=True)
    cap_row('HCRIS coverage (= hospitals w/ report / hospitals)',
            f'=IF(B{R_HOSP}>0,B{R_HCRISN}/B{R_HOSP},"n/a")', 'DERIVED',
            'Live formula: Panel B row / Panel A hospital count', '',
            fmt=lib.FMT_PCT1, fml=True)
    assert sb.r == R_COV, f'{tab}: row drift Panel B (r={sb.r})'

    sb.blank()
    sb.banner('Panel C. Share of the 20-metro footprint (green = cross-tab '
              'links to Metro_Index live totals)')
    sb.headers(['Share measure', 'Share', '', 'Basis', 'Denominator',
                '', 'Note'])
    for label, num, den_col, den_label in [
            ('Hospitals — share of footprint hospitals', f'B{R_HOSP}', 'E',
             'Metro_Index footprint total, live SUM of 20 metro tabs'),
            ('SNF certified beds — share of footprint SNF beds',
             f'B{R_SNFBEDS}', 'G',
             'Metro_Index footprint total, live SUM of 20 metro tabs'),
            ('O/D nodes — share of footprint O/D nodes', f'B{R_NODES}', 'N',
             'Metro_Index footprint total, live SUM of 20 metro tabs')]:
        sb.row([(label, 'label'),
                (f"={num}/'Metro_Index'!${den_col}${IDX_TOT}", 'link',
                 lib.FMT_PCT1), '',
                ('DERIVED', 'text'), (den_label, 'text'), '',
                ('', 'note')], wrap=True)

    sb.blank()
    sb.banner('Panel D. MMT presence evidence (NPPES sweep, pulled '
              '2026-07-10)')
    sb.headers(['Measure', 'Tier', '', 'Basis', 'Source dataset & locator',
                'Pull date', 'Evidence (verbatim from module)'])
    tier, ev = presence
    sb.row([('MMT presence tier in this metro', 'label'),
            (tier_labels[tier], 'src'), '',
            (PRESENCE_BASIS[tier], 'text'),
            ('CMS NPPES registry sweep (organizational NPIs, taxonomy '
             '341600000X family) + company-web check; tiers: NPI-VERIFIED / '
             'COMPANY-WEB / ADJACENT NPI / NOT VERIFIED', 'text'),
            ('2026-07-10', 'text'), (ev, 'note')], wrap=True, height=30)

    sb.blank()
    sb.banner('Panel E. Named hospital roster (geocoded CMS hospital roll)')
    sb.headers(['#', 'Hospital (facility_name)', '', 'Basis',
                'Source dataset & locator', 'Snapshot', 'Note'])
    for i, name in enumerate(m.hospital_names, start=1):
        first = (i == 1)
        sb.row([(i, 'text', lib.FMT_INT), (name, 'src'), '',
                ('SOURCED', 'text'),
                (ROLLS['hosp'][0] if first else '', 'text'),
                ('2026-05-23' if first else '', 'text'), ''])

    sb.blank()
    for cav in (m.data_caveats or []):
        sb.note('Caveat (module, verbatim): ' + cav)
    cities = ', '.join(getattr(m, 'cities', ()) or ())
    sb.note(f'Metro definition (city+state filter): {cities} '
            f'({states}). Anchor systems, named operators, insource reads '
            'and moat notes (PUBLIC-WEB) are deliberately NOT on this tab; '
            'modeled discharge-base and SAM/SOM dollars are excluded '
            '(see Excluded_Not_Sourced).')
    sb.note('Extraction: rcm_mc.market_reports.ift_geo.metro_structure('
            f'{m.name!r}) over vendored rcm_mc/data rolls '
            '(hospital_coords.csv; snf/irf/ltch/hospice/home_health/'
            'dialysis _providers.csv, each carrying source + source_date) '
            '+ HCRIS parquet join by CCN.')

    lib.add_chart(
        ws, 'I6', f'{m.name} — facility mix (counts by class)',
        f"'{tab}'!$A${R_HOSP}:$A${R_DIAL}",
        [('Facilities', f"'{tab}'!$B${R_HOSP}:$B${R_DIAL}")],
        kind='bar', width=17, height=9, y_title='Facilities',
        y_fmt='#,##0')
    return ws


def build(wb, ctx):
    lib = ctx['lib']
    accessed = ctx['accessed']
    facts, sources, excluded, notes = [], [], [], []

    from rcm_mc.market_reports import ift_geo

    all_tabs = [s['name'] for s in SHEETS]
    sources += [
        {'key': 'pdc_rolls_metrofam', 'publisher': 'CMS',
         'document': 'Care Compare / Provider Data Catalog provider rolls: '
                     'Hospital General Information (geocoded via US Census '
                     'Geocoder, 2026-05-23); NH_ProviderInfo (2026-04-01); '
                     'IRF Compare (2026-02-13); LTCH Compare (2026-02-13); '
                     'Hospice General Information yc9t-dgbk (2026-05-23); '
                     'Home Health Agencies 6jpm-sxkc (2026-05-23); '
                     'DFC_FACILITY (2026-03-25)',
         'vintage': 'Snapshots Feb-May 2026 (per-roll dates above)',
         'locator': 'Vendored rcm_mc/data CSVs (hospital_coords.csv, '
                    'snf/irf/ltch/hospice/home_health/dialysis '
                    '_providers.csv), each row carrying source + '
                    'source_date; city+state filtered per metro by '
                    'ift_geo.metro_structure()',
         'supplies': 'Per-metro facility counts + SNF certified beds, LTCH '
                     'beds, dialysis stations, and the named hospital '
                     'rosters — the origin/destination node structure',
         'url': 'https://data.cms.gov/provider-data/',
         'tier': 'B', 'accessed': accessed, 'powers': all_tabs},
        {'key': 'hcris_metrofam', 'publisher': 'CMS',
         'document': 'Healthcare Cost Report Information System (HCRIS), '
                     'hospital cost reports (Form CMS-2552-10)',
         'vintage': 'FY2020-FY2022 panel; latest filed report per CCN '
                    '(status-rank dedup: audited > settled > submitted)',
         'locator': 'Vendored parquet via rcm_mc/data/hcris.py; fields '
                    'beds, total_patient_days, bed_days_available, joined '
                    'to the metro hospital rolls by CCN',
         'supplies': 'Per-metro staffed beds, inpatient days, bed-days '
                     'available — the occupancy inputs',
         'url': 'https://www.cms.gov/data-research/statistics-trends-'
                'reports/cost-reports',
         'tier': 'B', 'accessed': accessed, 'powers': all_tabs},
        {'key': 'nppes_mmt_presence', 'publisher': 'CMS',
         'document': 'National Plan and Provider Enumeration System (NPPES) '
                     'registry — MMT organizational-NPI sweep (taxonomy '
                     '341600000X family) + company-web station check',
         'vintage': 'Pulled 2026-07-10',
         'locator': 'ift_geo.MMT_PRESENCE_EVIDENCE (20 metros, tiers '
                    'NPI-VERIFIED / COMPANY-WEB / ADJACENT NPI / NOT '
                    'VERIFIED; NPI numbers quoted in the evidence column)',
         'supplies': 'Per-metro MMT presence tier + evidence (Panel D of '
                     'each metro tab; index column)',
         'url': 'https://npiregistry.cms.hhs.gov/',
         'tier': 'B', 'accessed': accessed, 'powers': all_tabs},
    ]

    # ── 20 metro tabs ───────────────────────────────────────────────────────
    metros = {}
    for name, tab in METRO_TABS:
        m = ift_geo.metro_structure(name)
        if m is None or not getattr(m, 'available', False):
            notes.append(f'metro_structure({name!r}) unavailable — tab '
                         'built empty-safe SKIPPED')
            continue
        metros[tab] = m
        _metro_tab(wb, lib, m, tab, ift_geo.MMT_PRESENCE_EVIDENCE[name],
                   ift_geo.PRESENCE_TIER_LABEL)

    # ── Metro_Index ─────────────────────────────────────────────────────────
    roll = ift_geo.footprint_rollup()
    ws = wb.create_sheet('Metro_Index')
    NC = 17
    sb = lib.SheetBuilder(
        ws, NC,
        col_widths=[22, 20, 18, 8, 10, 8, 11, 6, 7, 9, 11, 9, 12, 9, 15,
                    15, 11],
        tab_color=PURPLE)
    sb.title('Metro index — the 20-metro footprint at a glance')
    sb.subtitle(
        'The question: across the 20 target metros, where do transfer '
        'origins and destinations cluster — which metros carry the deepest '
        'origin/destination node structure, and what share of the footprint '
        'does each carry? Every count cell is a live cross-tab link (green) '
        'to its metro tab; the tab-name column hyperlinks to the tab; '
        'footprint totals are live SUMs; the national universe rows are the '
        'same vendored rolls, unfiltered.', height=44)
    sb.blank()
    sb.banner('Panel A. Metro directory (counts link to each Metro_ tab; '
              'shares are live formulas)')
    sb.headers(['Metro', 'Tab (hyperlink)', 'Region', 'States', 'Hospitals',
                'SNFs', 'SNF certified beds', 'IRF', 'LTCH', 'Hospice',
                'Home health', 'Dialysis', 'HCRIS staffed beds',
                'O/D nodes', 'Density tier (module classification)',
                'MMT presence (NPPES tier)', 'Share of footprint O/D nodes'])
    assert sb.r + 1 == IDX_FIRST, 'Metro_Index header drift'

    from openpyxl.worksheet.hyperlink import Hyperlink
    cell_map = {'E': R_HOSP, 'F': R_SNF, 'G': R_SNFBEDS, 'H': R_IRF,
                'I': R_LTCH, 'J': R_HOSPICE, 'K': R_HH, 'L': R_DIAL,
                'M': R_BEDS, 'N': R_NODES}
    for name, tab in METRO_TABS:
        m = metros[tab]
        r = sb.r + 1
        tier, _ = ift_geo.MMT_PRESENCE_EVIDENCE[name]
        cells = [(m.name, 'text'), (tab, 'link'),
                 (m.region_label, 'text'), ('+'.join(m.states), 'text')]
        for col, src_row in cell_map.items():
            cells.append((f"='{tab}'!$B${src_row}", 'link', lib.FMT_INT))
        cells += [(m.density_tier, 'text'),
                  (ift_geo.PRESENCE_TIER_LABEL[tier], 'src'),
                  (f'=IF($N${IDX_TOT}>0,N{r}/$N${IDX_TOT},"n/a")', 'fml',
                   lib.FMT_PCT1)]
        sb.row(cells)
        c = ws.cell(row=r, column=2)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{tab}'!A1",
                                display=tab)
    assert sb.r == IDX_LAST, f'Metro_Index directory drift (r={sb.r})'

    sum_cells = [('Footprint total (20 metros) — live SUM', 'label'),
                 ('', 'text'), '', '']
    for col in cell_map:
        sum_cells.append((f'=SUM({col}{IDX_FIRST}:{col}{IDX_LAST})', 'fml',
                          lib.FMT_INT))
    sum_cells += ['', '', (f'=SUM(Q{IDX_FIRST}:Q{IDX_LAST})', 'fml',
                           lib.FMT_PCT1)]
    sb.row(sum_cells)
    assert sb.r == IDX_TOT, 'Metro_Index total-row drift'

    sb.row([('National universe (same vendored rolls, unfiltered US)',
             'label'), ('', 'text'), '', '',
            (roll.n_hospitals_national, 'src', lib.FMT_INT),
            (roll.n_snf_national, 'src', lib.FMT_INT),
            (roll.snf_beds_national, 'src', lib.FMT_INT),
            '', '', '', '', '', '', '', '',
            ('SOURCED (roll totals)', 'note'), ''])
    assert sb.r == IDX_NAT, 'Metro_Index national-row drift'
    sb.row([('20-metro share of the national universe', 'label'),
            ('', 'text'), '', '',
            (f'=E{IDX_TOT}/E{IDX_NAT}', 'fml', lib.FMT_PCT1),
            (f'=F{IDX_TOT}/F{IDX_NAT}', 'fml', lib.FMT_PCT1),
            (f'=G{IDX_TOT}/G{IDX_NAT}', 'fml', lib.FMT_PCT1),
            '', '', '', '', '', '', '', '',
            ('DERIVED', 'note'), ''])
    assert sb.r == IDX_SHARE, 'Metro_Index share-row drift'

    sb.blank()
    sb.note('Basis: count columns are DERIVED cross-tab links to the blue '
            'SOURCED counts on each Metro_ tab (CMS provider rolls, '
            'snapshots Feb-May 2026; HCRIS FY2020-22 panel). MMT presence '
            'column: NPPES sweep 2026-07-10 (tiers COMPANY-WEB / ADJACENT '
            'are PUBLIC-WEB evidence). Density tier is a module '
            'classification, not a measured value.')
    sb.note('Scope note: the 11-footprint-STATE shares of the national '
            'universe (24.9% of hospitals / 22.2% of SNF beds) live on '
            'Metro_Structure_20; the shares in row 28 here are the '
            '20-METRO share of the same national rolls — a different, '
            'narrower numerator.')
    sb.note('Extraction: rcm_mc.market_reports.ift_geo.all_metros() + '
            'footprint_rollup() (national roll totals) over vendored '
            'rcm_mc/data provider rolls.')

    lib.add_chart(
        ws, 'A32', 'O/D nodes by metro (hospitals + SNF + IRF + LTCH)',
        f"'Metro_Index'!$A${IDX_FIRST}:$A${IDX_LAST}",
        [('O/D nodes', f"'Metro_Index'!$N${IDX_FIRST}:$N${IDX_LAST}")],
        kind='bar', width=30, height=10, y_title='O/D nodes', y_fmt='#,##0')

    # python-side recompute audit for the identities the formulas assert
    for tab, m in metros.items():
        if m.n_postacute_destinations != m.n_snf + m.n_irf + m.n_ltch:
            notes.append(f'post-acute identity BROKE for {m.name}')
        if m.n_nodes != m.n_hospitals + m.n_postacute_destinations:
            notes.append(f'node identity BROKE for {m.name}')
        if len(m.hospital_names) != m.n_hospitals:
            notes.append(f'roster/count mismatch for {m.name}')
    tot_nodes = sum(m.n_nodes for m in metros.values())
    if sum(m.n_hospitals for m in metros.values()) != roll.n_hospitals:
        notes.append('hospital sum != footprint_rollup()')
    if sum(m.snf_beds for m in metros.values()) != roll.snf_beds:
        notes.append('SNF-bed sum != footprint_rollup()')

    # ── facts (headline figures across the family) ──────────────────────────
    SRC = ['pdc_rolls_metrofam']
    HCR = ['pdc_rolls_metrofam', 'hcris_metrofam']

    def fact(metric, ref, unit, basis, keys, locator, tab, cross=''):
        return {'metric': metric, 'year': 2026, 'value_ref': ref,
                'unit': unit, 'basis': basis, 'tier': 'B',
                'source_keys': keys, 'locator': locator, 'lives_on': tab,
                'cross_check': cross}

    facts += [
        fact('Omaha hospitals (transfer origins)', f'Metro_Omaha!B{R_HOSP}',
             'facilities', 'SOURCED', SRC,
             'Hospital General Information geocoded roll (2026-05-23), '
             'city+state filtered to Omaha metro', 'Metro_Omaha',
             'Metro_Structure_20 row Omaha; Nebraska Medicine/UNMC absent '
             'from the geocoded roll (undercount caveat printed)'),
        fact('Omaha O/D nodes (hospitals + SNF + IRF + LTCH)',
             f'Metro_Omaha!B{R_NODES}', 'nodes', 'DERIVED', SRC,
             'Live formula over Panel A SOURCED counts',
             'Metro_Omaha', 'module n_nodes = 48'),
        fact('Omaha SNF certified beds', f'Metro_Omaha!B{R_SNFBEDS}',
             'beds', 'SOURCED', SRC,
             'NH_ProviderInfo (2026-04-01), sum of certified_beds over '
             'Omaha-metro CCNs', 'Metro_Omaha', 'module snf_beds = 3,584'),
        fact('Omaha metro occupancy (HCRIS inpatient days / bed-days '
             'available)', f'Metro_Omaha!B{R_OCC}', 'share', 'DERIVED',
             HCR, 'HCRIS FY2020-22 panel, latest per CCN, 10 of 11 '
             'hospitals joined', 'Metro_Omaha',
             'python recompute 328,819/515,015 = 63.8%'),
        fact('Kansas City (bi-state) O/D nodes — densest footprint metro',
             f'Metro_KansasCity!B{R_NODES}', 'nodes', 'DERIVED', SRC,
             'Live formula over Panel A SOURCED counts',
             'Metro_KansasCity', 'module n_nodes = 105, density tier '
             'very-dense'),
        fact('Cincinnati O/D nodes', f'Metro_Cincinnati!B{R_NODES}',
             'nodes', 'DERIVED', SRC,
             'Live formula over Panel A SOURCED counts',
             'Metro_Cincinnati', 'module n_nodes = 94, density tier '
             'very-dense'),
        fact('North Platte O/D nodes — thinnest footprint metro',
             f'Metro_NorthPlatte!B{R_NODES}', 'nodes', 'DERIVED', SRC,
             'Live formula over Panel A SOURCED counts',
             'Metro_NorthPlatte', 'module n_nodes = 4, tier thin/long-leg'),
        fact('Footprint hospitals (20-metro live SUM)',
             f'Metro_Index!E{IDX_TOT}', 'facilities', 'DERIVED', SRC,
             'SUM over the 20 Metro_ tab hospital counts', 'Metro_Index',
             f'footprint_rollup() n_hospitals = {roll.n_hospitals}'),
        fact('Footprint SNF certified beds (20-metro live SUM)',
             f'Metro_Index!G{IDX_TOT}', 'beds', 'DERIVED', SRC,
             'SUM over the 20 Metro_ tab SNF-bed rows', 'Metro_Index',
             f'footprint_rollup() snf_beds = {roll.snf_beds:,}'),
        fact('Footprint O/D nodes (20-metro live SUM)',
             f'Metro_Index!N{IDX_TOT}', 'nodes', 'DERIVED', SRC,
             'SUM over the 20 Metro_ tab node totals', 'Metro_Index',
             f'python recompute = {tot_nodes}'),
        fact('Footprint HCRIS staffed beds (20-metro live SUM)',
             f'Metro_Index!M{IDX_TOT}', 'beds', 'DERIVED', HCR,
             'SUM over the 20 Metro_ tab HCRIS staffed-bed rows',
             'Metro_Index',
             f'footprint_rollup() hcris_beds = {int(roll.hcris_beds):,}; '
             'conservative floor (14 of 166 hospitals lack an HCRIS row)'),
        fact('20-metro share of national hospital universe',
             f'Metro_Index!E{IDX_SHARE}', 'share', 'DERIVED', SRC,
             'Footprint SUM / national roll total (4,630 hospitals)',
             'Metro_Index',
             'Distinct from the 24.9% 11-STATE share on Metro_Structure_20'),
    ]

    # ── excluded ────────────────────────────────────────────────────────────
    excluded += [
        {'figure': 'Per-metro discharge base (HCRIS beds x 53.3 '
                   'discharges/bed/yr)',
         'value': 'e.g. Omaha 75,206',
         'source_label': 'ift_geo.metro_structure().discharge_base — '
                         'ILLUSTRATIVE-with-basis (occ 0.657 x 365 / ALOS '
                         '4.5)',
         'why_excluded': 'The 53.3 multiplier is a modeled composite, not a '
                         'published per-metro measurement',
         'what_would_make_citable': 'Metro-level discharge counts from HCUP '
                                    'SID/SEDD or CMS MedPAR aggregated to '
                                    'the metro definition'},
        {'figure': 'Per-metro SAM/SOM dollars',
         'value': 'all dollarized metro opportunity figures',
         'source_label': 'ift_analytics.sam_formula / '
                         'ift_geo.footprint_sam_building_blocks() '
                         'dollarization layer',
         'why_excluded': 'Modeled serviceable-share and $/leg levers; no '
                         'published per-metro IFT market dollars exist',
         'what_would_make_citable': 'Claims-measured metro transport volumes '
                                    'x published fee schedules (CMS SAF/PSPS '
                                    'metro aggregation)'},
        {'figure': 'Anchor systems, named operators, insource reads, moat '
                   'notes (per metro)',
         'value': 'qualitative rosters/reads incl. embedded stats (e.g. '
                  '"~77% of county IFT moved to AMR 2022", "~34,000 '
                  'interfacility requests/2024")',
         'source_label': 'ift_geo.MARKETS MetroDef fields — PUBLIC-WEB, no '
                         'per-item citation attached in-module',
         'why_excluded': 'No formal per-item citation; embedded figures '
                         'unverifiable from the module',
         'what_would_make_citable': 'Per-claim citations (news/company URLs '
                                    'with dates) attached to each read'},
        {'figure': 'Metro density-moat prose (density_note)',
         'value': 'UHU/deadhead/load-chaining narrative per metro',
         'source_label': 'ift_geo.metro_structure().density_note — module '
                         'FRAMEWORK narrative',
         'why_excluded': 'Analytic framing, not a sourced measurement; only '
                         'the density tier label is carried (marked as '
                         'module classification)',
         'what_would_make_citable': 'Published UHU/deadhead studies tied to '
                                    'node density'},
        {'figure': 'MMT web-listed station addresses as facility counts',
         'value': '11 company-web stations',
         'source_label': 'ift_company.MMT_WEB_STATIONS — PUBLIC-WEB '
                         '(subparts often bill under a parent NPI)',
         'why_excluded': 'Company self-report; kept only as the COMPANY-WEB '
                         'presence tier label in Panel D, never as counts',
         'what_would_make_citable': 'NPPES subpart enumeration or state EMS '
                                    'licensure rosters per station'},
    ]

    row_counts = {s['name']: wb[s['name']].max_row for s in SHEETS
                  if s['name'] in wb.sheetnames}
    return {
        'facts': facts, 'sources': sources, 'excluded': excluded,
        'meta': {
            'notes': ('One tab per metro (20) + Metro_Index. All facility '
                      'counts SOURCED blue from vendored CMS rolls via '
                      'ift_geo.metro_structure(); HCRIS FY2020-22 join '
                      'live (pandas present). Node totals, shares, '
                      'occupancy, footprint SUMs are live formulas; index '
                      'counts are cross-tab links. python recompute audit: '
                      + ('; '.join(notes) if notes else 'all identities '
                         'hold (post-acute, nodes, roster length, '
                         'footprint sums vs footprint_rollup())')),
            'row_counts': row_counts},
    }
