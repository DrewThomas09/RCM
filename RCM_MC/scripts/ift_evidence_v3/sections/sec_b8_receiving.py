"""B.8 / X-B.2: the receiving-center registry.

The formal trauma / stroke / STEMI designation lists are behind JS portals
and per-state statutory pages (the research pass was cut short by a usage
limit before the rosters landed), so the designated-capacity rows ship as
bordered PENDING cells naming the exact public dataset that fills each - the
house style for evidence that is real but not yet retrieved.

What IS carried, from data already in hand: the emergency-capable acute
hospital universe by state (CMS Hospital General Information), the measured
receiving-capacity FLOOR into which every high-acuity interfacility transfer
must land, and its matched ratio against the 65+ demand base.
"""

SHEETS = [{'name': 'Receiving_Center_Registry',
           'question': 'Where is the receiving-side capacity that high-acuity '
                       'interfacility transfers must land in?'}]

FOOTPRINT = ['NE', 'IA', 'KS', 'MO', 'OH', 'WI', 'VA', 'MN', 'IN', 'KY']

PENDING_SOURCES = [
    ('ACS-verified trauma centers (Levels I-IV)',
     'American College of Surgeons verified trauma center list '
     '(facs.org/quality-programs/trauma/tqp/center-programs/vrc)',
     'JS-gated search UI; state trauma-system designation pages are the '
     'authoritative fallback per state'),
    ('Joint Commission stroke certifications (Comprehensive / '
     'Thrombectomy-Capable / Primary)',
     'TJC Quality Check certification list '
     '(qualitycheck.org data downloads)',
     'JS-heavy portal; DNV and state-only certifications are additional '
     'certifiers not in the TJC file'),
    ('Chest Pain / STEMI-receiving certifications',
     'TJC / ACC accreditation lists + state STEMI-system designations',
     'Multiple certifiers; several footprint states designate by statute '
     '(e.g. MO Time-Critical Diagnosis system)'),
]


def build(wb, ctx):
    lib = ctx['lib']
    cache = ctx['cache']
    facts, sources, findings = [], [], []

    hosp = lib.load_cache(cache, 'pdc2_hospitals')

    def is_acute(r):
        t = (r.get('hospital_type') or '').lower()
        return 'acute' in t and 'critical' not in t

    def is_cah(r):
        return 'critical access' in (r.get('hospital_type') or '').lower()

    def emerg(r):
        return (r.get('emergency_services') or '').strip().lower() == 'yes'

    by_state = {}
    for r in hosp:
        st = (r.get('state') or '').strip().upper()
        d = by_state.setdefault(st, {'acute': 0, 'acute_ed': 0, 'cah': 0,
                                     'cah_ed': 0})
        if is_acute(r):
            d['acute'] += 1
            if emerg(r):
                d['acute_ed'] += 1
        elif is_cah(r):
            d['cah'] += 1
            if emerg(r):
                d['cah_ed'] += 1

    sources.append(
        {'key': 'pdc_hosp_emerg', 'publisher': 'CMS',
         'document': 'Hospital General Information (Provider Data Catalog): '
                     'hospital type, ownership and emergency-services flag',
         'vintage': 'Current PDC release, pulled 2026 (Pull_Manifest)',
         'locator': 'Rows by state: acute-care and critical-access hospitals '
                    'with emergency_services = Yes',
         'supplies': 'The emergency-capable receiving-capacity floor by '
                     'state - the universe a high-acuity transfer can land in',
         'url': 'https://data.cms.gov/provider-data/dataset/xubh-q36u',
         'tier': 'A', 'accessed': ctx['accessed'],
         'powers': ['Receiving_Center_Registry']})

    ws = wb.create_sheet('Receiving_Center_Registry')
    sb = lib.SheetBuilder(ws, 7,
                          col_widths=[22, 16, 16, 16, 16, 16, 40],
                          tab_color='FF4C7C2C')
    sb.title('Receiving-center registry: the capacity that high-acuity '
             'transfers land in')
    sb.subtitle('The question: where is the receiving-side capacity - trauma, '
                'stroke and STEMI centers - that interfacility transfers must '
                'reach? The formal ACS/TJC/state designation rosters are '
                'behind JS portals and are carried as bordered PENDING rows '
                'naming the exact public dataset that fills each. What is '
                'MEASURED here from CMS Hospital General Information: the '
                'emergency-capable acute-hospital universe by state, the '
                'receiving-capacity floor. Join to Hub_Spoke_Map by CCN and '
                'state to place designated capacity on the hub map.')
    sb.note('DATA QUALITY: emergency-services capability is a FLOOR proxy for '
            'receiving capacity, not a trauma/stroke/STEMI designation; a '
            'hospital can run an ED without being a designated receiving '
            'center, and a designated center is a subset. Critical-access '
            'hospitals are counted separately (they are typically sending, '
            'not receiving, sites). The designation counts stay PENDING.')
    sb.blank()

    sb.banner('Panel A. Emergency-capable hospital universe by state '
              '(receiving-capacity floor)')
    sb.headers(['State', 'Acute hospitals', 'Acute with ED',
                'Critical-access', 'CAH with ED', 'Footprint', ''])
    a0 = sb.r + 1
    order = sorted(by_state.items(),
                   key=lambda kv: -kv[1]['acute_ed'])
    fp_rows = []
    for st, d in order:
        if not st or len(st) != 2:
            continue
        rn = sb.r + 1
        if st in FOOTPRINT:
            fp_rows.append(rn)
        sb.row([(st, 'src'), (d['acute'], 'src', lib.FMT_INT),
                (d['acute_ed'], 'src', lib.FMT_INT),
                (d['cah'], 'src', lib.FMT_INT),
                (d['cah_ed'], 'src', lib.FMT_INT),
                ('YES' if st in FOOTPRINT else '-', 'fml'), None])
    sb.blank()

    sb.banner('Panel B. Receiving-capacity floor vs the 65+ demand base '
              '(matched ratio, confound in-row)')
    sb.headers(['State', 'Acute with ED', '65+ population (link)',
                '65+ per acute-ED hospital', '', 'Footprint', 'Confound'])
    age = wb['State_Age_65plus'] if 'State_Age_65plus' in wb.sheetnames else None
    STNAME = {'NE': 'Nebraska', 'IA': 'Iowa', 'KS': 'Kansas', 'MO': 'Missouri',
              'OH': 'Ohio', 'WI': 'Wisconsin', 'VA': 'Virginia',
              'MN': 'Minnesota', 'IN': 'Indiana', 'KY': 'Kentucky'}

    def age_row(state_name):
        if age is None:
            return None
        for row in age.iter_rows(min_col=1, max_col=1):
            if (row[0].value or '') == state_name:
                return row[0].row
        return None

    b0 = sb.r + 1
    for st in FOOTPRINT:
        d = by_state.get(st, {'acute_ed': 0})
        ar = age_row(STNAME[st])
        rn = sb.r + 1
        age_ref = (f"='State_Age_65plus'!E{ar}" if ar else 'PENDING')
        sb.row([(st, 'src'), (d['acute_ed'], 'src', lib.FMT_INT),
                (age_ref, 'link', lib.FMT_INT) if ar else ('PENDING', 'note'),
                (f'=IF(B{rn}=0,"n/a",C{rn}/B{rn})', 'fml', lib.FMT_INT)
                if ar else None,
                None, ('YES', 'fml'),
                ('ED capability is a floor, not designated trauma/stroke '
                 'capacity', 'note') if st == FOOTPRINT[0] else None])
    sb.blank()

    sb.banner('Panel C. Formal designation registry (PENDING: dataset named '
              'per row)')
    sb.headers(['Designation class', 'Public dataset', 'Blocker / route',
                'Status', '', '', ''])
    from openpyxl.styles import Border, Side
    pend = Border(bottom=Side(style='dotted', color='FF8C1D40'),
                  top=Side(style='dotted', color='FF8C1D40'),
                  left=Side(style='dotted', color='FF8C1D40'),
                  right=Side(style='dotted', color='FF8C1D40'))
    for cls, ds, blk in PENDING_SOURCES:
        rn = sb.r + 1
        sb.row([(cls, 'label'), (ds, 'text'), (blk, 'text'),
                ('PENDING', 'note'), None, None, None])
        for col in range(1, 8):
            ws.cell(row=rn, column=col).border = pend
    sb.blank()
    sb.banner('Read panel')
    fp_ed = sum(by_state.get(st, {'acute_ed': 0})['acute_ed']
                for st in FOOTPRINT)
    us_ed = sum(d['acute_ed'] for st, d in by_state.items()
                if len(st) == 2)
    sb.prose('The receiving-capacity floor is countable now even though the '
             f'formal designations are PENDING: {fp_ed:,} emergency-capable '
             'acute hospitals sit in the ten footprint states, a subset of '
             f'the {us_ed:,} nationally, and the matched ratio against the '
             '65+ base (Panel B) shows where receiving capacity is thin per '
             'senior resident. Every high-acuity interfacility transfer must '
             'land in this universe; the designated trauma/stroke/STEMI '
             'centers are a subset that concentrates the highest-acuity '
             'volume, and their rosters are named for the next pass. Join to '
             'Hub_Spoke_Map to place designated capacity on the corridor '
             'hubs.')

    lib.add_chart(ws, 'I6', 'Emergency-capable acute hospitals, footprint '
                  'states', f"'Receiving_Center_Registry'!$A${b0}:$A${b0 + 9}",
                  [('Acute with ED',
                    f"'Receiving_Center_Registry'!$B${b0}:$B${b0 + 9}")],
                  kind='bar')

    facts += [
        {'metric': 'Emergency-capable acute hospitals, footprint states '
                   '(receiving-capacity floor)', 'year': 2026, 'value': fp_ed,
         'unit': 'hospitals', 'basis': 'GOV', 'tier': 'A',
         'source_keys': ['pdc_hosp_emerg'],
         'locator': 'Hospital General Information, acute type with '
                    'emergency_services = Yes, footprint states',
         'lives_on': 'Receiving_Center_Registry',
         'cross_check': 'A floor: designated trauma/stroke/STEMI centers are '
                        'a subset (PENDING Panel C)'},
        {'metric': 'Emergency-capable acute hospitals, national', 'year': 2026,
         'value': us_ed, 'unit': 'hospitals', 'basis': 'GOV', 'tier': 'A',
         'source_keys': ['pdc_hosp_emerg'],
         'locator': 'Hospital General Information, acute type with '
                    'emergency_services = Yes',
         'lives_on': 'Receiving_Center_Registry',
         'cross_check': 'Reconciles to the Hosp_Registry universe (same PDC '
                        'source family)'},
    ]
    findings.append({
        'id_hint': 92,
        'finding': 'The receiving-capacity floor is countable even before the '
                   f'formal designations land: {fp_ed:,} emergency-capable '
                   'acute hospitals in the footprint states form the universe '
                   'every high-acuity interfacility transfer must reach, and '
                   'the matched ratio against the 65+ base marks where that '
                   'capacity is thin per senior.',
        'numbers': f"='Receiving_Center_Registry'!B{a0}",
        'sources': 'pdc_hosp_emerg',
        'confidence': 'High on the ED-capability floor; the designated-center '
                      'subset is PENDING',
        'guardrail': 'Emergency-services capability is a floor proxy, not a '
                     'trauma/stroke/STEMI designation; the designated rosters '
                     '(ACS, TJC, state) are named PENDING in Panel C and are '
                     'the correct source for high-acuity receiving capacity.'})

    return {'facts': facts, 'sources': sources, 'excluded': [],
            'findings': findings, 'meta': {'states': len(by_state)}}
