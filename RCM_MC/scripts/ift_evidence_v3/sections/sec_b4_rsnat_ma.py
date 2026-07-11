"""B.4 + B.5: the RSNAT prior-authorization model, end to end from published
CMS documents (RSNAT_Series), and the GADCS MA-book calibrator with its
printed supersession clause (MA_Book_Calibrator).

Every URL below was fetched and read on 11 Jul 2026; every figure was read
from the primary document (CMS RSNAT model page, the updated CMS RSNAT FAQ,
the three Mathematica evaluation reports, GAO-18-341 full text, MLN6805343,
GADCS Year 1-4 appendix Table 2.30).
"""

SHEETS = [
    {'name': 'RSNAT_Series',
     'question': 'What did prior authorization do to scheduled non-emergent '
                 'ambulance transport, and when did each state enter the '
                 'regime?'},
    {'name': 'MA_Book_Calibrator',
     'question': 'What is the best published multiplier from a supplier\'s '
                 'Medicare FFS book to its Medicare Advantage book?'},
]

CMS_RSNAT_URL = ('https://www.cms.gov/data-research/monitoring-programs/'
                 'medicare-fee-service-compliance-programs/prior-authorization-'
                 'and-pre-claim-review-initiatives/prior-authorization-'
                 'repetitive-scheduled-non-emergent-ambulance-transport-rsnat')
FAQ_URL = ('https://www.cms.gov/research-statistics-data-and-systems/'
           'monitoring-programs/medicare-ffs-compliance-programs/'
           'prior-authorization-initiatives/downloads/'
           'ambulancepriorauthorization_externalfaq_121517.pdf')
EVAL1_URL = ('https://www.cms.gov/priorities/innovation/files/reports/'
             'rsnat-firstintevalrpt.pdf')
EVAL2_URL = ('https://www.cms.gov/priorities/innovation/data-and-reports/'
             '2020/rsnat-secondintevalrpt')
EVALF_URL = ('https://www.cms.gov/priorities/innovation/data-and-reports/'
             '2021/rsnat-finalevalrpt')
GAO_URL = 'https://www.gao.gov/products/gao-18-341'
MLN_URL = ('https://www.cms.gov/files/document/mln6805343-repetitive-'
           'scheduled-non-emergent-ambulance-transport-prior-authorization-'
           'model.pdf')
GADCS_URL = ('https://www.cms.gov/files/document/medicare-ground-ambulance-'
             'data-collection-system-gadcs-report-appendix-year-1-year-4-'
             'cohort-analysis.pdf')

FOOTPRINT = ['NE', 'IA', 'KS', 'MO', 'OH', 'WI', 'VA', 'MN', 'IN', 'KY']

# National expansion phases: CMS RSNAT model page + updated CMS RSNAT FAQ,
# both fetched 11 Jul 2026. Dates are transports-on-or-after dates.
PHASES = [
    ('1 Dec 2021', 'AR, CO, LA, MS, NM, OK, TX (MAC Jurisdiction H, Novitas)'),
    ('1 Feb 2022', 'AL, CA, GA, HI, NV, TN + AS, GU, MP'),
    ('1 Apr 2022', 'FL, IL, IA, KS, MN, MO, NE, WI + PR, USVI'),
    ('1 Jun 2022', 'CT, IN, ME, MA, MI, NH, NY, RI, VT'),
    ('1 Aug 2022', 'AK, AZ, ID, KY, MT, ND, OH, OR, SD, UT, WA, WY + '
                   'Railroad Retirement Board beneficiaries nationwide'),
]

# GADCS Table 2.30 (pp.64-65): average payer revenue per NPI, Traditional
# FFS Medicare vs Medicare Advantage, with every published cut. Read from
# the appendix text on 11 Jul 2026 (gadcs_text.txt lines 5407-5634).
T230_ALL = ('All NPIs', 1193046, 6957, 1238482, 6112)
T230_CUTS = [
    ('Provider vs supplier status', [
        ('Suppliers', 1096545, 6561, 1214893, 5775),
        ('Providers (hospital-based)', 2793852, 396, 1642902, 337)]),
    ('Medicare transport volume', [
        ('Low volume', 260296, 2686, 203536, 2213),
        ('Medium volume', 937824, 2106, 1036829, 1922),
        ('High volume', 2239114, 1290, 2393579, 1186),
        ('Very high volume', 3129109, 875, 2892195, 791)]),
    ('Ownership category', [
        ('Non-profit', 872538, 1857, 557051, 1639),
        ('For-profit / unknown', 2352190, 1239, 2548772, 1060),
        ('Government', 975045, 3860, 1158831, 3412)]),
    ('Service-area population density', [
        ('Urban', 1581403, 3865, 1675339, 3487),
        ('Rural', 900215, 1943, 900773, 1703),
        ('Super rural', 381594, 1149, 209096, 921)]),
    ('Public safety organization', [
        ('Public safety: yes', 508365, 3317, 542592, 2967),
        ('Public safety: no', 1817054, 3640, 1895215, 3144)]),
]


def _find_label(ws, needle, max_col=4):
    n = needle.lower()
    for row in ws.iter_rows(max_col=max_col):
        for c in row:
            if isinstance(c.value, str) and n in c.value.lower():
                return c.row, c.column
    return None


def _denial_pivot_cells(wb):
    """Locate the A0428 denial-rate pivot cells for 2014/2022/2024 on
    PSPS_Denial_Series Panel B by scanning, never by hardcoded address."""
    if 'PSPS_Denial_Series' not in wb.sheetnames:
        return {}
    ws = wb['PSPS_Denial_Series']
    hdr = None
    for row in ws.iter_rows(min_row=100, max_row=ws.max_row):
        vals = {c.column: c.value for c in row}
        if any(v == 'A0428' for v in vals.values()) and vals.get(1) == 'Year':
            hdr = row[0].row
            col = next(c for c, v in vals.items() if v == 'A0428')
            break
    if hdr is None:
        return {}
    from openpyxl.utils import get_column_letter
    out = {}
    for row in ws.iter_rows(min_row=hdr + 1, max_row=min(hdr + 20, ws.max_row),
                            max_col=1):
        y = row[0].value
        if y in (2014, 2022, 2024):
            out[int(y)] = f'{get_column_letter(col)}{row[0].row}'
    return out


def _dark_share_cells(wb):
    """Locate Imbalance_Ledger Panel C dark-share cells for 2013/2024/2025."""
    if 'Imbalance_Ledger' not in wb.sheetnames:
        return {}
    ws = wb['Imbalance_Ledger']
    anchor = _find_label(ws, 'Panel C')
    if not anchor:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=anchor[0], max_row=ws.max_row, max_col=1):
        y = row[0].value
        if y in (2013, 2024, 2025):
            out[int(y)] = f'B{row[0].row}'
    return out


def build(wb, ctx):
    lib = ctx['lib']
    accessed = ctx['accessed']
    facts, sources, excluded, findings = [], [], [], []

    sources += [
        {'key': 'cms_rsnat_model', 'publisher': 'CMS',
         'document': 'Prior Authorization of Repetitive, Scheduled '
                     'Non-Emergent Ambulance Transport (RSNAT) - model page',
         'vintage': 'Current page, fetched 11 Jul 2026',
         'locator': 'Model start and expansion dates; national expansion '
                    'phase schedule; links to all three evaluation reports',
         'supplies': 'The authoritative RSNAT timeline: 2014 start, 2016 '
                     'expansion, 2018 actuary certification, 2021-2022 '
                     'phased national rollout',
         'url': CMS_RSNAT_URL, 'tier': 'A', 'accessed': accessed,
         'powers': ['RSNAT_Series']},
        {'key': 'cms_rsnat_faq', 'publisher': 'CMS',
         'document': 'RSNAT Prior Authorization Model FAQs (updated; '
                     'original file 12/15/2017, content current through the '
                     'national expansion)',
         'vintage': 'Fetched and read 11 Jul 2026',
         'locator': 'Q2 (phase dates and MAC jurisdictions), Q3-Q5 '
                    '(statutory authority: SSA 1115A; SSA 1834(l)(16) as '
                    'added by MACRA sec. 515(b)), Q4 (original model ended '
                    '1 Dec 2020, transition 2 Dec 2020), Q6 (evaluation '
                    'summary: 72% use / 76% expenditure reduction, about '
                    '$750M; 2.4% and $1B total FFS over five years), Q10 '
                    '(MedPAC June 2013 basis for choosing NJ/PA/SC)',
         'supplies': 'Statutory chain and CMS\'s own summary of the '
                     'evaluation findings',
         'url': FAQ_URL, 'tier': 'A', 'accessed': accessed,
         'powers': ['RSNAT_Series']},
        {'key': 'rsnat_eval_first', 'publisher': 'CMS CMMI / Mathematica',
         'document': 'First Interim Evaluation Report of the Medicare Prior '
                     'Authorization Model for RSNAT (February 2018)',
         'vintage': 'Model years 2015-2016, report Feb 2018',
         'locator': 'Ch. IV: ESRD beneficiaries in Year 1 states: RSNAT use '
                    'probability -4.06pp from 5.09% baseline (-80%); trips '
                    '-87%; RSNAT payments -90% ($432 per beneficiary per '
                    'quarter), about $171M savings; suppliers per 100,000 '
                    'beneficiaries -15%',
         'supplies': 'The year-1 measured effect of the prior-authorization '
                     'screen', 'url': EVAL1_URL, 'tier': 'A',
         'accessed': accessed, 'powers': ['RSNAT_Series']},
        {'key': 'rsnat_eval_second', 'publisher': 'CMS CMMI / Mathematica',
         'document': 'Second Interim Evaluation Report (September 2020)',
         'vintage': 'Through 2018-2019, report Sep 2020',
         'locator': 'Exec summary and Ch. III: RSNAT use and expenditures '
                    'down over 60%; about $550M cumulative RSNAT savings '
                    '(average $138M per year); total Medicare expenditures '
                    '-$316 per beneficiary per quarter (-2%, about $650M); '
                    'Objective 3: RSNAT suppliers per 100,000 beneficiaries '
                    '-50% on implementation',
         'supplies': 'The mid-model savings and the supplier-exit measurement',
         'url': EVAL2_URL, 'tier': 'A', 'accessed': accessed,
         'powers': ['RSNAT_Series']},
        {'key': 'rsnat_eval_final', 'publisher': 'CMS CMMI / Mathematica',
         'document': 'Evaluation of the Medicare Prior Authorization Model '
                     'for RSNAT: Final Report (May 2021)',
         'vintage': 'Model period through 2019, report May 2021',
         'locator': 'Table ES.1 and Ch. III: RSNAT use -72%, RSNAT '
                    'expenditures -$746M across the nine model states; '
                    'total Medicare FFS -2.4% (-$381 per beneficiary per '
                    'quarter); Year 1 -74% vs Year 2 -64%; Table II.2 (p.9): '
                    'baseline RSNAT utilization 10.1% (Year 1 cohort); '
                    'Ch. VI Figure VI.8: 36%/45% of PARs approved on initial '
                    'submission (Year 1/Year 2 supplier survey), +31% after '
                    'resubmission; Ch. VII (pp.67-68): pre-model denial '
                    'baseline about 3.5% of non-emergency ambulance claims, '
                    'denial spike reverted to baseline within 8 quarters; no '
                    'adverse quality or access effects found',
         'supplies': 'The definitive measured effect of RSNAT prior '
                     'authorization', 'url': EVALF_URL, 'tier': 'A',
         'accessed': accessed, 'powers': ['RSNAT_Series']},
        {'key': 'gao_18_341', 'publisher': 'GAO',
         'document': 'GAO-18-341, Medicare: CMS Should Take Actions to '
                     'Continue Prior Authorization Efforts to Reduce '
                     'Spending (April 2018)',
         'vintage': 'Demonstrations through March 2017',
         'locator': 'Table 2 (p.16): RSNAT estimated potential savings '
                    '$349.5M in the 3 initial states plus $38.0M in the 6 '
                    'expansion states through March 2017; p.12: provisional '
                    'affirmation rate 28% (Dec 2014-May 2015) to 66% '
                    '(Oct 2016-Mar 2017)',
         'supplies': 'The independent savings estimate and the '
                     'affirmation-rate learning curve',
         'url': GAO_URL, 'tier': 'A', 'accessed': accessed,
         'powers': ['RSNAT_Series']},
        {'key': 'mln_rsnat', 'publisher': 'CMS MLN',
         'document': 'MLN6805343: RSNAT Prior Authorization Model fact '
                     'sheet (January 2026 edition)',
         'vintage': 'January 2026',
         'locator': 'HCPCS Codes section (p.6): A0426 and A0428 subject to '
                    'prior authorization; A0425 mileage treated as '
                    'associated and recouped if the transport is denied; '
                    'repetitive = 3+ transports in 10 days or weekly for 3+ '
                    'weeks',
         'supplies': 'The exact code scope of the screen',
         'url': MLN_URL, 'tier': 'A', 'accessed': accessed,
         'powers': ['RSNAT_Series']},
        # GADCS appendix: same document already registered by B.1
        # (sec_b1_facility_pay); the assembler dedupes on this key.
        {'key': 'gadcs_appendix', 'publisher': 'CMS / RAND Health Care',
         'document': 'Medicare Ground Ambulance Data Collection System '
                     '(GADCS) Report Appendix, Year 1-Year 4 Cohort '
                     'Analysis (PR-A2743-9, December 2025)',
         'vintage': 'GADCS Year 1-4 cohorts, data through 15 May 2025',
         'locator': 'Table 2.30 (pp.64-65): average payer revenue per NPI, '
                    'Traditional FFS Medicare and Medicare Advantage, with '
                    'provider/supplier, volume, ownership, density and '
                    'public-safety cuts',
         'supplies': 'The only national published MA-vs-FFS ambulance '
                     'revenue pairing', 'url': GADCS_URL, 'tier': 'A',
         'accessed': accessed, 'powers': ['MA_Book_Calibrator']},
    ]

    # ------------------------------------------------------------ TAB 1 ---
    den = _denial_pivot_cells(wb)
    ws = wb.create_sheet('RSNAT_Series')
    sb = lib.SheetBuilder(ws, 7, col_widths=[16, 54, 30, 14, 17, 11, 46],
                          tab_color='FF8C1D40')
    sb.title('RSNAT: the prior-authorization regime on repetitive scheduled '
             'non-emergent ambulance transport, 2014 to national')
    sb.subtitle('The question: what did prior authorization do to scheduled '
                'non-emergent transport, and when did each state enter the '
                'regime? Built end to end from published CMS documents, all '
                'fetched 11 Jul 2026: the CMS RSNAT model page and FAQ '
                '(timeline, statutory chain), the three Mathematica '
                'evaluation reports (Feb 2018, Sep 2020, May 2021), '
                'GAO-18-341, and MLN6805343 (code scope: A0426 and A0428 '
                'only, A0425 mileage associated). Joins to this workbook: '
                'the A0428 denial-rate series on PSPS_Denial_Series and the '
                'dialysis channel on Dialysis_ESRD_Channel.')
    sb.note('DATA QUALITY: evaluation effects are for the NINE 1115A model '
            'states and the ESRD and/or pressure-ulcer study population, '
            'not national totals; GAO savings are "as high as" potential '
            'estimates against a no-model counterfactual; affirmation rates '
            'come from two different instruments (MAC administrative data '
            'in GAO vs supplier survey in Mathematica) and are not one '
            'series; CMS stopped evaluating after 2019 because it had '
            'adequate evidence for the expansion decision.')
    sb.blank()

    sb.banner('Panel A. Milestones: every date is from the CMS RSNAT page '
              'or the CMS RSNAT FAQ, fetched 11 Jul 2026')
    sb.headers(['Date', 'Event', 'Authority / scope', '', '', '', 'Cite'])
    a0 = sb.r + 1
    milestones = [
        ('Jun 2013', 'MedPAC reports six states with far-above-average '
                     'non-emergent ambulance spending per dialysis '
                     'beneficiary; NJ, PA, SC rank top three by total '
                     'non-emergent ambulance spend - the selection basis '
                     'for the model states',
         'MedPAC June 2013 Report to Congress', 'CMS FAQ Q10'),
        ('15 Dec 2014', 'Model begins: prior authorization required for '
                        'repetitive scheduled non-emergent ambulance '
                        'transport in NJ, PA, SC (submissions from '
                        '1 Dec 2014)',
         'SSA sec. 1115A (CMMI model test)', 'CMS RSNAT page; GAO-18-341 '
         'p.10'),
        ('1 Jan 2016', 'Expansion to DE, DC, MD, NC, VA, WV (announced '
                       '15 Dec 2015). VA is the first footprint state in',
         'MACRA sec. 515(a) (Pub. L. 114-10)', 'CMS RSNAT page; FAQ Q10'),
        ('Mar 2018', 'CMS Chief Actuary certifies that nationwide expansion '
                     'would reduce (or not increase) net Medicare spending',
         'SSA sec. 1115A(c)(2) criterion', 'CMS RSNAT page; FAQ Q7'),
        ('1 Dec 2020', 'Original 1115A model ends in the nine states; they '
                       'transition without interruption to the expanded '
                       'MACRA model on 2 Dec 2020',
         'SSA sec. 1834(l)(16), added by MACRA sec. 515(b)', 'CMS FAQ Q4'),
    ] + [
        (d, f'National expansion phase {i + 1}: {states}',
         'SSA sec. 1834(l)(16); expansion required once 1115A(c)(1)-(3) '
         'criteria met' if i == 0 else 'same',
         'CMS RSNAT page; FAQ Q2')
        for i, (d, states) in enumerate(PHASES)
    ] + [
        ('Ongoing', 'Expanded model has no specified end date. Covered '
                    'codes: A0426 (ALS non-emergency) and A0428 (BLS '
                    'non-emergency) only; A0425 mileage is associated and '
                    'recouped if the transport is denied',
         'MLN6805343 (Jan 2026)', 'MLN p.6; FAQ Q4'),
    ]
    for i, (d, ev, auth, cite) in enumerate(milestones):
        sb.row([(d, 'src'), (ev, 'text'), (auth, 'text'), None, None, None,
                (cite, 'note')], wrap=True)
    sb.note('Footprint-state entry dates: IA, KS, MN, MO, NE, WI on '
            '1 Apr 2022 (phase 3); IN on 1 Jun 2022 (phase 4); KY and OH on '
            '1 Aug 2022 (phase 5); VA on 1 Jan 2016 (MACRA expansion). '
            'Phases were rolled out by MAC jurisdiction (phase 1 = '
            'Jurisdiction H, Novitas).')
    sb.blank()

    sb.banner('Panel B. Affirmation rates and review-volume evidence, as '
              'published (two instruments, never one series)')
    sb.headers(['Measure', 'Value', 'Period', '', 'Instrument', '',
                'Locator'])
    b0 = sb.r + 1
    b_rows = [
        ('Provisional affirmation rate, initial PAR submissions', 0.28,
         'Dec 2014 - May 2015', 'MAC administrative data',
         'GAO-18-341, p.12'),
        ('Provisional affirmation rate, initial PAR submissions', 0.66,
         'Oct 2016 - Mar 2017', 'MAC administrative data',
         'GAO-18-341, p.12'),
        ('PARs approved on initial submission, Year 1 states', 0.36,
         'Supplier survey, 14-20 months in', 'Mathematica supplier survey',
         'Final report, Figure VI.8'),
        ('PARs approved on initial submission, Year 2 states', 0.45,
         'Supplier survey, 14-20 months in', 'Mathematica supplier survey',
         'Final report, Figure VI.8'),
        ('Additional share of PARs approved after resubmission (Year 1)',
         0.31, 'Supplier survey', 'Mathematica supplier survey',
         'Final report, Ch. VI'),
        ('Pre-model denial baseline, non-emergency ambulance claims', 0.035,
         'Baseline (pre-Dec 2014)', 'Claims (about 7 denials per 100 '
         'beneficiaries per quarter)', 'Final report, p.67'),
    ]
    for lbl, v, per, inst, loc in b_rows:
        sb.row([(lbl, 'text'), (v, 'src', lib.FMT_PCT1), (per, 'src'),
                None, (inst, 'text'), None, (loc, 'note')], wrap=True)
    sb.note('Review volumes: no absolute PAR counts are published. What is '
            'published is the direction: MACs reported significant '
            'decreases in PAR volume for beneficiaries not meeting medical '
            'necessity and in poorly documented PARs (final report p.68), '
            'the claims-denial spike at implementation reverted to baseline '
            'within eight quarters (p.67), and RSNAT suppliers per 100,000 '
            'beneficiaries fell 50% on implementation (second interim, '
            'Objective 3).')
    sb.blank()

    sb.banner('Panel C. Measured effects, report by report (model states, '
              'ESRD and/or pressure-ulcer population)')
    sb.headers(['Report (date)', 'Headline effect', 'Savings estimate $',
                '', '', '', 'Locator'])
    c0 = sb.r + 1
    c_rows = [
        ('First interim (Feb 2018)',
         'ESRD beneficiaries, Year 1 states: RSNAT use probability -80% '
         'from a 5.09% quarterly baseline; trips -87%; RSNAT payments -90%; '
         'ambulance suppliers per 100,000 beneficiaries -15%',
         171000000, 'Ch. IV; $171M for ESRD beneficiaries'),
        ('GAO-18-341 (Apr 2018)',
         'Potential savings through Mar 2017: $349.5M in the 3 initial '
         'states plus $38.0M in the 6 expansion states ("as high as" '
         'estimates vs a flat-spend counterfactual)',
         349500000, 'Table 2, p.16'),
        ('Second interim (Sep 2020)',
         'RSNAT use and spend down over 60%; RSNAT suppliers per 100,000 '
         'beneficiaries -50%; total Medicare -2% ($316 per beneficiary per '
         'quarter, about $650M cumulative)',
         550000000, 'Exec summary; $550M RSNAT savings, avg $138M/yr'),
        ('Final report (May 2021)',
         'RSNAT use -72%; total Medicare FFS -2.4% ($381 per beneficiary '
         'per quarter); Year 1 -74% vs Year 2 -64%; no adverse effect on '
         'ED use, hospitalization or mortality',
         746000000, 'Table ES.1; $746M RSNAT expenditure reduction'),
        ('CMS FAQ summary (post-expansion)',
         'CMS\'s own framing: use -72% and expenditures -76%, about $750M; '
         'total FFS -2.4%, about $1B over the first five years',
         1000000000, 'FAQ Q6; $1B is the five-year TOTAL-FFS figure'),
    ]
    for rep, eff, sav, loc in c_rows:
        sb.row([(rep, 'label'), (eff, 'text'), (sav, 'src', lib.FMT_USD),
                None, None, None, (loc, 'note')], wrap=True)
    sb.row([('GAO initial-3-state savings as share of the final report '
             'RSNAT reduction (live)', 'label'), None,
            (f'=C{c0 + 1}/C{c0 + 3}', 'fml', lib.FMT_PCT1), None, None,
            None, ('different windows and methods; scale check only',
                   'note')])
    sb.blank()

    sb.banner('Panel D. The regime boundary in THIS workbook\'s series '
              '(green cells are live links into PSPS_Denial_Series)')
    sb.headers(['A0428 BLS non-emergency denial rate (submitted basis)',
                'Rate', '', '', '', '', 'Read'])
    d0 = sb.r + 1
    for yr, note in [(2014, 'last pre-model year: no state under prior '
                            'authorization'),
                     (2022, 'national rollout year: phases 1-5 complete by '
                            '1 Aug 2022'),
                     (2024, 'first full year of the national regime')]:
        ref = den.get(yr)
        sb.row([(f'A0428 denial rate, {yr}', 'label'),
                (f"='PSPS_Denial_Series'!{ref}", 'link', lib.FMT_PCT1)
                if ref else ('PENDING', 'note'),
                None, None, None, None, (note, 'note')])
    sb.note('Guardrail, printed where the join happens: prior authorization '
            'screens claims BEFORE submission, so its first-order effect is '
            'on the COMPOSITION of the submitted series (PSPS), not on the '
            'denial rate. Affirmed transports arrive pre-cleared (denials '
            'fall); deterred transports never enter the denominator at all '
            '(the dialysis-pair collapse on Dialysis_ESRD_Channel). The '
            'A0428 denial series must never be read across the 2014/2022 '
            'boundaries as one regime.')
    sb.blank()
    sb.banner('Read panel')
    sb.prose('What is established: a prior-authorization screen on exactly '
             'two HCPCS codes (A0426, A0428) removed roughly seven of every '
             'ten repetitive scheduled transports in the states it touched, '
             'cut RSNAT spending by about three quarters ($746M over the '
             'nine 1115A states through 2019), halved the RSNAT supplier '
             'density, and did so with no measured harm to quality or '
             'access - which is why the statute forced it national once '
             'the actuary certified savings. The rollout dates in Panel A '
             'are the regime boundaries for every scheduled-transport '
             'series in this workbook: IA/KS/MN/MO/NE/WI entered Apr 2022, '
             'IN Jun 2022, KY/OH Aug 2022, VA Jan 2016. What is NOT '
             'established: any national post-expansion affirmation series '
             '(CMS has not published one), and any read of PSPS denial '
             'rates across the boundary as demand.')

    facts += [
        {'metric': 'RSNAT baseline utilization, Year 1 cohort (share of '
                   'ESRD and/or pressure-ulcer beneficiaries using RSNAT, '
                   'baseline 2012-2014)', 'year': 2014, 'value': 0.101,
         'unit': 'share of study population', 'basis': 'GOV', 'tier': 'A',
         'source_keys': ['rsnat_eval_final'],
         'locator': 'Final report, Table II.2, p.9',
         'lives_on': 'RSNAT_Series',
         'cross_check': 'Year 2 cohort 4.7%, comparison states 4.9% on the '
                        'same table: the model states were picked for high '
                        'baseline use (FAQ Q10)'},
        {'metric': 'GAO estimated potential RSNAT savings, 3 initial '
                   'states, implementation through Mar 2017', 'year': 2017,
         'value': 349500000, 'unit': 'USD', 'basis': 'GOV', 'tier': 'A',
         'source_keys': ['gao_18_341'],
         'locator': 'GAO-18-341, Table 2, p.16', 'lives_on': 'RSNAT_Series',
         'cross_check': 'Plus $38.0M in the 6 expansion states, same '
                        'table; "as high as" counterfactual estimates'},
        {'metric': 'RSNAT provisional affirmation rate, initial PAR '
                   'submissions', 'year': 2017, 'value': 0.66,
         'unit': 'share of initial PARs affirmed, Oct 2016-Mar 2017',
         'basis': 'GOV', 'tier': 'A', 'source_keys': ['gao_18_341'],
         'locator': 'GAO-18-341, p.12', 'lives_on': 'RSNAT_Series',
         'cross_check': 'Up from 28% in the first six months (Dec 2014-'
                        'May 2015): a documentation learning curve, not a '
                        'loosening of the screen'},
        {'metric': 'RSNAT use reduction, final evaluation (nine model '
                   'states, ESRD and/or pressure-ulcer population)',
         'year': 2021, 'value': -0.72, 'unit': 'change vs comparison',
         'basis': 'GOV', 'tier': 'A', 'source_keys': ['rsnat_eval_final'],
         'locator': 'Final report, Table ES.1 and Ch. III',
         'lives_on': 'RSNAT_Series',
         'cross_check': 'RSNAT expenditures -$746M; total Medicare FFS '
                        '-2.4%; no adverse quality or access findings'},
        {'metric': 'RSNAT suppliers per 100,000 beneficiaries, change on '
                   'implementation (model states)', 'year': 2020,
         'value': -0.50, 'unit': 'change', 'basis': 'GOV', 'tier': 'A',
         'source_keys': ['rsnat_eval_second'],
         'locator': 'Second interim report, Objective 3 (exec summary)',
         'lives_on': 'RSNAT_Series',
         'cross_check': 'Concentrated almost entirely in Year 1 states; '
                        'exiting suppliers were smaller and dialysis-'
                        'dependent - the supply-side echo of finding 26 '
                        '(dialysis transports -63%)'},
    ]

    findings.append({
        'id_hint': 76,
        'finding': 'RSNAT prior authorization is the structural screen on '
                   'scheduled transport: a documentation gate on two HCPCS '
                   'codes removed about 72% of repetitive scheduled '
                   'non-emergent transport and $746M of spending in the '
                   'nine test states with no measured harm to quality, '
                   'halved RSNAT supplier density, and was then forced '
                   'national by statute in five MAC-jurisdiction phases '
                   'ending 1 Aug 2022. Every footprint state has operated '
                   'under the screen since Aug 2022 at the latest; VA '
                   'since Jan 2016.',
        'numbers': f"='RSNAT_Series'!C{c0 + 3}",
        'sources': 'rsnat_eval_final; gao_18_341; cms_rsnat_model; '
                   'cms_rsnat_faq',
        'confidence': 'High: three independent evaluations and GAO agree '
                      'on direction and rough magnitude',
        'guardrail': 'Prior authorization changes the COMPOSITION of the '
                     'SUBMITTED claims series (PSPS): affirmed transports '
                     'arrive pre-cleared and deterred transports never '
                     'enter the denominator. Never read the denial-rate '
                     'series across the RSNAT boundary (2014 for NJ/PA/SC, '
                     '2016 for the six MACRA states, 2021-2022 phases '
                     'elsewhere) as one regime, and never read the '
                     'scheduled-transport volume drop as demand '
                     'disappearing.'})

    # ------------------------------------------------------------ TAB 2 ---
    dark = _dark_share_cells(wb)
    ws2 = wb.create_sheet('MA_Book_Calibrator')
    sb = lib.SheetBuilder(ws2, 7, col_widths=[38, 15, 9, 15, 9, 11, 42],
                          tab_color='FF7A5195')
    sb.title('The MA book calibrator: Medicare Advantage revenue per NPI at '
             '1.04x Medicare FFS (GADCS Table 2.30)')
    sb.subtitle('The question: what is the best published multiplier from a '
                'supplier\'s measured Medicare FFS book to its invisible '
                'Medicare Advantage book? GADCS Table 2.30 (pp.64-65) is '
                'the only national pairing of FFS and MA ambulance revenue '
                'per NPI, self-reported by 6,000+ organizations. The '
                'all-NPI means put MA at 1.038x FFS - and the published '
                'cuts show that one number hides a spread from 0.55x '
                '(super rural) to 1.19x (government). This tab formalizes '
                'the calibrator the dark-share arithmetic on '
                'Imbalance_Ledger needs, and prints its own death clause.')
    sb.note('DATA QUALITY: GADCS revenue is self-reported survey data '
            '(MedPAC judged the COST side unusable for margins; the '
            'revenue-by-payer table is carried as the census it is); '
            'means per NPI, not medians, so large operators dominate; n '
            'differs by cell because not every organization reports every '
            'payer; ratios of means are NOT mean ratios; and the pairing '
            'is revenue per NPI, not price or volume - it cannot separate '
            'an MA rate discount from an MA utilization difference.')
    sb.blank()

    sb.banner('Panel A. The headline bound (blue = GADCS Table 2.30; '
              'ratio is live)')
    sb.headers(['Average payer revenue per NPI', 'Mean $', 'n', 'Mean $',
                'n', 'MA / FFS', 'Locator'])
    sb.row([('', 'text'), ('Traditional FFS Medicare', 'label'), None,
            ('Medicare Advantage', 'label'), None, None, None])
    a0 = sb.r + 1
    lbl, ffs, nf, ma, nm = T230_ALL
    sb.row([(lbl, 'label'), (ffs, 'src', lib.FMT_USD), (nf, 'src',
            lib.FMT_INT), (ma, 'src', lib.FMT_USD), (nm, 'src', lib.FMT_INT),
            (f'=D{a0}/B{a0}', 'fml', lib.FMT_X),
            ('Table 2.30, pp.64-65', 'note')])
    sb.note('The calibrator: multiply a measured Medicare FFS ambulance '
            'book by about 1.04x to bound the same supplier\'s MA book - '
            'PER NPI and ON AVERAGE. It is a revenue bound, not a rate '
            'bound: GADCS cannot say whether MA pays more per transport or '
            'MA members ride more.')
    sb.blank()

    sb.banner('Panel B. Every published cut of the same table (the '
              'calibrator is NOT one number)')
    sb.headers(['Cut (GADCS Table 2.30)', 'FFS mean $', 'n', 'MA mean $',
                'n', 'MA / FFS', 'Locator'])
    b0 = sb.r + 1
    for group, rows in T230_CUTS:
        sb.row([(group, 'label'), None, None, None, None, None, None])
        for lbl, ffs, nf, ma, nm in rows:
            r = sb.r + 1
            sb.row([('  ' + lbl, 'text'), (ffs, 'src', lib.FMT_USD),
                    (nf, 'src', lib.FMT_INT), (ma, 'src', lib.FMT_USD),
                    (nm, 'src', lib.FMT_INT),
                    (f'=D{r}/B{r}', 'fml', lib.FMT_X),
                    ('Table 2.30, pp.64-65', 'note')])
    b_end = sb.r
    sb.note('Read: the MA multiplier is a composition story. For-profit '
            '(1.08x), government (1.19x) and urban (1.06x) books carry MA '
            'at or above parity; hospital-based providers (0.59x), '
            'super-rural (0.55x) and low-volume (0.78x) books do not. '
            'Applying the 1.038x all-NPI mean to a specific operator '
            'ignores exactly the cuts this panel publishes - use the cut '
            'that matches the book.')
    sb.blank()

    sb.banner('Panel C. What it calibrates (green = live links)')
    sb.headers(['Series', 'Value', '', '', '', '', 'Read'])
    c0 = sb.r + 1
    for yr, note in [
            (2013, 'dark share when the MUP series starts'),
            (2024, 'the share of the Medicare transport market public '
                   'claims cannot see'),
            (2025, 'latest; drifting toward the CBO 63% by 2034')]:
        ref = dark.get(yr)
        sb.row([(f'Beneficiaries outside A and B FFS ("dark share"), {yr}',
                 'label'),
                (f"='Imbalance_Ledger'!{ref}", 'link', lib.FMT_PCT1)
                if ref else ('PENDING', 'note'),
                None, None, None, None, (note, 'note')])
    tam = None
    if 'TAM_Model_National' in wb.sheetnames:
        tam = _find_label(wb['TAM_Model_National'],
                          'MA allowed as a multiple')
    sb.row([('TAM model assumption this bound disciplines: MA allowed as a '
             'multiple of Medicare FFS (base case)', 'label'),
            (f"='TAM_Model_National'!C{tam[0]}", 'link', lib.FMT_DEC2)
            if tam else ('PENDING', 'note'),
            None, None, None, None,
            ('the 1.038x revenue-per-NPI bound sits inside the model\'s '
             '1.00-1.15x price-multiple range; different objects (revenue '
             'vs price), same discipline', 'note')], wrap=True)
    sb.blank()

    sb.banner('Panel D. Supersession clause (printed, not implied)')
    from openpyxl.styles import Border, Side
    pend = Border(bottom=Side(style='dotted', color='FF8C1D40'),
                  top=Side(style='dotted', color='FF8C1D40'),
                  left=Side(style='dotted', color='FF8C1D40'),
                  right=Side(style='dotted', color='FF8C1D40'))
    sup_r = sb.r + 1
    sb.row([('THIS CALIBRATOR DIES THE DAY MA ENCOUNTER DATA LANDS. The '
             'ResDAC MA encounter ambulance extract (receiving schema '
             'MA_Encounter_Recv, request drafted on Engagement_Data_Map) '
             'replaces a survey revenue ratio with measured encounters and '
             'paid amounts by HCPCS and origin-destination. When that file '
             'arrives, every use of 1.038x is superseded row by row.',
             'note'), ('PENDING', 'note'), None, None, None, None,
            ('supersession target: MA_Encounter_Recv', 'note')], wrap=True,
           height=44)
    for col in range(1, 8):
        ws2.cell(row=sup_r, column=col).border = pend
    sb.blank()
    sb.banner('Read panel')
    sb.prose('What is measured: the average ambulance NPI that reports '
             'both books earns $1,238,482 from Medicare Advantage against '
             '$1,193,046 from Medicare FFS - MA at 1.038x FFS - and the '
             'published cuts spread that multiplier from 0.55x (super '
             'rural) to 1.19x (government). Used with the dark-share '
             'series (59% of beneficiaries outside the FFS window in '
             '2024), the bound says the invisible MA transport book of a '
             'commercial operator is roughly the size of its visible FFS '
             'book, give or take composition. What is NOT measured: MA '
             'price vs utilization (revenue conflates them), and anything '
             'per transport. The clause in Panel D is the exit: this tab '
             'exists to be replaced by MA encounter data.')

    facts += [
        {'metric': 'MA-to-FFS ambulance revenue ratio per NPI (the MA '
                   'book calibrator)', 'year': 2025, 'value': 1.038,
         'unit': 'x (ratio of means per NPI)', 'basis': 'DERIVED',
         'tier': 'A', 'source_keys': ['gadcs_appendix'],
         'locator': 'GADCS Y1-4 appendix, Table 2.30, pp.64-65: '
                    '$1,238,482 / $1,193,046; live formula on '
                    'MA_Book_Calibrator Panel A',
         'lives_on': 'MA_Book_Calibrator',
         'cross_check': 'Published cuts run 0.55x (super rural) to 1.19x '
                        '(government); for-profit 1.08x; superseded the '
                        'day MA encounter data (MA_Encounter_Recv) lands'},
        {'metric': 'MA dark share this calibrator prices (beneficiaries '
                   'outside A and B FFS)', 'year': 2024, 'value': 0.593,
         'unit': 'share of Medicare beneficiaries', 'basis': 'DERIVED',
         'tier': 'A', 'source_keys': ['gadcs_appendix'],
         'locator': 'Imbalance_Ledger Panel C 2024 row (live link on '
                    'MA_Book_Calibrator Panel C); enrollment inputs S60',
         'lives_on': 'MA_Book_Calibrator',
         'cross_check': 'Engagement_Data_Map quotes the same 59.3%; CBO '
                        'projects 63% by 2034'},
    ]

    findings.append({
        'id_hint': 77,
        'finding': 'The MA book of a US ambulance supplier is, on the only '
                   'published national pairing, the same size as its FFS '
                   'book: GADCS Table 2.30 puts average MA revenue per NPI '
                   'at $1,238,482 vs $1,193,046 FFS, a 1.038x calibrator '
                   'for pricing the 59% dark share - with published cuts '
                   'from 0.55x (super rural) to 1.19x (government) that '
                   'make the all-NPI mean a starting point, not an answer.',
        'numbers': f"='MA_Book_Calibrator'!F{a0}",
        'sources': 'gadcs_appendix',
        'confidence': 'Moderate: self-reported survey means per NPI, '
                      'n>6,000, no external replication exists',
        'guardrail': 'SUPERSESSION: this calibrator is a survey revenue '
                     'ratio and it dies the day MA encounter data lands '
                     '(receiving schema MA_Encounter_Recv). Until then it '
                     'bounds revenue only - it cannot split MA price from '
                     'MA utilization, and it must be applied by cut '
                     '(ownership, density, volume), never as one number '
                     'to a specific operator.'})

    return {'facts': facts, 'sources': sources, 'excluded': excluded,
            'findings': findings,
            'meta': {'denial_links': den, 'dark_links': dark}}
