"""B.3: the Medicaid ambulance rate card for the footprint states.

Nine footprint-state Medicaid FFS fee schedules pulled and verified against
the actual published file or portal extract (medicaid_rates.json):
  Nebraska 471-000-504 (July 2026 xlsx), Iowa MedicaidFeeSched portal
  C11.csv, Kansas KMAP TXIX fee schedule files (7/1/2026), Missouri MHD
  Ambulance.xlsx (7/7/2026), Ohio OAC 5160-15-28 appendix (1/1/2024),
  Wisconsin ForwardHealth ambulance max-fee report (rates of 7/1/2008),
  Minnesota MHCP Fee Schedule xlsx (rates eff 1/1/2026), Indiana IHCP
  Professional Fee Schedule xlsx (eff 1/1/2026, 100% of CY2025 Medicare),
  Kentucky DMS 2026 Transportation Rates PDF (rev 1/9/2026).
Virginia is PARKED for dollar rates (broker-billed non-emergency ambulance;
the only public rate files are the 2009 / 2009-2012 vintage) and carried as
PENDING rows.

Comparison panel prices each state's rate against the CY2026 Medicare
national unadjusted base located by scanning Derived_Rate_Card for the
code labels (green links when found; blue AFS re-carry as fallback).
"""

SHEETS = [{'name': 'Medicaid_Rate_Card',
           'question': 'What does Medicaid actually pay for ground ambulance '
                       'in the footprint states, and how much scheduled '
                       'transport never touches these schedules?'}]

CODES = ['A0428', 'A0429', 'A0426', 'A0427', 'A0433', 'A0434', 'A0425']
DESC = {'A0428': 'BLS non-emergency', 'A0429': 'BLS emergency',
        'A0426': 'ALS1 non-emergency', 'A0427': 'ALS1 emergency',
        'A0433': 'ALS2', 'A0434': 'Specialty Care Transport',
        'A0425': 'Ground mileage'}
# CY2026 AFS fallback (blue re-carry) if Derived_Rate_Card scan fails:
AFS_RVU = {'A0428': 1.0, 'A0429': 1.6, 'A0426': 1.2, 'A0427': 1.9,
           'A0433': 2.75, 'A0434': 3.25}
AFS_CF, AFS_MILE = 284.56, 9.15

# Nine covered footprint states (VA parked separately). NE-WI unchanged from
# the original six; MN/IN/KY added and verified 12 Jul 2026.
STATES = ['NE', 'IA', 'KS', 'MO', 'OH', 'WI', 'MN', 'IN', 'KY']
SNAME = {'NE': 'Nebraska', 'IA': 'Iowa', 'KS': 'Kansas', 'MO': 'Missouri',
         'OH': 'Ohio', 'WI': 'Wisconsin', 'MN': 'Minnesota',
         'IN': 'Indiana', 'KY': 'Kentucky', 'VA': 'Virginia'}

# rate, effective (as printed), in-row note  (None rate = not covered / n/a)
RATES = {
 'NE': {'A0425': (6.35, '07/01/2026', 'per statute mile'),
        'A0426': (387.24, '07/01/2026', 'flat ALS rate: A0426=A0427=A0433=A0434'),
        'A0427': (387.24, '07/01/2026', ''),
        'A0428': (154.89, '07/01/2026', 'schedule comment: includes bariatric transfers'),
        'A0429': (189.93, '07/01/2026', ''),
        'A0433': (387.24, '07/01/2026', ''),
        'A0434': (387.24, '07/01/2026', '')},
 'IA': {'A0425': (2.61, '07/01/2014', 'per statute mile'),
        'A0426': (101.60, '10/01/2015', ''),
        'A0427': (127.01, '10/01/2015', ''),
        'A0428': (84.67, '07/01/2014', 'lowest true BLS A0428 in the footprint'),
        'A0429': (114.30, '10/01/2015', ''),
        'A0433': (232.84, '05/01/2020', ''),
        'A0434': (232.84, '12/01/2020', 'same rate as ALS2')},
 'KS': {'A0425': (13.20, '07/01/2023', 'Ambulance rate-type file; legacy $3.00 '
                  'row (eff 07/01/2022) persists on the general schedule'),
        'A0426': (306.84, '07/01/2023', ''),
        'A0427': (485.83, '07/01/2023', ''),
        'A0428': (255.70, '07/01/2023', 'SFY2024 rebase'),
        'A0429': (409.12, '07/01/2023', ''),
        'A0433': (703.18, '07/01/2023', ''),
        'A0434': (831.03, '07/01/2023', '')},
 'MO': {'A0425': (7.04, '07/01/2024', 'per statute mile'),
        'A0426': (172.26, '07/01/2019', 'HH (hospital-to-hospital) modifier '
                  'ONLY; unmodified code noncovered'),
        'A0427': (474.75, '07/01/2024', ''),
        'A0428': (105.62, '07/01/2019', 'HH and HD (hospital-to-diagnostic) '
                  'contexts ONLY; unmodified code noncovered'),
        'A0429': (379.69, '07/01/2024', ''),
        'A0433': (215.20, '07/01/2019', ''),
        'A0434': (None, '07/01/2002', 'NONCOVERED under MO HealthNet FFS')},
 'OH': {'A0425': (5.05, '01/01/2024', 'per statute mile'),
        'A0426': (244.50, '01/01/2024', ''),
        'A0427': (289.75, '01/01/2024', ''),
        'A0428': (203.75, '01/01/2024', ''),
        'A0429': (244.00, '01/01/2024', ''),
        'A0433': (349.50, '01/01/2024', ''),
        'A0434': (413.00, '01/01/2024', '')},
 'WI': {'A0425': (5.56, '07/01/2008', 'GM multi-patient half-rate $2.78'),
        'A0426': (113.88, '07/01/2008', ''),
        'A0427': (180.31, '07/01/2008', ''),
        'A0428': (94.90, '07/01/2008', 'unchanged for 18 years'),
        'A0429': (151.84, '07/01/2008', ''),
        'A0433': (260.97, '07/01/2008', ''),
        'A0434': (308.42, '07/01/2008', '')},
 # ── added 12 Jul 2026, verified against the primary files ──
 'MN': {'A0425': (9.51, '04/01/2026', 'per statute mile (ground codes eff '
                  '01/01/2026; mileage 04/01/2026)'),
        'A0426': (341.47, '01/01/2026', ''),
        'A0427': (540.66, '01/01/2026', ''),
        'A0428': (284.56, '01/01/2026', 'highest A0428 in the footprint - '
                  'exactly the CY2026 Medicare conversion factor ($284.56), '
                  'i.e. Medicare parity'),
        'A0429': (455.29, '01/01/2026', ''),
        'A0433': (782.54, '01/01/2026', ''),
        'A0434': (924.82, '01/01/2026', 'highest SCT in the footprint')},
 'IN': {'A0425': (9.15, '01/01/2026', 'per statute mile; reduced-modifier '
                  'variants (U3/U5, MRT/SE) pay $1.67'),
        'A0426': (322.83, '01/01/2026', ''),
        'A0427': (511.14, '01/01/2026', ''),
        'A0428': (269.02, '01/01/2026', '100% of CY2025 Medicare (BT2025156)'),
        'A0429': (430.44, '01/01/2026', ''),
        'A0433': (739.81, '01/01/2026', ''),
        'A0434': (None, '04/01/2004', 'NONCOVERED (NOCOV) under IHCP')},
 'KY': {'A0425': (2.50, '01/09/2026', 'BLS mileage (A0425 UB) Other $2.50 / '
                  'Hospital $3.00; ALS mileage $2.50/$4.00; stretcher and '
                  'return-trip mileage $2.00'),
        'A0426': (None, '01/09/2026', 'NOT LISTED - no ALS1 non-emergency '
                  'line on the KY transportation schedule'),
        'A0427': (60.00, '01/09/2026', 'Other (independent) $60.00 / '
                  'Hospital-based $110.00; A0427 GM add-patient $25'),
        'A0428': (55.00, '01/09/2026', 'NON-EMERGENCY STRETCHER line (PT56 '
                  'Specialty 16), flat $55 - a stretcher-van rate, NOT a full '
                  'BLS ambulance; formerly T2005, renamed A0428 eff 1/1/2025'),
        'A0429': (60.00, '01/09/2026', 'Other (independent) $60.00 / '
                  'Hospital-based $82.50; A0429 GM add-patient $20'),
        'A0433': (None, '01/09/2026', 'NOT LISTED - no ALS2 line'),
        'A0434': (None, '01/09/2026', 'NOT LISTED - no SCT line')},
}

# (effective, pa, broker, dialysis) per state
META = {
 'NE': ('Nebraska DHHS fee schedule 471-000-504, Ambulance Services, '
        'July 1 2026 (xlsx). Schedule cover: rates carry NO increase - no '
        'rate-increase appropriations this state fiscal year.',
        'PA column blank on every ground code (no PA flag); policy 471 NAC 4.',
        'YES - Heritage Health MCO brokers (MTM for NE Total Care and '
        'Molina; ModivCare for UHC, MTM Health from 1 Jan 2026); DHHS staff '
        'arrange residual FFS NET under 471-000-503.',
        'No dialysis restriction printed on the ambulance schedule; '
        'scheduled dialysis rides route through the NET benefit.'),
 'IA': ('Iowa Medicaid fee schedule portal extract, provider type 11 '
        'AMBULANCE (C11.csv), accessed 11 Jul 2026; per-code effective '
        'dates 2014-2020, open-ended.',
        'No PA; non-emergency requires bed-confinement or equivalent '
        'necessity criteria (Ambulance Services Provider Manual Ch. III, '
        '1 Oct 2015).',
        'YES - NEMT broker Access2Care renamed MTM Health 18 Sep 2025 '
        '(IA Health Link MCOs); MTM administers FFS NEMT.',
        'Dialysis NEMT rides schedulable 90 days ahead; ambulance claim '
        'destination modifiers G/J are dialysis facilities.'),
 'KS': ('KMAP TXIX fee schedule files, 7/1/2026 versions '
        '(FeeSchedule_20260701_TXIX_AMB.txt / _DEF.txt); ambulance rates '
        'effective 7/1/2023 (SFY2024 rebase).',
        'No PA, but ALL nonemergent transports need a Medical Necessity '
        'form attached to the claim (KMAP manual).',
        'YES - KanCare is near-total managed care; NEMT via MCO broker '
        'ModivCare (taxi to stretcher van and ambulance).',
        'Dialysis named among critical services eligible for NEMT in the '
        'Kansas NEMT provider manual.'),
 'MO': ('MHD Ambulance.xlsx, file dated 7 Jul 2026 (provider files updated '
        '07/08/2026); rate effective dates 2019-2024 per row.',
        'PA = No on all covered rows; A0428 billable ONLY with HH or HD '
        'modifier (MHD hot tip).',
        'YES - statewide NEMT broker MTM (mydss.mo.gov/mhd/nemt).',
        'Dialysis runs are NOT billable as FFS A0428 (HH/HD only); '
        'scheduled dialysis travel is an MTM broker trip.'),
 'OH': ('Appendix to OAC rule 5160-15-28 (file dated 22 Dec 2023), rule '
        'effective 1 Jan 2024.',
        'No PA in rule; necessity per OAC 5160-15-23(B): EMT care, oxygen, '
        'or supervised restraint during transport, or qualifying '
        'hospital-to-hospital transfer.',
        'NO statewide broker - county CDJFS NET programs (5160-15-10) plus '
        'MCO rides; ambulance stays a direct ODM/MCO claim. GEMT '
        'supplemental for public EMS (5160-15-30).',
        'Rules retrieved are silent on dialysis transport.'),
 'WI': ('ForwardHealth AMBULANCE MAXIMUM ALLOWABLE FEE SCHEDULE, live '
        'portal report accessed 11 Jul 2026; every ground code effective '
        '07/01/2008.',
        'No code-level PA flag; non-emergency needs ALS/BLS care or supine '
        'transport, and manager-coordinated trips need a trip ID within '
        'two business days (Update 2014-32).',
        'YES - statewide NEMT manager (MTM Inc.; Veyo from 11/2021 until '
        'MTM acquisition). Manager-coordinated NEMT ambulance claims are '
        'paid by the manager, NOT ForwardHealth.',
        'Recurring dialysis rides are standing rides with the NEMT '
        'manager; stretcher vans sit on the separate SMV schedule.'),
 'MN': ('Minnesota MHCP Fee Schedule xlsx (mn.gov/dhs), downloaded 12 Jul '
        '2026; ground rates effective 01/01/2026 (mileage A0425 04/01/2026); '
        'one max-fee row per code, status A.',
        'No PA flag on the ambulance rows; MHCP covers emergency and '
        'non-emergency ground/air ambulance billed directly to MHCP by '
        'enrolled providers (ambulance is not routed through the '
        'administered NEMT system).',
        'MIXED / STATE-ADMINISTERED (no single MCO broker) - split between '
        'local county/tribal agency-administered NEMT and State-Administered '
        'NEMT (Acentra Health / Kepro); ambulance itself stays FFS. NEMT '
        'providers need MnDOT STS certification.',
        'Explicit RETURN-LEG rule: for dialysis members eligible for local '
        'NEMT, State-Administered NEMT covers the RETURN trip - Kepro enters '
        'a single-day certification for dialysis return trips only.'),
 'IN': ('Indiana IHCP Professional Fee Schedule xlsx (in-session portal '
        'download), last updated 9 Jun 2026; ambulance rates effective '
        '01/01/2026, set to 100% of the CY2025 Medicare fee schedule '
        '(bulletin BT2025156).',
        'No PA flag on covered ambulance rows; A0434 SCT is Non-Covered '
        '(PA=Y, $0.00). ALS/BLS ambulance is EXEMPT from the FFS NEMT '
        'brokerage and billed FFS directly.',
        'PARTIAL - non-ambulance FFS NEMT brokered statewide through Verida '
        '(formerly Southeastrans); managed care via the MCE. ALS/BLS '
        'ambulance is carved OUT of the broker and billed FFS.',
        'Recurring dialysis trips are standing orders through the Verida '
        'FFS broker (or the MCE); the ambulance schedule is silent at code '
        'level.'),
 'KY': ('KY Medicaid Transportation Fee Schedule 2026, revised 1/9/2026 '
        '(no changes vs prior year). PT55 emergency ambulance, PT56 '
        'Specialty 16 non-emergency stretcher; Hospital vs Other variants, '
        'Other recorded as headline.',
        'No code-level PA on the schedule; emergency ambulance (PT55) billed '
        'FFS. Only non-emergency ambulance line is the PT56 stretcher; '
        'broader NEMT goes through the HSTD regional broker, off-schedule.',
        'YES - Human Service Transportation Delivery (HSTD): DMS contracts '
        'with the KY Transportation Cabinet, which assigns a regional broker '
        'per county; providers contract with the broker, trips booked 72h '
        'ahead. Five MCOs administer NEMT for their members.',
        'Recurring dialysis trips route through the HSTD regional broker (or '
        'the MCO); the ambulance/stretcher schedule is silent on dialysis.'),
}

URLS = {
 'NE': 'https://dhhs.ne.gov/Medicaid%20Practitioner%20Fee%20Schedules/'
       '471-000-504%20Ambulance%20-%20July%20%2726.xlsx',
 'IA': 'https://secureapp.dhs.state.ia.us/MedicaidFeeSched/C11.csv',
 'KS': 'https://portal.kmap-state-ks.us/PublicPage/ProviderPricing/'
       'FeeSchedules',
 'MO': 'https://apps.dss.mo.gov/fmsfeeschedules/dlfiles/Ambulance.xlsx',
 'OH': 'https://codes.ohio.gov/assets/laws/administrative-code/pdfs/5160/0/'
       '15/5160-15-28_PH_FF_A_APP1_20231222_0942.pdf',
 'WI': 'https://www.forwardhealth.wi.gov/WIPortal/Tab/42/icscontent/'
       'provider/maxfee/pdf/maxfee06_ambulance.pdf.spage',
 'MN': 'https://mn.gov/dhs/assets/mhcp-fee-schedule_tcm1053-293968.xlsx',
 'IN': 'https://provider.indianamedicaid.com/ihcp/Publications/MaxFee/'
       'Professional%20Fee%20Schedule.xlsx',
 'KY': 'https://www.chfs.ky.gov/agencies/dms/DMSFeeRateSchedules/'
       '2026TransportationRates.pdf',
 'VA': 'https://www.dmas.virginia.gov/media/1569/'
       'fee-schedules-for-emergency-ground-ambulance-a0427-and-a0429.pdf',
}

N = len(STATES)                       # 9 covered states
LAST_STATE_COL = 2 + N                # column of the last state in the matrix
# True-BLS A0428 comparison excludes KY (stretcher line) at the last column.
BLS_A0428_STATES = [s for s in STATES if s != 'KY']


def _mc_refs(wb):
    """Scan Derived_Rate_Card for the code labels; return {code: '$D$13'}."""
    if 'Derived_Rate_Card' not in wb.sheetnames:
        return {}
    ws = wb['Derived_Rate_Card']
    refs = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if not isinstance(v, str):
            continue
        s = v.strip()
        if s in AFS_RVU and s not in refs:
            refs[s] = f'$D${row[0].row}'          # national unadjusted base
        elif s.startswith('A0425 CY2026 national') and 'A0425' not in refs:
            refs['A0425'] = f'$B${row[0].row}'    # mileage national rate
    return refs


def build(wb, ctx):
    from openpyxl.utils import get_column_letter
    lib = ctx['lib']
    accessed = ctx['accessed']
    facts, sources, excluded, findings = [], [], [], []

    mc = _mc_refs(wb)
    scan_ok = all(c in mc for c in CODES)

    sources += [
        {'key': 'ne_amb_fs', 'publisher': 'Nebraska DHHS Medicaid',
         'document': 'Nebraska Medicaid Fee Schedule, Ambulance Services, '
                     'July 1 2026 (471-000-504 Ambulance - July 26.xlsx)',
         'vintage': 'Effective 1 Jul 2026 (SFY2027, no increase per cover)',
         'locator': 'Rows 000A0425-000A0434, MEDICAID ALLOWABLE column',
         'supplies': 'NE ground ambulance FFS rates + PA flags',
         'url': URLS['NE'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'ia_amb_fs', 'publisher': 'Iowa HHS (Iowa Medicaid)',
         'document': 'Iowa Medicaid fee schedule portal, provider type 11 '
                     'AMBULANCE, file C11.csv (via the I-Accept gate on the '
                     'hhs.iowa.gov fee-schedules page)',
         'vintage': 'Live extract 11 Jul 2026; code rates eff 2014-2020',
         'locator': 'C11.csv rows A0425-A0436, Factor column',
         'supplies': 'IA ground ambulance FFS rates + effective dates',
         'url': URLS['IA'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'ks_kmap_fs', 'publisher': 'Kansas KMAP (KDHE)',
         'document': 'KMAP public fee schedules: FeeSchedule_20260701_TXIX_'
                     'AMB.txt.zip and FeeSchedule_20260701_TXIX_DEF.txt.zip',
         'vintage': 'Posted versions of 7/1/2026; rates eff 7/1/2023',
         'locator': 'Fixed-width rows A0425-A0434, MAX FEE column; retrieved '
                    'by replicating the portal AJAX flow (list POST then '
                    'DownloadFSFile in-session)',
         'supplies': 'KS ground ambulance FFS rates (TXIX)',
         'url': URLS['KS'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'mo_mhd_fs', 'publisher': 'Missouri DSS MO HealthNet',
         'document': 'MHD fee schedule download Ambulance.xlsx, file dated '
                     '7 Jul 2026 (provider files updated 07/08/2026)',
         'vintage': 'Monthly refresh; row effective dates 2019-2024',
         'locator': 'Rows A0425-A0434 with modifier context, Rate and PA '
                    'columns; A0428 HH/HD restriction per MHD hot tip',
         'supplies': 'MO ground ambulance FFS rates, coverage restrictions',
         'url': URLS['MO'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'oh_odm_fs', 'publisher': 'Ohio Dept of Medicaid (OAC)',
         'document': 'Appendix to OAC rule 5160-15-28 (Transportation: '
                     'payment), file 5160-15-28_PH_FF_A_APP1_20231222',
         'vintage': 'Rule and rates effective 1 Jan 2024',
         'locator': 'Appendix pp.1-2, ground ambulance rows A0425-A0434; '
                    'necessity criteria in rule 5160-15-23(B)',
         'supplies': 'OH ground ambulance FFS rates + payment rule',
         'url': URLS['OH'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'wi_maxfee_amb', 'publisher': 'Wisconsin DHS ForwardHealth',
         'document': 'AMBULANCE MAXIMUM ALLOWABLE FEE SCHEDULE (portal-'
                     'generated live report, maxfee06_ambulance.pdf)',
         'vintage': 'Report as in effect 11 Jul 2026; rates eff 7/1/2008',
         'locator': 'Provider type 25 rows A0425-A0436, MAX FEE column; '
                    'NEMT-manager routing per ForwardHealth Update 2014-32',
         'supplies': 'WI ambulance max fees + NEMT manager carve',
         'url': URLS['WI'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'mn_mhcp_fs', 'publisher': 'Minnesota DHS (MHCP)',
         'document': 'Minnesota Health Care Programs (MHCP) Fee Schedule '
                     '(mhcp-fee-schedule_tcm1053-293968.xlsx)',
         'vintage': 'Downloaded 12 Jul 2026; ground rates eff 01/01/2026 '
                    '(mileage A0425 eff 04/01/2026)',
         'locator': 'Sheet1 rows A0425-A0434, Max Fee column, factor code B, '
                    'status A',
         'supplies': 'MN ground ambulance FFS rates + effective dates',
         'url': URLS['MN'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'in_ihcp_fs', 'publisher': 'Indiana FSSA (IHCP)',
         'document': 'IHCP Professional Fee Schedule (Professional Fee '
                     'Schedule.xlsx) with bulletin BT2025156',
         'vintage': 'Last updated 9 Jun 2026; ambulance rates eff 01/01/2026 '
                    '(100% of CY2025 Medicare per BT2025156)',
         'locator': 'Rows A0425-A0434, Max Fee/Non-Facility column, Rate '
                    'Effective Date 01/01/2026; retrieved in-session via the '
                    'maxfee agreement gate (maxfee_validate.asp?agreement='
                    'accept then fee_home.asp download)',
         'supplies': 'IN ground ambulance FFS rates; A0434 noncovered',
         'url': URLS['IN'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'ky_dms_fs', 'publisher': 'Kentucky CHFS DMS',
         'document': 'KY Medicaid Transportation Fee Schedule 2026 '
                     '(2026TransportationRates.pdf, revised 1/9/2026)',
         'vintage': 'Revised 1/9/2026 (no rate change vs prior year)',
         'locator': 'Provider Type 55 (emergency ambulance) and Provider '
                    'Type 56 Specialty 16 (non-emergency stretcher), '
                    'Reimbursement Rate column (Hospital vs Other)',
         'supplies': 'KY emergency ambulance + non-emergency stretcher FFS '
                     'rates (no ALS1-nonE / ALS2 / SCT lines)',
         'url': URLS['KY'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
        {'key': 'va_dmas_amb', 'publisher': 'Virginia DMAS',
         'document': 'DMAS Emergency Ground Ambulance rate PDF (A0427/A0429) '
                     'and Crossover rate table (A0425-A0434); Transportation '
                     'provider manual Chapter 5 (rev 12/5/2025)',
         'vintage': 'Rate PDFs are the 2009 and 2009-2012 vintages; manual '
                    'rev 12/5/2025',
         'locator': 'Emergency Ground PDF "DOS October 31 2009 and Before" '
                    'sliding scale; Crossover table "DOS Nov 1 2009 - Jun 30 '
                    '2012"; manual Ch.5 FFS NEMT Broker section for A0428',
         'supplies': 'VA broker carve + stale sliding-scale (current per-code '
                     'FFS rates PARKED as PENDING)',
         'url': URLS['VA'], 'tier': 'A', 'accessed': accessed,
         'powers': ['Medicaid_Rate_Card']},
    ]

    n_cols = 12
    ws = wb.create_sheet('Medicaid_Rate_Card')
    sb = lib.SheetBuilder(
        ws, n_cols, tab_color='FF8C1D40',
        col_widths=[32, 15, 11, 10, 11, 8, 8, 9, 9, 9, 9, 44])
    sb.title('Medicaid ground ambulance rate card - the footprint states')
    sb.subtitle('The question: what does each footprint state Medicaid FFS '
                'schedule actually pay for ground ambulance (A0425-A0434), '
                'and how much scheduled transport is carved out to NEMT '
                'brokers and never priced on these schedules? Sources: nine '
                'published state fee schedules named on each panel, each '
                'fetched and parsed from the primary file or portal (NE/IA/KS/'
                'MO/OH/WI on 11 Jul 2026; MN/IN/KY on 12 Jul 2026); Virginia '
                'is PARKED (only 2009-vintage public files). Join key is the '
                'HCPCS code. Comparison prices each state against the CY2026 '
                'Medicare national unadjusted base on Derived_Rate_Card.')
    sb.note('DATA QUALITY: Medicaid FFS schedule rates ONLY - managed-care '
            'plans negotiate their own rates and broker-carve states route '
            'scheduled volume outside these schedules, so the real price of '
            'scheduled transport is unobservable here. Effective dates differ '
            'by 18 years (WI 7/1/2008 vs MN/IN 1/1/2026): the cross-state '
            'comparison mixes vintages by construction. Rates are NOT '
            'apples-to-apples product-for-product: MO A0428 is HH/HD-context '
            'only; KY has NO ALS1-nonemergency, ALS2, or SCT line and its '
            'A0428 is a non-emergency STRETCHER-van line (formerly T2005), '
            'not a full BLS ambulance; IN and MO do not cover SCT (A0434). '
            'KY is therefore excluded from the A0428 BLS comparison range.')
    sb.blank()

    def _pad(cells):
        """Right-pad a cell list to n_cols so the last item lands in col 12."""
        return cells + [None] * (n_cols - len(cells))

    cell = {}     # (state, code) -> rate cell 'C12'
    for idx, st in enumerate(STATES):
        eff, pa, brk, dia = META[st]
        sb.banner(f'Panel {chr(65 + idx)}. {SNAME[st]} - '
                  f'{eff.split(";")[0].strip()[:82]}')
        sb.headers(_pad(['HCPCS', 'Service level', 'Rate $', 'Unit',
                         'Rate eff.']) [:-1] + ['In-row note'])
        for code in CODES:
            rate, reff, note = RATES[st][code]
            r = sb.r + 1
            if rate is None:
                lbl = 'NOT LISTED' if 'NOT LISTED' in note else 'NONCOVERED'
                sb.row(_pad([(code, 'text'), (DESC[code], 'text'),
                             (lbl, 'src'), ('base', 'text'),
                             (reff, 'src')])[:-1] + [(note, 'note')])
            else:
                sb.row(_pad([(code, 'text'), (DESC[code], 'text'),
                             (rate, 'src', lib.FMT_USD2),
                             ('per mile' if code == 'A0425' else 'base', 'text'),
                             (reff, 'src')])[:-1]
                       + [(note, 'note') if note else None])
                cell[(st, code)] = f'C{r}'
        sb.row(_pad([('Source file', 'label')])[:-1] + [(URLS[st], 'note')])
        sb.row([('PA rule', 'label'), (pa, 'note')], height=30)
        sb.row([('Broker carve', 'label'), (brk, 'note')], height=30)
        sb.row([('Dialysis policy', 'label'), (dia, 'note')], height=30)
        sb.blank()

    # ── Virginia: parked ─────────────────────────────────────────────
    va_letter = chr(65 + N)
    sb.banner(f'Panel {va_letter}. Virginia (extension state) - rates PARKED, '
              'carve verified')
    sb.headers(_pad(['HCPCS', 'Service level', 'Rate $'])[:-1]
               + ['Why parked / what fills it'])
    from openpyxl.styles import Border, Side
    va_pend_rows = []
    sb.row(_pad([('A0427/A0429', 'text'), ('Emergency codes', 'text'),
                 ('PENDING', 'note')])[:-1]
           + [('The only public DMAS emergency ground file is the 2009 '
               'sliding scale (VERIFIED "DOS October 31 2009 and Before"): '
               '1-5 mi $75, 6-10 mi $150, then +$2.50/mi. No current-'
               'effective FFS rate file is published; the DMAS procedure fee '
               'file search portal (interactive per-code lookup) would fill '
               'it.', 'note')], height=42)
    va_pend_rows.append(sb.r)
    sb.row(_pad([('A0428', 'text'), (DESC['A0428'], 'text'),
                 ('PENDING', 'note')])[:-1]
           + [('Not payable on a public FFS schedule: DMAS Transportation '
               'manual Ch.5 (rev 12/5/2025) requires A0428 to be '
               'preauthorized and billed to the FFS NEMT broker.', 'note')],
           height=30)
    va_pend_rows.append(sb.r)
    sb.row(_pad([('A0425/A0426/A0433/A0434', 'text'), ('Mileage / ALS / SCT',
                 'text'), ('PENDING', 'note')])[:-1]
           + [('Appear only on the 2009-2012 crossover rate table "for '
               'crossover calculation only"; no current standalone FFS '
               'rate is published.', 'note')], height=30)
    va_pend_rows.append(sb.r)
    # Bordered PENDING boxes (house rule: an unpriceable cell is a bordered
    # PENDING box naming the public dataset, visually distinct from priced
    # rate cells). The dataset filler is named in the adjacent note column.
    _pend_box = Border(*(Side(style='dotted', color='FF8C1D40'),) * 4)
    for rn in va_pend_rows:
        ws.cell(row=rn, column=3).border = _pend_box
    sb.row(_pad([('Source (verified)', 'label')])[:-1] + [(URLS['VA'], 'note')])
    sb.blank()

    # ── Comparison panel ─────────────────────────────────────────────
    cmp_letter = chr(65 + N + 1)
    sb.banner(f'Panel {cmp_letter}. Comparison - each state as a multiple of '
              'the CY2026 Medicare national unadjusted base (live formulas)')
    if scan_ok:
        sb.note('Medicare column is a GREEN link to Derived_Rate_Card '
                '(located by scanning that tab for the code labels: base '
                'codes column D, mileage cell B23).')
    else:
        sb.note('Derived_Rate_Card scan did NOT locate all code cells; the '
                'Medicare column re-carries the CY2026 AFS national base '
                '(RVU x $284.56 CF; mileage $9.15) as blue, per the CMS '
                'CY2026 Ambulance Fee Schedule public use file.')
    hdr_r = sb.r + 1
    sb.headers(['HCPCS (service)', 'Medicare CY26 $'] + STATES
               + ['Confound printed with the ratio'])
    cmp_row = {}
    for code in CODES:
        r = sb.r + 1
        cmp_row[code] = r
        if scan_ok:
            mc_cell = (f"='Derived_Rate_Card'!{mc[code]}", 'link',
                       lib.FMT_USD2)
        else:
            val = AFS_MILE if code == 'A0425' else AFS_RVU[code] * AFS_CF
            mc_cell = (round(val, 2), 'src', lib.FMT_USD2)
        row = [(f'{code} ({DESC[code]})', 'text'), mc_cell]
        for st in STATES:
            ref = cell.get((st, code))
            if ref is None:
                row.append(('n/a', 'note'))
            else:
                row.append((f'={ref}/$B${r}', 'fml', lib.FMT_X))
        conf = ('IA lowest true-BLS floor, MN at Medicare parity (1.00x); KY '
                'excluded (its A0428 is a stretcher-van line, not a BLS '
                'ambulance); MO HH/HD-context only'
                if code == 'A0428' else
                'MO and IN NONCOVERED - no ratio; KY has no SCT line'
                if code == 'A0434' else
                'KS mileage from the ambulance rate-type file, not the '
                'legacy $3.00 row; KY mileage is the BLS/Other rate'
                if code == 'A0425' else
                'MO A0426 is HH-modifier context; KY has no ALS1-nonE line'
                if code == 'A0426' else
                'KY has no ALS2 line (n/a)' if code == 'A0433' else
                'Same-code comparison; state schedules do not adjust for '
                'geography the way Medicare does')
        row.append((conf, 'note'))
        sb.row(row)
    # A0428 range over the true-BLS states (KY = last column, excluded)
    a_row = cmp_row['A0428']
    last_bls_col = get_column_letter(2 + len(BLS_A0428_STATES))   # skip KY
    r_min = sb.r + 1
    sb.row(_pad([('Footprint A0428 range - LOWEST multiple (Iowa)', 'label'),
                 None, (f'=MIN(C{a_row}:{last_bls_col}{a_row})', 'fml',
                        lib.FMT_X)])[:-1]
           + [('min over the eight true-BLS A0428 ratios (KY stretcher '
               'excluded)', 'note')])
    r_max = sb.r + 1
    sb.row(_pad([('Footprint A0428 range - HIGHEST multiple (Minnesota)',
                  'label'), None,
                 (f'=MAX(C{a_row}:{last_bls_col}{a_row})', 'fml',
                  lib.FMT_X)])[:-1]
           + [('max over the eight true-BLS A0428 ratios; MN = Medicare '
               'parity', 'note')])
    r_disp = sb.r + 1
    sb.row(_pad([('Dispersion: highest / lowest A0428 rate', 'label'), None,
                 (f'=C{r_max}/C{r_min}', 'fml', lib.FMT_X)])[:-1]
           + [('MN pays about 3.4x what IA pays for the same code', 'note')])
    ky_ref = cell.get(('KY', 'A0428'))
    if ky_ref:
        sb.row(_pad([('Memo - KY non-emergency STRETCHER (not BLS ambulance)',
                      'label'), None,
                     (f'={ky_ref}/$B${a_row}', 'fml', lib.FMT_X)])[:-1]
               + [('A0428 stretcher-van line (formerly T2005), $55 flat - a '
                   'different product; shown for context, excluded from the '
                   'range', 'note')])
    chart_anchor = f'{get_column_letter(n_cols + 2)}{hdr_r}'
    lib.add_chart(ws, chart_anchor,
                  'Medicaid A0428 as a multiple of Medicare CY2026',
                  f"'Medicaid_Rate_Card'!$C${hdr_r}:${get_column_letter(LAST_STATE_COL)}${hdr_r}",
                  [('A0428 multiple',
                    f"'Medicaid_Rate_Card'!$C${a_row}:"
                    f"${get_column_letter(LAST_STATE_COL)}${a_row}")],
                  kind='bar', y_fmt='0.00"x"')
    sb.blank()

    # ── Broker-carve summary ─────────────────────────────────────────
    carve_letter = chr(65 + N + 2)
    sb.banner(f'Panel {carve_letter}. Broker carve-out summary - who actually '
              'buys scheduled transport')
    sb.headers(_pad(['State', 'Carve?', 'Broker / manager'])[:-1]
               + ['What is carved'])
    b0 = sb.r + 1
    carves = [
        ('NE', 'YES', 'MTM / ModivCare (MTM Health from 1/1/26)',
         'NEMT via Heritage Health MCO brokers; FFS NET by DHHS staff'),
        ('IA', 'YES', 'MTM Health (was Access2Care to 9/18/25)',
         'NEMT for MCO and FFS members; ambulance stays FFS-billable'),
        ('KS', 'YES', 'ModivCare (KanCare MCOs)',
         'NEMT incl. stretcher van and NEMT ambulance coordination'),
        ('MO', 'YES', 'MTM (statewide)',
         'All scheduled NEMT; FFS ambulance only in HH/HD contexts'),
        ('OH', 'NO', 'County CDJFS NET + MCO vendors',
         'No statewide broker; ambulance is a direct ODM/MCO claim'),
        ('WI', 'YES', 'MTM Inc. (NEMT manager; Veyo 2021-22)',
         'Manager-coordinated NEMT ambulance is paid by the MANAGER, '
         'not ForwardHealth'),
        ('MN', 'YES', 'Acentra/Kepro (State-Admin) + county/tribal',
         'State-administered + local-agency NEMT (no single MCO broker); '
         'ambulance stays FFS'),
        ('IN', 'YES', 'Verida (formerly Southeastrans)',
         'Non-ambulance FFS NEMT brokered; ALS/BLS ambulance EXEMPT and '
         'billed FFS'),
        ('KY', 'YES', 'HSTD regional brokers (via KY Transp. Cabinet)',
         'NEMT via county regional brokers; emergency ambulance + PT56 '
         'stretcher billed FFS'),
    ]
    for st, y, who, what in carves:
        sb.row(_pad([(SNAME[st], 'text'), (y, 'src'), (who, 'src')])[:-1]
               + [(what, 'note')])
    r_cnt = sb.r + 1
    sb.row(_pad([('Broker/administered-carve states of the nine footprint '
                  'states', 'label'),
                 (f'=COUNTIF(B{b0}:B{b0 + N - 1},"YES")', 'fml',
                  lib.FMT_INT)])[:-1]
           + [('live count over the column above (OH is the only NO)',
               'note')])
    sb.row(_pad([('Memo - Virginia (extension state)', 'text'), ('YES', 'src'),
                 ('DMAS FFS NEMT broker', 'src')])[:-1]
           + [('Broker pays FFS NON-EMERGENCY AMBULANCE itself: no public '
               'FFS price for scheduled ambulance exists', 'note')])
    sb.blank()

    sb.banner('Read panel')
    sb.prose('What is measured: nine published state Medicaid FFS ambulance '
             'schedules, fetched and parsed from the primary file or portal '
             '(NE/IA/KS/MO/OH/WI on 11 Jul 2026; MN/IN/KY on 12 Jul 2026); '
             'Virginia is parked because its only public rate files are the '
             '2009 / 2009-2012 vintage. The BLS non-emergency base (A0428), '
             'the workhorse code of scheduled interfacility transport, ranges '
             'from $84.67 (IA, unchanged since 2014) to $284.56 (MN, exactly '
             'the Medicare conversion factor - full parity) - roughly 0.30x '
             'to 1.00x the CY2026 Medicare national base, a 3.4x spread, with '
             'WI frozen at $94.90 since 2008 and IN pinned to 100% of CY2025 '
             'Medicare. Kentucky is the outlier: it has NO ALS1-nonemergency, '
             'ALS2, or SCT line at all, and its only non-emergency ground '
             'line is a $55 stretcher-van rate (formerly T2005), so it is '
             'excluded from the BLS comparison. Eight of the nine states '
             'route scheduled NEMT through a broker, transportation manager, '
             'or a state-administered system (only OH runs county NET '
             'instead), and Missouri will not pay A0428 outside HH/HD '
             'contexts nor SCT at all (IN also does not cover SCT). What is '
             'NOT measured: managed-care negotiated rates, broker-paid trip '
             'prices, and GEMT supplemental payments - the real price of most '
             'scheduled Medicaid transport is set inside those channels and '
             'is not public.')

    A = 'Medicaid FFS ambulance fee schedule'
    fact_rows = [
        ('NE', 154.89, '1 Jul 2026', 'ne_amb_fs',
         '471-000-504 xlsx row 000A0428, MEDICAID ALLOWABLE',
         'Schedule cover states no increase for SFY2027'),
        ('IA', 84.67, 'eff 1 Jul 2014, current 11 Jul 2026', 'ia_amb_fs',
         'C11.csv row A0428, Factor column',
         'Lowest true-BLS A0428 in footprint; unchanged 12 years'),
        ('KS', 255.70, 'eff 1 Jul 2023, posted 7/1/2026', 'ks_kmap_fs',
         'FeeSchedule_20260701_TXIX_DEF.txt row A0428 MAX FEE',
         'SFY2024 rebase'),
        ('MO', 105.62, 'eff 1 Jul 2019, file 7 Jul 2026', 'mo_mhd_fs',
         'Ambulance.xlsx rows A0428 HH / A0428 HD',
         'Payable ONLY in HH/HD contexts; base code noncovered'),
        ('OH', 203.75, '1 Jan 2024', 'oh_odm_fs',
         'OAC 5160-15-28 appendix, row A0428',
         '2024 rebase; GEMT supplemental tops up public providers'),
        ('WI', 94.90, 'eff 1 Jul 2008, report 11 Jul 2026', 'wi_maxfee_amb',
         'maxfee06_ambulance.pdf row A0428 MAX FEE',
         'Unchanged 18 years; manager-coordinated NEMT paid off-schedule'),
        ('MN', 284.56, 'eff 1 Jan 2026', 'mn_mhcp_fs',
         'MHCP Fee Schedule xlsx row A0428, Max Fee column',
         'Highest A0428; exactly the CY2026 Medicare CF (parity)'),
        ('IN', 269.02, 'eff 1 Jan 2026', 'in_ihcp_fs',
         'IHCP Professional Fee Schedule xlsx row A0428, Rate Eff 01/01/2026',
         '100% of CY2025 Medicare (BT2025156); A0434 SCT noncovered'),
        ('KY', 55.00, 'eff 1 Jan 2026 (rev 1/9/2026)', 'ky_dms_fs',
         '2026 Transportation Rates PDF, PT56 Specialty 16 A0428',
         'Non-emergency STRETCHER-van line (formerly T2005), not a full BLS '
         'ambulance'),
    ]
    for st, val, eff, key, loc, xc in fact_rows:
        facts.append({'metric': f'{SNAME[st]} {A}: A0428 base rate ({eff})',
                      'year': 2026, 'value': val, 'unit': 'USD per transport',
                      'basis': 'GOV', 'tier': 'A', 'source_keys': [key],
                      'locator': loc, 'lives_on': 'Medicaid_Rate_Card',
                      'cross_check': xc})
    facts += [
        {'metric': 'Lowest footprint Medicaid true-BLS A0428 rate as a '
                   'multiple of the CY2026 Medicare national base (Iowa)',
         'year': 2026, 'value': 0.2975, 'unit': 'x Medicare',
         'basis': 'DERIVED', 'tier': 'A', 'source_keys': ['ia_amb_fs'],
         'locator': 'Medicaid_Rate_Card Panel, MIN over the true-BLS A0428 '
                    'ratios; 84.67 / 284.56 (KY stretcher excluded)',
         'lives_on': 'Medicaid_Rate_Card',
         'cross_check': 'Vintage confound printed in-row: IA rate is a 2014 '
                        'rate against a 2026 Medicare base'},
        {'metric': 'Highest footprint Medicaid A0428 rate as a multiple of '
                   'the CY2026 Medicare national base (Minnesota)',
         'year': 2026, 'value': 1.0000, 'unit': 'x Medicare',
         'basis': 'DERIVED', 'tier': 'A', 'source_keys': ['mn_mhcp_fs'],
         'locator': 'Medicaid_Rate_Card Panel, MAX over the true-BLS A0428 '
                    'ratios; 284.56 / 284.56 = exact parity',
         'lives_on': 'Medicaid_Rate_Card',
         'cross_check': 'MN A0428 equals the CY2026 Medicare conversion '
                        'factor to the cent'},
        {'metric': 'Footprint states routing scheduled NEMT through a broker, '
                   'transportation manager, or state-administered system',
         'year': 2026, 'value': 8, 'unit': 'of 9 states',
         'basis': 'DERIVED', 'tier': 'A',
         'source_keys': ['ne_amb_fs', 'ia_amb_fs', 'ks_kmap_fs', 'mo_mhd_fs',
                         'oh_odm_fs', 'wi_maxfee_amb', 'mn_mhcp_fs',
                         'in_ihcp_fs', 'ky_dms_fs'],
         'locator': 'Medicaid_Rate_Card carve panel, COUNTIF over the carve '
                    'column (OH is the only exception: county NET)',
         'lives_on': 'Medicaid_Rate_Card',
         'cross_check': 'VA (extension) goes further: the broker pays FFS '
                        'non-emergency ambulance itself'},
    ]

    findings.append({
        'id_hint': 75,
        'finding': 'The Medicaid price floor for scheduled ambulance '
                   'transport is low, wildly dispersed, and largely carved '
                   'away from the schedules entirely. Across the nine '
                   'footprint FFS schedules the same A0428 BLS non-emergency '
                   'transport pays $84.67 (IA, set 2014) to $284.56 (MN, '
                   '2026) - about 0.30x to a full 1.00x the CY2026 Medicare '
                   'national base and a 3.4x spread, with WI frozen since '
                   '2008 and IN pinned to 100% of CY2025 Medicare. Kentucky '
                   'has no ALS1-nonemergency, ALS2, or SCT line at all and '
                   'prices non-emergency ground as a $55 stretcher-van rate. '
                   'Eight of the nine states carve scheduled NEMT to brokers, '
                   'managers, or a state-administered system (only OH runs '
                   'county NET), so the schedule price is not even the paid '
                   'price for most scheduled volume.',
        'numbers': f"='Medicaid_Rate_Card'!C{r_disp}",
        'sources': 'ne_amb_fs; ia_amb_fs; ks_kmap_fs; mo_mhd_fs; oh_odm_fs; '
                   'wi_maxfee_amb; mn_mhcp_fs; in_ihcp_fs; ky_dms_fs; '
                   'va_dmas_amb',
        'confidence': 'High on the schedule values (each read from the '
                      'primary file); the multiples mix rate vintages by '
                      'construction',
        'guardrail': 'FFS schedule rates are NOT what managed-care plans '
                     'pay, and broker-carve states hide the real '
                     'scheduled-transport price inside broker and MCO '
                     'contracts; these floors bound the public price, not '
                     'the market price. Products are not identical across '
                     'states - MO A0428 exists only in HH/HD contexts, KY '
                     'A0428 is a stretcher-van line, and MO/IN pay nothing '
                     'for SCT.'})

    return {'facts': facts, 'sources': sources, 'excluded': excluded,
            'findings': findings,
            'meta': {'states_landed': STATES, 'states_parked': ['VA'],
                     'states_added_v34b': ['MN', 'IN', 'KY'],
                     'medicare_scan': 'green links' if scan_ok
                                      else 'blue AFS re-carry',
                     'research_record': 'medicaid_rates.json'}}
