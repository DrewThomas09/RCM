"""B.14: the application and request package - ready-to-send request drafts
appended to Engagement_Data_Map (cost and timeline stated) plus seven empty
bordered receiving schemas so licensed data lands into a fixed shape the
moment any request is approved.
"""

SHEETS = [{'name': n, 'question': 'Receiving schema for licensed data (B.14)'}
          for n in ('MA_Encounter_Recv', 'TAF_Ambulance_Recv',
                    'NEMSIS_State_IFT', 'HCUP_Transfer_Recv', 'AHA_Recv',
                    'Claims_Vendor_Recv', 'Commercial_Rate_MRF')]

SCHEMAS = {
    'MA_Encounter_Recv': (
        'Medicare Advantage encounter ambulance extract (ResDAC DUA)',
        ['Encounter year', 'HCPCS', 'Origin modifier', 'Destination modifier',
         'State', 'County FIPS', 'Encounters', 'Paid amount $',
         'Plan type', 'Notes'],
        'ResDAC MA encounter (EDGE-adjacent) ambulance slice; request drafted '
        'on Engagement_Data_Map; DUA + months of lead time; this schema is '
        'the landing shape'),
    'TAF_Ambulance_Recv': (
        'Medicaid T-MSIS TAF ambulance extract (ResDAC)',
        ['Claim year', 'State', 'HCPCS', 'Origin-destination modifiers',
         'Claims', 'Paid $', 'Managed-care vs FFS flag', 'Broker flag',
         'Dual status', 'Notes'],
        'TAF OT claims, ambulance HCPCS; the Medicaid volume the public '
        'fee-schedule layer cannot show'),
    'NEMSIS_State_IFT': (
        'NEMSIS state research extracts / TAC research request',
        ['State', 'Year', 'Activations (911)', 'Activations (interfacility)',
         'Transport disposition share', 'ALS share', 'Response mode',
         'Data-use terms', 'Vintage', 'Notes'],
        'State-level interfacility activation counts with the IFT flag the '
        'public national report aggregates away'),
    'HCUP_Transfer_Recv': (
        'HCUP Central Distributor order: SID + SEDD, footprint states',
        ['State', 'Year', 'Discharges with transfer-out', 'ED transfers',
         'Origin hospital ID', 'Destination class', 'Payer class',
         'DRG/CCSR group', 'Weight', 'Notes'],
        'Hospital-identified transfer flows for footprint states; closes the '
        'hospital-pair gap named on Dataset_Linkage_Map'),
    'AHA_Recv': (
        'AHA Annual Survey license extract',
        ['AHA ID', 'CCN', 'System ID', 'State', 'Ambulance service owned '
         '(yes/no)', 'Beds', 'Admissions', 'Survey year', 'Notes', ''],
        'The ownership flag that turns the insourcing bounds into a measured '
        'split'),
    'Claims_Vendor_Recv': (
        'Commercial claims panel (any vendor; spec is vendor-neutral)',
        ['Billing entity NPI', 'Billing TIN', 'HCPCS', 'Origin modifier',
         'Destination modifier', 'Payer class', 'Service year-month',
         'Allowed $', 'Sample completeness % (by CBSA)', 'Notes'],
        'Exact fields stated so any claims vendor can quote: billing NPI/TIN, '
        'HCPCS, O-D modifiers, payer class, service dates, plus required '
        'sample-completeness statistics by CBSA'),
    'Commercial_Rate_MRF': (
        'Transparency-in-Coverage negotiated-rate slices (5 payers x '
        'footprint states)',
        ['Payer', 'State', 'TIN', 'NPI (where given)', 'HCPCS',
         'Negotiated rate $', 'Rate type', 'File month', 'Slice SHA-256',
         'Medicare multiple (live)'],
        'The receiving shape for the MRF pilot and scale-out; every slice '
        'hashed; Medicare-multiple column computes against Derived_Rate_Card'),
}

REQUESTS = [
    ('ResDAC: MA encounter ambulance extract', 'MA_Encounter_Recv',
     'CMS ResDAC research request; DUA required',
     'Fee schedule per ResDAC (thousands); 3-6 months',
     'Closes the MA dark share measured on Imbalance_Ledger'),
    ('ResDAC: T-MSIS TAF ambulance extract', 'TAF_Ambulance_Recv',
     'CMS ResDAC; DUA', 'Thousands; 3-6 months',
     'Closes the Medicaid volume gap; broker-carve states flagged'),
    ('NEMSIS: state data requests (footprint) + TAC research request',
     'NEMSIS_State_IFT', 'Per-state data-use agreements; TAC research form',
     'Free; paperwork weeks to months',
     'State interfacility activation counts'),
    ('HCUP Central Distributor: SID + SEDD, footprint states, latest two '
     'years', 'HCUP_Transfer_Recv', 'AHRQ DUA + purchase',
     'Low four figures per state-year', 'Hospital-pair transfer flows'),
    ('AHA Annual Survey license', 'AHA_Recv', 'Commercial license',
     'Four to five figures annually',
     'Ambulance-ownership flag per hospital; SAM filter hardening'),
    ('Claims vendor panel (vendor-neutral spec)', 'Claims_Vendor_Recv',
     'Commercial quote against the stated field spec',
     'Five figures typical; weeks',
     'Commercial allowed + facility-remit visibility; the spec is written '
     'so any vendor can fill it'),
    ('Biospatial: quote request', 'NEMSIS_State_IFT',
     'Commercial; b.iQ platform', 'Quote-based',
     'Near-real-time EMS activation feed as a NEMSIS alternative'),
]


def build(wb, ctx):
    lib = ctx['lib']
    from openpyxl.styles import Border, Side
    pend = Border(bottom=Side(style='dotted', color='FF8C1D40'),
                  top=Side(style='dotted', color='FF8C1D40'),
                  left=Side(style='dotted', color='FF8C1D40'),
                  right=Side(style='dotted', color='FF8C1D40'))

    for name, (title, cols, note) in SCHEMAS.items():
        if name in wb.sheetnames:
            continue
        ws = wb.create_sheet(name)
        sb = lib.SheetBuilder(ws, len(cols),
                              col_widths=[18] * len(cols),
                              tab_color='FF6B7C93')
        sb.title(f'Receiving schema: {title}')
        sb.subtitle('An empty, bordered landing shape (handoff B.14). No '
                    'cell here carries a value until the licensed or '
                    'requested data arrives; the bordered rows are the '
                    'contract for what lands. ' + note + '.')
        sb.blank()
        sb.headers(cols)
        r0 = sb.r + 1
        for r in range(r0, r0 + 12):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).border = pend
        sb.r += 12
        sb.blank()
        sb.note('PENDING by design: this schema ships empty. The request '
                'that fills it is drafted on Engagement_Data_Map with cost '
                'and timeline.')

    # append the request register to Engagement_Data_Map
    target = ('Engagement_Data_Map' if 'Engagement_Data_Map' in wb.sheetnames
              else 'Dataset_Linkage_Map')
    if target in wb.sheetnames:
        ws = wb[target]
        r = ws.max_row + 2
        import v3lib as _v3
        c = ws.cell(row=r, column=1,
                    value='v3.4 request package (B.14): ready-to-send '
                          'drafts; each lands into its named receiving '
                          'schema tab')
        c.font = _v3.F_BANNER
        c.fill = _v3.FILL_BANNER
        for i in range(2, 6):
            ws.cell(row=r, column=i).fill = _v3.FILL_BANNER
        r += 1
        for i, h in enumerate(['Request', 'Receiving schema', 'Access route',
                               'Cost / timeline', 'What it closes'], start=1):
            hc = ws.cell(row=r, column=i, value=h)
            hc.font = _v3.F_HDR
            hc.fill = _v3.FILL_HDR
        r += 1
        for req in REQUESTS:
            for i, v in enumerate(req, start=1):
                cc = ws.cell(row=r, column=i, value=v)
                cc.font = _v3.F_TXT
                cc.alignment = _v3.AL_WRAP
            ws.row_dimensions[r].height = 28
            r += 1

    return {'facts': [], 'sources': [], 'excluded': [], 'findings': [],
            'meta': {'schemas': len(SCHEMAS), 'requests': len(REQUESTS),
                     'register_tab': target}}
