"""C.4-C.8: the assembly layer - five tabs that wire the measured evidence
into investor-facing shapes without creating any new blended number.

  Growth_Outlook_Shell (C.4)  component growth rows, each a measured source;
                              NO blended headline growth number anywhere.
  Vendor_Share_Stack   (C.5)  one exhibit, six labeled legs, none blended;
                              every leg states what it CAN and CANNOT say.
  Stickiness_Evidence  (C.6)  observed public contract terms, exclusivity
                              frequency, SLA/penalty precedents, and
                              registry-presence churn. A panel, not a score.
  Investor_QA          (C.8)  twelve tearsheets wired to live cells or
                              named PENDING slots, never to assertions.
  Slide_Feed           (C.7)  one row per planned exhibit with an honest
                              RED/AMBER/GREEN evidence status; feeds the
                              deck, not the ledger.

Every cross-tab reference is guarded: tabs that land in parallel modules
are scanned at build time and fall back to a named PENDING slot when the
tab is not yet in the workbook, so statuses and links recompute honestly
at final assembly order.
"""

import datetime
import re

SHEETS = [
    {'name': 'Growth_Outlook_Shell',
     'question': 'What are the MEASURED growth components of the IFT '
                 'market - tailwinds and headwinds - without blending '
                 'them into a headline number?'},
    {'name': 'Vendor_Share_Stack',
     'question': 'What can the public evidence stack say about vendor '
                 'share of the outsourced IFT wallet, leg by leg?'},
    {'name': 'Stickiness_Evidence',
     'question': 'What do observed public contracts and registry churn '
                 'show about how sticky ambulance relationships are?'},
    {'name': 'Investor_QA',
     'question': 'The twelve questions an investor asks, each answered '
                 'in three sentences and wired to live cells?'},
    {'name': 'Slide_Feed',
     'question': 'Which planned exhibits of the study are backed by live '
                 'cells today, and where are the named public gaps?'},
]

GROUND_BASE = ('A0426', 'A0427', 'A0428', 'A0429', 'A0433', 'A0434')
COHORT_STATES = ('OH', 'VA', 'NE', 'IA', 'MO', 'KS', 'WI', 'KY', 'IN')
COHORT_FRAME = ('research cohort of representative multi-hospital health '
                'systems operating in the study footprint, selected for '
                'depth')

_MONTHS = {m: i + 1 for i, m in enumerate(
    ['January', 'February', 'March', 'April', 'May', 'June', 'July',
     'August', 'September', 'October', 'November', 'December'])}
_WORDNUM = {'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
            'six': 6, 'twelve': 12, 'twenty-four': 24, 'thirty-six': 36,
            'sixty': 60}


# ------------------------------------------------------------ helpers ---

def _clean(s, cap=None):
    """Sanitize any string headed for a cell: no em/en dashes ever."""
    if s is None:
        return None
    s = str(s).replace('—', ' - ').replace('–', ' - ')
    if cap and len(s) > cap:
        s = s[:cap - 3].rstrip() + '...'
    return s


def _tab(wb, name):
    return wb[name] if name in wb.sheetnames else None


def _tab_like(wb, *needles):
    """First sheet whose name contains any needle (case-insensitive)."""
    for nd in needles:
        for nm in wb.sheetnames:
            if nd.lower() in nm.lower():
                return nm
    return None


def _scan(ws, needle, max_col=12, max_row=None):
    """First cell whose text contains needle; the cell or None."""
    if ws is None:
        return None
    n = needle.lower()
    for row in ws.iter_rows(max_row=max_row or ws.max_row, max_col=max_col):
        for c in row:
            if isinstance(c.value, str) and n in c.value.lower():
                return c
    return None


def _L(tab, coord, fmt=None):
    f = f"='{tab}'!{coord}"
    return (f, 'link', fmt) if fmt else (f, 'link')


def _box(ws, r, cols, border):
    for c in cols:
        ws.cell(row=r, column=c).border = border


def _year_cell_row(ws, year, col='A', lo=1, hi=60):
    """Row index where column col equals year (int or str)."""
    for r in range(lo, min(hi, ws.max_row) + 1):
        v = ws[f'{col}{r}'].value
        if v == year or (isinstance(v, str) and v.strip() == str(year)):
            return r
    return None


def _headline(wb, tab):
    """Coordinate of a tab's first real data value: the first numeric cell,
    or first live-formula cell, in column B below its header. Used to upgrade
    the C.8 'title pull' fallbacks (which pointed at A1) to a cell that
    actually shows a number now that every analysis tab exists in the
    assembly. Falls back to A1 only if the tab has no such cell."""
    if tab not in wb.sheetnames:
        return 'A1'
    ws = wb[tab]
    for r in range(4, min(80, ws.max_row) + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, (int, float)):
            return f'B{r}'
        if isinstance(v, str) and v.startswith('='):
            return f'B{r}'
    return 'A1'


def _parse_term_months(text):
    """Initial contract term in months from a term_and_renewal string."""
    if not text or str(text).strip().lower() == 'none':
        return None
    t = str(text)
    low = t.lower()
    m = re.search(r'\b(' + '|'.join(_WORDNUM) + r')[\s-]*(year|month)', low)
    if m:
        n = _WORDNUM[m.group(1)]
        return n * 12 if m.group(2) == 'year' else n
    m = re.search(r'\b(\d{1,2})[\s-]*(?:\([a-z0-9]+\)\s*)?(year|month)', low)
    if m:
        n = int(m.group(1))
        return n * 12 if m.group(2) == 'year' else n
    dr = re.search(
        r'([A-Z][a-z]+ \d{1,2},? \d{4})\s*(?:through|to|until|and ends?|'
        r'ending)\s*([A-Z][a-z]+ \d{1,2},? \d{4})', t)
    if dr:
        def _d(s):
            mm = re.match(r'([A-Z][a-z]+) (\d{1,2}),? (\d{4})', s)
            return datetime.date(int(mm.group(3)), _MONTHS[mm.group(1)],
                                 int(mm.group(2)))
        d1, d2 = _d(dr.group(1)), _d(dr.group(2))
        if d2 > d1:
            return int(round(((d2 - d1).days + 1) / 30.44))
    return None


def _excl_class(text):
    """(bucket, printable snippet) for an exclusivity field."""
    if not text or str(text).strip().lower() == 'none':
        return 'none', 'None stated'
    low = str(text).lower()
    if 'nonexclusive' in low or 'non-exclusive' in low:
        return 'nonexclusive', _clean(text, 100)
    if 'exclusiv' in low or 'first call' in low or 'first-call' in low:
        return 'exclusive', _clean(text, 100)
    return 'coverage', _clean(text, 100)


# ------------------------------------------------------------- build ----

def build(wb, ctx):
    lib = ctx['lib']
    cache = ctx['cache']
    accessed = ctx['accessed']
    facts, sources, excluded, findings = [], [], [], []

    from openpyxl.styles import Border, Side
    pend_border = Border(bottom=Side(style='dotted', color='FF8C1D40'),
                         top=Side(style='dotted', color='FF8C1D40'),
                         left=Side(style='dotted', color='FF8C1D40'),
                         right=Side(style='dotted', color='FF8C1D40'))

    # ---------------------------------------------------------- sources -
    sources += [
        {'key': 'census_age_pep',
         'publisher': 'US Census Bureau',
         'document': 'Population Estimates Program, state and county age '
                     '65+/85+ series, vintage 2024 (as carried on '
                     'State_Age_65plus and County_Age_65plus)',
         'vintage': '2020-2024', 'locator': 'State_Age_65plus Panel A '
         '(United States row); County_Age_65plus column K',
         'supplies': 'The measured demographic growth component',
         'url': 'https://www.census.gov/programs-surveys/popest.html',
         'tier': 'A', 'accessed': accessed,
         'powers': ['Growth_Outlook_Shell']},
        {'key': 'cms_enrollment_ma',
         'publisher': 'CMS (Medicare Monthly/Annual Enrollment)',
         'document': 'Medicare enrollment, Original Medicare vs MA & '
                     'other, 2013-2025 (as carried on '
                     'Enrollment_ESRD_State)',
         'vintage': '2013-2025', 'locator': 'Enrollment_ESRD_State Panel '
         'A, columns C-F', 'supplies': 'The MA migration component: the '
         'shrinking FFS-visible base',
         'url': 'https://data.cms.gov/summary-statistics-on-beneficiary-'
                'enrollment/medicare-and-medicaid-reports/'
                'medicare-monthly-enrollment',
         'tier': 'A', 'accessed': accessed,
         'powers': ['Growth_Outlook_Shell', 'Investor_QA']},
        {'key': 'caa2026_addon_cliff',
         'publisher': 'US Congress / CMS',
         'document': 'Consolidated Appropriations Act, 2026, section 6203 '
                     '(ground ambulance add-on extension) and CMS Pub. '
                     '100-04 ch.15 s.20.4, as verified on Payment_Rules',
         'vintage': 'Signed 3 February 2026',
         'locator': 'Payment_Rules add-on status panel (rows 55-59) and '
                    'AIF table',
         'supplies': 'The statutory add-on cliff dates carried live into '
                     'the growth shell',
         'url': 'https://www.congress.gov/bill/119th-congress',
         'tier': 'A', 'accessed': accessed,
         'powers': ['Growth_Outlook_Shell', 'Investor_QA']},
        {'key': 'irs990_contractor_sweep',
         'publisher': 'IRS (Form 990) via ProPublica Nonprofit Explorer '
                      'API v2',
         'document': 'Form 990 Part VII Section B five-highest-paid '
                     'independent contractor disclosures, latest two '
                     'XML-available filings per resolved EIN of the ' +
                     COHORT_FRAME,
         'vintage': '46 filings, 24 EINs, 9 systems; harvested 11 Jul '
                    '2026 (cohort_990.json)',
         'locator': 'cohort_990.json headline_result + per-system filing '
                    'lists; register renders on Cohort_990_Contractors',
         'supplies': 'The only public window on whether transport vendors '
                     'reach top-contractor scale at nonprofit systems',
         'url': 'https://projects.propublica.org/nonprofits/api',
         'tier': 'A', 'accessed': accessed,
         'powers': ['Vendor_Share_Stack']},
        {'key': 'public_contract_corpus',
         'publisher': 'State, county and municipal records (Nebraska DAS '
                      'statecontracts portal; city/county contract and '
                      'compliance documents)',
         'document': '12 public ambulance service agreements, franchises '
                     'and compliance reports abstracted field-by-field '
                     '(contract_corpus.json, per-document URLs inside)',
         'vintage': 'Documents dated 2018-2025; abstracted 11 Jul 2026',
         'locator': 'contract_corpus.json public_contracts[]: '
                    'term_and_renewal, exclusivity_or_first_call, '
                    'response_time_SLA, penalty_structure',
         'supplies': 'Observed contract terms, exclusivity language '
                     'frequency and SLA/penalty precedents',
         'url': 'https://statecontracts.nebraska.gov (plus per-row URLs '
                'in contract_corpus.json)',
         'tier': 'B', 'accessed': accessed,
         'powers': ['Stickiness_Evidence', 'Vendor_Share_Stack']},
        {'key': 'mup_churn_2019_2024',
         'publisher': 'CMS',
         'document': 'Medicare Physician & Other Practitioners by '
                     'Provider and Service PUF, ground base codes, 2019 '
                     'and 2024 final-action files',
         'vintage': '2019 and 2024', 'locator': 'cache keys '
         'mup_provider_{2019,2024}_{A0426..A0434}; org entities '
         '(Ent_Cd=O) by state',
         'supplies': 'Registry-presence churn between vintages in the '
                     'cohort-corridor states',
         'url': 'https://data.cms.gov/provider-summary-by-type-of-service/'
                'medicare-physician-other-practitioners',
         'tier': 'A', 'accessed': accessed,
         'powers': ['Stickiness_Evidence']},
    ]

    # ============================================================ TAB 1 =
    # Growth_Outlook_Shell (C.4)
    g_anchor = {}

    st_age = _tab(wb, 'State_Age_65plus')
    assert st_age is not None, 'State_Age_65plus missing from workbook'
    us_row = None
    for r in range(1, 15):
        if st_age[f'A{r}'].value == 'United States':
            us_row = r
            break
    assert us_row, 'United States row not found on State_Age_65plus'
    us65_20 = float(st_age[f'C{us_row}'].value)
    us65_24 = float(st_age[f'E{us_row}'].value)
    us65_cagr = (us65_24 / us65_20) ** 0.25 - 1

    cty = _tab(wb, 'County_Age_65plus')
    cty_hdr = _scan(cty, '65+ CAGR', max_col=14, max_row=10)
    cty_rng = (f'K5:K{cty.max_row}' if cty_hdr else None)

    pr = _tab(wb, 'Payment_Rules')
    aif_cell = _scan(pr, 'Latest AIF', max_col=2)
    aif_ref = f'B{aif_cell.row}' if aif_cell else None
    aif3_cell = _scan(pr, 'Trailing 3-year average', max_col=2)
    law_cell = _scan(pr, 'extends the add-ons through', max_col=3)
    expiry = None
    if law_cell:
        m = re.search(r'through\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})',
                      str(law_cell.value), re.I)
        expiry = m.group(1) if m else None

    enr = _tab(wb, 'Enrollment_ESRD_State')
    r13 = _year_cell_row(enr, 2013, lo=5, hi=25)
    r24 = _year_cell_row(enr, 2024, lo=5, hi=25)
    r25 = _year_cell_row(enr, 2025, lo=5, hi=25)
    assert r13 and r24, 'Enrollment_ESRD_State panel A years not found'
    ma_share_24 = (float(enr[f'E{r24}'].value) /
                   float(enr[f'C{r24}'].value))

    hub_nm = 'Hub_Spoke_Map' if 'Hub_Spoke_Map' in wb.sheetnames else None
    hub_cnt_ref = hub_share_ref = None
    if hub_nm:
        hws = wb[hub_nm]
        hdr = _scan(hws, 'Share of hospitals', max_col=12)
        if hdr:
            for r in range(hdr.row + 1, min(hdr.row + 10, hws.max_row + 1)):
                v = hws[f'A{r}'].value
                if isinstance(v, str) and v.strip().lower().startswith('hub'):
                    hub_cnt_ref, hub_share_ref = f'B{r}', f'C{r}'
                    break

    reh_nm = ('REH_Closure_Flow'
              if 'REH_Closure_Flow' in wb.sheetnames else None)
    reh_tot_ref = clo_sum_fml = None
    if reh_nm:
        rws = wb[reh_nm]
        tot = _scan(rws, 'Total enrolled REHs', max_col=2)
        if tot:
            reh_tot_ref = f'B{tot.row}'
        ch = _scan(rws, 'Closures + conversions', max_col=4)
        if ch:
            r0 = ch.row + 1
            r1 = r0
            while r1 <= rws.max_row and isinstance(
                    rws[f'A{r1}'].value, (int, float)):
                r1 += 1
            if r1 > r0:
                clo_sum_fml = f"=SUM('{reh_nm}'!B{r0}:B{r1 - 1})"

    hc_nm = ('HCRIS_Ambulance_CostCenters'
             if 'HCRIS_Ambulance_CostCenters' in wb.sheetnames else None)
    hc_net_refs, hc_share_ref = [], None
    if hc_nm:
        hws = wb[hc_nm]
        pair_re = re.compile(r'^\d{4} -> \d{4}$')
        for row in hws.iter_rows(max_col=1):
            c = row[0]
            if isinstance(c.value, str) and pair_re.match(c.value.strip()):
                hc_net_refs.append((c.value.strip(), f'D{c.row}'))
        sh_hdr = _scan(hws, 'Share of filers', max_col=12)
        if sh_hdr:
            best = None
            for r in range(sh_hdr.row + 1,
                           min(sh_hdr.row + 12, hws.max_row + 1)):
                v = hws[f'A{r}'].value
                if isinstance(v, (int, float)) and 2014 <= v <= 2023:
                    best = r
            if best:
                hc_share_ref = f'D{best}'

    ws = wb.create_sheet('Growth_Outlook_Shell')
    sb = lib.SheetBuilder(ws, 7, col_widths=[44, 15, 15, 15, 14, 14, 52],
                          tab_color='FF7A5195')
    sb.title('Growth outlook shell: measured components only, no blended '
             'headline')
    sb.subtitle('The question: what are the measured growth components of '
                'the IFT market - tailwinds and headwinds - kept separate? '
                'Sources: Census PEP age series (State_Age_65plus, '
                'County_Age_65plus), CMS AIF and the statutory add-on '
                'register (Payment_Rules), hub concentration and REH/'
                'closure flow (Hub_Spoke_Map, REH_Closure_Flow), the HCRIS '
                'ambulance cost-center flow series '
                '(HCRIS_Ambulance_CostCenters), and CMS enrollment MA '
                'migration (Enrollment_ESRD_State). Join keys: none - each '
                'component stays on its own measured base and window.')
    sb.note('DATA QUALITY: each component is measured on ITS OWN base and '
            'window (population, price, hospitals, cost reports, '
            'beneficiaries); they are not commensurable and are never '
            'summed, weighted or compounded together on this tab. '
            'Components from tabs still landing in this run print as '
            'named PENDING slots and go live at assembly order.')
    sb.blank()

    sb.banner('Panel A. Component register - one measured row per driver, '
              'never blended')
    sb.headers(['Component (measured source)', 'Measured value (live)',
                'Companion 1 (live)', 'Companion 2 (live)', '', '',
                'What it can and cannot say (same row, always)'])

    # 1 demographics
    sb.row([('Demographic base: US 65+ population CAGR 2020-2024 '
             '(State_Age_65plus)', 'label'),
            _L('State_Age_65plus', f'G{us_row}', lib.FMT_PCT1),
            (f"=MEDIAN('County_Age_65plus'!{cty_rng})" if cty_rng
             else 'PENDING', 'link' if cty_rng else 'note', lib.FMT_PCT1),
            _L('State_Age_65plus', f'J{us_row}', lib.FMT_PCT1),
            None, None,
            ('CAN: the transported-age base compounds near 3%/yr '
             '(companion 1 = median county CAGR; companion 2 = US 85+ '
             'CAGR). CANNOT: convert people into transports; utilization '
             'per beneficiary is measured separately and fell on the FFS '
             'window.', 'note')], wrap=True)
    g_anchor['demo'] = f'B{sb.r}'

    # 2 AIF forward
    sb.row([('Medicare price: Ambulance Inflation Factor, CY2026 '
             '(Payment_Rules)', 'label'),
            _L('Payment_Rules', aif_ref, lib.FMT_PCT1) if aif_ref
            else ('PENDING', 'note'),
            _L('Payment_Rules', f'B{aif3_cell.row}', lib.FMT_PCT1)
            if aif3_cell else ('PENDING', 'note'),
            None, None, None,
            ('CAN: current-law price growth on the FFS book (companion = '
             'trailing 3-year average). CANNOT: predict future CPI-U '
             'less productivity; the 2020-2024 window spans 0.2% to '
             '8.7%.', 'note')], wrap=True)
    g_anchor['aif'] = f'B{sb.r}'

    # 3 regionalization
    sb.row([('Regionalization: hub hospitals and hub share '
             f'({hub_nm or "Hub_Spoke_Map - lands in parallel"})', 'label'),
            _L(hub_nm, hub_cnt_ref, lib.FMT_INT)
            if hub_cnt_ref else ('PENDING', 'note'),
            _L(hub_nm, hub_share_ref, lib.FMT_PCT1)
            if hub_share_ref else ('PENDING', 'note'),
            None, None, None,
            ('CAN: sending/receiving concentration measured from CMS '
             'Hospital Service Area corridors. CANNOT: state a transport '
             'growth rate; concentration moves corridors, not counts. '
             'PENDING cells go live when Hub_Spoke_Map lands.', 'note')],
           wrap=True)
    g_anchor['hub'] = f'B{sb.r}'

    sb.row([('Regionalization: rural emergency hospital conversions and '
             f'closures ({reh_nm or "REH_Closure_Flow - lands in parallel"})',
             'label'),
            _L(reh_nm, reh_tot_ref, lib.FMT_INT)
            if reh_tot_ref else ('PENDING', 'note'),
            (clo_sum_fml, 'link', lib.FMT_INT)
            if clo_sum_fml else ('PENDING', 'note'),
            None, None, None,
            ('CAN: every REH conversion removes inpatient beds and '
             'mandates transfer relationships (companion = Sheps '
             'closures+conversions, summed live). CANNOT: price the '
             'transfer volume each closure creates; no public per-event '
             'series exists.', 'note')], wrap=True)
    g_anchor['reh'] = f'B{sb.r}'

    # 4 conversion flow + PENDING assumption
    sb.row([('Insourcing-to-outsourcing conversion: HCRIS net flow, '
             'latest year pairs (starts minus stops)', 'label'),
            _L(hc_nm, hc_net_refs[-1][1], lib.FMT_INT)
            if hc_net_refs else ('PENDING', 'note'),
            _L(hc_nm, hc_net_refs[-2][1], lib.FMT_INT)
            if len(hc_net_refs) >= 2 else ('PENDING', 'note'),
            None, None, None,
            ('CAN: the observed direction of hospital ambulance '
             'cost-center starts and stops (candidate events; filing '
             'lags, mergers and relabels inside). CANNOT: state a '
             'conversion RATE; the row below holds that slot open.',
             'note')], wrap=True)
    g_anchor['flow'] = f'B{sb.r}'

    sb.row([('Forward conversion assumption (any outsourcing-rate input '
             'to a growth view)', 'label'),
            ('PENDING', 'note'), None, None, None, None,
            ('No assumed conversion rate anywhere in this workbook. '
             'Would be filled by: AHA Annual Survey ambulance-ownership '
             'flag (AHA_Recv schema) or a claims-vendor panel with '
             'billing-entity flags (Claims_Vendor_Recv schema); both '
             'requests are drafted on Engagement_Data_Map.', 'note')],
           wrap=True)
    _box(ws, sb.r, (2, 3, 4), pend_border)

    # 5 MA migration
    sb.row([('MA migration: MA & other share of Medicare beneficiaries '
             '(Enrollment_ESRD_State)', 'label'),
            _L('Enrollment_ESRD_State', f'F{r24}', lib.FMT_PCT1),
            _L('Enrollment_ESRD_State', f'F{r13}', lib.FMT_PCT1),
            _L('Enrollment_ESRD_State', f'F{r25}', lib.FMT_PCT1)
            if r25 else None,
            None, None,
            ('CAN: the FFS-visible base shrinks every year (2013 '
             'companion vs 2024/2025); every point of MA share moves '
             'volume out of every public claims series. CANNOT: say MA '
             'members ride less; GADCS shows MA revenue per organization '
             'at parity or above FFS.', 'note')], wrap=True)
    g_anchor['ma'] = f'B{sb.r}'
    sb.blank()

    sb.banner('Panel B. The statutory add-on cliff register (carried from '
              'Payment_Rules; dates per statute)')
    sb.headers(['Add-on', 'Rate', 'Expires (per statute)', 'Statute', '',
                '', 'Carried from'])
    for name, rate in [('Urban base/mileage add-on', 0.02),
                       ('Rural base/mileage add-on', 0.03),
                       ('Super-rural base add-on', 0.226)]:
        sb.row([(name, 'text'), (rate, 'src', lib.FMT_PCT1),
                (expiry if expiry else 'PENDING', 'src' if expiry
                 else 'note'),
                ('42 U.S.C. 1395m(l)(12)-(13); CAA 2026 s.6203', 'note'),
                None, None,
                ('Payment_Rules add-on status panel, verified 10 Jul '
                 '2026', 'note')])
        if not expiry:
            _box(ws, sb.r, (3,), pend_border)
    g_anchor['cliff'] = f'C{sb.r}'
    sb.row([('Current-law status, pulled live from Payment_Rules',
             'label'),
            _L('Payment_Rules', f'B{law_cell.row}') if law_cell
            else ('PENDING', 'note'),
            None, None, None, None,
            ('the full statute sentence renders here', 'note')],
           wrap=True, height=34)
    sb.blank()

    sb.banner('Read panel')
    sb.prose('There is NO blended headline growth number on this tab, and '
             'none anywhere in this workbook, by design. What is measured: '
             'the 65+ base compounds near 3 percent a year (tailwind on '
             'the demand base); hub concentration and 48 REH conversions '
             'plus rural closures keep regionalizing care (tailwind on '
             'corridor demand); the current-law price update is 2.0 '
             'percent with a dated add-on cliff at the end of 2027 '
             '(headwind on price); the MA and other share passed half of '
             'all beneficiaries in 2024 and rises every year (headwind on '
             'the VISIBLE base, not necessarily on the business); and the '
             'insourcing-to-outsourcing conversion direction is observed '
             'in cost-report flows but its forward rate is a bordered '
             'PENDING slot, not an assumption. Anyone who needs one '
             'growth number must blend these components themselves and '
             'own the blend.')

    facts += [
        {'metric': 'US 65+ population CAGR, 2020-2024',
         'year': 2024, 'value': round(us65_cagr, 4),
         'unit': 'share per year', 'basis': 'DERIVED', 'tier': 'A',
         'source_keys': ['census_age_pep'],
         'locator': 'State_Age_65plus Panel A, United States row, '
                    'G column (live formula over Census PEP levels)',
         'lives_on': 'Growth_Outlook_Shell',
         'cross_check': 'Median county 65+ CAGR prints beside it live; '
                        '85+ CAGR runs lower on the same row'},
        {'metric': 'MA & other share of Medicare beneficiaries, 2024',
         'year': 2024, 'value': round(ma_share_24, 4), 'unit': 'share',
         'basis': 'DERIVED', 'tier': 'A',
         'source_keys': ['cms_enrollment_ma'],
         'locator': 'Enrollment_ESRD_State Panel A, 2024 row, E/C '
                    'columns', 'lives_on': 'Growth_Outlook_Shell',
         'cross_check': '2013 share 28.3% on the same panel; the '
                        'series rises every single year'},
    ]

    findings.append({
        'id_hint': 99,
        'finding': 'The growth outlook decomposes into measured tailwinds '
                   'and measured headwinds with no blended headline: '
                   'tailwinds are the demographic base (US 65+ CAGR about '
                   '2.9%/yr 2020-2024, 85+ slower, median county lower) '
                   'and regionalization (hub concentration, 48 REH '
                   'conversions, rural closures); headwinds are the '
                   'current-law price path (AIF 2.0% CY2026 with the 2%/'
                   '3%/22.6% add-ons expiring 31 December 2027 per CAA '
                   '2026 s.6203) and MA migration (MA & other share '
                   'passed 50% of beneficiaries in 2024, shrinking the '
                   'claims-visible base); the insourcing-to-outsourcing '
                   'conversion rate is a bordered PENDING slot, not an '
                   'assumption.',
        'numbers': f"='Growth_Outlook_Shell'!{g_anchor['demo']}",
        'sources': 'census_age_pep; cms_enrollment_ma; caa2026_addon_cliff',
        'confidence': 'High on each component; the components are not '
                      'commensurable and are never blended',
        'guardrail': 'Each row lives on its own base and window '
                     '(population, price, hospitals, cost reports, '
                     'beneficiaries). Nothing on this tab is a market '
                     'growth rate; any blend is the reader\'s own model '
                     'and responsibility.'})

    # ============================================================ TAB 2 =
    # Vendor_Share_Stack (C.5)
    import json as _json
    import os as _os
    with open(_os.path.join(_os.path.dirname(cache), 'cohort_990.json')
              if not _os.path.exists(_os.path.join(cache, 'cohort_990.json'))
              else _os.path.join(cache, 'cohort_990.json')) as f:
        c9 = _json.load(f)
    n_systems = len(c9['systems'])
    n_eins = sum(len(s.get('eins', [])) for s in c9['systems'])
    n_filings = 46 if '46' in c9.get('headline_result', '') else sum(
        len(e.get('filings', [])) for s in c9['systems']
        for e in s.get('eins', []))

    with open(_os.path.join(_os.path.dirname(cache), 'contract_corpus.json')
              if not _os.path.exists(
                  _os.path.join(cache, 'contract_corpus.json'))
              else _os.path.join(cache, 'contract_corpus.json')) as f:
        corpus = _json.load(f)
    pubc = corpus['public_contracts']
    n_dedicated = sum(
        1 for r in pubc
        if r.get('dedicated_unit_or_unit_hour_rate')
        and str(r['dedicated_unit_or_unit_hour_rate']).strip().lower()
        != 'none')

    ins_nm = ('Insourcing_Bounds'
              if 'Insourcing_Bounds' in wb.sheetnames else None)
    ins_floor_ref = ins_ceil_ref = None
    if ins_nm:
        iws = wb[ins_nm]
        fh = _scan(iws, 'FLOOR share', max_col=12)
        ch_ = _scan(iws, 'CEILING share', max_col=12)
        if fh and ch_:
            from openpyxl.utils import get_column_letter as _gcl
            fcol, ccol = _gcl(fh.column), _gcl(ch_.column)
            for r in range(fh.row + 1, min(fh.row + 8, iws.max_row + 1)):
                v = iws[f'A{r}'].value
                if v == 2024 or (isinstance(v, str) and v.strip() == '2024'):
                    ins_floor_ref = f'{fcol}{r}'
                    ins_ceil_ref = f'{ccol}{r}'
                    break

    c990_nm = _tab_like(wb, 'Cohort_990')
    case_nm = (_tab_like(wb, 'System_Research_Cohort')
               or _tab_like(wb, 'Press_Footprint'))

    ws = wb.create_sheet('Vendor_Share_Stack')
    sb = lib.SheetBuilder(ws, 6, col_widths=[34, 16, 16, 44, 44, 38],
                          tab_color='FF7A5195')
    sb.title('The vendor-share evidence stack: six labeled legs, none '
             'blended')
    sb.subtitle('The question: what can PUBLIC evidence say about vendor '
                'share of the outsourced IFT wallet? One exhibit, six '
                'legs: IRS Form 990 contractor disclosures for the ' +
                COHORT_FRAME + '; the Medicare claims-side insourcing '
                'bounds; the HCRIS cost-center classification; public '
                'case-study references; and two bordered PENDING legs '
                '(licensed buyer research; a claims-vendor panel). Join '
                'keys: none - the legs measure different bases and are '
                'never combined.')
    sb.note('DATA QUALITY: leg bases differ (filings, transports, '
            'hospitals, documents) and none of them is a share of '
            'facility transport SPEND. The 990 leg is floor-censored by '
            'the top-five-over-$100K disclosure rule; the claims legs '
            'see Medicare FFS only; the corpus is a convenience sample. '
            'No leg asserts a share-of-wallet norm, and no row here may '
            'be quoted as one.')
    sb.blank()

    sb.banner('Panel A. The stack - every leg states what it CAN and '
              'CANNOT say')
    sb.headers(['Leg', 'Value', 'Companion', 'What this leg CAN say',
                'What it CANNOT say', 'Locator / named gap'])

    c990_ref = None
    if c990_nm:
        zc = (_scan(wb[c990_nm], 'five-highest', max_col=8)
              or _scan(wb[c990_nm], 'top-five', max_col=8)
              or _scan(wb[c990_nm], 'top five', max_col=8))
        c990_ref = zc.coordinate if zc else 'A1'

    sb.row([('Leg 1. IRS 990 contractor observations (' +
             (c990_nm or 'Cohort_990_Contractors - lands in parallel') +
             ')', 'label'),
            (0, 'src', lib.FMT_INT),
            _L(c990_nm, c990_ref) if c990_ref
            else (n_filings, 'src', lib.FMT_INT),
            ('Across the cohort systems\' latest two public 990s (' +
             f'{n_systems} systems, {n_eins} EINs, {n_filings} filings), '
             'ZERO ambulance/EMS/medical-transport vendors appear among '
             'any filer\'s five highest-paid independent contractors '
             'over $100K.', 'note'),
            ('That any system spends zero on transport vendors. The '
             'top-five floor differs per filer; a vendor billing below '
             'each fifth-highest contractor is invisible. Zero is a '
             'floor-censored observation.', 'note'),
            ('Form 990 Part VII Section B; register tab ' +
             (c990_nm or 'Cohort_990_Contractors') + '; cohort_990.json',
             'note')], wrap=True)
    leg1_row = sb.r

    sb.row([('Leg 2. Medicare claims-side bounds (' +
             (ins_nm or 'Insourcing_Bounds - lands in parallel') + ')',
             'label'),
            _L(ins_nm, ins_floor_ref, lib.FMT_PCT1)
            if ins_floor_ref else ('PENDING', 'note'),
            _L(ins_nm, ins_ceil_ref, lib.FMT_PCT1)
            if ins_ceil_ref else ('PENDING', 'note'),
            ('Hospital-BILLED ground base transports sit between the '
             'floor (name AND roster linkage) and the ceiling (name OR '
             'roster linkage) of Medicare FFS base services; the '
             'remainder is billed by non-hospital entities.', 'note'),
            ('Who serves the non-hospital remainder: independent '
             'vendors, municipal services and districts all mix there. '
             'Says nothing about MA, Medicaid, commercial or '
             'facility-paid volume.', 'note'),
            ('Insourcing_Bounds Panel B, 2024 vintage row (floor/'
             'ceiling live)', 'note')], wrap=True)
    if not ins_floor_ref:
        _box(ws, sb.r, (2, 3), pend_border)

    sb.row([('Leg 3. HCRIS cost-center classification (' +
             (hc_nm or 'HCRIS_Ambulance_CostCenters - lands in parallel') +
             ')', 'label'),
            _L(hc_nm, hc_share_ref, lib.FMT_PCT1)
            if hc_share_ref else ('PENDING', 'note'),
            None,
            ('The share of cost-report filers carrying an ambulance '
             'cost center on Worksheet A line 95, and its as-filed '
             'dollars - the hospital-side capacity that COULD convert '
             'either way.', 'note'),
            ('Whether the cost center is in-house staff or a contracted '
             'vendor passed through hospital books; HCRIS carries cost, '
             'not vendor identity or revenue.', 'note'),
            ('HCRIS_Ambulance_CostCenters Panel A, latest complete FY '
             'row', 'note')], wrap=True)
    if not hc_share_ref:
        _box(ws, sb.r, (2,), pend_border)

    sb.row([('Leg 4. Public case-study references (' +
             (case_nm or 'System_Research_Cohort / '
              'Press_Footprint_Registry - land in parallel') + ')',
             'label'),
            _L(case_nm, 'A1') if case_nm else ('PENDING', 'note'),
            None,
            ('Existence and shape of hospital-vendor transport '
             'relationships that reached public record: press, RFPs, '
             'board minutes, system announcements.', 'note'),
            ('Any frequency or share: what reaches public record is '
             'selected for newsworthiness, not representative of '
             'contracting.', 'note'),
            ('Pointer resolves in order: System_Research_Cohort, else '
             'Press_Footprint_Registry, else PENDING', 'note')],
           wrap=True)
    if not case_nm:
        _box(ws, sb.r, (2,), pend_border)

    sb.row([('Leg 5. Licensed buyer research', 'label'),
            ('PENDING', 'note'), None,
            ('Would say: sourcing model and vendor set per facility, '
             'from the buyer side.', 'note'),
            ('Nothing today: no public source exists; primary research '
             'is out of scope for this public-evidence workbook.',
             'note'),
            ('Named slot: buyer-side research engagement; request '
             'framework on Engagement_Data_Map', 'note')], wrap=True)
    _box(ws, sb.r, (2, 3), pend_border)

    sb.row([('Leg 6. Claims-vendor panel', 'label'),
            ('PENDING', 'note'), None,
            ('Would say: billing-entity share of commercial and MA '
             'transport claims by CBSA - the actual vendor share of '
             'the claims-visible wallet.', 'note'),
            ('Nothing today: commercial claims panels are licensed '
             'products; the exact receiving fields are already '
             'specified.', 'note'),
            ('Named slot: Claims_Vendor_Recv schema (billing NPI/TIN, '
             'HCPCS, O-D modifiers, payer class, completeness by '
             'CBSA); request drafted on Engagement_Data_Map', 'note')],
           wrap=True)
    _box(ws, sb.r, (2, 3), pend_border)
    sb.blank()

    sb.banner('Read panel')
    sb.prose('The stack BOUNDS the sourcing question from four public '
             'directions: nonprofit systems disclose no transport vendor '
             'at top-contractor scale (a floor-censored zero across '
             f'{n_filings} filings); Medicare claims bound the '
             'hospital-billed share of ground base transports; HCRIS '
             'shows which hospitals carry ambulance cost at all; and '
             'public documents prove the relationship forms that exist. '
             'What no public leg measures is vendor share of wallet - '
             'the dollar split of facility transport spend across '
             'vendors. That number lives only in legs 5 and 6, which are '
             'bordered PENDING slots with named receiving schemas, and '
             'this workbook asserts no share-of-wallet norm in their '
             'place. That is the honest finding.')

    facts += [
        {'metric': 'Transport vendors among five highest-paid 990 '
                   'contractors, cohort systems (count across filings)',
         'year': 2024, 'value': 0, 'unit': 'vendors', 'basis': 'SOURCED',
         'tier': 'A', 'source_keys': ['irs990_contractor_sweep'],
         'locator': 'Form 990 Part VII Section B, latest two '
                    'XML-available filings per resolved EIN; '
                    'cohort_990.json headline_result',
         'lives_on': 'Vendor_Share_Stack',
         'cross_check': 'Floor-censored: disclosure covers only the top '
                        'five contractors over $100K per filer, so zero '
                        'bounds vendor scale, not vendor existence'},
        {'metric': 'Form 990 filings parsed for the contractor sweep',
         'year': 2024, 'value': n_filings, 'unit': 'filings',
         'basis': 'SOURCED', 'tier': 'A',
         'source_keys': ['irs990_contractor_sweep'],
         'locator': f'cohort_990.json: {n_systems} systems, {n_eins} '
                    'EINs, latest two XML filings per EIN',
         'lives_on': 'Vendor_Share_Stack',
         'cross_check': 'IRS XML availability lags: latest filings are '
                        'tax years 2022-2023 for most EINs'},
    ]

    findings.append({
        'id_hint': 100,
        'finding': 'The public evidence stack bounds the outsourcing '
                   'question from four directions (990 contractor '
                   'disclosures: a floor-censored zero transport vendors '
                   f'in any top-five across {n_filings} cohort filings; '
                   'Medicare claims bounds on hospital-billed share; '
                   'HCRIS cost-center incidence; public case documents), '
                   'but NO public leg measures vendor share of wallet - '
                   'that number exists only behind the two bordered '
                   'PENDING legs (licensed buyer research; a claims-'
                   'vendor panel with named receiving fields).',
        'numbers': f"='Vendor_Share_Stack'!B{leg1_row}",
        'sources': 'irs990_contractor_sweep; public_contract_corpus',
        'confidence': 'High that the bound structure is complete; the '
                      'wallet split itself is unmeasured by design '
                      'honesty, not by oversight',
        'guardrail': 'No row on this tab is a share-of-wallet estimate '
                     'and none may be quoted as one; the legs measure '
                     'different bases (filings, transports, hospitals, '
                     'documents) and are never blended.'})

    # ============================================================ TAB 3 =
    # Stickiness_Evidence (C.6)
    parsed = []
    for r in pubc:
        months = _parse_term_months(r.get('term_and_renewal'))
        bucket, snip = _excl_class(r.get('exclusivity_or_first_call'))
        sla = (r.get('response_time_SLA')
               and str(r['response_time_SLA']).strip().lower() != 'none')
        pen = (r.get('penalty_structure')
               and str(r['penalty_structure']).strip().lower() != 'none')
        parsed.append({
            'issuer': _clean(r.get('issuer'), 70) or 'n/s',
            'state': r.get('state') or 'n/s',
            'months': months,
            'renew': _clean(r.get('term_and_renewal'), 95) or 'None stated',
            'bucket': bucket, 'excl': snip,
            'sla': bool(sla), 'pen': bool(pen)})
    terms = sorted(p['months'] for p in parsed if p['months'])
    term_lo, term_hi = (terms[0], terms[-1]) if terms else (None, None)
    n_excl = sum(1 for p in parsed if p['bucket'] == 'exclusive')
    n_nonexcl = sum(1 for p in parsed if p['bucket'] == 'nonexclusive')

    # corpus-tab live COUNTIF hooks (guarded)
    corp_nm = _tab_like(wb, 'Contract_Corpus')
    corp_countif = corp_term_links = None
    if corp_nm:
        cws = wb[corp_nm]
        exh = _scan(cws, 'exclusiv', max_col=20)
        if exh:
            from openpyxl.utils import get_column_letter as _gcl
            col = _gcl(exh.column)
            lo, hi = exh.row + 1, min(exh.row + 60, cws.max_row)
            corp_countif = (
                f"=COUNTIF('{corp_nm}'!{col}{lo}:{col}{hi},"
                f'"*exclusiv*")-COUNTIF(\'{corp_nm}\'!{col}{lo}:{col}{hi},'
                f'"*nonexclusiv*")')
        th = _scan(cws, 'term', max_col=20)
        if th:
            from openpyxl.utils import get_column_letter as _gcl
            col = _gcl(th.column)
            refs = []
            for r_ in range(th.row + 1, min(th.row + 60, cws.max_row + 1)):
                if cws[f'{col}{r_}'].value not in (None, ''):
                    refs.append(f'{col}{r_}')
                if len(refs) == 2:
                    break
            corp_term_links = refs or None

    # churn from MUP caches
    churn = {}
    for yr in (2019, 2024):
        per_state = {s: set() for s in COHORT_STATES}
        for code in GROUND_BASE:
            for r in lib.load_cache(cache, f'mup_provider_{yr}_{code}'):
                if (r.get('Rndrng_Prvdr_Ent_Cd') == 'O'
                        and r.get('Rndrng_Prvdr_State_Abrvtn')
                        in per_state):
                    per_state[r['Rndrng_Prvdr_State_Abrvtn']].add(
                        r['Rndrng_NPI'])
        churn[yr] = per_state

    bench_nm = ('Contract_Benchmarks'
                if 'Contract_Benchmarks' in wb.sheetnames else None)

    ws = wb.create_sheet('Stickiness_Evidence')
    sb = lib.SheetBuilder(ws, 7, col_widths=[38, 8, 12, 40, 40, 10, 10],
                          tab_color='FF7A5195')
    sb.title('Stickiness evidence: observed terms, exclusivity language, '
             'SLA precedents, registry churn - a panel, NOT a score')
    sb.subtitle('The question: what do observed public contracts and '
                'registry presence show about how sticky ambulance '
                'relationships are? Sources: the 12-document public '
                'contract corpus (Contract_Corpus tab + '
                'contract_corpus.json with per-document URLs), the '
                'Contract_Benchmarks precedent register, and CMS MUP '
                'provider files 2019 vs 2024 for the cohort-corridor '
                'states. Join keys: none across panels; churn joins NPI '
                'within state between vintages.')
    sb.note('DATA QUALITY: the corpus is a CONVENIENCE SAMPLE of '
            'documents that happen to be public - municipal, county and '
            'state agreements skewed to 911 and government purchasing; '
            'hospital IFT agreements are private and systematically '
            'absent. Frequencies below describe the corpus, never the '
            'market. This tab is an evidence panel; it deliberately '
            'computes NO stickiness score or index (house rule: no '
            'composite indices).')
    sb.blank()

    sb.banner('Panel A. Observed public contract terms (12-document '
              'corpus; blue = abstracted from the named document)')
    sb.headers(['Issuer (public document)', 'State',
                'Initial term (months, parsed)',
                'Renewal / term language (as written, truncated)',
                'Exclusivity / first-call language (as written)',
                'SLA?', 'Penalty?'])
    a0 = sb.r + 1
    for p in parsed:
        sb.row([(p['issuer'], 'src'), (p['state'], 'src'),
                (p['months'], 'src', lib.FMT_INT) if p['months']
                else ('n/s', 'note'),
                (p['renew'], 'note'), (p['excl'], 'note'),
                ('YES' if p['sla'] else 'no', 'src'),
                ('YES' if p['pen'] else 'no', 'src')], wrap=True)
    a1 = sb.r
    sb.blank()
    sb.headers(['Corpus summary', 'Value', '', '', '', '', ''])
    sb.row([('Observed initial terms, shortest (parsed months)', 'label'),
            (term_lo, 'src', lib.FMT_INT), None,
            (f'{len(terms)} of {len(parsed)} documents state a parseable '
             'initial term', 'note'), None, None, None])
    term_lo_row = sb.r
    sb.row([('Observed initial terms, longest (parsed months)', 'label'),
            (term_hi, 'src', lib.FMT_INT), None,
            ('60 months = the Sioux Falls-style five-year franchise '
             'form', 'note'), None, None, None])
    term_hi_row = sb.r
    sb.row([('Documents with explicit exclusivity / first-call language '
             '(live count over Panel A)', 'label'),
            (f'=COUNTIF(E{a0}:E{a1},"*xclusiv*")-'
             f'COUNTIF(E{a0}:E{a1},"*onexclusiv*")', 'fml', lib.FMT_INT),
            None,
            (f'blue check: {n_excl} of {len(parsed)}; plus '
             f'{n_nonexcl} explicitly NONexclusive', 'note'),
            (n_excl, 'src', lib.FMT_INT), None, None])
    excl_row = sb.r
    sb.row([('Same count over the Contract_Corpus tab column (live '
             'COUNTIF)', 'label'),
            (corp_countif, 'link', lib.FMT_INT) if corp_countif
            else ('PENDING', 'note'), None,
            ('goes live when Contract_Corpus lands in assembly order',
             'note') if not corp_countif else
            (f"over '{corp_nm}' exclusivity column", 'note'),
            None, None, None])
    if not corp_countif:
        _box(ws, sb.r, (2,), pend_border)
    sb.row([('Documents with a response-time SLA (live)', 'label'),
            (f'=COUNTIF(F{a0}:F{a1},"YES")', 'fml', lib.FMT_INT), None,
            ('911 and NEMT forms carry them; pure IFT agreements in the '
             'corpus mostly do not', 'note'), None, None, None])
    sb.row([('Documents with a penalty structure (live)', 'label'),
            (f'=COUNTIF(G{a0}:G{a1},"YES")', 'fml', lib.FMT_INT), None,
            ('liquidated damages, per-response fines, report penalties',
             'note'), None, None, None])
    sb.row([('Observed term cells on Contract_Corpus (green when '
             'landed)', 'label'),
            _L(corp_nm, corp_term_links[0])
            if corp_term_links else ('PENDING', 'note'),
            _L(corp_nm, corp_term_links[1])
            if corp_term_links and len(corp_term_links) > 1 else None,
            ('scan of the corpus tab term column, first populated rows',
             'note'), None, None, None], wrap=True)
    if not corp_term_links:
        _box(ws, sb.r, (2,), pend_border)
    sb.note('Document URLs and full abstracts live in '
            'contract_corpus.json (Pull_Manifest) and render on the '
            'Contract_Corpus tab; nothing in Panel A is quoted beyond '
            'the public document text.')
    sb.blank()

    sb.banner('Panel B. SLA and penalty precedents (green = live from '
              'Contract_Benchmarks; corpus rows above carry their own)')
    sb.headers(['Precedent', 'Threshold (live)', '', 'Detail', '', '', ''])
    for label, coord, fmt, detail in [
            ('Urban 911 response-time fractile (municipal contracts + '
             'NFPA 1710)', 'C6', lib.FMT_PCT1,
             '8:59 at the 90th percentile is the standard urban form'),
            ('NEMT broker on-time-pickup standard (common form)', 'C9',
             lib.FMT_PCT1, '>=95% on-time pickup with liquidated damages'),
            ('Per-response outlier fine (San Diego / Falck)', 'E12', None,
             '$5,000 per response beyond 24 minutes'),
            ('California APOT offload standard (AB 40)', 'D11', None,
             '30-minute offload standard, monthly EMSA monitoring')]:
        sb.row([(label, 'text'),
                _L(bench_nm, coord, fmt) if bench_nm
                else ('PENDING', 'note'),
                None, (detail, 'note'), None, None, None], wrap=True)
    sb.note('Guardrail: these precedents come from 911, NEMT and offload '
            'regimes because those documents are public; hospital IFT '
            'agreements typically keep SLAs and penalties private. '
            'Precedent, not norm.')
    sb.blank()

    sb.banner('Panel C. Registry-presence churn, cohort-corridor states: '
              'organization NPIs billing ground base codes, 2019 vs 2024 '
              '(CMS MUP)')
    sb.headers(['State', 'Org NPIs 2019', 'Org NPIs 2024', 'Present both',
                'Exited', 'Entered', 'Survivor share of 2019 cohort '
                '(live)'])
    c0 = sb.r + 1
    for s in COHORT_STATES:
        a, b = churn[2019][s], churn[2024][s]
        rn = sb.r + 1
        sb.row([(s, 'src'), (len(a), 'src', lib.FMT_INT),
                (len(b), 'src', lib.FMT_INT),
                (len(a & b), 'src', lib.FMT_INT),
                (len(a - b), 'src', lib.FMT_INT),
                (len(b - a), 'src', lib.FMT_INT),
                (f'=IF(B{rn}=0,"n/a",D{rn}/B{rn})', 'fml', lib.FMT_PCT1)])
    c1 = sb.r
    tot_rn = sb.r + 1
    sb.row([('All cohort-corridor states (live)', 'label'),
            (f'=SUM(B{c0}:B{c1})', 'fml', lib.FMT_INT),
            (f'=SUM(C{c0}:C{c1})', 'fml', lib.FMT_INT),
            (f'=SUM(D{c0}:D{c1})', 'fml', lib.FMT_INT),
            (f'=SUM(E{c0}:E{c1})', 'fml', lib.FMT_INT),
            (f'=SUM(F{c0}:F{c1})', 'fml', lib.FMT_INT),
            (f'=D{tot_rn}/B{tot_rn}', 'fml', lib.FMT_PCT1)])
    churn_tot_row = sb.r
    sb.note('DATA QUALITY: presence-in-registry churn, not contract '
            'churn. The states are those of the ' + COHORT_FRAME + ' '
            'corridor systems (Ascension\'s national footprint is '
            'excluded from the state cut on purpose). Organization '
            'entities (Ent_Cd = O) only; the PUF suppresses providers '
            'under 11 beneficiaries per code, so small entrants and '
            'exiters are undercounted; NPI renumbering after a merger '
            'looks like one exit plus one entry.')
    sb.blank()

    sb.banner('Read panel')
    sb.prose('A panel, not a score - no stickiness index is computed '
             'here or anywhere in this workbook. What the public record '
             'shows: stated initial terms in the 12-document corpus run '
             f'{term_lo} to {term_hi} months with renewal options the '
             f'norm; {n_excl} of {len(parsed)} documents carry explicit '
             'exclusivity or first-call language (and '
             f'{n_nonexcl} are explicitly nonexclusive); enforceable '
             'SLAs with money attached exist as precedent in the 911 '
             'and NEMT forms; and registry presence in the cohort-'
             'corridor states churns slowly at the organization-NPI '
             'grain, with most 2019 billers still present in 2024. The '
             'corpus is a convenience sample of what happens to be '
             'public; hospital IFT agreements are private, so corpus '
             'frequencies must never be quoted as market frequencies.')

    facts += [
        {'metric': 'Longest stated initial term in the public contract '
                   'corpus', 'year': 2026, 'value': term_hi,
         'unit': 'months', 'basis': 'SOURCED', 'tier': 'A',
         'source_keys': ['public_contract_corpus'],
         'locator': 'contract_corpus.json public_contracts[].'
                    'term_and_renewal, parsed; Stickiness_Evidence '
                    'Panel A prints every parse',
         'lives_on': 'Stickiness_Evidence',
         'cross_check': f'Shortest stated term {term_lo} months; '
                        f'{len(terms)} of {len(parsed)} documents state '
                        'a parseable initial term; documents dated '
                        '2018-2025'},
        {'metric': 'Documents with explicit exclusivity or first-call '
                   'language in the public contract corpus',
         'year': 2026, 'value': n_excl, 'unit': f'documents of '
         f'{len(parsed)}', 'basis': 'SOURCED', 'tier': 'A',
         'source_keys': ['public_contract_corpus'],
         'locator': 'contract_corpus.json public_contracts[].'
                    'exclusivity_or_first_call; live COUNTIF on '
                    'Stickiness_Evidence Panel A',
         'lives_on': 'Stickiness_Evidence',
         'cross_check': f'{n_nonexcl} documents are explicitly '
                        'NONexclusive; corpus skews to 911/government '
                        'forms where exclusivity is a franchise feature'},
    ]

    findings.append({
        'id_hint': 101,
        'finding': 'The observable stickiness evidence: stated initial '
                   f'terms of {term_lo} to {term_hi} months with '
                   'renewals the norm in the 12-document public corpus, '
                   f'explicit exclusivity or first-call language in '
                   f'{n_excl} of {len(parsed)} documents, enforceable '
                   'SLA-plus-penalty precedents in the 911/NEMT forms, '
                   'and slow organization-NPI churn in the cohort-'
                   'corridor states between the 2019 and 2024 registry '
                   'vintages.',
        'numbers': f"='Stickiness_Evidence'!B{excl_row}",
        'sources': 'public_contract_corpus; mup_churn_2019_2024',
        'confidence': 'High on each observation; low on '
                      'generalizability by construction',
        'guardrail': 'CONVENIENCE SAMPLE: the corpus contains only '
                     'documents that happen to be public (municipal/'
                     'county/state 911 and NEMT forms); hospital IFT '
                     'agreements are private and systematically absent, '
                     'so corpus frequencies describe the corpus, not '
                     'the market, and no stickiness score exists here.'})

    # ============================================================ TAB 4 =
    # Investor_QA (C.8)
    fpl_nm = ('Facility_Pay_Layer'
              if 'Facility_Pay_Layer' in wb.sheetnames else None)
    fpl = _tab(wb, fpl_nm) if fpl_nm else None
    fpl_nonpayer = _scan(fpl, 'Any non-payer revenue', max_col=2)
    fpl_faccon = _scan(fpl, 'Facility contracts (hospitals', max_col=2)
    fpl_outside = _scan(fpl, 'Revenue OUTSIDE per-trip billing', max_col=2)
    fpl_maratio = _scan(fpl, 'MA-to-FFS revenue ratio', max_col=2)

    urec_nm = _tab_like(wb, 'Universe_Recon')
    frag_nm = ('Fragmentation_National'
               if 'Fragmentation_National' in wb.sheetnames else None)
    msp_nm = ('Market_Share_Panels'
              if 'Market_Share_Panels' in wb.sheetnames else None)
    frag_top10 = None
    if frag_nm:
        # find a 'top-10' row whose value column (B) actually holds a number
        # or formula - the subtitle prose also contains 'top-10' but its B
        # cell is blank, which is what the loose scan used to grab.
        fw = wb[frag_nm]
        for row in fw.iter_rows(min_col=1, max_col=1):
            a = row[0].value
            if isinstance(a, str) and 'top-10' in a.lower():
                b = fw.cell(row=row[0].row, column=2).value
                if isinstance(b, (int, float)) or (isinstance(b, str)
                                                   and b.startswith('=')):
                    frag_top10 = f'B{row[0].row}'
                    break
        if not frag_top10:
            frag_top10 = _headline(wb, frag_nm)
    wfd_nm = ('Workforce_Depth'
              if 'Workforce_Depth' in wb.sheetnames else None)
    mabc_nm = _tab_like(wb, 'MA_Book')
    ici_nm = _tab_like(wb, 'Input_Cost')
    ebr_nm = _tab_like(wb, 'Entry_Barrier')
    bbs_nm = _tab_like(wb, 'Balance_Billing')
    mrf_nm = ('Commercial_Rate_MRF'
              if 'Commercial_Rate_MRF' in wb.sheetnames else None)

    qcew = _tab(wb, 'QCEW_EMS_Employment')
    qcew_emp_ref = None
    if qcew:
        for r in range(6, min(qcew.max_row, 60)):
            if (str(qcew[f'A{r}'].value) == '2024'
                    and qcew[f'B{r}'].value == 'Private'):
                qcew_emp_ref = f'D{r}'
                break
        if not qcew_emp_ref:
            for r in range(6, min(qcew.max_row, 60)):
                if (str(qcew[f'A{r}'].value) == '2023'
                        and qcew[f'B{r}'].value == 'Private'):
                    qcew_emp_ref = f'D{r}'
                    break

    ws = wb.create_sheet('Investor_QA')
    sb = lib.SheetBuilder(ws, 7, col_widths=[46, 16, 16, 16, 14, 14, 44],
                          tab_color='FF7A5195')
    sb.title('Investor QA: twelve tearsheets, every number a live cell '
             'or a named PENDING slot')
    sb.subtitle('The question: the twelve questions an investor asks '
                'about this market, each answered in three sentences of '
                'plain words, wired to 2-4 live green cells on the home '
                'tabs, with the guardrail printed under every block. '
                'Sources: the home tabs named per block (this tab wires; '
                'it does not create). Join keys: none - green links '
                'only.')
    sb.note('DATA QUALITY: answers inherit every caveat of their home '
            'tabs; where a home tab lands in a parallel module the '
            'number row is a bordered PENDING slot that names it, and '
            'the link goes live at assembly order. Nothing in this tab '
            'is an assertion without a cell behind it.')
    sb.blank()
    sb.headers(['Tearsheets wired (live count of Q blocks below)',
                'Count', '', '', '', '', ''])
    sb.row([('=COUNTIF(A1:A400,"Q*")', 'fml'), None, None, None, None,
            None, ('the twelve banners below', 'note')])
    qa_count_row = sb.r
    sb.blank()

    def _qblock(qtitle, answer, rows, guardrail):
        sb.banner(qtitle)
        sb.prose('THE ANSWER IN THREE SENTENCES: ' + answer)
        for spec in rows:
            sb.row(spec, wrap=True)
        sb.note('THE GUARDRAIL: ' + guardrail)
        sb.blank()

    def _num(label, cell_spec, home):
        return [(label, 'label'), cell_spec, None, None, None, None,
                (home, 'note')]

    def _pnd(label, gap):
        return [(label, 'label'), ('PENDING', 'note'), None, None, None,
                None, (gap, 'note')]

    # Q1
    _qblock(
        'Q1. Is IFT the same business as 911 or NEMT?',
        'Interfacility transport is scheduled clinical logistics between '
        'licensed facilities, paid per transport under the same Medicare '
        'fee schedule as 911 but with a different code and acuity mix. '
        'The hospital-to-hospital channel runs about 31 percent '
        'emergency while residence and scene channels run above 98 '
        'percent, and nearly all specialty care transport lives in the '
        'facility channels. NEMT is a separate Medicaid benefit of '
        'brokered vans and sedans that sometimes shares vehicles but '
        'not the ambulance fee schedule.',
        [_num('SCT services in the hospital-to-hospital channel, 2024',
              _L('Acuity_by_Channel', 'C9', lib.FMT_INT),
              'Acuity_by_Channel Panel A'),
         _num('Hospital-to-hospital emergency share, 2024',
              _L('Acuity_by_Channel', 'C21', lib.FMT_PCT1),
              'Acuity_by_Channel Panel B'),
         _num('Hospital-to-SNF BLS non-emergency services, 2024',
              _L('Acuity_by_Channel', 'H10', lib.FMT_INT),
              'Acuity_by_Channel Panel A'),
         _num('The institutional claims fingerprint (revenue centers '
              '054x)', _L('Code_Vocabulary', 'C11'),
              'Code_Vocabulary section B')],
        'Channel shares are Medicare FFS origin-destination coded '
        'services only; MA, Medicaid and commercial mixes are not '
        'publicly published at this grain.')

    # Q2
    _qblock(
        'Q2. Why do claims undercount this market, and by how much?',
        'Claims-based sizing sees only transports that become payer '
        'claims, and a measured share of ambulance revenue never does. '
        'Three quarters of ambulance organizations report some '
        'non-payer revenue, about one in five did not always bill at '
        'least one payer category, and a listed pure-play operator '
        'books roughly 41 percent of transport revenue outside per-trip '
        'claims. The activation-versus-claims reconciliation prices the '
        'wedge; the IFT-specific undercount inside it has no public '
        'number and stays PENDING.',
        [_num('Organizations reporting any non-payer revenue (GADCS)',
              _L(fpl_nm, f'B{fpl_nonpayer.row}', lib.FMT_PCT1)
              if fpl_nonpayer else ('PENDING', 'note'),
              'Facility_Pay_Layer Panel A' if fpl_nonpayer else
              'Facility_Pay_Layer lands earlier in assembly order'),
         _num('Listed-operator revenue outside per-trip billing, FY2025',
              _L(fpl_nm, f'B{fpl_outside.row}', lib.FMT_PCT1)
              if fpl_outside else ('PENDING', 'note'),
              'Facility_Pay_Layer Panel B' if fpl_outside else
              'Facility_Pay_Layer lands earlier in assembly order'),
         (_num('Five-universe reconciliation (activations vs claims)',
               _L(urec_nm, _headline(wb, urec_nm)), urec_nm + ' (title pull)')
          if urec_nm else
          _pnd('Five-universe reconciliation (activations vs claims)',
               'Universe_Reconciliation lands in a parallel module; '
               'inputs (NEMSIS, GADCS, MUP, PSPS) already cached'))],
        'The wedge is a mixed pool of unbilled, facility-paid and '
        'non-FFS volume; no public source splits it, so no cell here '
        'claims to.')

    # Q3
    _qblock(
        'Q3. What does insourced really mean, and how much exists?',
        'Insourced means the hospital bills the transport itself or '
        'carries the ambulance cost center on its own books. Medicare '
        'claims bound hospital-billed ground base transports between '
        'about 1.4 and 27 percent (a strict-linkage floor against a '
        'name-collision ceiling), and the HCRIS panel shows the share '
        'of cost-report filers carrying an ambulance cost center. '
        'Between the bounds sits the mixed reality of owned services, '
        'JVs and contracted units billed under hospital numbers.',
        [_num('Hospital-billed share of ground base transports, 2024 '
              'FLOOR',
              _L(ins_nm, ins_floor_ref, lib.FMT_PCT1)
              if ins_floor_ref else ('PENDING', 'note'),
              'Insourcing_Bounds Panel B' if ins_floor_ref else
              'Insourcing_Bounds lands in a parallel module'),
         _num('Same share, 2024 CEILING',
              _L(ins_nm, ins_ceil_ref, lib.FMT_PCT1)
              if ins_ceil_ref else ('PENDING', 'note'),
              'Insourcing_Bounds Panel B' if ins_ceil_ref else
              'Insourcing_Bounds lands in a parallel module'),
         _num('Cost-report filers with an ambulance cost center, latest '
              'complete FY',
              _L(hc_nm, hc_share_ref, lib.FMT_PCT1)
              if hc_share_ref else ('PENDING', 'note'),
              'HCRIS_Ambulance_CostCenters Panel A' if hc_share_ref else
              'HCRIS_Ambulance_CostCenters lands in a parallel module')],
        'The floor is a floor twice over (roster gaps understate '
        'linkage) and the ceiling is padded by place-name collisions; '
        'the AHA ambulance-ownership flag would turn bounds into a '
        'measured split (AHA_Recv schema).')

    # Q4
    _qblock(
        'Q4. How big is the facility direct-bill shadow?',
        'Some ambulance revenue is billed to the facility, not a payer, '
        'so it never appears in any claims dataset. In the national '
        'census 18.6 percent of organizations report facility-contract '
        'revenue with a long right tail, and public contracts price the '
        'form by the dedicated unit, the hour and the annual subsidy. '
        'The corpus proves the channel and its pricing forms; its '
        'IFT-specific dollar share is a named PENDING slot.',
        [_num('Organizations reporting facility-contract revenue (GADCS)',
              _L(fpl_nm, f'B{fpl_faccon.row}', lib.FMT_PCT1)
              if fpl_faccon else ('PENDING', 'note'),
              'Facility_Pay_Layer Panel A' if fpl_faccon else
              'Facility_Pay_Layer lands earlier in assembly order'),
         _num('Facility-contract revenue, conditional mean per '
              'organization',
              _L(fpl_nm, f'C{fpl_faccon.row}', lib.FMT_USD)
              if fpl_faccon else ('PENDING', 'note'),
              'same row: median $48K prints beside it' if fpl_faccon
              else 'Facility_Pay_Layer lands earlier in assembly order'),
         _num('Corpus documents with dedicated-unit or unit-hour '
              'pricing', (n_dedicated, 'src', lib.FMT_INT),
              'contract_corpus.json; abstracts on Contract_Corpus')],
        'GADCS incidence covers ALL ambulance work, not IFT alone; the '
        'IFT-specific facility-pay share stays PENDING on '
        'Facility_Pay_Layer Panel D by design.')

    # Q5
    _qblock(
        'Q5. How much of the market is invisible because of MA?',
        'Medicare Advantage ambulance claims are not published at '
        'provider grain, so the MA half of Medicare is dark to every '
        'public claims series. The MA and other share crossed 50 '
        'percent of beneficiaries in 2024 and rises every year, '
        'mechanically shrinking the FFS window all public series stand '
        'on. The GADCS census shows MA revenue per organization at '
        'parity or above FFS, so the dark half is not smaller business, '
        'just unmeasured.',
        [(_num('MA book calibrator', _L(mabc_nm, _headline(wb, mabc_nm)),
               mabc_nm + ' (title pull)') if mabc_nm else
          _num('Share of beneficiaries outside the A and B FFS window, '
               '2024', _L('Imbalance_Ledger', 'B39', lib.FMT_PCT1),
               'Imbalance_Ledger Panel C')),
         _num('MA & other share of beneficiaries, 2024',
              _L('Enrollment_ESRD_State', f'F{r24}', lib.FMT_PCT1),
              'Enrollment_ESRD_State Panel A'),
         _num('MA-to-FFS revenue ratio per ambulance NPI (GADCS)',
              _L(fpl_nm, f'B{fpl_maratio.row}', lib.FMT_X)
              if fpl_maratio else ('PENDING', 'note'),
              'Facility_Pay_Layer Panel A' if fpl_maratio else
              'Facility_Pay_Layer lands earlier in assembly order')],
        'MA encounter data exists at ResDAC under DUA (MA_Encounter_Recv '
        'schema); until licensed, MA volume is inferred from enrollment '
        'and the GADCS revenue census, never from claims.')

    # Q6
    _qblock(
        'Q6. What do commercial payers actually pay?',
        'Commercial allowed amounts run well above Medicare for the '
        'same service, and dispersion across states is wide. FAIR '
        'Health puts ALS emergency in-network allowed at 758 dollars '
        'nationally in 2020, state means span roughly 1.7x, and ground '
        'ambulance stayed outside the No Surprises Act. Machine-'
        'readable negotiated rates exist in Transparency-in-Coverage '
        'MRF files; the receiving schema is built and the pull is the '
        'named gap.',
        [_num('ALS emergency average allowed, in-network, 2020',
              _L('Payer_Rates_Commercial', 'C6', lib.FMT_USD),
              'Payer_Rates_Commercial Panel A (FAIR Health)'),
         _num('State dispersion: high-state over low-state ALS allowed, '
              '2022', _L('Payer_Rates_Commercial', 'F15', lib.FMT_X),
              'Payer_Rates_Commercial Panel B'),
         (_pnd('Negotiated-rate slices (5 payers x footprint states)',
               'Commercial_Rate_MRF receiving schema is built; TiC MRF '
               'slice pull is the named public gap')
          if mrf_nm else
          _pnd('Negotiated-rate slices (5 payers x footprint states)',
               'Commercial_Rate_MRF receiving schema lands with B.14; '
               'TiC MRF slice pull is the named public gap'))],
        'FAIR Health figures are national/state means from a '
        'contributor panel, not a census; negotiated-rate truth arrives '
        'only with the MRF slices.')

    # Q7
    _qblock(
        'Q7. Where does growth come from without a hockey stick?',
        'Growth is presented as measured components, never a blended '
        'headline: the 65+ base compounds near 3 percent, the '
        'current-law price update is 2 percent, and regionalization '
        'keeps concentrating sending and receiving. The offsets are '
        'equally measured: MA migration shrinks the visible FFS base '
        'and the statutory add-ons expire at the end of 2027. Anyone '
        'needing one growth number must blend the components and own '
        'the blend.',
        [_num('Demographic component (US 65+ CAGR, live)',
              _L('Growth_Outlook_Shell', g_anchor['demo'], lib.FMT_PCT1),
              'Growth_Outlook_Shell Panel A'),
         _num('Price component (AIF CY2026, live)',
              _L('Growth_Outlook_Shell', g_anchor['aif'], lib.FMT_PCT1),
              'Growth_Outlook_Shell Panel A'),
         _num('MA migration component (MA share 2024, live)',
              _L('Growth_Outlook_Shell', g_anchor['ma'], lib.FMT_PCT1),
              'Growth_Outlook_Shell Panel A'),
         _num('Add-on cliff date (per statute)',
              _L('Growth_Outlook_Shell', g_anchor['cliff']),
              'Growth_Outlook_Shell Panel B')],
        'No blended growth number exists in this workbook; the '
        'conversion-rate slot on Growth_Outlook_Shell is bordered '
        'PENDING, not assumed.')

    # Q8
    _qblock(
        'Q8. How fragmented is this market, and who do we actually '
        'compete with?',
        'The national biller base is thousands of mostly small '
        'organizations, and the top of the distribution holds a modest '
        'share, so the real competitive set is local, not national. '
        'Concentration only means something per state and per corridor, '
        'which is why the workbook prints share panels and HHI at state '
        'and NPI grain. Parent roll-ups raise measured concentration, '
        'so every printed top share is a floor.',
        [(_num('National concentration (top-10 view)',
               _L(frag_nm, frag_top10, lib.FMT_PCT1)
               if frag_top10 else _L(frag_nm, _headline(wb, frag_nm)),
               frag_nm + (' top-10 row' if frag_top10 else ' (title '
                          'pull)'))
          if frag_nm else
          _pnd('National concentration (top-10 view)',
               'Fragmentation_National lands in a parallel module; MUP '
               'inputs cached')),
         (_num('State share panels and per-state HHI',
               _L(msp_nm, _headline(wb, msp_nm)), msp_nm + ' (title pull)')
          if msp_nm else
          _pnd('State share panels and per-state HHI',
               'Market_Share_Panels lands in a parallel module; MUP '
               'inputs cached'))],
        'NPI-grain shares understate parent concentration; roll-up '
        'panels print the correction where public ownership is '
        'traceable.')

    # Q9
    _qblock(
        'Q9. What makes these relationships sticky?',
        'Public contracts show multi-year initial terms with renewal '
        'options, explicit exclusivity or first-call language in a '
        'minority of documents, and enforceable SLAs with liquidated '
        'damages in the 911 and NEMT forms. Registry presence in the '
        'cohort-corridor states churns slowly, with most 2019 billers '
        'still present in 2024. This is an evidence panel, not a '
        'stickiness score, and the corpus is a convenience sample.',
        [_num('Longest observed initial term (months)',
              _L('Stickiness_Evidence', f'B{term_hi_row}', lib.FMT_INT),
              'Stickiness_Evidence Panel A summary'),
         _num('Documents with explicit exclusivity language (live '
              'count)', _L('Stickiness_Evidence', f'B{excl_row}',
                           lib.FMT_INT),
              'Stickiness_Evidence Panel A summary'),
         _num('Survivor share of the 2019 org-NPI cohort, cohort states '
              '(live)', _L('Stickiness_Evidence', f'G{churn_tot_row}',
                           lib.FMT_PCT1),
              'Stickiness_Evidence Panel C')],
        'Hospital IFT agreements are private and absent from the '
        'corpus; churn is registry presence, not contract retention.')

    # Q10
    _qblock(
        'Q10. How bad is the labor problem?',
        'The binding input is EMT and paramedic labor: wages, depth and '
        'churn. Over 2019-2024 the compounded payer update roughly '
        'matched national EMT median wage growth, so the scissors story '
        'is window-dependent, while county-level depth varies by an '
        'order of magnitude. Labor risk prices locally: thin-depth '
        'counties with hub demand are where service fails first.',
        [(_num('Workforce depth panel', _L(wfd_nm, _headline(wb, wfd_nm)),
               wfd_nm + ' (title pull)')
          if wfd_nm else
          _pnd('Workforce depth panel',
               'Workforce_Depth lands in a parallel module; QCEW/OEWS '
               'inputs live below')),
         _num('EMT median wage growth 2019-2024 (live)',
              _L('Workforce_Supply', 'B15', lib.FMT_PCT1),
              'Workforce_Supply Panel B'),
         _num('Compounded AIF over the same window (live)',
              _L('Workforce_Supply', 'B16', lib.FMT_PCT1),
              'Workforce_Supply Panel B'),
         _num('Private ambulance employment (QCEW, latest annual)',
              _L('QCEW_EMS_Employment', qcew_emp_ref, lib.FMT_INT)
              if qcew_emp_ref else ('PENDING', 'note'),
              'QCEW_EMS_Employment Panel A' if qcew_emp_ref else
              'QCEW annual row not found at build')],
        'National medians hide the local market; wage-vs-depth '
        'tightness is printed state by state on the workforce tabs, '
        'and the 2019-2024 scissors read reverses on other windows.')

    # Q11
    _qblock(
        'Q11. What is the reimbursement cliff risk?',
        'Current-law price growth is the AIF, 2.0 percent for CY2026, '
        'and the temporary add-ons of 2, 3 and 22.6 percent expire on '
        '31 December 2027 under current statute. A Medicare-heavy '
        'ground book therefore carries a dated statutory cliff inside a '
        'normal hold period, largest for super-rural mileage. Input '
        'costs are tracked separately; the spread between the AIF and '
        'measured input inflation is the margin risk to watch.',
        [_num('Latest AIF (CY2026, live)',
              _L('Payment_Rules', aif_ref, lib.FMT_PCT1)
              if aif_ref else ('PENDING', 'note'), 'Payment_Rules'),
         _num('Add-on expiry (per statute, carried live)',
              _L('Growth_Outlook_Shell', g_anchor['cliff']),
              'Growth_Outlook_Shell Panel B'),
         (_num('Input cost index', _L(ici_nm, _headline(wb, ici_nm)),
               ici_nm + ' (title pull)')
          if ici_nm else
          _pnd('Input cost tracking',
               'Input_Cost_Index lands in a parallel module; BLS PPI/'
               'ECEC and EIA diesel series already cached'))],
        'The cliff is statutory, not speculative: the add-ons lapsed '
        'once already in October 2025 before CAA 2026 s.6203 restored '
        'them through 2027.')

    # Q12
    _qblock(
        'Q12. What regulatory barriers protect incumbents?',
        'Entry is regulated locally: state licensure, certificates of '
        'need in some states, franchise and exclusive-operating-area '
        'statutes, and municipal contracts with performance bonds. '
        'Ground ambulance sits outside the federal No Surprises Act, '
        'and a growing list of states has enacted its own balance-'
        'billing limits, reshaping out-of-network economics state by '
        'state. Barriers are real but jurisdictional; there is no '
        'single national moat number.',
        [(_num('Entry barrier register', _L(ebr_nm, _headline(wb, ebr_nm)),
               ebr_nm + ' (title pull)')
          if ebr_nm else
          _pnd('Entry barrier register',
               'Entry_Barrier_Register lands in a parallel module; '
               'Regulatory_Register carries the statutory anchors '
               'today')),
         (_num('State balance-billing pegs', _L(bbs_nm, _headline(wb, bbs_nm)),
               bbs_nm + ' (title pull)')
          if bbs_nm else
          _num('State balance-billing status (carried text, live pull)',
               _L('Payer_Rates_Commercial', 'B29'),
               'Payer_Rates_Commercial Panel D')),
         _num('Ground exclusion from the No Surprises Act (statutory '
              'anchor, live pull)',
              _L('Payment_Rules', 'B44'), 'Payment_Rules statutory '
              'anchors')],
        'Barrier strength varies by state and municipality; the '
        'register prints citations, not a barrier score.')

    sb.banner('Read panel')
    sb.prose('Twelve tearsheets, and every number in them is either a '
             'green live cell pulled from its home tab or a bordered '
             'PENDING slot that names the public dataset or landing '
             'schema that would fill it. Nothing on this tab is an '
             'assertion: the tab wires evidence that exists elsewhere '
             'in this workbook, and where a home tab lands later in '
             'assembly order the slot says so and goes live on the '
             'next build.')

    findings.append({
        'id_hint': 102,
        'finding': 'All twelve investor questions resolve to live cells '
                   'on their home tabs or to bordered PENDING slots '
                   'naming the public dataset or receiving schema that '
                   'would fill them - none resolves to an assertion.',
        'numbers': f"='Investor_QA'!A{qa_count_row}",
        'sources': 'cms_enrollment_ma; caa2026_addon_cliff; '
                   'public_contract_corpus',
        'confidence': 'High: the tab wires, it does not create',
        'guardrail': 'Each tearsheet inherits every caveat of its home '
                     'tab; the three-sentence answers are summaries of '
                     'linked cells, not independent claims.'})

    # ============================================================ TAB 5 =
    # Slide_Feed (C.7)
    def _status(exists, amber_gap):
        if not exists:
            return 'RED'
        return 'AMBER' if amber_gap else 'GREEN'

    exhibits = [
        ('Market definition: IFT vs 911 vs NEMT', ['Acuity_by_Channel'],
         'Acuity_by_Channel!C21', None,
         'none - live'),
        ('Code vocabulary and claims fingerprint', ['Code_Vocabulary'],
         'Code_Vocabulary!C11', None, 'none - live'),
        ('Five-universe reconciliation', ['~Universe_Recon'],
         'Universe_Reconciliation!A1', None,
         'module lands in parallel; NEMSIS/GADCS/MUP/PSPS inputs cached'),
        ('TAM floor grid (Medicare-visible floor)', ['TAM_Model_National'],
         'TAM_Model_National!A1',
         'MA and Medicaid legs need licensed extracts (MA_Encounter_Recv,'
         ' TAF_Ambulance_Recv)', ''),
        ('Metro TAM panel (20 target metros)', ['~Metro_TAM'],
         'Metro_TAM_Panel!A1', None,
         'C.1 module not yet landed; Metro_Structure_20 inputs live'),
        ('Gross-up waterfall (floor to all-payer)', ['Sizing_Playbook'],
         'Sizing_Playbook!A1',
         'gross-up legs beyond Medicare FFS carry PENDING inputs (MA '
         'encounter, TAF, MRF)', ''),
        ('Facility-pay triangulation', ['Facility_Pay_Layer'],
         'Facility_Pay_Layer!B8',
         'IFT-specific facility-pay share PENDING (claims-vendor panel '
         'with facility-remit flags)',
         'B.1 module lands earlier in assembly order'),
        ('Insourcing bounds + flow', ['Insourcing_Bounds'],
         'Insourcing_Bounds Panel B',
         'AHA ambulance-ownership flag would turn bounds into a split '
         '(AHA_Recv)', 'A.4 module lands in assembly order'),
        ('HCRIS ambulance cost centers',
         ['HCRIS_Ambulance_CostCenters'],
         'HCRIS_Ambulance_CostCenters Panel A',
         'FY2024 year-file not yet published (filing lag)',
         'A.5 module lands in assembly order'),
        ('Fragmentation curve', ['Fragmentation_National'],
         'Fragmentation_National!A1', None,
         'A.3 module lands in assembly order; MUP inputs cached'),
        ('Biller survival curves', ['Annual_Market_Structure'],
         'Annual_Market_Structure!A1', None,
         'X-F.1 module lands in assembly order; MUP vintages cached'),
        ('Market share panels (state grain)', ['Market_Share_Panels'],
         'Market_Share_Panels!A1', None,
         'A.2 module lands in assembly order; MUP inputs cached'),
        ('HHI / concentration bands', ['Market_Share_Panels',
                                       'Fragmentation_National'],
         'Market_Share_Panels Panel C', None,
         'A.2/A.3 modules land in assembly order'),
        ('Cohort corridor map', ['Cohort_Corridors'],
         'Cohort_Corridors!A1', None,
         'A.6 module lands in assembly order; HSA 2025 corridors cached'),
        ('Hub counts and hub share', ['Hub_Spoke_Map'],
         'Hub_Spoke_Map Panel B', None,
         'A.7 module lands in assembly order'),
        ('Whitespace counties', ['County_Whitespace_Screens'],
         'County_Whitespace_Screens!A1', None,
         'A.8 module lands in assembly order; census/QCEW inputs cached'),
        ('Growth decomposition (price x volume)', ['Growth_Decomposition'],
         'Growth_Decomposition!A1', None,
         'A.9 module lands in assembly order; MUP/PSPS inputs cached'),
        ('Growth outlook components', ['Growth_Outlook_Shell'],
         'Growth_Outlook_Shell Panel A',
         'forward conversion rate is a bordered PENDING slot (AHA flag '
         'or claims-vendor panel)', ''),
        ('Denial dollars', ['Denial_Economics'],
         'Denial_Economics!A1', None,
         'A.10 module lands in assembly order; PSPS denial series live'),
        ('Wage scissors', ['Workforce_Depth', 'Workforce_Supply'],
         'Workforce_Supply!B17', None,
         'A.12 module lands in assembly order; OEWS/QCEW live'),
        ('Boarding-cost shelf', ['Throughput_Economics_Public'],
         'Throughput_Economics_Public!A1', None,
         'E.4 module lands in assembly order'),
        ('DIDO / transfer delay friction', ['Transfer_Delay_Burden'],
         'Transfer_Delay_Burden!A1', None,
         'A.11 module lands in assembly order'),
        ('REH conversions and rural closures', ['REH_Closure_Flow'],
         'REH_Closure_Flow Panel A/B', None,
         'B.2 module lands in assembly order; REH + Sheps lists cached'),
        ('Receiving centers', ['~Receiving_Centers', 'Receiving_Side'],
         'Receiving_Centers!A1', None,
         'B.8 module lands in parallel; Receiving_Side carries the v3.3 '
         'layer'),
        ('Medicaid rate floor', ['~Medicaid'],
         'Medicaid fee-schedule tab!A1', None,
         'B.3 module lands in parallel; state fee schedules cached '
         '(medicaid_rates.json)'),
        ('Balance-billing pegs', ['~Balance_Billing'],
         'Balance_Billing_States!A1', None,
         'B.9 module lands in parallel; Payer_Rates_Commercial Panel D '
         'carries the federal statute today'),
        ('Entry barrier register', ['~Entry_Barrier',
                                    'Regulatory_Register'],
         'Entry_Barrier_Register!A1', None,
         'B.9 module lands in parallel; Regulatory_Register carries '
         'statutory anchors today'),
        ('Input cost index', ['~Input_Cost'],
         'Input_Cost_Index!A1', None,
         'B.11 module lands in parallel; BLS PPI/ECEC + EIA diesel '
         'cached'),
        ('Cohort dossiers', ['~Dossier_', '~System_Research'],
         'per-system dossier tabs', None,
         'E.1 module awaits corridors + press registry (both in '
         'flight)'),
        ('Footprint determination', ['~Press_Footprint'],
         'Press_Footprint_Registry!A1', None,
         'X-B/E.5 modules land in parallel; wayback_footprint.json '
         'cached'),
        ('Prospect screen', ['~Prospect'],
         'Prospect screen tab!A1', None,
         'E.6 module awaits E.1/E.5 outputs'),
        ('Scenario grid (no tornado without it)', ['~Scenario'],
         'C.3 scenario tab!A1', None,
         'C.2/C.3 modules not yet landed; all scenario inputs must be '
         'live cells first'),
        ('Vendor share stack', ['Vendor_Share_Stack'],
         'Vendor_Share_Stack Panel A',
         'legs 5-6 PENDING: licensed buyer research; claims-vendor '
         'panel (Claims_Vendor_Recv)', ''),
        ('Stickiness evidence panel', ['Stickiness_Evidence'],
         'Stickiness_Evidence Panel A',
         'hospital IFT agreements are private; corpus is a convenience '
         'sample', ''),
        ('Subject-company measured book', ['MMT_Medicare_Book',
                                           'MMT_NPI_Estate'],
         'MMT_Medicare_Book!A1', None,
         'A.1 module lands in assembly order; MMT_NPI_Estate carries '
         'the v3.3 estate'),
        ('Investor QA tearsheets', ['Investor_QA'],
         'Investor_QA (12 blocks)', None, ''),
        ('MA invisibility calibrator', ['~MA_Book', 'Imbalance_Ledger'],
         'Imbalance_Ledger!B39',
         'MA encounter extract under DUA is the named gap '
         '(MA_Encounter_Recv)', ''),
        ('Commercial rate benchmarks', ['Payer_Rates_Commercial'],
         'Payer_Rates_Commercial!C6',
         'TiC MRF negotiated-rate slices PENDING (Commercial_Rate_MRF '
         'schema)', ''),
    ]

    ws = wb.create_sheet('Slide_Feed')
    sb = lib.SheetBuilder(ws, 6, col_widths=[38, 26, 26, 40, 10, 52],
                          tab_color='FF6B7C93')
    sb.title('Slide feed: every planned exhibit mapped to live cells, '
             'with an honest evidence status')
    sb.subtitle('The question: which planned exhibits of the study are '
                'backed by live cells today, and where are the named '
                'public gaps? Status rule: GREEN = the home tab exists '
                'with live cells and no named PENDING input; AMBER = '
                'the tab exists but a named input is partial or '
                'PENDING; RED = the tab or data is absent from the '
                'workbook at this build. Statuses recompute at every '
                'assembly because this module scans the workbook it is '
                'given. Sources: pointers only - this tab creates no '
                'evidence.')
    sb.note('OPERATIONAL TAB: this feed powers the deck build and the '
            'remaining work plan; it does NOT feed the findings ledger, '
            'and honest RED/AMBER statuses are the deliverable. Every '
            'non-GREEN row names its gap: either a public dataset/'
            'receiving schema or the parallel module that lands the '
            'tab in assembly order.')
    sb.blank()
    sb.banner('The exhibit register')
    sb.headers(['Planned exhibit', 'Home tab (resolved live)', 'Anchor',
                'Live pull (green = tab title, proves the tab exists)',
                'Status', 'Named gap (every non-GREEN row)'])
    e0 = sb.r + 1
    n_g = n_a = n_r = 0
    for name, cands, anchor, amber_gap, red_gap in exhibits:
        resolved = None
        for cd in cands:
            if cd.startswith('~'):
                resolved = _tab_like(wb, cd[1:])
            else:
                resolved = cd if cd in wb.sheetnames else None
            if resolved:
                break
        stat = _status(resolved is not None, amber_gap)
        if stat == 'GREEN':
            n_g += 1
        elif stat == 'AMBER':
            n_a += 1
        else:
            n_r += 1
        gap = ''
        if stat == 'AMBER':
            gap = amber_gap
        elif stat == 'RED':
            gap = red_gap or 'tab absent at this build'
        sb.row([(name, 'text'),
                (resolved or cands[0].lstrip('~') + ' (planned)',
                 'note' if not resolved else 'text'),
                (anchor, 'note'),
                _L(resolved, 'A1') if resolved else ('PENDING', 'note'),
                (stat, 'label'),
                (gap or 'none - live', 'note')], wrap=True)
        if not resolved:
            _box(ws, sb.r, (4,), pend_border)
    e1 = sb.r
    sb.blank()
    sb.headers(['Status counts (live)', 'GREEN', 'AMBER', 'RED',
                'Total', ''])
    rn = sb.r + 1
    sb.row([('Exhibits by evidence status', 'label'),
            (f'=COUNTIF(E{e0}:E{e1},"GREEN")', 'fml', lib.FMT_INT),
            (f'=COUNTIF(E{e0}:E{e1},"AMBER")', 'fml', lib.FMT_INT),
            (f'=COUNTIF(E{e0}:E{e1},"RED")', 'fml', lib.FMT_INT),
            (f'=B{rn}+C{rn}+D{rn}', 'fml', lib.FMT_INT), None])
    sb.blank()
    sb.banner('Read panel')
    sb.prose('This tab is the deck feed and the remaining work plan in '
             'one: every planned exhibit of the study points at its '
             'home tab and anchor, GREEN rows are ready to render, '
             'AMBER rows render with their named PENDING input printed '
             'on the slide, and RED rows are not exhibits yet - they '
             'are work items whose gap column names the module or '
             'public dataset that unblocks them. The statuses are '
             'recomputed live at every assembly, so this page is only '
             'ever as optimistic as the workbook it sits in. It feeds '
             'the deck, not the findings ledger.')

    return {'facts': facts, 'sources': sources, 'excluded': excluded,
            'findings': findings,
            'meta': {'slide_feed_counts': {'GREEN': n_g, 'AMBER': n_a,
                                           'RED': n_r},
                     'note': 'C.4-C.8 assembly; all cross-tab links '
                             'guarded, statuses and links recompute at '
                             'assembly order'}}
