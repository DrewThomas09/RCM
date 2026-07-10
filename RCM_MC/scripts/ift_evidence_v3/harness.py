"""Test harness for v3 section-builder modules: python3 harness.py <key>"""
import importlib.util
import json
import os
import re
import sys

SCRATCH = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRATCH)
import v3lib  # noqa: E402

BANNED = [
    (re.compile(r'\$6\.5\s*B', re.I), 'illustrative $6.5B TAM'),
    (re.compile(r'\$18[-–]22\s*B', re.I), 'illustrative $18-22B market'),
    (re.compile(r'165\.8'), 'SAM/SOM 165.8x artifact'),
]


def main(key):
    path = os.path.join(SCRATCH, 'sections', f'sec_{key}.py')
    spec = importlib.util.spec_from_file_location(f'sec_{key}', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    wb.remove(wb.active)
    ctx = {'lib': v3lib, 'repo': '/home/user/RCM',
           'cache': os.path.join(SCRATCH, 'ift_v3_cache'), 'accessed': '10 Jul 2026'}
    out = mod.build(wb, ctx)

    problems, warnings = [], []
    declared = [s['name'] for s in mod.SHEETS]
    if list(wb.sheetnames) != declared:
        problems.append(f'sheetnames {wb.sheetnames} != declared {declared}')

    facts = out.get('facts', [])
    sources = out.get('sources', [])
    skeys = {s['key'] for s in sources}
    for i, f in enumerate(facts):
        for req in ('metric', 'basis', 'source_keys', 'locator', 'lives_on'):
            if not f.get(req):
                problems.append(f'fact[{i}] missing {req}: {f.get("metric")}')
        for sk in f.get('source_keys', []):
            if sk not in skeys:
                problems.append(f'fact[{i}] unknown source_key {sk!r}')
        ref = f.get('value_ref')
        if ref:
            m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", ref)
            if not m:
                problems.append(f'fact[{i}] bad value_ref {ref!r}')
            else:
                sh, col, row = m.groups()
                if sh not in wb.sheetnames:
                    problems.append(f'fact[{i}] value_ref sheet missing {sh!r}')
                elif wb[sh][f'{col}{row}'].value is None:
                    problems.append(f'fact[{i}] value_ref {ref} is EMPTY')
        elif f.get('value') is None:
            problems.append(f'fact[{i}] has neither value nor value_ref')
    for i, s in enumerate(sources):
        for req in ('key', 'publisher', 'document', 'vintage', 'tier', 'accessed', 'powers'):
            if not s.get(req):
                problems.append(f'source[{i}] {s.get("key")} missing {req}')

    n_charts = 0
    for name in wb.sheetnames:
        ws = wb[name]
        n_charts += len(ws._charts)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for rx, why in BANNED:
                        if rx.search(cell.value) and name != 'Excluded_Not_Sourced':
                            warnings.append(f'{name}!{cell.coordinate}: {why}: {cell.value[:60]!r}')
    outp = os.path.join(SCRATCH, f'test_{key}.xlsx')
    wb.calculation.fullCalcOnLoad = True
    wb.save(outp)
    load_workbook(outp)  # round-trip check

    rc = {n: wb[n].max_row for n in wb.sheetnames}
    print(json.dumps({'sheets': rc, 'facts': len(facts), 'sources': len(sources),
                      'excluded': len(out.get('excluded', [])), 'charts': n_charts,
                      'pages_est': sum(v3lib.estimate_print_pages(wb[n]) for n in wb.sheetnames),
                      'saved': outp}, indent=1))
    if warnings:
        print('WARNINGS:')
        [print('  -', w) for w in warnings]
    if problems:
        print('PROBLEMS:')
        [print('  -', p) for p in problems]
        sys.exit(1)
    print('OK')


if __name__ == '__main__':
    main(sys.argv[1])
