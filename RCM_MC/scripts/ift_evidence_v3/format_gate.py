"""Format gate (Run 3, Block U.6) - the committed CIM-presentation check.

Runs beside verify.py. Asserts the deck-ready standard on EVERY tab:
  - gridlines off, a tab colour set, freeze panes present;
  - title in A1, bold;
  - no numeric cell left on General format holding a raw float or a large
    non-year integer (the U.2 number-format rule);
  - no literal None / nan / NaN text (Python artifacts).
Writes format_gate.json and exits non-zero on any violation, so no revision
ships without the gate passing.
"""
import json
import os
import sys

from openpyxl import load_workbook

SCRATCH = os.path.dirname(os.path.abspath(__file__))
DEFAULT = os.path.join(SCRATCH, 'IFT_Sourced_Evidence_Master_v3_10.xlsx')
_NONE_TEXT = {'None', 'nan', 'NaN', 'NaT', 'NULL', 'null'}
_YEAR_MIN, _YEAR_MAX = 1990, 2035


def _is_year(v):
    return (isinstance(v, (int, float)) and not isinstance(v, bool)
            and float(v).is_integer() and _YEAR_MIN <= v <= _YEAR_MAX)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(path)
    v = {'gridlines_on': [], 'no_tabcolor': [], 'no_freeze': [],
         'no_title': [], 'unformatted_numeric': [], 'none_text': []}
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_view.showGridLines:
            v['gridlines_on'].append(name)
        tc = ws.sheet_properties.tabColor
        if not (tc and getattr(tc, 'rgb', None)):
            v['no_tabcolor'].append(name)
        if not ws.freeze_panes:
            v['no_freeze'].append(name)
        t = ws.cell(row=1, column=1).value
        if not (isinstance(t, str) and t.strip() and ws.cell(row=1, column=1).font.bold):
            v['no_title'].append(name)
        for row in ws.iter_rows():
            for c in row:
                val = c.value
                if isinstance(val, str) and val.strip() in _NONE_TEXT:
                    v['none_text'].append(f'{name}!{c.coordinate}')
                elif (isinstance(val, (int, float)) and not isinstance(val, bool)
                      and c.number_format in (None, 'General')
                      and not _is_year(val)
                      and (abs(val) >= 1000
                           or (isinstance(val, float) and val != int(val)))):
                    v['unformatted_numeric'].append(f'{name}!{c.coordinate}')
    counts = {k: len(x) for k, x in v.items()}
    clean = all(n == 0 for n in counts.values())
    result = {'clean': clean, 'counts': counts,
              'samples': {k: x[:8] for k, x in v.items() if x}}
    json.dump(result, open(os.path.join(os.path.dirname(path),
                                        'format_gate.json'), 'w'), indent=1)
    print(f'format gate: {len(wb.sheetnames)} tabs')
    for k, n in counts.items():
        print(f'  {k}: {n}')
    if not clean:
        for k, x in v.items():
            for s in x[:8]:
                print('   FLAG', k, s)
        sys.exit(1)
    print('FORMAT GATE PASS')


if __name__ == '__main__':
    main()
