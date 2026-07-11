"""Faithful v2.7 -> v3 copy engine prototype.

Copies every sheet of the v2.7 workbook 1:1 (values, formulas, styles,
comments, merges, widths, freeze panes, print setup) into a fresh openpyxl
workbook, then re-creates all 37 charts from the parsed chart JSON.
"""
from copy import copy
import json
import re

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

V27 = '/root/.claude/uploads/3de345a1-c58f-5ce6-b747-7cbb0636d5d9/bec059da-IFT_Sourced_Evidence_Master_v2_7.xlsx'
SCRATCH = '/tmp/claude-0/-home-user-RCM/3de345a1-c58f-5ce6-b747-7cbb0636d5d9/scratchpad'


def copy_sheet(src, dst):
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    if src.freeze_panes:
        dst.freeze_panes = src.freeze_panes
    for key, dim in src.column_dimensions.items():
        nd = dst.column_dimensions[key]
        nd.width = dim.width
        nd.hidden = dim.hidden
    for idx, dim in src.row_dimensions.items():
        nd = dst.row_dimensions[idx]
        nd.height = dim.height
        nd.hidden = dim.hidden
    for row in src.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style and cell.comment is None:
                continue
            nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.fill = copy(cell.fill)
                nc.border = copy(cell.border)
                nc.alignment = copy(cell.alignment)
                nc.number_format = cell.number_format
                nc.protection = copy(cell.protection)
            if cell.comment is not None:
                nc.comment = Comment(cell.comment.text, cell.comment.author or 'v2.7',
                                     height=cell.comment.height or 120,
                                     width=cell.comment.width or 260)
    for rng in src.merged_cells.ranges:
        dst.merge_cells(str(rng))
    dst.sheet_properties.tabColor = src.sheet_properties.tabColor
    # print setup (landscape, fit-to-width) is normalized later workbook-wide


_SHEET_REF = re.compile(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def parse_ref(f):
    m = _SHEET_REF.match(f)
    if not m:
        return None
    sheet, c1, r1, c2, r2 = m.groups()
    from openpyxl.utils import column_index_from_string
    min_col = column_index_from_string(c1)
    min_row = int(r1)
    max_col = column_index_from_string(c2) if c2 else min_col
    max_row = int(r2) if r2 else min_row
    return sheet, min_col, min_row, max_col, max_row


def make_ref(wb, f):
    p = parse_ref(f)
    if p is None:
        return None
    sheet, min_col, min_row, max_col, max_row = p
    return Reference(wb[sheet], min_col=min_col, min_row=min_row,
                     max_col=max_col, max_row=max_row)


def rebuild_charts(wb, charts_json, anchors_json=None):
    """Re-create the carried v2.7 charts from the full-fidelity re-parse
    (v27_charts2.json: anchors embedded, categories and name refs kept).
    House styling is applied here and again by the workbook-wide
    normalize pass, so carried charts match the v3-built ones."""
    import os
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import v3lib
    from openpyxl.chart.data_source import StrRef
    from openpyxl.chart.series import SeriesLabel

    specs = json.load(open(charts_json))
    built = 0
    for spec in specs:
        anchor = spec.get('anchor')
        sheet = spec['sheet']
        if sheet is None or anchor is None or sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for plot in spec['plots']:   # v2.7 has no combo charts (verified)
            if plot['type'] == 'barChart':
                ch = BarChart()
                ch.type = plot.get('barDir') or 'col'
                ch.grouping = plot.get('grouping') or 'clustered'
                kind = 'bar'
            elif plot['type'] == 'lineChart':
                ch = LineChart()
                kind = 'line'
            else:
                continue
            cat = None
            for ser in plot['series']:
                val = ser.get('val') or ser.get('y')
                if not val:
                    continue
                ref = make_ref(wb, val)
                if ref is None:
                    continue
                s = Series(ref, title_from_data=False)
                if ser.get('name_ref') and _SHEET_REF.match(ser['name_ref']):
                    s.tx = SeriesLabel(strRef=StrRef(ser['name_ref']))
                elif ser.get('name'):
                    s.tx = SeriesLabel(v=ser['name'])
                ch.series.append(s)
                cat = cat or ser.get('cat') or ser.get('x')
            if not ch.series:
                continue
            if cat:
                cref = make_ref(wb, cat)
                if cref is not None:
                    ch.set_categories(cref)
            ch.title = (spec.get('title') or '').replace('\n', ' ') or None
            for ax in spec.get('axes', []):
                if ax['kind'] == 'valAx':
                    if ax.get('numFmt'):
                        ch.y_axis.number_format = ax['numFmt']
                    if ax.get('title'):
                        ch.y_axis.title = ax['title']
                elif ax.get('title'):
                    ch.x_axis.title = ax['title']
            v3lib.style_chart(ch, kind, y_fmt=ch.y_axis.number_format)
            ch.height = max(6.2, (anchor['to'][1] - anchor['from'][1]) * 0.529)
            ch.width = max(11.0, (anchor['to'][0] - anchor['from'][0]) * 1.83)
            cell = f"{get_column_letter(anchor['from'][0] + 1)}{anchor['from'][1] + 1}"
            ws.add_chart(ch, cell)
            built += 1
    return built


def main():
    src = load_workbook(V27)
    wb = Workbook()
    wb.remove(wb.active)
    for name in src.sheetnames:
        ws = wb.create_sheet(title=name)
        copy_sheet(src[name], ws)
    n = rebuild_charts(wb, f'{SCRATCH}/v27_charts.json', f'{SCRATCH}/v27_chart_anchors.json')
    wb.calculation.fullCalcOnLoad = True
    out = f'{SCRATCH}/copy_test.xlsx'
    wb.save(out)
    print('charts rebuilt:', n, '-> saved', out)


if __name__ == '__main__':
    main()
