"""Deck feed extract (Run 4, outcome 6): a small standalone workbook holding
just the deck-facing tabs - Study_Synthesis, Slide_Feed, Investor_QA - with
every formula resolved to its computed VALUE, so a deck can be built from a
~50 KB file instead of the 31 MB master.

Reads the LibreOffice-recalced copy (data_only) that the verify gate produces
in recalc_out/, so the values are the same numbers the master recomputes to.
Preserves number formats, fonts, fills, alignment and column widths, and
carries a provenance banner naming the master build it was cut from.

Usage: python3 deck_extract.py <recalced_master.xlsx> <out.xlsx>
"""
import sys
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

DECK_TABS = ['Study_Synthesis', 'Slide_Feed', 'Investor_QA']
NAVY = 'FF00294C'


def _copy_sheet(src, dst):
    for col, dim in src.column_dimensions.items():
        if dim.width:
            dst.column_dimensions[col].width = dim.width
    for mc in src.merged_cells.ranges:
        dst.merge_cells(str(mc))
    for row in src.iter_rows():
        for c in row:
            if c.value is None:
                continue
            d = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                d.font = copy(c.font)
                d.fill = copy(c.fill)
                d.alignment = copy(c.alignment)
                d.number_format = c.number_format
    # match row heights so nothing clips
    for r, dim in src.row_dimensions.items():
        if dim.height:
            dst.row_dimensions[r].height = dim.height
    dst.sheet_view.showGridLines = False
    if src.sheet_properties.tabColor:
        dst.sheet_properties.tabColor = src.sheet_properties.tabColor
    dst.freeze_panes = 'A4'


def main():
    src_path = sys.argv[1]
    out_path = sys.argv[2]
    rc = load_workbook(src_path, data_only=True)
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet('Deck_Feed_Cover')
    cover.sheet_view.showGridLines = False
    cover.column_dimensions['A'].width = 118
    cover['A1'] = ('IFT deck feed extract: the three deck-facing tabs, values '
                   'resolved, cut from the verified master')
    cover['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
    cover['A3'] = ('What this is: a small standalone workbook holding only '
                   'Study_Synthesis (the thesis spine), Slide_Feed (every '
                   'planned exhibit with its live evidence status) and '
                   'Investor_QA (the twelve investor-question tearsheets). '
                   'Every formula has been resolved to the value the master '
                   'recomputes to, so a deck can be built from this file '
                   'without the 31 MB evidence master. Nothing here is a new '
                   'claim; each number traces to a tab in the master, and each '
                   'caveat travels with it.')
    cover['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    cover['A3'].font = Font(name='Arial', size=10)
    cover.row_dimensions[3].height = 90
    cover['A5'] = ('Source of truth: IFT_Sourced_Evidence_Master (the master '
                   'workbook and its Verification_Log). This extract is a '
                   'convenience cut, not the record; when they disagree, the '
                   'master governs.')
    cover['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    cover['A5'].font = Font(name='Arial', size=9, italic=True, color='FF5A5A5A')
    cover.row_dimensions[5].height = 40
    cover.sheet_properties.tabColor = NAVY

    kept = []
    for name in DECK_TABS:
        if name not in rc.sheetnames:
            continue
        dst = wb.create_sheet(name)
        _copy_sheet(rc[name], dst)
        kept.append(name)

    wb.active = 0
    wb.save(out_path)
    print(f'deck extract: {len(kept)} tabs {kept} -> {out_path}')


if __name__ == '__main__':
    main()
