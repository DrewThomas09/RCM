"""Render proof (Run 3, Block U.5): build a small proof workbook holding only a
representative sample of exhibit tabs (all others hidden so LibreOffice's PDF
export skips them), set landscape fit-to-width print areas, and emit it for
soffice --convert-to pdf. The rendered pages are then eyeballed for clipping,
overflow, stray artifacts and unreadable density.
"""
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties

SAMPLE = [
    'README', 'Study_Synthesis', 'Style_Standard', 'Methodology',
    'Macro_Demand_Drivers', 'Fragmentation_National',
    'MMT_Medicare_Book', 'SNF_ReturnLeg_Structure', 'SNF_Return_Leg_Quality',
    'Facility_Pay_Layer', 'Medicaid_Rate_Card', 'Scenario_Matrix',
    'TAM_Assembly_State', 'Fact_Ledger', 'Verification_Log',
    'Slide_Feed', 'Investor_QA',
    'Input_Cost_Index', 'Press_Footprint_Registry',
]


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else \
        'IFT_Sourced_Evidence_Master_v3_10.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else '_proof.xlsx'
    wb = load_workbook(src)
    keep = [n for n in SAMPLE if n in wb.sheetnames]
    for name in wb.sheetnames:
        ws = wb[name]
        if name in keep:
            ws.sheet_state = 'visible'
            # landscape, fit-to-width one page across
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.print_options.horizontalCentered = True
        else:
            ws.sheet_state = 'hidden'
    # a visible sheet must be active
    wb.active = wb.sheetnames.index(keep[0])
    wb.save(out)
    print(f'proof workbook: {len(keep)} visible / {len(wb.sheetnames)} total -> {out}')
    print('visible:', ', '.join(keep))


if __name__ == '__main__':
    main()
