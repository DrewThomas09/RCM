"""Presentation polish pass (v3.12): semantic status fills + zebra banding.

Pure formatting - it changes no cell value, so it is invisible to the V1
carried-cell fidelity gate, the recompute gate and the firewall leak check.
Idempotent: re-running produces the same styling.

  1. RAG status fills - every cell whose entire value is a GREEN / AMBER / RED /
     PENDING token gets the matching semantic fill and font. The match is on the
     whole stripped value, so legend sentences ("Status rule: GREEN = ...") are
     left alone. Existing borders are preserved, so the bordered-PENDING markers
     keep their border.
  2. Zebra banding - the marquee record-tables (Source_Register, Source_Index,
     Findings, Slide_Feed) get alternating row shading using the same grey already
     shipped on Fact_Ledger and Verification_Log, so the whole book bands the same
     colour. Banding is block-aware (it restarts under each navy panel header) and
     only touches cells that carry no fill of their own, so panel headers and the
     RAG fills stay intact. Prose tabs are excluded on purpose.
  3. Header font - navy and teal section-header cells are set white + bold so
     every table header reads the same.

Importable (polish(wb, log)) so assemble.py can call it after the professional-
tone pass; also runnable standalone on a built file for iteration.
"""
from openpyxl.styles import PatternFill, Font

BAND = 'FFF2F4F6'          # the banding grey already on Fact_Ledger / Verify log
HEADER_NAVY = 'FF00294C'   # navy panel header
HEADER_TEAL = 'FF1F6F8B'   # teal section sub-head
_HEADER_FILLS = (HEADER_NAVY, HEADER_TEAL)

# Semantic status palette - print-friendly, matches the workbook's RAG legend.
# (fill, font, force-bold)
_RAG = {
    'GREEN':   ('FFC6EFCE', 'FF1B6B3A', True),
    'AMBER':   ('FFFFE4B0', 'FF8A5000', True),
    'RED':     ('FFF7C9C4', 'FF9E1B14', True),
    'PENDING': ('FFE8EAED', 'FF5A5A5A', False),
}

# Clean one-record-per-row tables only. Prose tabs (Investor_QA) are excluded -
# banding narrative paragraphs reads as noise, not structure.
BAND_TABS = ('Source_Register', 'Source_Index', 'Findings', 'Slide_Feed')


def _fill_rgb(cell):
    f = cell.fill
    if f is not None and f.patternType:
        return getattr(f.fgColor, 'rgb', None)
    return None


def rag_fills(wb, log=None):
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                spec = _RAG.get(c.value.strip().upper())
                if spec is None:
                    continue
                fill, font_rgb, bold = spec
                c.fill = PatternFill('solid', fgColor=fill)
                f0 = c.font
                c.font = Font(name=f0.name, size=f0.size,
                              bold=bold or bool(f0.bold),
                              italic=f0.italic, color=font_rgb)
                n += 1
    if log:
        log(f'polish: {n} RAG status cells filled')
    return n


def band_tables(wb, log=None):
    n = 0
    band = PatternFill('solid', fgColor=BAND)
    for name in BAND_TABS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        maxc = ws.max_column
        seen_header = False
        stripe = 0
        for r in range(1, ws.max_row + 1):
            cells = [ws.cell(row=r, column=c) for c in range(1, maxc + 1)]
            if any(_fill_rgb(c) == HEADER_NAVY for c in cells):
                seen_header = True
                stripe = 0
                continue
            if not seen_header:
                continue                       # skip title / legend rows up top
            if all(c.value is None or (isinstance(c.value, str)
                                       and not c.value.strip()) for c in cells):
                stripe = 0                     # blank row ends a block
                continue
            stripe += 1
            if stripe % 2 == 0:                # band every other record row
                for c in cells:
                    if _fill_rgb(c) is None:    # never overwrite a header / RAG fill
                        c.fill = band
                        n += 1
    if log:
        log(f'polish: banded {n} cells across '
            f'{sum(1 for t in BAND_TABS if t in wb.sheetnames)} tables')
    return n


def header_fonts(wb, log=None):
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if _fill_rgb(c) in _HEADER_FILLS:
                    f0 = c.font
                    if f0.bold and getattr(f0.color, 'rgb', None) == 'FFFFFFFF':
                        continue
                    c.font = Font(name=f0.name, size=f0.size, bold=True,
                                  italic=f0.italic, color='FFFFFFFF')
                    n += 1
    if log:
        log(f'polish: {n} section-header cells set white + bold')
    return n


def polish(wb, log=None):
    r = rag_fills(wb, log=log)
    b = band_tables(wb, log=log)
    h = header_fonts(wb, log=log)
    return {'rag': r, 'band': b, 'header': h}


if __name__ == '__main__':
    import sys
    from openpyxl import load_workbook
    path = sys.argv[1]
    wb = load_workbook(path)
    res = polish(wb, log=print)
    wb.save(path)
    print('saved', path, res)
