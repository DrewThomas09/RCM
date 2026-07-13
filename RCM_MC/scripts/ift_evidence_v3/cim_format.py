"""CIM-grade presentation pass (Run 3, Block U).

Enforces the deck-ready standard on EVERY tab, including the carried v2.7 and
early-v3 tabs that were built before the SheetBuilder set gridlines-off /
tab-colour / freeze-panes:
  U.2 number formatting - no raw floats, no unformatted large integers, no
      literal None/nan; counts get thousands separators, the Fact_Ledger value
      column gets a value-preserving grouped format.
  U.3 layout - gridlines off, freeze panes below the title block, a section
      tab colour on every tab, cursor parked at A1, 100% zoom.
Idempotent: safe to run on an already-clean tab.

Importable (cim_pass(wb, log)) so assemble.py can call it at the end of the
build; also runnable standalone on a built file for iteration.
"""
import re

# Section tab colours (match the existing palette used by the SheetBuilder).
NAVY = 'FF00294C'        # governance / cover
DEMAND = 'FF1F6F8B'      # teal - demand evidence
SUPPLY = 'FF4C7C2C'      # green - supply evidence
ANALYSIS = 'FF6B7C93'    # slate - analysis
ASSEMBLY = 'FFC58F00'    # amber - assembly / TAM
REFERENCE = 'FF555555'   # grey - reference / audit

# Explicit colour assignment for the carried tabs that ship without one.
CARRIED_COLOUR = {
    'README': NAVY,
    'Engagement_Data_Map': REFERENCE, 'Dataset_Linkage_Map': REFERENCE,
    'Code_Crosswalks': REFERENCE, 'Fault_Audit': REFERENCE,
    'IFT_Alt_Measures': REFERENCE,
    'Macro_Demand_Drivers': DEMAND, 'Demand_Stack': DEMAND,
    'Condition_Transfer_Anchors': DEMAND, 'Medicare_OD_Matrix': DEMAND,
    'Utilization_Normalized': DEMAND, 'Acuity_by_Channel': DEMAND,
    'Air_Ambulance_IFT': DEMAND, 'Dialysis_ESRD_Channel': DEMAND,
    'Supplier_Landscape': SUPPLY, 'Supplier_Series_Raw': SUPPLY,
    'Supplier_Trend': SUPPLY, 'Workforce_Supply': SUPPLY, 'Supply_Stack': SUPPLY,
    'State_Saturation_Raw': SUPPLY, 'State_Saturation': SUPPLY,
    'MSA_Landscape': SUPPLY, 'Payer_Rates_Commercial': SUPPLY,
    'Imbalance_Ledger': ANALYSIS,
    'Sizing_Playbook': ASSEMBLY, 'TAM_Model_National': ASSEMBLY,
}

# Tabs whose numeric cells are years, not counts (no thousands separator).
_YEAR_MIN, _YEAR_MAX = 1990, 2035
FL_VALUE_FMT = '#,##0.####'    # value-preserving: grouping + up to 4 decimals
COUNT_FMT = '#,##0'
_NONE_TEXT = {'None', 'nan', 'NaN', 'NaT', 'none', 'NULL', 'null'}


def _looks_like_year(v):
    if isinstance(v, int) and _YEAR_MIN <= v <= _YEAR_MAX:
        return True
    if isinstance(v, float) and v.is_integer() and _YEAR_MIN <= v <= _YEAR_MAX:
        return True
    return False


def _title_block_end(ws):
    """Row index of the last title/subtitle/banner row before the data grid;
    freeze just below it. Heuristic: the header row is the first row whose
    cells are mostly short bold labels, else default to row 4."""
    # find first row after row 1 that has >=3 non-empty cells (a header row)
    for r in range(3, min(ws.max_row, 12) + 1):
        n = sum(1 for c in range(1, min(ws.max_column, 14) + 1)
                if ws.cell(row=r, column=c).value not in (None, ''))
        if n >= 3:
            return r
    return min(4, ws.max_row)


def cim_pass(wb, log=None):
    def _log(m):
        if log:
            log(m)
    n_fmt = n_none = n_grid = n_colour = n_freeze = 0
    for name in wb.sheetnames:
        ws = wb[name]
        # U.3 gridlines off
        if ws.sheet_view.showGridLines:
            ws.sheet_view.showGridLines = False
            n_grid += 1
        # U.3 tab colour
        tc = ws.sheet_properties.tabColor
        if not (tc and getattr(tc, 'rgb', None)):
            colour = CARRIED_COLOUR.get(name, ANALYSIS)
            ws.sheet_properties.tabColor = colour
            n_colour += 1
        # U.3 freeze panes below the title block
        if not ws.freeze_panes:
            end = _title_block_end(ws)
            ws.freeze_panes = f'A{end + 1}'
            n_freeze += 1
        # U.3 cursor to A1, 100% zoom
        ws.sheet_view.topLeftCell = 'A1'
        try:
            ws.sheet_view.selection[0].activeCell = 'A1'
            ws.sheet_view.selection[0].sqref = 'A1'
        except (IndexError, AttributeError):
            pass
        ws.sheet_view.zoomScale = 100
        ws.sheet_view.zoomScaleNormal = 100
        # U.2 number formats + None/nan cleanup
        is_fl = (name == 'Fact_Ledger')
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip() in _NONE_TEXT:
                    c.value = None
                    n_none += 1
                    continue
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                if c.number_format not in (None, 'General'):
                    continue
                if _looks_like_year(v):
                    continue                       # years stay bare
                if is_fl:
                    c.number_format = FL_VALUE_FMT
                    n_fmt += 1
                elif abs(v) >= 1000 or (isinstance(v, float) and v != int(v)):
                    c.number_format = (COUNT_FMT if float(v).is_integer()
                                       else FL_VALUE_FMT)
                    n_fmt += 1
    _log(f'cim pass: {n_grid} gridlines off, {n_colour} tab colours set, '
         f'{n_freeze} freeze panes, {n_fmt} number formats, '
         f'{n_none} None/nan cleared')
    return {'gridlines': n_grid, 'colours': n_colour, 'freeze': n_freeze,
            'formats': n_fmt, 'none_cleared': n_none}


if __name__ == '__main__':
    import sys
    from openpyxl import load_workbook
    path = sys.argv[1]
    wb = load_workbook(path)
    r = cim_pass(wb, log=print)
    wb.save(path)
    print('saved', path, r)
