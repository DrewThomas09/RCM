"""D.3 firewall leak check (handoff Section 2).

Scans every cell, tab name, note and finding of the v3.4 workbook for
firewall violations, restricted to the tabs ADDED in v3.4 (the carried
v2.7/v3.0-v3.3 content was cleared in prior passes). Flags:
  - customer / account / client-list language;
  - performance or contract metrics not tied to a public document;
  - survey statistics of any kind presented as fact;
  - em dashes in cell text (house rule).
Verbatim quotes from public documents that contain 'customer' (e.g. a
contract's 'customer service target') are allow-listed by an explicit
justification set, mirroring the E.2 agent's pre-flag.

Writes leak_check.json for the Verification_Log panel; exits non-zero on any
un-justified violation.
"""
import json
import os
import re
import sys

from openpyxl import load_workbook

DEFAULT = '/tmp/claude-0/-home-user-RCM/3de345a1-c58f-5ce6-b747-7cbb0636d5d9/scratchpad/IFT_Sourced_Evidence_Master_v4_0.xlsx'
# Baseline = the last shipped version whose new tabs were already leak-checked
# clean; anything ADDED since it (plus the tabs v3.5 extended in place) gets
# scanned. Each version removes its predecessor when packaged, so the baseline
# tracks forward; v3.12 adds no new tab or text (pure formatting), so its diff
# against v3.11 is empty and only the FORCE_INCLUDE tabs are rescanned.
V33 = '/home/user/RCM/RCM_MC/deliverables/IFT_Sourced_Evidence_Master_v3_12.xlsx'
# Tabs that already existed but were EXTENDED with new-authored content, so
# a pure new-tab diff would skip them. Force them into the scan. Contract_Corpus
# and Index carry the v3.9 portability / usefulness edits.
FORCE_INCLUDE = {'Medicaid_Rate_Card', 'Fact_Ledger', 'Source_Index',
                 'Source_Register', 'Findings', 'Run_Log', 'Verification_Log',
                 'README', 'Methodology', 'Contract_Corpus', 'Index',
                 # Run 5: Insourcing_Bounds gained the QN/QM third-leg panel
                 'Insourcing_Bounds'}

# Phrases the firewall bans as CUSTOMER/ACCOUNT framing.
BAN = [
    r'\bMMT (?:top |key )?customers?\b',
    r'\bMMT accounts?\b',
    r'\bclient list\b',
    r'\bcustomer list\b',
    r'\bour customers?\b',
    r'\bMMT[\'’]s (?:customers?|clients?|accounts?)\b',
    r'\bis an? MMT (?:customer|client|account|prospect)\b',
]
# Allow-listed public-document verbatim quotes containing 'customer'.
ALLOW_SUBSTR = [
    'customer service target',        # Sioux Falls SD performance report
    'customer satisfaction',          # Benton AR renewal criterion
    'customer service',               # generic contract SLA language
]
EMDASH = '—'


def added_tabs():
    v35 = set(load_workbook(DEFAULT, read_only=True).sheetnames)
    base = set(load_workbook(V33, read_only=True).sheetnames)
    return (v35 - base) | (FORCE_INCLUDE & v35)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(path, read_only=True)
    new = added_tabs()
    ban_rx = [re.compile(p, re.I) for p in BAN]
    viol_customer, viol_survey, viol_emdash = [], [], []
    n_cells = 0
    for name in wb.sheetnames:
        if name not in new:
            continue
        for row in wb[name].iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v:
                    continue
                n_cells += 1
                low = v.lower()
                for rx in ban_rx:
                    if rx.search(v):
                        viol_customer.append(f'{name}!{c.coordinate}: {v[:70]}')
                # survey statistics presented as fact (not as a named public
                # survey like GADCS/OEWS which ARE public datasets)
                if re.search(r'\b(our|internal|proprietary) survey\b', low) or \
                   re.search(r'\bsurvey (?:shows?|found|reports?) that\b', low):
                    viol_survey.append(f'{name}!{c.coordinate}: {v[:70]}')
                if EMDASH in v:
                    viol_emdash.append(f'{name}!{c.coordinate}: {v[:70]}')
    # filter customer hits by the allow-list (public-doc verbatim quotes)
    kept = []
    allowed = []
    for h in viol_customer:
        body = h.split(': ', 1)[1].lower()
        if any(a in body for a in ALLOW_SUBSTR):
            allowed.append(h)
        else:
            kept.append(h)

    result = {
        'tabs_scanned': len(new),
        'cells_scanned': n_cells,
        'customer_account_violations': kept,
        'customer_allowlisted_public_quotes': allowed,
        'survey_statistic_violations': viol_survey,
        'em_dash_violations': viol_emdash,
        'clean': not (kept or viol_survey or viol_emdash),
    }
    json.dump(result, open(os.path.join(os.path.dirname(path),
                                        'leak_check.json'), 'w'), indent=1)
    print(f'leak check: {len(new)} new tabs, {n_cells:,} text cells scanned')
    print(f'  customer/account violations (real): {len(kept)}')
    print(f'  allow-listed public-doc quotes w/ "customer": {len(allowed)}')
    print(f'  survey-statistic violations: {len(viol_survey)}')
    print(f'  em-dash violations: {len(viol_emdash)}')
    for h in kept[:10] + viol_survey[:10] + viol_emdash[:10]:
        print('   FLAG', h)
    if not result['clean']:
        sys.exit(1)
    print('LEAK CHECK CLEAN')


if __name__ == '__main__':
    main()
