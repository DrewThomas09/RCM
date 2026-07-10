"""House-style sheet builder for IFT Sourced Evidence Master v3.

Renders new v3 tabs in the exact v2.7 visual language:
  - title row: Arial 15 bold navy #00294C
  - subtitle row: Arial 9 grey #555555, wrapped
  - column headers: Arial 9 bold white on navy fill, medium bottom border
  - section banners: Arial 10 bold white on teal #1F6F8B
  - data cells: Arial 9; BLUE #0000FF = hardcoded from a source document,
    BLACK = Excel formula, GREEN #1F7A33 = link to another tab
  - thin bottom borders on data rows; landscape / fit-to-width print setup
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = 'FF00294C'
TEAL = 'FF1F6F8B'
GREY = 'FF555555'
BLUE = 'FF0000FF'   # hardcoded from source
BLACK = 'FF000000'  # formula
GREEN = 'FF1F7A33'  # cross-tab link
PURPLE = 'FF7A5195'

FMT_INT = '#,##0;\\(#,##0\\);\\-'
FMT_DEC1 = '#,##0.0;\\(#,##0.0\\);\\-'
FMT_DEC2 = '#,##0.00;\\(#,##0.00\\);\\-'
FMT_USD = '$#,##0;\\($#,##0\\);\\-'
FMT_USD2 = '$#,##0.00;\\($#,##0.00\\);\\-'
FMT_PCT1 = '0.0%;\\(0.0%\\);\\-'
FMT_PCT2 = '0.00%;\\(0.00%\\);\\-'
FMT_X = '0.00"x"'

_thin = Border(bottom=Side(style='thin', color='FFD9D9D9'))
_medium = Border(bottom=Side(style='medium', color=NAVY))

F_TITLE = Font(name='Arial', size=15, bold=True, color=NAVY)
F_SUB = Font(name='Arial', size=9, color=GREY)
F_HDR = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
F_BANNER = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
F_SRC = Font(name='Arial', size=9, color=BLUE)
F_FML = Font(name='Arial', size=9, color=BLACK)
F_LINK = Font(name='Arial', size=9, color=GREEN)
F_TXT = Font(name='Arial', size=9, color=BLACK)
F_LABEL = Font(name='Arial', size=9, bold=True, color=BLACK)
F_NOTE = Font(name='Arial', size=8, italic=True, color=GREY)

FILL_HDR = PatternFill('solid', fgColor=NAVY)
FILL_BANNER = PatternFill('solid', fgColor=TEAL)

AL_HDR = Alignment(horizontal='center', vertical='top', wrap_text=True)
AL_WRAP = Alignment(vertical='top', wrap_text=True)
AL_TOP = Alignment(vertical='top')


class SheetBuilder:
    """Row-cursor builder that writes v2.7-house-style content."""

    def __init__(self, ws, n_cols, col_widths=None, tab_color=None):
        self.ws = ws
        self.n = n_cols
        self.r = 0
        ws.sheet_view.showGridLines = False
        if col_widths:
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = 5

    def _row(self):
        self.r += 1
        return self.r

    def title(self, text):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_TITLE
        c.alignment = AL_TOP
        self.ws.row_dimensions[r].height = 24
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def subtitle(self, text, height=30):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_SUB
        c.alignment = AL_WRAP
        self.ws.row_dimensions[r].height = height
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def blank(self):
        r = self._row()
        self.ws.row_dimensions[r].height = 13.5
        return self

    def banner(self, text):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_BANNER
        c.fill = FILL_BANNER
        c.alignment = AL_TOP
        self.ws.row_dimensions[r].height = 16
        for i in range(2, self.n + 1):
            self.ws.cell(row=r, column=i).fill = FILL_BANNER
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def headers(self, cols, freeze=True, height=33.75):
        r = self._row()
        for i, h in enumerate(cols, start=1):
            c = self.ws.cell(row=r, column=i, value=h)
            c.font = F_HDR
            c.fill = FILL_HDR
            c.alignment = AL_HDR
            c.border = _medium
        for i in range(len(cols) + 1, self.n + 1):
            c = self.ws.cell(row=r, column=i)
            c.fill = FILL_HDR
            c.border = _medium
        self.ws.row_dimensions[r].height = height
        if freeze and self.ws.freeze_panes is None:
            self.ws.freeze_panes = f'A{r + 1}'
        return self

    def row(self, cells, height=None, wrap=False):
        """cells: list of (value, kind[, numfmt]) or plain values (kind='text').

        kinds: 'src' blue source-hardcoded | 'fml' black formula |
               'link' green cross-tab | 'text' | 'label' bold | 'note'
        """
        r = self._row()
        for i, spec in enumerate(cells, start=1):
            if spec is None:
                continue
            if not isinstance(spec, tuple):
                spec = (spec, 'text')
            value, kind = spec[0], spec[1]
            numfmt = spec[2] if len(spec) > 2 else None
            c = self.ws.cell(row=r, column=i, value=value)
            c.font = {'src': F_SRC, 'fml': F_FML, 'link': F_LINK,
                      'label': F_LABEL, 'note': F_NOTE}.get(kind, F_TXT)
            c.border = _thin
            c.alignment = AL_WRAP if wrap else AL_TOP
            if numfmt:
                c.number_format = numfmt
        if height:
            self.ws.row_dimensions[r].height = height
        elif wrap:
            self.ws.row_dimensions[r].height = 30
        return self

    def note(self, text, height=24):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_NOTE
        c.alignment = AL_WRAP
        self.ws.row_dimensions[r].height = height
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self


import re as _re

_REF = _re.compile(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def _mkref(wb, f):
    from openpyxl.chart import Reference
    from openpyxl.utils import column_index_from_string
    m = _REF.match(f)
    if not m:
        raise ValueError(f'bad ref {f!r}')
    sheet, c1, r1, c2, r2 = m.groups()
    return Reference(wb[sheet], min_col=column_index_from_string(c1), min_row=int(r1),
                     max_col=column_index_from_string(c2) if c2 else column_index_from_string(c1),
                     max_row=int(r2) if r2 else int(r1))


def add_chart(ws, anchor, title, cat_ref, series, kind='line', width=24, height=10,
              y_title=None, x_title=None, y_fmt=None, stacked=False, secondary=None):
    """Add a native chart whose series are live range references.

    series: list of (series_name:str, values_ref:'Sheet!$B$5:$B$20').
    secondary: optional list of (name, ref, kind) plotted on a right axis.
    """
    from openpyxl.chart import BarChart, LineChart
    from openpyxl.chart.data_source import StrRef
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.series_factory import SeriesFactory as Series

    wb = ws.parent

    def mk(k):
        if k == 'bar':
            ch = BarChart()
            ch.type = 'col'
            ch.grouping = 'stacked' if stacked else 'clustered'
            if stacked:
                ch.overlap = 100
        else:
            ch = LineChart()
        return ch

    main = mk(kind)
    for name, ref in series:
        s = Series(_mkref(wb, ref), title_from_data=False)
        s.tx = SeriesLabel(strRef=StrRef(name)) if _REF.match(name) else SeriesLabel(v=name)
        if kind == 'line':
            s.smooth = False
        main.series.append(s)
    if cat_ref:
        main.set_categories(_mkref(wb, cat_ref))
    main.title = title
    if y_title:
        main.y_axis.title = y_title
    if x_title:
        main.x_axis.title = x_title
    if y_fmt:
        main.y_axis.number_format = y_fmt
        main.y_axis.majorGridlines = None
    if secondary:
        sec = mk(secondary[0][2] if len(secondary[0]) > 2 else 'line')
        for item in secondary:
            name, ref = item[0], item[1]
            s = Series(_mkref(wb, ref), title_from_data=False)
            s.tx = SeriesLabel(strRef=StrRef(name)) if _REF.match(name) else SeriesLabel(v=name)
            sec.series.append(s)
        sec.y_axis.axId = 200
        sec.y_axis.crosses = 'max'
        main += sec
    main.width = width
    main.height = height
    ws.add_chart(main, anchor)
    return main


def cagr_formula(end_ref, start_ref, years):
    """CAGR as a live Excel formula string."""
    return f'=({end_ref}/{start_ref})^(1/{years})-1'


def load_cache(cache_dir, key):
    """Load a pull artifact; accepts plain .json (scratchpad) or .json.gz (repo)."""
    import gzip
    import json as _json
    import os as _os
    p = _os.path.join(cache_dir, key + '.json')
    if _os.path.exists(p):
        return _json.load(open(p))
    with gzip.open(p + '.gz', 'rt') as f:
        return _json.load(f)


def estimate_print_pages(ws, rows_per_page=42.0):
    """Printed-page estimate under landscape fit-to-width-1 setup.

    Counts weighted rows (taller rows consume more of a page) so dense wrapped
    sheets estimate honestly. A page at 100% zoom fits ~42 rows of 13.5pt;
    fit-to-width shrink adds rows per page for wide sheets, so this is a
    conservative (low) page estimate for narrow sheets and roughly right for
    wide ones.
    """
    used = 0.0
    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        h = ws.row_dimensions[r].height if r in ws.row_dimensions and ws.row_dimensions[r].height else 13.5
        used += h / 13.5
    if used == 0:
        return 0
    return max(1, round(used / rows_per_page + 0.499))
