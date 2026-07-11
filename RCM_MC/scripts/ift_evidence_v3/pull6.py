"""v3.1 pulls: Census state age estimates (65+/85+ denominators, closes the
state-age gap behind pending P20) and BLS OEWS May 2024 state wages for the
EMS occupations (closes pending P4).
"""
import csv
import io
import os
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pull import OPENER, get, load_manifest, log, record  # noqa: E402

CENSUS_URL = ('https://www2.census.gov/programs-surveys/popest/datasets/'
              '2020-2024/state/asrh/sc-est2024-agesex-civ.csv')
OEWS_URL = 'https://www.bls.gov/oes/special-requests/oesm24st.zip'
EMS_SOCS = {'29-2042': 'Emergency medical technicians',
            '29-2043': 'Paramedics',
            '53-3011': 'Ambulance drivers and attendants, except EMTs'}


def pull_census_age(man):
    key = 'census_state_age_2024'
    if key in man:
        return
    log('Census SC-EST2024-AGESEX-CIV state age estimates')
    raw = get(CENSUS_URL).decode('utf-8', 'replace')
    rdr = csv.DictReader(io.StringIO(raw))
    keep = []
    for r in rdr:
        # SEX 0 = total; keep total rows for every age (0-85, 85 = 85+) and
        # the AGE=999 total, at national (STATE 0) and state grain.
        if r.get('SEX') == '0':
            keep.append({k: r[k] for k in (
                'STATE', 'NAME', 'AGE', 'POPEST2020_CIV', 'POPEST2021_CIV',
                'POPEST2022_CIV', 'POPEST2023_CIV', 'POPEST2024_CIV')})
    record(man, key, keep, {
        'dataset': 'Census Bureau Vintage 2024 State Population Estimates by '
                   'Age and Sex, civilian (SC-EST2024-AGESEX-CIV)',
        'data_year': '2020-2024', 'endpoint': CENSUS_URL,
        'filters': {'client_side': 'SEX=0 (total) rows, all ages incl. 85 = 85+'},
        'rows': len(keep)})


def pull_oews(man):
    key = 'oews_ems_state_2024'
    if key in man:
        return
    log('BLS OEWS May 2024 state file (EMS occupations)')
    # bls.gov 403s the default urllib agent; send a browser-style UA via
    # opener addheaders (Request-object headers trip its filter).
    raw_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'ift_v3_cache', '_oews_raw.zip')
    if os.path.exists(raw_path):
        blob = open(raw_path, 'rb').read()
    else:
        import time
        blob = None
        OPENER.addheaders = [('User-Agent',
                              'Mozilla/5.0 (research; data pipeline)')]
        for i in range(5):
            try:
                blob = OPENER.open(OEWS_URL, timeout=300).read()
                break
            except Exception as e:  # noqa: BLE001
                log(f'  retry {i + 1}: {type(e).__name__}: {str(e)[:100]}')
                time.sleep(2 ** i)
        if blob is None:
            raise RuntimeError('OEWS download failed after retries')
    zf = zipfile.ZipFile(io.BytesIO(blob))
    xlsx_names = [n for n in zf.namelist() if n.lower().endswith('.xlsx')]
    log(f'  zip members: {xlsx_names}')
    from openpyxl import load_workbook
    keep = []
    cols = None
    for name in xlsx_names:
        wb = load_workbook(io.BytesIO(zf.read(name)), read_only=True)
        ws = wb[wb.sheetnames[0]]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().upper() if c else '' for c in row]
                idx = {h: i for i, h in enumerate(header)}
                occ_i = idx.get('OCC_CODE')
                if occ_i is None:
                    break
                cols = ['AREA', 'AREA_TITLE', 'PRIM_STATE', 'OCC_CODE',
                        'OCC_TITLE', 'TOT_EMP', 'H_MEAN', 'A_MEAN', 'H_MEDIAN',
                        'A_MEDIAN', 'A_PCT10', 'A_PCT90']
                continue
            occ = row[occ_i]
            if occ in EMS_SOCS:
                keep.append({c: (row[idx[c]] if c in idx and idx[c] < len(row)
                                 else None) for c in cols})
        wb.close()
    record(man, key, keep, {
        'dataset': 'BLS Occupational Employment and Wage Statistics (OEWS), '
                   'May 2024, state file — EMS occupations',
        'data_year': 'May 2024', 'endpoint': OEWS_URL,
        'filters': {'client_side': f'OCC_CODE in {sorted(EMS_SOCS)}'},
        'rows': len(keep),
        'note': 'Closes v2.7 pending register P4 (OEWS state wage files). '
                '* and # cells are BLS suppression/top-code markers, kept as '
                'published.'})


if __name__ == '__main__':
    man = load_manifest()
    pull_census_age(man)
    pull_oews(man)
    log('ALL DONE 6')
