"""Verification gates for the assembled v3 workbook.

Gate V1: carried v2.7 cells identical (recalc values vs v2.7 cached values),
         excluding rebuilt sheets (README, Source_Index), logged correction
         cells, and rows appended below the original extent.
Gate V2: whole-workbook LibreOffice recalc -> zero Excel error cells.
Gate V3: python recompute of derived cells on the pull-backed tabs.
Gate V7: printed-page estimate.
Gate V8: ledger/source integrity (IDs contiguous, tabs exist, counts true).

Writes verify_results.json for the second assemble pass (Panel K numbers).
"""
import json
import os
import subprocess
import sys

SCRATCH = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRATCH)
import v3lib  # noqa: E402

def _default(env, *candidates):
    v = os.environ.get(env)
    if v:
        return v
    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[-1]


_REPO_REF = '/home/user/RCM/RCM_MC/rcm_mc/market_reports/reference'
V27 = _default('IFT_V27_XLSX',
               '/root/.claude/uploads/3de345a1-c58f-5ce6-b747-7cbb0636d5d9/bec059da-IFT_Sourced_Evidence_Master_v2_7.xlsx',
               os.path.join(_REPO_REF, 'IFT_Sourced_Evidence_Master_v2_7.xlsx'))
V3 = _default('IFT_V3_OUT', os.path.join(SCRATCH, 'IFT_Sourced_Evidence_Master_v3_5.xlsx'))
CACHE = _default('IFT_V3_CACHE', os.path.join(SCRATCH, 'ift_v3_cache'),
                 os.path.join(_REPO_REF, 'ift_v3_cache'))
ERR = {'#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NULL!', '#NUM!'}
REBUILT = {'README', 'Source_Index', 'Methodology'}


def recalc(path):
    out = os.path.join(SCRATCH, 'recalc_out')
    target = os.path.join(out, os.path.basename(path))
    if os.path.exists(target):
        os.remove(target)
    subprocess.run(
        ['soffice', '--headless',
         f'-env:UserInstallation=file://{SCRATCH}/lo_profile/user',
         '--convert-to', 'xlsx', '--outdir', out, path],
        check=True, capture_output=True, timeout=1200)
    return target


def close(a, b):
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= max(1e-9, abs(a) * 1e-9)
    return a == b


def main():
    from openpyxl import load_workbook
    results = {}
    problems = []

    print('recalculating v3 in LibreOffice...', flush=True)
    rc_path = recalc(V3)
    rc = load_workbook(rc_path, data_only=True)
    v3 = load_workbook(V3)  # formulas
    orig = load_workbook(V27, data_only=True)
    orig_f = load_workbook(V27)

    # V2: zero error cells + formula count
    n_err, n_formula = 0, 0
    err_list = []
    for n in rc.sheetnames:
        for row in rc[n].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERR:
                    n_err += 1
                    err_list.append((n, c.coordinate, c.value))
    for n in v3.sheetnames:
        for row in v3[n].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    n_formula += 1
    results['n_formulas'] = f'{n_formula:,}'
    results['n_errors'] = n_err
    if n_err:
        problems.append(f'V2 FAIL: {n_err} error cells, first: {err_list[:10]}')

    # V1: carried-cell fidelity
    corrections = set()
    if 'V3_Change_Log' in v3.sheetnames:
        for row in v3['V3_Change_Log'].iter_rows(min_row=5):
            tab, cell = row[1].value, row[2].value
            if tab and cell and cell != '(sheet)':
                corrections.add((tab, cell))
    checked = diffs = 0
    for name in orig.sheetnames:
        if name in REBUILT or name not in rc.sheetnames:
            continue
        o = orig[name]
        r = rc[name]
        for row in o.iter_rows():
            for cell in row:
                if (name, cell.coordinate) in corrections:
                    continue
                ov = cell.value
                rv = r.cell(row=cell.row, column=cell.column).value
                if ov is None and rv is None:
                    continue
                checked += 1
                if not close(ov, rv):
                    # string cells: original formulas render as cached text; compare loosely
                    diffs += 1
                    if diffs <= 10:
                        problems.append(f'V1 diff {name}!{cell.coordinate}: '
                                        f'{ov!r} -> {rv!r}')
    results['copy_cells'] = f'{checked:,}'
    results['copy_diffs'] = diffs
    results['copy_errors'] = 0 if not n_err else n_err
    if diffs:
        problems.append(f'V1 FAIL: {diffs} carried-cell diffs')

    # V3: recompute audit on pull-backed tabs
    n_re, n_mm = 0, 0

    def approx(tab, coord, expected, label):
        nonlocal n_re, n_mm
        n_re += 1
        got = rc[tab][coord].value
        if not (isinstance(got, (int, float)) and isinstance(expected, (int, float))
                and abs(got - expected) <= max(1e-6, abs(expected) * 1e-6)):
            n_mm += 1
            problems.append(f'V3 mismatch {tab}!{coord} ({label}): '
                            f'expected {expected!r} got {got!r}')

    # PSPS denial rates: recompute 2024 + 2010 A0428 from cache
    for yr in (2010, 2017, 2024):
        if not (os.path.exists(os.path.join(CACHE, f'psps_agg_{yr}_A0428.json'))
                or os.path.exists(os.path.join(CACHE, f'psps_agg_{yr}_A0428.json.gz'))):
            continue
        d = v3lib.load_cache(CACHE, f'psps_agg_{yr}_A0428')
        sub = sum(a.get('SUBMITTED_SERVICE_CNT', 0) for a in d['by_initial_modifier'].values())
        den = sum(a.get('DENIED_SERVICES_CNT', 0) for a in d['by_initial_modifier'].values())
        ws = rc['PSPS_Denial_Series']
        found = False
        for row in ws.iter_rows():
            if row[0].value == yr and len(row) > 5 and row[1].value == 'A0428':
                approx('PSPS_Denial_Series', row[3].coordinate, sub, f'{yr} submitted')
                if sub:
                    approx('PSPS_Denial_Series', row[5].coordinate, den / sub,
                           f'{yr} denial rate')
                found = True
                break
        if not found:
            n_mm += 1
            problems.append(f'V3: PSPS {yr} A0428 row not found on tab')

    # Enrollment: national 2024 MA share
    nat = v3lib.load_cache(CACHE, 'enrollment_national_year')
    r24 = [r for r in nat if r['YEAR'] == '2024'][0]
    ma = float(r24['MA_AND_OTH_BENES']) / float(r24['TOT_BENES'])
    ws = rc['Enrollment_ESRD_State']
    for row in ws.iter_rows(max_row=25):
        if row[0].value == 2024:
            approx('Enrollment_ESRD_State', row[5].coordinate, ma, '2024 MA share')
            break

    # QCEW private employment CAGR
    q = {}
    for yr in range(2014, 2026):
        if (os.path.exists(os.path.join(CACHE, f'qcew_621910_{yr}.json'))
                or os.path.exists(os.path.join(CACHE, f'qcew_621910_{yr}.json.gz'))):
            for r in v3lib.load_cache(CACHE, f'qcew_621910_{yr}'):
                if r['area_fips'] == 'US000' and r['own_code'] == '5':
                    q[yr] = float(r['annual_avg_emplvl'])
    if len(q) >= 2:
        yrs = sorted(q)
        cagr = (q[yrs[-1]] / q[yrs[0]]) ** (1 / (len(yrs) - 1)) - 1
        ws = rc['QCEW_EMS_Employment']
        for row in ws.iter_rows():
            if row[0].value == 'CAGR (full window)':
                approx('QCEW_EMS_Employment', row[2].coordinate, cagr,
                       'QCEW employment CAGR')
                break

    # Facility universe: hospital total
    hosp = v3lib.load_cache(CACHE, 'pdc_hospitals')
    ws = rc['Facility_Universe_State']
    for row in ws.iter_rows():
        if row[0].value == 'US total':
            approx('Facility_Universe_State', row[1].coordinate,
                   sum(1 for r in hosp if (r.get('state') or '').strip()),
                   'hospital total')
            break

    results['n_recomputed'] = n_re
    results['n_mismatch'] = n_mm

    # V7: pages, tab count, file size
    pages = sum(v3lib.estimate_print_pages(v3[n]) for n in v3.sheetnames)
    results['pages'] = pages
    if pages < 200:
        problems.append(f'V7 FAIL: page estimate {pages} < 200')
    if len(v3.sheetnames) < 200:
        problems.append(f'V7 FAIL: {len(v3.sheetnames)} tabs < 200')
    size_mb = os.path.getsize(V3) / 1e6
    results['file_mb'] = round(size_mb, 1)
    if size_mb < 29:
        problems.append(f'V7 FAIL: file {size_mb:.1f}MB < 29MB')

    # V8: ledger integrity
    fl = v3['Fact_Ledger']
    fids = []
    for row in fl.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.startswith('F') and v[1:].isdigit():
            fids.append(int(v[1:]))
    fmax = max(fids)
    missing = sorted(set(range(166, fmax + 1)) - set(fids))
    if missing:
        problems.append(f'V8: missing fact IDs {missing[:10]}')
    si = v3['Source_Index']
    sids = []
    for row in si.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.startswith('S') and v[1:].isdigit():
            sids.append(int(v[1:]))
    smax = max(sids)
    smissing = sorted(set(range(1, smax + 1)) - set(sids))
    if smissing:
        problems.append(f'V8: missing source IDs {smissing[:10]}')
    lives = set()
    for row in fl.iter_rows(min_col=10, max_col=10, min_row=186):
        if row[0].value:
            lives.add(str(row[0].value))
    ghost = [t for t in lives if t not in v3.sheetnames]
    if ghost:
        problems.append(f'V8: fact lives_on ghost tabs {ghost}')
    results['facts_max'] = fmax
    results['sources_max'] = smax
    results['tabs'] = len(v3.sheetnames)
    results['charts'] = sum(len(v3[n]._charts) for n in v3.sheetnames)

    # V9: chart integrity - house axes, no combo/secondary, no overlaps,
    # no smoothed lines, categories present on multi-point series
    import zipfile
    import chart_audit as CA
    z = zipfile.ZipFile(V3)
    smap = CA.sheet_map(z)
    n_ch = 0
    v9 = []
    geo_by_sheet = {}
    for sheet, spath in smap.items():
        dpath = CA.drawing_for(z, spath)
        if not dpath:
            continue
        for col, row, cx, cy, cpath in CA.charts_in_drawing(z, dpath):
            if not cpath:
                continue
            info, _ = CA.audit_chart(z, cpath)
            n_ch += 1
            n_val = sum(1 for a in info['axes'] if a['kind'] == 'valAx')
            if len(info['plots']) > 1 or n_val > 1:
                v9.append(f'combo/secondary chart on {sheet} ({cpath})')
            horiz = any(p.get('barDir') == 'bar' for p in info['plots'])
            cat_pos, val_pos = ('l', 'b') if horiz else ('b', 'l')
            for a in info['axes']:
                if a['kind'] in ('catAx', 'dateAx') and a['pos'] != cat_pos:
                    v9.append(f'category axis pos={a["pos"]} on {sheet}')
                if a['kind'] == 'valAx' and a['pos'] != val_pos:
                    v9.append(f'value axis pos={a["pos"]} on {sheet}')
                if a['delete'] not in ('0', 'false'):
                    v9.append(f'axis delete not explicit-0 on {sheet}')
            for p in info['plots']:
                if p['type'] == 'lineChart' and any(
                        s not in ('0', 'false') for s in p['ser_smooth']):
                    v9.append(f'smoothed line series on {sheet}')
                if p['n_series'] == 0:
                    v9.append(f'zero-series chart on {sheet}')
    # exact overlap geometry from the loaded workbook (true col/row sizes)
    del geo_by_sheet
    for name in v3.sheetnames:
        ws3 = v3[name]
        rects = []
        for ch in ws3._charts:
            a = ch.anchor
            frm = getattr(a, '_from', None)
            ext = getattr(a, 'ext', None)
            if frm is None or ext is None:
                continue
            x = v3lib._col_x_cm(ws3, frm.col + 1) + (frm.colOff or 0) / 360000
            y = v3lib._row_y_cm(ws3, frm.row + 1) + (frm.rowOff or 0) / 360000
            rects.append((x, y, ext.cx / 360000, ext.cy / 360000))
        for i, (ax, ay, aw, ah) in enumerate(rects):
            for bx, by, bw, bh in rects[i + 1:]:
                ox = min(ax + aw, bx + bw) - max(ax, bx)
                oy = min(ay + ah, by + bh) - max(ay, by)
                if ox > 0.6 and oy > 0.6:
                    v9.append(f'chart overlap on {name} (~{ox:.1f}x{oy:.1f}cm)')
    results['charts_checked'] = n_ch
    results['chart_defects'] = len(v9)
    for msg in v9[:20]:
        problems.append(f'V9: {msg}')
    if len(v9) > 20:
        problems.append(f'V9: ... +{len(v9) - 20} more chart defects')

    json.dump(results, open(os.path.join(SCRATCH, 'verify_results.json'), 'w'),
              indent=1)
    print(json.dumps(results, indent=1))
    if problems:
        print(f'\nPROBLEMS ({len(problems)}):')
        for p in problems:
            print(' -', p)
        sys.exit(1)
    print('ALL GATES PASS')


if __name__ == '__main__':
    main()
