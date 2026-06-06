#!/usr/bin/env python3
"""Ingest HCPEA (Healthcare Private Equity Association) deal-tracker workbooks
into the Deal Library.

These are *transaction* trackers the user holds under their own access: one
row per healthcare-PE deal (target, acquirer/sponsor, sector, deal type,
date, deal size). They are NOT company screens, so this loads the
transaction fields added to ``deal_library_companies`` (deal_date,
deal_quarter, deal_type, buyer_name, seller_name, transaction_value,
post_valuation) alongside the company fields it already had.

Compliance posture mirrors ``ingest_deal_library_exports.py``:
  * Missing means unavailable — blanks / "-" / "NM" become None, never 0.
  * Nothing is inferred from a blank; no value is invented.
  * Every row keeps source_file / source_sheet / source_row_id + a
    completeness_score + a missing_fields list, so provenance is auditable.
  * Licensed source workbooks stay out of git (data/vendor/** is gitignored);
    only this loader is tracked.

Two layouts are handled (auto-detected per sheet):
  * HCPEA quarterly  — a "Data" sheet, header row begins with "Deal Date":
      Deal Date | Target | Acquirer | Deal Size | Post Valuation | Deal Type | Sector
  * Annual tracker   — one sheet per quarter ("Q1 2019" ...), header row
    begins with "Date":
      Date | Target Company | Company(buyer) | Private Equity Firm |
      Selling Entity | Size ($ M) | Sector | Deal Type

Usage::

    python scripts/ingest_hcpea_trackers.py --in data/vendor/deal_library --db prod.db
    python scripts/ingest_hcpea_trackers.py --files a.xlsx b.xlsx --db prod.db
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

# Make rcm_mc importable when run from the repo root or scripts/.
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from rcm_mc.data import deal_library  # noqa: E402
from rcm_mc.portfolio.store import PortfolioStore  # noqa: E402

_MISSING = {"", "-", "--", "n/a", "na", "nm", "null", "none", "nan", "tbd", "n.a."}
_SOURCE_SYSTEM = "HCPEA Deal Tracker"

# Key fields used to score completeness (not every column — provenance/derived
# columns don't count toward "how complete is this deal row").
_KEY_FIELDS = [
    "company_name", "buyer_name", "sponsor_owner", "industry",
    "deal_type", "deal_date", "transaction_value",
]


def _s(v: Any) -> Optional[str]:
    """Trim + collapse whitespace; map missing-tokens to None. Never invents."""
    if v is None:
        return None
    s = " ".join(str(v).split())
    return None if s.lower() in _MISSING else s


def _money(v: Any) -> Optional[float]:
    """Parse a $USDmm cell. Blank/missing -> None (never 0)."""
    s = _s(v)
    if s is None:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").rstrip("xX ").strip()
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _iso_date(v: Any) -> Optional[str]:
    """Normalise a date cell to ISO YYYY-MM-DD. Tolerates datetime or string."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = _s(v)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s[:10] if len(s) >= 10 else s


def _quarter(iso: Optional[str]) -> Optional[str]:
    """ISO date -> 'YYYYQn' (the house quarter format). None-safe."""
    if not iso or len(iso) < 7:
        return None
    try:
        y, m = int(iso[:4]), int(iso[5:7])
        return f"{y}Q{(m - 1) // 3 + 1}"
    except ValueError:
        return None


_DEAL_TYPE_ACRONYMS = {"LBO", "IPO", "M&A", "PE", "PIPE"}


def _norm_deal_type(v: Any) -> Optional[str]:
    """Canonical casing/spacing for the deal-type category so source variants
    collapse for clean aggregation. 'Add-On'/'Add-on' -> 'Add-On';
    'Merger / Acquisition'/'Merger/Acquisition' -> 'Merger/Acquisition';
    'Buyout/LBO' kept. Display normalization only — the meaning is unchanged
    and provenance (source_file/row) is preserved."""
    s = _s(v)
    if s is None:
        return None
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s*-\s*", "-", s)
    parts = re.split(r"([/-])", s)
    out = []
    for p in parts:
        if p in ("/", "-"):
            out.append(p)
        elif p.upper() in _DEAL_TYPE_ACRONYMS:
            out.append(p.upper())
        else:
            out.append(p.capitalize())
    return "".join(out)


def _sponsor_from_acquirer(acquirer: Optional[str]) -> Optional[str]:
    """HCPEA packs the lead sponsor as the acquirer, often with the deal
    contact in parens — 'CVC Capital Partners (Jane Doe)'. Keep the lead
    firm (text before the first '('); fall back to the whole string."""
    if not acquirer:
        return None
    head = acquirer.split("(", 1)[0].strip(" ,;")
    return head or acquirer


def _find_header(ws, first_label: str) -> Optional[int]:
    """Row index (1-based) whose row contains a cell == first_label."""
    target = first_label.lower()
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            cell = ws.cell(row=r, column=c).value
            if cell is not None and str(cell).strip().lower() == target:
                return r
    return None


def _col_map(ws, header_row: int) -> Dict[str, int]:
    """Map a normalized header label -> column index for the header row."""
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            key = " ".join(str(v).split()).lower()
            out.setdefault(key, c)
    return out


def _rec(source_file: str, sheet: str, row_id: int,
         fields: Dict[str, Any]) -> Dict[str, str]:
    """Assemble one normalized deal_library row (string-valued for CSV)."""
    company_id = hashlib.sha1(
        f"{_SOURCE_SYSTEM}|{source_file}|{sheet}|{row_id}".encode()
    ).hexdigest()[:20]
    deal_date = fields.get("deal_date")
    rec: Dict[str, Any] = {
        "company_id": company_id,
        "source_system": _SOURCE_SYSTEM,
        "source_file": source_file,
        "source_sheet": sheet,
        "source_row_id": str(row_id),
        "company_name": fields.get("company_name"),
        "industry": fields.get("industry"),
        "sponsor_owner": fields.get("sponsor_owner"),
        "deal_date": deal_date,
        "deal_quarter": _quarter(deal_date),
        "deal_type": fields.get("deal_type"),
        "buyer_name": fields.get("buyer_name"),
        "seller_name": fields.get("seller_name"),
        "transaction_value": fields.get("transaction_value"),
        "post_valuation": fields.get("post_valuation"),
    }
    present = sum(1 for k in _KEY_FIELDS if rec.get(k) not in (None, ""))
    missing = [k for k in _KEY_FIELDS if rec.get(k) in (None, "")]
    rec["completeness_score"] = round(present / len(_KEY_FIELDS), 3)
    rec["missing_fields"] = ";".join(missing)
    # CSV wants strings; None -> "" (loader._coerce maps "" back to None).
    return {k: ("" if v is None else str(v)) for k, v in rec.items()}


def _rows_hcpea(ws, source_file: str) -> List[Dict[str, str]]:
    hr = _find_header(ws, "Deal Date")
    if hr is None:
        return []
    cm = _col_map(ws, hr)
    out: List[Dict[str, str]] = []
    for r in range(hr + 1, ws.max_row + 1):
        get = lambda label: (ws.cell(row=r, column=cm[label]).value
                             if label in cm else None)  # noqa: E731
        target = _s(get("target"))
        if not target:
            continue
        acquirer = _s(get("acquirer"))
        out.append(_rec(source_file, ws.title, r, {
            "company_name": target,
            "buyer_name": acquirer,
            "sponsor_owner": _sponsor_from_acquirer(acquirer),
            "industry": _s(get("sector")),
            "deal_type": _norm_deal_type(get("deal type")),
            "deal_date": _iso_date(get("deal date")),
            "transaction_value": _money(
                next((get(k) for k in cm if k.startswith("deal size")), None)),
            "post_valuation": _money(
                next((get(k) for k in cm if k.startswith("post valuation")), None)),
        }))
    return out


def _rows_annual(ws, source_file: str) -> List[Dict[str, str]]:
    hr = _find_header(ws, "Date")
    if hr is None:
        return []
    cm = _col_map(ws, hr)
    # Position-based for the ambiguous bare "Company" (buyer) column, which
    # sits immediately right of "Target Company".
    tc = cm.get("target company")
    buyer_col = (tc + 1) if tc else cm.get("company")
    out: List[Dict[str, str]] = []
    for r in range(hr + 1, ws.max_row + 1):
        cell = lambda col: (ws.cell(row=r, column=col).value if col else None)  # noqa: E731
        target = _s(cell(tc))
        if not target:
            continue
        out.append(_rec(source_file, ws.title, r, {
            "company_name": target,
            "buyer_name": _s(cell(buyer_col)),
            "sponsor_owner": _s(cell(cm.get("private equity firm"))),
            "seller_name": _s(cell(cm.get("selling entity"))),
            "industry": _s(cell(cm.get("sector"))),
            "deal_type": _norm_deal_type(cell(cm.get("deal type"))),
            "deal_date": _iso_date(cell(cm.get("date"))),
            "transaction_value": _money(
                cell(next((c for k, c in cm.items() if k.startswith("size")), None))),
        }))
    return out


def parse_workbook(path: Path) -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, str]] = []
    for sheet in wb.sheetnames:
        if sheet.strip().lower() in ("disclaimer", "notes", "readme", "key"):
            continue
        ws = wb[sheet]
        if _find_header(ws, "Deal Date") is not None:
            rows.extend(_rows_hcpea(ws, path.name))
        elif _find_header(ws, "Date") is not None:
            rows.extend(_rows_annual(ws, path.name))
    wb.close()
    return rows


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="indir", help="dir of .xlsx tracker files")
    ap.add_argument("--files", nargs="*", help="explicit .xlsx file paths")
    ap.add_argument("--db", required=True, help="SQLite DB path to load into")
    ap.add_argument("--dry-run", action="store_true",
                    help="parse + report, do not write to the DB")
    args = ap.parse_args(argv)

    paths: List[Path] = []
    if args.indir:
        paths.extend(sorted(Path(args.indir).glob("*.xlsx")))
    if args.files:
        paths.extend(Path(p) for p in args.files)
    paths = [p for p in paths if p.exists() and not p.name.startswith("~$")]
    if not paths:
        print("No .xlsx files found.", file=sys.stderr)
        return 2

    all_rows: List[Dict[str, str]] = []
    per_file: List[Tuple[str, int]] = []
    for p in paths:
        rows = parse_workbook(p)
        per_file.append((p.name, len(rows)))
        all_rows.extend(rows)

    print(f"Parsed {len(all_rows)} deals from {len(paths)} workbook(s):")
    for name, n in per_file:
        print(f"  {n:>5}  {name}")

    if args.dry_run:
        return 0

    # Write a normalized CSV, then reuse the audited loader (upsert on
    # company_id -> idempotent re-import).
    tmp = Path(tempfile.mkdtemp()) / "deal_library_companies.csv"
    cols = deal_library._ALL_COLS
    with tmp.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for row in all_rows:
            w.writerow({c: row.get(c, "") for c in cols})

    store = PortfolioStore(args.db)
    written = deal_library.load_companies_csv(store, tmp)
    total = deal_library.count(store)
    print(f"Loaded {written} rows into deal_library_companies "
          f"({total} total now in {args.db}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
