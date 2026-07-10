"""Self-contained NPI claims-cleaner engine — stdlib only.

Backs the ``/npi-cleaner`` page. Reads a delimited claims file (CSV / TSV),
auto-detects NPI-bearing columns, validates every NPI against the official
Luhn check, de-duplicates exact-duplicate rows, normalizes whitespace, flags
missing / malformed / checksum-failing billing NPIs, and emits a cleaned CSV
plus a scorecard.

Why stdlib-only, and why not the uploaded ``npi_recovery`` package
--------------------------------------------------------------------
The uploaded ``NPI_Recovery_and_Cleaner_v48`` package ships 43 modules but is
**missing its engine core** — its ``__init__`` imports ``pipeline.py`` and
``entity.py``, neither of which is in the archive, so ``run_pipeline`` (the
full Steps 0–8 recovery orchestrator) cannot be imported. It also needs
pandas / numpy / live CMS network access, which the stdlib ``rcm-mc serve``
server deliberately avoids. Those modules are kept for provenance under
``vendor_v48/`` (see its README).

This engine implements the genuinely deliverable, offline cleaning steps with
zero third-party dependencies, so the page works end-to-end. ``run_pipeline``
is attempted first (guarded) — if a complete ``npi_recovery`` is ever dropped
in, it is used automatically; otherwise we fall back to the built-in cleaner.

The Luhn rule mirrors ``vendor_v48/npi_recovery/field_validators.py`` exactly:
Luhn over the constant prefix ``80840`` plus the first 9 NPI digits.
"""
from __future__ import annotations

import csv
import io
import shutil
import re
import threading
import time
import traceback
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

# Scratch area for in-flight jobs and their cleaned output. /tmp keeps this off
# the repo tree and matches the vendored webapp's WORKDIR convention.
WORKDIR = Path("/tmp/npi_cleaner_web")
WORKDIR.mkdir(parents=True, exist_ok=True)

# Column names (case/space/punct-insensitive) that carry an NPI.
_NPI_HINTS = (
    "npi", "billingnpi", "renderingnpi", "referringnpi", "providernpi",
    "attendingnpi", "billingprovidernpi", "facilitynpi", "servicingnpi",
)
# Tokens that mark a column as a DERIVED ATTRIBUTE OF an NPI — a boolean flag,
# a status, a type, a name, a date — rather than a column that HOLDS the NPI
# value itself. A real 600k-row extract carried "OPTIONCARE_NPI_FLAG" (a Y/N
# flag): it contains "npi", was scored as 600k malformed NPIs, and tanked the
# Validity dimension to 0.0%. "NPI Number"/"NPI_ID"/"NPI Code" are deliberately
# NOT excluded — those DO hold the identifier; only clearly-derived companions
# are dropped. Tokens are matched against split header words (see
# ``_header_tokens``); the suffix fallback (len ≥ 4) handles delimiter-free
# blobs like "OPTIONCARENPIFLAG".
_NPI_MODIFIER_TOKENS = frozenset({
    "flag", "flg", "indicator", "ind", "status", "stat",
    "valid", "validity", "validated", "verify", "verified", "verification",
    "match", "matched", "type", "typ", "kind", "category", "cat",
    "desc", "description", "name", "count", "cnt", "date", "dt",
    "reason", "qual", "qualifier", "present", "missing", "exists",
})
# A "billing" NPI is the one that must be present for a claim to be payable —
# missing values here are the headline recovery target.
_BILLING_HINTS = ("billing", "billingprovider", "pay")
# Ordering / referring provider NPI columns — a distinct eligibility concern
# (CMS PECOS ordering/referring enrollment) screened separately from billing.
_ORDER_REFER_HINTS = ("ordering", "referring", "orderprov", "referprov",
                      "orderingprovider", "referringprovider", "orderphys",
                      "referphys")

# Columns that carry a provider / organization name, used to recover a missing
# NPI by searching NPPES. Order = priority.
_NAME_HINTS = (
    "organizationname", "orgname", "providername", "billingprovidername",
    "facilityname", "practicename", "provider", "name",
)
# PERSON provider-name columns for the re-case + credential-parse repair.
# Deliberately narrower than _NAME_HINTS (which also feeds NPPES lookups):
# re-casing must never touch an organization-name column, so org-flavored
# headers are excluded outright and org-looking VALUES (LLC, CLINIC …) are
# skipped again per cell inside the cleaner.
_PNAME_HINTS = ("providername", "physicianname", "renderingname",
                "attendingname", "referringname", "servicingname",
                "providerfirst", "providerlast", "doctorname")
_PNAME_EXCLUDE = ("organization", "org", "facility", "practice", "group",
                  "patient", "payer", "member", "subscriber", "insur")
# "st" is exact-match only (see _detect_state): as a substring it matched
# ClaimStatus on 835 tables and fed a claim-status digit to NPPES as the
# state.
_STATE_HINTS = ("providerstate", "billingstate", "state", "provstate")
# Columns carrying a drug identifier, for the RxNorm / openFDA connectors.
_NDC_HINTS = ("ndc11", "ndc", "drugndc", "ndccode")
_DRUG_HINTS = ("drugname", "drug", "productname", "medication", "labelname")
# Roles for the deterministic normalization pass.
_MONEY_HINTS = ("allowedamt", "allowed", "paidamt", "paid", "billedamt",
                "billed", "chargeamt", "charge", "amount", "cost", "fee")
# Headers that CONTAIN a money hint but are not money — verified corruption
# without this: "CostCenter" matched "cost" and cost-center code "0100" was
# rewritten to "100.00"; "ChargeDescription"/"FeeScheduleName" false-flagged
# money-unparseable on every row. Same class as the DischargeDate incident.
_MONEY_EXCLUDE = ("description", "desc", "center", "centre", "code", "name",
                  "note", "comment", "schedule", "flag", "type", "status",
                  "reason")
_DATE_HINTS = ("dateofservice", "servicedate", "dos", "paiddate", "date",
               "dob", "birthdate", "fromdate", "thrudate")
# Date roles that can NEVER legitimately be in the future — a service was
# rendered, a patient was born, a claim was adjudicated. Used to flag
# future-dated rows. Deliberately excludes generic "date"/coverage-end/
# authorization-expiry columns, which CAN be future.
# Deliberately excludes "paiddate": 835 remits stamp the payer's production
# date on every row, and forward-dated scheduled payments are legitimate —
# flagging them tanked the consistency grade on ordinary ERAs.
_SERVICE_DATE_HINTS = ("dateofservice", "servicedate", "svcdate", "dos",
                       "fromdate", "thrudate", "servicefromdate",
                       "servicetodate", "admitdate", "admissiondate",
                       "dischargedate", "adjudicationdate")
_DOB_HINTS = ("dateofbirth", "dob", "birthdate", "birthdt", "patientdob")
_ZIP_HINTS = ("zip", "postalcode", "postal")
_HCPCS_HINTS = ("hcpcs", "cpt", "proccode", "procedurecode")
_SEX_HINTS = ("patientsex", "sex", "gender", "patientgender")
_DX_HINTS = ("diagnosis", "diagnosiscode", "icd", "icd10", "dx", "dxcode")
_MOD_HINTS = ("modifier", "modifiers", "mod1", "mod2", "hcpcsmodifier")
_PHONE_HINTS = ("phone", "fax", "telephone", "phonenumber")
_TAXO_HINTS = ("taxonomy", "taxonomycode", "providertaxonomy", "nucc")
_REV_HINTS = ("revenuecode", "revcode", "revenuecd", "revcd")
_ADMIT_HINTS = ("admitdate", "admissiondate", "admitdt")
_DISCH_HINTS = ("dischargedate", "dischargedt", "dischdate")
# Payer column, priority order — deliberately excludes bare "plan" (too many
# false matches: planid, plan_year, careplan).
_PAYER_HINTS = ("payername", "payer", "payor", "insurancename",
                "insurancecompany", "insurance", "carrier", "healthplan")

_TOB_HINTS = ("typeofbill", "tobcode", "billtype", "tob")
_MEMBER_HINTS = ("mbi", "medicareid", "medicarenumber", "memberid",
                 "subscriberid", "insuredid", "beneficiaryid", "beneid",
                 "hicn", "hic")
_COND_HINTS = ("conditioncode", "conditioncodes", "condcode")
_OCC_HINTS = ("occurrencecode", "occurrencecodes", "occcode")
_VALUE_HINTS = ("valuecode", "valuecodes", "valcode")
_DISCH_STATUS_HINTS = ("dischargestatus", "patientdischargestatus",
                       "dischargedisposition", "patientstatus")
_ADMIT_TYPE_HINTS = ("admissiontype", "admittype")
_SUBMIT_HINTS = ("receiveddate", "submissiondate", "submitdate",
                 "filedate", "receiptdate", "datereceived")
# Claim adjustment / denial reason column (CARC). "denial" last — broadest.
_CARC_HINTS = ("carc", "denialcode", "denialreason", "adjustmentreason",
               "adjreason", "reasoncode", "denial")
# A CARC is 1-3 digits or letter+1-2 digits (A1, B7, P1, W1 …).
_CARC_VALID_RE = re.compile(r"^(\d{1,3}|[A-Z]\d{1,2})$")
# 835 exports routinely prefix the CARC with its X12 group code ("CO-45",
# "PR1", "OA23"): strip the group before shape-checking and tallying so a
# CO-45 file doesn't flag carc-invalid on every row (tanking the validity
# dimension) and CO-45 / 45 count as ONE denial reason with an intact
# playbook join. Group-only tokens ("CO") pass through unchanged — the
# lookahead requires a code to follow.
_CARC_GROUP_RE = re.compile(r"^(?:CO|OA|PI|PR|CR)[-\s]?(?=[A-Z]?\d)")

# The official CMS two-digit Place of Service code set. Report-only domain
# check — a POS outside this list is a denial waiting to happen.
_POS_VALID = {
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
    "21", "22", "23", "24", "25", "26", "27", "31", "32", "33",
    "34", "41", "42", "49", "50", "51", "52", "53", "54", "55",
    "56", "57", "58", "60", "61", "62", "65", "66", "71", "72",
    "81", "99",
}

# Payer-family prefixes for variant clustering (report-only — cells are never
# rewritten across brands). Matching is on the punctuation-stripped compact
# uppercase key; startswith keeps "BCBS OF TEXAS" and "BCBSTX" together while
# leaving unknown payers as their own normalized key.
_PAYER_FAMILIES = (
    ("BLUECROSS", "BLUE CROSS BLUE SHIELD"),
    ("BCBS", "BLUE CROSS BLUE SHIELD"),
    ("UNITEDHEALTH", "UNITEDHEALTHCARE"),
    ("UHC", "UNITEDHEALTHCARE"),
    ("AETNA", "AETNA"),
    ("CIGNA", "CIGNA"),
    ("HUMANA", "HUMANA"),
    ("KAISER", "KAISER PERMANENTE"),
    ("MEDICARE", "MEDICARE"),
    ("MEDICAID", "MEDICAID"),
    ("TRICARE", "TRICARE"),
)

# Tokens that mean "missing" and are normalized to a blank cell.
_NULL_TOKENS = {"na", "n/a", "null", "none", "nan", "nil", "-", "--", "#n/a",
                "unknown", "unk", ".", "?"}

# Common mojibake fixups from UTF-8 mis-decoded as latin-1/cp1252.
_MOJIBAKE = {
    "â": "'", "â": "'",
    "â": '"', "â": '"',
    "â": "-", "â": "-",
    "Â ": " ",
    # Accented-letter artifacts (utf-8 read as cp1252) — the most common
    # letters in Spanish/French/German provider and practice names. "Ã "
    # (a-grave) is keyed with a PLAIN space because the whitespace step
    # folds the NBSP to a space before this table runs.
    "Ã©": "é", "Ã¨": "è", "Ãª": "ê", "Ã«": "ë",
    "Ã¡": "á", "Ã¢": "â", "Ã£": "ã", "Ã¤": "ä", "Ã¥": "å",
    "Ã³": "ó", "Ã²": "ò", "Ã´": "ô", "Ãµ": "õ", "Ã¶": "ö",
    "Ãº": "ú", "Ã¹": "ù", "Ã»": "û", "Ã¼": "ü",
    "Ã­": "í", "Ã¬": "ì", "Ã®": "î", "Ã¯": "ï",
    "Ã±": "ñ", "Ã§": "ç",
    "Ã‰": "É", "Ã‘": "Ñ", "Ã–": "Ö", "Ãœ": "Ü",
    "Ã ": "à", "Ã ": "à",
}

# Full state / territory names → USPS 2-letter code.
_STATE_NAMES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "district of columbia": "DC", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN",
    "iowa": "IA", "kansas": "KS", "kentucky": "KY", "louisiana": "LA",
    "maine": "ME", "maryland": "MD", "massachusetts": "MA", "michigan": "MI",
    "minnesota": "MN", "mississippi": "MS", "missouri": "MO", "montana": "MT",
    "nebraska": "NE", "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "puerto rico": "PR", "guam": "GU", "virgin islands": "VI",
}
_STATE_CODES = set(_STATE_NAMES.values()) | {
    "AS", "MP", "FM", "MH", "PW", "AA", "AE", "AP"}
_SEX_MAP = {"m": "M", "male": "M", "1": "M", "f": "F", "female": "F",
            "2": "F", "u": "U", "unknown": "U", "0": "U"}
_EXCEL_FLOAT_RE = re.compile(r"^(\d+)\.0+$")
_MONEY_PARENS_RE = re.compile(r"^\((.*)\)$")
# Trailing-sign negatives from mainframe/EBCDIC-lineage extracts and some
# 835 flatteners: "500.00-" and "500.00CR" both mean -500.00. Matched on the
# $-and-space-stripped value.
_MONEY_TRAIL_NEG_RE = re.compile(r"^([\d.,]*\d)\s*(-|CR)$", re.IGNORECASE)
# Unambiguously EUROPEAN-formatted amounts: dot-grouped thousands with a
# comma-decimal ("1.234,56"), or a bare comma-decimal with exactly two
# decimal digits ("1234,56"). "1,234" alone stays a US thousands reading —
# ambiguity always resolves to the US interpretation, so nothing that parsed
# correctly before changes meaning.
_MONEY_EU_RE = re.compile(r"^-?(\d{1,3}(\.\d{3})+|\d+),\d{1,2}$")
# Scientific-notation NPI damage (spreadsheet round-trip): a 10-digit NPI
# stored as a float renders as d.dddddddddE+09. When the mantissa carries all
# 10 digits the value is exactly recoverable; when it doesn't, the export
# destroyed the identifier and the cell is flagged, never guessed at.
_NPI_SCI_RE = re.compile(r"^\d(\.\d+)?[eE]\+?0*9$")


def _clean_generic(cell: str) -> Tuple[str, List[str]]:
    """Generic cell cleanups applied to every column. Returns (value, rules)."""
    rules: List[str] = []
    v = cell
    # Zero-width + non-breaking spaces → normal space.
    v2 = (v.replace("​", "").replace("﻿", "")
          .replace(" ", " ").replace("\t", " "))
    if v2 != v:
        rules.append("whitespace-chars")
        v = v2
    # Mojibake repair.
    for bad, good in _MOJIBAKE.items():
        if bad in v:
            v = v.replace(bad, good)
            rules.append("mojibake")
    # Collapse runs of internal whitespace.
    v2 = re.sub(r"  +", " ", v).strip()
    if v2 != v:
        if "whitespace-chars" not in rules:
            rules.append("collapse-space")
        v = v2
    # Excel leading text-marker apostrophe: '01234 → 01234.
    if v.startswith("'") and len(v) > 1:
        v = v[1:]
        rules.append("leading-apostrophe")
    # Null tokens → blank.
    if v.lower() in _NULL_TOKENS:
        if v != "":
            rules.append("null-token")
        v = ""
    return v, rules


def _clean_npi_cell(v: str) -> Tuple[str, List[str]]:
    rules: List[str] = []
    m = _EXCEL_FLOAT_RE.match(v)
    if m:  # 1234567890.0 (Excel float coercion)
        v = m.group(1)
        rules.append("npi-excel-float")
    if _NPI_SCI_RE.match(v.strip()):
        # Spreadsheet scientific notation ("1.679576722E+09"). Recover ONLY
        # when the mantissa carries ALL 10 digits — "1.68E+09" could have
        # been any NPI starting 168, so expanding it would fabricate an
        # identifier ("1680000000") and stripping non-digits would
        # fabricate a 12-digit junk value ("16809"). The lossy case is
        # left untouched for the npi-scientific-lossy flag (see the row
        # loop) so the user learns the export destroyed identifiers.
        _mant = v.strip().upper().split("E", 1)[0].replace(".", "")
        if len(_mant) == 10 and _mant.isdigit():
            return _mant, rules + ["npi-scientific-notation"]
        return v, rules
    digits = "".join(ch for ch in v if ch.isdigit())
    if digits != v and v != "":
        rules.append("npi-strip-nondigits")
    return (digits if v != "" else ""), rules


def _clean_money_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    rules: List[str] = []
    raw = v
    neg = False
    m = _MONEY_PARENS_RE.match(v.strip())
    if m:  # (123.45) accounting negative
        v = m.group(1)
        neg = True
    v2 = v.replace("$", "").replace(" ", "").strip()
    trail = _MONEY_TRAIL_NEG_RE.match(v2)
    if trail:  # 500.00- / 500.00CR (mainframe trailing sign)
        v2 = trail.group(1)
        neg = True
    eu = False
    if _MONEY_EU_RE.match(v2):
        # Unambiguous European format: "1.234,56" → 1234.56. Without this,
        # stripping commas turned a €1.234,56 charge into $1.23 and logged
        # it as a repair. Anything ambiguous keeps the US reading below.
        v2 = v2.replace(".", "").replace(",", ".")
        eu = True
    else:
        v2 = v2.replace(",", "")
    try:
        num = float(v2)
    except ValueError:
        return raw, []
    if neg:
        num = -abs(num)
    out = "%.2f" % num
    if out != raw:
        if eu:
            rules.append("money-eu-decimal")
        if trail:
            rules.append("money-trailing-negative")
        if not rules:
            rules.append("money-normalize")
    return out, rules


def _modifier_malformed(v: str) -> bool:
    """True when a cleaned modifier cell still carries a token that isn't a
    2-character alphanumeric modifier ("LT4", "2", "GYX"). The normalizer
    deliberately KEEPS such tokens (content is never silently discarded);
    this flag is how the user learns they exist."""
    s = v.strip()
    if not s:
        return False
    return any(not (len(p) == 2 and p.isalnum())
               for p in s.split(",") if p)


def _money_unparseable(v: str) -> bool:
    """True when a non-blank amount cell can't be read as a number.

    Mirrors ``_clean_money_cell``'s parse ($, commas, spaces, accounting
    parens) — that cleaner returns the cell unchanged with no repair when it
    fails, so an unparseable amount (``pending``, ``1,2OO`` with a letter O, a
    stray ``$``) would otherwise slip through the money column unflagged. Null
    tokens are already blanked by ``_clean_generic`` before this runs, so a
    blank never flags."""
    s = v.strip()
    if not s:
        return False
    m = _MONEY_PARENS_RE.match(s)
    if m:
        s = m.group(1)
    s = s.replace("$", "").replace(" ", "").strip()
    t = _MONEY_TRAIL_NEG_RE.match(s)
    if t:
        s = t.group(1)
    if _MONEY_EU_RE.match(s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    if s == "":
        return True
    try:
        float(s)
        return False
    except ValueError:
        return True


def _excel_serial_to_iso(n: float) -> Optional[str]:
    # Excel serial: 1 = 1900-01-01, with the well-known 1900 leap bug (>59).
    if not (20000 <= n <= 60000):  # ~1954..2064, the plausible claims window
        return None
    from datetime import date, timedelta
    base = date(1899, 12, 30)
    try:
        return (base + timedelta(days=int(n))).isoformat()
    except (OverflowError, ValueError):
        return None


_DATE_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")
# US dates tolerate a trailing time component ("01/02/2024 10:30",
# "1/2/24 3:45:00 PM") — datetime-stamped exports are routine and the time
# part is discarded exactly like date-iso-trim discards it from ISO stamps.
_DATE_US_RE = re.compile(
    r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})"
    r"(?:[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AaPp][Mm])?)?$")
# Compact CCYYMMDD ("20240211") — the HIPAA X12 date format leaking into a
# tabular extract. Century-bounded so a stray 8-digit id can't false-match.
_DATE_COMPACT_RE = re.compile(r"^(19|20)(\d{2})(\d{2})(\d{2})$")


def _clean_date_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    s = v.strip()
    if _DATE_ISO_RE.match(s):
        return s[:10], ([] if s[:10] == v else ["date-iso-trim"])
    m = _DATE_COMPACT_RE.match(s)
    if m:
        # 8-digit CCYYMMDD → ISO, but only when it is a REAL calendar date;
        # an implausible month/day falls through to the unparseable flag.
        try:
            from datetime import date as _d
            iso = _d(int(m.group(1) + m.group(2)), int(m.group(3)),
                     int(m.group(4))).isoformat()
            return iso, ["date-compact-to-iso"]
        except ValueError:
            return v, []
    # Excel serial number.
    try:
        n = float(s)
        iso = _excel_serial_to_iso(n)
        if iso:
            return iso, ["date-excel-serial"]
    except ValueError:
        pass
    m = _DATE_US_RE.match(s)
    if m:
        mo, da, yr = m.group(1), m.group(2), m.group(3)
        if len(yr) == 2:
            yr = ("20" + yr) if int(yr) < 50 else ("19" + yr)
        try:
            from datetime import date as _d
            iso = _d(int(yr), int(mo), int(da)).isoformat()
            return iso, ["date-us-to-iso"]
        except ValueError:
            return v, []
    return v, []


def _date_unparseable(v: str) -> bool:
    """True when a CLEANED date cell is non-blank and still not ISO-shaped.

    Every parseable form (ISO, ISO datetime, Excel serial, US slash/dash
    with or without a time, compact CCYYMMDD) is normalized to YYYY-MM-DD by
    ``_clean_date_cell`` before this runs, so anything left is a value the
    chronology checks silently skip — the money column has
    ``money-unparseable``; this is the date analogue."""
    s = v.strip()
    if not s:
        return False
    return _DATE_ISO_RE.match(s) is None


# ZIP3-prefix ranges are approximate for a few split prefixes and for the
# territory/military pseudo-states, so we never flag a mismatch involving one
# of these — the reliable signal is the 50 states + DC.
_ZIP_STATE_SKIP = {"AA", "AE", "AP", "PR", "VI", "GU", "MP", "AS"}
_ZIP3_STATE_CACHE: Optional[Dict[str, str]] = None


def _zip3_state_map() -> Dict[str, str]:
    """The vendored USPS ZIP3→state crosswalk (pure-Python, no pandas), cached.

    Reused rather than duplicated so the offline engine and the full v49
    pipeline agree on the same prefix table."""
    global _ZIP3_STATE_CACHE
    if _ZIP3_STATE_CACHE is None:
        try:
            from .vendor_v49.npi_recovery.geo import ZIP3_TO_STATE
            _ZIP3_STATE_CACHE = dict(ZIP3_TO_STATE)
        except Exception:  # noqa: BLE001 — vendor tree missing → feature off
            _ZIP3_STATE_CACHE = {}
    return _ZIP3_STATE_CACHE


def _zip_state_pairs(headers: List[str]) -> List[Tuple[int, int]]:
    """Pair a state column with a ZIP column that refers to the SAME entity.

    Cross-entity pairs (PatientState vs ProviderZip) would false-positive, so
    columns are matched by the header text left after removing the
    state/zip role tokens (ProviderState↔ProviderZip → both "provider"). When
    nothing matches by entity but there is exactly one of each, they are
    paired as a fallback."""
    def _entity(folded: str, drop: tuple) -> str:
        for d in drop:
            folded = folded.replace(d, "")
        return folded

    states, zips = [], []
    for i, h in enumerate(headers):
        k = _norm_key(h)
        # Bare "St" is a real state header in the wild; match it by exact
        # key (a substring "st" would grab ClaimStatus/PostDate), same as
        # the primary state detector, so state-from-zip fill reaches it.
        if "state" in k or "province" in k or k == "st":
            states.append((i, _entity(k, ("state", "province"))))
        if any(x in k for x in ("zipcode", "zip", "postalcode", "postal")):
            zips.append((i, _entity(k, ("zipcode", "zip", "postalcode",
                                        "postal", "code"))))
    pairs: List[Tuple[int, int]] = []
    used = set()
    for si, se in states:
        for zi, ze in zips:
            if zi not in used and se and ze and se == ze:
                pairs.append((si, zi))
                used.add(zi)
                break
    if not pairs and len(states) == 1 and len(zips) == 1:
        pairs.append((states[0][0], zips[0][0]))
    return pairs


def _date_after(iso: str, today) -> bool:
    """True when a cleaned cell is a valid ISO date strictly after ``today``.

    Only recognizes the normalized ``YYYY-MM-DD`` form the date normalizer
    emits — anything unparseable is treated as not-future (report-only flag,
    so we never guess)."""
    m = _DATE_ISO_RE.match(iso.strip()) if iso else None
    if not m:
        return False
    try:
        from datetime import date as _d
        return _d(int(m.group(1)), int(m.group(2)), int(m.group(3))) > today
    except ValueError:
        return False


def _clean_state_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    up = v.strip().upper()
    if up in _STATE_CODES:
        return up, ([] if up == v else ["state-upper"])
    low = v.strip().lower()
    if low in _STATE_NAMES:
        return _STATE_NAMES[low], ["state-name-to-code"]
    return v, []


def _clean_zip_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    m = _EXCEL_FLOAT_RE.match(v.strip())
    s = m.group(1) if m else v.strip()
    # zip+4 → keep 5; short zip → pad leading zeros (Excel drops them).
    digits = "".join(ch for ch in s if ch.isdigit())
    if 1 <= len(digits) < 5:
        return digits.zfill(5), ["zip-pad"]
    if len(digits) == 9:
        return digits[:5] + "-" + digits[5:], (["zip5+4"] if v != digits[:5] + "-" + digits[5:] else [])
    if digits and digits != v:
        return digits, ["zip-clean"]
    return v, []


def _clean_hcpcs_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    up = v.strip().upper()
    return up, ([] if up == v else ["hcpcs-upper"])


def _clean_sex_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    key = v.strip().lower()
    if key in _SEX_MAP:
        out = _SEX_MAP[key]
        return out, ([] if out == v else ["sex-normalize"])
    return v, []


_ICD_SHAPE_RE = re.compile(r"^[A-Z][0-9][0-9A-Z]{1,5}$")

# Well-formed code shapes for the report-only validity screens (not auto-fixed
# — a bad shape is flagged, never guessed). A HCPCS/CPT base code is 5 digits
# (CPT Cat I), a letter A–V + 4 digits (HCPCS Level II), or 4 digits + a letter
# (CPT Cat II "F" / Cat III "T"). ICD-10-CM is letter · digit · alnum, then an
# optional dotted 1–4-char subclass.
_HCPCS_VALID_RE = re.compile(r"^(\d{5}|[A-V]\d{4}|\d{4}[A-Z])$")
_ICD10_VALID_RE = re.compile(r"^[A-Z][0-9][0-9A-Z](\.[0-9A-Z]{1,4})?$")


def _hcpcs_malformed(v: str) -> bool:
    """True when a non-blank procedure code isn't a valid HCPCS/CPT shape.
    A trailing modifier (``99213-25``/``J1885 59``) is stripped first so a
    valid code with an appended modifier is not falsely flagged."""
    s = v.strip().upper()
    if not s:
        return False
    core = re.split(r"[-\s]", s, 1)[0]
    return _HCPCS_VALID_RE.match(core) is None


def _icd10_malformed(v: str) -> bool:
    """True when a non-blank diagnosis code isn't a valid ICD-10-CM shape."""
    s = v.strip().upper()
    if not s:
        return False
    return _ICD10_VALID_RE.match(s) is None


def _clean_dx_cell(v: str) -> Tuple[str, List[str]]:
    """Uppercase an ICD-10-CM code and insert the decimal after the 3rd
    character (E1165 → E11.65). Codes of length ≤ 3 (e.g. I10) stay as-is."""
    if v == "":
        return v, []
    up = v.strip().upper().replace(" ", "")
    rules: List[str] = []
    if up != v.strip():
        rules.append("dx-upper")
    if "." not in up and _ICD_SHAPE_RE.match(up) and len(up) > 3:
        up = up[:3] + "." + up[3:]
        rules.append("dx-decimal")
    return up, (rules if up != v else ([] if not rules else rules))


def _clean_modifier_cell(v: str) -> Tuple[str, List[str]]:
    """Normalize a claim-line modifier field: split on common delimiters,
    upper-case, de-dup, and re-join with commas. Valid 2-char alphanumeric
    modifiers come first; unrecognized tokens ("LT4", a keyed LT+4) are KEPT
    after them rather than silently discarded — flag, never delete (the
    modifier-malformed flag tells the user they exist)."""
    if v == "":
        return v, []
    parts = re.split(r"[,;|/\s]+", v.strip().upper())
    mods, extras, seen = [], [], set()
    for p in parts:
        p = p.strip()
        if not p or p in seen:
            continue
        seen.add(p)
        if len(p) == 2 and p.isalnum():
            mods.append(p)
        else:
            extras.append(p)
    out = ",".join(mods + extras)
    return out, ([] if out == v else ["modifier-normalize"])


def _clean_phone_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    digits = "".join(ch for ch in v if ch.isdigit())
    if len(digits) == 11 and digits[0] == "1":
        digits = digits[1:]
    if len(digits) == 10:
        out = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return out, ([] if out == v else ["phone-format"])
    return v, []


def _clean_taxonomy_cell(v: str) -> Tuple[str, List[str]]:
    if v == "":
        return v, []
    up = v.strip().upper()
    return up, ([] if up == v else ["taxonomy-upper"])


# A NUCC taxonomy code is exactly 10 alphanumeric characters (the 10th is
# usually "X", but not universally, so we only require the length/charset —
# requiring a trailing X would false-flag legacy codes).
_TAXONOMY_VALID_RE = re.compile(r"^[A-Z0-9]{10}$")
_VALID_SEX = {"M", "F", "U"}


def _taxonomy_malformed(v: str) -> bool:
    s = v.strip().upper()
    if not s:
        return False
    return _TAXONOMY_VALID_RE.match(s) is None


def _sex_invalid(v: str) -> bool:
    """True when a non-blank sex cell didn't resolve to M/F/U. ``_clean_sex_cell``
    returns unmapped values unchanged with no repair, so a stray code (X, 3, N)
    would otherwise pass through the sex column unflagged."""
    s = v.strip().upper()
    if not s:
        return False
    return s not in _VALID_SEX


def _clean_revcode_cell(v: str) -> Tuple[str, List[str]]:
    """Zero-pad a numeric UB-04 revenue code to 4 digits (Excel strips the
    leading zero: 450 → 0450). Non-numeric / over-long values pass through
    for the shape flag to catch."""
    if v == "":
        return v, []
    s = v.strip()
    if s.isdigit() and 1 <= len(s) < 4:
        return s.zfill(4), ["revcode-pad"]
    return s, ([] if s == v else [])


def _revcode_malformed(v: str) -> bool:
    s = v.strip()
    if not s:
        return False
    return not (s.isdigit() and len(s) == 4)


def _clean_pos_cell(v: str) -> Tuple[str, List[str]]:
    """Zero-pad a 1-digit Place of Service (Excel again: 11 stays, 1 → 01)."""
    if v == "":
        return v, []
    s = v.strip()
    if s.isdigit() and len(s) == 1:
        return s.zfill(2), ["pos-pad"]
    return s, ([] if s == v else [])


def _pos_invalid(v: str) -> bool:
    """True when a non-blank POS cell isn't in the official CMS code set."""
    s = v.strip()
    if not s:
        return False
    return s not in _POS_VALID


def _clean_drg_cell(v: str) -> Tuple[str, List[str]]:
    """Zero-pad a numeric MS-DRG to 3 digits (Excel strips leading zeros:
    87 → 087). Non-numeric / over-long values pass through for the shape
    flag to catch."""
    if v == "":
        return v, []
    s = v.strip()
    if s.isdigit() and 1 <= len(s) < 3:
        return s.zfill(3), ["drg-pad"]
    return s, ([] if s == v else [])


def _drg_malformed(v: str) -> bool:
    """MS-DRG is exactly 3 digits, 001-999 (000 is not assigned)."""
    s = v.strip()
    if not s:
        return False
    return not (s.isdigit() and len(s) == 3 and s != "000")


# Clinical credentials recognized at the end of a person provider name.
# Parsed (period-insensitively: "M.D." → MD) for the credential-mix report
# and kept uppercase during re-casing. Mirrored by refdata.CREDENTIALS,
# which carries the display meanings — a test keeps the two sets in sync.
_CREDENTIALS = frozenset((
    "MD", "DO", "MBBS", "NP", "PA", "PAC", "PA-C", "APRN", "ARNP", "FNP",
    "CRNA", "CNM", "CNS", "RN", "LPN", "LVN", "DDS", "DMD", "DPM", "OD",
    "AUD", "PHD", "PSYD", "DC", "PT", "DPT", "OT", "OTR", "SLP", "RD",
    "RDN", "PHARMD", "RPH", "LCSW", "LMSW", "LMFT", "LPC", "BCBA", "MPH",
    "MSN", "BSN", "FACS", "FACP", "FAAP",
))
# Generational suffixes keep canonical casing rather than plain title case.
_NAME_SUFFIXES = {"JR": "Jr", "SR": "Sr", "II": "II", "III": "III",
                  "IV": "IV"}
# A value containing any of these words is a billing ENTITY, not a person —
# "SMITH FAMILY CLINIC LLC" can land in a provider-name column and must
# pass through untouched (re-casing an org name is always wrong).
_ORG_NAME_TOKENS = frozenset((
    "INC", "LLC", "PLLC", "LLP", "LTD", "CORP", "CORPORATION", "PC", "PLC",
    "HOSPITAL", "HOSP", "CENTER", "CTR", "CLINIC", "ASSOCIATES", "ASSOC",
    "GROUP", "GRP", "PARTNERS", "LABORATORY", "LABORATORIES", "LAB", "LABS",
    "PHARMACY", "HEALTHCARE", "MEDICAL", "SERVICES", "SVCS", "SYSTEM",
    "SYSTEMS", "INSTITUTE", "FOUNDATION", "UNIVERSITY", "COLLEGE", "DEPT",
    "DEPARTMENT", "AGENCY", "IMAGING", "RADIOLOGY",
))


def _recase_name_word(w: str) -> str:
    """Proper-case one name token: single letters stay capital (initials),
    Mc gets its second cap (MCDONALD → McDonald), and apostrophe/hyphen
    parts re-case independently (O'BRIEN → O'Brien, SMITH-JONES →
    Smith-Jones). Deliberately no Mac- rule — Macias/Mackey/Macon would be
    corrupted far more often than MacDonald would be helped."""
    out = []
    for p in re.split(r"([\-'’])", w):
        if p in ("-", "'", "’"):
            out.append(p)
        elif len(p) <= 1:
            out.append(p.upper())
        elif len(p) > 2 and p[:2].upper() == "MC":
            out.append("Mc" + p[2].upper() + p[3:].lower())
        else:
            out.append(p[0].upper() + p[1:].lower())
    return "".join(out)


def _clean_provider_name_cell(v: str) -> Tuple[str, List[str], List[str]]:
    """Re-case a shouting/whispering person name and parse credentials.

    Returns ``(value, rules_hit, credentials_seen)``. The re-case fires
    ONLY when the whole cell is upper- or lower-case — mixed case means a
    human already curated it — and never when the value looks like an
    organization or contains digits. Credentials are reported either way.
    """
    if not v:
        return v, [], []
    up_toks = [t for t in
               (t.replace(".", "") for t in re.split(r"[ ,;/]+", v.upper()))
               if t]
    # Credentials never lead a name — "DO, HANH" / "PA, MINH" are
    # surnames (Do, Pa), not a Doctor of Osteopathy or a PA.
    creds = [c for c in dict.fromkeys(up_toks[1:]) if c in _CREDENTIALS]
    if any(t in _ORG_NAME_TOKENS for t in up_toks):
        return v, [], []
    if any(ch.isdigit() for ch in v):
        return v, [], creds
    if not any(ch.isalpha() for ch in v) or v not in (v.upper(), v.lower()):
        return v, [], creds
    out = []
    _widx = 0
    for piece in re.split(r"([ ,;/]+)", v):
        if not piece or piece[0] in " ,;/":
            out.append(piece)
            continue
        plain = piece.replace(".", "").upper()
        if _widx > 0 and plain in _CREDENTIALS:
            out.append(plain)                       # "m.d." → "MD"
        elif _widx > 0 and plain in _NAME_SUFFIXES:
            out.append(_NAME_SUFFIXES[plain])
        else:
            out.append(_recase_name_word(piece))
        _widx += 1
    new = "".join(out)
    if new != v:
        return new, ["provider-name-format"], creds
    return v, [], creds


def _payer_key(v: str) -> str:
    """Fold a payer name to a clustering key: uppercase, punctuation → space,
    collapsed whitespace, then family-prefix aliasing (BCBS OF TX and
    Blue Cross Texas land in the same cluster). Used for the report-only
    variant analysis — the cell itself is never rewritten."""
    k = re.sub(r"[^A-Z0-9 ]", " ", v.upper())
    k = re.sub(r"\s+", " ", k).strip()
    compact = k.replace(" ", "")
    for tok, family in _PAYER_FAMILIES:
        if compact.startswith(tok):
            return family
    return k


def _quantile(sorted_vals: List[float], q: float) -> float:
    """Linear-interpolation quantile (type-7, numpy's default) on a
    pre-sorted ascending list. Mirrors the analysis page's JS helper so the
    outlier fences here agree with the box-plot there."""
    n = len(sorted_vals)
    if n == 1:
        return sorted_vals[0]
    pos = (n - 1) * q
    base = int(pos)
    rest = pos - base
    lo = sorted_vals[base]
    hi = sorted_vals[base + 1] if base + 1 < n else sorted_vals[base]
    return lo + rest * (hi - lo)


# Which hyphenated segment to zero-pad, by segment-length signature. Mirrors
# npi_recovery.field_validators.normalize_ndc11 so verdicts match the package.
_NDC_PAD_MAP = {(4, 4, 2): 0, (5, 3, 2): 1, (5, 4, 1): 2}


def _clean_ndc_cell(v: str) -> Tuple[str, List[str]]:
    """Normalize an NDC to the 11-digit billing form. Segment-aware padding of
    hyphenated 10-digit forms (4-4-2 / 5-3-2 / 5-4-1 → 5-4-2). An unhyphenated
    10-digit NDC is AMBIGUOUS (segmentation unknown) and left unchanged — the
    caller flags it rather than guess a left-pad."""
    if v == "":
        return v, []
    s = v.strip()
    if "-" in s:
        segs = s.split("-")
        if len(segs) == 3 and all(seg.isdigit() for seg in segs):
            lens = tuple(len(seg) for seg in segs)
            if lens == (5, 4, 2):
                out = "".join(segs)
                return out, ([] if out == v else ["ndc-normalize-11"])
            if lens in _NDC_PAD_MAP:
                i = _NDC_PAD_MAP[lens]
                segs = list(segs)
                segs[i] = segs[i].zfill(len(segs[i]) + 1)
                return "".join(segs), ["ndc-pad-11"]
        digits = "".join(c for c in s if c.isdigit())
        if len(digits) == 11:
            return digits, (["ndc-normalize-11"] if digits != v else [])
        return v, []
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 11 and digits != v:
        return digits, ["ndc-normalize-11"]
    return v, []


def _ndc_malformed(v: str) -> bool:
    """True when a present NDC can't be a valid drug code: not 10 or 11
    digits after stripping separators, or a 3-part hyphenated form whose
    segment lengths aren't a known NDC layout. 10-digit is valid (just
    segmentation-ambiguous, flagged separately); blank passes."""
    s = v.strip()
    if not s:
        return False
    if "-" in s:
        segs = s.split("-")
        if len(segs) == 3 and all(seg.isdigit() for seg in segs):
            lens = tuple(len(seg) for seg in segs)
            return lens != (5, 4, 2) and lens not in _NDC_PAD_MAP
        # A hyphenated value that isn't a clean 3-part numeric NDC.
        digits = "".join(c for c in s if c.isdigit())
        return len(digits) not in (10, 11)
    digits = "".join(c for c in s if c.isdigit())
    # A pure-digit NDC must be 10 or 11; anything else is keying damage.
    return (not s.replace(" ", "").isdigit()) or len(digits) not in (10, 11)


import hashlib

# Distinct-row digests tracked for cross-chunk dedupe on a streamed run
# (bigfile.py). 2M ints is ~100 MB peak — bounded, honest: past the cap,
# NEW distinct rows stop being tracked and the run says so.
_STREAM_SEEN_CAP = 2_000_000

# Change-log entries kept IN MEMORY on the result (UI preview + workbook
# Issues context). The change-log FILE is never capped: entries past this
# preview spill straight to disk during the row loop, so a 10M-row run
# gets a complete audit trail without holding it in RAM. The old design
# capped the file itself at this number — users with big files got a
# silently truncated audit CSV.
_CHANGELOG_PREVIEW = 20_000


class _ChangelogSpill:
    """Disk spill for change-log entries past the in-memory preview.

    Owns its temp file for the WHOLE run: if the run dies before
    _write_output consumes the spill — user cancel (JobCancelled rides
    the progress callback), a network error in enrich mode, disk full —
    the finalizer closes the handle and unlinks the file. The spill
    holds PRE-de-identification cell values, so an orphaned copy in
    WORKDIR would be unmasked PHI sitting in /tmp; cleanup cannot
    depend on the happy path reaching _write_output.
    """

    def __init__(self) -> None:
        self.fh = None
        self.writer = None
        self.path: Optional[Path] = None
        self.failed = False

    def write(self, row: List[object]) -> None:
        if self.failed:
            return
        try:
            if self.writer is None:
                WORKDIR.mkdir(parents=True, exist_ok=True)
                self.path = WORKDIR / f"spill_{uuid.uuid4().hex}.csv"
                self.fh = open(self.path, "w", newline="",
                               encoding="utf-8")
                self.writer = csv.writer(self.fh)
            self.writer.writerow(row)
        except OSError:
            # Disk over quota mid-run: drop the partial file so a broken
            # half-spill never masquerades as a complete audit trail.
            self.failed = True
            self.discard()

    def close(self) -> None:
        """Flush + close the write handle, keeping the file for
        _write_output to consume. A close failure counts as a spill
        failure (the tail of the file may be missing)."""
        if self.fh is not None and not self.fh.closed:
            try:
                self.fh.close()
            except OSError:
                self.failed = True
        if self.failed:
            self.discard()

    def discard(self) -> None:
        if self.fh is not None and not self.fh.closed:
            try:
                self.fh.close()
            except OSError:
                pass
        self.fh = None
        self.writer = None
        if self.path is not None:
            try:
                Path(self.path).unlink(missing_ok=True)
            except OSError:
                pass
        self.path = None

    __del__ = discard
# Flagged OUTPUT rows captured per rule (and per payer family) for the
# worklist downloads. 500 was actionable-list sized; at 10M-row scale a
# single rule can legitimately fire on hundreds of thousands of rows and
# a 500-row worklist reads as data loss. 50k ints per rule keeps worst-
# case memory bounded (~70 rules × 50k × 28B ≈ 100 MB absolute worst)
# while the sanity counters always report the TRUE totals.
_WORKLIST_CAP = 50_000


def _phi_kind(hn: str) -> Optional[str]:
    """Classify a header (folded key) as a patient-PHI field, or None.

    Deliberately conservative: provider name / NPI / practice phone are NOT
    PHI (a provider is a public entity, and NPI recovery relies on them), so
    generic phone/zip/address only count when the header is patient-scoped.
    SSN / DOB / MRN / email are treated as PHI regardless of scope.
    """
    if "ssn" in hn or "socialsecurity" in hn:
        return "redact"
    if "dateofbirth" in hn or hn == "dob" or "birthdate" in hn or "birthdt" in hn \
            or "dateofbirth" in hn:
        return "dob"
    if "mrn" in hn or "medicalrecord" in hn:
        return "hashid"
    if "email" in hn:
        return "redact"
    patient = any(p in hn for p in (
        "patient", "member", "subscriber", "beneficiary", "insured",
        "guarantor", "enrollee", "bene"))
    if patient:
        if "name" in hn:
            return "redact"
        if "dob" in hn or "birth" in hn:
            return "dob"
        if "zip" in hn or "postal" in hn:
            return "zip3"
        if "phone" in hn or "fax" in hn:
            return "redact"
        if "addr" in hn or "street" in hn:
            return "redact"
        if "id" in hn or "acct" in hn or "account" in hn or "number" in hn:
            return "hashid"
    return None


def _deid_value(v: str, kind: str, salt: str) -> str:
    """Apply a de-identification transform, preserving referential integrity
    for ids (same value → same token within a run) via a per-run salt."""
    if v == "":
        return v
    if kind == "hashid":
        h = hashlib.sha256((salt + v).encode("utf-8")).hexdigest()[:10].upper()
        return "PT-" + h
    if kind == "dob":
        m = _DATE_ISO_RE.match(v.strip())
        if m:
            return m.group(1)  # keep year only
        digs = "".join(c for c in v if c.isdigit())
        if len(digs) >= 4:
            for token in re.findall(r"(19|20)\d{2}", v):
                return token  # first plausible 4-digit year
        return "REDACTED"
    if kind == "zip3":
        digs = "".join(c for c in v if c.isdigit())
        return (digs[:3] + "XX") if len(digs) >= 3 else "REDACTED"
    return "REDACTED"


def _to_number(s: str) -> Optional[float]:
    """Parse a possibly-formatted numeric cell, or None."""
    if s is None or s == "":
        return None
    t = str(s).replace("$", "").replace(",", "").strip()
    m = _MONEY_PARENS_RE.match(t)
    if m:
        t = "-" + m.group(1)
    tn = _MONEY_TRAIL_NEG_RE.match(t)
    if tn:  # trailing-sign negative ("500.00-" / "500.00CR")
        t = "-" + tn.group(1)
    try:
        return float(t)
    except ValueError:
        return None


def _row_sanity_flags(row, allowed_i, billed_i, paid_i, units_i,
                      high_units: int = 0) -> List[str]:
    """Cross-field impossibilities on a cleaned row (report-only).

    ``high_units`` (>0) turns on an MUE-style over-utilization flag for a
    units-per-line ceiling — opt-in via the profile because no ceiling is
    universal across procedures."""
    flags: List[str] = []

    def val(i):
        return _to_number(row[i]) if (i is not None and i < len(row)) else None

    a, b, p, u = val(allowed_i), val(billed_i), val(paid_i), val(units_i)
    if a is not None and b is not None and a > b + 1e-6:
        flags.append("allowed-exceeds-billed")
    if p is not None and a is not None and p > a + 1e-6:
        flags.append("paid-exceeds-allowed")
    if p is not None and b is not None and p > b + 1e-6:
        flags.append("paid-exceeds-billed")
    if a is not None and a < 0:
        flags.append("negative-allowed")
    if p is not None and p < 0:
        flags.append("negative-paid")
    if u is not None and u <= 0:
        flags.append("nonpositive-units")
    elif u is not None and abs(u - round(u)) > 1e-9:
        flags.append("fractional-units")
    if high_units > 0 and u is not None and u > high_units:
        flags.append("units-exceed-threshold")
    return flags

def _rule_catalog() -> List[Dict[str, str]]:
    """Rule registry for the UI/report — lazy so engine import stays cheap."""
    try:
        from . import rules as _rules
        return _rules.catalog()
    except Exception:  # noqa: BLE001
        return []


ProgressCb = Callable[[str, float], None]


def _norm_key(name: str) -> str:
    """Fold a header to a comparison key: lowercase, strip non-alphanumerics."""
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


def luhn_npi_valid(npi: object) -> bool:
    """True when a 10-digit NPI passes the Luhn check over ``80840`` + first 9.

    Mirrors ``npi_recovery.field_validators.luhn_npi_valid`` byte-for-byte so
    this offline engine agrees with the full package's verdicts.
    """
    s = "".join(ch for ch in str(npi) if ch.isdigit())
    if len(s) != 10:
        return False
    full = "80840" + s[:9]
    total = 0
    for i, ch in enumerate(reversed(full)):
        d = int(ch)
        if i % 2 == 0:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return (10 - total % 10) % 10 == int(s[9])


def _digits(v: object) -> str:
    return "".join(ch for ch in str(v) if ch.isdigit())


def classify_npi(raw: object) -> str:
    """One of: ``blank`` · ``malformed`` (not 10 digits, or an implausible
    first digit) · ``checksum`` (10 digits but Luhn-fails) · ``valid``.

    CMS only ever issues NPIs beginning with 1 (individual/entity) or 2
    (reserved) — a Luhn-valid 10-digit value starting 3-9 cannot be a real
    NPI and reads ``malformed`` rather than lending false confidence."""
    s = str(raw).strip() if raw is not None else ""
    if s == "" or s.lower() in ("nan", "none", "null", "na"):
        return "blank"
    d = _digits(s)
    if len(d) != 10:
        return "malformed"
    if d[0] not in "12":
        return "malformed"
    return "valid" if luhn_npi_valid(d) else "checksum"


# ---------------------------------------------------------------- data model --
@dataclass
class CleanResult:
    n_rows_in: int = 0
    n_rows_out: int = 0
    n_dupes_removed: int = 0
    npi_columns: List[str] = field(default_factory=list)
    billing_column: Optional[str] = None
    # Per-column tallies: col -> {"valid","blank","malformed","checksum","cells"}
    column_stats: Dict[str, Dict[str, int]] = field(default_factory=dict)
    n_cells_trimmed: int = 0
    # Deterministic normalization fixes applied, rule → count of cells changed.
    repairs: Dict[str, int] = field(default_factory=dict)
    # Cross-field data-sanity flags (not auto-fixed), rule → count of rows.
    sanity: Dict[str, int] = field(default_factory=dict)
    # PHI de-identification (opt-in): whether it ran, cells masked, columns hit.
    deid_applied: bool = False
    deid_cells: int = 0
    deid_columns: List[str] = field(default_factory=list)
    # Data-quality raw counters (drive the 0-100 report card).
    n_cells_total: int = 0
    n_cells_filled: int = 0
    # Cell-level audit trail: every change the cleaner made, capped so a
    # pathological file can't hold the log in memory. n_changes counts ALL
    # changes even past the cap.
    changelog: List[Tuple[int, str, str, str, str]] = field(default_factory=list)
    n_changes: int = 0
    changelog_truncated: bool = False
    changelog_path: Optional[str] = None
    changelog_name: str = "changelog.csv"
    # Disk spill of change-log entries past the in-memory preview
    # (_CHANGELOG_PREVIEW): a _ChangelogSpill guard that owns the temp
    # file. Consumed and discarded by _write_output; self-cleans via its
    # finalizer if the run dies first (the spill holds pre-de-id values).
    changelog_spill: Optional[object] = None
    # Payer-name variant clusters (report-only), per-HCPCS charge outliers,
    # and structural findings (duplicate headers / empty columns).
    payer_variants: Optional[Dict[str, object]] = None
    outliers: Optional[List[Dict[str, object]]] = None
    structure: Dict[str, object] = field(default_factory=dict)
    # Per-column fill rate (completeness profile): column → filled count + %.
    column_fill: List[Dict[str, object]] = field(default_factory=list)
    # Top denial / adjustment reason codes (revenue-cycle visibility).
    denials: Optional[Dict[str, object]] = None
    # Worklists: rule → 1-based OUTPUT row indices that fired it (capped per
    # rule) so flagged rows can be exported as actionable per-rule CSVs.
    flag_rows: Dict[str, List[int]] = field(default_factory=dict)
    # Credential tokens parsed from provider-name columns (MD, NP, PA …) →
    # count of cells carrying each. Report-only; counted per input row.
    credentials: Dict[str, int] = field(default_factory=dict)
    # Specialty mix: top taxonomy codes on kept rows with NUCC display
    # names where known. Report-only.
    specialties: List[Dict[str, object]] = field(default_factory=list)
    # Claim rollup (when a claim-id column exists): claim count, lines per
    # claim, per-claim charge distribution. Report-only.
    claims: Optional[Dict[str, object]] = None
    # Per-payer quality split: rows / flagged / clean % / top rules for
    # the top payer families. Report-only.
    payer_quality: List[Dict[str, object]] = field(default_factory=list)
    # Flagged OUTPUT row indices per payer family (capped) — powers the
    # per-payer worklist download (?fmt=worklist&payer=FAMILY).
    payer_flag_rows: Dict[str, List[int]] = field(default_factory=dict)
    # Zip batch mode: per-file summaries (grade, rows, repairs, findings).
    batch: List[Dict[str, object]] = field(default_factory=list)
    # Regression warnings vs the previous run of the same file (history).
    trend_alerts: List[str] = field(default_factory=list)
    # Data dictionary: per column — detected role, fill %, distinct count,
    # PHI-safe sample values. Exported via ?fmt=dictionary.
    dictionary: List[Dict[str, object]] = field(default_factory=list)
    # Profile applied to this run (see profiles.py): accepted rules still
    # report but don't count against the grade.
    accepted_rules: List[str] = field(default_factory=list)
    profile_name: Optional[str] = None
    delimiter: str = ","
    headers: List[str] = field(default_factory=list)
    out_path: Optional[str] = None
    out_name: str = "cleaned.csv"
    workbook_path: Optional[str] = None
    workbook_name: str = "report.xlsx"
    # Corrections companion (v49 suggested_fixes): row-level current→suggested
    # fixes with provenance. Written to its own CSV for download.
    companion_path: Optional[str] = None
    companion_name: str = "corrections.csv"
    suggestions_records: List[Dict[str, str]] = field(default_factory=list)
    # NPPES-recovered NPIs written into the cleaned output, keyed by the
    # 1-based row index → recovered NPI. Populated only when enrich resolves
    # a single confident candidate for a row's provider name+state.
    recovered_rows: Dict[int, str] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    # Real-engine findings from the vendored v48 field/consistency/dedup
    # screens (via vendor_adapter). None when pandas / the modules are
    # unavailable and we ran the stdlib path only.
    advanced: Optional[Dict[str, object]] = None
    # Live NPPES verify/recover results (via nppes_bridge). None unless the
    # user opted into the CMS cross-check. See nppes_bridge.py.
    nppes: Optional[Dict[str, object]] = None
    # Live drug connectors (RxNorm / openFDA) results, and the available-source
    # catalog. None unless online mode ran. See connectors.py.
    connectors: Optional[List[Dict[str, object]]] = None
    catalog: Optional[List[Dict[str, object]]] = None
    # Per-connector recommendation for THIS file (which sources apply + why),
    # computed from the detected columns even on an offline run so the panel
    # can explain "2 of 20 used" instead of looking broken. See
    # connectors.plan().
    connector_plan: Optional[List[Dict[str, object]]] = None
    # Compliance screens (OIG LEIE + Medicare PECOS). None unless online. See
    # compliance.py.
    compliance: Optional[List[Dict[str, object]]] = None
    # Ordering/referring provider NPI columns detected, and their NPPES
    # active-status screen (enrich mode) — a claim can deny when the ordering
    # or referring provider isn't a valid, active/enrolled provider.
    order_referring_columns: List[str] = field(default_factory=list)
    order_referring: Optional[Dict[str, object]] = None
    # Deep recovery (full v49 run_pipeline) result. None unless deep mode ran.
    deep: Optional[Dict[str, object]] = None
    deep_workbook_path: Optional[str] = None
    deep_workbook_name: str = "recovered.xlsx"
    # Streaming-chunk handoff (see bigfile.py): the cleaned table is returned
    # in memory instead of written to WORKDIR so the streamer can append it
    # to ONE master output file. Never set on a normal run.
    chunk_payload: Optional[Tuple[List[str], List[List[str]]]] = None
    # Population analytics (analytics.py): service mix, encounters,
    # chronic-condition prevalence, volume integrity, readmissions, coding
    # intensity. Report-only; None when the columns aren't there.
    population: Optional[Dict[str, object]] = None
    # Selectable enrichment (enrich.py): appended-column summary + report
    # marts (top codes/trend, key providers, MSA mix, Medicare benchmark).
    # None unless the upload selected enrichments.
    enrichment: Optional[Dict[str, object]] = None

    @property
    def total_npi_cells(self) -> int:
        return sum(c.get("cells", 0) for c in self.column_stats.values())

    @property
    def total_valid(self) -> int:
        return sum(c.get("valid", 0) for c in self.column_stats.values())

    @property
    def total_issues(self) -> int:
        return sum(
            c.get("blank", 0) + c.get("malformed", 0) + c.get("checksum", 0)
            for c in self.column_stats.values()
        )

    def billing_issue_count(self) -> int:
        if not self.billing_column:
            return 0
        c = self.column_stats.get(self.billing_column, {})
        return c.get("blank", 0) + c.get("malformed", 0) + c.get("checksum", 0)

    # Sanity rules that speak to VALIDITY (a value that can't be right) vs
    # CONSISTENCY (values that contradict each other). Drives the report card.
    _VALIDITY_RULES = ("hcpcs-malformed", "icd10-malformed", "money-unparseable",
                       "sex-invalid", "taxonomy-malformed", "pos-invalid",
                       "revenue-code-malformed", "carc-invalid",
                       "mbi-malformed", "condition-code-malformed",
                       "occurrence-code-malformed", "value-code-malformed",
                       "drg-malformed", "tob-malformed",
                       "discharge-status-invalid", "admission-type-invalid",
                       "modifier-unknown", "icd10-unknown-code",
                       "hcpcs-unknown-code", "admission-source-invalid",
                       "ndc-malformed", "taxonomy-unknown-code",
                       "date-unparseable", "ragged-row",
                       "npi-scientific-lossy", "modifier-malformed",
                       "leie-excluded-npi")
    _CONSISTENCY_RULES = ("allowed-exceeds-billed", "paid-exceeds-allowed",
                          "paid-exceeds-billed", "negative-allowed",
                          "negative-paid", "nonpositive-units",
                          "fractional-units", "date-in-future", "date-stale",
                          "zip-state-mismatch", "service-before-birth",
                          "discharge-before-admit", "ndc-ambiguous-10digit",
                          "charge-outlier", "jw-zero-units", "bilateral-units",
                          "conflicting-amount-claim",
                          "anesthesia-units-implausible",
                          "revenue-tob-mismatch", "timely-filing-risk",
                          "service-date-order", "discharge-status-final-bill",
                          "units-exceed-threshold", "npi-name-conflict",
                          "drg-on-professional")

    def quality(self) -> Dict[str, object]:
        """The 0-100 data-quality report card over the five classic DQ
        dimensions. Each dimension is a 0-1 ratio; the score is a weighted
        blend. Deliberately simple, deterministic math — a partner should be
        able to recompute the grade from the visible counts."""
        rows = max(self.n_rows_out, 1)
        completeness = (self.n_cells_filled / self.n_cells_total
                        if self.n_cells_total else 1.0)
        npi_bad = self.total_issues - sum(
            c.get("blank", 0) for c in self.column_stats.values())
        _accepted = set(self.accepted_rules)
        validity_hits = sum(self.sanity.get(r, 0)
                            for r in self._VALIDITY_RULES
                            if r not in _accepted)
        validity = max(0.0, 1.0 - min(1.0, (validity_hits + npi_bad) / rows))
        consistency_hits = sum(self.sanity.get(r, 0)
                               for r in self._CONSISTENCY_RULES
                               if r not in _accepted)
        consistency = max(0.0, 1.0 - min(1.0, consistency_hits / rows))
        dup = (self.n_dupes_removed
               + self.sanity.get("suspected-duplicate-claim", 0)
               + self.sanity.get("near-duplicate-row", 0)
               + self.sanity.get("possible-duplicate-service", 0))
        uniqueness = max(0.0, 1.0 - min(1.0, dup / max(self.n_rows_in, 1)))
        conformity = max(0.0, 1.0 - min(1.0, sum(self.repairs.values())
                                        / max(self.n_cells_filled, 1)))
        dims = {"completeness": completeness, "validity": validity,
                "consistency": consistency, "uniqueness": uniqueness,
                "conformity": conformity}
        score = round(100 * (0.25 * completeness + 0.25 * validity
                             + 0.20 * consistency + 0.15 * uniqueness
                             + 0.15 * conformity))
        letter = ("A" if score >= 93 else "B" if score >= 85 else
                  "C" if score >= 70 else "D" if score >= 55 else "F")
        return {"score": score, "letter": letter,
                "dimensions": {k: round(v * 100, 1) for k, v in dims.items()}}

    def _population_public(self) -> Optional[Dict[str, object]]:
        """Population marts minus the per-encounter records — the browser
        polls the scorecard, and 200k encounter rows belong in the CSV
        download, not the status JSON."""
        if not self.population:
            return None
        pub = dict(self.population)
        enc = pub.get("encounters")
        if isinstance(enc, dict):
            enc = dict(enc)
            enc.pop("records", None)
            pub["encounters"] = enc
        return pub

    def as_scorecard(self) -> Dict[str, object]:
        cells = self.total_npi_cells
        valid = self.total_valid
        health = round(100.0 * valid / cells, 1) if cells else 0.0
        return {
            "rows_in": self.n_rows_in,
            "rows_out": self.n_rows_out,
            "duplicates_removed": self.n_dupes_removed,
            "cells_trimmed": self.n_cells_trimmed,
            "repairs": dict(self.repairs),
            "repairs_total": sum(self.repairs.values()),
            "sanity": dict(self.sanity),
            "deid": ({"cells": self.deid_cells,
                      "columns": self.deid_columns} if self.deid_applied else None),
            "quality": self.quality(),
            "payer": self.payer_variants,
            "outliers": self.outliers,
            "structure": self.structure or None,
            "changelog_name": (self.changelog_name
                               if self.changelog_path else None),
            "changes_logged": self.n_changes,
            "changelog_truncated": self.changelog_truncated,
            "fill_rates": self.column_fill,
            "denials": self.denials,
            "credentials": self.credentials or None,
            "specialties": self.specialties or None,
            "claims": self.claims,
            "population": self._population_public(),
            "payer_quality": self.payer_quality or None,
            "payer_worklists": ({k: len(v) for k, v
                                 in self.payer_flag_rows.items()} or None),
            "batch": self.batch or None,
            "trend_alerts": self.trend_alerts or None,
            "dictionary": self.dictionary or None,
            "worklists": {k: len(v) for k, v in self.flag_rows.items()},
            "accepted_rules": self.accepted_rules,
            "profile": self.profile_name,
            "rule_catalog": _rule_catalog(),
            "npi_columns": self.npi_columns,
            "billing_column": self.billing_column,
            "npi_cells": cells,
            "npi_valid": valid,
            "npi_issues": self.total_issues,
            "billing_issues": self.billing_issue_count(),
            "health_pct": health,
            "column_stats": self.column_stats,
            "delimiter": {",": "comma", "\t": "tab", ";": "semicolon",
                          "|": "pipe", "xlsx": "xlsx (Excel)",
                          "x12": "X12 837 (EDI)",
                          "x835": "X12 835 (ERA)",
                          "zip": "zip batch"}.get(
                              self.delimiter, self.delimiter),
            "out_name": self.out_name,
            "workbook_name": self.workbook_name if self.workbook_path else None,
            "companion_name": self.companion_name if self.companion_path else None,
            "companion_n": len(self.suggestions_records),
            "recovered_written": len(self.recovered_rows),
            "warnings": self.warnings,
            "advanced": self.advanced,
            "nppes": self.nppes,
            "connectors": self.connectors,
            "catalog": self.catalog,
            "connector_plan": self.connector_plan,
            "compliance": self.compliance,
            "order_referring": self.order_referring,
            "order_referring_columns": self.order_referring_columns or None,
            "deep": self.deep,
            "deep_workbook_name": (self.deep_workbook_name
                                   if self.deep_workbook_path else None),
            "enrichment": self.enrichment,
        }


# ------------------------------------------------------------------ decoding --
def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        return dialect.delimiter
    except csv.Error:
        # Fall back to whichever candidate the header line uses most.
        first = sample.splitlines()[0] if sample else ""
        return max(",\t;|", key=first.count) if first else ","


# Leading characters Excel/Sheets treat as a formula — a CSV export of
# untrusted claims data must neutralize these to avoid CSV-injection.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


_NUMERIC_RE = re.compile(r"^[+-]?\.?\d[\d,]*\.?\d*$")


def _defang_cell(value: str) -> str:
    """Prefix a lone quote when a CSV cell would otherwise start a formula.

    Pure numbers (including a leading ``-``/``+``/``.`` on a value like
    ``-50.00``) are left alone — Excel reads them as numbers, not formulas —
    so normalized money keeps its natural form.
    """
    if value and value[0] in _FORMULA_LEAD:
        if value[0] in "+-." and _NUMERIC_RE.match(value):
            return value
        return "'" + value
    return value


# A zip bomb defeats the HTTP upload cap because that cap sees COMPRESSED
# bytes: a ~200 KB archive of zeros expands to gigabytes and OOM-kills the
# process (reproduced before this guard existed). Enforced twice — from
# the declared sizes before any read, and again with a hard cap while
# decompressing, because central-directory sizes can lie.
_BATCH_MAX_UNCOMPRESSED = 200 * 1024 * 1024


def zip_batch_probe(data: bytes) -> bool:
    """Metadata-only "is this a zip batch?" — names and declared sizes,
    never member decompression. The detect route needs only the boolean
    and was paying full extraction (twice, with the upload) to get it."""
    if data[:4] != b"PK\x03\x04":
        return False
    import zipfile as _zf
    try:
        names = _zf.ZipFile(io.BytesIO(data)).namelist()
    except Exception:  # noqa: BLE001
        return False
    if "[Content_Types].xml" in names:
        return False
    return any(not n.endswith("/") and not n.startswith("__MACOSX")
               and n.rsplit("/", 1)[-1][:1] != "."
               and n.lower().endswith((".csv", ".tsv", ".txt", ".837",
                                       ".835", ".edi", ".x12"))
               for n in names)


# Zip-batch member cap: files past this many are skipped (and now SAID so).
_BATCH_MEMBER_CAP = 50


def zip_batch_members(data: bytes) -> Optional[List[Tuple[str, bytes]]]:
    """Members of a multi-file zip batch, or None when this isn't one.
    Back-compat wrapper over ``zip_batch_members_ex`` (which also reports
    what was skipped)."""
    members, _info = zip_batch_members_ex(data)
    return members


def zip_batch_members_ex(data: bytes) -> Tuple[
        Optional[List[Tuple[str, bytes]]], Dict[str, object]]:
    """(members, skip-info) of a multi-file zip batch; members is None when
    this isn't one. ``skip-info`` carries ``skipped_unsupported`` (member
    names excluded by format) and ``over_cap`` (count past the
    {cap}-member batch cap) so the batch banner can say what was NOT
    cleaned instead of silently dropping files.

    An .xlsx is ALSO a zip — the tell is ``[Content_Types].xml`` in the
    archive root, so that (and any other Office package) is excluded
    before the cheaper suffix scan. Only claim-shaped members count;
    directory entries and macOS resource forks are skipped silently
    (they are packaging noise, not data).

    Raises ``ValueError`` when the archive IS a batch but expands past
    the uncompressed cap — the caller turns that into a clear warning
    instead of sniffing the zip as CSV garbage."""
    info: Dict[str, object] = {"skipped_unsupported": [], "over_cap": 0,
                               "cap": _BATCH_MEMBER_CAP}
    if data[:4] != b"PK\x03\x04":
        return None, info
    import zipfile as _zf
    try:
        zf = _zf.ZipFile(io.BytesIO(data))
        infos = zf.infolist()
    except Exception:  # noqa: BLE001 — truncated/hostile zip → not a batch
        return None, info
    names = [i.filename for i in infos]
    if "[Content_Types].xml" in names:
        return None, info                 # Office package (xlsx/docx/…)
    real_files = [i for i in infos
                  if not i.filename.endswith("/")
                  and not i.filename.startswith("__MACOSX")
                  and i.filename.rsplit("/", 1)[-1][:1] != "."]
    members = [i for i in real_files
               if i.filename.lower().endswith(
                   (".csv", ".tsv", ".txt", ".837", ".835", ".edi", ".x12"))]
    if not members:
        return None, info
    info["skipped_unsupported"] = sorted(
        i.filename.rsplit("/", 1)[-1] for i in real_files
        if i not in members)
    members = sorted(members, key=lambda i: i.filename)
    info["over_cap"] = max(0, len(members) - _BATCH_MEMBER_CAP)
    members = members[:_BATCH_MEMBER_CAP]
    declared = sum(max(i.file_size, 0) for i in members)
    if declared > _BATCH_MAX_UNCOMPRESSED:
        raise ValueError(
            f"This zip declares {declared / 1e6:,.0f} MB of uncompressed "
            f"data — the batch limit is "
            f"{_BATCH_MAX_UNCOMPRESSED / 1e6:,.0f} MB total. Split the "
            "archive and upload in parts.")
    out = []
    budget = _BATCH_MAX_UNCOMPRESSED
    for m_info in members:
        try:
            # Chunked read with a hard cap: declared sizes can be forged,
            # so never trust file_size alone.
            chunks = []
            got = 0
            with zf.open(m_info) as fh:
                while True:
                    chunk = fh.read(1 << 20)
                    if not chunk:
                        break
                    got += len(chunk)
                    if got > budget:
                        raise ValueError(
                            "A zip member expanded past its declared size "
                            "beyond the batch limit — refusing to "
                            "decompress further.")
                    chunks.append(chunk)
            budget -= got
            out.append((m_info.filename.rsplit("/", 1)[-1],
                        b"".join(chunks)))
        except ValueError:
            raise
        except Exception:  # noqa: BLE001 — skip the unreadable member
            continue
    return (out or None), info


def _looks_like_xlsx(data: bytes) -> bool:
    # .xlsx is a zip; the local-file-header magic is "PK\x03\x04". A CSV that
    # happens to start with "PK" is vanishingly unlikely to also be valid zip.
    return data[:4] == b"PK\x03\x04"


def _pick_xlsx_sheet(wb) -> Tuple[str, List[Tuple[str, int, int]]]:
    """Choose the worksheet that actually holds the claims table.

    Vendor extracts routinely lead with a cover/'Detail'/notes sheet; the
    real data lives on a later tab and is, by far, the sheet with the
    most populated cells. Reading only the first sheet silently cleaned
    a 3-column cover page while a 13M-cell 'DATA' tab sat ignored — the
    incident this function exists for. Scored from declared dimensions
    plus a 50-row sample scan (a sheet with a stale/absent dimension
    record still registers), so huge sheets aren't fully parsed twice.

    Returns (chosen sheet name, [(name, est_rows, est_cols), …]).
    """
    stats: List[Tuple[str, int, int]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        sampled = 0
        max_seen_cols = 0
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 50:
                break
            if r and any(c is not None and str(c).strip() for c in r):
                sampled += 1
                max_seen_cols = max(
                    max_seen_cols,
                    sum(1 for c in r if c is not None and str(c).strip()))
        if sampled == 0:
            stats.append((name, 0, 0))
            continue
        est_rows = max(int(ws.max_row or 0), sampled)
        est_cols = max(int(ws.max_column or 0), max_seen_cols)
        stats.append((name, est_rows, est_cols))
    best = max(stats, key=lambda t: t[1] * max(t[2], 1))
    return best[0], stats


def _xlsx_best_sheet(data: bytes) -> Optional[str]:
    """The chosen sheet name for a workbook's raw bytes, or None when the
    workbook can't be inspected. Shared with the pandas paths
    (vendor_adapter) so the mapping editor, the v49 engine and the stdlib
    pipeline all read the SAME sheet."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            name, _stats = _pick_xlsx_sheet(wb)
        finally:
            wb.close()
        return name
    except Exception:  # noqa: BLE001
        return None


def _read_xlsx(data: bytes) -> Tuple[List[str], List[List[str]],
                                     Optional[str]]:
    """Read the data-bearing worksheet of an .xlsx via openpyxl
    (read-only mode) — see _pick_xlsx_sheet for how it's chosen. Returns
    (headers, rows, note); the note names the sheet used and the sheets
    skipped whenever the workbook has more than one non-empty sheet.

    Raises if openpyxl is unavailable or the workbook is unreadable — the
    caller turns that into a friendly warning telling the user to export CSV.
    """
    from openpyxl import load_workbook  # base dependency

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet_name, stats = _pick_xlsx_sheet(wb)
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for r in ws.iter_rows(values_only=True):
            if r is None:
                continue
            cells = ["" if c is None else str(c) for c in r]
            if any(c.strip() for c in cells):
                rows.append(cells)
    finally:
        wb.close()
    note = None
    others = [(n, er, ec) for n, er, ec in stats if n != sheet_name]
    if others:
        skipped = ", ".join(
            (f"'{n}' ({er:,}×{ec})" if er else f"'{n}' (empty)")
            for n, er, ec in others)
        note = (f"Workbook has {len(stats)} sheets — cleaned "
                f"'{sheet_name}' ({max(len(rows) - 1, 0):,} data rows, the "
                f"largest table). Skipped: {skipped}. If a different sheet "
                "is the one you meant, export it as CSV and re-upload.")
    if not rows:
        return [], [], note
    headers, body, shape_notes = _shape_rows(rows)
    if shape_notes:
        note = "; ".join(([note] if note else []) + shape_notes)
    return headers, body, note


# BOM → codec for wide-character uploads. UTF-32 BOMs are checked FIRST:
# the UTF-32-LE BOM starts with the UTF-16-LE BOM bytes. Excel's "Unicode
# Text" export is exactly UTF-16-LE + BOM + tabs — without this sniff those
# files decoded to NUL-riddled garbage and were "cleaned" anyway.
_WIDE_BOMS: Tuple[Tuple[bytes, str, str], ...] = (
    (b"\xff\xfe\x00\x00", "utf-32-le", "UTF-32 (little-endian)"),
    (b"\x00\x00\xfe\xff", "utf-32-be", "UTF-32 (big-endian)"),
    (b"\xff\xfe", "utf-16-le", "UTF-16 (little-endian)"),
    (b"\xfe\xff", "utf-16-be", "UTF-16 (big-endian)"),
)
# BOM-less UTF-16 heuristic: NUL bytes above this share of the first 4 KB
# mean a wide encoding (ASCII text has essentially none).
_NUL_DENSITY_MIN = 0.20


def _wide_probe(head: bytes) -> Optional[Tuple[str, int, str]]:
    """(codec, bom_length, human label) when the head bytes say UTF-16/32,
    else None. BOM first; failing that, NUL-density + NUL position decide
    (text bytes land at even offsets for LE, odd for BE)."""
    for bom, enc, label in _WIDE_BOMS:
        if head.startswith(bom):
            return enc, len(bom), label
    probe = head[:4096]
    if probe and probe.count(0) / len(probe) >= _NUL_DENSITY_MIN:
        even_nuls = probe[0::2].count(0)
        odd_nuls = probe[1::2].count(0)
        if odd_nuls >= even_nuls:
            return "utf-16-le", 0, "UTF-16 (little-endian, no BOM)"
        return "utf-16-be", 0, "UTF-16 (big-endian, no BOM)"
    return None


def _decode_table_bytes(data: bytes) -> Tuple[str, Optional[str]]:
    """bytes → (text, note). UTF-16/32 (BOM or NUL-density) decodes to real
    text with a note naming the encoding; then utf-8-sig; latin-1 last."""
    wide = _wide_probe(data[:4096])
    if wide is not None:
        enc, bom_len, label = wide
        try:
            return (data[bom_len:].decode(enc, errors="replace"),
                    f"File was {label}-encoded (Excel's 'Unicode Text' "
                    "export produces this) — decoded automatically. "
                    "Prefer CSV UTF-8 exports.")
        except Exception:  # noqa: BLE001 — fall through to the classic path
            pass
    try:
        return data.decode("utf-8-sig"), None
    except UnicodeDecodeError:
        return data.decode("latin-1", errors="replace"), None


def _populated_cells(row: List[str]) -> int:
    return sum(1 for c in row if c and c.strip())


_NUMERIC_CELL_RE = re.compile(r"^-?\d+(\.\d+)?$")


def _shape_rows(all_rows: List[List[str]]
                ) -> Tuple[List[str], List[List[str]], List[str]]:
    """(headers, body, notes) from raw parsed rows, with two bounded
    heuristics for files that don't lead with a clean header:

      * PREAMBLE: cover/title lines ("Claims Extract Q3 2025") with ≤2
        populated cells above a table whose rows consistently carry ≥4 are
        skipped (max 3) — otherwise every real column vanishes into a
        one-column table.
      * HEADERLESS: when the would-be header row is ≥60% numeric-shaped
        cells, it is DATA — ``column_1..N`` names are synthesized and the
        row is kept, instead of promoting NPIs to header text.

    Both are warned about; neither ever fires on an ordinary headered file.
    """
    notes: List[str] = []
    skipped = 0
    while (skipped < 3 and len(all_rows) >= 4
           and _populated_cells(all_rows[0]) <= 2):
        # Row 0 is a title only when a STABLE ≥4-wide table starts within
        # the next few rows: everything before the first wide row is
        # narrow (more title lines), everything from it on is wide. An
        # ordinary 2-column file never matches (no wide row follows).
        widths = [_populated_cells(r) for r in all_rows[1:5]]
        wide_at = next((i for i, w in enumerate(widths) if w >= 4), None)
        if (wide_at is None
                or any(w > 2 for w in widths[:wide_at])
                or any(w < 4 for w in widths[wide_at:])):
            break
        all_rows = all_rows[1:]
        skipped += 1
    if skipped:
        notes.append(
            f"Skipped {skipped} title row(s) above the header — the file "
            "leads with a cover line, not column names.")
    if not all_rows:
        return [], [], notes
    hdr = [h.strip() for h in all_rows[0]]
    populated = [c for c in hdr if c]
    numeric = sum(1 for c in populated if _NUMERIC_CELL_RE.match(c))
    if len(populated) >= 2 and numeric / len(populated) >= 0.6:
        headers = [f"column_{i + 1}" for i in range(len(hdr))]
        notes.append(
            "File appears to have no header row — column_1…column_"
            f"{len(hdr)} names were synthesized and the first row was "
            "kept as data. Add a header row to name the columns.")
        return headers, all_rows, notes
    return hdr, all_rows[1:], notes


def _read_table(data: bytes, *, reshape: bool = True
                ) -> Tuple[List[str], List[List[str]], str,
                           Optional[str]]:
    """Decode bytes → (headers, rows, format, note). Handles CSV/TSV and
    .xlsx. ``format`` is the delimiter for text files, or ``"xlsx"`` for
    spreadsheets; ``note`` is a user-facing remark about how the file was
    read (worksheet chosen, encoding detected, title rows skipped, …).

    ``reshape=False`` disables the title-row / headerless heuristics.
    The streaming path needs that: bigfile prepends the FILE's first
    record to every chunk as its header line, so shaping must happen
    ONCE at the stream level (bigfile._shape_stream_head) — running it
    per chunk made chunk 2+ treat that prepended record as a title/data
    row, silently eating or replaying one record per chunk.
    """
    if _looks_like_xlsx(data):
        headers, body, note = _read_xlsx(data)
        return headers, body, "xlsx", note
    from . import x12 as _x12
    if _x12.looks_like_x12(data):
        parsed = _x12.x12_to_table(data)
        if parsed is not None:
            return parsed[0], parsed[1], "x12", None
        parsed = _x12.x835_to_table(data)
        if parsed is not None:
            return parsed[0], parsed[1], "x835", None
        return [], [], "x12", None    # X12 but not 837/835 — warning later
    text, enc_note = _decode_table_bytes(data)
    sample = text[:8192]
    delim = _sniff_delimiter(sample)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    all_rows = [r for r in reader if r != []]
    notes = [enc_note] if enc_note else []
    if not all_rows:
        return [], [], delim, ("; ".join(notes) or None)
    if reshape:
        headers, body, shape_notes = _shape_rows(all_rows)
        notes.extend(shape_notes)
    else:
        headers, body = [h.strip() for h in all_rows[0]], all_rows[1:]
    return headers, body, delim, ("; ".join(notes) or None)


def _header_tokens(header: str) -> List[str]:
    """Split a header into lowercase word tokens across delimiters, camelCase,
    and letter↔digit boundaries so ``OptionCareNPIFlag`` and
    ``OPTIONCARE_NPI_FLAG`` both surface a ``flag`` token. Used to tell an NPI
    VALUE column from a derived NPI flag/status column."""
    if not header:
        return []
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", header)          # camelCase
    s = re.sub(r"(?<=[A-Za-z])(?=[0-9])|(?<=[0-9])(?=[A-Za-z])", " ", s)
    return [p.lower() for p in re.split(r"[^A-Za-z0-9]+", s) if p]


def _is_npi_modifier_column(header: str, key: str) -> bool:
    """True when a header that mentions "npi" is actually a derived companion
    (a flag/status/type/name/date) rather than a column holding the NPI value.
    Prevents boolean/indicator columns from being scored as malformed NPIs."""
    toks = _header_tokens(header)
    if any(t in _NPI_MODIFIER_TOKENS for t in toks):
        return True
    # Delimiter-free blob (e.g. "OPTIONCARENPIFLAG" → one token): fall back to
    # a suffix test with tokens long enough to be unambiguous.
    if len(toks) <= 1:
        for t in _NPI_MODIFIER_TOKENS:
            if len(t) >= 4 and key.endswith(t):
                return True
    return False


def _detect_npi_columns(headers: List[str]) -> Tuple[List[int], Optional[int]]:
    """Return (indices of NPI columns, index of the billing NPI column or None).

    A header must both mention "npi" AND not be a derived NPI companion
    (``*_NPI_FLAG``, ``NPI_Status``, ``NPI_Type`` …) — see
    ``_is_npi_modifier_column``. Without that guard a Y/N flag column was
    scored as 600k malformed NPIs and drove Validity to 0% on a real file."""
    npi_idx: List[int] = []
    billing_idx: Optional[int] = None
    for i, h in enumerate(headers):
        key = _norm_key(h)
        if not (any(hint in key for hint in _NPI_HINTS) or key == "npi"):
            continue
        if _is_npi_modifier_column(h, key):
            continue
        npi_idx.append(i)
        if billing_idx is None and any(b in key for b in _BILLING_HINTS):
            billing_idx = i
    # No explicit billing column? Treat the first NPI column as billing.
    if billing_idx is None and npi_idx:
        billing_idx = npi_idx[0]
    return npi_idx, billing_idx


# HCPCS Level II drug code ("J-code", e.g. J1745 = infliximab). These live in
# the procedure/HCPCS column and ARE drug identifiers even when a file carries
# no NDC and no free-text drug name — the signal that an infusion-pharmacy or
# oncology extract needs the drug connectors.
_JCODE_RE = re.compile(r"^J\d{4}$")


def _jcode_stats(rows: List[List[str]], hcpcs_idx: Optional[int], *,
                 scan_cap: int = 50000, distinct_cap: int = 400
                 ) -> Tuple[float, List[str]]:
    """(% of populated HCPCS cells that are J-codes over a bounded scan,
    per-CELL J-code values for the RxNorm crosswalk). Bounded so a 10M-row
    file can't turn connector planning into a full extra pass: the scan cap
    bounds the list length and ``distinct_cap`` bounds the distinct-value
    universe. The list is cell-grain (not distinct) on purpose —
    resolve_drugs dedupes before any lookup, and passing cells makes its
    ``rows_seen`` coverage counter mean CELLS for J-codes exactly as it
    does for NDCs and drug names (one canonical unit across the scorecard)."""
    if hcpcs_idx is None:
        return 0.0, []
    seen: List[str] = []
    seen_set: set = set()
    populated = jcount = 0
    for n, row in enumerate(rows):
        if n >= scan_cap:
            break
        if hcpcs_idx >= len(row):
            continue
        val = (row[hcpcs_idx] or "").strip().upper()
        if not val:
            continue
        populated += 1
        if _JCODE_RE.match(val):
            jcount += 1
            if val in seen_set or len(seen_set) < distinct_cap:
                seen_set.add(val)
                seen.append(val)
    pct = (100.0 * jcount / populated) if populated else 0.0
    return pct, seen


def _detect_one(headers: List[str], hints: tuple) -> Optional[int]:
    """First header whose folded key contains any hint (hint order = priority)."""
    keys = [_norm_key(h) for h in headers]
    for hint in hints:
        for i, k in enumerate(keys):
            if hint in k:
                return i
    return None


# How many rows the value-based amount sniffer reads. Bounded so the fallback
# stays cheap at 10M-row scale (sample, never full-scan).
_AMOUNT_SNIFF_SAMPLE = 4000

# Value-based NPI column sniff bounds — mirrors the amount sniffer's
# conservatism: bounded sample, high thresholds, unclaimed columns only.
_NPI_SNIFF_SAMPLE = 4000
_NPI_SNIFF_MIN_POPULATED = 3
_NPI_SNIFF_TEN_FRAC = 0.90
_NPI_SNIFF_LUHN_FRAC = 0.85


def _sniff_npi_columns(headers: List[str], rows: List[List[str]],
                       claimed: set, *,
                       sample: int = _NPI_SNIFF_SAMPLE) -> List[int]:
    """Columns that hold NPI VALUES under headers the hint list misses
    (PROV_ID, ProviderNumber, Perf_Prov). Over a bounded sample, an
    unclaimed column whose populated cells are ≥90% bare 10-digit values
    and ≥85% Luhn-valid is adopted for stats + screens (never for cell
    rewriting, and never displacing an explicit billing header). Without
    this, such a column got no Luhn screen, no column stats, and no LEIE
    screen at all."""
    out: List[int] = []
    for ci in range(len(headers)):
        if ci in claimed:
            continue
        populated = ten = luhn = 0
        for n, row in enumerate(rows):
            if n >= sample:
                break
            if ci >= len(row):
                continue
            cell = row[ci].strip()
            if not cell:
                continue
            populated += 1
            if len(cell) == 10 and cell.isdigit():
                ten += 1
                if cell[0] in "12" and luhn_npi_valid(cell):
                    luhn += 1
        if (populated >= _NPI_SNIFF_MIN_POPULATED
                and ten / populated >= _NPI_SNIFF_TEN_FRAC
                and luhn / populated >= _NPI_SNIFF_LUHN_FRAC):
            out.append(ci)
    return out


def _sniff_amount_column(
    headers: List[str], rows: List[List[str]], claimed: set, *,
    sample: int = _AMOUNT_SNIFF_SAMPLE,
) -> Optional[int]:
    """Find the column most likely to hold claim AMOUNTS by value, for the
    Population/analytics charge role ONLY — used as a last resort when no
    billed/allowed/paid/other money header was detected (e.g. a charge column
    named ``NET_SVC_VAL`` that no header hint matches).

    Deliberately conservative so it never *invents* money:

      * A candidate must be predominantly parseable as money (``_to_number``
        over a bounded sample) AND carry monetary evidence — decimal cents on
        a real share of cells, or a ``$``/thousands-grouped format. A bare
        integer column (units, counts, ages, sequence numbers, numeric codes)
        therefore does NOT qualify, and a de-identified extract whose amounts
        were stripped keeps the role empty (charges stay ``$0``).
      * Columns already claimed by any other role (NPI / date / zip / phone /
        HCPCS / dx / modifier / taxonomy / NDC / revenue / POS / DRG / state /
        units / patient / member / claim / …) are never adopted — ``claimed``
        carries every such index.

    Returns the winning column index, or ``None`` when nothing qualifies.
    NEVER feeds the deterministic cleaning pass (money_set / date_set / the
    money-unparseable screen) — the caller uses it for analytics only.
    """
    ncols = len(headers)
    best: Optional[int] = None
    best_score: Optional[tuple] = None
    for ci in range(ncols):
        if ci in claimed:
            continue
        nonblank = money = decimal = nonzero = 0
        total_abs = 0.0
        for n, row in enumerate(rows):
            if n >= sample:
                break
            if ci >= len(row):
                continue
            cell = row[ci].strip()
            if not cell:
                continue
            nonblank += 1
            num = _to_number(cell)
            if num is None:
                continue
            money += 1
            if "." in cell or "$" in cell or "," in cell:
                decimal += 1
            if abs(num) > 1e-9:
                nonzero += 1
                total_abs += abs(num)
        if nonblank < 3 or money == 0:
            continue
        money_frac = money / nonblank
        decimal_frac = decimal / nonblank
        # Predominantly money-shaped, with real monetary formatting evidence,
        # and not an all-zero column (which would add nothing anyway).
        if money_frac < 0.9 or decimal_frac < 0.3 or nonzero == 0:
            continue
        mean_abs = total_abs / nonzero
        score = (round(money_frac, 3), round(decimal_frac, 3), mean_abs)
        if best_score is None or score > best_score:
            best_score = score
            best = ci
    return best


def _analytics_amount_col(
    headers: List[str], rows: List[List[str]], *,
    billed_i: Optional[int], allowed_i: Optional[int], paid_i: Optional[int],
    money_set: set, claimed: set,
) -> Optional[int]:
    """Pick the single amount column the Population/analytics marts sum for
    charges. Priority: billed → allowed → paid (header-detected) → any other
    header-detected money column (``amount``/``cost``/``fee``-flavored) →
    value-based sniff over unclaimed columns.

    Report-only: the returned index feeds ONLY analytics; it never changes the
    deterministic cleaning money_set/date_set or any existing behavior. The
    value sniff runs only when NO money header matched, so a genuinely
    amount-less (or de-identified) file yields ``None`` and charges stay ``$0``.
    """
    for i in (billed_i, allowed_i, paid_i):
        if i is not None:
            return i
    if money_set:
        # A money header exists but none of the billed/allowed/paid roles
        # matched (e.g. "NetAmount", "ServiceFee"). Prefer the most charge-like
        # by header wording so the marts sum a gross figure when there is one.
        _pri = ("charge", "billed", "gross", "total", "amount", "cost", "fee",
                "allowed", "paid")

        def _rank(idx: int) -> tuple:
            k = _norm_key(headers[idx])
            for r, tok in enumerate(_pri):
                if tok in k:
                    return (r, idx)
            return (len(_pri), idx)

        return min(money_set, key=_rank)
    return _sniff_amount_column(headers, rows, claimed)


# -------------------------------------------------------------------- cleaner --
def detect_columns_preview(data: bytes) -> Optional[Dict[str, object]]:
    """Detect the column→role mapping for the pre-clean mapping editor.

    Delegates to the v49 detector via ``vendor_adapter``; returns None when
    pandas / the vendored engine is unavailable (the page then skips the
    confirm step and cleans directly on auto-detection).
    """
    try:
        from . import vendor_adapter
        return vendor_adapter.detect(data)
    except Exception:  # noqa: BLE001
        return None


def clean_bytes(
    data: bytes,
    src_name: str,
    *,
    drop_duplicates: bool = True,
    enrich: bool = False,
    deep: bool = False,
    deid: bool = False,
    enrichments: Optional[List[str]] = None,
    profile: Optional[Dict[str, object]] = None,
    overrides: Optional[Dict[str, str]] = None,
    progress: Optional[ProgressCb] = None,
    _stream_chunk: bool = False,
    _stream_seen: Optional[set] = None,
    _stream_fold_seen: Optional[set] = None,
    _changelog_sink: Optional[Callable[[Tuple[int, str, str, str, str]],
                                       None]] = None,
) -> CleanResult:
    """Clean a delimited claims file given as raw bytes.

    Offline by default. When ``enrich`` is set, the distinct NPIs are verified
    against the live NPPES registry and rows with a missing/bad billing NPI but
    a provider name are run through NPPES recovery — both via the app's shared
    CMS connection (``nppes_bridge``), fully guarded.

    ``_stream_chunk`` is the internal huge-file mode (bigfile.py): the cleaned
    table is handed back on ``res.chunk_payload`` instead of written to disk,
    and per-run side effects that belong to the WHOLE file — history record,
    trend alerts, the pandas suggestions companion — are skipped so a 10 GB
    upload doesn't record 200 history rows for one run.

    ``_stream_seen`` is the streamer's SHARED duplicate-tracking set: when
    given, exact-dup membership uses compact row digests in that set instead
    of full-tuple keys in a per-call set, so duplicates die across chunk
    boundaries too. ``_stream_fold_seen`` is the same idea for the
    near-duplicate (case/whitespace-folded) signal, so a "SMITH" row in
    chunk 1 and its "Smith" twin in chunk 9 still count. The caller bounds
    both sets' growth (see bigfile.py).
    """
    def cb(msg: str, frac: float) -> None:
        if progress:
            progress(msg, frac)

    cb("Reading file", 0.05)
    try:
        _members, _zip_info = zip_batch_members_ex(data)
    except ValueError as _zexc:
        _res = CleanResult(delimiter="zip", headers=[])
        _res.warnings.append(str(_zexc))
        _res.out_name = _out_name(src_name)
        _write_output(_res, [], [])
        cb("Done", 1.0)
        return _res
    if _members is not None:
        return _clean_batch(_members, src_name,
                            drop_duplicates=drop_duplicates, deid=deid,
                            profile=profile, cb=cb, skip_info=_zip_info)
    try:
        headers, rows, delim, _read_note = _read_table(
            data, reshape=not _stream_chunk)
    except ImportError:
        _res = CleanResult(delimiter="xlsx", headers=[])
        _res.warnings.append(
            "This looks like an .xlsx file but the Excel reader isn't "
            "available on the server. Please export the sheet as CSV and "
            "re-upload.")
        _res.out_name = _out_name(src_name)
        _write_output(_res, [], [])
        cb("Done", 1.0)
        return _res
    except Exception as exc:  # noqa: BLE001 — malformed spreadsheet
        _res = CleanResult(delimiter="xlsx", headers=[])
        _res.warnings.append(f"Could not read the file: {exc}. If this is an "
                             "Excel file, try exporting it as CSV.")
        _res.out_name = _out_name(src_name)
        _write_output(_res, [], [])
        cb("Done", 1.0)
        return _res
    res = CleanResult(delimiter=delim, headers=headers)
    res.n_rows_in = len(rows)
    if _read_note:
        # e.g. "Workbook has 3 sheets — cleaned 'DATA' …". A user staring
        # at a wrong-sheet result must be able to see which tab was read.
        res.warnings.append(_read_note)
    if not headers and delim == "x12":
        res.warnings.append(
            "This is an X12 interchange but it contains no 837 claim (CLM) "
            "or 835 payment (CLP) segments — a 999 acknowledgment or "
            "270/276 inquiry can't be cleaned as claims.")
        res.out_name = _out_name(src_name)
        _write_output(res, headers, [])
        if not _stream_chunk:
            # This early return used to bypass the auto-file loop, so the
            # most common "we can't handle your file" outcome never
            # reached the backlog.
            _autofile_gaps(res)
        cb("Done", 1.0)
        return res
    if not headers:
        res.warnings.append("File appears to be empty — no header row found.")
        res.out_name = _out_name(src_name)
        _write_output(res, headers, [])
        cb("Done", 1.0)
        return res

    cb("Detecting NPI columns", 0.15)
    npi_idx, billing_idx = _detect_npi_columns(headers)
    res.npi_columns = [headers[i] for i in npi_idx]
    res.billing_column = headers[billing_idx] if billing_idx is not None else None
    # Ordering / referring provider NPI columns among the detected NPI columns.
    # CMS denies claims whose ordering/referring provider isn't a valid, active
    # (PECOS-enrolled) provider, so these get their own eligibility screen —
    # not just the billing NPI.
    order_refer_idx = [i for i in npi_idx
                       if any(h in _norm_key(headers[i])
                              for h in _ORDER_REFER_HINTS)]
    res.order_referring_columns = [headers[i] for i in order_refer_idx]
    # The "no NPI column" warning waits until after the value-based sniff
    # below — a PROV_ID column full of Luhn-valid NPIs is an NPI column.

    name_idx = _detect_one(headers, _NAME_HINTS)
    state_idx = _detect_one(headers, _STATE_HINTS)
    if state_idx is None:
        # Bare "St" headers exist in the wild but "st" as a SUBSTRING hint
        # matched ClaimStatus/PostDate-style columns — exact key only.
        for _i0, _h0 in enumerate(headers):
            if _norm_key(_h0) == "st":
                state_idx = _i0
                break
    ndc_idx = _detect_one(headers, _NDC_HINTS)
    drug_idx = _detect_one(headers, _DRUG_HINTS)
    # Column-role sets for the normalization pass (by header hint).
    npi_set = set(npi_idx)
    money_set = {i for i, h in enumerate(headers)
                 if any(x in _norm_key(h) for x in _MONEY_HINTS)
                 and not any(x in _norm_key(h) for x in _MONEY_EXCLUDE)}
    date_set = {i for i, h in enumerate(headers)
                if any(x in _norm_key(h) for x in _DATE_HINTS)}
    # A date column always wins over money: "DischargeDate" contains the
    # substring "charge", and without this a discharge-date column would be
    # routed to the money cleaner (never date-normalized) and false-flagged
    # as unparseable money.
    money_set -= date_set
    zip_set = {i for i, h in enumerate(headers)
               if any(x in _norm_key(h) for x in _ZIP_HINTS)}
    hcpcs_set = {i for i, h in enumerate(headers)
                 if any(x in _norm_key(h) for x in _HCPCS_HINTS)}
    sex_set = {i for i, h in enumerate(headers)
               if any(x in _norm_key(h) for x in _SEX_HINTS)}
    dx_set = {i for i, h in enumerate(headers)
              if any(x in _norm_key(h) for x in _DX_HINTS)}
    mod_set = {i for i, h in enumerate(headers)
               if any(x in _norm_key(h) for x in _MOD_HINTS)}
    phone_set = {i for i, h in enumerate(headers)
                 if any(x in _norm_key(h) for x in _PHONE_HINTS)}
    taxo_set = {i for i, h in enumerate(headers)
                if any(x in _norm_key(h) for x in _TAXO_HINTS)}
    ndc_norm_set = {i for i, h in enumerate(headers)
                    if any(x in _norm_key(h) for x in _NDC_HINTS)}
    state_set = {i for i in ([state_idx] if state_idx is not None else [])}
    rev_set = {i for i, h in enumerate(headers)
               if any(x in _norm_key(h) for x in _REV_HINTS)}
    # POS by exact "pos" key or the unambiguous long forms — bare substring
    # "pos" would false-match deposits/positions.
    pos_set = {i for i, h in enumerate(headers)
               if _norm_key(h) == "pos" or "placeofservice" in _norm_key(h)
               or "poscode" in _norm_key(h)}
    # DRG columns by exact/suffix match — bare substring "drg" would false-
    # match headers like "DrGroup", so the key must BE drg, END with drg
    # (MSDRG, FinalDRG), or carry the unambiguous long forms.
    drg_set = {i for i, h in enumerate(headers)
               if _norm_key(h) == "drg" or _norm_key(h).endswith("drg")
               or "msdrg" in _norm_key(h) or "drgcode" in _norm_key(h)}
    # Person provider-name columns for the re-case + credential parse. Any
    # role already claimed above wins — a header like "RenderingProviderNPI"
    # must stay with the NPI cleaner.
    pname_set = {i for i, h in enumerate(headers)
                 if any(x in _norm_key(h) for x in _PNAME_HINTS)
                 and not any(x in _norm_key(h) for x in _PNAME_EXCLUDE)}
    pname_set -= (npi_set | money_set | date_set | zip_set | phone_set
                  | taxo_set)
    # Date columns that must never be in the future (service / birth / paid).
    past_date_cols = {i for i, h in enumerate(headers)
                      if any(x in _norm_key(h)
                             for x in (_SERVICE_DATE_HINTS + _DOB_HINTS))}
    service_date_cols = {i for i, h in enumerate(headers)
                         if any(x in _norm_key(h) for x in _SERVICE_DATE_HINTS)}
    dob_i = _detect_one(headers, _DOB_HINTS)
    admit_i = _detect_one(headers, _ADMIT_HINTS)
    disch_i = _detect_one(headers, _DISCH_HINTS)
    payer_i = _detect_one(headers, _PAYER_HINTS)
    carc_i = _detect_one(headers, _CARC_HINTS)
    tob_i = _detect_one(headers, _TOB_HINTS)
    dstat_i = _detect_one(headers, _DISCH_STATUS_HINTS)
    # "DischargeStatus" contains the substring "charge" — without this it
    # would be routed to the money cleaner ("01" → "1.00") and every status
    # false-flagged. Same class of bug as DischargeDate vs the date role.
    if dstat_i is not None:
        money_set.discard(dstat_i)
    atype_i = _detect_one(headers, _ADMIT_TYPE_HINTS)
    asrc_i = _detect_one(headers, ("admissionsource", "admitsource",
                                   "pointoforigin", "sourceofadmission",
                                   "admsrc", "priorityvisittype"))
    # Service thru / to date — pairs with dos_i (from date) for the
    # from>thru ordering screen. Exact/suffix hints only so it never
    # steals the from-date column.
    todate_i = _detect_one(headers, ("servicetodate", "thrudate",
                                     "todate", "servicethrudate",
                                     "dateofservicethru", "dosthru",
                                     "statementthrudate"))
    submit_i = _detect_one(headers, _SUBMIT_HINTS)
    member_i = _detect_one(headers, _MEMBER_HINTS)
    cond_i = _detect_one(headers, _COND_HINTS)
    occ_i = _detect_one(headers, _OCC_HINTS)
    valc_i = _detect_one(headers, _VALUE_HINTS)
    try:
        from . import refdata as _rd
    except Exception:  # noqa: BLE001 — refdata missing → those checks off
        _rd = None
    # Reference-data packs (refdata_packs.py): the FULL public code sets,
    # loaded once per run when the user has pulled them. None → the
    # corresponding pack-gated checks simply stay off; nothing about the
    # zero-setup path changes.
    _icd_pack = _hcpcs_pack = _leie_pack = _taxo_pack = None
    try:
        from . import refdata_packs as _packs
        _icd_pack = _packs.icd10_codes()
        _hcpcs_pack = _packs.hcpcs_codes()
        _leie_pack = _packs.leie_npis()
        _taxo_pack = _packs.taxonomy_codes()
    except Exception:  # noqa: BLE001 — packs never block cleaning
        pass
    from datetime import datetime as _dt, timezone as _tz
    _today = _dt.now(_tz.utc).date()
    # Profile-tunable thresholds (see profiles.py) with safe defaults.
    _prof = profile or {}
    _thr = _prof.get("thresholds") or {}
    _stale_years = int(_thr.get("stale_years", 10) or 10)
    _timely_days = int(_thr.get("timely_filing_days", 365) or 365)
    _iqr_mult = float(_thr.get("outlier_iqr_mult", 3.0) or 3.0)
    _dup_window = int(_thr.get("dup_window_days", 3) or 3)
    _future_grace = int(_thr.get("future_grace_days", 0) or 0)
    _high_units = int(_thr.get("high_units_threshold", 0) or 0)
    _readmit_window = int(_thr.get("readmit_window_days", 30) or 30)
    # Future-date grace: a service date this many days past today is still OK
    # (near-future scheduled visits); beyond it flags. 0 = flag past today.
    from datetime import timedelta as _timedelta
    _future_cut = _today + _timedelta(days=_future_grace)
    # Staleness horizon for service dates — a DOS this old in a working
    # claims extract is almost always a key-entry or century error.
    _stale_cut = f"{_today.year - _stale_years:04d}-01-01"
    # ZIP↔state consistency: pair same-entity state/zip columns and flag rows
    # whose ZIP3 prefix resolves to a different state than the state cell.
    zs_pairs = _zip_state_pairs(headers)
    zs_map = _zip3_state_map() if zs_pairs else {}
    # PHI columns for opt-in de-identification (patient direct identifiers).
    phi_cols: Dict[int, str] = {}
    if deid:
        for i, h in enumerate(headers):
            k = _phi_kind(_norm_key(h))
            if k:
                phi_cols[i] = k
        res.deid_applied = True
        res.deid_columns = [headers[i] for i in sorted(phi_cols)]
    _deid_salt = uuid.uuid4().hex
    # Specific columns for cross-field sanity flags.
    allowed_i = _detect_one(headers, ("allowedamt", "allowed"))
    billed_i = _detect_one(headers, ("billedamt", "billed", "chargeamt",
                                     "charge", "submittedamt"))
    paid_i = _detect_one(headers, ("paidamt", "paymentamt", "planpaid", "paid"))
    units_i = _detect_one(headers, ("units", "unit", "quantity", "qty", "srvccnt"))
    # Composite-key columns for suspected duplicate-claim detection.
    patient_i = _detect_one(headers, ("patientid", "memberid", "subscriberid",
                                      "patientaccount", "patientacct",
                                      "patient", "memberidnumber", "beneid"))
    dos_i = _detect_one(headers, ("dateofservice", "servicedate", "svcdate",
                                  "dos", "fromdate", "servicefromdate"))
    hcpcs_i = _detect_one(headers, _HCPCS_HINTS)

    # User column-mapping overrides (canonical role → header) win over
    # auto-detection. Only roles the stdlib path acts on are honored here; the
    # rest flow to the v49 engine via vendor_adapter.
    if overrides:
        hidx = {h: i for i, h in enumerate(headers)}

        def _ov(role):
            col = overrides.get(role)
            return hidx.get(col) if col else None

        _b = _ov("billing_npi")
        if _b is not None:
            billing_idx = _b
            if _b not in npi_idx:
                npi_idx = sorted(set(npi_idx) | {_b})
                res.npi_columns = [headers[i] for i in npi_idx]
            res.billing_column = headers[_b]
        for role, setter in (("billing_name", "name"), ("state", "state"),
                             ("ndc", "ndc"), ("drug_name", "drug")):
            _i = _ov(role)
            if _i is not None:
                if setter == "name":
                    name_idx = _i
                elif setter == "state":
                    state_idx = _i
                elif setter == "ndc":
                    ndc_idx = _i
                elif setter == "drug":
                    drug_idx = _i

    claim_i = _detect_one(headers, ("claimid", "claimnumber", "claimno",
                                    "patientcontrolnumber", "icn", "dcn"))
    # Value-based NPI sniff over columns NO role claimed: a PROV_ID /
    # ProviderNumber column full of Luhn-valid 10-digit values is an NPI
    # column even though its header never says so. Adopted for stats and
    # screens only (repair-free — cells are never rewritten), and it never
    # displaces an explicit billing header.
    try:
        _sniff_claimed: set = set(npi_idx)
        for _rs0 in (money_set, date_set, zip_set, hcpcs_set, sex_set,
                     dx_set, mod_set, phone_set, taxo_set, ndc_norm_set,
                     state_set, rev_set, pos_set, drg_set, pname_set):
            _sniff_claimed |= _rs0
        for _ix0 in (name_idx, state_idx, ndc_idx, drug_idx, dob_i, admit_i,
                     disch_i, payer_i, carc_i, tob_i, dstat_i, atype_i,
                     asrc_i, todate_i, submit_i, member_i, cond_i, occ_i,
                     valc_i, allowed_i, billed_i, paid_i, units_i,
                     patient_i, dos_i, hcpcs_i, claim_i):
            if _ix0 is not None:
                _sniff_claimed.add(_ix0)
        _sniffed_npi = _sniff_npi_columns(headers, rows, _sniff_claimed)
    except Exception:  # noqa: BLE001 — the sniff must never block a run
        _sniffed_npi = []
    if _sniffed_npi:
        npi_idx = sorted(set(npi_idx) | set(_sniffed_npi))
        res.npi_columns = [headers[i] for i in npi_idx]
        res.structure["npi_sniffed_columns"] = \
            [headers[i] for i in _sniffed_npi]
        if billing_idx is None:
            # No header-detected NPI column at all → the first sniffed
            # column is the de-facto billing column (screened, not edited).
            billing_idx = _sniffed_npi[0]
            res.billing_column = headers[billing_idx]
        order_refer_idx = [i for i in npi_idx
                           if any(h in _norm_key(headers[i])
                                  for h in _ORDER_REFER_HINTS)]
        res.order_referring_columns = [headers[i] for i in order_refer_idx]
        res.warnings.append(
            f"{len(_sniffed_npi)} column(s) hold NPI-shaped values under "
            "non-NPI headers ("
            + ", ".join(headers[i] for i in _sniffed_npi[:5])
            + ") — screened as NPI columns (values are never rewritten). "
            "Rename the headers to include 'NPI' to make this explicit.")
    if not npi_idx:
        res.warnings.append(
            "No NPI column detected (looked for headers containing 'NPI'). "
            "Rows were still trimmed and de-duplicated.")

    ncols = len(headers)
    for i in npi_idx:
        res.column_stats[headers[i]] = {
            "valid": 0, "blank": 0, "malformed": 0, "checksum": 0, "cells": 0}

    # Structural hygiene before the row loop: a header that appears twice
    # makes every downstream mapping ambiguous, so it's surfaced up front.
    _hdr_counts: Dict[str, int] = {}
    for h in headers:
        _hdr_counts[h] = _hdr_counts.get(h, 0) + 1
    _dup_headers = sorted(h for h, n in _hdr_counts.items() if n > 1 and h)
    if _dup_headers:
        res.structure["duplicate_headers"] = _dup_headers

    # Accumulators for the post-loop analyses.
    _col_filled = [0] * ncols                     # empty-column detection
    _col_distinct = [set() for _ in range(ncols)]  # dictionary (capped 1000)
    _payer_raw: Dict[str, int] = {}               # payer variant clustering
    _carc_counts: Dict[str, int] = {}             # top denial reasons
    _carc_variants: Dict[str, set] = {}           # bare code → raw spellings
    _code_charges: Dict[str, List[tuple]] = {}   # per-HCPCS (amt, row) pairs
    _taxo_counts: Dict[str, int] = {}             # taxonomy → specialty mix
    _svc_seen: Dict[tuple, List[tuple]] = {}      # dup-service window scan
    _payer_q: Dict[str, Dict] = {}                # per-payer quality split
    _claim_agg: Dict[str, List[float]] = {}       # claim rollup [lines, charge]
    _claim_trunc = False
    _CLAIM_CAP = 50_000                           # distinct claims tracked
    # (claim_i was detected above, before the NPI value sniff.)
    _charge_i = billed_i if billed_i is not None else allowed_i
    # Change-log spill: entries past the in-memory preview stream to a
    # temp CSV so the audit-trail FILE is complete at any scale. The
    # guard owns the temp file and self-cleans if the run dies early
    # (cancel / exception) — see _ChangelogSpill.
    _spill = _ChangelogSpill()
    _phi_log_skipped = 0

    cb("Cleaning rows", 0.30)
    cleaned: List[List[str]] = []
    seen = set()
    _seen_fold = set()   # case/whitespace-folded keys → near-dup signal
    # Ragged-row accounting: pads/trims used to happen with ZERO signal, so
    # an unquoted comma in an address shifted every later cell AND silently
    # destroyed the overflow. xlsx / X12 rows are structurally variable-
    # width by construction, so only delimited text counts as ragged.
    _ragged_track = delim not in ("xlsx", "x12", "x835")
    _ragged_padded = _ragged_truncated = 0
    total = max(len(rows), 1)
    for ri, row in enumerate(rows):
        # Pad / trim ragged rows to the header width (counted below).
        _ragged_row = False
        if len(row) < ncols:
            row = row + [""] * (ncols - len(row))
            if _ragged_track:
                _ragged_padded += 1
                _ragged_row = True
        elif len(row) > ncols:
            row = row[:ncols]
            if _ragged_track:
                _ragged_truncated += 1
                _ragged_row = True
        # Deterministic normalization pass on every cell: generic cleanups
        # (whitespace, mojibake, null-tokens, Excel apostrophe) plus per-role
        # fixes (NPI, money, date, state, zip, HCPCS). Each fix is tallied.
        new_row = []
        _cell_flags: set = set()   # row-level flags raised by cell cleaners
        for ci, cell in enumerate(row):
            stripped = cell.strip()
            if stripped != cell:
                res.n_cells_trimmed += 1
            val, hits = _clean_generic(stripped)
            if ci in npi_set:
                _pre_npi = val
                val, r = _clean_npi_cell(val); hits += r
                if ("npi-scientific-notation" not in r and _pre_npi
                        and _NPI_SCI_RE.match(_pre_npi.strip())):
                    # Scientific notation whose mantissa lost digits — the
                    # identifier is unrecoverable; flagged, never guessed.
                    _cell_flags.add("npi-scientific-lossy")
            elif ci in money_set:
                val, r = _clean_money_cell(val); hits += r
            elif ci in date_set:
                val, r = _clean_date_cell(val); hits += r
            elif ci in state_set:
                val, r = _clean_state_cell(val); hits += r
            elif ci in zip_set:
                val, r = _clean_zip_cell(val); hits += r
            elif ci in hcpcs_set:
                val, r = _clean_hcpcs_cell(val); hits += r
            elif ci in sex_set:
                val, r = _clean_sex_cell(val); hits += r
            elif ci in dx_set:
                val, r = _clean_dx_cell(val); hits += r
            elif ci in mod_set:
                val, r = _clean_modifier_cell(val); hits += r
            elif ci in phone_set:
                val, r = _clean_phone_cell(val); hits += r
            elif ci in taxo_set:
                val, r = _clean_taxonomy_cell(val); hits += r
            elif ci in ndc_norm_set:
                val, r = _clean_ndc_cell(val); hits += r
            elif ci in rev_set:
                val, r = _clean_revcode_cell(val); hits += r
            elif ci in pos_set:
                val, r = _clean_pos_cell(val); hits += r
            elif ci in drg_set:
                val, r = _clean_drg_cell(val); hits += r
            elif ci in pname_set:
                val, r, _crd = _clean_provider_name_cell(val); hits += r
                for _c in _crd:
                    res.credentials[_c] = res.credentials.get(_c, 0) + 1
            for rule in hits:
                res.repairs[rule] = res.repairs.get(rule, 0) + 1
            # Cell-level audit trail. On a de-identified run, changes in
            # patient-identifier columns are COUNTED but never logged —
            # the change log is a downloadable artifact, and its
            # before-values would hand back exactly the PHI the user
            # asked to remove (the old 20k cap bounded that exposure by
            # accident; uncapped it would be the whole file). Everything
            # else: first _CHANGELOG_PREVIEW entries stay in memory for
            # the UI; the rest go to the caller's sink (streaming chunks)
            # or spill to disk, so the FILE is complete at any scale.
            if val != cell:
                res.n_changes += 1
                if phi_cols and ci in phi_cols:
                    _phi_log_skipped += 1
                else:
                    _entry = (ri + 1,
                              headers[ci] if ci < ncols else str(ci),
                              cell, val, ";".join(hits) or "trim")
                    if _changelog_sink is not None:
                        _changelog_sink(_entry)
                        if len(res.changelog) < _CHANGELOG_PREVIEW:
                            res.changelog.append(_entry)
                    elif len(res.changelog) < _CHANGELOG_PREVIEW:
                        res.changelog.append(_entry)
                    elif _stream_chunk:
                        # Chunk without a sink (direct callers): no file
                        # will be written; spilling would leak temp files.
                        res.changelog_truncated = True
                    else:
                        _spill.write(
                            [_entry[0], _defang_cell(_entry[1]),
                             _defang_cell(_entry[2]),
                             _defang_cell(_entry[3]), _entry[4]])
                        if _spill.failed:
                            res.changelog_truncated = True
            # Fill-rate counters for the data-quality report card.
            res.n_cells_total += 1
            if val != "":
                res.n_cells_filled += 1
                _col_filled[ci] += 1
            # PHI de-identification (opt-in) — applied last so the value that
            # lands in the output (and the pivot) is masked. Stable id hashing
            # keeps within-file referential integrity for counts/joins.
            if phi_cols and ci in phi_cols and val != "":
                masked = _deid_value(val, phi_cols[ci], _deid_salt)
                if masked != val:
                    res.deid_cells += 1
                val = masked
            new_row.append(val)
        # Snapshot for worklist row-capture AND per-payer quality: any rule
        # whose count rises during this row's checks fired on this row.
        _sanity_pre = dict(res.sanity)
        # Row-shape and cell-cleaner flags (counted after the snapshot so
        # they land in the worklists like every other row-level flag).
        if _ragged_row:
            res.sanity["ragged-row"] = res.sanity.get("ragged-row", 0) + 1
        for _cf in _cell_flags:
            res.sanity[_cf] = res.sanity.get(_cf, 0) + 1
        # Cross-field sanity flags (report-only) on the cleaned row.
        for f in _row_sanity_flags(new_row, allowed_i, billed_i, paid_i,
                                   units_i, high_units=_high_units):
            res.sanity[f] = res.sanity.get(f, 0) + 1
        # Impossible future date on a service/birth/paid column (counted once
        # per row, like the other row-level sanity flags). A profile can allow
        # a grace window for near-future scheduled dates.
        if any(ci < len(new_row) and _date_after(new_row[ci], _future_cut)
               for ci in past_date_cols):
            res.sanity["date-in-future"] = res.sanity.get("date-in-future", 0) + 1
        # ZIP prefix disagrees with the same-entity state cell (counted
        # once) — and a BLANK state next to a resolvable ZIP is FILLED:
        # ZIP3→state is deterministic (military/territory prefixes
        # excluded), the same truth the mismatch flag already trusts, so
        # filling is safe by construction and fully audited.
        for si, zi in zs_pairs:
            if si >= len(new_row) or zi >= len(new_row):
                continue
            sv = new_row[si].strip().upper()
            zdig = "".join(c for c in new_row[zi] if c.isdigit())
            if not sv and len(zdig) >= 3:
                exp = zs_map.get(zdig[:3])
                if exp and exp not in _ZIP_STATE_SKIP:
                    res.repairs["state-from-zip"] = \
                        res.repairs.get("state-from-zip", 0) + 1
                    res.n_changes += 1
                    _fill_entry = (ri + 1,
                                   headers[si] if si < ncols else str(si),
                                   "", exp, "state-from-zip")
                    if _changelog_sink is not None:
                        _changelog_sink(_fill_entry)
                        if len(res.changelog) < _CHANGELOG_PREVIEW:
                            res.changelog.append(_fill_entry)
                    elif len(res.changelog) < _CHANGELOG_PREVIEW:
                        res.changelog.append(_fill_entry)
                    elif not _stream_chunk:
                        _spill.write([_fill_entry[0],
                                      _defang_cell(_fill_entry[1]), "",
                                      exp, "state-from-zip"])
                    new_row[si] = exp
                    res.n_cells_filled += 1
                    _col_filled[si] += 1
                continue
            if len(sv) == 2 and sv in _STATE_CODES and len(zdig) >= 3:
                exp = zs_map.get(zdig[:3])
                if (exp and exp != sv
                        and exp not in _ZIP_STATE_SKIP
                        and sv not in _ZIP_STATE_SKIP):
                    res.sanity["zip-state-mismatch"] = \
                        res.sanity.get("zip-state-mismatch", 0) + 1
                    break
        # NDC that couldn't be safely normalized (unhyphenated 10-digit).
        for ci in ndc_norm_set:
            if ci < len(new_row):
                nd = new_row[ci]
                if nd and nd.isdigit() and len(nd) == 10:
                    res.sanity["ndc-ambiguous-10digit"] = \
                        res.sanity.get("ndc-ambiguous-10digit", 0) + 1
        # Code-shape validity (report-only, once per row) — malformed HCPCS/CPT
        # and ICD-10 codes are a top denial driver.
        if hcpcs_set and any(ci < len(new_row) and _hcpcs_malformed(new_row[ci])
                             for ci in hcpcs_set):
            res.sanity["hcpcs-malformed"] = res.sanity.get("hcpcs-malformed", 0) + 1
        if dx_set and any(ci < len(new_row) and _icd10_malformed(new_row[ci])
                          for ci in dx_set):
            res.sanity["icd10-malformed"] = res.sanity.get("icd10-malformed", 0) + 1
        # Amount cell that survived cleaning but still isn't a number.
        if money_set and any(ci < len(new_row) and _money_unparseable(new_row[ci])
                             for ci in money_set):
            res.sanity["money-unparseable"] = \
                res.sanity.get("money-unparseable", 0) + 1
        # Date cell that survived cleaning but still isn't ISO-shaped —
        # every chronology check on this row silently skipped it, so the
        # user must be told (the date analogue of money-unparseable).
        if date_set and any(ci < len(new_row)
                            and _date_unparseable(new_row[ci])
                            for ci in date_set):
            res.sanity["date-unparseable"] = \
                res.sanity.get("date-unparseable", 0) + 1
        # Modifier cell still carrying a token that isn't a 2-char
        # alphanumeric modifier — kept in the output (never discarded),
        # flagged here.
        if mod_set and any(ci < len(new_row) and new_row[ci]
                           and _modifier_malformed(new_row[ci])
                           for ci in mod_set):
            res.sanity["modifier-malformed"] = \
                res.sanity.get("modifier-malformed", 0) + 1
        # A DRG on a professional-shaped row (POS populated, no Type of
        # Bill) is structurally wrong — DRGs price institutional stays.
        if drg_set and pos_set:
            _drg_val = any(ci < len(new_row) and new_row[ci].strip()
                           for ci in drg_set)
            _pos_val = any(ci < len(new_row) and new_row[ci].strip()
                           for ci in pos_set)
            _tob_val = (tob_i is not None and tob_i < len(new_row)
                        and new_row[tob_i].strip())
            if _drg_val and _pos_val and not _tob_val:
                res.sanity["drg-on-professional"] = \
                    res.sanity.get("drg-on-professional", 0) + 1
        # Value-domain validity: sex not M/F/U, taxonomy not 10-char alnum.
        if sex_set and any(ci < len(new_row) and _sex_invalid(new_row[ci])
                           for ci in sex_set):
            res.sanity["sex-invalid"] = res.sanity.get("sex-invalid", 0) + 1
        if taxo_set and any(ci < len(new_row) and _taxonomy_malformed(new_row[ci])
                            for ci in taxo_set):
            res.sanity["taxonomy-malformed"] = \
                res.sanity.get("taxonomy-malformed", 0) + 1
        # Revenue-code / Place-of-Service domain validity (report-only).
        if rev_set and any(ci < len(new_row) and _revcode_malformed(new_row[ci])
                           for ci in rev_set):
            res.sanity["revenue-code-malformed"] = \
                res.sanity.get("revenue-code-malformed", 0) + 1
        if pos_set and any(ci < len(new_row) and _pos_invalid(new_row[ci])
                           for ci in pos_set):
            res.sanity["pos-invalid"] = res.sanity.get("pos-invalid", 0) + 1
        if drg_set and any(ci < len(new_row) and _drg_malformed(new_row[ci])
                           for ci in drg_set):
            res.sanity["drg-malformed"] = \
                res.sanity.get("drg-malformed", 0) + 1
        # Modifier↔units coding edits: a JW line (discarded drug) bills the
        # wasted units so units must be positive, and a bilateral (50) line
        # bills 1 unit per CMS MUE guidance. Modifiers were normalized to a
        # comma-joined upper-case list by _clean_modifier_cell, so the split
        # is reliable.
        if mod_set and units_i is not None and units_i < len(new_row):
            _mods: set = set()
            for ci in mod_set:
                if ci < len(new_row) and new_row[ci]:
                    _mods.update(new_row[ci].split(","))
            if _mods:
                _u_val = _to_number(new_row[units_i])
                if "JW" in _mods and (_u_val is None or _u_val <= 0):
                    res.sanity["jw-zero-units"] = \
                        res.sanity.get("jw-zero-units", 0) + 1
                if "50" in _mods and _u_val is not None and _u_val > 1:
                    res.sanity["bilateral-units"] = \
                        res.sanity.get("bilateral-units", 0) + 1
        # Anesthesia lines (CPT 00100-01999) bill time units — more than 24
        # hours' worth of minutes on one line is a keying error (extra digit,
        # column shift), not a marathon case.
        if (hcpcs_i is not None and units_i is not None
                and hcpcs_i < len(new_row) and units_i < len(new_row)):
            _ac = new_row[hcpcs_i]
            if (len(_ac) == 5 and _ac.isdigit()
                    and "00100" <= _ac <= "01999"):
                _au = _to_number(new_row[units_i])
                if _au is not None and _au > 24 * 60:
                    res.sanity["anesthesia-units-implausible"] = \
                        res.sanity.get("anesthesia-units-implausible", 0) + 1
        # Denial / adjustment reason (CARC) domain validity. Cells may carry
        # several codes ("16, 97") — every part must be a valid CARC shape.
        if carc_i is not None and carc_i < len(new_row) and new_row[carc_i]:
            _cparts = [p for p in re.split(r"[,;|\s]+",
                                           new_row[carc_i].strip().upper()) if p]
            if _cparts and any(
                    _CARC_VALID_RE.match(_CARC_GROUP_RE.sub("", p)) is None
                    for p in _cparts):
                res.sanity["carc-invalid"] = \
                    res.sanity.get("carc-invalid", 0) + 1
        # Institutional-claim domains (UB-04): Type of Bill, discharge
        # status, admission type — clearinghouse front-door edits.
        if _rd is not None:
            if (tob_i is not None and tob_i < len(new_row)
                    and _rd.tob_invalid(new_row[tob_i])):
                res.sanity["tob-malformed"] = \
                    res.sanity.get("tob-malformed", 0) + 1
            if (dstat_i is not None and dstat_i < len(new_row)
                    and _rd.discharge_status_invalid(new_row[dstat_i])):
                res.sanity["discharge-status-invalid"] = \
                    res.sanity.get("discharge-status-invalid", 0) + 1
            if (atype_i is not None and atype_i < len(new_row)
                    and _rd.admission_type_invalid(new_row[atype_i])):
                res.sanity["admission-type-invalid"] = \
                    res.sanity.get("admission-type-invalid", 0) + 1
            if (asrc_i is not None and asrc_i < len(new_row)
                    and _rd.admission_source_invalid(new_row[asrc_i])):
                res.sanity["admission-source-invalid"] = \
                    res.sanity.get("admission-source-invalid", 0) + 1
            # Discharge status 30 ("still a patient") on a FINAL bill — the
            # TOB frequency digit 1 (admit-through-discharge) or 4 (last
            # interim) says the stay is closed, so the patient can't still
            # be admitted. Classic NUBC contradiction.
            if (tob_i is not None and dstat_i is not None
                    and tob_i < len(new_row) and dstat_i < len(new_row)):
                _fc2 = _rd.tob_facility_class(new_row[tob_i])
                _ds = new_row[dstat_i].strip().lstrip("0") or "0"
                _tob_raw = new_row[tob_i].strip()
                _freq = _tob_raw[-1] if len(_tob_raw) >= 3 else ""
                if _fc2 is not None and _ds == "30" and _freq in ("1", "4"):
                    res.sanity["discharge-status-final-bill"] = \
                        res.sanity.get("discharge-status-final-bill", 0) + 1
            # Accommodation (room & board / ICU, revenue 0100-0219) revenue
            # on an OUTPATIENT bill type — hospital outpatient 013x/014x,
            # clinic 07xx, ASC 083x. Inpatient room charges can't ride an
            # outpatient claim; this is a front-door clearinghouse edit.
            if (tob_i is not None and tob_i < len(new_row)
                    and new_row[tob_i] and rev_set):
                _fc = _rd.tob_facility_class(new_row[tob_i])
                if _fc is not None and (
                        _fc[0] == "7"
                        or (_fc[0] == "1" and _fc[1] in ("3", "4"))
                        or (_fc[0] == "8" and _fc[1] == "3")):
                    for _rci in rev_set:
                        _rv2 = (new_row[_rci]
                                if _rci < len(new_row) else "")
                        if (_rv2 and _rv2.isdigit() and len(_rv2) == 4
                                and "0100" <= _rv2 <= "0219"):
                            res.sanity["revenue-tob-mismatch"] = \
                                res.sanity.get("revenue-tob-mismatch", 0) + 1
                            break
            # Unknown (but well-formed) modifiers — typo signal.
            if mod_set:
                for ci in mod_set:
                    if ci < len(new_row) and new_row[ci] and any(
                            _rd.modifier_unknown(m)
                            for m in new_row[ci].split(",")):
                        res.sanity["modifier-unknown"] = \
                            res.sanity.get("modifier-unknown", 0) + 1
                        break
        # Malformed NDC — a drug code that can't be 10/11 digits (keying
        # damage the normalizer couldn't fix). Uses the engine's own helper,
        # so it runs independent of the refdata (_rd) guard.
        if ndc_norm_set:
            for _ni in ndc_norm_set:
                if _ni < len(new_row) and _ndc_malformed(new_row[_ni]):
                    res.sanity["ndc-malformed"] = \
                        res.sanity.get("ndc-malformed", 0) + 1
                    break
        # Service from-date after thru-date — an impossible span.
        if (dos_i is not None and todate_i is not None
                and dos_i < len(new_row) and todate_i < len(new_row)):
            _fd, _td = new_row[dos_i], new_row[todate_i]
            if (_DATE_ISO_RE.match(_fd) and _DATE_ISO_RE.match(_td)
                    and _fd[:10] > _td[:10]):
                res.sanity["service-date-order"] = \
                    res.sanity.get("service-date-order", 0) + 1
        # Taxonomy code not in the full NUCC set — pack-gated, same as
        # icd10/hcpcs unknown-code (off until the taxonomy pack is pulled).
        if _taxo_pack is not None and taxo_set:
            for _txi in taxo_set:
                _txu = (new_row[_txi] if _txi < len(new_row)
                        else "").strip().upper()
                if (len(_txu) == 10 and _txu.isalnum()
                        and _txu not in _taxo_pack):
                    res.sanity["taxonomy-unknown-code"] = \
                        res.sanity.get("taxonomy-unknown-code", 0) + 1
                    break
        # Timely-filing risk: days between service and received dates over
        # the limit. When the row names a payer with a published limit
        # (Medicare 365, most commercial 90-180 — see refdata), that limit
        # applies; the profile threshold covers unknown payers.
        if (submit_i is not None and dos_i is not None
                and submit_i < len(new_row) and dos_i < len(new_row)):
            _sv, _rv = new_row[dos_i], new_row[submit_i]
            if _DATE_ISO_RE.match(_sv) and _DATE_ISO_RE.match(_rv):
                _tf_limit = _timely_days
                if (_rd is not None and payer_i is not None
                        and payer_i < len(new_row) and new_row[payer_i]):
                    _pd = _rd.timely_filing_days(_payer_key(new_row[payer_i]))
                    if _pd is not None:
                        _tf_limit = _pd
                try:
                    from datetime import date as _d2
                    _delta = (_d2.fromisoformat(_rv[:10])
                              - _d2.fromisoformat(_sv[:10])).days
                    if _delta > _tf_limit:
                        res.sanity["timely-filing-risk"] = \
                            res.sanity.get("timely-filing-risk", 0) + 1
                except ValueError:
                    pass
        # Medicare MBI shape — only when the payer on THIS row is Medicare
        # (family key), and never under de-identification (member IDs are
        # hashed to PT-… tokens which would all false-flag).
        if (_rd is not None and not deid and member_i is not None
                and payer_i is not None
                and member_i < len(new_row) and payer_i < len(new_row)
                and new_row[member_i] and new_row[payer_i]
                and _payer_key(new_row[payer_i]) == "MEDICARE"
                and _rd.mbi_malformed(new_row[member_i])):
            res.sanity["mbi-malformed"] = \
                res.sanity.get("mbi-malformed", 0) + 1
        # UB-04 condition / occurrence / value code shape (multi-code cells
        # split like modifiers; catalog membership NOT required — payers
        # define proprietary codes, so only keying damage flags).
        if _rd is not None:
            for _ci2, _rule2 in ((cond_i, "condition-code-malformed"),
                                 (occ_i, "occurrence-code-malformed"),
                                 (valc_i, "value-code-malformed")):
                if _ci2 is None or _ci2 >= len(new_row) or not new_row[_ci2]:
                    continue
                _parts2 = [p for p in re.split(r"[,;|\s]+",
                                               new_row[_ci2].strip()) if p]
                if _parts2 and any(_rd.ub_code_malformed(p)
                                   for p in _parts2):
                    res.sanity[_rule2] = res.sanity.get(_rule2, 0) + 1
        # Pack-gated validity: codes that pass every SHAPE check but do
        # not exist in the authoritative set (counted once per row, like
        # the other row-level flags). Off until the pack is pulled.
        if _icd_pack is not None and dx_set:
            for _dxi in dx_set:
                _dxv = new_row[_dxi] if _dxi < len(new_row) else ""
                _dxn = _dxv.strip().upper().replace(".", "")
                if (_dxn and not _icd10_malformed(_dxv)
                        and _dxn not in _icd_pack):
                    res.sanity["icd10-unknown-code"] = \
                        res.sanity.get("icd10-unknown-code", 0) + 1
                    break
        if _hcpcs_pack is not None and hcpcs_set:
            for _hxi in hcpcs_set:
                _hxv = (new_row[_hxi] if _hxi < len(new_row) else "").strip()
                _hxv = _hxv.upper()
                # Level II only: letter + 4 digits. Numeric CPT-4 codes
                # are AMA-licensed — shape-checked elsewhere, never
                # membership-checked here.
                if (len(_hxv) == 5 and _hxv[0].isalpha()
                        and _hxv[1:].isdigit() and _hxv not in _hcpcs_pack):
                    res.sanity["hcpcs-unknown-code"] = \
                        res.sanity.get("hcpcs-unknown-code", 0) + 1
                    break
        if (_leie_pack is not None and billing_idx is not None
                and billing_idx < len(new_row)):
            _bnd = "".join(c for c in new_row[billing_idx] if c.isdigit())
            if len(_bnd) == 10 and _bnd in _leie_pack:
                res.sanity["leie-excluded-npi"] = \
                    res.sanity.get("leie-excluded-npi", 0) + 1
        # Chronology impossibilities. Normalized ISO dates compare correctly
        # as strings, so no re-parsing per row.
        if (dob_i is not None and dos_i is not None
                and dob_i < len(new_row) and dos_i < len(new_row)):
            _dob, _dos = new_row[dob_i], new_row[dos_i]
            if (_DATE_ISO_RE.match(_dob) and _DATE_ISO_RE.match(_dos)
                    and _dos[:10] < _dob[:10]):
                res.sanity["service-before-birth"] = \
                    res.sanity.get("service-before-birth", 0) + 1
        if (admit_i is not None and disch_i is not None
                and admit_i < len(new_row) and disch_i < len(new_row)):
            _adm, _dis = new_row[admit_i], new_row[disch_i]
            if (_DATE_ISO_RE.match(_adm) and _DATE_ISO_RE.match(_dis)
                    and _dis[:10] < _adm[:10]):
                res.sanity["discharge-before-admit"] = \
                    res.sanity.get("discharge-before-admit", 0) + 1
        # Service date older than the 10-year horizon (likely century error).
        if any(ci < len(new_row) and new_row[ci]
               and _DATE_ISO_RE.match(new_row[ci])
               and new_row[ci][:10] < _stale_cut
               for ci in service_date_cols):
            res.sanity["date-stale"] = res.sanity.get("date-stale", 0) + 1
        # Tally NPI health per detected column.
        for i in npi_idx:
            cstat = res.column_stats[headers[i]]
            cstat["cells"] += 1
            cstat[classify_npi(new_row[i])] += 1
        if drop_duplicates:
            if _stream_seen is not None:
                # Streaming mode: membership lives in the caller's shared
                # set as a 96-bit digest — cross-chunk dedupe in bounded
                # memory (collision odds are negligible at claims scale).
                _dk = int.from_bytes(
                    hashlib.blake2b("\x1f".join(new_row).encode("utf-8"),
                                    digest_size=12).digest(), "big")
                if _dk in _stream_seen:
                    res.n_dupes_removed += 1
                    continue
                if len(_stream_seen) < _STREAM_SEEN_CAP:
                    _stream_seen.add(_dk)
            else:
                key = tuple(new_row)
                if key in seen:
                    res.n_dupes_removed += 1
                    continue
                seen.add(key)
        # Near-duplicate (report-only, row kept): identical after case
        # folding + whitespace collapse — "SMITH" vs "Smith" rows a human
        # would merge. Runs regardless of dedupe mode: with dedupe off,
        # exact repeats ARE kept near-duplicates worth surfacing. Stored
        # as a hash — it's only a membership probe for a counter, and
        # keeping folded copies of every row doubled peak memory at scale.
        # Streaming runs inject a SHARED digest set (mirroring
        # _stream_seen) so folded twins die across chunk boundaries too.
        if _stream_fold_seen is not None:
            _fk = int.from_bytes(
                hashlib.blake2b(
                    "\x1f".join(" ".join(c.split()).casefold()
                                for c in new_row).encode("utf-8"),
                    digest_size=12).digest(), "big")
            if _fk in _stream_fold_seen:
                res.sanity["near-duplicate-row"] = \
                    res.sanity.get("near-duplicate-row", 0) + 1
            elif len(_stream_fold_seen) < _STREAM_SEEN_CAP:
                _stream_fold_seen.add(_fk)
        else:
            _fold_h = hash(tuple(" ".join(c.split()).casefold()
                                 for c in new_row))
            if _fold_h in _seen_fold:
                res.sanity["near-duplicate-row"] = \
                    res.sanity.get("near-duplicate-row", 0) + 1
            else:
                _seen_fold.add(_fold_h)
        # Kept-row accumulators for payer clustering and charge outliers.
        # Taxonomy → specialty mix: count well-formed codes on kept rows
        # (capped distinct codes so a garbage column can't grow the dict).
        for _ti in taxo_set:
            _tv = new_row[_ti] if _ti < len(new_row) else ""
            if _tv and not _taxonomy_malformed(_tv) and (
                    _tv in _taxo_counts or len(_taxo_counts) < 300):
                _taxo_counts[_tv] = _taxo_counts.get(_tv, 0) + 1
        # Duplicate-service window scan: same patient + provider + code on
        # kept rows, dates compared post-loop against the profile window.
        if (patient_i is not None and hcpcs_i is not None
                and dos_i is not None and billing_idx is not None
                and max(patient_i, hcpcs_i, dos_i, billing_idx) < len(new_row)):
            _pt, _hc = new_row[patient_i], new_row[hcpcs_i]
            _np2, _ds2 = new_row[billing_idx], new_row[dos_i]
            if _pt and _hc and _np2 and _DATE_ISO_RE.match(_ds2):
                _svc_seen.setdefault((_pt, _np2, _hc), []).append(
                    (_ds2[:10], len(cleaned) + 1))
        # Claim rollup: lines + charge per claim id (bounded distinct set).
        if claim_i is not None and claim_i < len(new_row) and new_row[claim_i]:
            _cid = new_row[claim_i]
            _agg = _claim_agg.get(_cid)
            if _agg is None:
                if len(_claim_agg) >= _CLAIM_CAP:
                    _claim_trunc = True
                else:
                    _agg = _claim_agg[_cid] = [0, 0.0]
            if _agg is not None:
                _agg[0] += 1
                if _charge_i is not None and _charge_i < len(new_row):
                    _cv = _to_number(new_row[_charge_i])
                    if _cv is not None:
                        _agg[1] += _cv
        # Distinct-value tracking for the data dictionary (capped per
        # column so an ID column can't hold millions of strings).
        for _ci4 in range(min(len(new_row), ncols)):
            _cv4 = new_row[_ci4]
            if _cv4:
                _ds4 = _col_distinct[_ci4]
                if len(_ds4) <= 1000:
                    _ds4.add(_cv4)
        if payer_i is not None and payer_i < len(new_row) and new_row[payer_i]:
            _payer_raw[new_row[payer_i]] = \
                _payer_raw.get(new_row[payer_i], 0) + 1
        if carc_i is not None and carc_i < len(new_row) and new_row[carc_i]:
            for _cp in re.split(r"[,;|\s]+", new_row[carc_i].strip().upper()):
                if _cp:
                    # Tally by the BARE code so "CO-45" and "45" merge into
                    # one denial reason (and hit the playbook); the raw
                    # spelling is kept as a display variant.
                    _cn = _CARC_GROUP_RE.sub("", _cp)
                    _carc_counts[_cn] = _carc_counts.get(_cn, 0) + 1
                    if _cn != _cp:
                        _carc_variants.setdefault(_cn, set()).add(_cp)
        if (_charge_i is not None and hcpcs_i is not None
                and _charge_i < len(new_row) and hcpcs_i < len(new_row)):
            _amt = _to_number(new_row[_charge_i])
            _code = new_row[hcpcs_i]
            if _amt is not None and _code:
                # (amount, 1-based output row) — the row index feeds the
                # charge-outlier worklist so flagged rows are downloadable.
                _code_charges.setdefault(_code, []).append(
                    (_amt, len(cleaned) + 1))
        cleaned.append(new_row)
        # Rules that fired on this row → worklist row indices and the
        # per-payer quality split ("which payer's feed is dirtiest").
        _fired = {_rk for _rk, _rv2 in res.sanity.items()
                  if _rv2 > _sanity_pre.get(_rk, 0)}
        # Per-rule cap only — an earlier global shut-off (_wl_open) meant a
        # rule first firing late in the file got a sanity count but ZERO
        # worklist rows once every rule seen so far was full.
        for _rk in _fired:
            _lst = res.flag_rows.setdefault(_rk, [])
            if len(_lst) < _WORKLIST_CAP:
                _lst.append(len(cleaned))
        if payer_i is not None and payer_i < len(new_row) and new_row[payer_i]:
            _fam2 = _payer_key(new_row[payer_i])
            _pq = _payer_q.get(_fam2)
            if _pq is None and len(_payer_q) < 50:
                _pq = _payer_q[_fam2] = {"rows": 0, "flagged": 0, "rules": {}}
            if _pq is not None:
                _pq["rows"] += 1
                if _fired:
                    _pq["flagged"] += 1
                    _ri = _pq.setdefault("rows_idx", [])
                    if len(_ri) < _WORKLIST_CAP:
                        _ri.append(len(cleaned))
                    for _rk in _fired:
                        _pq["rules"][_rk] = _pq["rules"].get(_rk, 0) + 1
        if ri % 500 == 0:
            cb("Cleaning rows", 0.30 + 0.55 * (ri / total))

    _spill.close()
    if _spill.failed:
        res.changelog_truncated = True
    elif _spill.path is not None:
        # Hand the GUARD to _write_output (not a bare path): the guard's
        # finalizer keeps covering the window between here and there, so
        # a cancel/crash in enrich/vendor code can't orphan the file.
        res.changelog_spill = _spill
    if _phi_log_skipped:
        res.warnings.append(
            f"De-identification is on: {_phi_log_skipped:,} change(s) in "
            "patient-identifier columns are counted but omitted from the "
            "change log — the audit trail must not carry the unmasked "
            "values you asked to remove.")
    res.n_rows_out = len(cleaned)

    # Ragged-row visibility: the pads/trims above, surfaced as structure
    # counts (and a warning when data was actually DROPPED). The sanity
    # counter/worklist already carry the per-row detail.
    if _ragged_padded or _ragged_truncated:
        res.structure["ragged_rows"] = {"padded": _ragged_padded,
                                        "truncated": _ragged_truncated}
        if _ragged_truncated:
            res.warnings.append(
                f"{_ragged_truncated:,} row(s) carried MORE cells than the "
                "header — the overflow cells were dropped (see the "
                "ragged-row worklist). Unquoted delimiters inside a text "
                "field are the usual cause; fix the export's quoting.")

    # Suspected duplicate CLAIMS — distinct rows (already past exact-row dedup)
    # that share the same billing provider · patient · date-of-service ·
    # procedure · amount. This is the double-billing signal exact dedup misses;
    # reported, never auto-removed (a repeat key can be legitimate). BOTH
    # members of each colliding key land in the worklist — the count stays
    # "extra rows" while the download shows the full collision.
    _dup_key_idx = [i for i in (billing_idx, patient_i, dos_i, hcpcs_i,
                                units_i, allowed_i) if i is not None]
    _has_when = dos_i is not None
    _has_who = billing_idx is not None or patient_i is not None
    _has_what = hcpcs_i is not None
    if _has_when and _has_who and _has_what and len(_dup_key_idx) >= 3:
        _first_row: Dict[str, int] = {}
        _dup = 0
        _dup_rows: set = set()
        for _rn, r in enumerate(cleaned, 1):
            parts = [r[i] for i in _dup_key_idx if i < len(r)]
            if all(p == "" for p in parts):
                continue
            k = "||".join(parts)
            _f = _first_row.get(k)
            if _f is not None:
                _dup += 1
                if len(_dup_rows) < _WORKLIST_CAP:
                    _dup_rows.add(_f)
                    _dup_rows.add(_rn)
            else:
                _first_row[k] = _rn
        if _dup:
            res.sanity["suspected-duplicate-claim"] = _dup
            res.flag_rows["suspected-duplicate-claim"] = \
                sorted(_dup_rows)[:_WORKLIST_CAP]

    # Conflicting-amount claims: the SAME who·when·what key billed at TWO OR
    # MORE different amounts. Disjoint from suspected-duplicate-claim (which
    # keys on the amount too) — this is the corrected-claim / re-bill signal.
    # Every row of a conflicted key (bounded per key) lands in the worklist.
    _amt_key_idx = [i for i in (billing_idx, patient_i, dos_i, hcpcs_i)
                    if i is not None]
    if (_has_when and _has_who and _has_what and len(_amt_key_idx) >= 3
            and _charge_i is not None):
        _key_amts: Dict[str, set] = {}
        _key_n: Dict[str, int] = {}
        _key_rows: Dict[str, List[int]] = {}
        for _rn, r in enumerate(cleaned, 1):
            parts = [r[i] for i in _amt_key_idx if i < len(r)]
            if all(p == "" for p in parts):
                continue
            _amt = _to_number(r[_charge_i]) if _charge_i < len(r) else None
            if _amt is None:
                continue
            k = "||".join(parts)
            _key_amts.setdefault(k, set()).add(round(_amt, 2))
            _key_n[k] = _key_n.get(k, 0) + 1
            _kr = _key_rows.setdefault(k, [])
            if len(_kr) < 200:            # bounded per key
                _kr.append(_rn)
        _conf = sum(_key_n[k] - 1 for k, amts in _key_amts.items()
                    if len(amts) > 1)
        if _conf:
            res.sanity["conflicting-amount-claim"] = _conf
            _conf_rows: set = set()
            for k, amts in _key_amts.items():
                if len(amts) > 1:
                    _conf_rows.update(_key_rows.get(k, ()))
                    if len(_conf_rows) >= _WORKLIST_CAP:
                        break
            res.flag_rows["conflicting-amount-claim"] = \
                sorted(_conf_rows)[:_WORKLIST_CAP]

    # NPI ↔ name conflict: the same billing NPI carrying materially
    # different provider names across rows — a keying/identity error and
    # the top cause of bad NPPES recovery input. Names are folded
    # (uppercase, punctuation stripped, whitespace collapsed) so case and
    # formatting variants never flag. Counted per extra row; bounded.
    if billing_idx is not None and name_idx is not None and cleaned:
        _npi_first_name: Dict[str, str] = {}
        _nc_rows: List[int] = []
        for _rn, r in enumerate(cleaned, 1):
            if billing_idx >= len(r) or name_idx >= len(r):
                continue
            _npi_v = r[billing_idx]
            _nm_v = r[name_idx]
            if len(_npi_v) != 10 or not _npi_v.isdigit() or not _nm_v.strip():
                continue
            _nk = " ".join(re.sub(r"[^A-Z0-9 ]", " ",
                                  _nm_v.upper()).split())
            if not _nk:
                continue
            _fk2 = _npi_first_name.get(_npi_v)
            if _fk2 is None:
                if len(_npi_first_name) < 50_000:
                    _npi_first_name[_npi_v] = _nk
            elif _nk != _fk2:
                if len(_nc_rows) < _WORKLIST_CAP:
                    _nc_rows.append(_rn)
                res.sanity["npi-name-conflict"] = \
                    res.sanity.get("npi-name-conflict", 0) + 1
        if _nc_rows:
            res.flag_rows["npi-name-conflict"] = _nc_rows

    # Top denial / adjustment reasons (revenue-cycle visibility), enriched
    # with the playbook: category (preventable / process / contractual /
    # patient-responsibility), the upstream screen that catches it, and
    # what to do. preventable_pct answers "how much of this denial volume
    # was catchable before submission?".
    if _carc_counts and carc_i is not None:
        _top = sorted(_carc_counts.items(), key=lambda kv: -kv[1])[:10]
        _top_entries = []
        for _c, _n in _top:
            _e: Dict[str, object] = {"code": _c, "count": _n}
            _vs = _carc_variants.get(_c)
            if _vs:
                # Group-prefixed spellings seen in the file ("CO-45") —
                # the code above is the bare CARC they normalized to.
                _e["as_seen"] = sorted(_vs)[:4]
            _pb = _rd.carc_playbook(_c) if _rd is not None else None
            if _pb:
                _e["category"] = _pb["category"]
                _e["linked_rule"] = _pb["rule"]
                _e["action"] = _pb["action"]
            _top_entries.append(_e)
        _prev_n = _known_n = 0
        for _c, _n in _carc_counts.items():
            _pb = _rd.carc_playbook(_c) if _rd is not None else None
            if _pb:
                _known_n += _n
                if _pb["category"] == "preventable":
                    _prev_n += _n
        res.denials = {"column": headers[carc_i],
                       "distinct": len(_carc_counts),
                       "top": _top_entries}
        if _known_n:
            res.denials["preventable_pct"] = round(100 * _prev_n / _known_n, 1)

    # Specialty mix from taxonomy codes on kept rows. Names come from the
    # NUCC display catalog; codes outside it still report (name None) —
    # catalog membership is deliberately NOT a validity domain.
    if _taxo_counts:
        _tops = sorted(_taxo_counts.items(), key=lambda kv: (-kv[1], kv[0]))
        res.specialties = [
            {"code": c, "n": n,
             "name": (_rd.taxonomy_specialty(c) if _rd is not None else None)}
            for c, n in _tops[:15]]

    # Data dictionary: what each column IS (detected role), how filled it
    # is, and what its values look like. PHI-safe: patient-identifier
    # columns never expose raw samples unless de-id already masked them.
    if headers:
        _role_of: Dict[int, str] = {}
        for _rset, _rname in (
                (npi_set, "npi"), (money_set, "money"), (date_set, "date"),
                (state_set, "state"), (zip_set, "zip"),
                (hcpcs_set, "hcpcs/cpt"), (sex_set, "sex"),
                (dx_set, "diagnosis"), (mod_set, "modifier"),
                (phone_set, "phone"), (taxo_set, "taxonomy"),
                (ndc_norm_set, "ndc"), (rev_set, "revenue-code"),
                (pos_set, "place-of-service"), (drg_set, "drg"),
                (pname_set, "provider-name")):
            for _ci4 in _rset:
                _role_of.setdefault(_ci4, _rname)
        for _ci4, _rname in (
                (payer_i, "payer"), (carc_i, "carc/denial"),
                (tob_i, "type-of-bill"), (dstat_i, "discharge-status"),
                (atype_i, "admission-type"), (member_i, "member-id"),
                (patient_i, "patient-id"), (claim_i, "claim-id"),
                (units_i, "units")):
            if _ci4 is not None:
                _role_of.setdefault(_ci4, _rname)
        _base = max(res.n_rows_in, 1)
        for _ci4, _h4 in enumerate(headers):
            _phi_k = _phi_kind(_norm_key(_h4))
            _ds4 = _col_distinct[_ci4]
            if _phi_k and not deid:
                _samples = ["(redacted — patient identifier)"]
            else:
                _samples = sorted(_ds4)[:3]
            res.dictionary.append({
                "column": _h4,
                "role": (_role_of.get(_ci4)
                         or (f"patient-phi:{_phi_k}" if _phi_k else "")),
                "fill_pct": round(100 * _col_filled[_ci4] / _base, 1),
                "distinct": len(_ds4) if len(_ds4) <= 1000 else None,
                "samples": _samples,
            })

    # Per-payer quality: which payer's feed is dirtiest — rows, share of
    # rows with at least one finding, and each payer's top rules. Top 10
    # payer families by volume (capped 50 tracked).
    if _payer_q:
        _tops_p = sorted(_payer_q.items(), key=lambda kv: -kv[1]["rows"])[:10]
        for _fam2, _pq in _tops_p:
            _tr = sorted(_pq["rules"].items(), key=lambda kv: -kv[1])[:3]
            res.payer_quality.append({
                "payer": _fam2,
                "rows": _pq["rows"],
                "flagged": _pq["flagged"],
                "clean_pct": round(
                    100 * (1 - _pq["flagged"] / _pq["rows"]), 1),
                "top_rules": [{"rule": r, "n": n} for r, n in _tr],
            })
            if _pq.get("rows_idx"):
                res.payer_flag_rows[_fam2] = _pq["rows_idx"]

    # Possible duplicate service: same patient + provider + code within the
    # profile window (default 3 days) on DIFFERENT dates. Same-date repeats
    # already belong to suspected-duplicate-claim / conflicting-amount.
    if _svc_seen and _dup_window > 0:
        from datetime import date as _dd
        _dup_rows: set = set()
        # Group by UNIQUE date first: a key repeated heavily on one date
        # (interface replays) made the old pairwise loop O(k²) with two
        # date parses per pair — hours of post-loop work on a big file.
        # Unique dates per key are bounded (~365/yr), each parsed once.
        for _entries in _svc_seen.values():
            if len(_entries) < 2:
                continue
            _by_date: Dict[str, List[int]] = {}
            for _d1, _r1 in _entries:
                _by_date.setdefault(_d1, []).append(_r1)
            if len(_by_date) < 2:
                continue        # same-date repeats belong to the dup-claim rules
            _dates = sorted(_by_date)
            try:
                _ords = [_dd.fromisoformat(_d1).toordinal()
                         for _d1 in _dates]
            except ValueError:
                continue
            for _i1 in range(len(_dates) - 1):
                for _i2 in range(_i1 + 1, len(_dates)):
                    if _ords[_i2] - _ords[_i1] > _dup_window:
                        break   # dates unique + sorted → nothing closer follows
                    _dup_rows.update(_by_date[_dates[_i1]])
                    _dup_rows.update(_by_date[_dates[_i2]])
        if _dup_rows:
            res.sanity["possible-duplicate-service"] = len(_dup_rows)
            res.flag_rows["possible-duplicate-service"] = \
                sorted(_dup_rows)[:_WORKLIST_CAP]

    # Claim rollup — the claim-level shape of a line-level file: how many
    # claims, how deep, and how the money distributes per claim.
    if _claim_agg:
        _lines = [int(a[0]) for a in _claim_agg.values()]
        _charges = sorted(a[1] for a in _claim_agg.values()
                          if a[1] > 0)
        _n_claims = len(_claim_agg)
        res.claims = {
            "column": headers[claim_i],
            "n_claims": _n_claims,
            "avg_lines": round(sum(_lines) / _n_claims, 2),
            "max_lines": max(_lines),
            "truncated": _claim_trunc,
        }
        if _charges:
            res.claims["charge"] = {
                "min": round(_charges[0], 2),
                "median": round(_quantile(_charges, 0.5), 2),
                "mean": round(sum(_charges) / len(_charges), 2),
                "max": round(_charges[-1], 2),
            }

    # Headered-but-empty columns — an extract/mapping defect worth surfacing.
    if res.n_rows_out:
        _empty = [headers[i] for i in range(ncols) if _col_filled[i] == 0]
        if _empty:
            res.structure["empty_columns"] = _empty
        # Per-column completeness profile ("which fields are 60% blank").
        # Denominator is rows IN: the cell counters run before the dedupe
        # skip, so dividing by rows_out would exceed 100% on files with
        # duplicates removed.
        res.column_fill = [
            {"column": headers[i], "filled": _col_filled[i],
             "pct": round(100.0 * _col_filled[i] / max(res.n_rows_in, 1), 1)}
            for i in range(ncols)]

    # Payer-name variant clustering (report-only). "BCBS", "Blue Cross of TX"
    # and "B.C.B.S." are the same payer spelled three ways — the single most
    # common grouping defect in real claims extracts. Cells are never
    # rewritten; the clusters are reported for the user to reconcile.
    if _payer_raw and payer_i is not None:
        _clusters: Dict[str, Dict[str, int]] = {}
        for raw, n in _payer_raw.items():
            _clusters.setdefault(_payer_key(raw), {})[raw] = n
        _multi = []
        for canon, variants in _clusters.items():
            if len(variants) >= 2:
                _multi.append({
                    "canonical": canon,
                    "total": sum(variants.values()),
                    "n_variants": len(variants),
                    "variants": [
                        {"value": v, "count": c} for v, c in
                        sorted(variants.items(), key=lambda kv: -kv[1])[:6]],
                })
        _multi.sort(key=lambda d: -d["total"])
        res.payer_variants = {
            "column": headers[payer_i],
            "distinct_raw": len(_payer_raw),
            "clusters": len(_clusters),
            "multi_spelling": _multi[:10],
        }

    # Statistical charge outliers per HCPCS code: values beyond 3×IQR fences
    # (Tukey far-out) within codes seen ≥10 times. A $25,000 office visit is
    # a data error or a story either way. Same type-7 quantile as the
    # analysis page's box plot, so the two agree.
    _out_total = 0
    _out_detail: List[Dict[str, object]] = []
    _out_rows: List[int] = []
    for _code, _pairs in _code_charges.items():
        if len(_pairs) < 10:
            continue
        _vals = sorted(a for a, _r in _pairs)
        _q1, _q3 = _quantile(_vals, 0.25), _quantile(_vals, 0.75)
        _iqr = _q3 - _q1
        if _iqr <= 0:
            continue
        _lo, _hi = (_q1 - _iqr_mult * _iqr, _q3 + _iqr_mult * _iqr)
        _n_out = 0
        for _a, _r in _pairs:
            if _a < _lo or _a > _hi:
                _n_out += 1
                if len(_out_rows) < _WORKLIST_CAP:
                    _out_rows.append(_r)
        if _n_out:
            _out_total += _n_out
            _out_detail.append({
                "code": _code, "n": len(_vals), "outliers": _n_out,
                "median": round(_quantile(_vals, 0.5), 2),
                "max": round(_vals[-1], 2)})
    if _out_total:
        res.sanity["charge-outlier"] = _out_total
        res.flag_rows["charge-outlier"] = sorted(_out_rows)[:_WORKLIST_CAP]
        _out_detail.sort(key=lambda d: -int(d["outliers"]))  # type: ignore[arg-type]
        res.outliers = _out_detail[:8]

    # Population analytics — the Tuva-class marts (service mix, encounters,
    # chronic conditions, volume integrity, readmissions, coding intensity).
    # Whole-table by nature, so streaming chunks skip it; report-only and
    # guarded so a mart failure never blocks the cleaning result.
    if not _stream_chunk and res.n_rows_out:
        cb("Population analytics (service mix · encounters · conditions)",
           0.56)
        try:
            from . import analytics as _analytics
            # Amount column the marts sum for charges. Header-detected
            # billed/allowed/paid win; failing those, any money-header or a
            # value-based sniff feeds ONLY analytics (never the cleaning pass)
            # so a charge column with a nonstandard header still yields
            # non-zero Population volume instead of a misleading $0.
            _claimed_cols: set = set()
            for _rs in (npi_set, money_set, date_set, zip_set, phone_set,
                        hcpcs_set, dx_set, mod_set, taxo_set, ndc_norm_set,
                        sex_set, rev_set, pos_set, drg_set, state_set,
                        pname_set):
                _claimed_cols |= _rs
            for _ix in (units_i, patient_i, member_i, claim_i, dos_i, todate_i,
                        dob_i, admit_i, disch_i, submit_i, payer_i, carc_i,
                        tob_i, dstat_i, atype_i, asrc_i, cond_i, occ_i, valc_i,
                        name_idx, ndc_idx, drug_idx, billing_idx, hcpcs_i,
                        allowed_i, billed_i, paid_i):
                if _ix is not None:
                    _claimed_cols.add(_ix)
            _amount_i = _analytics_amount_col(
                headers, cleaned, billed_i=billed_i, allowed_i=allowed_i,
                paid_i=paid_i, money_set=money_set, claimed=_claimed_cols)
            if _amount_i is not None and _amount_i != billed_i:
                res.structure["population_amount_column"] = headers[_amount_i]
            # Rendering/servicing/attending individual-NPI column, when one
            # exists among the detected NPI columns: coding intensity is
            # about who CODED the visit — the org-grain billing NPI washes
            # hot coders out in group practices. analytics.build prefers
            # rendering_i and labels provider_basis so renderers caption
            # the grain honestly.
            _rend_i = None
            for _ni in sorted(npi_set):
                if _ni == billing_idx or _ni >= len(headers):
                    continue
                _hh = re.sub(r"[^a-z0-9]", "", str(headers[_ni]).lower())
                if any(k in _hh for k in ("rendering", "servicing",
                                          "attending", "individualnpi")):
                    _rend_i = _ni
                    break
            res.population = _analytics.build(headers, cleaned, {
                "rev_set": rev_set, "tob_i": tob_i, "pos_set": pos_set,
                "hcpcs_i": hcpcs_i, "billed_i": _amount_i, "dx_set": dx_set,
                "patient_i": patient_i, "dos_i": dos_i, "admit_i": admit_i,
                "disch_i": disch_i, "billing_idx": billing_idx,
                "rendering_i": _rend_i,
                "readmit_window": _readmit_window,
            })
        except Exception:  # noqa: BLE001 — analytics never block cleaning
            res.population = None

    # Connector recommendation — computed from THIS file's detected columns
    # even on an offline run, so the Live-connectors panel can explain which
    # of the catalog apply and why ("openFDA — no NDC/J-code column here")
    # instead of a run looking broken when it lights only 2 sources. J-codes
    # are the load-bearing signal for infusion/oncology extracts that have no
    # NDC column: they still obviously need drug matching.
    _jcode_pct, _jcodes = _jcode_stats(cleaned, hcpcs_i)
    try:
        _bstat = (res.column_stats.get(headers[billing_idx])
                  if billing_idx is not None
                  and billing_idx < len(headers) else None) or {}
        _bcells = int(_bstat.get("cells") or 0)
        _bmiss = (int(_bstat.get("blank") or 0)
                  + int(_bstat.get("malformed") or 0)
                  + int(_bstat.get("checksum") or 0))
        _blank_npi_pct = (100.0 * _bmiss / _bcells) if _bcells else 0.0
    except Exception:  # noqa: BLE001
        _blank_npi_pct = 0.0
    try:
        from . import connectors as _conn
        # Static source metadata: rendered by the Live-connectors panel even
        # on an offline run, so it is set unconditionally (it used to ride
        # the enrich flag, leaving the panel empty on plain runs).
        res.catalog = _conn.catalog()
        res.connector_plan = _conn.plan({
            "has_npi": bool(npi_idx),
            "has_billing": billing_idx is not None,
            "blank_npi_pct": _blank_npi_pct,
            "has_ndc": ndc_idx is not None,
            "has_drug_name": drug_idx is not None,
            "jcode_pct": _jcode_pct,
            "has_hcpcs": hcpcs_i is not None,
            "has_dx": bool(dx_set),
            "has_taxonomy": bool(taxo_set),
            "rows": res.n_rows_out,
        })
    except Exception:  # noqa: BLE001 — planning never blocks cleaning
        res.connector_plan = None

    # Billed dollars per billing NPI (digits-keyed): lets the bounded online
    # screens spend their per-run caps on the HIGHEST-dollar providers
    # (NPPES verify) and lets the LEIE screen report billed-dollar exposure
    # behind an exclusion match, not just a count. Billed wins over allowed —
    # exposure is what was billed through the NPI.
    _npi_dollars: Dict[str, float] = {}
    if (enrich or deep) and billing_idx is not None:
        _amt_i2 = billed_i if billed_i is not None else allowed_i
        if _amt_i2 is not None:
            for _drow in cleaned:
                if billing_idx < len(_drow) and _amt_i2 < len(_drow):
                    _dk = _digits(_drow[billing_idx])
                    if len(_dk) == 10:
                        _dv = _to_number(_drow[_amt_i2])
                        if _dv is not None:
                            _npi_dollars[_dk] = (_npi_dollars.get(_dk, 0.0)
                                                 + _dv)

    # Optional live NPPES cross-check via the app's shared CMS connection.
    # Guarded end-to-end: any failure leaves res.nppes with a note and the
    # offline results stand.
    if enrich:
        cb("Verifying NPIs against the live NPPES registry", 0.58)
        try:
            _taxo_idx = min(taxo_set) if taxo_set else None
            res.nppes, res.recovered_rows, _nppes_fills = _enrich_via_nppes(
                cleaned, npi_idx, billing_idx, name_idx, state_idx,
                taxo_idx=_taxo_idx, headers=headers,
                weights=_npi_dollars or None)
            # Record the blanks NPPES filled: audited in the change log,
            # counted as repairs, credited to completeness. Bounded (≤ the
            # verify NPI cap), so appending to the preview list is safe.
            for _fe in _nppes_fills:
                _rule = _fe[4]
                res.repairs[_rule] = res.repairs.get(_rule, 0) + 1
                res.n_changes += 1
                res.n_cells_filled += 1
                if len(res.changelog) < _CHANGELOG_PREVIEW:
                    res.changelog.append(_fe)
        except Exception as exc:  # noqa: BLE001
            res.nppes = {"error": f"{type(exc).__name__}: {exc}"}
        # Drug connectors (RxNorm / openFDA) + the available-source catalog.
        cb("Resolving drugs via RxNorm / openFDA", 0.68)
        try:
            from . import connectors
            ndcs = ([row[ndc_idx] for row in cleaned
                     if ndc_idx is not None and ndc_idx < len(row)]
                    if ndc_idx is not None else [])
            drugs = ([row[drug_idx] for row in cleaned
                      if drug_idx is not None and drug_idx < len(row)]
                     if drug_idx is not None else [])
            # J-codes (HCPCS drug codes) resolve via RxNav's HCPCS crosswalk —
            # the path that lets an all-J-code infusion file light up RxNorm
            # with no NDC/drug-name column present.
            if ndcs or drugs or _jcodes:
                res.connectors = connectors.resolve_drugs(
                    ndcs, drugs, hcpcs=_jcodes)
                # Deterministic blanks-only drug fills from the resolved
                # concepts (drug-name ← NDC/J-code, and NDC ← name/J-code when
                # unambiguous). Audited + graded exactly like the NPPES fills.
                _rx = next((c for c in res.connectors
                            if c.get("id") == "rxnorm"), None)
                if _rx and _rx.get("concepts"):
                    for _fe in _apply_drug_fills(cleaned, headers, drug_idx,
                                                 ndc_idx, hcpcs_i,
                                                 _rx["concepts"]):
                        _rule = _fe[4]
                        res.repairs[_rule] = res.repairs.get(_rule, 0) + 1
                        res.n_changes += 1
                        res.n_cells_filled += 1
                        if len(res.changelog) < _CHANGELOG_PREVIEW:
                            res.changelog.append(_fe)
            else:
                res.connectors = []
        except Exception as exc:  # noqa: BLE001
            res.connectors = [{"id": "error",
                               "note": f"{type(exc).__name__}: {exc}"}]
        # Ordering / referring provider eligibility — verify those NPIs are
        # active in NPPES (a deactivated ordering/referring provider is a
        # frequent, avoidable denial). Bounded + guarded like the rest.
        if order_refer_idx:
            cb("Screening ordering / referring provider NPIs", 0.78)
            try:
                from . import nppes_bridge as _nb
                _or_npis = [row[i] for i in order_refer_idx
                            for row in cleaned if i < len(row) and row[i]]
                _orv = _nb.verify_npis(_or_npis)
                _recs = _orv.get("records") or {}
                _deact = sum(1 for r in _recs.values()
                             if r.get("status") == "not_found")
                res.order_referring = {
                    "columns": res.order_referring_columns,
                    "checked": _orv.get("checked", 0),
                    "active": _orv.get("active", 0),
                    "not_found": _deact,
                    "errors": _orv.get("errors", 0),
                    "degraded": bool(_orv.get("degraded")),
                    "note": _orv.get("note", ""),
                }
            except Exception as exc:  # noqa: BLE001
                res.order_referring = {"error": f"{type(exc).__name__}: {exc}"}

    # Compliance — OIG LEIE (offline) + Medicare PECOS (networked, bounded).
    # Deliberately OUTSIDE the enrich block: deep implies online consent and
    # the LEIE half needs no network at all, so a deep-without-enrich run
    # must not silently drop the exclusions screen (it used to leave
    # res.compliance None). PECOS still rides the deep flag inside screen().
    if enrich or deep:
        cb("Screening billing NPIs (OIG LEIE · PECOS)", 0.74)
        try:
            from . import compliance
            billing = ([row[billing_idx] for row in cleaned
                        if billing_idx is not None and billing_idx < len(row)]
                       if billing_idx is not None else [])
            res.compliance = (compliance.screen(
                billing, run_cms=deep,
                dollars_by_npi=_npi_dollars or None)
                if billing else [])
        except Exception as exc:  # noqa: BLE001
            res.compliance = [{"id": "error",
                               "note": f"{type(exc).__name__}: {exc}"}]
        # A stale installed LEIE pack silently weakens the screen it just
        # ran — say so on the run itself, not only in the packs panel. Env
        # CSV deployments manage their own freshness, so only the pack path
        # warns. Guarded: freshness reporting must never fail a run.
        try:
            import os as _os
            if not _os.environ.get("RCM_MC_LEIE_CSV"):
                from . import refdata_packs as _rpk
                _lst = next((p for p in _rpk.status()
                             if p.get("id") == "leie"), None)
                if _lst and _lst.get("installed") and _lst.get("stale"):
                    res.warnings.append(
                        "Compliance: the installed LEIE reference pack is "
                        f"{_lst.get('age_days', '?')} days old (refresh "
                        f"cadence {_lst.get('cadence_days', '?')} days) — "
                        "exclusion screening may miss recent OIG actions. "
                        "Refresh the leie pack.")
        except Exception:  # noqa: BLE001
            pass

    # Real vendored-engine pass: run the actual v48 field_validators +
    # consistency + dedup screens when pandas and the modules are available.
    # Guarded end-to-end — any failure just leaves res.advanced None and the
    # stdlib results stand on their own.
    if _stream_chunk:
        # Chunk mode: the pandas suggestions pass would re-run per chunk and
        # produce row indices no one can offset reliably — skipped, matching
        # zip-batch mode's "no fan-out" rule.
        res.advanced = None
    else:
        cb("Running the v49 deterministic engine (repairs · screens · issues)",
           0.82)
        try:
            from . import vendor_adapter
            adv = vendor_adapter.run(data, overrides)
            if adv:
                # The full companion can be large — keep it out of the JSON
                # the browser polls; retain it for the CSV/workbook downloads
                # and expose only a small preview inline.
                res.suggestions_records = adv.pop(
                    "suggestions_records", []) or []
                adv["suggestions_sample"] = res.suggestions_records[:25]
            res.advanced = adv
        except Exception:  # noqa: BLE001
            res.advanced = None

    # Selectable enrichment (enrich.py) — appends analysis-ready columns
    # (care setting, specialty, CBSA, Medicare benchmark) to the output the
    # pivot/workbook/CSV all read, plus report marts. Guarded end-to-end:
    # any failure leaves res.enrichment carrying the error note and the
    # cleaned output stands. Skipped in chunk mode like the other
    # whole-file passes (per-chunk marts would double-count).
    if enrichments and not _stream_chunk:
        cb("Enriching the cleaned table", 0.86)
        try:
            from . import enrich as _enrich
            _eout = _enrich.apply(
                headers, cleaned, {
                    "rev_set": rev_set, "tob_i": tob_i, "pos_set": pos_set,
                    "hcpcs_i": hcpcs_i, "billed_i": billed_i,
                    "dos_i": dos_i, "zip_set": zip_set,
                    "taxo_set": taxo_set, "billing_idx": billing_idx,
                    "name_idx": name_idx,
                }, list(enrichments), progress=cb)
            _eh = list(_eout.get("added_headers") or [])
            if _eh:
                _ecols = _eout.get("added_columns") or []
                for _ri, _erow in enumerate(cleaned):
                    _erow.extend(_ec[_ri] for _ec in _ecols)
                headers = list(headers) + _eh
            res.enrichment = {
                "marts": _eout.get("marts") or {},
                "results": _eout.get("results") or [],
                "requested": _eout.get("requested") or [],
                "columns_added": _eh,
            }
        except JobCancelled:
            raise
        except Exception as exc:  # noqa: BLE001 — enrichment never blocks
            res.enrichment = {
                "error": f"{type(exc).__name__}: {exc}",
                "requested": [str(x) for x in enrichments],
                "marts": {}, "results": [], "columns_added": [],
            }

    cb("Writing cleaned file", 0.90)
    res.out_name = _out_name(src_name)
    if _stream_chunk:
        res.chunk_payload = (headers, cleaned)
    else:
        _write_output(res, headers, cleaned)

    # Deep recovery — the full networked v49 pipeline, opt-in, guarded, with
    # its own timeout. Runs last so the fast deterministic results are already
    # complete; failure/timeout leaves them untouched.
    if deep:
        cb("Deep recovery — starting", 0.5)
        try:
            from . import deep_pipeline
            res.deep = deep_pipeline.run(data, src_name, progress=cb)
            if res.deep.get("workbook_path"):
                res.deep_workbook_path = res.deep["workbook_path"]
                res.deep_workbook_name = res.deep.get(
                    "workbook_name", "recovered.xlsx")
        except Exception as exc:  # noqa: BLE001
            res.deep = {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

    # Profile post-processing: disabled flags vanish entirely;
    # accepted flags stay visible but stop counting against the grade.
    _disabled = set(_prof.get("disabled_rules") or [])
    for _dr in _disabled:
        res.sanity.pop(_dr, None)
        res.flag_rows.pop(_dr, None)
    res.accepted_rules = sorted(set(_prof.get("accepted_rules") or []))
    res.profile_name = (str(_prof.get("name")) if _prof.get("name")
                        else None)

    # Longitudinal observability: aggregate summary only, never PHI, and
    # guarded so history storage can never fail the cleaning job. Trend
    # alerts compare against the PREVIOUS run of this file, so they must
    # be computed before this run is recorded. A streaming chunk is not a
    # run — bigfile.py records the merged result once.
    if not _stream_chunk:
        try:
            from . import history as _history
            res.trend_alerts = _history.trend_alerts(res.as_scorecard(),
                                                     src_name)
            _history.record_run(res.as_scorecard(), src_name)
        except Exception:  # noqa: BLE001
            pass
        _autofile_gaps(res)
    cb("Done", 1.0)
    return res


def _autofile_gaps(res: "CleanResult") -> None:
    """File machine-detected capability gaps onto the wishlist (source='auto',
    deduped by title) so patterns the cleaner can't yet handle surface on the
    build backlog instead of vanishing. Guarded end-to-end and generic-titled
    (no filename/PHI) so it can never block a run or leak claim data. Runs once
    per completed run, beside history recording."""
    try:
        gaps: List[Dict[str, str]] = []
        # Headers must exist for "no NPI column" to be meaningful — an X12
        # ack or empty file has no columns to detect anything in.
        if res.headers and not res.npi_columns:
            gaps.append({
                "category": "field",
                "title": "Detector found no NPI column in an upload",
                "details": "A cleaned file had no header the NPI detector "
                           "recognized. The NPI header hints may need "
                           "extending for this source system."})
        rows_out = res.n_rows_out or 0
        # A high unknown-code rate (only possible when the matching reference
        # pack is installed) points at a stale/incomplete pack, not dirty data.
        if rows_out >= 200:
            for rule, label in (("hcpcs-unknown-code", "HCPCS"),
                                ("icd10-unknown-code", "ICD-10-CM"),
                                ("taxonomy-unknown-code", "NUCC taxonomy")):
                n = int(res.sanity.get(rule) or 0)
                if n and n / rows_out >= 0.05:
                    gaps.append({
                        "category": "integration",
                        "title": f"High {label} unknown-code rate — "
                                 "reference pack may be stale",
                        "details": f"{label} codes are failing the installed "
                                   "reference set at a high rate; refreshing "
                                   "the pack may resolve it."})
        # Signals the survey found completing runs WITHOUT leaving a trace
        # on the backlog: header hints an NPI sniff had to rescue, ragged
        # files, unparseable date formats, headerless/X12-mismatch shapes,
        # zip members skipped. Titles are generic (dedupe-by-title, no PHI
        # or filenames); headers are metadata and ride in the details.
        _sniffed = res.structure.get("npi_sniffed_columns") or []
        if _sniffed:
            gaps.append({
                "category": "field",
                "title": "NPI column detected by value under a non-NPI "
                         "header",
                "details": "The value sniff adopted column(s) the header "
                           "hints missed: "
                           + ", ".join(str(s) for s in _sniffed[:5])
                           + ". Consider extending the NPI header hints."})
        rows_in = res.n_rows_in or 0
        _rr = res.structure.get("ragged_rows") or {}
        _rr_n = int(_rr.get("padded") or 0) + int(_rr.get("truncated") or 0)
        if rows_in >= 100 and _rr_n and _rr_n / rows_in >= 0.05:
            gaps.append({
                "category": "format",
                "title": "High ragged-row rate on an upload",
                "details": "Many rows had a different cell count than the "
                           "header — a delimiter/quoting defect in the "
                           "source export. A smarter re-alignment repair "
                           "may be worth building."})
        _du = int(res.sanity.get("date-unparseable") or 0)
        if rows_out >= 200 and _du and _du / rows_out >= 0.05:
            gaps.append({
                "category": "format",
                "title": "High date-unparseable rate — an unrecognized "
                         "date format",
                "details": "Date cells failed every parser (ISO, US, "
                           "Excel serial, CCYYMMDD) at a high rate; the "
                           "source may use a format worth adding."})
        _wblob = " ".join(res.warnings)
        if "no 837 claim" in _wblob:
            gaps.append({
                "category": "format",
                "title": "X12 upload carried no 837/835 claims",
                "details": "An X12 interchange without CLM/CLP segments "
                           "was uploaded (999/270/276 …). Support for "
                           "more transaction sets may be wanted."})
        if "no header row" in _wblob:
            gaps.append({
                "category": "format",
                "title": "Headerless file cleaned with synthesized "
                         "column names",
                "details": "A file with no header row was cleaned as "
                           "column_1..N. A saved mapping template for "
                           "this source would restore role detection."})
        _bs = res.structure.get("batch_skipped") or {}
        if _bs:
            gaps.append({
                "category": "format",
                "title": "Zip batch members were skipped",
                "details": f"{_bs.get('unsupported', 0)} unsupported-format "
                           f"member(s) and {_bs.get('over_cap', 0)} over "
                           "the batch cap were not cleaned. Batch-format "
                           "support / a bigger cap may be wanted."})
        # Integration gaps — online/reference screens that could not do
        # their job on THIS run. Missing data and missing egress should
        # land on the backlog exactly like missing parsers do.
        for _cmp in (res.compliance or []):
            if (isinstance(_cmp, dict) and _cmp.get("id") == "oig_leie"
                    and _cmp.get("available") is False):
                gaps.append({
                    "category": "integration",
                    "title": "LEIE exclusion screen ran without data",
                    "details": "Billing NPIs went through the compliance "
                               "screen but no LEIE dataset is loaded — "
                               "install the leie reference pack (or set "
                               "RCM_MC_LEIE_CSV) so exclusions actually "
                               "screen."})
                break
        _derr = (str(res.deep.get("error") or "").lower()
                 if isinstance(res.deep, dict) else "")
        if "timed out" in _derr or "could not reach" in _derr:
            gaps.append({
                "category": "integration",
                "title": "Deep recovery timed out — likely no egress",
                "details": "The networked deep pipeline hit its preflight/"
                           "watchdog limit. If this deployment has no "
                           "outbound network, deep mode needs an offline "
                           "story (or should say so up front)."})
        if any("could not reach" in str(_cn.get("note") or "").lower()
               for _cn in (res.connectors or []) if isinstance(_cn, dict)):
            gaps.append({
                "category": "integration",
                "title": "Drug connectors unreachable on an enrich run",
                "details": "RxNorm/openFDA lookups failed as a transport "
                           "error. If this deployment is offline, a "
                           "vendored drug reference pack would cover the "
                           "same checks without egress."})
        for _pr in (res.connector_plan or []):
            if (isinstance(_pr, dict)
                    and str(_pr.get("id") or "").startswith("pack_")
                    and _pr.get("applies")
                    and _pr.get("state") == "needs_data"):
                gaps.append({
                    "category": "integration",
                    "title": f"{_pr.get('name')} needed but not installed",
                    "details": str(_pr.get("reason") or "")})
        if not gaps:
            return
        from . import wishlist as _wishlist
        for g in gaps:
            _wishlist.auto_file(g["category"], g["title"], g["details"])
    except Exception:  # noqa: BLE001 — auto-file never blocks a run
        pass


def _clean_batch(members, src_name, *, drop_duplicates, deid, profile,
                 cb, skip_info: Optional[Dict[str, object]] = None
                 ) -> "CleanResult":
    """Clean every member of a zip batch and merge into one parent result.

    Each file runs through the full single-file pipeline (its own history
    record included); the parent carries summed counters so the report
    card blends all rows, plus per-file grades in ``batch``. Online modes
    (enrich/deep) are deliberately off in batch — one upload must not fan
    out into N network sweeps. ``skip_info`` (from zip_batch_members_ex)
    names the members that were NOT cleaned so the banner never implies
    full coverage of a partially-processed archive.
    """
    import zipfile as _zf
    res = CleanResult(delimiter="zip", headers=[])
    subs = []
    n = len(members)
    for i, (name, blob) in enumerate(members):
        cb(f"Cleaning {name} ({i + 1}/{n})", 0.05 + 0.85 * (i / max(n, 1)))
        sub = clean_bytes(blob, name, drop_duplicates=drop_duplicates,
                          deid=deid, profile=profile)
        subs.append((name, sub))
        res.n_rows_in += sub.n_rows_in
        res.n_rows_out += sub.n_rows_out
        res.n_dupes_removed += sub.n_dupes_removed
        res.n_cells_trimmed += sub.n_cells_trimmed
        res.n_changes += sub.n_changes
        res.n_cells_total += sub.n_cells_total
        res.n_cells_filled += sub.n_cells_filled
        for k, v in sub.repairs.items():
            res.repairs[k] = res.repairs.get(k, 0) + v
        for k, v in sub.sanity.items():
            res.sanity[k] = res.sanity.get(k, 0) + v
        q = sub.quality()
        res.batch.append({"file": name, "rows_in": sub.n_rows_in,
                          "rows_out": sub.n_rows_out,
                          "score": q["score"], "letter": q["letter"],
                          "repairs": sum(sub.repairs.values()),
                          "findings": sum(sub.sanity.values())})
        res.warnings.extend(f"{name}: {w}" for w in sub.warnings)
        # De-id confirmation must survive the merge — a masked batch that
        # reports deid: null tells the user redaction did NOT happen.
        if sub.deid_applied:
            res.deid_applied = True
            res.deid_cells += sub.deid_cells
            for _dc in sub.deid_columns:
                if _dc not in res.deid_columns:
                    res.deid_columns.append(_dc)
    # Skipped members surface BEFORE the banner: a mixed archive must never
    # imply full coverage. Unsupported formats are named (capped); the
    # member cap is stated with the count over it.
    _skips: List[str] = []
    _si = skip_info or {}
    _uns = list(_si.get("skipped_unsupported") or [])
    _over = int(_si.get("over_cap") or 0)
    if _uns:
        _named = ", ".join(str(u) for u in _uns[:8])
        if len(_uns) > 8:
            _named += f", … ({len(_uns) - 8} more)"
        _skips.append(f"{len(_uns)} member(s) in an unsupported format "
                      f"were not cleaned: {_named} — export those as CSV "
                      "and re-upload.")
    if _over:
        _skips.append(f"{_over} member(s) over the "
                      f"{_si.get('cap', _BATCH_MEMBER_CAP)}-file batch cap "
                      "were not cleaned — split the archive and upload "
                      "the rest separately.")
    if _uns or _over:
        res.structure["batch_skipped"] = {"unsupported": len(_uns),
                                          "over_cap": _over}
    for _sk in reversed(_skips):
        res.warnings.insert(0, _sk)
    res.warnings.insert(0, (
        f"Batch mode: {n} file(s) cleaned from the zip — per-file grades "
        "on the Quality tab; the download is a zip of the cleaned files. "
        "NPPES verification and deep recovery are skipped in batch mode."))
    out_path = WORKDIR / f"{uuid.uuid4().hex}_batch.zip"
    with _zf.ZipFile(out_path, "w", _zf.ZIP_DEFLATED) as z:
        for name, sub in subs:
            if sub.out_path:
                z.write(sub.out_path, sub.out_name)
    res.out_path = str(out_path)
    _stem = Path(src_name).stem or "batch"
    res.out_name = f"{_stem}_cleaned.zip"
    try:
        from . import history as _history
        res.trend_alerts = _history.trend_alerts(res.as_scorecard(),
                                                 src_name)
        _history.record_run(res.as_scorecard(), src_name)
    except Exception:  # noqa: BLE001 — observability never blocks cleaning
        pass
    _autofile_gaps(res)
    cb("Done", 1.0)
    return res


def dictionary_csv(result: "CleanResult") -> str:
    """The data dictionary as CSV text — one row per column. Shared by the
    ?fmt=dictionary download and the everything-bundle."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["column", "detected_role", "fill_pct", "distinct_values",
                "sample_values"])
    for e in result.dictionary:
        w.writerow([e["column"], e["role"], e["fill_pct"],
                    e["distinct"] if e["distinct"] is not None else "1000+",
                    " | ".join(str(s) for s in e["samples"])])
    return buf.getvalue()


def _apply_drug_fills(cleaned, headers, drug_idx, ndc_idx, hcpcs_i, concepts):
    """Blanks-only, audited drug fills from resolved RxNorm concepts.

    Two safe directions: (1) a blank drug-name filled from the row's NDC or
    J-code concept (a code maps to exactly one ingredient — unambiguous);
    (2) a blank NDC filled from the row's drug-name / J-code concept ONLY when
    that concept resolves to a single NDC (otherwise ambiguous, skipped).
    Mirrors the NPPES fill contract: returns ``(row_1based, col, "", new,
    rule)`` tuples and mutates the cleaned rows in place. Never overwrites a
    populated cell."""
    fills: List[Tuple[int, str, str, str, str]] = []
    if not concepts:
        return fills
    from . import connectors as _conn
    ndc_c = concepts.get("ndc") or {}
    name_c = concepts.get("name") or {}
    hcpcs_c = concepts.get("hcpcs") or {}

    def _col(ci):
        return headers[ci] if headers and ci < len(headers) else str(ci)

    for ridx, row in enumerate(cleaned):
        # (1) blank drug-name ← NDC concept, else J-code concept
        if (drug_idx is not None and drug_idx < len(row)
                and not row[drug_idx].strip()):
            name = ""
            if ndc_idx is not None and ndc_idx < len(row) and row[ndc_idx].strip():
                c = ndc_c.get(_conn._fold_drug_key(row[ndc_idx], "ndc"))
                name = (c or {}).get("name", "")
            if not name and hcpcs_i is not None and hcpcs_i < len(row):
                jk = _conn._fold_drug_key(row[hcpcs_i], "hcpcs")
                if _JCODE_RE.match(jk):
                    name = (hcpcs_c.get(jk) or {}).get("name", "")
            if name:
                fills.append((ridx + 1, _col(drug_idx), "", name,
                              "drug-name-from-code"))
                row[drug_idx] = name
        # (2) blank NDC ← drug-name / J-code concept, only when unambiguous
        if (ndc_idx is not None and ndc_idx < len(row)
                and not row[ndc_idx].strip()):
            src = None
            if drug_idx is not None and drug_idx < len(row) and row[drug_idx].strip():
                src = name_c.get(_conn._fold_drug_key(row[drug_idx], "name"))
            if src is None and hcpcs_i is not None and hcpcs_i < len(row):
                jk = _conn._fold_drug_key(row[hcpcs_i], "hcpcs")
                if _JCODE_RE.match(jk):
                    src = hcpcs_c.get(jk)
            ndcs = (src or {}).get("ndcs") or []
            if len(ndcs) == 1 and str(ndcs[0]).strip():
                fills.append((ridx + 1, _col(ndc_idx), "", str(ndcs[0]),
                              "ndc-from-drug"))
                row[ndc_idx] = str(ndcs[0])
    return fills


def _enrich_via_nppes(cleaned, npi_idx, billing_idx, name_idx, state_idx,
                      taxo_idx=None, headers=None, weights=None):
    """Run NPPES verify + recover over the cleaned rows via nppes_bridge.

    ``weights`` (optional, NPI digits → summed billed dollars) makes the
    bridge spend its bounded verify cap on the highest-dollar NPIs instead
    of first-seen order — on multi-thousand-provider extracts the cap
    covers <1% of NPIs and first-seen made that 1% arbitrary.

    Returns ``(payload, recovered_rows, fills)`` where ``fills`` is a list
    of ``(row_1based, column_header, before, after, rule)`` — blank
    provider name / state / taxonomy cells filled from the AUTHORITATIVE
    NPPES record of a verified (active) billing NPI. Provider identifiers
    are never PHI, the fill is deterministic (one canonical record per
    NPI) and non-destructive (blanks only), so it is safe and audited —
    the same contract as the recovered-NPI write-back and state-from-zip.
    ``fills`` is empty when the bridge is unavailable.
    """
    from . import nppes_bridge
    if not nppes_bridge.available():
        return ({"note": "NPPES connection unavailable in this deployment."},
                {}, [])

    # Distinct NPIs across every detected NPI column, for verification.
    all_npis: List[str] = []
    for row in cleaned:
        for i in npi_idx:
            if i < len(row):
                all_npis.append(row[i])
    verify = nppes_bridge.verify_npis(all_npis, weights=weights)

    # Fill blank name / state / taxonomy from the verified billing NPI's
    # canonical record. Bounded by verify's distinct-NPI cap (≤40) and
    # blanks-only, so it can't rewrite anyone's data or run away.
    fills: List[Tuple[int, str, str, str, str]] = []
    _recs = (verify.get("records") or {}) if isinstance(verify, dict) else {}
    if billing_idx is not None and _recs:
        _fill_cols = []  # (col_index, record_key, rule_id)
        if name_idx is not None:
            _fill_cols.append((name_idx, "name", "name-from-nppes"))
        if state_idx is not None:
            _fill_cols.append((state_idx, "state", "state-from-nppes"))
        if taxo_idx is not None:
            _fill_cols.append((taxo_idx, "taxonomy", "taxonomy-from-nppes"))
        for ridx, row in enumerate(cleaned):
            if billing_idx >= len(row):
                continue
            rec = _recs.get(_digits(row[billing_idx]))
            if not rec or rec.get("status") != "active":
                continue
            for _ci, _key, _rule in _fill_cols:
                if _ci >= len(row) or row[_ci].strip():
                    continue  # blanks only — never overwrite
                _new = str(rec.get(_key) or "").strip()
                if not _new:
                    continue
                _col = (headers[_ci] if headers and _ci < len(headers)
                        else str(_ci))
                fills.append((ridx + 1, _col, "", _new, _rule))
                row[_ci] = _new

    # Recovery queries: rows whose billing NPI is not valid but which carry a
    # provider name (+ state) we can search on. Track which rows each query
    # covers so a resolved candidate can be written back to every matching row.
    recover: Dict[str, object] = {"note": "No provider-name column to recover from."}
    recovered_rows: Dict[int, str] = {}
    if billing_idx is not None and name_idx is not None:
        queries: List[Dict[str, str]] = []
        key_to_rows: Dict[tuple, List[int]] = {}
        for ridx, row in enumerate(cleaned):
            if billing_idx >= len(row):
                continue
            if classify_npi(row[billing_idx]) == "valid":
                continue
            name = row[name_idx] if name_idx < len(row) else ""
            if not name.strip():
                continue
            state = row[state_idx] if (state_idx is not None
                                       and state_idx < len(row)) else ""
            queries.append({"row": str(ridx + 1), "name": name, "state": state})
            key = (name.strip().lower(), (state or "").strip().upper()[:2])
            key_to_rows.setdefault(key, []).append(ridx + 1)
        if queries:
            recover = nppes_bridge.recover_candidates(queries)
            # A match with exactly one candidate is confident enough to write
            # back to every row that shared that provider name + state.
            for m in recover.get("matches", []):
                cands = m.get("candidates") or []
                if len(cands) != 1:
                    continue
                key = ((m.get("query") or "").strip().lower(),
                       (m.get("state") or "").strip().upper()[:2])
                for rownum in key_to_rows.get(key, []):
                    recovered_rows[rownum] = cands[0]["npi"]
        else:
            recover = {"note": "No rows needed NPI recovery."}

    payload = {"verify": verify, "recover": recover,
               "recovered_written": len(recovered_rows),
               "filled_from_nppes": len(fills),
               "source": "NPPES via rcm_mc.data_public.nppes_api_client"}
    return payload, recovered_rows, fills


def sample_csv() -> str:
    """A small illustrative claims file exercising every check: a clean row, a
    duplicate, a blank billing NPI, a malformed NPI, a checksum failure, a
    whitespace case, a future service date and a money-ordering violation.
    NPIs here are Luhn-valid synthetic examples — not real providers.
    """
    return (
        "ClaimID,BillingProviderNPI,RenderingNPI,OrganizationName,"
        "ProviderState,ChargeAmt,AllowedAmt,PaidAmt,DateOfService,HCPCS\n"
        "1001,1679576722,1234567893,Mercy General Hospital,OH,420,300,240,2024-02-11,99213\n"
        "1001, 1679576722 ,1234567893,Mercy General Hospital,OH,420,300,240,2024-02-11,99213\n"
        "1003,,1245319599,Riverbend Clinic,TX,180,140,110,2024-03-02,99214\n"
        "1004,99999,1245319599,Riverbend Clinic,TX,180,140,110,2024-03-02,99214\n"
        "1005,1234567890,1679576722,Summit Surgical Center,CA,900,700,760,2024-04-19,ABCDE\n"
        "1006,1699999984,1234567893,Lakeside Imaging,FL,250,200,150,2099-01-01,70450\n"
    )


def _out_name(src_name: str) -> str:
    stem = Path(src_name or "claims").stem.replace(" ", "_") or "claims"
    return f"{stem}_cleaned.csv"


def _write_output(res: CleanResult, headers: List[str],
                  rows: List[List[str]]) -> None:
    # When NPPES recovery filled in NPIs, append a non-destructive column so
    # the original billing-NPI column is preserved and the recovery is visible.
    out_headers = list(headers)
    out_rows = [list(r) for r in rows]
    if res.recovered_rows and headers:
        out_headers = out_headers + ["recovered_billing_npi"]
        for i, r in enumerate(out_rows):
            r.append(res.recovered_rows.get(i + 1, ""))

    token = uuid.uuid4().hex
    out_path = WORKDIR / f"{token}_{res.out_name}"
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        if out_headers:
            writer.writerow([_defang_cell(h) for h in out_headers])
        for r in out_rows:
            writer.writerow([_defang_cell(c) for c in r])
    res.out_path = str(out_path)

    # Corrections companion (v49 suggested_fixes) as its own CSV download.
    if res.suggestions_records:
        cols = list(res.suggestions_records[0].keys())
        stem = res.out_name[:-len("_cleaned.csv")] if res.out_name.endswith(
            "_cleaned.csv") else res.out_name.rsplit(".", 1)[0]
        res.companion_name = f"{stem or 'claims'}_corrections.csv"
        comp_path = WORKDIR / f"{token}_{res.companion_name}"
        with open(comp_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow([_defang_cell(c) for c in cols])
            for rec in res.suggestions_records:
                writer.writerow([_defang_cell(str(rec.get(c, ""))) for c in cols])
        res.companion_path = str(comp_path)

    # Cell-level audit trail: every change the cleaner made (row · column ·
    # before → after · rule), as its own CSV download. On de-identified
    # runs, patient-identifier columns never reach this file at all
    # (skipped at log time) — the artifact must stay as clean as the
    # output the user asked for.
    # res.changelog holds the first _CHANGELOG_PREVIEW entries; anything
    # past that was spilled to disk during the row loop and is appended
    # here in BINARY mode — text mode's universal-newline translation
    # would rewrite CR/LF inside quoted cells (multiline addresses) and
    # make identical inputs read differently across the preview boundary.
    if res.changelog:
        stem = res.out_name[:-len("_cleaned.csv")] if res.out_name.endswith(
            "_cleaned.csv") else res.out_name.rsplit(".", 1)[0]
        res.changelog_name = f"{stem or 'claims'}_changelog.csv"
        log_path = WORKDIR / f"{token}_{res.changelog_name}"
        with open(log_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["row", "column", "before", "after", "rule"])
            for row_i, col, before, after, rule in res.changelog:
                writer.writerow([row_i, _defang_cell(col),
                                 _defang_cell(before), _defang_cell(after),
                                 rule])
        spill = res.changelog_spill
        if spill is not None:
            try:
                if getattr(spill, "path", None):
                    with open(log_path, "ab") as out_b, \
                            open(spill.path, "rb") as in_b:
                        shutil.copyfileobj(in_b, out_b, 1 << 20)
            except OSError:
                res.changelog_truncated = True
            finally:
                spill.discard()
                res.changelog_spill = None
        res.changelog_path = str(log_path)

    # A styled multi-tab .xlsx workbook (Cleaned · issues · scorecard) via the
    # app's stdlib xlsx writer — no new dependency. Guarded: a workbook build
    # failure never blocks the CSV download.
    try:
        from . import report
        stem = res.out_name[:-len("_cleaned.csv")] if res.out_name.endswith(
            "_cleaned.csv") else res.out_name.rsplit(".", 1)[0]
        res.workbook_name = f"{stem or 'claims'}_report.xlsx"
        wb_bytes = report.build_workbook(res, out_headers, out_rows)
        wb_path = WORKDIR / f"{token}_{res.workbook_name}"
        with open(wb_path, "wb") as fh:
            fh.write(wb_bytes)
        res.workbook_path = str(wb_path)
    except Exception:  # noqa: BLE001
        res.workbook_path = None


# ------------------------------------------------------------- job management --
class JobCancelled(Exception):
    """Raised inside a job's progress callback when the user cancels — the
    only sanctioned way to stop a multi-hour streaming run mid-flight."""


@dataclass
class Job:
    job_id: str
    name: str
    frac: float = 0.0
    msg: str = "Queued"
    done: bool = False
    error: Optional[str] = None
    result: Optional[CleanResult] = None
    created: float = 0.0
    cancelled: bool = False

    def status_dict(self) -> Dict[str, object]:
        d: Dict[str, object] = {
            "job_id": self.job_id, "frac": round(self.frac, 3),
            "msg": self.msg, "done": self.done, "error": self.error,
        }
        # A 10 GB job runs for hours — the page needs "how much longer",
        # not just a percentage. Linear projection over elapsed/frac is
        # honest enough once real work is under way.
        if not self.done and self.created and self.frac >= 0.03:
            elapsed = max(0.0, time.time() - self.created)
            d["elapsed_secs"] = round(elapsed)
            d["eta_secs"] = round(elapsed * (1.0 - self.frac) / self.frac)
        if self.result is not None:
            d["scorecard"] = self.result.as_scorecard()
            d["download"] = f"/npi-cleaner/download/{self.job_id}"
        return d


class JobManager:
    """Thread-safe registry of cleaning jobs. One instance per server."""

    def __init__(self, max_jobs: int = 200) -> None:
        self._jobs: Dict[str, Job] = {}
        self._lock = threading.Lock()
        self._max = max_jobs

    def _evict(self) -> None:
        if len(self._jobs) <= self._max:
            return
        # Drop the oldest completed jobs first.
        finished = sorted(
            (j for j in self._jobs.values() if j.done),
            key=lambda j: j.created)
        for j in finished[: len(self._jobs) - self._max]:
            self._jobs.pop(j.job_id, None)

    def submit(self, data: bytes, name: str, *,
               drop_duplicates: bool = True, enrich: bool = False,
               deep: bool = False, deid: bool = False,
               enrichments: Optional[List[str]] = None,
               overrides: Optional[Dict[str, str]] = None,
               profile: Optional[Dict[str, object]] = None) -> str:
        job_id = uuid.uuid4().hex
        job = Job(job_id=job_id, name=name, created=time.time())
        with self._lock:
            self._jobs[job_id] = job
            self._evict()

        def _run() -> None:
            def cb(msg: str, frac: float) -> None:
                if job.cancelled:
                    raise JobCancelled()
                job.msg, job.frac = msg, float(frac)
            try:
                job.result = clean_bytes(
                    data, name, drop_duplicates=drop_duplicates,
                    enrich=enrich, deep=deep, deid=deid,
                    enrichments=enrichments,
                    overrides=overrides, profile=profile, progress=cb)
                job.frac, job.msg, job.done = 1.0, "Done", True
            except JobCancelled:
                job.error = "Cancelled by user."
                job.msg, job.done = "Cancelled", True
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                job.error = f"{type(exc).__name__}: {exc}"
                job.done = True

        threading.Thread(target=_run, daemon=True).start()
        return job_id

    def submit_path(self, path: str, name: str, *,
                    drop_duplicates: bool = True, enrich: bool = False,
                    deep: bool = False, deid: bool = False,
                    enrichments: Optional[List[str]] = None,
                    overrides: Optional[Dict[str, str]] = None,
                    profile: Optional[Dict[str, object]] = None,
                    cleanup: bool = True) -> str:
        """Start a cleaning job over a file already ON DISK (a spooled
        upload). Huge files stream through bigfile.clean_path in bounded
        memory instead of being read whole; the spool file is removed when
        the job finishes either way (``cleanup``)."""
        job_id = uuid.uuid4().hex
        job = Job(job_id=job_id, name=name, created=time.time())
        with self._lock:
            self._jobs[job_id] = job
            self._evict()

        def _run() -> None:
            def cb(msg: str, frac: float) -> None:
                if job.cancelled:
                    raise JobCancelled()
                job.msg, job.frac = msg, float(frac)
            try:
                from . import bigfile
                job.result = bigfile.clean_path(
                    path, name, drop_duplicates=drop_duplicates,
                    enrich=enrich, deep=deep, deid=deid,
                    enrichments=enrichments,
                    overrides=overrides, profile=profile, progress=cb)
                job.frac, job.msg, job.done = 1.0, "Done", True
            except JobCancelled:
                job.error = "Cancelled by user."
                job.msg, job.done = "Cancelled", True
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                job.error = f"{type(exc).__name__}: {exc}"
                job.done = True
            finally:
                if cleanup:
                    try:
                        Path(path).unlink(missing_ok=True)
                    except OSError:
                        pass

        threading.Thread(target=_run, daemon=True).start()
        return job_id

    def cancel(self, job_id: str) -> bool:
        """Flag a running job for cancellation. The worker notices at its
        next progress tick (between chunks on a streaming run) — there is
        no thread kill, so partial artifacts are cleaned up normally."""
        with self._lock:
            job = self._jobs.get(job_id)
            if job is None or job.done:
                return False
            job.cancelled = True
            return True

    def get(self, job_id: str) -> Optional[Job]:
        with self._lock:
            return self._jobs.get(job_id)


# One process-wide manager, created lazily so importing this module is cheap.
_MANAGER: Optional[JobManager] = None
_MANAGER_LOCK = threading.Lock()


def manager() -> JobManager:
    global _MANAGER
    if _MANAGER is None:
        with _MANAGER_LOCK:
            if _MANAGER is None:
                _MANAGER = JobManager()
    return _MANAGER
