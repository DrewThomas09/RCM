"""Turn a pipeline Result into the formatted multi-tab output workbook."""

import datetime as dt
import re

import numpy as np
import pandas as pd

from . import config, excelio


def _money(x):
    try:
        return f"${float(x):,.0f}"
    except Exception:
        return str(x)



def _v43_sheets(result):
    """v43 diligence-signal sheets: recovery calibration head-to-head, capture /
    completeness, and taxonomy-coherence flags. Built defensively so a missing
    piece yields an informative one-row frame rather than a crash."""
    import pandas as pd
    sheets = []
    pc = getattr(result, "prob_calibration", {}) or {}
    if pc.get("status") == "ok":
        inc, mod = pc["incumbent"], pc["model"]
        head = pd.DataFrame([
            {"method": "incumbent_confidence (tier_weight x purity)",
             "brier": inc["brier"], "ece": inc["ece"], "auc": inc["auc"],
             "held_out_rows": inc["n"], "base_rate": inc["base_rate"]},
            {"method": "calibrated_model (learned + isotonic)",
             "brier": mod["brier"], "ece": mod["ece"], "auc": mod["auc"],
             "held_out_rows": mod["n"], "base_rate": mod["base_rate"]},
        ])
        sheets.append(("Recovery_Calibration", head,
                       "v43 - Is the stated recovery confidence calibrated?",
                       pc.get("verdict", "") + " Brier and ECE lower is better; AUC "
                       "higher is better. ECE is how far the stated confidence sits "
                       "from the actual hit rate, dollar-weighted. This is measured on "
                       "held-out rows whose true billing NPI we hid and recovered."))
        rel = mod.get("reliability")
        if isinstance(rel, pd.DataFrame) and not rel.empty:
            sheets.append(("Calibration_Reliability", rel,
                           "v43 - Reliability curve of the calibrated model",
                           "Per confidence bin: mean stated confidence vs actual hit "
                           "rate. A calibrated model sits on the diagonal (gap near 0)."))
        coef = pc.get("coefficients")
        if isinstance(coef, pd.DataFrame) and not coef.empty:
            sheets.append(("Calibration_Coefficients", coef,
                           "v43 - What the model learned each signal is worth",
                           "Standardized logistic coefficients over the recovery "
                           "evidence signals. Positive lifts the probability a "
                           "recovered NPI is correct. This is why the score is "
                           "explainable, not a black box."))
    cap = getattr(result, "capture", {}) or {}
    if cap:
        ch = cap.get("channel_completeness")
        if isinstance(ch, pd.DataFrame) and not ch.empty:
            band = cap.get("implied_capture_band", {})
            sheets.append(("Capture_ByChannel", ch,
                           "v43 - Panel completeness by payer channel",
                           (band.get("note", "") if isinstance(band, dict) else "")
                           + " under_captured flags channels a medical claims panel "
                           "systematically misses (VA/military, cash, managed "
                           "Medicaid). The true book is larger there by an unknown "
                           "amount."))
        dr = cap.get("drug_capture_flags")
        if isinstance(dr, pd.DataFrame) and not dr.empty:
            sheets.append(("Capture_ByDrug", dr,
                           "v43 - Panel completeness by drug",
                           "well_captured = clinician-administered buy-and-bill "
                           "J-codes a medical panel sees. Poorly captured codes are "
                           "pharmacy-benefit or self-administered and under-represented "
                           "here vs the true book."))
    tax = getattr(result, "taxonomy_flags", None)
    if isinstance(tax, pd.DataFrame) and not tax.empty:
        sheets.append(("Recovery_Taxonomy_Flags", tax,
                       "v43 - Recovered NPIs whose specialty cannot bill the drug",
                       tax.attrs.get("note", "") + " A recovered NPI resolving to a "
                       "provider who cannot plausibly administer the billed drug is a "
                       "recovery-precision risk. Treat these as leads to verify."))

    # ---- v44: agreement, ledger, seller request, manifest ----
    led = getattr(result, "evidence_ledger", None)
    if isinstance(led, pd.DataFrame) and not led.empty:
        from . import ledger as _LED
        roll = _LED.basecase_rollup(led)
        if not roll.empty:
            sheets.append(("Ledger_BaseCase_Rollup", roll,
                           "v44 - Recovered dollars: base-case safe vs leads",
                           "base_case_safe = point attribution, not demoted, not a "
                           "two-method disagreement, clears the calibrated-probability "
                           "floor when available. lead_verify = everything else. This "
                           "is the number to cite: how much of the recovered book is "
                           "defensible in a base case."))
        sheets.append(("Evidence_Ledger", led,
                       "v44 - Per-value audit record",
                       led.attrs.get("note", "") + " One row per recovered or repaired "
                       "value: what changed, by what method, how sure (incumbent "
                       "confidence, calibrated probability, two-method agreement), on "
                       "what basis and vintage, and whether it is safe for a base case."))
    ag = getattr(result, "agreement_summary", None)
    if isinstance(ag, pd.DataFrame) and not ag.empty:
        sheets.append(("Recovery_Agreement", ag,
                       "v44 - Two-method recovery agreement",
                       ag.attrs.get("note", "") + " Each blank is recovered two "
                       "independent ways (in-file pattern vs CMS pool). Agreement is a "
                       "model-free precision signal; disagreement is a lead, not a fact."))
    dq = getattr(result, "disagreement_queue", None)
    if isinstance(dq, pd.DataFrame) and not dq.empty:
        sheets.append(("Recovery_Disagreements", dq,
                       "v44 - Recoveries where the two methods disagree",
                       dq.attrs.get("note", "") + " Ranked by dollars. These are the "
                       "recovered NPIs most worth verifying before relying on them."))
    sr = getattr(result, "seller_request", None)
    if isinstance(sr, pd.DataFrame) and not sr.empty:
        sheets.append(("Seller_Data_Request", sr,
                       "v44 - Prioritized data-request list for the seller",
                       sr.attrs.get("note", "") + " What to ask the seller for and the "
                       "specific check or recovery each field unblocks. Send this with "
                       "the data-room follow-up."))
    man = getattr(result, "run_manifest", None)
    if isinstance(man, dict) and man:
        from . import run_manifest as _RMAN
        sheets.append(("Run_Manifest_v44", _RMAN.manifest_frame(man),
                       "v44 - Reproducibility record",
                       "Run id, input file hash, tool version, and the dated CMS "
                       "reference seeds this run used. The answer to why did this flag."))
    sdm = getattr(result, "specialty_drug_mix", None)
    if isinstance(sdm, pd.DataFrame) and not sdm.empty:
        sheets.append(("Specialty_Drug_Mix", sdm,
                       "v48 - Which specialties bill each top drug",
                       sdm.attrs.get("note", "") + " A data-driven read of who bills "
                       "each drug, used to score whether a recovered NPI's specialty is "
                       "a plausible biller. Refit from real Medicare utilization or RIF "
                       "carrier data for measured frequencies."))
    return sheets


def write_report(result, out_path):
    s = result.stats
    bt = result.bt

    # v23: operator-leaderboard headline for the README (top recoverable + UNMEASURABLE count)
    _lb = getattr(result, "backtest_by_operator", pd.DataFrame())
    _top_operators_str, _unmeasurable_count = "(none measured)", 0
    if _lb is not None and not _lb.empty and "verdict" in _lb.columns:
        _ranked = _lb[_lb["verdict"].isin(["HIGH_CONFIDENCE_RECOVERABLE", "MODERATE"])]
        _ranked = _ranked.sort_values("precision_lcb", ascending=False).head(3)
        if not _ranked.empty:
            _top_operators_str = ", ".join(
                f"{str(r['parent_operator'])[:28]} {float(r['precision_lcb'])*100:.0f}%"
                for _, r in _ranked.iterrows() if str(r.get("precision_lcb", "")) != "")
        _unmeasurable_count = int((_lb["verdict"] == "UNMEASURABLE").sum())

    summary_kv = [
        ("Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Rows in file", f"{s.get('rows_total', 0):,}"),
        ("Field repairs applied (rows)", f"{s.get('rows_with_repairs', 0):,}  "
                                         f"({s.get('field_repairs_total', 0):,} field fixes)"),
        ("Providers enriched from NPPES", f"{s.get('providers_enriched', 0):,}  "
                                          f"({s.get('providers_found_in_nppes', 0):,} found)"),
        ("   of which Medicare-enrolled (PECOS)", f"{s.get('providers_medicare_enrolled', 0):,}"),
        ("   of which opted out of Medicare", f"{s.get('providers_opted_out', 0):,}"),
        ("Referring NPIs verified (Order & Referring)", f"{s.get('referring_npis_verified', 0):,}"),
        ("340B providers (eligibility signal / registered)",
         f"{s.get('providers_340b_signal', 0):,} signal / {s.get('providers_340b_registered', 0):,} registered"),
        ("Rows missing a billing NPI", f"{s.get('rows_blank_billing', 0):,}  ({s.get('pct_rows_blank', 0)}% of rows)"),
        ("Dollars on blank rows", f"{_money(s.get('dollars_blank', 0))}  ({s.get('pct_dollars_blank', 0)}% of $)"),
        ("Billers recovered (point + distributional)", f"{s.get('rows_recovered', 0):,}  ({s.get('pct_blanks_recovered', 0)}% of blanks)"),
        ("   of which point attribution (tier T1-T3)", f"{s.get('rows_point_attribution', 0):,}"),
        ("Routed to gross-up (Part D / NOC / no Part-B / no match)", f"{s.get('rows_grossup', 0):,}"),
        ("Honest expected top-1 accuracy (blanks-weighted)", _pct(bt.get("honest_top1"))),
        ("Honest expected top-3 accuracy", _pct(bt.get("honest_top3"))),
        ("   validated-only top-1 (attributable share)", _pct(bt.get("honest_top1_validated_only"))),
        ("HONEST top-1 on drugs we can validate (drug-stratified)", _pct(bt.get("drug_honest_top1"))),
        ("   blank $ share we CAN validate (have holdout data)", _pct(bt.get("blank_validated_drug_share"))),
        ("   blank $ share we CANNOT validate (no holdout data)", _pct(bt.get("blank_unvalidated_drug_share"))),
        ("Raw random-holdout top-1 (for reference only)", _pct(bt.get("holdout_top1"))),
        ("Top recoverable operators (precision_lcb)", _top_operators_str),
        ("Operators flagged UNMEASURABLE (pharmacy / no holdout)", _unmeasurable_count),
        ("CMS Physician dataset (resolved latest)", s.get("cms_physician_url", "")),
    ]
    if s.get("note"):
        summary_kv.insert(2, ("Note", s["note"]))
    summary_notes = [
        "",
        "HOW TO READ THIS FILE",
        "• Cleaned_Claims = your original rows, with fixable fields repaired in place and",
        "  Billing_NPI_Final / recovery columns appended. Repairs_Applied lists every fix per row.",
        "• Data_Quality = field coverage before vs. after repair. Repairs_Log = each fix and its method.",
        "• Provider_Directory = a deduped, fully-enriched NPPES record for every billing NPI (keyed for merge).",
        "• Drug_Reference = each HCPCS cross-referenced across CMS (benchmark + Part-D footprint).",
        "• Recovery_Detail = one row per blank, with the tier it resolved at, top-3 candidates, confidence.",
        "• Backtest = the honest accuracy number above, with its full derivation. Use the blanks-weighted figure.",
        "• See the Caveats tab before quoting any recovered name as fact.",
    ]

    # tidy the recovery detail for output
    rec = result.recovery.copy()
    rec_cols = ["orig_row", "hcpcs", "drug_name", "referring_npi", "pos", "zip3", "state",
                "blank_allowed", "blank_units", "benefit", "channel", "reason",
                "recovered_npi", "recovered_name", "recovered_operator",
                "recovered_top3", "tier", "attribution", "confidence", "support"]
    if rec.empty or not all(c in rec.columns for c in rec_cols):
        rec_out = pd.DataFrame(columns=rec_cols)
    else:
        rec_out = rec[rec_cols].copy()
    rec_out = rec_out.rename(columns={
        "orig_row": "row", "blank_allowed": "allowed_amt", "blank_units": "units",
        "recovered_npi": "recovered_billing_npi"})

    per_tier = bt.get("per_tier", pd.DataFrame())
    blanks_mix = bt.get("blanks_mix", pd.DataFrame())

    # before/after field coverage
    cb = result.coverage_before.rename(columns={"present": "present_before", "pct_present": "pct_before"})
    ca = result.coverage.rename(columns={"present": "present_after", "pct_present": "pct_after"})
    data_quality = cb.merge(ca, on="field", how="outer")
    data_quality["delta_rows"] = (data_quality["present_after"].fillna(0)
                                  - data_quality["present_before"].fillna(0)).astype(int)

    # v18: Recoverability Map + Acquisition Plan — operationalize the external-data
    # strategy into the deliverable so a consultant sees, per N/A field, what's
    # recoverable from free public data, what this tool already does, what paid
    # source closes the rest, and what no source can reach.
    recoverability = pd.DataFrame([
        {"field": "PAYER_NAME (branded carrier)",
         "from_free_public_data": "None",
         "what_this_tool_does": "Cannot recover branded carrier; documents the floor. Uses payer in cluster key where present.",
         "closes_the_rest_paid": "Komodo KPI / MapEnhance (segment + some branded); HealthVerity OPEN claims (named payer, token-linked).",
         "structurally_unrecoverable": "Branded carrier + true allowed amount on self-funded ERISA volume (Gobeille). No public source."},
        {"field": "PAYER_SEGMENT (Commercial / MA / Mgd Medicaid)",
         "from_free_public_data": "Moot — extract is already commercial-LOB only",
         "what_this_tool_does": "On a commercial-filtered extract, segment is known by construction (it is 'Commercial'), so 'recovering' it adds nothing on THIS book. Benefit channel (medical vs pharmacy) is still emitted per claim. The only real payer gap here is the branded carrier (row above).",
         "closes_the_rest_paid": "Komodo KPI gives segment on mixed-LOB data; irrelevant for a single-segment book.",
         "structurally_unrecoverable": "Branded carrier (see PAYER_NAME). Segment itself is not a gap on this extract."},
        {"field": "BILLING_PROVIDER_NPI",
         "from_free_public_data": "Moderate-High",
         "what_this_tool_does": "Recovers via your own-book prior + CMS pools (DMEPOS supplier, Part D, Physician); Luhn-validated; rolled to parent.",
         "closes_the_rest_paid": "NCPDP dataQ (pharmacy NPIs); IQVIA OneKey / Definitive for commercial-only billers.",
         "structurally_unrecoverable": "~$12M Part-D self-administered/pharmacy slice has no medical NPI to attribute to."},
        {"field": "BILLING NAME / ENTITY_TYPE / SPECIALTY",
         "from_free_public_data": "High (once NPI resolved)",
         "what_this_tool_does": "NPPES name, entity type, full taxonomy set; CMS facility affiliation + hospital ownership.",
         "closes_the_rest_paid": "Commercial reference license only for residual commercial-only providers.",
         "structurally_unrecoverable": "—"},
        {"field": "REFERRING_PROVIDER_NPI",
         "from_free_public_data": "Low for FILL — Open Payments adds a KOL flag, not a fill",
         "what_this_tool_does": "_Referring_PaidByDrugMaker flags rows whose EXISTING referring NPI is paid by the billed drug's maker (a KOL / steering signal). It does NOT populate a MISSING referring NPI — a blank row has no NPI to check — so missing referrers stay N/A.",
         "closes_the_rest_paid": "IQVIA OneKey / Definitive / Clarify referral-pattern data.",
         "structurally_unrecoverable": "No public source records the referrer on a private claim, so a blank referring NPI is largely unrecoverable from public data."},
        {"field": "REFERRING NAME / SPECIALTY / AFFILIATION",
         "from_free_public_data": "High (once NPI resolved)",
         "what_this_tool_does": "NPPES specialty + CMS facility affiliation once the referring NPI is known.",
         "closes_the_rest_paid": "Commercial affiliation license for group/IDN hierarchy.",
         "structurally_unrecoverable": "—"},
        {"field": "AFFILIATION / ENTITY ROLL-UP (economic owner)",
         "from_free_public_data": "Mixed",
         "what_this_tool_does": "PECOS PAC ID + exact shared-address/name/DBA clustering, now with Splink probabilistic fuzzy linkage layered on top (catches typo / suffix / spacing variants in names and addresses that exact matching misses) -> _Billing_Parent_Group. Probabilistic merges are flagged match_basis='fuzzy' for audit, and can only ADD to the deterministic clusters, never split them.",
         "closes_the_rest_paid": "PitchBook / D&B / Secretary-of-State for PE-platform parentage and unregistered DBAs.",
         "structurally_unrecoverable": "Physician-group equity ownership / PE parentage never appears in any CMS file."},
    ], columns=["field", "from_free_public_data", "what_this_tool_does",
                "closes_the_rest_paid", "structurally_unrecoverable"])

    acquisition_plan = pd.DataFrame([
        {"stage": "1 — Free public (this tool)", "cost": "~$0 / days",
         "action": "PECOS PAC ID + reassignment, DMEPOS supplier, Part D Prescriber, Open Payments, NPPES, MA/Medicaid MCO enrollment.",
         "trigger_or_benchmark": "Already integrated. Advance only if a material dollar-weighted residual remains after this."},
        {"stage": "2 — Komodo-native payer", "cost": "Incremental to Komodo",
         "action": "License Komodo KPI / MapEnhance to attach primary/secondary payer + segment to the same de-identified lives.",
         "trigger_or_benchmark": "Measure share of the PAYER_NAME rows that get a confident branded carrier vs only a segment; fall back to segment + market-share priors."},
        {"stage": "3 — One commercial reference license", "cost": "Mid-6 to low-7 fig / yr",
         "action": "Pick ONE: IQVIA OneKey (corporate-parent/IDN hierarchy) OR Definitive (physician-group ownership + payer mix).",
         "trigger_or_benchmark": "Adopt only if Stage 1 leaves a material residual in NAME / ENTITY_TYPE / AFFILIATION."},
        {"stage": "4 — PE-ownership roll-up", "cost": "Subscription",
         "action": "PitchBook (and/or D&B + TX Secretary of State) to map the platform parent behind sibling NPIs/DBAs at shared addresses.",
         "trigger_or_benchmark": "Trigger when affiliation fields need true parent-operator attribution for the thesis (volume concentration under one PE platform)."},
        {"stage": "5 — External named-payer open claims (optional)", "cost": "Multi-6 fig / 3-6 mo",
         "action": "HealthVerity OPEN claims, token-linked via Datavant, with Expert Determination re-certification.",
         "trigger_or_benchmark": "Only if KPI branded-carrier fill is insufficient and the thesis demands per-claim carrier. Confirm license permits payer-name enrichment."},
    ], columns=["stage", "cost", "action", "trigger_or_benchmark"])

    # v21: legend for the complete-labeling tokens. Every DATA cell is non-blank;
    # the unrecoverable ones carry one of these self-documenting tokens. This maps
    # each to plain English, its confidence band, and whether paid data could
    # recover it — so "100% populated" never means "100% observed".
    token_legend = pd.DataFrame([
        {"token": "(a real value)", "appears_in": "any data field",
         "means": "Observed in your source, or recovered/derived and trustworthy.",
         "confidence": "Observed = ground truth. Recovered billing NPI carries its measured hit-rate in _NPI_Confidence. NPPES-derived name/specialty/affiliation = high-confidence deterministic lookup.",
         "recoverable_with_paid_data": "n/a — already populated"},
        {"token": "UNRECOVERABLE_PAYER_ERISA", "appears_in": "PAYER_NAME",
         "means": "No public dataset records the branded carrier on a closed claim; self-funded ERISA volume is outside every state APCD.",
         "confidence": "Structural floor — not a tooling gap.",
         "recoverable_with_paid_data": "Partly — Komodo KPI / MapEnhance recovers SOME branded carriers, never all; self-funded stays out of reach."},
        {"token": "PHARMACY_BENEFIT_NO_MEDICAL_NPI", "appears_in": "BILLING_PROVIDER_NPI",
         "means": "Part-D self-administered drug billed by a pharmacy on NCPDP rails — there is no medical billing NPI to attribute.",
         "confidence": "Structural — a medical NPI does not exist for this claim.",
         "recoverable_with_paid_data": "NCPDP dataQ (paid) can name the pharmacy, but it is not a medical NPI."},
        {"token": "BESTGUESS_BELOW_BAR", "appears_in": "BILLING_PROVIDER_NPI",
         "means": "Ranked candidate billers exist in _NPI_BestGuess, but none cleared the measured accuracy bar, so none is written as fact.",
         "confidence": "Low — use _NPI_BestGuess for sizing/sensitivity, never for naming.",
         "recoverable_with_paid_data": "A commercial reference license (OneKey / Definitive) may resolve some."},
        {"token": "NO_CANDIDATE_IN_PUBLIC_DATA", "appears_in": "BILLING_PROVIDER_NPI",
         "means": "No plausible biller for this drug + geography was found in public data.",
         "confidence": "Unknown — no signal.",
         "recoverable_with_paid_data": "Possibly, with commercial claims/reference data."},
        {"token": "UNCLASSIFIED_HCPCS_CODE", "appears_in": "BILLING_PROVIDER_NPI",
         "means": "Unclassified J-code (e.g. J3490/J3590) — drug and biller are indeterminate from the code.",
         "confidence": "Unknown — the code itself is non-specific.",
         "recoverable_with_paid_data": "No — requires the underlying NDC, not in this extract."},
        {"token": "NO_TRUSTED_BILLING_NPI", "appears_in": "BILLING name / ENTITY_TYPE / affiliation / specialty",
         "means": "These identity fields are derived from the billing NPI; the NPI here was not trusted (a guess or absent), so nothing is looked up off it.",
         "confidence": "Correct by design — refuses to derive identity from a guess.",
         "recoverable_with_paid_data": "Only once the billing NPI itself is resolved."},
        {"token": "NPI_NOT_IN_NPPES", "appears_in": "billing / referring identity fields",
         "means": "The row's NPI is a real 10-digit number, but NPPES had no record (deactivated, foreign, or never enumerated).",
         "confidence": "High that the gap is real, not a lookup error.",
         "recoverable_with_paid_data": "A commercial provider-reference file may carry it."},
        {"token": "NO_REFERRING_NPI_ON_CLAIM", "appears_in": "REFERRING NPI / name / affiliation / specialty",
         "means": "The source claim carried no referring NPI, so there is nothing to look up.",
         "confidence": "Structural — no public source records the referrer on a private claim.",
         "recoverable_with_paid_data": "Largely no for the NPI; identity follows only if the NPI is recovered."},
        {"token": "NOT_POINT_ATTRIBUTABLE", "appears_in": "BILLING_PROVIDER_NPI (verified output only)",
         "means": "Verified output writes only observed values; a blank billing NPI is left unrecovered here by design (see the statistical output for an estimate).",
         "confidence": "By design — verified output never imputes.",
         "recoverable_with_paid_data": "See the statistical output / commercial data."},
        {"token": "DATA_MISSING_IN_SOURCE", "appears_in": "any source column the tool does not fill",
         "means": "A blank in a source column outside the recovery scope (safety net so no cell is ever empty).",
         "confidence": "Reflects a gap in the original extract.",
         "recoverable_with_paid_data": "Depends on the field."},
        {"token": "\u2014 (em-dash)", "appears_in": "audit columns (_-prefixed) only",
         "means": "Not applicable to this row (e.g. _NPI_BestGuess is em-dash when the billing NPI is trusted, so no best-guess is needed).",
         "confidence": "n/a — metadata, not data.",
         "recoverable_with_paid_data": "n/a"},
    ], columns=["token", "appears_in", "means", "confidence", "recoverable_with_paid_data"])

    # v22: connector audit -> a sheet proving every public API is live and used.
    cs_rows = getattr(result, "connector_status", []) or []
    if cs_rows:
        connector_status = pd.DataFrame([
            {"connector": r.get("source", ""),
             "status": "LIVE" if r.get("ok") else "FAIL",
             "detail": r.get("detail", ""),
             "seconds": r.get("seconds", "")} for r in cs_rows])
    else:
        connector_status = pd.DataFrame([{"connector": "(audit skipped)", "status": "",
                                          "detail": "Run without --no-audit to probe every API.",
                                          "seconds": ""}])
    _n_live = int((connector_status["status"] == "LIVE").sum())
    _n_tot = int((connector_status["status"].isin(["LIVE", "FAIL"])).sum())

    # v23: the genuine fill report — census across the three certainty tiers,
    # a truthful methodology note, and a landscape pivot sanity-check.
    from . import tiers as _tiers
    _closed = getattr(result, "filled_verified", pd.DataFrame())
    _recov = getattr(result, "filled", pd.DataFrame())
    _stat = getattr(result, "filled_statistical_full", pd.DataFrame())
    _map = getattr(result, "mapping", {}) or {}
    try:
        if not _recov.empty and not _stat.empty and not _closed.empty:
            cell_census = _tiers.cell_census(_closed, _recov, _stat, _map)
        else:
            cell_census = pd.DataFrame([{"note": "census needs all three tiers (run with default outputs)"}])
    except Exception as e:
        cell_census = pd.DataFrame([{"note": f"census unavailable: {type(e).__name__}"}])
    method_doc = _tiers.statistical_method()
    try:
        if not _recov.empty and not _stat.empty:
            pivot_tbl = _tiers.pivot_landscape(_recov, _stat, _map)
            _pivot_verdict = pivot_tbl.attrs.get("verdict", "")
        else:
            pivot_tbl = pd.DataFrame([{"note": "pivot needs the recovered + statistical tiers"}])
            _pivot_verdict = ""
    except Exception as e:
        pivot_tbl = pd.DataFrame([{"note": f"pivot unavailable: {type(e).__name__}"}])
        _pivot_verdict = ""

    sheets = [
        ("Cleaned_Claims", result.cleaned, "Cleaned Claims (repaired fields + recovered NPIs)",
         "Fixable fields are repaired in place; Billing_NPI_Final fills the gaps; Repairs_Applied logs each fix."),
        ("Data_Quality", data_quality, "Step 0 — Field coverage before vs. after repair",
         "What share of rows carried each field before cleaning, and after deriving/repairing it."),
        ("Repairs_Log", result.repairs_log, "Step 0.5 — Field repairs applied",
         "Every automated fix, the field it touched, how many rows, and the method used."),
        ("Column_Mapping", result.map_report, "Detected column mapping",
         "How your file's headers were matched onto the canonical fields. Override via the CLI if any are wrong."),
        ("Blank_Profile", result.blank_profile, "Step 0 — Blank rows by channel and benefit",
         "Confirms the gap is channel-driven (specialty-pharmacy / DME / Part-D), not random."),
        ("Blank_NPI_Decomposition", result.blank_decomposition,
         "WHY each billing NPI is blank — three different problems, not one",
         "A blank billing NPI is not one problem with one filter. It is: (1) RECOVERABLE from referral/identifier "
         "matches — filled here, point matches in the NPI column and distributional best-guesses in a separate "
         "labeled column; (2) NEVER POPULATED BY DESIGN — pharmacy-benefit / non-medical drugs have no medical "
         "NPI, so the cell stays empty on purpose (a closed-claims filter does NOT populate these — they are blank "
         "in closed claims too); (3) GENUINELY MISSING or still OPEN — left empty, closes only with the data room. "
         "Restricting to closed claims discards bucket 1 (the recoverable open claims) and so LOWERS coverage on "
         "the target exactly when you are worried it is undercounted. The fix: use closed claims as the validation "
         "set (Backtest sheets), recover bucket 1 on the open claims, keep buckets 2-3 empty."),
        ("CMS_Candidate_Pool", result.pool_table, "Step 3 — CMS candidate billers (live, by HCPCS x state)",
         "Real Medicare providers/suppliers who bill each code there. Source of the plausible-biller universe."),
        ("Recovery_Detail", rec_out, "Step 4-5 — Per-blank recovery with confidence tier",
         "One row per blank. Tier = how it resolved; attribution = point vs distributional."),
        ("Provider_Directory", result.provider_directory,
         "Enrichment — deduped NPPES profile for every billing NPI",
         "One row per NPI (original-valid + recovered): entity type, specialty, license, address, phone."),
        ("Drug_Reference", result.drug_reference,
         "Enrichment — each HCPCS cross-referenced across CMS",
         "Benefit track + channel, national Medicare benchmark (Geography), and Part-D footprint."),
        ("Coverage_340B", result.coverage_340b,
         "340B coverage — per-NPI eligibility signal + registered status",
         "Taxonomy-based 340B eligibility signal (always on); registered status if an OPAIS file was supplied."),
        ("Entity_Rollup", result.entity_rollup, "Step 5 — Recovered NPIs rolled to parent operator",
         "Sibling NPIs grouped via shared authorized official, mailing/location address, OR matching "
         "organization name vs NPPES other/former/DBA name (catches rebranded & acquired practices). "
         "match_basis shows which signal linked each cluster."),
        ("Entity_Crosswalk", result.entity_crosswalk, "Step 5 — legacy NPI -> parent operator crosswalk",
         "Join this onto a claims panel before measuring share so acquired/rebranded sibling NPIs collapse "
         "onto their parent. multi_npi_parent flags the genuine roll-ups; feed it to the HHI confirm/deny test."),
        ("Backtest_PerTier", per_tier, "Step 6 — Closed claims used as the VALIDATION set (accuracy by tier)",
         "This is the right use of closed/complete claims: not as the working dataset (which throws away the "
         "recoverable open claims) but as ground truth. We hide the billing NPI on rows that HAVE one, recover it, "
         "and measure dollar-weighted top-1/top-3 accuracy by tier. That measured accuracy is what the open-claim "
         "recovery inherits — completeness without discarding coverage."),
        ("Backtest_BlanksMix", blanks_mix, "Step 6 — Tier mix of the real blanks",
         "The blanks' tier shares; blended with per-tier accuracy to get the honest expected number."),
        ("Backtest_UnvalidatedDrugs", bt.get("unvalidated_drugs", pd.DataFrame()),
         "Step 6 — Blanks in drugs with too little holdout to validate",
         "These drugs are mostly blank in your file (e.g. pharmacy-benefit, military/VA), so the backtest "
         "has no held-out rows to measure recovery accuracy on them. Recovery here is UNVALIDATED — treat "
         "those recovered NPIs as leads, and close the gap with the CIM / data room, not this tool."),
        ("Backtest_ByOperator", getattr(result, "backtest_by_operator", pd.DataFrame()),
         "Step 6 — Per-operator recovery reliability (which named platforms recover well)",
         "Same held-out rows, aggregated by the TRUE biller's parent operator. recall_* = how findable an "
         "operator is; precision_* = when the engine says this operator, how often it's right (the diligence "
         "number). Ranked by precision_lcb (Wilson 95% lower bound) so small samples sink; operators with "
         "<30 held-out rows are INSUFFICIENT_HOLDOUT and pharmacy/no-holdout operators are UNMEASURABLE — "
         "never a spuriously high score. blanks_attributed / blanks_dollars connect that measured reliability "
         "to real recovered exposure. Dollar-weighted figures are the headline; count-weighted is the companion."),
        *_v43_sheets(result),
        ("GrossUp_ByDrug", result.grossup_table, "Step 7 — Coverage gross-up (non-attributable volume)",
         "Magnitude by drug for the slice we deliberately do not name a biller for."),
        ("Recoverability_Map", recoverability,
         "External-data strategy — what each N/A field is recoverable from",
         "Per field: recoverability from FREE public data, what this tool already does, the paid source that "
         "closes the rest, and what no source can reach. The branded payer on closed/self-funded claims is the "
         "hard floor (ERISA / Gobeille) — segment is recoverable, the carrier is not."),
        ("Acquisition_Plan", acquisition_plan,
         "External-data strategy — sequenced data acquisition",
         "Buy in order, cheapest-first: free CMS (this tool) -> Komodo KPI for payer -> one commercial reference "
         "license -> PE-ownership data -> optional named-payer open claims. Each stage lists the benchmark that "
         "justifies moving to the next, so spend is earned, not assumed."),
        ("Token_Legend", token_legend,
         "v21 — what every cell means when it isn't an observed value",
         "The Filled_Data sheets have ZERO blank data cells. Cells that have no observable true value carry a "
         "specific token from this legend (not a fabricated value), with its confidence band and whether paid "
         "data could recover it. '100% populated' here means every cell is in a known, labeled state — never that "
         "every cell is observed."),
        ("Connector_Status", connector_status,
         f"v22 — live audit of every data connector ({_n_live}/{_n_tot} reachable this run)",
         "Each public API the tool relies on is probed with a small bounded request at the end of the run. "
         "LIVE = reachable and returning data; FAIL = unreachable or errored (the pipeline degrades around it). "
         "This is the proof the CMS / NPPES / DMEPOS / Open Payments / 340B / RxNorm / Coverage connectors are "
         "actually used, not silently skipped."),
        ("Cell_Census", cell_census,
         "v23 — how many cells were filled, and at what certainty",
         "Every fillable cell, classified by the highest-certainty tier that populates it: certain (in "
         "Closed_Claims) / recovered via API (added in Recovered_Claims) / statistically estimated (added in "
         "Statistically_Filled, flagged for review) / still unfillable. The literal answer to 'how much got "
         "filled vs estimated vs left missing', per field and overall."),
        ("Statistical_Method", method_doc,
         "v23 — how the statistical estimation actually works",
         "What is estimated (only the residual billing NPI), the referral-anchored estimator, the evidence "
         "tiers, the k-fold MEASURED accuracy behind every confidence number, and why the estimates live in a "
         "separate file. Read before relying on the Statistically_Filled output."),
        ("Pivot_Landscape", pivot_tbl,
         "v23 — do the estimates distort the market-share picture?",
         (_pivot_verdict + "  |  Table: operator $ and share, Recovered tier vs Statistically_Filled tier.")
         if _pivot_verdict else
         "Operator $ and share compared between the Recovered tier and the Statistically_Filled tier."),
    ]

    # v25: readout analytics. Each is a deterministic cut over the cleaned +
    # recovered frame. Descriptions carry the honesty framing (exposure vs
    # measured; reference-sourced columns).
    _an_desc = {
        "Referral_Concentration": ("v25 — referral dependency and leading-indicator risk",
            "Per referring provider: downstream billing concentration, dominant operator, referral HHI, and a "
            "single-source flag. This is CURRENT exposure and an early-warning signal (is a key referrer already "
            "routing share to a competitor), NOT a measured future decline."),
        "Referral_Concentration_IVIG": ("v25 — same cut, restricted to IVIG / immune globulin",
            "The IVIG-specific referral concentration view. Use to read a specific referrer's IVIG routing "
            "(e.g. a large health-system referrer) without other therapies diluting the signal."),
        "Submarket_Landscape": ("v25 — single-year market composition by submarket",
            "Volume, units and allowed by submarket x drug category x site of care, latest single year by "
            "default (coverage barriers by J-code make year-over-year unreliable). Observed Komodo subset only; "
            "VRDC gross-up is a separate labeled scaffold and is NOT applied here."),
        "Submarket_Saturation": ("v25 — relative saturation and build-vs-acquire read",
            "Operators per submarket, operator HHI, and a RELATIVE saturation percentile (no absolute per-capita "
            "benchmark — the 2-per-100k idea is too market-dependent). Read column suggests build vs acquire."),
        "Referral_Target_Map": ("v25 — who to target by submarket",
            "Ranked targetable referrers per submarket with openness (open / partially open / closed, from "
            "current operator concentration) and an org-type heuristic (health system / PPM / independent). "
            "Org type is a name/specialty heuristic, not a verified ownership field."),
        "Benefit_Shift_Exposure": ("v25 — white/brown-bag forward risk by drug",
            "Per drug: measured book share, plus reference-sourced flags (suffix __ref) for whether an approved "
            "subcutaneous/self-admin version exists and the resulting forward white-bag risk. The total row is "
            "the share of book at forward risk. Reference columns are curated, not derived from claims."),
        "Formulary_Scorecard": ("v25 — formulary mix read",
            "Per top drug: MEASURED book share and demand trend, plus curated reference flags (suffix __ref) for "
            "buy-and-bill economics, IRA exposure, biosimilar risk, benefit-shift, therapeutic area, and the "
            "team read. Reference columns are not provable from claims; update as policy moves."),
        "SiteOfCare_Trend": ("v26 — office-to-home shift (measured)",
            "Site-of-care mix (home / office / outpatient) by year from the panel. The home_share_change_pp "
            "column flags the net shift over the window. Measured from the panel."),
        "Payer_Landscape": ("v26 — payer concentration by submarket (network-adequacy lens)",
            "Per submarket: number of payers, top payer and its share, payer HHI, and a concentration flag. "
            "Use to spot single-payer dependence and network-adequacy / leakage openings. Measured from the panel."),
        "Regulatory_Scorecard": ("v26 — single-state regulatory risk vs the formulary",
            "Per drug: forward white-bag risk tied to whether an approved SC version exists, against the Texas "
            "regulatory shield (white/brown-bag bar + AG opinion) and the risk if that opinion is overturned. "
            "Tailwinds row at the bottom. Reference-sourced (suffix __ref)."),
        "Market_Size_Estimate": ("v26 — VRDC-free market size (bottom-up payer-segment build)",
            "Estimated all-payer market by drug for the state, built by SUMMING measured segments (commercial "
            "from the panel, Medicare FFS from CMS actual utilization) and grossing up only MA and Medicaid with "
            "stated multipliers. measured_share_of_estimate_pct says how much is measured vs grossed. "
            "panel_capture_vs_medicare_pct is a per-drug completeness cross-check. An ESTIMATE, not a census; "
            "this is the VRDC stand-in for this project."),
        "Market_Size_Method": ("v26 — market-sizing method and assumptions",
            "Every multiplier and assumption behind Market_Size_Estimate, and what is measured vs grossed up. "
            "Read alongside the estimate; the multipliers are adjustable parameters, not magic numbers."),
        "Growth_Pockets": ("v27 — where to build (growing + under-served submarkets)",
            "Submarkets ranked by a build-opportunity score blending growth, fragmentation/whitespace, and size. "
            "The intra-state geographic-prioritization view for Phase 2. Measured from the panel."),
        "Referral_Leakage": ("v27 — BD target list (addressable leaked volume)",
            "Per referring provider: total referred dollars, the leading operator and its share, and the "
            "addressable volume NOT held by the leader, ranked. Use to chase PPMs and health systems with "
            "leakage. Measured from the panel."),
        "Therapeutic_Adjacency": ("v27 — therapeutic-area mix and whitespace",
            "Book composition by therapeutic area, plus adjacency rows the book does not yet serve (oncology, "
            "cardiology) flagged as whitespace with context. Areas are reference-mapped from the drug."),
        "Market_Attractiveness": ("v27 — cross-state attractiveness (Phase 2 geographic screen)",
            "States scored 0-100 on a weighted blend of size, growth, and headroom from CMS Market Saturation "
            "(Part B Drugs). A Medicare fee-for-service proxy and first-pass geographic screen, NOT a commercial "
            "market size. The national-model starting point."),
        "Market_Attractiveness_Method": ("v27 — attractiveness method and weights",
            "The weights and fields behind the attractiveness score, and the explicit caveat that it is a "
            "Medicare FFS proxy. Weights are adjustable."),
        "Pharmacy_Benefit_Estimate": ("v28 — pharmacy-benefit section (CMS Part D analog gross-up)",
            "Estimates the pharmacy-benefit channel the medical panel cannot see. Per dual-channel drug, the real "
            "Medicare Part D / Part B ratio is applied to the commercial medical panel. measured columns carry "
            "__measured, the CMS ratio and Part B/D carry __ref. A LOWER BOUND: pharmacy-only drugs (Stelara, pure "
            "SCIG) are flagged to size from the pharmacy pull or data room, not grossed up. Immune globulin is the "
            "headline: ~1.36x, so the visible IVIG is only ~42% of the immune-globulin market."),
        "Pharmacy_Benefit_Method": ("v28 — pharmacy gross-up method and assumptions",
            "The Part D analog approach, the commercial-adjustment knob, and the explicit lower-bound caveat. "
            "Ratios are measured from CMS; the channel map and commercial assumption are reference-sourced."),
        "Analytics_Note": ("v25 — analytics note", "Status note from the analytics stage."),
        "SiteOfCare_Reclassification": ("v29 — site-of-care reclassification (office->AIC, AIS broken out)",
            "Corrects the as-billed POS view, which collapses ambulatory infusion centers (AIC) and "
            "home-infusion pharmacy (AIS) into 'office' and mis-stamps true AICs as office. Resolves each "
            "drug line by, strongest first: POS home; a home-infusion code (Sxxxx/99601/G00xx) anywhere in "
            "the claim (overrides a mis-stamped office POS); infusion-clinic taxonomy -> AIC; "
            "infusion/home-pharmacy or DME taxonomy -> AIS; home-health taxonomy -> Home; then a hand-rolled "
            "logistic office->AIC propensity for the residual office bucket. Shows as-billed vs reclassified "
            "share and the pp move, plus a probability-weighted expected office->AIC reattribution. NOTE row "
            "states when admin codes are absent (expand the extract) or the propensity is untrained."),
        "SiteOfCare_ProofPoint": ("v29 — site reclassification proof point (Ocrevus-style anchor)",
            "For an AIC-heavy anchor drug (default ocrelizumab/Ocrevus J2350), the as-billed office vs AIC "
            "split and how much reclassification moves to AIC — the 'the volume we see as office is paltry "
            "vs what AICs actually do' check from the readout."),
        "Operator_Case_Mix": ("v29 — case mix: premium mix vs rate (confirm/deny)",
            "Allowed-per-unit by operator vs market, decomposed into a MIX effect (the operator's drug "
            "composition priced at market unit rates) and a RATE effect (actual price vs that mix-expected "
            "price). Answers 'does the platform deliver higher-value drugs than the street?'. Because major "
            "commercial payers pay near-identical unit rates per drug, a high allowed-per-unit is almost "
            "always richer mix, not better contracts. Reversals excluded; allowed-per-unit is "
            "price-interested volume, not a contracted rate."),
        "Share_Shift_Markets": ("v29 — share-shift markets (top/bottom by share-point change)",
            "For the focal operator (largest by allowed), its allowed-weighted share by submarket in the "
            "first vs last dated year and the change in share POINTS (pp), ranked into top gainers and top "
            "decliners. Basis = patient location at the submarket grain. The Richmond +pp share-shift cut."),
        "Growth_Reconciliation": ("v29 — growth reconciliation (YoY vs CAGR vs cumulative)",
            "Membership-adjustable allowed by year with year-over-year growth, plus CAGR (annualised) and "
            "cumulative (first->last) computed and LABELLED separately so a cumulative figure is never "
            "mis-read as an annual decline. NOTE row carries the bridge: medical-only vs +pharmacy, and "
            "whether the read is covered-lives-adjusted."),
        "Channel_Reconciliation": ("v29 — medical vs pharmacy channel reconciliation",
            "Per operator and in total: medical-only allowed vs medical+pharmacy allowed, the pharmacy share "
            "of combined, and the year-over-year growth on each basis with the swing in pp. The medical-only "
            "decline that flips flat/positive once the pharmacy channel is added back is the headline. "
            "Operator attribution of the pharmacy feed uses the medical entity rollup; pharmacy NPIs not seen "
            "in the medical book attribute to their own NPI unless added to the crosswalk."),
        "Pharmacy_Feed_Reconciliation": ("v29 — pharmacy feed union + dedup report",
            "Row counts for the medical and pharmacy feeds, rows dropped as same-claim-id duplicates "
            "(medical kept), and rows flagged as possible exact cross-source duplicates (not dropped, for "
            "analyst review). The audit trail for the dual-feed union."),
        "DualChannel_Drug_Reconciliation": ("v30 — per-molecule medical vs pharmacy stitch + double-count guard",
            "For each drug class (consistent across feeds via the NDC/name -> class bridge): medical allowed "
            "vs pharmacy allowed vs combined, the pharmacy share, and a double-count guard. A dual-benefit "
            "molecule can bill on both the medical (J-code) and pharmacy (NDC) feed; the tab counts patients "
            "present on BOTH benefits in the same year and estimates the overlapping dollars (the smaller "
            "side per patient-year), flags classes where that overlap is material, and shows combined net of "
            "overlap. Net the overlap out before quoting combined."),
        "Rx_Bridge_Coverage": ("v30 — pharmacy NDC/name -> drug-class resolution coverage",
            "How the pharmacy rows resolved onto the medical drug taxonomy — by NDC (live RxNorm), then "
            "free-text name, then the brand/ingredient seed — with rows and dollars in each bucket. "
            "UNMAPPED_RX rows still count in dollar totals but are excluded from per-drug reconciliation; a "
            "large unmapped share is the signal to extend the crosswalk or supply a fuller NDC map."),
        "Pharmacy_Provider_Attribution": ("v30 — pharmacy dispensing-NPI enrichment + operator rollup",
            "The pharmacy dispensing NPIs enriched via NPPES (their taxonomy drives the AIS site "
            "classification) and rolled up to an operator by organisation name: exact or contains match to a "
            "known medical operator, else their own org name. Reports how many NPIs matched a medical "
            "operator and how much pharmacy allowed rolls up to a known operator vs stands alone. "
            "Deterministic name match only — no fuzzy merges, so it cannot fabricate a link."),
    }
    _an = getattr(result, "analytics", {}) or {}
    for _name, _df in _an.items():
        _hdr, _desc = _an_desc.get(_name, (f"v25 — {_name}", "Readout analytics cut."))
        sheets.append((_name, _df, _hdr, _desc))

    caveats = [
        "1. CANDIDATE UNIVERSE, NOT COMMERCIAL VOLUME. All CMS provider files are 100% Medicare",
        "   fee-for-service. The set of providers who bill a drug transfers to commercial claims; the",
        "   volumes and dollars do NOT. Candidate pools rank who plausibly bills a code, not commercial share.",
        "",
        "2. TIERS ARE THE CONFIDENCE SIGNAL. Tier T1-T3 (in-panel referral patterns) are point attribution.",
        "   T4-T5 are weaker in-panel signals; T6 is the CMS pool with NO in-panel validation. Never present",
        "   a T6 (or distributional) recovery as a confirmed biller — it is a pooled/synthetic-operator guess.",
        "",
        "3. THE NUMBER TO REPORT is the blanks-weighted honest expected accuracy on the README/Backtest tabs,",
        "   not the random-holdout figure. The validated-only figure normalises over the attributable share.",
        "",
        "4. SELF-ADMINISTERED (PART D) DRUGS ARE NOT IMPUTED. SAD-list codes (e.g. subcutaneous Stelara",
        "   J3357) and unresolved NOC codes (J3490/J3590/C9399) are routed to the gross-up and sized by drug,",
        "   because the biller is a pharmacy under the drug benefit, not a medical biller. Dual-channel",
        "   molecules are routed by HCPCS, not name: vedolizumab IV J3380 / ustekinumab IV J3358 / golimumab",
        "   IV J1602 stay Part B (attributable) even though the same brand's SC form is a SAD gross-up.",
        "",
        "5. MAIL-ORDER GEOGRAPHY IS WEAK. National specialty-pharmacy / DME suppliers report one central",
        "   address, so geographic matching degrades for them; recovery leans on HCPCS mix + referrer.",
        "",
        "6. STRUCTURALLY ABSENT VOLUME. VA / TRICARE-direct and commercial pharmacy-benefit (NCPDP) volume",
        "   are not in any of these medical files and are not recoverable from public data.",
        "",
        "7. SEEDS ARE STARTERS. The SAD and home/DME HCPCS lists shipped with the tool are seeded from the",
        "   live CMS Coverage DB plus curated entries; extend reference/*.csv for full production coverage.",
        "",
        "8. FIELD REPAIRS ARE DOCUMENTED PROXIES. Missing state is derived from the ZIP3->state crosswalk;",
        "   missing drug names from the HCPCS description; invalid NPIs (failing the 10-digit Luhn check) are",
        "   cleared and re-recovered. Every change is in Repairs_Log and the Repairs_Applied column — review",
        "   before treating a derived field as ground truth.",
        "",
        "9. INFERRED FILLS ARE NOT LOOKUPS. Rows tagged inferred-cluster or inferred-continuity in",
        "   _NPI_Confidence were filled by borrowing a biller from sibling rows (a dominant operator in the",
        "   referrer/drug/area cluster, or the same patient+drug+site within 90 days), not by looking up this",
        "   row. They sit in the billing column so the blank rate can fall to a few percent, but they are a",
        "   high-probability inference, not a verified biller. Cluster inference runs on entity-ROLLED operators",
        "   and fails closed across a change-of-ownership; continuity never bridges a >90-day gap or a site change.",
        "",
        "10. PAYER NAME IS UNRECOVERABLE. No public CMS dataset (NPPES, PECOS, NCPDP, any PUF) carries the",
        "    commercial payer on a claim. Where a payer cell is blank it stays N/A — no inference reaches it,",
        "    because the key does not exist in public data. Payer present in the source IS used as a cluster key.",
        "",
        "11. CONFIDENCE IS MEASURED, NOT ASSERTED. The masking backtest holds out a fraction of rows that",
        "    have a biller and re-recovers them, measuring each tier's dollar-weighted top-1 accuracy. If a",
        "    tier the imputer treats as high-confidence underperforms on that holdout (and has enough holdout",
        "    support), its recoveries are DEMOTED to best-guess — the tier ranking is a prior, the holdout is",
        "    evidence, and the evidence wins. Each _NPI_Confidence cell shows the measured hit-rate for its tier.",
        "",
        "12. DEACTIVATED BILLERS ARE FLAGGED. A final billing NPI that the CMS Deactivated-NPI file shows was",
        "    deactivated AS OF the claim's service date is flagged in _NPI_Deactivated (and noted in",
        "    _NPI_Confidence). For a recovered biller that means the recovery is almost certainly wrong; for an",
        "    original it's a data anomaly. It is flagged, not dropped — CMS deactivations can be reversed.",
        "",
        "13. NDC FILLS ARE CONSERVATIVE. When the file carries an NDC, a blank drug name is filled from the FDA",
        "    NDC Directory and an NDC that disagrees with a present drug name is flagged (ndc_drug_mismatch) —",
        "    a present drug name is never overwritten. The filled generic name also lets the NOC router classify",
        "    an otherwise-unclassifiable J3490/J3590 line.",
    ]

    return excelio.build_workbook(out_path, sheets, summary_kv, summary_notes, caveats)


def _pct(x):
    try:
        if x is None or (isinstance(x, float) and (x != x)):
            return "n/a (insufficient in-panel training)"
        return f"{100 * float(x):.1f}%"
    except Exception:
        return "n/a"


# ---- illegal-character / type sanitizer so a hostile source value can never
# ---- break the workbook on save -------------------------------------------
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _safe_cell(v):
    """Coerce any value into something openpyxl can write safely: strips the
    control characters Excel rejects, truncates over-long text, converts numpy
    scalars / NaT / NA to native types, leaving formula handling to the caller."""
    if v is None or v is pd.NA or v is pd.NaT:
        return None
    try:
        if isinstance(v, float) and np.isnan(v):
            return None
    except Exception:
        pass
    if hasattr(v, "item") and not isinstance(v, (str, bytes, bytearray)):
        try:
            v = v.item()
        except Exception:
            pass
    if isinstance(v, pd.Timestamp):
        try:
            return v.to_pydatetime()
        except Exception:
            v = str(v)
    if isinstance(v, (bytes, bytearray)):
        v = v.decode("utf-8", "replace")
    if isinstance(v, str):
        s = _ILLEGAL_XLSX.sub("", v)
        if len(s) > 32767:
            s = s[:32760] + "…"
        return s
    return v


def _is_formula_like(s):
    return isinstance(s, str) and len(s) > 0 and (
        s[0] in "=+@" or (s[0] == "-" and len(s) > 1 and not s[1].isdigit()))


def _wcell(ws, value, bold=False):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font
    cv = _safe_cell(value)
    cell = WriteOnlyCell(ws, value=cv)
    if _is_formula_like(cv):
        cell.data_type = "s"   # force literal text — never an executable formula
    if bold:
        cell.font = Font(bold=True)
    return cell


def write_filled(result, out_path, which="statistical"):
    """Write a 'filled' deliverable as its own workbook: the original columns
    (deduplicated) with missing cells filled, "N/A" where unfillable, and audit
    columns. which="statistical" (full recovery) or "verified" (direct lookups
    only). A header on Fill_Summary states exactly how the values were formed."""
    # v22: prefer the formatted xlsxwriter build (frozen header, autofilter,
    # fitted widths, colour-grouped by trust, leading Summary). Falls back to the
    # plain openpyxl writer below if xlsxwriter isn't installed.
    try:
        from . import prettyxl
        if prettyxl.xlsxwriter_available():
            return prettyxl.write_filled_pretty(result, out_path, which=which)
    except Exception:
        pass  # any issue -> fall through to the dependable openpyxl writer

    from openpyxl import Workbook
    import textwrap

    if which == "verified":
        filled = result.filled_verified
        fs = result.fill_summary_verified
        gaps = result.fill_gaps_verified
    elif which == "statistical_full":
        filled = getattr(result, "filled_statistical_full", None)
        fs = result.fill_summary
        gaps = result.fill_gaps
    else:
        filled = result.filled
        fs = result.fill_summary
        gaps = result.fill_gaps
    wb = Workbook(write_only=True)

    # --- Filled_Data (the deliverable) ---
    ws = wb.create_sheet("Filled_Data")
    if filled is not None and len(filled):
        ws.append([_wcell(ws, str(c), bold=True) for c in filled.columns])
        for row in filled.itertuples(index=False, name=None):
            ws.append([_wcell(ws, v) for v in row])
    else:
        ws.append([_wcell(ws, "(no rows)")])

    # --- Column_Guide (so a consultant knows exactly what every field is) ---
    wsg = wb.create_sheet("Column_Guide")
    present = set(filled.columns) if (filled is not None and len(filled)) else set()
    is_verified = (which == "verified")
    wsg.append([_wcell(wsg, "HOW TO READ THIS FILE", bold=True)])
    for line in [
        "This is your original data, cleaned and made analysis-ready. Every column you "
        "uploaded is preserved. Blank cells were filled from public data where a real "
        "source exists; where none does, the cell reads \"N/A\" (never a guess).",
        "Columns whose names start with \"_\" were ADDED by this tool for analysis. Your "
        "original columns are never silently overwritten with an estimate — any recovered "
        "billing NPI is marked in _NPI_Source and scored in _NPI_Confidence.",
        ("This is the VERIFIED build: billing NPIs are only what you provided (validated); "
         "blanks stay N/A because no public source records who billed a private claim."
         if is_verified else
         "This is the STATISTICAL build: blank billing NPIs are recovered/estimated, each "
         "tagged with its method and a calibrated confidence so you can filter on it."),
        "Tip: filter _NPI_Source to 'original' for a maximally clean subset, or keep "
        "recovered rows and threshold on _NPI_Confidence for your analysis.",
    ]:
        for w in textwrap.wrap(line, 116):
            wsg.append([_wcell(wsg, w)])
    wsg.append([])

    guide = [
        ("ORIGINAL DATA", "Billing_NPI_Final",
         "The cleaned billing NPI for the row: your value where present and check-digit-valid, "
         "otherwise the recovered one. Type-clean (no trailing .0), Luhn-validated.", "your data + recovery"),
        ("BILLING NPI — recovery", "_NPI_Source",
         "How this row's billing NPI was set: original (you provided it, validated) / recovered "
         "(point match in your own claims or CMS) / inferred (a sibling NPI under the same operator) "
         "/ best-guess (low confidence) / missing.", "recovery engine"),
        ("BILLING NPI — recovery", "_NPI_Confidence",
         "0–1 confidence for a recovered/inferred NPI, calibrated against held-out accuracy. "
         "Blank for 'original' rows — those aren't guesses.", "k-fold backtest"),
        ("BILLING NPI — recovery", "_NPI_BestGuess",
         "Marked when the billing NPI is a low-confidence best guess that did not clear the "
         "evidence bar. Treat as a lead to verify, not a fact.", "recovery engine"),
        ("BILLING NPI — recovery", "_NPI_Deactivated",
         "NPPES deactivation date if this NPI has been retired (blank = active). Flags NPIs that "
         "are no longer valid providers so they don't corrupt counts.", "CMS NPPES deactivation file"),
        ("ECONOMIC OWNER — roll-up", "_Billing_Parent_Group",
         "The parent operator this billing NPI rolls up to. Scattered NPIs that share a practice "
         "address, org name, or DBA/former-name thread are grouped under one label (the platform "
         "roll-up). Shows the operator name, or the anchor NPI when NPPES has no org name. A "
         "separate column — your original NPI/name are untouched.", "NPPES address + name clustering"),
        ("ECONOMIC OWNER — roll-up", "_Billing_Parent_NPI_Count",
         "How many distinct NPIs roll up to that same parent. >1 is a real multi-entity operator; "
         "1 is standalone. Use it to spot platforms hiding behind acquired-practice names.", "NPPES clustering"),
        ("DRUG identity", "_Drug_Ingredient",
         "Active ingredient(s) for the drug on this row, resolved from the NDC or the drug-name text.", "RxNorm / NLM"),
        ("DRUG identity", "_Drug_Brand",
         "Brand name for the drug (e.g. Stelara, Ocrevus, Entyvio), from the RxNorm branded record.", "RxNorm"),
        ("DRUG identity", "_Drug_Class",
         "Therapeutic class of the drug (e.g. Immunoglobulins, Monoclonal antibodies).", "RxNorm ATC / DailyMed"),
        ("ROUTING", "_Benefit_Channel",
         "Whether this drug is billed under the medical benefit (Part B) or the pharmacy benefit — "
         "this drives how the billing NPI is interpreted during recovery.", "HCPCS routing"),
        ("PROVIDER context", "_Billing_Facility_Affil",
         "The billing provider's facility affiliation(s) and facility type.", "CMS Facility Affiliation file"),
        ("PROVIDER context", "_Referring_Facility_Affil",
         "The referring provider's facility affiliation(s) and type.", "CMS Facility Affiliation file"),
        ("PROVIDER context", "_Facility_Ownership",
         "Ownership type of the provider's affiliated hospital (Proprietary / Voluntary non-profit / "
         "Government) — a PE-relevant signal.", "CMS Hospital General Information"),
        ("PROVIDER context", "_Billing_All_Specialties",
         "The billing provider's full specialty/taxonomy set from NPPES, not just the primary one.", "CMS NPPES"),
        ("PROVIDER context", "_Referring_PaidByDrugMaker",
         "\"Y\" when the referring physician received manufacturer payments tied to the drug on this claim — "
         "a free signal for the referring gap and a KOL/steering relationship flag. Blank = no such payment "
         "found (not proof of none).", "CMS Open Payments"),
        ("AUDIT", "_Cells_Filled",
         "Per-cell provenance: which cells in THIS row the tool filled, and how. Your audit trail.", "this tool"),
    ]
    cur = None
    for group, col, desc, src in guide:
        if col not in present and col != "Billing_NPI_Final":
            continue
        if group != cur:
            wsg.append([])
            wsg.append([_wcell(wsg, group, bold=True)])
            wsg.append([_wcell(wsg, "column", bold=True), _wcell(wsg, "what it means", bold=True),
                        _wcell(wsg, "source", bold=True)])
            cur = group
        wrapped = textwrap.wrap(desc, 92) or [""]
        wsg.append([_wcell(wsg, col), _wcell(wsg, wrapped[0]), _wcell(wsg, src)])
        for extra in wrapped[1:]:
            wsg.append([_wcell(wsg, ""), _wcell(wsg, extra), _wcell(wsg, "")])

    # --- Fill_Summary ---
    ws2 = wb.create_sheet("Fill_Summary")
    a = fs.attrs if (fs is not None and not fs.empty) else {}

    def _money(x):
        try:
            return "$" + format(float(x), ",.0f")
        except (TypeError, ValueError):
            return "$0"

    # provenance header — how every value in THIS file was formed
    mode = a.get("mode", which)
    note = a.get("method_note", "")
    ws2.append([_wcell(ws2, f"HOW THESE VALUES WERE FORMED  ({mode.upper()} OUTPUT)", bold=True)])
    for line in textwrap.wrap(note, 112):
        ws2.append([_wcell(ws2, line)])
    ws2.append([_wcell(ws2, "")])

    ws2.append([_wcell(ws2, "Filled output — summary", bold=True)])
    ws2.append([_wcell(ws2, "Rows in (original)"), _wcell(ws2, a.get("rows_in", 0))])
    ws2.append([_wcell(ws2, "Exact duplicate rows removed"), _wcell(ws2, a.get("n_dupes_removed", 0))])
    ws2.append([_wcell(ws2, "Rows out (deduplicated)"), _wcell(ws2, a.get("rows_out", 0))])
    ws2.append([])

    # Billing-NPI recovery, split by confidence and dollar-weighted. This is the
    # honest headline: a point-attributed recovery is a near-certain biller, a
    # distributional guess is not, and a Part D / unclassified row should never
    # carry an imputed biller at all. Blending them into one "% reduction" (as
    # v5 did) hid the guesses, so we report the three channels separately and
    # weight by allowed dollars, since that is what the diligence actually cares
    # about.
    ws2.append([_wcell(ws2, "Billing-NPI recovery — by confidence (dollar-weighted)", bold=True)])
    ws2.append([_wcell(ws2, "Billing NPIs blank in source"),
                _wcell(ws2, a.get("bill_blank_rows", 0)),
                _wcell(ws2, _money(a.get("bill_blank_dollars", 0)))])
    ws2.append([_wcell(ws2, "  -> filled, high confidence (point-attributed)"),
                _wcell(ws2, a.get("bill_high_rows", 0)),
                _wcell(ws2, _money(a.get("bill_high_dollars", 0))),
                _wcell(ws2, f"{a.get('bill_high_pct_dollars', 0)}% of blank $")])
    ws2.append([_wcell(ws2, "  -> filled, inferred from sibling rows (cluster / continuity)"),
                _wcell(ws2, a.get("bill_inferred_rows", 0)),
                _wcell(ws2, _money(a.get("bill_inferred_dollars", 0))),
                _wcell(ws2, f"{a.get('bill_inferred_pct_dollars', 0)}% of blank $")])
    ws2.append([_wcell(ws2, "  -> best-guess only, NOT in billing column (distributional)"),
                _wcell(ws2, a.get("bill_lowguess_rows", 0)),
                _wcell(ws2, _money(a.get("bill_lowguess_dollars", 0))),
                _wcell(ws2, f"{a.get('bill_lowguess_pct_dollars', 0)}% of blank $")])
    ws2.append([_wcell(ws2, "  -> not attributed (gross-up, never imputed)"),
                _wcell(ws2, a.get("bill_na_rows", 0)),
                _wcell(ws2, _money(a.get("bill_na_dollars", 0))),
                _wcell(ws2, f"{a.get('bill_na_pct_dollars', 0)}% of blank $")])
    by_reason = a.get("bill_na_by_reason", {}) or {}
    for reason, d in by_reason.items():
        ws2.append([_wcell(ws2, f"        {reason}"),
                    _wcell(ws2, d.get("rows", 0)),
                    _wcell(ws2, _money(d.get("dollars", 0)))])
    ws2.append([_wcell(ws2, "Billing NPIs present in source"),
                _wcell(ws2, a.get("bill_orig_present_rows", 0)),
                _wcell(ws2, _money(a.get("bill_orig_present_dollars", 0)))])
    ws2.append([])

    # Per-column raw fill counts (all fillable columns, not just billing NPI).
    if fs is not None and not fs.empty:
        ws2.append([_wcell(ws2, "Per-column fill counts", bold=True)])
        ws2.append([_wcell(ws2, c, bold=True) for c in fs.columns])
        for row in fs.itertuples(index=False, name=None):
            ws2.append([_wcell(ws2, v) for v in row])
    ws2.append([])
    ws2.append([_wcell(ws2, "Audit columns in Filled_Data:", bold=True)])
    ws2.append([_wcell(ws2, "  _NPI_Confidence"), _wcell(ws2, "billing-NPI grade per row: original / high [tier] / low-guess [tier] / not-attributed [reason]")])
    ws2.append([_wcell(ws2, "  _NPI_BestGuess"), _wcell(ws2, "top-3 candidate billers for distributional rows (these are NOT written to the real billing column)")])
    ws2.append([_wcell(ws2, "  _Benefit_Channel"), _wcell(ws2, "why a billing cell is what it is: Part B medical biller / Part B DME-home / Part D self-administered / unclassified code")])
    ws2.append([_wcell(ws2, "  _Cells_Filled"), _wcell(ws2, "which columns were filled in that row (blank = nothing filled)")])
    ws2.append([_wcell(ws2, "  _NPI_Source"), _wcell(ws2, "billing NPI: original / recovered (point) / inferred (sibling) / best-guess / missing")])
    ws2.append([_wcell(ws2, "  N/A"), _wcell(ws2, "value could not be found in public data (see Could_Not_Fill)")])

    # --- Could_Not_Fill ---
    ws3 = wb.create_sheet("Could_Not_Fill")
    ws3.append([_wcell(ws3, "Cells set to N/A — couldn't be found in public data", bold=True)])
    if gaps is not None and not gaps.empty:
        ws3.append([_wcell(ws3, c, bold=True) for c in gaps.columns])
        for row in gaps.itertuples(index=False, name=None):
            ws3.append([_wcell(ws3, v) for v in row])
    else:
        ws3.append([_wcell(ws3, "(nothing — everything fillable was filled)")])

    wb.save(out_path)
