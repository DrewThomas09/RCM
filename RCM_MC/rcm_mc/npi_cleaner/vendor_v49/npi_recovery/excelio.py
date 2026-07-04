"""Read the input claims file (keeping NPIs/ZIPs as strings) and write the
cleaned, multi-tab output workbook with professional formatting."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Calibri"
HDR_FILL = PatternFill("solid", fgColor="1F3864")     # deep navy
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SUB_FONT = Font(name=FONT, italic=True, size=9, color="595959")
CELL_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
INT_FMT = '#,##0;(#,##0);"-"'


def _dedupe_columns(df):
    """Make column names unique (a__2, a__3, ...) so df[name] never returns a
    DataFrame, and drop fully-empty unnamed columns that messy exports add."""
    seen, out = {}, []
    for c in df.columns:
        name = "" if c is None else str(c).strip()
        if name == "" or name.lower().startswith("unnamed"):
            name = "_unnamed"
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        out.append(name)
    df = df.copy()
    df.columns = out
    # drop columns that are entirely empty AND were auto-named
    drop = [c for c in df.columns if c.startswith("_unnamed") and df[c].isna().all()]
    return df.drop(columns=drop) if drop else df


CLAIM_KEYWORDS = (
    "npi", "hcpcs", "cpt", "hipps", "ndc", "jcode", "j-code", "j_code", "code",
    "procedure", "proc", "state", "zip", "postal", "allowed", "paid", "charge",
    "amount", "amt", "cost", "revenue", "units", "unit", "qty", "quantity",
    "drug", "brand", "generic", "payer", "plan", "insurer", "claim", "date",
    "dos", "referring", "rendering", "ordering", "billing", "provider",
    "prescriber", "specialty", "taxonomy", "place", "pos", "site", "county",
    "city", "member", "patient", "service",
)


def _looks_numeric(v):
    t = (v.replace(",", "").replace("$", "").replace("%", "").replace("(", "")
          .replace(")", "").replace("-", "").replace(".", "").strip())
    return t.isdigit() and t != ""


def _score_header_row(cells):
    """How header-like is this row? Returns (keyword_hits, text_cells)."""
    vals = [str(c).strip().lower() for c in cells
            if c is not None and str(c).strip() != "" and str(c).strip().lower() != "nan"]
    if not vals:
        return (0, 0)
    kw = sum(1 for v in vals if any(k in v for k in CLAIM_KEYWORDS))
    numeric = sum(1 for v in vals if _looks_numeric(v))
    return (kw, len(vals) - numeric)


def _grid_to_frame(grid):
    """Find the header row in a raw (header=None) grid and return a clean frame,
    or None if the grid has nothing usable. Skips banner/title rows above the
    real headers. Returns (frame, keyword_hits, n_data_rows)."""
    n = len(grid)
    if n == 0:
        return None
    best = None  # (kw, score, row_index)
    for i in range(min(n, 25)):
        kw, txt = _score_header_row(list(grid.iloc[i].values))
        if txt <= 0:
            continue
        cand = (kw, txt, -i)  # prefer keyword-rich, text-rich, earliest row
        if best is None or cand > best[0]:
            best = (cand, i)
    if best is None:
        return None
    hdr_i = best[1]
    kw_hits = best[0][0]
    header = grid.iloc[hdr_i].tolist()
    cols, seen = [], {}
    for j, h in enumerate(header):
        name = "" if h is None or str(h).strip().lower() in ("", "nan") else str(h).strip()
        if not name:
            name = f"_unnamed_{j}"
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 0
        cols.append(name)
    body = grid.iloc[hdr_i + 1:].copy()
    body.columns = cols
    # drop fully-empty rows
    body = body.dropna(how="all")
    # drop only UNNAMED empty columns (banner junk); keep named columns even if
    # empty in this extract — the caller may want to fill them.
    keep = [c for c in body.columns
            if (not str(c).startswith("_unnamed_")) or not body[c].isna().all()]
    body = body.loc[:, keep]
    n_data = len(body)
    return (body.reset_index(drop=True), kw_hits, n_data)


def _sniff_delim(path, default=","):
    """Detect the delimiter from the first non-empty line (handles , ; tab |)."""
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            for line in fh:
                if line.strip():
                    counts = {d: line.count(d) for d in [",", ";", "\t", "|"]}
                    best = max(counts, key=counts.get)
                    return best if counts[best] > 0 else default
    except Exception:
        pass
    return default


def _excel_engine():
    """Prefer python-calamine (Rust-backed, ~10x faster than openpyxl on large
    workbooks — 33s vs minutes on a 600K-row file); fall back to openpyxl when
    calamine isn't installed."""
    import importlib.util
    return "calamine" if importlib.util.find_spec("python_calamine") else None


def _read_excel(path, **kw):
    eng = _excel_engine()
    try:
        return pd.read_excel(path, engine=eng, **kw) if eng else pd.read_excel(path, **kw)
    except Exception:
        return pd.read_excel(path, **kw)   # any calamine hiccup -> openpyxl


def read_claims(path):
    """Read a claims file. For Excel, scans EVERY sheet and auto-detects the
    sheet + header row that actually hold the claims (handles cover/title sheets
    and banner rows above the headers). For CSV/TSV, sniffs the delimiter and
    detects the header row too."""
    path = Path(path)

    if path.suffix.lower() in (".csv", ".tsv", ".txt"):
        default = "\t" if path.suffix.lower() == ".tsv" else ","
        sep = _sniff_delim(path, default=default)
        grid = pd.read_csv(path, dtype=str, header=None, sep=sep, engine="python",
                           encoding="utf-8-sig", on_bad_lines="skip",
                           keep_default_na=False).replace({"": None})
        picked = _grid_to_frame(grid)
        if picked is None:
            df = pd.read_csv(path, dtype=str, sep=sep, engine="python",
                             encoding="utf-8-sig", on_bad_lines="skip",
                             keep_default_na=False)
            return _dedupe_columns(df.replace({"": pd.NA}))
        return _dedupe_columns(picked[0].replace({"": pd.NA}))

    # Excel: read all sheets raw, score each, pick the best
    sheets = _read_excel(path, sheet_name=None, header=None, dtype=str)
    best = None  # (kw, n_data, frame, sheet_name)
    for name, grid in sheets.items():
        picked = _grid_to_frame(grid)
        if picked is None:
            continue
        frame, kw, n_data = picked
        if frame.shape[1] == 0 or n_data == 0:
            continue
        cand = (kw, n_data)
        if best is None or cand > best[:2]:
            best = (kw, n_data, frame, name)
    if best is None:
        # nothing scored — fall back to the plain first-sheet read
        return _dedupe_columns(_read_excel(path, dtype=str).replace({"": pd.NA}))
    return _dedupe_columns(best[2].replace({"": pd.NA}))


def _fmt_for(colname):
    c = str(colname).lower()
    if any(k in c for k in ("dollar", "allowed", "amt", "paid", "charge", "revenue")):
        return CUR_FMT
    if c.startswith("pct") or c.endswith("pct_present") or c == "pct_present":
        return '0.0'
    if any(k in c for k in ("share", "acc", "top1", "top3", "confidence", "rate", "purity")):
        return PCT_FMT
    if any(k in c for k in ("rows", "count", "support", "srvcs", "units", "benes", "n_", "qty")):
        return INT_FMT
    return None


def write_table(ws, df, start_row=1, title=None, subtitle=None):
    r = start_row
    if title:
        ws.cell(r, 1, title).font = TITLE_FONT
        r += 1
    if subtitle:
        c = ws.cell(r, 1, subtitle)
        c.font = SUB_FONT
        c.alignment = Alignment(wrap_text=False)
        r += 2
    if df is None or len(df) == 0:
        ws.cell(r, 1, "(no rows)").font = Font(name=FONT, italic=True, color="999999")
        return
    df = df.copy()
    df = df.loc[:, ~df.columns.duplicated()]   # guard against duplicate headers
    # header
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(r, j, str(col))
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(r + 1, 1)
    fmts = {j: _fmt_for(col) for j, col in enumerate(df.columns, start=1)}
    # body
    for i, (_, row) in enumerate(df.iterrows(), start=r + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            elif hasattr(val, "item"):
                val = val.item()
            cell = ws.cell(i, j, val)
            cell.font = CELL_FONT
            cell.border = BORDER
            if fmts[j] and isinstance(val, (int, float)):
                cell.number_format = fmts[j]
    # widths
    for j, col in enumerate(df.columns, start=1):
        try:
            lengths = df[col].astype(str).str.len()
            maxlen = int(lengths.max()) if lengths.notna().any() else 8
        except Exception:
            maxlen = 8
        width = min(max(len(str(col)) + 2, maxlen) + 1, 48)
        ws.column_dimensions[get_column_letter(j)].width = max(width, 10)


def write_kv(ws, title, pairs, notes=None):
    ws.cell(1, 1, title).font = TITLE_FONT
    r = 3
    for k, v in pairs:
        kc = ws.cell(r, 1, k); kc.font = Font(name=FONT, bold=True, size=10)
        vc = ws.cell(r, 2, v); vc.font = CELL_FONT
        vc.alignment = Alignment(wrap_text=True)
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90
    if notes:
        r += 1
        for line in notes:
            c = ws.cell(r, 1, line); c.font = CELL_FONT
            c.alignment = Alignment(wrap_text=True)
            r += 1


def build_workbook(out_path, sheets, summary_kv, summary_notes, caveats):
    """sheets: ordered list of (sheet_name, df, title, subtitle)."""
    wb = Workbook()
    cover = wb.active
    cover.title = "README"
    write_kv(cover, "Billing-NPI Recovery — Output Workbook", summary_kv, summary_notes)

    for name, df, title, subtitle in sheets:
        ws = wb.create_sheet(name[:31])
        write_table(ws, df, title=title, subtitle=subtitle)

    cav = wb.create_sheet("Caveats")
    cav.cell(1, 1, "What this output can and cannot claim").font = TITLE_FONT
    for i, line in enumerate(caveats, start=3):
        c = cav.cell(i, 1, line); c.font = CELL_FONT
        c.alignment = Alignment(wrap_text=True)
    cav.column_dimensions["A"].width = 120

    wb.save(out_path)
    return out_path
