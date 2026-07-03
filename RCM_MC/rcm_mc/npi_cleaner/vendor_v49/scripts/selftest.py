"""End-to-end self-test. Generates a small messy multi-state sample, runs the
full pipeline against the live CMS/NPPES APIs, and checks the key invariants.

    python selftest.py

Exits 0 on success, 1 on failure. Useful to confirm the tool works in your
environment and that the public APIs are reachable.
"""

import sys
import tempfile
from pathlib import Path

import pandas as pd

from make_sample import build
from npi_recovery import run_pipeline, write_report
from npi_recovery.repair import npi_is_valid
from npi_recovery.geo import state_from_zip


def _check(name, cond):
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}")
    return bool(cond)


def main():
    ok = True

    print("Unit checks:")
    ok &= _check("NPI Luhn valid(real NPI passes)", npi_is_valid("1003914151"))
    ok &= _check("NPI Luhn valid( bad NPI fails)", not npi_is_valid("1234567890"))
    ok &= _check("ZIP->state (77002=TX)", state_from_zip("77002") == "TX")
    ok &= _check("ZIP->state (10001=NY)", state_from_zip("10001") == "NY")

    # v20: Splink fuzzy entity layer — works WITH splink installed, degrades to a
    # clean no-op WITHOUT it. Uses a mock NPPES so this check needs no network.
    from npi_recovery import entity, splink_entity
    class _MockN:
        D = {"1000000001": dict(name="Clinical Specialty Infusions Pharmacy", postal="75019", taxonomy_code="3336C0003X", addr_line="1700 Lakewood Dr", official="Jane Roe", other_org_names="CSI"),
             "1000000002": dict(name="CSI Pharmacy", postal="75019", taxonomy_code="3336C0003X", addr_line="1700 Lakewood Drive", official="Jane Roe", other_org_names="Clinical Specialty Infusions"),
             "1000000003": dict(name="Clinical Speciality Infusion", postal="75019", taxonomy_code="3336C0003X", addr_line="1700 Lakewood Dr Ste 2", official="", other_org_names=""),
             "1000000004": dict(name="KabaFusion Texas", postal="77002", taxonomy_code="3336C0003X", addr_line="900 Main St", official="", other_org_names=""),
             "1000000005": dict(name="Kaba Fusion TX LLC", postal="77002", taxonomy_code="3336C0003X", addr_line="900 Main Street", official="", other_org_names=""),
             "1000000006": dict(name="Houston Infusion Associates", postal="77030", taxonomy_code="261QM0850X", addr_line="6400 Fannin St", official="", other_org_names="")}
        def lookup(self, npi):
            d = dict(self.D.get(str(npi), {})); d["npi"] = str(npi); return d
    _mn = _MockN(); _npis = list(_MockN.D.keys())
    _ent = pd.DataFrame([{"npi": n, "name": _MockN.D[n]["name"].upper(), "type": "", "taxonomy": "",
                          "state": "TX", "other_org_names": _MockN.D[n]["other_org_names"],
                          "parent_key": n, "parent_operator": _MockN.D[n]["name"].upper(),
                          "match_basis": "singleton"} for n in _npis])
    _edges = splink_entity.build_fuzzy_links(_npis, _mn)
    ok &= _check("splink layer returns a list (no crash)", isinstance(_edges, list))
    _ent2, _roll = entity.merge_fuzzy_links(_ent, _edges)
    ok &= _check("merge_fuzzy_links never increases operator count (can only merge)",
                 _ent2["parent_operator"].nunique() <= _ent["parent_operator"].nunique())
    _cw = entity.make_crosswalk(_ent2)
    _prf = splink_entity.pairwise_prf(_cw, [{"1000000001", "1000000002", "1000000003"}, {"1000000004", "1000000005"}])
    if splink_entity.splink_available():
        ok &= _check("splink installed: fuzzy edges found on typo/suffix variants", len(_edges) > 0)
        ok &= _check("splink installed: no FALSE merges of distinct operators (precision==1.0)",
                     _prf["precision"] in (None, 1.0))
    else:
        ok &= _check("splink absent: graceful no-op (zero edges)", len(_edges) == 0)

    # v25: readout analytics — deterministic cuts, no network. Build a synthetic
    # canonical std frame and confirm every cut populates and the honesty
    # invariants hold (single-source flag fires; reference columns are flagged;
    # forward-risk total row exists; landscape defaults to a single year).
    from npi_recovery import analytics as _an
    import numpy as _np
    _rng = _np.random.default_rng(11); _n = 1500
    _z = _rng.choice(["770", "750", "787"], _n)
    _std_an = pd.DataFrame({
        "referring_npi": _rng.choice([f"REF{i:03d}" for i in range(25)], _n),
        "referring_specialty": _rng.choice(["Neurology", "Gastroenterology", "Hospital"], _n),
        "hcpcs": _rng.choice(["J1569", "J3241", "J3358", "J1745"], _n),
        "drug_name": _rng.choice(["immune globulin", "ustekinumab", "vedolizumab", "daptomycin"], _n),
        "pos": _rng.choice(["11", "12", "22"], _n), "zip3": _z, "zip": [z + "01" for z in _z],
        "state": "TX", "allowed_amt": _rng.gamma(3, 1200, _n).round(2),
        "units": _rng.integers(1, 30, _n),
        "payer": _rng.choice(["BCBS-TX", "UHC", "TRICARE"], _n),
        "date": pd.to_datetime([f"{y}-05-15" for y in _rng.choice([2023, 2024, 2025], _n)]),
        "billing_npi": _rng.choice([f"{1000000000 + i}" for i in range(10)], _n)})
    _std_an["is_blank_billing"] = False
    _mm = _std_an["referring_npi"] == "REF003"
    _std_an.loc[_mm, "billing_npi"] = "1000000009"; _std_an.loc[_mm, "drug_name"] = "immune globulin"
    _pm = {f"{1000000000 + i}": ("KabaFusion" if i == 9 else f"Op_{i % 4}") for i in range(10)}
    _ana = _an.build_analytics(_std_an, _std_an["billing_npi"], _pm, {}, granularity="zip3")
    _need = {"Referral_Concentration", "Submarket_Landscape", "Submarket_Saturation",
             "Referral_Target_Map", "Benefit_Shift_Exposure", "Formulary_Scorecard"}
    ok &= _check("v25 analytics: all readout cuts present", _need.issubset(set(_ana)))
    ok &= _check("v25 analytics: referral single-source flag fires",
                 bool((_ana["Referral_Concentration"]["single_source_flag"] == "Y").any()))
    ok &= _check("v25 analytics: formulary reference columns are flagged (__ref)",
                 any(c.endswith("__ref") for c in _ana["Formulary_Scorecard"].columns))
    ok &= _check("v25 analytics: benefit-shift forward-risk total row exists",
                 bool(_ana["Benefit_Shift_Exposure"]["drug_label"].astype(str).str.contains("TOTAL").any()))
    ok &= _check("v25 analytics: landscape defaults to single year",
                 bool(_ana["Submarket_Landscape"]["window"].astype(str).str.contains("single year").any()))
    ok &= _check("v25 analytics: saturation is relative (percentile, no fixed benchmark)",
                 "operator_density_pctile" in _ana["Submarket_Saturation"].columns)
    # v26: additional panel cuts + VRDC-free sizing graceful degradation
    ok &= _check("v26 analytics: site-of-care, payer landscape, regulatory cuts present",
                 {"SiteOfCare_Trend", "Payer_Landscape", "Regulatory_Scorecard"}.issubset(set(_ana)))
    ok &= _check("v26 analytics: payer landscape has HHI + concentration flag",
                 {"payer_hhi", "concentration_flag"}.issubset(set(_ana["Payer_Landscape"].columns)))
    ok &= _check("v26 analytics: regulatory scorecard is reference-flagged with a tailwinds row",
                 any(c.endswith("__ref") for c in _ana["Regulatory_Scorecard"].columns)
                 and bool(_ana["Regulatory_Scorecard"]["drug_label"].astype(str).str.contains("TAILWIND").any()))
    from npi_recovery import market_sizing as _ms
    _est, _meth = _ms.build_market_size(_std_an, None, state="TX")   # cms=None -> graceful note, no crash
    ok &= _check("v26 market sizing: degrades cleanly without a CMS client",
                 isinstance(_est, pd.DataFrame) and isinstance(_meth, pd.DataFrame) and len(_meth) > 0)
    # v27: growth pockets, referral leakage, therapeutic adjacency + attractiveness degradation
    ok &= _check("v27 analytics: growth pockets, referral leakage, adjacency cuts present",
                 {"Growth_Pockets", "Referral_Leakage", "Therapeutic_Adjacency"}.issubset(set(_ana)))
    ok &= _check("v27 analytics: growth pockets carries a build-opportunity score",
                 "build_opportunity_score" in _ana["Growth_Pockets"].columns)
    ok &= _check("v27 analytics: referral leakage quantifies addressable leaked volume",
                 "addressable_leaked_allowed" in _ana["Referral_Leakage"].columns)
    ok &= _check("v27 analytics: therapeutic adjacency flags whitespace",
                 bool(_ana["Therapeutic_Adjacency"]["status"].astype(str).str.contains("Whitespace").any()))
    _att, _attm = _ms.build_market_attractiveness(None, ["TX", "FL"], ["J1569"])  # no client -> graceful
    ok &= _check("v27 attractiveness: degrades cleanly without a CMS client / codes",
                 isinstance(_att, pd.DataFrame) and isinstance(_attm, pd.DataFrame) and len(_attm) > 0)
    _ph, _phm = _ms.build_pharmacy_benefit_grossup(None, _std_an)  # no client -> graceful note
    ok &= _check("v28 pharmacy gross-up: degrades cleanly without a CMS client",
                 isinstance(_ph, pd.DataFrame) and isinstance(_phm, pd.DataFrame) and len(_phm) > 0)
    ok &= _check("v28 pharmacy gross-up: method documents the lower-bound caveat",
                 bool(_phm["value"].astype(str).str.contains("Excludes pharmacy-only").any()))

    # v29: site-of-care reclassification + pharmacy dual-feed + new analytics.
    # All deterministic, no network.
    from npi_recovery import site_of_care as _soc, pharmacy_feed as _pf
    # the v25 synthetic frame (_std_an) is multi-year with POS + billing NPI, so
    # build_analytics now also emits the v29 cuts on it.
    ok &= _check("v29 analytics: site reclassification + proof point present",
                 {"SiteOfCare_Reclassification", "SiteOfCare_ProofPoint"}.issubset(set(_ana)))
    ok &= _check("v29 analytics: case-mix, share-shift, growth-recon present",
                 {"Operator_Case_Mix", "Share_Shift_Markets", "Growth_Reconciliation"}.issubset(set(_ana)))
    ok &= _check("v29 reclassification: keeps site dollars conserved (as-billed total == reclassified total)",
                 True)  # exercised numerically below on a controlled frame

    # controlled reclassification frame: a home-infusion code must override an
    # office POS; an infusion-clinic taxonomy must lift office->AIC; per-diem
    # support rows must be flagged out of drug volume.
    def _mk(n, biller, pos, code, drug, claimbase, yr=2025):
        return pd.DataFrame({
            "claim_id": [f"{claimbase}{i}" for i in range(n)],
            "patient_id": [f"P{i%10}" for i in range(n)],
            "date": pd.to_datetime([f"{yr}-06-15"] * n),
            "billing_npi": [biller] * n, "hcpcs": [code] * n, "drug_name": [drug] * n,
            "pos": [pos] * n, "zip": ["77030"] * n, "zip3": ["770"] * n, "state": ["TX"] * n,
            "allowed_amt": [800.0] * n, "units": [10] * n})
    _hf = pd.concat([
        _mk(30, "1000000004", "11", "J1561", "IVIG", "H"),          # office POS...
        _mk(30, "1000000004", "11", "S9494", "", "H"),              # ...with a home per-diem in the SAME claim
        _mk(30, "1000000009", "11", "J2350", "OCREVUS", "A"),       # AIC operator at office POS
        _mk(30, "1000000009", "11", "96413", "", "A"),              # in-clinic admin companion
        _mk(40, "1000000001", "11", "J1745", "INFLIXIMAB", "C"),    # true physician office
    ], ignore_index=True)
    _tax = {"1000000004": "3336H0001X", "1000000009": "261QI0500X", "1000000001": "207RH0003X"}
    _sc, _diag = _soc.reclassify(_hf, taxonomy_of=_tax, drug_label=_hf["drug_name"])
    _home_rows = _hf[(_hf["hcpcs"] == "J1561")].index
    ok &= _check("v29 reclassification: home-infusion code overrides office POS -> Home",
                 bool((_sc.loc[_home_rows, "site_reclassified"] == "Home").all()))
    _aic_rows = _hf[(_hf["hcpcs"] == "J2350")].index
    ok &= _check("v29 reclassification: infusion-clinic taxonomy lifts office J2350 -> AIC",
                 bool((_sc.loc[_aic_rows, "site_reclassified"] == "AIC").all()))
    ok &= _check("v29 reclassification: per-diem/admin lines flagged as support (out of drug volume)",
                 bool(_sc.loc[_hf[_hf["hcpcs"].isin(["S9494", "96413"])].index, "is_support_code"].all()))
    _summ = _soc.site_reclassification_summary(_hf, _sc, _diag)
    _buckets = ["Office", "AIC", "AIS", "Home", "Outpatient hospital", "Inpatient", "Other / unknown"]
    _summ_b = _summ[_summ["site_of_care"].isin(_buckets)]
    _ab = pd.to_numeric(_summ_b["as_billed_allowed"], errors="coerce").fillna(0).sum()
    _rc = pd.to_numeric(_summ_b["reclassified_allowed"], errors="coerce").fillna(0).sum()
    ok &= _check("v29 reclassification: dollars conserved (as-billed == reclassified, support excluded)",
                 abs(_ab - _rc) < 1.0)

    # dual-feed union + channel reconciliation: declining medical + growing
    # pharmacy on a shared operator -> combined growth must exceed medical-only.
    def _ff(n, billers, codes, yrs, cb):
        import numpy as _np2
        r = _np2.random.default_rng(5)
        yy = r.choice(yrs, n)
        return pd.DataFrame({
            "claim_id": [f"{cb}{i}" for i in range(n)], "patient_id": [f"P{i%30}" for i in range(n)],
            "date": pd.to_datetime([f"{y}-06-15" for y in yy]),
            "billing_npi": r.choice(billers, n), "hcpcs": r.choice(codes, n),
            "drug_name": r.choice(["IVIG", "OCREVUS"], n), "pos": r.choice(["11", "12"], n),
            "zip": ["77030"] * n, "zip3": ["770"] * n, "state": ["TX"] * n,
            "allowed_amt": r.uniform(400, 1200, n).round(2), "units": r.integers(1, 12, n),
            "is_blank_billing": [False] * n})
    _med = _ff(500, ["2000000001"], ["J1561", "J2350"], [2023, 2024, 2025], "M")
    _med = pd.concat([_med[_med.date.dt.year == 2023],
                      _med[_med.date.dt.year == 2024].sample(frac=0.75, random_state=2),
                      _med[_med.date.dt.year == 2025].sample(frac=0.55, random_state=2)], ignore_index=True)
    _phf = _ff(400, ["2000000001"], ["J3357", "J1551"], [2023, 2024, 2025], "R")
    _phf = pd.concat([_phf[_phf.date.dt.year == 2023].sample(frac=0.6, random_state=2),
                      _phf[_phf.date.dt.year == 2024].sample(frac=0.8, random_state=2),
                      _phf[_phf.date.dt.year == 2025]], ignore_index=True)
    _union, _dd = _pf.union_feeds(_med, _phf)
    ok &= _check("v29 dual-feed: union row count == medical + pharmacy (no spurious drops here)",
                 len(_union) == len(_med) + len(_phf))
    ok &= _check("v29 dual-feed: _claim_source present with both medical and pharmacy",
                 set(_union["_claim_source"].unique()) == {"medical", "pharmacy"})
    _opu = _an._operator_series(pd.Series(_union["billing_npi"].values), {"2000000001": "Onyx"})
    _recon = _pf.channel_reconciliation(_union, _opu, _an._year(_union))
    _tot = _recon[_recon["operator"].astype(str).str.startswith("TOTAL")].iloc[0]
    ok &= _check("v29 channel recon: combined YoY beats medical-only (pharmacy add-back lifts growth)",
                 float(_tot["combined_yoy_pct"]) > float(_tot["medical_only_yoy_pct"]))
    ok &= _check("v29 channel recon: pharmacy share of combined is reported and > 0",
                 float(_tot["pharmacy_share_of_combined_pct"]) > 0)
    # growth reconciliation labels CAGR and cumulative distinctly
    _gr = _an.growth_reconciliation(_union, _union["allowed_amt"], _an._year(_union), has_pharmacy=True)
    ok &= _check("v29 growth recon: separate CAGR and cumulative rows (no mislabel)",
                 bool(_gr["year"].astype(str).str.contains("CAGR").any()) and
                 bool(_gr["year"].astype(str).str.contains("Cumulative").any()))
    # the v29 synthetic builders run
    try:
        from make_sample import build_with_admin, build_pharmacy
        _bw = build_with_admin(rows=300, seed=5)
        _bp = build_pharmacy(rows=120, seed=5)
        ok &= _check("v29 sample builders: medical-with-admin carries 9xxxx/Sxxxx codes",
                     bool(_bw["ProcedureCode"].astype(str).str.match(r"(96\d{3}|99601|99602|S9\d{3}|G00\d\d)").any()))
        ok &= _check("v29 sample builders: pharmacy feed builds with rows", len(_bp) > 0)
    except Exception as _e:
        ok &= _check(f"v29 sample builders run ({type(_e).__name__})", False)


    # v30: NDC/name -> drug-class bridge, dual-channel per-molecule reconciliation
    # with a double-count guard, pharmacy money synthesis, unit basis, idempotency.
    from npi_recovery import rx_bridge as _rxb
    _xw = _rxb.load_crosswalk()
    ok &= _check("v30 crosswalk: seed loads with ingredient/brand/hcpcs maps",
                 len(_xw["by_ingredient"]) > 10 and len(_xw["by_hcpcs"]) > 10 and len(_xw["brand_tokens"]) > 10)
    # offline name-only RX -> class (brand in parens, bare brand, ingredient text)
    _rxn = pd.DataFrame({
        "_claim_source": ["pharmacy"] * 5,
        "hcpcs": [pd.NA] * 5, "ndc": [pd.NA] * 5,
        "drug_name": ["HIZENTRA (immune globulin SC)", "CUVITRU SCIG", "STELARA SUBCUTANEOUS",
                      "ENTYVIO", "TOTALLY UNKNOWN DRUG"],
        "allowed_amt": [1000, 1200, 3000, 2500, 500.0], "units": [5, 6, 45, 300, 1]})
    _rxr = _rxb.resolve_feed(_rxn, rxnorm=None)
    ok &= _check("v30 bridge: bare/paren brand names resolve to a drug class offline",
                 (_rxr.loc[:3, "drug_class_rx"] != "UNMAPPED_RX").all())
    ok &= _check("v30 bridge: immune globulin SCIG brands collapse to the IVIG/SCIG class",
                 (_rxr.loc[:1, "drug_class_rx"] == "Immune globulin (IVIG/SCIG)").all())
    ok &= _check("v30 bridge: unknown drug stays UNMAPPED_RX but is not dropped",
                 _rxr.loc[4, "drug_class_rx"] == "UNMAPPED_RX" and len(_rxr) == 5)
    ok &= _check("v30 bridge: pharmacy rows tagged units_basis=DISPENSED",
                 (_rxr["units_basis"] == "DISPENSED").all())
    # idempotency: re-running does not change already-resolved rows
    _rxr2 = _rxb.resolve_feed(_rxr, rxnorm=None)
    ok &= _check("v30 bridge: resolve_feed is idempotent (stable class on re-run)",
                 (_rxr2["drug_class_rx"].values == _rxr["drug_class_rx"].values).all())
    # medical J-codes (incl. alternate IVIG codes) map via the HCPCS table
    _medr = _rxb.resolve_feed(pd.DataFrame({
        "_claim_source": ["medical"] * 3, "hcpcs": ["J1561", "J1569", "J3358"],
        "ndc": [pd.NA] * 3, "drug_name": ["IVIG", "IVIG", "STELARA IV"],
        "allowed_amt": [4000, 5000, 3000.0], "units": [8, 10, 3000]}), rxnorm=None)
    ok &= _check("v30 bridge: alternate IVIG J-codes (J1561/J1569) both -> immune globulin",
                 (_medr.loc[:1, "drug_class_rx"] == "Immune globulin (IVIG/SCIG)").all())
    ok &= _check("v30 bridge: medical rows tagged units_basis=HCPCS",
                 (_medr["units_basis"] == "HCPCS").all())
    # dual-channel reconciliation + double-count guard on a shared patient
    _u = pd.concat([
        pd.DataFrame({"_claim_source": ["medical"] * 2, "patient_id": ["A", "A"],
                      "date": pd.to_datetime(["2025-02-01", "2025-03-01"]),
                      "hcpcs": ["J1569", "J1569"], "drug_class_rx": ["Immune globulin (IVIG/SCIG)"] * 2,
                      "allowed_amt": [5000.0, 5000.0]}),
        pd.DataFrame({"_claim_source": ["pharmacy"], "patient_id": ["A"],
                      "date": pd.to_datetime(["2025-04-01"]),
                      "hcpcs": ["J1559"], "drug_class_rx": ["Immune globulin (IVIG/SCIG)"],
                      "allowed_amt": [1500.0]})], ignore_index=True)
    _dc = _an.dual_channel_drug_reconciliation(_u, _u["allowed_amt"], _an._year(_u))
    _ivig = _dc[_dc["drug_class"] == "Immune globulin (IVIG/SCIG)"].iloc[0]
    ok &= _check("v30 dual-channel: combined = medical + pharmacy per molecule",
                 abs(float(_ivig["combined_allowed"]) - 11500.0) < 1.0)
    ok &= _check("v30 dual-channel: same patient on both benefits is flagged (double-count guard)",
                 _ivig["double_count_flag"] == "Y" and int(_ivig["both_benefit_patients"]) == 1)
    ok &= _check("v30 dual-channel: overlap is the smaller side ($1,500), netted from combined",
                 abs(float(_ivig["potential_double_count_allowed"]) - 1500.0) < 1.0 and
                 abs(float(_ivig["combined_net_of_overlap"]) - 10000.0) < 1.0)
    # pharmacy schema: dispensing NPI / quantity / days-supply detected; money
    # synthesised from ingredient cost + dispensing fee; bridge runs at read time
    try:
        import tempfile, os as _os
        from make_sample import build_pharmacy as _bpf
        _df = _bpf(rows=120, seed=9, style="name", cost_fee=True)
        _tmp = _os.path.join(tempfile.gettempdir(), "v30_rx.xlsx")
        _df.to_excel(_tmp, index=False)
        _stdph, _rep = _pf.read_and_standardize(_tmp, rxnorm=None)
        ok &= _check("v30 schema: dispensing-pharmacy NPI maps to billing_npi",
                     _stdph["billing_npi"].notna().any())
        ok &= _check("v30 schema: quantity-dispensed maps to units, days-supply captured",
                     _stdph["units"].notna().any() and "days_supply" in _stdph.columns and _stdph["days_supply"].notna().any())
        ok &= _check("v30 money: allowed synthesised from ingredient cost + dispensing fee",
                     pd.to_numeric(_stdph["allowed_amt"], errors="coerce").fillna(0).sum() > 0)
        ok &= _check("v30 read-time bridge: most name-only RX rows resolve to a class",
                     (_stdph["drug_class_rx"] != "UNMAPPED_RX").mean() > 0.6)
        _os.remove(_tmp)
    except Exception as _e:
        ok &= _check(f"v30 pharmacy schema/money/bridge run ({type(_e).__name__}: {_e})", False)


    # v21: complete labeling — zero blank DATA cells, specific reason tokens.
    from npi_recovery import fill as _fill
    _fr = pd.DataFrame({
        "BILLING_PROVIDER_NPI": ["1003914151", "", ""],
        "BILLING_PROVIDER_NAME": ["Acme Infusion", "", ""],
        "PAYER_NAME": ["", "", ""],
        "REFERRING_PROVIDER_NPI": ["1003914151", "", ""],
        "REFERRING_PROVIDER_NAME": ["", "", ""],
        "_NPI_Source": ["original", "best-guess", "missing"],
        "_Benefit_Channel": ["Part B — medical biller", "Part B — medical biller", "Part D — self-administered"],
        "_NPI_BestGuess": ["", "1111111111; 2222222222", ""]})
    _map = {"billing_npi": "BILLING_PROVIDER_NPI", "billing_name": "BILLING_PROVIDER_NAME",
            "payer": "PAYER_NAME", "referring_npi": "REFERRING_PROVIDER_NPI",
            "referring_name": "REFERRING_PROVIDER_NAME"}
    _out = _fill.relabel_unrecoverable(_fr.copy(), _map, verified=False)
    _data_cols = [c for c in _out.columns if not str(c).startswith("_")]
    _blanks = sum(int(_fill._is_blank(_out[c]).sum()) for c in _data_cols)
    ok &= _check("v21: ZERO blank data cells after relabel", _blanks == 0)
    ok &= _check("v21: empty payer -> UNRECOVERABLE_PAYER_ERISA",
                 bool((_out["PAYER_NAME"] == "UNRECOVERABLE_PAYER_ERISA").all()))
    ok &= _check("v21: Part-D blank billing -> PHARMACY_BENEFIT_NO_MEDICAL_NPI",
                 _out["BILLING_PROVIDER_NPI"].iloc[2] == "PHARMACY_BENEFIT_NO_MEDICAL_NPI")
    ok &= _check("v21: best-guess row -> BESTGUESS_BELOW_BAR",
                 _out["BILLING_PROVIDER_NPI"].iloc[1] == "BESTGUESS_BELOW_BAR")
    ok &= _check("v21: missing referrer -> NO_REFERRING_NPI_ON_CLAIM",
                 _out["REFERRING_PROVIDER_NPI"].iloc[1] == "NO_REFERRING_NPI_ON_CLAIM")

    print("\nEnd-to-end (small messy multi-state sample, live APIs):")
    tmp = Path(tempfile.mkdtemp())
    src = tmp / "selftest_claims.xlsx"
    out = tmp / "selftest_out.xlsx"
    df = build(rows=900, seed=99, national=True, messy=True)
    df.to_excel(src, index=False)
    n_blank_raw = int((df["BillingProviderNPI"].astype(str).str.strip() == "").sum())

    try:
        res = run_pipeline(str(src), top_hcpcs=5, do_entity=False, do_connectors=False,
                           do_health_audit=False, cache_dir=str(tmp / ".cache"), progress=None)
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"\nFAILED: pipeline raised {type(e).__name__}: {e}")
        return 1

    s = res.stats
    ok &= _check("pipeline returned a Result", res is not None)
    ok &= _check("rows preserved (cleaned == input length)", len(res.cleaned) == len(df))
    ok &= _check("found blank/invalid billing NPIs to fix", s["rows_blank_billing"] >= n_blank_raw)
    ok &= _check("recovered at least some billers", s["rows_recovered"] > 0)
    ok &= _check("field repairs were applied", s["field_repairs_total"] > 0)
    ok &= _check("repairs log is populated", not res.repairs_log.empty)
    ok &= _check("CMS candidate pool is non-empty (live data)", len(res.pool_table) > 0)
    ok &= _check("candidate pool spans multiple states", res.pool_table["state"].nunique() > 1)
    ok &= _check("honest accuracy computed", s.get("honest_top1") is not None)
    # enrichment built
    ok &= _check("provider directory built (NPPES enrichment)",
                 hasattr(res, "provider_directory") and res.provider_directory is not None
                 and set(["NPI", "Found_In_NPPES", "Primary_Specialty"]).issubset(set(res.provider_directory.columns)))
    ok &= _check("drug reference built (cross-source)",
                 hasattr(res, "drug_reference") and not res.drug_reference.empty
                 and "Natl_Medicare_Providers" in res.drug_reference.columns)
    ok &= _check("340B coverage built (eligibility signal)",
                 hasattr(res, "coverage_340b") and res.coverage_340b is not None
                 and "B340_Eligibility_Signal" in getattr(res.coverage_340b, "columns", []))
    # every recovered row must carry a tier (no un-tiered names)
    rec = res.recovery
    recovered = rec[rec["reason"] == "recovered"]
    ok &= _check("every recovered row has a confidence tier",
                 bool((recovered["tier"].astype(str).str.len() > 0).all()))
    # SAD rows must NOT carry a recovered NPI
    sad = rec[rec["reason"] == "sad"]
    ok &= _check("self-administered rows are not attributed",
                 bool(sad["recovered_npi"].isna().all()) if len(sad) else True)

    try:
        write_report(res, str(out))
        import openpyxl
        wb = openpyxl.load_workbook(out)
        need = {"README", "Cleaned_Claims", "Repairs_Log", "Data_Quality",
                "CMS_Candidate_Pool", "Recovery_Detail", "Provider_Directory",
                "Drug_Reference", "Coverage_340B", "Caveats",
                "Referral_Concentration", "Submarket_Landscape", "Formulary_Scorecard"}
        ok &= _check("workbook builds with all key tabs", need.issubset(set(wb.sheetnames)))
    except Exception as e:
        import traceback; traceback.print_exc()
        ok &= _check("workbook builds", False)

    # v22: connector-audit wiring + formatted (xlsxwriter) output
    ok &= _check("Result carries a connector_status field", hasattr(res, "connector_status"))
    try:
        from npi_recovery.report import write_filled
        from npi_recovery import prettyxl
        fout = tmp / "selftest_filled.xlsx"
        write_filled(res, str(fout), which="statistical")
        import openpyxl
        wbf = openpyxl.load_workbook(fout)
        if prettyxl.xlsxwriter_available():
            ok &= _check("formatted output: leading Summary sheet", wbf.sheetnames[0] == "Summary")
            fd = wbf["Filled_Data"]
            ok &= _check("formatted output: header frozen + autofilter on",
                         fd.freeze_panes is not None and fd.auto_filter.ref is not None)
        else:
            ok &= _check("plain output builds (xlsxwriter absent)", "Filled_Data" in wbf.sheetnames)
    except Exception as e:
        import traceback; traceback.print_exc()
        ok &= _check("filled output builds", False)

    # v23: three certainty tiers + census + pivot
    from npi_recovery import tiers
    sf = getattr(res, "filled_statistical_full", None)
    ok &= _check("tier-3 statistical-full frame exists", sf is not None and not sf.empty)
    if sf is not None and not sf.empty:
        ok &= _check("tier-3 has a _Review_Required column", "_Review_Required" in sf.columns)
        bcol = res.mapping.get("billing_npi")
        if bcol and bcol in sf.columns and bcol in res.filled.columns:
            from npi_recovery.tiers import _is_real
            more = int(_is_real(sf[bcol]).sum()) >= int(_is_real(res.filled[bcol]).sum())
            ok &= _check("tier-3 fills >= recovered tier (more billing cells populated)", more)
        try:
            cen = tiers.cell_census(res.filled_verified, res.filled, sf, res.mapping)
            tot = cen.attrs.get("totals", {})
            consistent = tot and (tot["certain"] + tot["recovered"] + tot["estimated"] + tot["unfillable"] == tot["cells"])
            ok &= _check("cell census totals are internally consistent", bool(consistent))
        except Exception as e:
            ok &= _check("cell census runs", False)
        try:
            piv = tiers.pivot_landscape(res.filled, sf, res.mapping)
            ok &= _check("landscape pivot returns a verdict", bool(piv.attrs.get("verdict")))
        except Exception:
            ok &= _check("landscape pivot runs", False)
        try:
            from npi_recovery.report import write_filled as _wf
            fp3 = tmp / "tier3.xlsx"
            _wf(res, str(fp3), which="statistical_full")
            import openpyxl as _ox
            ok &= _check("tier-3 file builds", "Filled_Data" in _ox.load_workbook(fp3).sheetnames)
        except Exception:
            ok &= _check("tier-3 file builds", False)

    # v23: operator-stratified leaderboard (deterministic unit checks)
    from npi_recovery import backtest as _bt
    ok &= _check("wilson lcb bounded and below point estimate", 0.0 <= _bt._wilson_lcb(8, 10) < 0.8)
    ok &= _check("wilson lcb of 0/0 == 0", _bt._wilson_lcb(0, 0) == 0.0)
    _hd = pd.DataFrame({
        "tier": ["T1"] * 40 + ["T2"] * 5, "amt": [100.0] * 45,
        "t1": [True] * 45, "t3": [True] * 45,
        "true_npi": ["1111111111"] * 40 + ["2222222222"] * 5,
        "pred_npi": ["1111111111"] * 38 + ["3333333333"] * 2 + ["2222222222"] * 5,
        "pred_top3": ["1111111111"] * 40 + ["2222222222"] * 5})
    _pt = pd.DataFrame({"tier": ["T1", "T2"], "top1_acc": [0.9, 0.5]})
    _fk = pd.DataFrame({"_Billing_Parent_Group": ["1111111111", "4444444444"],
                        "_NPI_Source": ["recovered", "recovered"], "AMT": [10.0, 20.0]})
    _lb = _bt.operator_leaderboard(_hd, _pt, parent_map={}, parent_size={"1111111111": 3},
                                   filled=_fk, mapping={"allowed_amt": "AMT"})
    ok &= _check("leaderboard builds rows", not _lb.empty)
    _op1 = _lb[_lb["parent_operator"] == "1111111111"]
    ok &= _check("ACCEPT#4: operator with >=30 holdout is ranked",
                 not _op1.empty and _op1.iloc[0]["verdict"] in ("HIGH_CONFIDENCE_RECOVERABLE", "MODERATE", "LOW_CONFIDENCE"))
    _op2 = _lb[_lb["parent_operator"] == "2222222222"]
    ok &= _check("ACCEPT#4: operator with <30 holdout -> INSUFFICIENT_HOLDOUT",
                 not _op2.empty and _op2.iloc[0]["verdict"] == "INSUFFICIENT_HOLDOUT")
    _op4 = _lb[_lb["parent_operator"] == "4444444444"]
    ok &= _check("ACCEPT#3: blank-only operator (no holdout) -> UNMEASURABLE",
                 not _op4.empty and _op4.iloc[0]["verdict"] == "UNMEASURABLE")
    ok &= _check("Result carries backtest_by_operator", hasattr(res, "backtest_by_operator"))
    ok &= _check("inherited operator columns present on filled",
                 "_Operator_Recovery_Verdict" in res.filled.columns)

    # v30: live dual-feed end-to-end — medical + a pharmacy file sharing billers,
    # so the union, NDC/name bridge, dual-channel reconciliation, pharmacy-NPI
    # enrichment and operator rollup all run through the real pipeline.
    print("\nv30 dual-feed end-to-end (medical + pharmacy, live APIs):")
    try:
        from make_sample import build_pharmacy as _bpf2
        _shared = [str(x) for x in pd.Series(df["BillingProviderNPI"]).astype(str)
                   if x.strip() and x.strip().lower() != "nan"][:5]
        _phx = _bpf2(rows=240, seed=77, national=True, shared_billers=_shared, style="ndc")
        psrc = tmp / "selftest_pharmacy.xlsx"
        _phx.to_excel(psrc, index=False)
        pout = tmp / "selftest_dualfeed_out.xlsx"
        res2 = run_pipeline(str(src), top_hcpcs=5, do_entity=False, do_connectors=False,
                            do_health_audit=False, pharmacy_path=str(psrc),
                            cache_dir=str(tmp / ".cache"), progress=None)
        a2 = res2.analytics if hasattr(res2, "analytics") else {}
        ok &= _check("v30 e2e: dual-channel drug reconciliation tab built",
                     "DualChannel_Drug_Reconciliation" in a2 and not a2["DualChannel_Drug_Reconciliation"].empty)
        ok &= _check("v30 e2e: Rx bridge coverage tab built",
                     "Rx_Bridge_Coverage" in a2)
        ok &= _check("v30 e2e: pharmacy provider attribution tab built",
                     "Pharmacy_Provider_Attribution" in a2)
        ok &= _check("v30 e2e: channel reconciliation present (pharmacy unioned)",
                     "Channel_Reconciliation" in a2)
        # the dual-channel tab actually stitched both sides for >=1 molecule
        _dct = a2.get("DualChannel_Drug_Reconciliation", pd.DataFrame())
        if "pharmacy_allowed" in _dct.columns:
            ok &= _check("v30 e2e: at least one molecule carries pharmacy dollars",
                         (pd.to_numeric(_dct["pharmacy_allowed"], errors="coerce").fillna(0) > 0).any())
        write_report(res2, str(pout))
        import openpyxl as _ox2
        wb2 = _ox2.load_workbook(pout)
        ok &= _check("v30 e2e: workbook includes the dual-feed tabs",
                     {"DualChannel_Drug_Reconciliation", "Rx_Bridge_Coverage",
                      "Pharmacy_Provider_Attribution"}.issubset(set(wb2.sheetnames)))
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v30 dual-feed e2e runs ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v31: the six data-engineering fixes from the Onyx / Project Infusion
    # transcript. All offline (seed-only, no network).
    # ------------------------------------------------------------------ #
    print("\nv31 checks (transcript fixes):")
    try:
        from npi_recovery import common_name as _cn31
        from npi_recovery import ndc_jcode as _nj31
        from npi_recovery import formulary as _fm31
        from npi_recovery import npi_channel as _nc31
        from npi_recovery import control_total as _ct31
        from npi_recovery import rx_bridge as _rxb31
        from npi_recovery import config as _cfg31
        _ref = "npi_recovery/reference"
        _xw = _rxb31.load_crosswalk(_ref)

        # (P1) common-name grouper: Stelara across J3357/J3358/J3590 collapses to
        # one molecule; a truly-unknown J3590 row does NOT merge into it.
        _s = pd.DataFrame([
            {"drug_name": "STELARA 90 MG", "hcpcs": "J3357", "ndc": "", "allowed": 25000},
            {"drug_name": "STELARA IV", "hcpcs": "J3358", "ndc": "", "allowed": 20000},
            {"drug_name": "ustekinumab-auub (WEZLANA)", "hcpcs": "J3590", "ndc": "", "allowed": 9000},
            {"drug_name": "totally unknown drug", "hcpcs": "J3590", "ndc": "", "allowed": 1500},
        ])
        _named = _cn31.assign_common_name(_s, crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _uste = _named[_named["drug_common_key"].str.contains("ustekinumab", case=False, na=False)]
        ok &= _check("v31 P1: Stelara collapses across 3 J-codes ($54k)",
                     abs(float(_uste["allowed"].sum()) - 54000) < 1)
        ok &= _check("v31 P1: unknown J3590 not merged into a real molecule",
                     not _named.loc[_named["drug_name"] == "totally unknown drug",
                                    "drug_common_key"].str.contains("ustekinumab", case=False, na=False).any())
        _roll = _cn31.common_name_rollup(_named, allowed=_named["allowed"])
        _sr = _roll[_roll["drug_common_key"] == "ustekinumab"].iloc[0]
        ok &= _check("v31 P1: rollup flags split_across_codes with first-code undercount",
                     bool(_sr["split_across_codes"]) and float(_sr["undercount_if_first_code_only"]) > 0)

        # (P2/P4) NDC -> best J-code: brand-specific wins, unknown -> NO_JCODE
        # (never force-mapped), unknown-brand IVIG -> class NOS code.
        _d = pd.DataFrame([
            {"drug_name": "CUVITRU", "ndc": "", "hcpcs": ""},
            {"drug_name": "some unknown biologic", "ndc": "", "hcpcs": ""},
            {"drug_name": "IVIG (unspecified brand)", "ndc": "", "hcpcs": ""},
        ])
        _res = _nj31.resolve_best_jcode(_d, crosswalk=_xw, ref_dir=_ref)
        _bs = dict(zip(_res["drug_name"], _res["_best_disposition"]))
        ok &= _check("v31 P2: branded NDC resolves to its own J-code (BRAND_SPECIFIC)",
                     _bs.get("CUVITRU") == "BRAND_SPECIFIC")
        ok &= _check("v31 P4: unknown drug is NOT force-mapped (NO_JCODE_KEEP_AS_NDC)",
                     _bs.get("some unknown biologic") == "NO_JCODE_KEEP_AS_NDC")

        # (P3) formulary gate: Keytruda excluded, OPAT antibiotic kept off-formulary,
        # crosswalk molecule in formulary.
        _fp = pd.DataFrame([
            {"drug_name": "KEYTRUDA 100 MG", "hcpcs": "J9271", "ndc": "", "allowed": 88000},
            {"drug_name": "vancomycin 1 gm", "hcpcs": "J3370", "ndc": "", "allowed": 3000},
            {"drug_name": "STELARA 90 MG", "hcpcs": "J3357", "ndc": "", "allowed": 10000},
        ])
        _fn = _cn31.assign_common_name(_fp, crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _spec = _fm31.load_formulary(_ref)
        _tag = _fm31.assign_formulary_disposition(_fn, _spec, _ref)
        _st = dict(zip(_tag["drug_name"], _tag["formulary_status"]))
        ok &= _check("v31 P3: Keytruda -> EXCLUDE_CONFIRMED", _st.get("KEYTRUDA 100 MG") == "EXCLUDE_CONFIRMED")
        ok &= _check("v31 P3: OPAT vancomycin -> KEEP_OFF_FORMULARY", _st.get("vancomycin 1 gm") == "KEEP_OFF_FORMULARY")
        ok &= _check("v31 P3: crosswalk molecule -> IN_FORMULARY", _st.get("STELARA 90 MG") == "IN_FORMULARY")

        # payer-type classifier underpins the channel call
        ok &= _check("v31 P5: TRICARE -> military_va", _cfg31.classify_payer_type("TRICARE West") == "military_va")
        ok &= _check("v31 P5: BCBS -> commercial", _cfg31.classify_payer_type("Blue Cross Blue Shield of Texas") == "commercial")
        ok &= _check("v31 P5: UHC Medicare Advantage -> medicare (not commercial)",
                     _cfg31.classify_payer_type("UnitedHealthcare Medicare Advantage") == "medicare")

        # (P5) NPI channel: pharmacy-supplier + government book -> GOVERNMENT_PHARMACY;
        # reconciliation against a client list flags CONFIRMED_GOV and not-in-panel.
        _cp = pd.DataFrame([
            {"billing_npi": "1111111116", "payer": "TRICARE West", "allowed": 6000},
            {"billing_npi": "1111111116", "payer": "Texas Medicaid STAR", "allowed": 4000},
            {"billing_npi": "2222222221", "payer": "Blue Cross Blue Shield of Texas", "allowed": 8000},
        ])
        _tax = {"1111111116": "3336C0002X", "2222222221": "207RH0003X"}
        _chan = _nc31.classify_npi_channels(_cp, allowed=_cp["allowed"], taxonomy_of=_tax)
        _cmap = dict(zip(_chan["billing_npi"], _chan["channel"]))
        ok &= _check("v31 P5: gov pharmacy NPI -> GOVERNMENT_PHARMACY", _cmap.get("1111111116") == "GOVERNMENT_PHARMACY")
        ok &= _check("v31 P5: commercial office NPI -> COMMERCIAL_MEDICAL", _cmap.get("2222222221") == "COMMERCIAL_MEDICAL")
        _rec = _nc31.government_reconciliation(_chan, {"1111111116", "9999999999"})
        _rmap = dict(zip(_rec["billing_npi"], _rec["reconciliation"]))
        ok &= _check("v31 P5: client-listed + panel-agrees -> CONFIRMED_GOV", _rmap.get("1111111116") == "CONFIRMED_GOV")
        ok &= _check("v31 P5: client NPI absent from panel -> CLIENT_LIST_NOT_IN_PANEL",
                     _rmap.get("9999999999") == "CLIENT_LIST_NOT_IN_PANEL")

        # (P6) control total: captured vs control math + honest exposure (buckets
        # already inside the panel are shown, not double-added).
        _ctr = _ct31.reconcile_control_total(captured=900_000, control_total=1_000_000, entity_label="Onyx")
        _lines = dict(zip(_ctr["line"], _ctr["amount"]))
        ok &= _check("v31 P6: captured vs control shortfall = control - captured",
                     abs(float(_lines.get("shortfall (control - captured)", 0)) - 100_000) < 1)
        _exp = _ct31.exposure_summary(panel_total=1_000_000, split_code_undercount=32_000,
                                      government_channel=41_000, control_total=1_050_000)
        ok &= _check("v31 P6: exposure tab renders split + government at-risk dollars",
                     len(_exp) >= 3 and "note" not in _exp.columns)
        ok &= _check("v31 P6: no control total -> honest note (not a fake match)",
                     "note" in _ct31.reconcile_control_total(captured=1, control_total=None).columns)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v31 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v32: the Onyx / Project Infusion comprehensive-report fixes. All offline
    # (seed-only, no network). Built on top of v31, not repeating it.
    # ------------------------------------------------------------------ #
    print("\nv32 checks (comprehensive-report fixes):")
    try:
        import numpy as _np32
        from npi_recovery import vrdc_suppression as _supp
        from npi_recovery import deficit_diagnostics as _dd32
        from npi_recovery import crosswalk_builder as _xb32
        from npi_recovery import npi_enrollment as _en32
        from npi_recovery import universe as _uni32
        from npi_recovery import therapy_area as _ta32
        from npi_recovery import coverage_grossup as _gu32
        from npi_recovery import common_name as _cn32
        from npi_recovery import rx_bridge as _rxb32
        _ref = "npi_recovery/reference"
        _xw = _rxb32.load_crosswalk(_ref)

        # (v32-1) suppression: single-suppressed cell recovered exactly; ceiling
        # below expected -> upstream deficit dominates.
        _grid = pd.DataFrame([
            {"prov": "A", "allowed": 40000, "benes": 50, "row_total": 100000},
            {"prov": "A", "allowed": 35000, "benes": 30, "row_total": 100000},
            {"prov": "A", "allowed": "*", "benes": 8, "row_total": 100000},
        ])
        _rec = _supp.reconcile_scope(_grid, value_col="allowed", scope_cols=["prov"],
                                     total_col="row_total", count_col="benes")
        ok &= _check("v32-1 suppression: single-cell residual is exact (complementary)",
                     bool(_rec.iloc[0]["exact_recoverable"]) and abs(float(_rec.iloc[0]["suppressed_residual"]) - 25000) < 1)
        _imp = _supp.impute_suppressed_cells(_grid, value_col="allowed", scope_cols=["prov"],
                                             total_col="row_total", count_col="benes")
        ok &= _check("v32-1 suppression: imputed scope sums to the aggregate ceiling",
                     abs(float(_imp["value_filled"].sum()) - 100000) < 1)
        _cr = _supp.ceiling_report(reconstructed_total=160000, external_expected=250000,
                                   visible_sum=125000, suppressed_residual=35000)
        ok &= _check("v32-1 suppression: ceiling below expected -> upstream deficit dominates",
                     _cr.attrs.get("verdict") == "UPSTREAM DEFICIT DOMINATES")

        # (v32-2) deficit: home-infusion codes land in DME/HIT/Part D; a medical-only
        # pull misses them; the dominant cause is claim-type scope.
        _hi = pd.DataFrame([
            {"hcpcs": "J1561", "allowed": 30000}, {"hcpcs": "G0068", "allowed": 5000},
            {"hcpcs": "J3357", "allowed": 8000}, {"hcpcs": "J1569", "allowed": 40000},
        ])
        _chn = _dd32.classify_claim_channel(_hi, ref_dir=_ref)
        _cm = dict(zip(_hi["hcpcs"], _chn))
        ok &= _check("v32-2 deficit: SCIG->DME, HIT G-code->HIT, SAD->Part D, IVIG->Part B",
                     _cm.get("J1561") == "DME_SUPPLY" and _cm.get("G0068") == "HIT_PROF"
                     and _cm.get("J3357") == "PARTD_SAD" and _cm.get("J1569") == "PARTB_MEDICAL")
        _cov = _dd32.claim_scope_coverage(_hi, allowed=_hi["allowed"], ref_dir=_ref,
                                          pull_channels=("PARTB_MEDICAL",))
        ok &= _check("v32-2 deficit: medical-only pull misses DME+HIT+Part D ($43k)",
                     abs(float(_cov.attrs.get("missed_share_allowed", 0)) - 43000) < 1)
        _dg = _dd32.diagnose_deficit(captured_total=79000, expected_total=254000,
                                     claim_scope_deficit=120000, book_structure_deficit=40000,
                                     filter_deficit=15000)
        ok &= _check("v32-2 deficit: dominant cause is claim-type / file scope",
                     "claim-type" in str(_dg.attrs.get("verdict", "")))

        # (v32-3) top-down crosswalk: infliximab carries originator + biosimilars;
        # dual-flag captures Inflectra by molecule while NOC is not admitted; NDC gaps found.
        _cw = _xb32.build_crosswalk(seed_ref_dir=_ref, use_example=True)
        _inflx = _cw["codes_by_molecule"].get("infliximab", {"hcpcs": set()})
        ok &= _check("v32-3 crosswalk: infliximab has originator + biosimilar codes (top-down)",
                     len(_inflx["hcpcs"]) >= 3)
        _panel = _cn32.assign_common_name(pd.DataFrame([
            {"drug_name": "Inflectra 100 MG", "hcpcs": "Q5103", "ndc": "59676-0310-20", "allowed": 12000},
            {"drug_name": "mystery NOC", "hcpcs": "J3590", "ndc": "99999-9999-99", "allowed": 1000},
        ]), crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _mem = _xb32.any_match_membership(_panel, _cw, {"hcpcs": {"J1745"}, "ndc": set()})
        _mm = dict(zip(_panel["drug_name"], _mem["capture_in"]))
        ok &= _check("v32-3 crosswalk: Inflectra captured via molecule any-match (J1745)",
                     bool(_mm.get("Inflectra 100 MG")) is True)
        ok &= _check("v32-3 crosswalk: shared NOC not admitted wholesale",
                     bool(_mm.get("mystery NOC")) is False)
        _gaps = _xb32.ndc_gap_targets(_panel, _cw, allowed=_panel["allowed"], molecules=["infliximab"])
        ok &= _check("v32-3 crosswalk: NDC-level gaps found under an included J-code",
                     ("gap_flag" in _gaps.columns) and bool(_gaps["gap_flag"].any()))

        # (v32-4) enrollment: pharmacy NPI billing SCIG+HIT+IVIG -> 3 gov channels;
        # NPI absent from roster flagged JV.
        _es = pd.DataFrame([
            {"billing_npi": "1111111116", "hcpcs": "J1561", "allowed": 30000},
            {"billing_npi": "1111111116", "hcpcs": "G0068", "allowed": 5000},
            {"billing_npi": "1111111116", "hcpcs": "J1569", "allowed": 10000},
        ])
        _em = _en32.map_enrollment_channels(_es, allowed=_es["allowed"],
                                            taxonomy_of={"1111111116": "3336C0002X"},
                                            ref_dir=_ref, roster_npis={"9999999999"})
        _r1 = _em.iloc[0]
        ok &= _check("v32-4 enrollment: pharmacy NPI implies DME + HIT + Part D files",
                     all(x in _r1["cms_files_implied"] for x in ("DME", "HIT", "Part D")))
        ok &= _check("v32-4 enrollment: NPI absent from roster flagged JV/missing",
                     bool(_r1["jv_or_missing_candidate"]) is True)

        # (v32-5) universe: Stelara grouped-in; Keytruda market-excluded; tiny below
        # floor; freeze keeps tiny OUT on a second source (no flicker).
        _up = _cn32.assign_common_name(pd.DataFrame([
            {"drug_name": "GAMMAGARD", "hcpcs": "J1569", "ndc": "", "allowed": 5_000_000},
            {"drug_name": "STELARA 90 MG", "hcpcs": "J3357", "ndc": "", "allowed": 1_200_000},
            {"drug_name": "STELARA IV", "hcpcs": "J3358", "ndc": "", "allowed": 900_000},
            {"drug_name": "tiny drug", "hcpcs": "J3490", "ndc": "", "allowed": 200_000},
            {"drug_name": "KEYTRUDA", "hcpcs": "J9271", "ndc": "", "allowed": 9_000_000},
        ]), crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _u = _uni32.define_universe(_up, allowed=_up["allowed"], floor=1_000_000, ref_dir=_ref)
        _st = dict(zip(_u["name"], _u["status"]))
        _stel = [k for k in _st if "Stelara" in k]
        _keyt = [k for k in _st if "eytruda" in k.lower() or "embroli" in k.lower()]
        ok &= _check("v32-5 universe: grouped Stelara in; Keytruda market-excluded despite $9M",
                     bool(_stel) and _st[_stel[0]] == "IN_UNIVERSE"
                     and bool(_keyt) and _st[_keyt[0]] == "EXCLUDED_MARKET_DEF")
        _keys = _uni32.frozen_universe_keys(_u)
        _src2 = _cn32.assign_common_name(pd.DataFrame([
            {"drug_name": "tiny drug", "hcpcs": "J3490", "ndc": "", "allowed": 800_000}]),
            crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _ap = _uni32.apply_frozen_universe(_src2, _keys)
        ok &= _check("v32-5 universe: freeze keeps below-floor drug OUT on 2nd source (no flicker)",
                     bool(_ap["in_universe"].iloc[0]) is False)
        ok &= _check("v32-5 universe: exclusion register documents Keytruda with a rationale",
                     "rationale" in _uni32.exclusion_register(_u).columns)

        # (v32-6) therapy area: acute share 22% inside 20-26 band; rituximab straddle
        # surfaced; chronic splits IVIG vs rare/orphan.
        _tp = _cn32.assign_common_name(pd.DataFrame([
            {"drug_name": "daptomycin", "hcpcs": "J0878", "ndc": "", "allowed": 12000},
            {"drug_name": "vancomycin", "hcpcs": "J3370", "ndc": "", "allowed": 10000},
            {"drug_name": "GAMMAGARD", "hcpcs": "J1569", "ndc": "", "allowed": 40000},
            {"drug_name": "efgartigimod", "hcpcs": "J9332", "ndc": "", "allowed": 20000},
            {"drug_name": "STELARA 90 MG", "hcpcs": "J3357", "ndc": "", "allowed": 10000},
            {"drug_name": "rituximab", "hcpcs": "J9312", "ndc": "", "allowed": 8000},
        ]), crosswalk=_xw, drug_ident=None, ref_dir=_ref)
        _tag = _ta32.assign_therapy_area(_tp, ref_dir=_ref)
        _ac = _ta32.acute_share_check(_tag, allowed=_tp["allowed"], band=(0.20, 0.26))
        ok &= _check("v32-6 therapy: acute share 22% is in the client band (joint grouper+map test)",
                     bool(_ac.attrs.get("in_band")) and abs(float(_ac.attrs.get("acute_share_pct")) - 22.0) < 0.5)
        _rev = _ta32.dominant_therapy_review(_tag, allowed=_tp["allowed"])
        ok &= _check("v32-6 therapy: rituximab straddle surfaced for hand-check",
                     ("molecule" in _rev.columns) and _rev["molecule"].str.contains("ituximab", case=False).any())
        _sub = _ta32.chronic_subdivision(_tag, allowed=_tp["allowed"])
        ok &= _check("v32-6 therapy: chronic subdivides IVIG vs rare/orphan",
                     _sub["chronic_subclass"].str.contains("IVIG").any()
                     and _sub["chronic_subclass"].str.contains("Rare").any())

        # (v32-7) gross-up: 14.1% -> 7.1x; 2-pt error swings ~15-17%; lives-vs-dollars flagged.
        _ge = _gu32.grossup_estimate(captured=50_000_000, coverage_ratio=0.141, basis="lives")
        _gm = dict(zip(_ge["line"], _ge["value"])).get("gross-up multiplier (1 / ratio)")
        ok &= _check("v32-7 grossup: 14.1% ratio is a 7.1x multiplier",
                     _gm is not None and abs(float(_gm) - 7.092) < 0.01)
        ok &= _check("v32-7 grossup: lives-vs-dollars assumption is flagged",
                     any("LIVES" in f for f in _ge.attrs.get("flags", [])))
        _gs = _gu32.grossup_sensitivity(captured=50_000_000, coverage_ratio=0.141, delta_points=2.0)
        ok &= _check("v32-7 grossup: 2-point ratio error swings the estimate 15-17%",
                     14.0 <= float(_gs.attrs.get("max_abs_swing_pct")) <= 18.0)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v32 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v33: close the loop, pin the ratio, protect the fallback. All offline.
    # ------------------------------------------------------------------ #
    print("\nv33 checks (loop-closing fixes):")
    try:
        from npi_recovery import trend_integrity as _ti33
        from npi_recovery import cross_source as _cs33
        from npi_recovery import roster_forensics as _rf33
        from npi_recovery import calibration as _cal33
        from npi_recovery import ma_proxy as _mp33
        from npi_recovery import market_view as _mv33
        from npi_recovery import coverage_grossup as _gu333
        from npi_recovery import crosswalk_builder as _xb333
        from npi_recovery import common_name as _cn33
        from npi_recovery import rx_bridge as _rxb33
        _ref = "npi_recovery/reference"
        _xw33 = _rxb33.load_crosswalk(_ref)

        # trend integrity: Q-code entry under a snapshot match bends the trend
        _t = pd.DataFrame([
            {"drug_common_name": "infliximab", "hcpcs": "J1745", "ndc": "", "allowed": 100, "y": 2024},
            {"drug_common_name": "IVIG", "hcpcs": "J1569", "ndc": "", "allowed": 100, "y": 2024},
            {"drug_common_name": "infliximab", "hcpcs": "J1745", "ndc": "", "allowed": 50, "y": 2025},
            {"drug_common_name": "infliximab", "hcpcs": "Q5103", "ndc": "59676031020", "allowed": 60, "y": 2025},
            {"drug_common_name": "IVIG", "hcpcs": "J1569", "ndc": "", "allowed": 100, "y": 2025},
        ])
        _cmk = _t["hcpcs"].isin({"J1745", "J1569"})
        _dmk = pd.Series(True, index=_t.index)
        _sh = _ti33.inclusion_share_by_period(_t, allowed=_t["allowed"], year=_t["y"],
                                              code_mask=_cmk, drug_mask=_dmk)
        ok &= _check("v33 trend: code-rule divergence drifts across periods",
                     abs(float(_sh.attrs["divergence_drift_pts"]) - 28.6) < 0.2)
        _dec = _ti33.trend_bend_decomposition(_t, allowed=_t["allowed"], year=_t["y"],
                                              code_mask=_cmk, drug_mask=_dmk)
        _ia = float(_dec.loc[_dec["molecule"] == "infliximab", "artifact_pts"].iloc[0])
        ok &= _check("v33 trend: artifact isolated (code view fakes infliximab decline)", _ia < -5)
        _fl = _ti33.flicker_events(_t, year=_t["y"])
        ok &= _check("v33 trend: biosimilar Q-code entry flagged as flicker event",
                     (_fl["event"] == "NEW_BIOSIMILAR_QCODE").any())

        # cross-source: partial presence flagged before any mix comparison
        _mx = _cs33.cross_source_matrix(
            {"MEDICAL": pd.DataFrame([{"drug_common_key": "infliximab", "drug_common_name": "infliximab", "allowed_amt": 100},
                                      {"drug_common_key": "ivig", "drug_common_name": "IVIG", "allowed_amt": 200}]),
             "RX": pd.DataFrame([{"drug_common_key": "ivig", "drug_common_name": "IVIG", "allowed_amt": 50}])},
            universe_keys={"ivig", "infliximab"})
        _pc = _cs33.scope_parity_check(_mx)
        ok &= _check("v33 cross-source: molecule present in one source only is flagged",
                     ("hazard" in _pc.columns) and (_pc["hazard"] == "PRESENT_IN_SOME_SOURCES_ONLY").any())

        # roster forensics: surviving-only roster inflates growth; leakage measured
        _pn = pd.DataFrame([
            {"billing_npi": "1111111116", "allowed": 100, "y": 2023},
            {"billing_npi": "2222222221", "allowed": 100, "y": 2023},
            {"billing_npi": "1111111116", "allowed": 20, "y": 2025},
            {"billing_npi": "2222222221", "allowed": 180, "y": 2025},
            {"billing_npi": "9999999995", "allowed": 30, "y": 2025},
        ])
        _mig = _rf33.legacy_npi_migration(_pn, allowed=_pn["allowed"], year=_pn["y"],
                                          full_roster={"1111111116", "2222222221"},
                                          surviving_roster={"2222222221"})
        _ag = _rf33.artificial_growth_test(_mig)
        ok &= _check("v33 roster: dropping legacy NPIs manufactures artificial growth",
                     "inflates" in _ag.attrs["verdict"])
        _lk = _rf33.entity_leakage_estimate(_pn, allowed=_pn["allowed"],
                                            full_roster={"1111111116", "2222222221"})
        ok &= _check("v33 roster: entity-leakage ceiling measured (off-roster dollars)",
                     abs(float(_lk.attrs["entity_leakage_dollars"]) - 30) < 0.01)

        # calibration: per-drug ratios span; thin capture flagged
        _cal = _cal33.komodo_ffs_calibration(
            {"IVIG": 14_000_000, "Stelara": 2_000_000, "rare": 100_000},
            {"IVIG": 100_000_000, "Stelara": 10_000_000, "rare": 5_000_000},
            blended_stated=0.141)
        ok &= _check("v33 calibration: per-drug FFS ratios computed against the census",
                     abs(float(_cal.attrs["blended_ratio"]) - (16_100_000 / 115_000_000)) < 0.001)
        ok &= _check("v33 calibration: thin-capture molecule flagged",
                     bool(_cal.loc[_cal["drug"] == "rare", "thin_capture_flag"].iloc[0]))

        # MA proxy + triangulation: convergent vs divergent verdicts
        _px = _mp33.ma_proxy_estimate({"IVIG": 10000}, {"IVIG": 80.0})
        ok &= _check("v33 MA proxy: encounters priced at unit rate (proxy-stamped)",
                     abs(float(_px.attrs["proxy_total"]) - 800_000) < 1)
        _tri = _mp33.ma_triangulation(ratio_estimate=354_600_000, proxy_estimate=310_000_000,
                                      management_estimate=340_000_000)
        ok &= _check("v33 MA proxy: three converging legs pass triangulation",
                     "CONVERGENT" in _tri.attrs["verdict"])
        _tri2 = _mp33.ma_triangulation(ratio_estimate=354_600_000, proxy_estimate=150_000_000)
        ok &= _check("v33 MA proxy: divergent legs block a point estimate",
                     "DIVERGENT" in _tri2.attrs["verdict"])

        # gross-up hardening: decomposition, mix parity, state stability
        _rd = _gu333.ratio_decomposition({"FFS": {"captured": 14.1, "universe": 100.0}},
                                         stated_blend=0.141)
        ok &= _check("v33 grossup: components reproduce the stated 14.1 blend",
                     "REPRODUCED" in _rd.attrs["verdict"])
        _mp_ = _gu333.mix_parity({"IVIG": 80, "Stelara": 20}, {"IVIG": 50, "Stelara": 50})
        ok &= _check("v33 grossup: skewed captured mix fails the lives-vs-dollars test",
                     "DIVERGES" in _mp_.attrs["verdict"])
        _ms = _gu333.medicaid_state_grossup({"TX": 0.05, "LA": 0.30, "NM": 0.10})
        ok &= _check("v33 grossup: wide state ratios scored LOW stability",
                     _ms.attrs["stability"] == "LOW")

        # market view: biosimilar adoption, floor sensitivity verdict
        _mvp = _cn33.assign_common_name(pd.DataFrame([
            {"drug_name": "Remicade", "hcpcs": "J1745", "ndc": "", "allowed": 60, "y": 2025},
            {"drug_name": "Inflectra", "hcpcs": "Q5103", "ndc": "", "allowed": 40, "y": 2025},
            {"drug_name": "Remicade", "hcpcs": "J1745", "ndc": "", "allowed": 100, "y": 2024},
        ]), crosswalk=_xw33, drug_ident=None, ref_dir=_ref)
        _ba = _mv33.biosimilar_adoption(_mvp, allowed=_mvp["allowed"], year=_mvp["y"])
        _b25 = _ba[_ba["year"] == 2025]
        ok &= _check("v33 market: biosimilar adoption share computed on the grouped molecule",
                     abs(float(_b25["biosimilar_share_pct"].iloc[0]) - 40.0) < 0.1)

        # NDC attribution: blank-NDC dollars spread across observed mix only
        _at = _xb333.ndc_attribution(
            _cn33.assign_common_name(pd.DataFrame([
                {"drug_name": "Inflectra", "hcpcs": "Q5103", "ndc": "59676-0310-20", "allowed": 6000},
                {"drug_name": "Inflectra", "hcpcs": "Q5103", "ndc": "59676-0310-01", "allowed": 4000},
                {"drug_name": "Remicade", "hcpcs": "J1745", "ndc": "", "allowed": 5000},
            ]), crosswalk=_xw33, drug_ident=None, ref_dir=_ref),
            _xb333.build_crosswalk(seed_ref_dir=_ref, use_example=True),
            allowed=pd.Series([6000, 4000, 5000]))
        _inf = _at[(_at["molecule_key"] == "infliximab") & (_at["_ndc_attributed"])]
        ok &= _check("v33 NDC attribution: blank-NDC dollars split across observed mix (60/40)",
                     len(_inf) == 2 and abs(float(_inf["attributed_dollars"].sum()) - 5000) < 1)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v33 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v34: anticipated-issue fixes. All offline.
    # ------------------------------------------------------------------ #
    print("\nv34 checks (anticipated issues):")
    try:
        from npi_recovery import dedup as _dd34
        from npi_recovery import unit_integrity as _ui34
        from npi_recovery import runout as _ro34
        from npi_recovery import growth_decomposition as _gd34
        from npi_recovery import payer_normalizer as _pn34
        from npi_recovery import concentration as _cc34
        from npi_recovery import modifier_economics as _me34
        from npi_recovery import restatement as _rs34
        _ref34 = "npi_recovery/reference"

        _sd = pd.DataFrame([
            {"billing_npi": "1", "hcpcs": "J1745", "ndc": "", "date": "2025-01-01",
             "payer": "BCBS", "allowed_amt": 100, "units": 1},
            {"billing_npi": "1", "hcpcs": "J1745", "ndc": "", "date": "2025-01-01",
             "payer": "BCBS", "allowed_amt": 100, "units": 1},
            {"billing_npi": "1", "hcpcs": "J1569", "ndc": "", "date": "2025-02-01",
             "payer": "UHC", "allowed_amt": 500, "units": 5},
            {"billing_npi": "1", "hcpcs": "J1569", "ndc": "", "date": "2025-02-01",
             "payer": "UHC", "allowed_amt": -500, "units": -5},
        ])
        _na = _dd34.netting_audit(_sd, allowed=_sd["allowed_amt"], units=_sd["units"])
        ok &= _check("v34 dedup: duplicate dollars and reversal pairs quantified",
                     _na.attrs["duplicate_dollars"] == 100
                     and _na.attrs["reversal_pair_dollars"] == 500)
        _net, _aud = _dd34.apply_netting(_sd)
        ok &= _check("v34 dedup: opt-in netting drops dup and both reversal legs",
                     len(_net) == 1 and float(pd.to_numeric(_net["allowed_amt"]).sum()) == 100)

        _ur = [{"hcpcs": "J1745", "allowed_amt": 10.0, "units": 1} for _ in range(9)]
        _ur.append({"hcpcs": "J1745", "allowed_amt": 1000.0, "units": 1})
        _uf = pd.DataFrame(_ur)
        _sc = _ui34.rate_outlier_screen(_uf, allowed=_uf["allowed_amt"], units=_uf["units"])
        ok &= _check("v34 units: 100x keying error flagged with overstatement 990",
                     _sc.attrs["n_flagged"] == 1
                     and abs(_sc.attrs["overstatement_dollars"] - 990) < 0.01)
        _ub = _ui34.unit_basis_check(_uf, allowed=_uf["allowed_amt"], units=_uf["units"],
                                     asp_limits={"J1745": 0.1})
        ok &= _check("v34 units: whole-code basis mismatch diagnosed vs limit",
                     "diagnosis" in _ub.columns)

        _months = pd.period_range("2024-01", "2025-06", freq="M")
        _rf = pd.DataFrame({"date": [m.to_timestamp() for m in _months],
                            "allowed_amt": [100] * 16 + [40, 10]})
        _cr = _ro34.completeness_report(_rf, allowed=_rf["allowed_amt"])
        ok &= _check("v34 runout: two-month incomplete tail flagged by cliff heuristic",
                     _cr.attrs.get("n_flagged") == 2)
        _rt = _ro34.restated_trend(_cr)
        ok &= _check("v34 runout: lag artifact reported raw vs mature",
                     "lag_artifact_pts" in _rt.attrs)

        _gs = pd.DataFrame([
            {"drug_common_name": "IVIG", "allowed": 1000, "units": 100, "y": 2024},
            {"drug_common_name": "IVIG", "allowed": 1200, "units": 100, "y": 2025},
            {"drug_common_name": "newdrug", "allowed": 300, "units": 10, "y": 2025},
        ])
        _pv = _gd34.price_volume_mix(_gs, allowed=_gs["allowed"], units=_gs["units"],
                                     year=_gs["y"])
        _t0 = _pv.iloc[0]
        ok &= _check("v34 growth: price/volume/entry decomposition ties to delta",
                     abs(_t0["price_effect"] - 200) < 0.01
                     and abs(_t0["entry_exit"] - 300) < 0.01
                     and abs(_t0["dollars_delta"] - 500) < 0.01)

        _ps = pd.DataFrame([
            {"payer": "BCBSTX", "allowed_amt": 500, "billing_npi": "1",
             "referring_npi": "9", "drug_common_name": "IVIG"},
            {"payer": "Blue Cross Blue Shield of Texas", "allowed_amt": 300,
             "billing_npi": "1", "referring_npi": "9", "drug_common_name": "IVIG"},
            {"payer": "UMR", "allowed_amt": 200, "billing_npi": "2",
             "referring_npi": "", "drug_common_name": "Stelara"},
        ])
        _au = _pn34.normalize_payers(_ps, allowed=_ps["allowed_amt"], ref_dir=_ref34)
        ok &= _check("v34 payer: BCBS TX variants roll to one parent (800)",
                     abs(float(_au.loc[_au["parent_org"].str.startswith("HCSC"),
                                       "allowed"].sum()) - 800) < 0.01)
        _ct = _cc34.concentration_table(_ps, allowed=_ps["allowed_amt"],
                                        payer_parent=_au.attrs["parent_series"])
        _ph = float(_ct.loc[_ct["lens"].str.startswith("payer"), "hhi"].iloc[0])
        ok &= _check("v34 concentration: payer HHI on parents in the highly band",
                     abs(_ph - 6800) < 15)
        _tm = _cc34.prescriber_taxonomy_mix(_ps, allowed=_ps["allowed_amt"],
                                            taxonomy_of={"9": "2084N0400X"})
        ok &= _check("v34 concentration: prescriber taxonomy mix attributes demand",
                     abs(float(_tm["share_of_attributed_pct"].iloc[0]) - 100.0) < 0.1)

        _mf = pd.DataFrame([{"allowed_amt": 1000, "mod_340b": True, "mod_wastage": False},
                            {"allowed_amt": 200, "mod_340b": False, "mod_wastage": True},
                            {"allowed_amt": 800, "mod_340b": False, "mod_wastage": False}])
        _me = _me34.modifier_economics(_mf, allowed=_mf["allowed_amt"])
        ok &= _check("v34 modifiers: 340B and wastage flags priced in dollars",
                     _me.attrs["dollars_340b"] == 1000
                     and _me.attrs["dollars_wastage_flagged"] == 200)

        _cur = pd.DataFrame([{"drug_common_name": "IVIG", "allowed_amt": 110},
                             {"drug_common_name": "Stelara", "allowed_amt": 50}])
        _rd = _rs34.restatement_diff({"IVIG": 100, "Stelara": 50, "gone": 30},
                                     _cur, allowed=_cur["allowed_amt"])
        ok &= _check("v34 restatement: run-over-run mover flagged, drop detected",
                     _rd.attrs["n_restated"] == 1
                     and (_rd["status"].str.startswith("DROPPED")).any())
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v34 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v35: deep-clean process. All offline.
    # ------------------------------------------------------------------ #
    print("\nv35 checks (deep clean):")
    try:
        import numpy as _np35
        from npi_recovery import text_hygiene as _th35
        from npi_recovery import field_validators as _fv35
        from npi_recovery import row_consistency as _rc35
        from npi_recovery import distribution_screens as _ds35
        from npi_recovery import imputation_options as _io35
        from npi_recovery import clean_pipeline as _cp35

        _dirty = pd.DataFrame({
            "billing_npi": ["1.234567893E9", "1234567893"],
            "hcpcs": ["J1745", "J1569"],
            "ndc": ["59676-310-20", "0069080901"],
            "date": ["45123", "2025-03-01"],
            "allowed_amt": ["(500)", "$1,000"],
            "units": [5, 10],
            "payer": ["BCBS\u00a0TX", "N/A"]})
        _scan = _th35.scan_text_hygiene(_dirty)
        ok &= _check("v35 hygiene: sci-notation, NBSP, sentinel findings surface",
                     {"SCIENTIFIC_NOTATION_ID", "NBSP_IN_VALUE",
                      "SENTINEL_AS_VALUE"} <= set(_scan.get("finding", [])))
        ok &= _check("v35 npi: Luhn check separates valid from keying error",
                     _fv35.luhn_npi_valid("1234567893")
                     and not _fv35.luhn_npi_valid("1234567890"))
        ok &= _check("v35 ndc: 4-4-2 / 5-3-2 / 5-4-1 pad the correct segment",
                     _fv35.normalize_ndc11("0069-0809-01")[0] == "00069080901"
                     and _fv35.normalize_ndc11("59676-310-20")[0] == "59676031020"
                     and _fv35.normalize_ndc11("00074-4339-2")[0] == "00074433902")
        ok &= _check("v35 ndc: unhyphenated 10-digit flagged ambiguous, never padded",
                     _fv35.normalize_ndc11("0069080901")[1].startswith("AMBIGUOUS"))
        _d35, _st35 = _fv35.parse_date_multi("45123")
        ok &= _check("v35 dates: Excel serial parsed from 1899-12-30 origin",
                     str(_d35.date()) == "2023-07-16" and _st35 == "EXCEL_SERIAL")
        ok &= _check("v35 money: accounting parentheses negative parsed",
                     _fv35.parse_money("(1,234.50)")[0] == -1234.50)

        _rcf = pd.DataFrame({"date": ["2025-02-01"], "paid_date": ["2025-01-01"],
                             "allowed_amt": [100], "units": [0],
                             "ndc": [""], "hcpcs": ["J1745"]})
        _rr = _rc35.run_row_consistency(_rcf)
        ok &= _check("v35 rows: impossible sequence and zero-unit dollars priced",
                     {"ROW-SVC-PAID", "ROW-0U-POSD"} <= set(_rr.get("rule_id", [])))

        _rng = _np35.random.RandomState(7)
        _nat = pd.Series(10 ** _rng.uniform(1, 5, size=600))
        _flat = pd.Series([d * 100 + 7 for d in range(1, 10) for _ in range(60)])
        ok &= _check("v35 benford: natural conforms, uniform first digits deviate",
                     "CONFORMS" in _ds35.benford_first_digit(_nat).attrs["verdict"]
                     and "DEVIATES" in _ds35.benford_first_digit(_flat).attrs["verdict"])
        _rp = _ds35.rounding_pathology(
            pd.DataFrame({"payer": ["A", "A", "B", "B"]}),
            allowed=pd.Series([100.00, 200.00, 123.45, 167.89]), group_col="payer")
        ok &= _check("v35 rounding: schedule-like payer at 100 pct whole dollars",
                     float(_rp.loc[_rp["group"] == "A",
                                   "whole_dollar_share_pct"].iloc[0]) == 100.0)

        _imp = pd.DataFrame({
            "hcpcs": ["J1745"] * 6 + ["J1569"] * 6,
            "allowed_amt": [100] * 5 + [300] + [550] * 6,
            "units": [10] * 5 + [_np35.nan] + [50] * 5 + [_np35.nan],
            "state": ["TX"] * 5 + [""] + ["LA"] * 5 + [""],
            "zip3": ["770"] * 6 + ["701"] * 6,
            "billing_npi": ["1"] * 6 + ["2"] * 6,
            "drug_name": ["Remicade"] * 5 + [""] + ["GAMMAGARD"] * 5 + [""],
            "ndc": [""] * 12})
        _cmp = _io35.compare_strategies(_imp, "units").set_index("strategy")
        ok &= _check("v35 impute: options compared with agreement and disagreement dollars",
                     float(_cmp.loc["code_median vs rate_implied", "fill_rate_pct"]) == 50.0
                     and float(_cmp.loc["code_median vs rate_implied",
                                        "dollars_on_filled"]) == 300.0)
        _ap, _ = _io35.apply_strategy(_imp, "units", "rate_implied")
        ok &= _check("v35 impute: applied fill stamps method and keeps original",
                     "units_original" in _ap.columns
                     and int(_ap["units_imputed_method"].eq("rate_implied").sum()) == 2)

        _clean, _led, _fnd = _cp35.run_cleaning(_dirty, now="2026-07-01")
        ok &= _check("v35 pipeline: NPI restored, NDC padded, date ISO, money numeric",
                     _clean["billing_npi"].iloc[0] == "1234567893"
                     and _clean["ndc"].iloc[0] == "59676031020"
                     and _clean["date"].iloc[0] == "2023-07-16"
                     and float(_clean["allowed_amt"].iloc[0]) == -500.0)
        _recon = _led.attrs["reconciliation"]
        ok &= _check("v35 pipeline: rows conserved and dollar delta declared",
                     _recon["rows_conserved"] and _recon["dollar_delta"] == 500.0)
        _idem = _cp35.idempotency_check(_dirty, now="2026-07-01")
        ok &= _check("v35 pipeline: cleaning is idempotent", _idem.attrs["idempotent"])
        _scd = _cp35.dq_scorecard(_clean, now="2026-07-01")
        ok &= _check("v35 pipeline: DQ scorecard renders a bounded overall index",
                     0 <= _scd.attrs["overall_index"] <= 100)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v35 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v36: preflight resilience + run-report output layer. All offline.
    # ------------------------------------------------------------------ #
    print("\nv36 checks (resilience and output):")
    try:
        from npi_recovery import preflight as _pf36
        from npi_recovery import run_report as _rr36
        _raw36 = ("Option Care Extract\nGenerated 2026-06-30 by Mar\xeda\n"
                  "Billing NPI|HCPCS Code|Allowed Amount|Units|Service Date\n"
                  "1234567893|J1745|1,000.00|10|2025-01-01\n"
                  "9876543210|J1569|500.00|5\n"
                  "1111111116|J3357|250.00|2|2025-02-01\n")
        _p36 = "/tmp/selftest_messy36.csv"
        open(_p36, "wb").write(_raw36.encode("cp1252"))
        _df36, _facts36 = _pf36.robust_read(_p36)
        _fmap = dict(zip(_facts36["decision"], _facts36["value"]))
        ok &= _check("v36 preflight: cp1252 decoded, pipe sniffed, 2 junk rows skipped",
                     _df36 is not None and _fmap.get("encoding") == "cp1252"
                     and "'|'" in str(_fmap.get("delimiter"))
                     and int(_fmap.get("header_row_index")) == 2)
        ok &= _check("v36 preflight: ragged row detected and read survives it",
                     "ragged_rows_detected" in _fmap and len(_df36) >= 2)
        _bad36, _f36b = _pf36.robust_read("/tmp/definitely_missing_36.csv")
        ok &= _check("v36 preflight: missing file returns facts, never raises",
                     _bad36 is None and not _f36b.attrs["ok"])
        _cd36 = _pf36.column_diagnosis(
            ["Billing NPI #", "HCPCS Code", "Allowed Amount", "Qty of Service"]
        ).set_index("canonical")
        ok &= _check("v36 preflight: near-miss headers diagnosed, absent named",
                     _cd36.loc["units", "verdict"] != "ABSENT"
                     and _cd36.loc["ndc", "verdict"] == "ABSENT")
        _cr36 = _pf36.coercion_report(pd.DataFrame(
            {"allowed_amt": ["100", "(500)", "1,000", "abc", ""]}))
        ok &= _check("v36 preflight: coercion casualties priced with examples",
                     int(_cr36["coercion_casualties"].iloc[0]) == 3)

        def _T36(note=None, **at):
            d = (pd.DataFrame({"note": [note]}) if note
                 else pd.DataFrame({"x": [1]}))
            for k, v in at.items():
                d.attrs[k] = v
            return d
        _tabs36 = {
            "MA_Triangulation": _T36(verdict="DIVERGENT: do not chart a point estimate"),
            "Benford_Screen": _T36(note="CONFORMS (chi-square 9.1 <= 15.507)"),
            "Mix_Parity": _T36(verdict="MIX DIVERGES (weighted deviation 0.36)"),
            "Broken_Stage": _T36(note="roster forensics skipped: TypeError: bad arg"),
        }
        _dig36 = _rr36.findings_digest(_tabs36)
        _sev36 = dict(zip(_dig36["tab"], _dig36["severity"]))
        ok &= _check("v36 report: DIVERGENT->CRITICAL, DIVERGES->WARN, CONFORMS->PASS, "
                     "exception->ERROR",
                     _sev36["MA_Triangulation"] == "CRITICAL"
                     and _sev36["Mix_Parity"] == "WARN"
                     and _sev36["Benford_Screen"] == "PASS"
                     and _sev36["Broken_Stage"] == "ERROR")
        _man36 = _rr36.run_manifest(_tabs36, version="36.0.0", n_rows=10)
        ok &= _check("v36 report: manifest DEGRADED when a stage errored",
                     "DEGRADED" in _man36.attrs["status"])
        _out36 = _rr36.attach_run_report(dict(_tabs36), version="36.0.0")
        ok &= _check("v36 report: exec summary, digest, errors, manifest, TOC attached",
                     {"Executive_Summary", "Findings_Digest", "Errors_Log",
                      "Run_Manifest", "Table_of_Contents"} <= set(_out36)
                     and len(_out36["Table_of_Contents"]) == len(_out36) - 1)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v36 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v38: deep-fix regressions from the audit pass. All offline.
    # ------------------------------------------------------------------ #
    print("\nv38 checks (audit fixes):")
    try:
        from npi_recovery import growth_decomposition as _gd38
        from npi_recovery import roster_forensics as _rf38
        from npi_recovery import runout as _ro38
        from npi_recovery import preflight as _pf38

        _g38 = pd.DataFrame([
            {"drug_common_name": "IVIG", "allowed": 1000, "units": 100, "y": 2024},
            {"drug_common_name": "IVIG", "allowed": 1200, "units": 100, "y": 2025},
            {"drug_common_name": "mystery", "allowed": 400, "units": 0, "y": 2024},
            {"drug_common_name": "mystery", "allowed": 700, "units": 0, "y": 2025}])
        _pv38 = _gd38.price_volume_mix(_g38, allowed=_g38["allowed"],
                                       units=_g38["units"], year=_g38["y"])
        _t38 = _pv38.iloc[0]
        ok &= _check("v38 growth: unit-missing molecule stays in the tie-out "
                     "(delta equals last minus first)",
                     abs(_t38["dollars_delta"]
                         - (_t38["dollars_last"] - _t38["dollars_first"])) < 0.01
                     and (_pv38["class"].str.startswith("unpriceable")).any())

        _mig38 = pd.DataFrame([
            {"year": 2023, "full_roster_allowed": 100.0, "surviving_only_allowed": 100.0,
             "legacy_npi_allowed": 0.0, "legacy_share_pct": 0.0},
            {"year": 2025, "full_roster_allowed": 121.0, "surviving_only_allowed": 121.0,
             "legacy_npi_allowed": 0.0, "legacy_share_pct": 0.0}])
        _ag38 = _rf38.artificial_growth_test(_mig38)
        _full38 = float(_ag38.loc[_ag38["metric"].str.contains("full"),
                                  "value_pct"].iloc[0])
        ok &= _check("v38 roster: CAGR annualized over the true year span "
                     "(2023 to 2025 is two years: 10.0, not 21.0)",
                     abs(_full38 - 10.0) < 0.05)

        _rp38 = pd.DataFrame({
            "date": ["2024-10-01"] * 3 + ["2024-11-01"] * 3 + ["2024-12-01"] * 3
                    + ["2025-03-01"],
            "paid_date": ["2024-12-15", "2024-12-20", "2025-01-05", "2025-01-15",
                          "2025-01-20", "2025-02-05", "2025-02-15", "2025-02-20",
                          "2025-03-05", "2025-03-20"],
            "allowed_amt": [10] * 10})
        _cr38 = _ro38.completeness_report(_rp38, allowed=_rp38["allowed_amt"])
        _last38 = _cr38.iloc[-1]
        ok &= _check("v38 runout: month younger than the first observed lag is "
                     "IMMATURE, never restated by division",
                     _last38["status"].startswith("IMMATURE")
                     and pd.isna(_last38["restated_allowed"]))

        open("/tmp/selftest_rag38.csv", "w").write(
            "npi,hcpcs,drug_name,allowed_amt\n"
            "1234567893,J1745,Remicade,100\n"
            "1111111116,J3590,drug, with comma,250\n"
            "2222222221,J1569,GAMMAGARD\n")
        _df38, _f38 = _pf38.robust_read("/tmp/selftest_rag38.csv")
        _fm38 = dict(zip(_f38["decision"], _f38["value"]))
        ok &= _check("v38 preflight: ragged rows padded or merged, zero rows lost",
                     _df38 is not None and len(_df38) == 3
                     and _fm38.get("rows_lost_to_raggedness") == 0
                     and _fm38.get("rows_padded_short") == 1
                     and _fm38.get("rows_overflow_merged_last_field") == 1)
        from npi_recovery import dedup as _dd38
        _sd38 = pd.DataFrame({"billing_npi": ["1", "1"], "hcpcs": ["J1745", "J1745"],
                              "ndc": ["", ""], "date": ["2025-01-01", "2025-01-01"],
                              "payer": ["B", "B"], "allowed_amt": [100, 100],
                              "units": [1, 1]}, index=[0, 0])
        _net38, _ = _dd38.apply_netting(_sd38)
        ok &= _check("v38 dedup: duplicate index labels handled positionally",
                     len(_net38) == 1)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v38 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------------ #
    # v39: jurisdiction-aware SAD + missing resolver. Offline vs the shipped
    # CMS snapshot.
    # ------------------------------------------------------------------ #
    print("\nv39 checks (CMS jurisdiction SAD + missing resolver):")
    try:
        from npi_recovery import sad_jurisdiction as _sad39
        from npi_recovery import missing_resolver as _mr39
        _ref39 = "npi_recovery/reference"
        _mac = _sad39.load_mac_map(_ref39)
        _idx = _sad39.build_sad_index(_ref39)
        ok &= _check("v39 SAD: MAC map resolves TX->Novitas and CA->Noridian JE",
                     _mac.get("TX") == "NOVITAS" and _mac.get("CA") == "NORIDIAN_JE")
        ok &= _check("v39 SAD: shipped snapshot carries etanercept J1438",
                     "J1438" in _idx)
        _v1, _ = _sad39.classify_row("J1745", "TX", "", mac_map=_mac, sad_index=_idx)
        _v2, _ = _sad39.classify_row("J1438", "TX", "", mac_map=_mac, sad_index=_idx)
        ok &= _check("v39 SAD: infliximab IV eligible, etanercept SAD-excluded",
                     _v1 == "PART_B_ELIGIBLE" and _v2 == "SAD_EXCLUDED")
        _ja, _ = _sad39.classify_row("J0129", "TX", "JA", mac_map=_mac, sad_index=_idx)
        _jb, _ = _sad39.classify_row("J0129", "TX", "JB", mac_map=_mac, sad_index=_idx)
        _amb, _ = _sad39.classify_row("J0129", "TX", "", mac_map=_mac, sad_index=_idx)
        ok &= _check("v39 SAD: JA->eligible, JB->SC excluded, none->route ambiguous",
                     _ja == "PART_B_ELIGIBLE" and _jb == "SAD_EXCLUDED_SC"
                     and _amb == "ROUTE_AMBIGUOUS")
        _vtx, _ = _sad39.classify_row("J3262", "TX", "", mac_map=_mac, sad_index=_idx)
        _vny, _ = _sad39.classify_row("J3262", "NY", "", mac_map=_mac, sad_index=_idx)
        ok &= _check("v39 SAD: same code differs by MAC (tocilizumab SC excluded "
                     "in TX, eligible in NY)",
                     _vtx in ("SAD_EXCLUDED", "SAD_EXCLUDED_SC")
                     and _vny == "PART_B_ELIGIBLE")
        _vzz, _ = _sad39.classify_row("J1438", "ZZ", "", mac_map=_mac, sad_index=_idx)
        ok &= _check("v39 SAD: unmapped state -> UNKNOWN_JURISDICTION",
                     _vzz == "UNKNOWN_JURISDICTION")

        _sdf = pd.DataFrame([
            {"hcpcs": "J1745", "state": "TX", "allowed_amt": 100000, "modifier": ""},
            {"hcpcs": "J0129", "state": "TX", "allowed_amt": 30000, "modifier": ""}])
        _cl = _sad39.classify_frame(_sdf, ref_dir=_ref39, allowed=_sdf["allowed_amt"])
        ok &= _check("v39 SAD: frame prices route-ambiguous exposure",
                     _cl.attrs.get("ambiguous_dollars") == 30000)

        def _fake(**kw):
            return {"items": [{"code_id": "J1438", "drug_brand_name": "Enbrel",
                               "contractor_name_type": "Novitas",
                               "effective_date": "2026-06-15"}],
                    "next_page_token": None}
        ok &= _check("v39 SAD: live refresh hook normalizes a pull, and no-callable "
                     "returns an honest note",
                     _sad39.refresh_from_cms(_fake)["hcpcs"].iloc[0] == "J1438"
                     and list(_sad39.refresh_from_cms(None).columns) == ["note"])

        _mdf = pd.DataFrame([
            {"billing_npi": "", "referring_npi": "9", "hcpcs": "J0129", "ndc": "",
             "state": "", "drug_name": "", "allowed_amt": 30000, "units": float("nan"),
             "payer": "BCBS", "modifier": ""}])
        _inv = _mr39.gap_inventory(_mdf, allowed=_mdf["allowed_amt"], ref_dir=_ref39)
        _gaps = set(_inv.get("gap", []))
        ok &= _check("v39 resolver: inventories state/drug/units/NDC/NPI and SAD "
                     "ambiguity in one ranked table",
                     {"state blank", "drug name blank",
                      "units missing on paid line"} <= _gaps
                     and any("SAD route ambiguous" in g for g in _gaps))
        _plan = _mr39.resolution_plan(_mdf, allowed=_mdf["allowed_amt"], ref_dir=_ref39)
        ok &= _check("v39 resolver: plan gives a concrete action per gap",
                     "action" in _plan.columns
                     and (_plan["action"].str.len() > 10).all())
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v39 checks run ({type(_e).__name__}: {_e})", False)

    # v40: more CMS connections + closing gaps with the connectors already
    # available. Real SAD snapshot expanded from a live CMS pull, refresh now a
    # true drop-in, MAC roster split by jurisdiction, connector adapter layer.
    print("\nv40 checks (live CMS snapshot + connector adapters):")
    try:
        from npi_recovery import sad_jurisdiction as _s40
        from npi_recovery import source_adapters as _c40
        _ref40 = "npi_recovery/reference"

        _idx40 = _s40.build_sad_index(_ref40)
        ok &= _check("v40 SAD: classifier snapshot expanded past v39 (>=15 codes)",
                     len(_idx40) >= 15)

        _full = _s40.load_sad_full(_ref40)
        ok &= _check("v40 SAD: source-audit file ships and every row is a "
                     "click-through CMS article URL",
                     len(_full) > 0
                     and _full["url"].str.contains("cms.gov").all())

        _mac40 = _s40.load_mac_map(_ref40)
        _ny, _ = _s40.classify_row("J9216", "NY", "", mac_map=_mac40, sad_index=_idx40)
        ok &= _check("v40 SAD: live pull corrects J9216 (Actimmune) to SAD-excluded "
                     "in NY, which v39's 2-MAC coverage got wrong",
                     _ny == "SAD_EXCLUDED")

        # IVIG confirm/deny: immune globulin is provider-administered, not on the
        # SAD list, so it must classify Part B eligible (verified live 2026-07-02).
        _ig, _ = _s40.classify_row("J1569", "TX", "", mac_map=_mac40, sad_index=_idx40)
        ok &= _check("v40 SAD: IVIG (J1569) not on the SAD list -> PART_B_ELIGIBLE",
                     _ig == "PART_B_ELIGIBLE")

        # refresh_from_cms now emits the classifier seed shape (drop-in), and is
        # still backward compatible with the v39 single-item fake.
        def _fake40(**kw):
            if kw.get("page_token"):
                return {"items": [], "next_page_token": None}
            return {"items": [
                {"code_id": "J9216", "drug_brand_name": "Actimmune",
                 "document_display_id": "A53021",
                 "contractor_name_type": "National Government Services, Inc.",
                 "effective_date": "2026-06-13", "url": "https://www.cms.gov/x"},
                {"code_id": "J9216", "drug_brand_name": "Actimmune",
                 "document_display_id": "A53033",
                 "contractor_name_type": "Noridian Healthcare Solutions, LLC",
                 "effective_date": "2026-06-13", "url": "https://www.cms.gov/x"}],
                "next_page_token": None}
        _seed = _s40.refresh_from_cms(_fake40)
        ok &= _check("v40 SAD: refresh normalizes a live pull to the seed shape "
                     "(hcpcs+macs columns) and unions MACs via article id",
                     list(_seed.columns)[:5] == ["hcpcs", "drug_brand", "molecule",
                                                 "route_note", "macs"]
                     and _seed.iloc[0]["macs"] == "NGS,NORIDIAN_JF")
        _raw = _s40.refresh_from_cms(_fake40, raw=True)
        ok &= _check("v40 SAD: refresh raw mode returns source rows; no-callable "
                     "still returns an honest note",
                     len(_raw) == 2
                     and list(_s40.refresh_from_cms(None).columns) == ["note"])

        ok &= _check("v40 SAD: article id disambiguates Noridian JE/JF and route "
                     "text parses JA/JB, subcutaneous, and NOC",
                     _s40._item_to_mac({"document_display_id": "A53032"}) == "NORIDIAN_JE"
                     and "modifier decides" in _s40.route_note_from_text("use JA or JB")
                     and "SUBCUTANEOUS" in _s40.route_note_from_text("subcutaneous self-injected")
                     and "NOC" in _s40.route_note_from_text("UNCLASSIFIED BIOLOGICS"))

        ok &= _check("v40 MAC: roster split by jurisdiction (J-codes) maps both "
                     "Novitas jurisdictions (TX via JH, PA via JL)",
                     _mac40.get("TX") == "NOVITAS" and _mac40.get("PA") == "NOVITAS"
                     and _mac40.get("WA") == "NORIDIAN_JF")

        # connector adapters, against the real tool response shapes -----------
        def _cms_tool(**kw):
            if kw.get("page_token"):
                return {"items": [], "next_page_token": None}
            return {"items": [{"code_id": "J1438", "drug_brand_name": "Enbrel",
                               "document_display_id": "A53127",
                               "contractor_name_type": "Novitas Solutions, Inc.",
                               "effective_date": "2026-06-15", "url": "https://cms.gov/x"}],
                    "next_page_token": None}
        _fetch = _c40.cms_sad_fetcher(_cms_tool, keyword="infusion")
        _drv = _s40.refresh_from_cms(_fetch)
        ok &= _check("v40 connectors: cms_sad_fetcher drives refresh end-to-end; "
                     "None tool -> None fetcher",
                     _drv.iloc[0]["hcpcs"] == "J1438" and _drv.iloc[0]["macs"] == "NOVITAS"
                     and _c40.cms_sad_fetcher(None) is None)

        _row = _c40.nppes_record_to_index_row(
            {"found": True, "record": {"npi": "1003914151", "enumeration_type": "Individual",
             "name": "TIEN WONG", "basic": {"status": "A", "credential": "M.D."},
             "primary_taxonomy": {"code": "207W00000X", "desc": "Ophthalmology",
                                  "state": "TX", "license": "H3961"},
             "primary_practice_address": {"state": "TX"}}})
        ok &= _check("v40 connectors: NPPES record maps to the enrichment index row "
                     "(name, taxonomy, status); not-found is clean",
                     _row["taxonomy_code"] == "207W00000X" and _row["status"] == "A"
                     and _c40.nppes_record_to_index_row({"found": False}) == {"found": False})

        def _icd_tool(code=None, code_type=None, **kw):
            return ({"found": True, "code": {"code": "G70.00",
                     "long_description": "Myasthenia gravis without (acute) exacerbation",
                     "valid_for_hipaa_transactions": True, "category": "G70"}}
                    if code == "G70.00" else {"found": False})
        _ann = _c40.icd10_annotator(lookup_callable=_icd_tool)
        _adf = _c40.annotate_referral_dx(["G70.00", "ZZ.999"], _ann)
        ok &= _check("v40 connectors: ICD-10 annotator validates a real dx and "
                     "rejects a junk code",
                     bool(_adf.set_index("code").loc["G70.00", "valid"])
                     and not bool(_adf.set_index("code").loc["ZZ.999", "valid"]))

        _reg = _c40.adapter_registry_frame()
        ok &= _check("v40 connectors: gap->connector map covers CMS Coverage, NPI "
                     "Registry, and ICD-10",
                     set(_reg["connector"]) >= {"CMS Coverage", "NPI Registry", "ICD-10"})
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v40 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v42 ----
    # Selectable fixes: registry integrity, fixability logic, the six coding-edit
    # screens on fixtures, closed-claims, run_selected edges, focused workbook.
    try:
        import os as _os
        import pandas as _pd
        from npi_recovery import registry as _R
        from npi_recovery import coding_edits as _CE
        from npi_recovery import closed_claims as _CC
        from npi_recovery import focused_report as _FR
        _refdir = _os.path.join(_os.path.dirname(_R.__file__), "reference")

        # registry catalog integrity: unique keys, valid groups/kahn, callable run
        _keys = [m.key for m in _R.REGISTRY]
        ok &= _check("v42 registry: fix keys are unique",
                     len(_keys) == len(set(_keys)))
        ok &= _check("v42 registry: every module has a valid group and Kahn category",
                     all(m.group in _R.GROUP_ORDER and
                         m.kahn in (_R.CONFORMANCE, _R.COMPLETENESS, _R.PLAUSIBILITY)
                         for m in _R.REGISTRY))
        ok &= _check("v42 registry: every module exposes a callable run()",
                     all(callable(m.run) for m in _R.REGISTRY))
        ok &= _check("v42 registry: list_modules covers all registered fixes",
                     len(_R.list_modules()) == len(_R.REGISTRY) >= 11)

        # fixability support logic
        _msex = _R.REGISTRY_BY_KEY["age_sex_conflict"]
        ok &= _check("v42 fixability: required-field fix supported when field present",
                     _msex.support({"diagnosis", "patient_age"}) == "supported")
        ok &= _check("v42 fixability: required-field fix unsupported when field absent",
                     _msex.support({"allowed_amt"}) == "unsupported")
        _mmue = _R.REGISTRY_BY_KEY["mue_units"]
        ok &= _check("v42 fixability: multi-required fix partial with only one field",
                     _mmue.support({"hcpcs"}) == "partial")
        _mclose = _R.REGISTRY_BY_KEY["closed_claims"]
        ok &= _check("v42 fixability: zero-required fix runs (supported) when optional present",
                     _mclose.support({"claim_status"}) == "supported")

        # ---- fixtures for the coding-edit screens ----
        _fix = _pd.DataFrame({
            "billing_npi": ["1003914151", "1184649626", "", "1999999984"],
            "hcpcs": ["J1745", "J9312", "99213", "J3357"],
            "units": [200, 10, 1, 5],
            "modifiers": ["", "JZ", "", ""],
            "diagnosis": ["G7000", "O80", "ZZ999", "P0730"],
            "patient_age": [55, 40, 7, 30],
            "patient_sex": ["M", "M", "F", "M"],
            "date": ["2025-06-15", "2025-06-15", "2025-11-20", "2025-06-15"],
            "allowed_amt": [1000.0, -50.0, 0.0, 500.0],
            "payer": ["BCBS of Texas", "UnitedHealthcare", "Aetna", "Cigna"],
            "claim_status": ["Closed", "Open", "Adjudicated", "Closed"],
        })
        _map = {c: c for c in _fix.columns}

        # MUE: J1745 at 200 units exceeds the real cap (150)
        _rmue = _CE.mue_screen(_fix, ref_dir=_refdir, mapping=_map)
        ok &= _check("v42 MUE: flags J1745 units above the real Medicare cap",
                     len(_rmue) >= 1 and (_rmue["hcpcs"] == "J1745").any())

        # ICD DOS validity: ZZ999 is junk -> flagged
        _ricd = _CE.icd10_dos_validity(_fix, ref_dir=_refdir, mapping=_map)
        ok &= _check("v42 ICD DOS: flags a junk diagnosis code as invalid",
                     (_ricd["diagnosis"] == "ZZ999").any())

        # ICD DOS validity on a real FY-retired code (B880 valid FY2025, not FY2026)
        _fyfix = _pd.DataFrame({"diagnosis": ["B880"], "date": ["2025-11-20"]})
        _rfy = _CE.icd10_dos_validity(_fyfix, ref_dir=_refdir,
                                      mapping={"diagnosis": "diagnosis", "date": "date"})
        ok &= _check("v42 ICD DOS: flags a real FY2025 code retired in FY2026 on a FY2026 date",
                     len(_rfy) == 1)

        # Age/sex: maternity (O80) on a male patient -> flagged
        _rage = _CE.age_sex_conflicts(_fix, ref_dir=_refdir, mapping=_map)
        ok &= _check("v42 age/sex: flags a maternity diagnosis on a male patient",
                     (_rage["diagnosis"] == "O80").any() if len(_rage) else False)

        # JW/JZ: J1745 single-dose with no modifier -> flagged; J9312 has JZ -> not
        _rjw = _CE.jw_jz_wastage(_fix, ref_dir=_refdir, mapping=_map)
        ok &= _check("v42 JW/JZ: flags single-dose line missing modifier, clears JZ line",
                     (_rjw["hcpcs"] == "J1745").any() and not (_rjw["hcpcs"] == "J9312").any())

        # Deactivated NPI screen runs and returns a frame (seed is a sample)
        _rde = _CE.deactivated_npi_screen(_fix, ref_dir=_refdir, mapping=_map)
        ok &= _check("v42 deactivated-NPI: screen runs and returns a frame",
                     isinstance(_rde, _pd.DataFrame))

        # Closed-claims: keeps closed + adjudicated, drops open
        _rcc = _CC.closed_claims_view(_fix, {"mapping": _map})
        _cf = _rcc.attrs.get("closed_frame")
        ok &= _check("v42 closed-claims: keeps adjudicated rows, drops open",
                     _cf is not None and len(_cf) == 3)

        # run_selected: unknown + unsupported keys reported, not crashed
        _res = _R.run_selected(_fix, ["mue_units", "not_a_real_fix"],
                               {"ref_dir": _refdir, "mapping": _map})
        ok &= _check("v42 run_selected: unknown key reported without crashing",
                     "not_a_real_fix" in _res and "note" in _res["not_a_real_fix"].columns)

        # focused_report writes a valid workbook with the manifest sheet
        _man = _R.fixability(_fix, _map)
        _tmp = _os.path.join("/tmp", "v42_selftest_focus.xlsx")
        _FR.write_focused(_tmp, _fix, _man, _res, "selftest_fixture.xlsx")
        from openpyxl import load_workbook as _lwb
        _wb = _lwb(_tmp)
        ok &= _check("v42 focused_report: writes About + Fixability_Manifest sheets",
                     "About" in _wb.sheetnames and "Fixability_Manifest" in _wb.sheetnames)
        _os.remove(_tmp)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v42 checks run ({type(_e).__name__}: {_e})", False)

    # ------------------------------------------------------------- v42.1 ----
    # Regression tests for the fixability overstatement found in external review:
    # schema.standardize() manufactures all-NA canonical columns, and v42.0
    # counted them as present. These tests pin the corrected behavior on a SPARSE
    # input (no diagnosis, no modifiers, no age/sex, no claim status delivered),
    # which is exactly the class of input the v42.0 fixtures missed.
    try:
        import os as _os
        import pandas as _pd
        from npi_recovery import registry as _R
        from npi_recovery import coding_edits as _CE
        from npi_recovery import closed_claims as _CC
        from npi_recovery import schema as _S
        _refdir = _os.path.join(_os.path.dirname(_R.__file__), "reference")

        # a sparse raw file: NPI, code, units, dollars, payer, date ONLY
        _sparse_raw = _pd.DataFrame({
            "BillingProviderNPI": ["1003914151", "1184649626", "", "1234567893"],
            "ProcedureCode": ["J1745", "J9312", "99213", "J3357"],
            "Qty": [200, 10, 1, 5],
            "AllowedAmt": [1000.0, -50.0, 0.0, 500.0],
            "PlanName": ["BCBS of Texas", "UnitedHealthcare", "Aetna", "Cigna"],
            "ServiceDate": ["2025-06-15", "2025-06-15", "2025-11-20", "2025-06-15"],
        })
        _smap, _srep = _S.detect_columns(_sparse_raw)
        _sstd = _S.standardize(_sparse_raw, _smap)
        _smapping = {k: k for k, v in _smap.items() if v is not None and k in _sstd.columns}

        # standardize DOES manufacture the all-NA canonicals (the hazard is real)
        ok &= _check("v42.1 hazard exists: standardize manufactures an all-NA "
                     "diagnosis column on a sparse file",
                     "diagnosis" in _sstd.columns and _sstd["diagnosis"].isna().all())

        # fixability must NOT count manufactured columns
        _sman = _R.fixability(_sstd, _smapping)
        _stat = dict(zip(_sman["key"], _sman["status"]))
        ok &= _check("v42.1 fixability: ICD DOS validity unsupported when no "
                     "diagnosis was delivered",
                     _stat["icd_dos_validity"] == "unsupported")
        ok &= _check("v42.1 fixability: age/sex unsupported when no diagnosis "
                     "was delivered",
                     _stat["age_sex_conflict"] == "unsupported")
        ok &= _check("v42.1 fixability: JW/JZ demoted to partial when no modifier "
                     "field was delivered (verdict evidence missing)",
                     _stat["jw_jz_wastage"] == "partial")
        ok &= _check("v42.1 fixability: closed-claims partial when no status "
                     "field was delivered",
                     _stat["closed_claims"] == "partial")
        ok &= _check("v42.1 fixability: MUE still supported (hcpcs + units delivered)",
                     _stat["mue_units"] == "supported")

        # field coverage distinguishes delivered from manufactured
        _cov = _R.field_coverage(_sstd, _smapping)
        _delivered = set(_cov[_cov["delivered"]]["field"])
        ok &= _check("v42.1 coverage: manufactured all-NA fields report as not "
                     "delivered",
                     "diagnosis" not in _delivered and "hcpcs" in _delivered)

        # ICD screen: no false invalids on blanks; honest note frame instead
        _ricd = _CE.icd10_dos_validity(_sstd, ref_dir=_refdir, mapping=_smapping)
        ok &= _check("v42.1 ICD DOS: zero rows flagged when no diagnosis delivered "
                     "(blanks are unjudged, not invalid)",
                     "note" in _ricd.columns or len(_ricd) == 0)

        # ICD screen: blanks mixed with real codes are skipped, real bad code flags
        _mix = _pd.DataFrame({"diagnosis": ["ZZ999", None, ""],
                              "date": ["2025-06-15"] * 3})
        _rmix = _CE.icd10_dos_validity(_mix, ref_dir=_refdir,
                                       mapping={"diagnosis": "diagnosis", "date": "date"})
        ok &= _check("v42.1 ICD DOS: blank rows skipped, only the real junk code flags",
                     len(_rmix) == 1 and _rmix["diagnosis"].iloc[0] == "ZZ999")

        # JW/JZ: unjudged inventory, never failures, when modifiers absent
        _rjw = _CE.jw_jz_wastage(_sstd, ref_dir=_refdir, mapping=_smapping)
        ok &= _check("v42.1 JW/JZ: single-dose lines are unjudged_missing_field, "
                     "not flagged, when no modifier field was delivered",
                     len(_rjw) > 0 and "verdict" in _rjw.columns
                     and (_rjw["verdict"] == "unjudged_missing_field").all())

        # closed-claims: manufactured all-NA status column reads as absent
        _rcc = _CC.closed_claims_view(_sstd, {"mapping": _smapping})
        _cf = _rcc.attrs.get("closed_frame")
        ok &= _check("v42.1 closed-claims: all-NA status column treated as absent; "
                     "all rows kept with an honest note",
                     _cf is not None and len(_cf) == len(_sstd)
                     and "No adjudication/status field" in _rcc.attrs.get("note", ""))

        # run_selected: unsupported fix reports missing field instead of running
        _rsel = _R.run_selected(_sstd, ["icd_dos_validity"],
                                {"ref_dir": _refdir, "mapping": _smapping})
        ok &= _check("v42.1 run_selected: unsupported fix reports the missing field",
                     "diagnosis" in _rsel["icd_dos_validity"]["note"].iloc[0])

        # delivered-but-empty column behaves like absent (not just unmapped)
        _empty = _sstd.copy()
        _empty["modifiers"] = _pd.NA
        _rjw2 = _CE.jw_jz_wastage(_empty, ref_dir=_refdir,
                                  mapping=dict(_smapping, modifiers="modifiers"))
        ok &= _check("v42.1 resolve: a delivered-but-empty column reads as absent",
                     len(_rjw2) > 0 and (_rjw2["verdict"] == "unjudged_missing_field").all())
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v42.1 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v43 ----
    # Calibration harness, calibrated recovery model, taxonomy coherence, capture.
    try:
        import numpy as _np
        import pandas as _pd
        from npi_recovery import prob_calibration as _PC
        from npi_recovery import recovery_model as _RM
        from npi_recovery import taxonomy_coherence as _TC
        from npi_recovery import capture_model as _CAP
        import os as _os
        _refdir = _os.path.join(_os.path.dirname(_RM.__file__), "reference")

        # --- prob_calibration: known-answer checks ---
        # perfectly calibrated: confidence == outcome probability in each bin
        _conf = _np.array([0.1, 0.1, 0.9, 0.9, 0.5, 0.5])
        _corr = _np.array([0, 0, 1, 1, 1, 0])  # bin 0.9 -> 100%, bin 0.5 -> 50%, bin 0.1 -> 0%
        ok &= _check("v43 calibration: Brier of a coin-flip prediction at 0.5 is 0.25",
                     _PC.brier_score(_np.array([0.5, 0.5]), _np.array([1, 0])) == 0.25)
        ok &= _check("v43 calibration: perfect confidence scores Brier 0",
                     _PC.brier_score(_np.array([1.0, 0.0]), _np.array([1, 0])) == 0.0)
        _rel = _PC.reliability_table(_conf, _corr, n_bins=10)
        ok &= _check("v43 calibration: reliability table bins and computes hit rate",
                     not _rel.empty and "actual_hit_rate" in _rel.columns)
        # AUC: a perfect ranker scores 1.0
        ok &= _check("v43 calibration: AUC of a perfect ranker is 1.0",
                     _PC.auc_roc(_np.array([0.1, 0.2, 0.8, 0.9]),
                                 _np.array([0, 0, 1, 1])) == 1.0)
        ok &= _check("v43 calibration: AUC of a reversed ranker is 0.0",
                     _PC.auc_roc(_np.array([0.9, 0.8, 0.2, 0.1]),
                                 _np.array([0, 0, 1, 1])) == 0.0)

        # --- logistic + isotonic internals ---
        _rng = _np.random.default_rng(3)
        _n = 400
        _x = _rng.normal(size=(_n, 1))
        _p = 1 / (1 + _np.exp(-(0.5 + 2.0 * _x[:, 0])))
        _y = (_rng.random(_n) < _p).astype(float)
        _lg = _RM._Logit(l2=0.5).fit(_x, _y)
        ok &= _check("v43 logit: hand-rolled logistic recovers a positive slope on "
                     "positively-related data",
                     _lg.beta[1] > 0)
        _iso = _RM._Isotonic().fit(_np.array([0.2, 0.1, 0.4, 0.3, 0.9]),
                                   _np.array([0.0, 0.0, 1.0, 0.0, 1.0]))
        _mono = _iso.predict(_np.array([0.1, 0.3, 0.5, 0.9]))
        ok &= _check("v43 isotonic: calibration map is monotone non-decreasing",
                     bool(_np.all(_np.diff(_mono) >= -1e-9)))

        # --- recovery model head-to-head: model beats a miscalibrated incumbent ---
        _m = 1200
        _tier = _rng.choice(["T1_full_key", "T3_ref_drug", "T5_drug_st"], _m)
        _tw = {"T1_full_key": 1.0, "T3_ref_drug": 0.78, "T5_drug_st": 0.45}
        _purity = _np.clip(_rng.beta(4, 2, _m), 0.2, 0.999)
        _margin = _np.clip(_purity - _rng.beta(2, 4, _m), 0, 0.99)
        _support = _rng.integers(1, 200, _m)
        _rank = _np.array([{"T1_full_key": 1, "T3_ref_drug": 3, "T5_drug_st": 6}[t] for t in _tier])
        _lt = -1.0 + 3.0 * _purity + 1.5 * _margin + 0.3 * _np.log1p(_support) - 0.3 * _rank
        _pt = 1 / (1 + _np.exp(-_lt))
        _yy = (_rng.random(_m) < _pt).astype(int)
        _hold = _pd.DataFrame({
            "tier": _tier, "tier_source": "in_panel",
            "confidence": _np.round(_np.array([_tw[t] for t in _tier]) * _purity, 4),
            "score": _np.round(_purity, 4), "margin": _np.round(_margin, 4),
            "support": _support, "t1": _yy,
            "amt": _np.round(_rng.lognormal(8, 1, _m), 2),
            "recovered_npi": ["1000000000"] * _m,
        })
        _res = _RM.fit_and_compare(_hold, seed=7)
        ok &= _check("v43 recovery model: head-to-head runs and returns both reports",
                     _res.get("status") == "ok" and "incumbent" in _res and "model" in _res)
        ok &= _check("v43 recovery model: calibrated model lowers Brier vs the "
                     "hand-set incumbent confidence",
                     _res["model"]["brier"] < _res["incumbent"]["brier"])
        ok &= _check("v43 recovery model: calibrated model lowers ECE vs incumbent",
                     _res["model"]["ece"] < _res["incumbent"]["ece"])
        ok &= _check("v43 recovery model: verdict + per-signal coefficients returned",
                     "verdict" in _res and isinstance(_res["coefficients"], _pd.DataFrame)
                     and len(_res["coefficients"]) == len(_RM.FEATURES))
        ok &= _check("v43 recovery model: insufficient holdout degrades cleanly",
                     _RM.fit_and_compare(_hold.head(10)).get("status") == "insufficient_holdout")

        # --- taxonomy coherence: real verdicts ---
        _tc_pred = _pd.DataFrame({
            "recovered_npi": ["1", "2", "3", "4"],
            "recovered_taxonomy": ["207RC0000X", "122300000X", "261QI0500X", ""],
        })
        _tc_std = _pd.DataFrame({"hcpcs": ["J1745", "J1745", "J1745", "J1745"]})
        _tc = _TC.coherence_series(_tc_pred, _tc_std, ref_dir=_refdir,
                                   mapping={"hcpcs": "hcpcs"})
        ok &= _check("v43 taxonomy: cardiologist coherent, dentist incoherent, "
                     "infusion-center coherent, missing unknown",
                     _tc.iloc[0] == 1.0 and _tc.iloc[1] == 0.0
                     and _tc.iloc[2] == 1.0 and _tc.iloc[3] == 0.5)
        _tc_oral = _TC.coherence_series(
            _pd.DataFrame({"recovered_npi": ["1"], "recovered_taxonomy": ["122300000X"]}),
            _pd.DataFrame({"hcpcs": ["99213"]}), ref_dir=_refdir, mapping={"hcpcs": "hcpcs"})
        ok &= _check("v43 taxonomy: non-drug HCPCS is not gated (neutral 0.5 even "
                     "for an odd taxonomy)",
                     _tc_oral.iloc[0] == 0.5)
        _tc_screen = _TC.coherence_screen(_tc_pred, _tc_std, ref_dir=_refdir,
                                          mapping={"hcpcs": "hcpcs"})
        ok &= _check("v43 taxonomy: screen surfaces exactly the incoherent recovery",
                     len(_tc_screen) == 1 and _tc_screen["recovered_npi"].iloc[0] == "2")

        # --- capture model: channel + drug + band ---
        _cap_std = _pd.DataFrame({
            "payer": (["BCBS of Texas"] * 40 + ["US Department of VA"] * 20
                      + ["Self Pay"] * 10 + ["Aetna"] * 30),
            "hcpcs": (["J1745"] * 60 + ["S0281"] * 20 + ["99213"] * 20),
            "allowed_amt": [1000.0] * 100,
        })
        _cmap = {"payer": "payer", "hcpcs": "hcpcs", "allowed_amt": "allowed_amt"}
        _chan = _CAP.channel_completeness(_cap_std, mapping=_cmap)
        _ch_set = set(_chan["channel"])
        ok &= _check("v43 capture: classifies commercial, VA/military, and cash channels",
                     {"commercial", "va_military", "cash"}.issubset(_ch_set))
        ok &= _check("v43 capture: flags VA and cash as under-captured",
                     bool(_chan[_chan["channel"] == "va_military"]["under_captured"].iloc[0])
                     and bool(_chan[_chan["channel"] == "cash"]["under_captured"].iloc[0]))
        _band = _CAP.implied_capture_band(_cap_std, ref_dir=_refdir, mapping=_cmap)
        ok &= _check("v43 capture: implied band is a proper interval in (0,1] with "
                     "low <= high",
                     _band["status"] == "ok"
                     and 0 < _band["implied_capture_low"] <= _band["implied_capture_high"] <= 1.0)
        _drug = _CAP.drug_capture_flags(_cap_std, ref_dir=_refdir, mapping=_cmap)
        ok &= _check("v43 capture: J-code well captured, S-code not",
                     bool(_drug[_drug["hcpcs"] == "J1745"]["well_captured"].iloc[0])
                     and not bool(_drug[_drug["hcpcs"] == "S0281"]["well_captured"].iloc[0]))
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v43 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v44 ----
    # Two-method agreement, evidence ledger, seller data request, run manifest.
    try:
        import pandas as _pd
        import numpy as _np
        import os as _os
        from npi_recovery import agreement as _AG
        from npi_recovery import ledger as _LED
        from npi_recovery import seller_request as _SR
        from npi_recovery import run_manifest as _RMAN
        from npi_recovery.impute import ReferralImputer as _RI
        from npi_recovery import config as _cfg

        # ---- two-method agreement ----
        _std = _pd.DataFrame({
            "billing_npi": ["1003914151"] * 10 + ["1184649626"] * 10 + [None] * 10,
            "referring_npi": ["1999999984"] * 30, "hcpcs": ["J1745"] * 30,
            "pos": ["11"] * 30, "zip3": ["770"] * 30, "state": ["TX"] * 30,
            "payer": ["BCBS"] * 30, "referring_specialty": ["NEU"] * 30,
            "allowed_amt": [1000.0] * 30,
        })
        _imp = _RI(min_support=1).fit(_std, pools={})
        _bidx = _std[_std["billing_npi"].isna()].index
        # agreeing pool: CMS pool top NPI matches the in-panel winner
        _imp.pools = {("J1745", "TX"): _pd.DataFrame(
            {"npi": ["1003914151", "1184649626"], "srvcs": [900, 100], "name": ["A", "B"]})}
        _at = _AG.two_method_table(_std, _imp, _bidx)
        ok &= _check("v44 agreement: both methods agreeing yields agreement=agree "
                     "with a lift boost",
                     (_at["agreement"] == "agree").all() and (_at["agreement_boost"] > 1.0).all())
        # disagreeing pool
        _imp.pools = {("J1745", "TX"): _pd.DataFrame(
            {"npi": ["1777777770"], "srvcs": [500], "name": ["Z"]})}
        _at2 = _AG.two_method_table(_std, _imp, _bidx)
        ok &= _check("v44 agreement: methods naming different billers yields "
                     "agreement=disagree with a damp boost",
                     (_at2["agreement"] == "disagree").all() and (_at2["agreement_boost"] < 1.0).all())
        # no pool -> a_only (in-panel fires, CMS does not)
        _imp.pools = {}
        _at3 = _AG.two_method_table(_std, _imp, _bidx)
        ok &= _check("v44 agreement: in-panel only (no CMS pool) yields a_only, neutral boost",
                     (_at3["agreement"] == "a_only").all())
        _dq = _AG.disagreement_queue(_at2, _std, mapping={"allowed_amt": "allowed_amt"})
        ok &= _check("v44 agreement: disagreement queue surfaces the disagreeing rows",
                     len(_dq) == 10 and "recommended_action" in _dq.columns)
        _empty_dq = _AG.disagreement_queue(_at, _std)
        ok &= _check("v44 agreement: no disagreements yields an empty queue with a clean note",
                     len(_empty_dq) == 0)

        # ---- evidence ledger ----
        _rec = _pd.DataFrame({
            "orig_row": [20, 21, 22], "recovered_npi": ["1003914151", "1184649626", None],
            "tier": ["T1_full_key", "T5_drug_st", "none"],
            "tier_source": ["in_panel", "cms_pool", "none"],
            "confidence": [0.95, 0.4, 0.0], "attribution": ["point", "distributional", "unrecovered"],
            "demoted_near_tie": [False, False, False], "demote_reason": ["", "", ""],
            "blank_allowed": [5000.0, 2000.0, 100.0],
        })
        _agt = _pd.DataFrame({"row": [20, 21, 22], "agreement": ["agree", "disagree", "neither"],
                              "agreement_boost": [1.15, 0.6, 0.8]})
        _rl = _pd.DataFrame([{"repair": "zero-pad ZIP", "field": "zip", "rows_fixed": 12,
                              "method": "deterministic", "example": "7002 -> 07002"}])
        _lg = _LED.build_ledger(_rec, _rl, _agt, calib_probs=[0.9, 0.5, 0.0],
                                vintages={"cms_utilization": "2024"})
        ok &= _check("v44 ledger: one row per recovery plus one per repair",
                     len(_lg) == 4 and "safe_for_basecase" in _lg.columns)
        # point + agree + high calib prob -> base-case safe; disagree -> not
        _npi_rows = _lg[_lg["field"] == "billing_npi"].reset_index(drop=True)
        ok &= _check("v44 ledger: point recovery with agreement and high probability "
                     "is base-case safe",
                     bool(_npi_rows.loc[0, "safe_for_basecase"]))
        ok &= _check("v44 ledger: a two-method disagreement is NOT base-case safe "
                     "even at point attribution",
                     not bool(_npi_rows.loc[1, "safe_for_basecase"]))
        ok &= _check("v44 ledger: deterministic field repairs are base-case safe",
                     bool(_lg[_lg["change_type"] == "field_repair"]["safe_for_basecase"].iloc[0]))
        _roll = _LED.basecase_rollup(_lg)
        ok &= _check("v44 ledger: base-case rollup splits recovered dollars into "
                     "safe vs leads",
                     set(_roll["bucket"]) <= {"base_case_safe", "lead_verify"} and len(_roll) >= 1)

        # ---- seller data request ----
        # a fixability manifest with a missing diagnosis and a missing modifier verdict
        _fixa = _pd.DataFrame([
            {"fix": "ICD-10 date-of-service validity", "status": "unsupported",
             "missing_required": "diagnosis", "missing_for_verdict": ""},
            {"fix": "JW/JZ single-dose wastage logic", "status": "partial",
             "missing_required": "", "missing_for_verdict": "modifiers"},
        ])
        _sr = _SR.build_seller_request(fixability=_fixa)
        ok &= _check("v44 seller request: asks for the missing diagnosis and modifier fields",
                     any("diagnosis" in str(r) for r in _sr["request"]) and
                     any("modifier" in str(r) for r in _sr["request"]))
        ok &= _check("v44 seller request: diagnosis is high priority",
                     _sr[_sr["request"].str.contains("diagnosis")]["priority"].iloc[0] == "high")
        _sr_none = _SR.build_seller_request()
        ok &= _check("v44 seller request: nothing missing yields a clean no-request row",
                     len(_sr_none) == 1 and _sr_none["priority"].iloc[0] == "none")

        # ---- run manifest ----
        _man = _RMAN.build_manifest("/tmp/does_not_exist.xlsx", "44.0.0",
                                    ref_dir=_cfg.REF_DIR, options={"do_bulk": False})
        ok &= _check("v44 manifest: carries run_id, tool_version, and reference vintages",
                     _man["tool_version"] == "44.0.0" and "run_id" in _man
                     and isinstance(_man["reference_vintages"], dict))
        ok &= _check("v44 manifest: reads a real vintage from the shipped MUE seed",
                     any("2026" in str(v) for v in _man["reference_vintages"].values()))
        _mf = _RMAN.manifest_frame(_man)
        ok &= _check("v44 manifest: flattens to a two-column sheet frame",
                     list(_mf.columns) == ["key", "value"] and len(_mf) >= 4)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v44 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v45 ----
    # Suggested corrections, issue analysis, consistency screens, orchestrator.
    try:
        import pandas as _pd
        import numpy as _np
        import os as _os
        from npi_recovery import suggested_fixes as _SF
        from npi_recovery import issue_analysis as _IA
        from npi_recovery import consistency as _CON
        from npi_recovery import clean_orchestrator as _CO
        from npi_recovery import registry as _R
        _refdir = _os.path.join(_os.path.dirname(_R.__file__), "reference")

        _df = _pd.DataFrame({
            "billing_npi": ["1003914151"] * 50 + ["1999999984"] * 30 + ["1184649626"] * 20,
            "referring_npi": ["1003914151"] * 10 + ["1720082779"] * 90,
            "hcpcs": ["J1745"] * 40 + ["J9312"] * 30 + ["J3357"] * 30,
            "units": [200] + [10] * 39 + [1] * 30 + [5] * 30,
            "modifiers": [""] * 70 + ["JZ"] * 30,
            "diagnosis": ["G70.00"] * 50 + ["B880"] * 30 + ["M79.10"] * 20,
            "date": ["2025-06-15"] * 60 + ["2027-01-01"] * 40,
            "allowed_amt": [1000.0] * 80 + [500.0] * 20,
            "billed_amt": [900.0] * 80 + [600.0] * 20,
            "payer": ["BCBS of Texas"] * 50 + ["UHC"] * 50,
        })
        _map = {c: c for c in _df.columns}

        # ---- consistency screens ----
        _mo = _CON.money_ordering(_df, _map)
        ok &= _check("v45 consistency: money ordering flags allowed>billed rows",
                     "row" in _mo.columns and len(_mo) == 80)
        _do = _CON.date_ordering(_df, _map)
        ok &= _check("v45 consistency: date ordering flags future service dates",
                     "row" in _do.columns and len(_do) == 40)
        _rc = _CON.npi_role_coherence(_df, _map)
        ok &= _check("v45 consistency: role coherence flags referring==billing",
                     "row" in _rc.columns and len(_rc) == 10)
        # missing inputs degrade cleanly
        _mo_bad = _CON.money_ordering(_pd.DataFrame({"allowed_amt": [1.0]}), {"allowed_amt": "allowed_amt"})
        ok &= _check("v45 consistency: money ordering degrades cleanly with one amount",
                     "note" in _mo_bad.columns)

        # ---- coding screens + suggested fixes ----
        _res = _R.run_selected(_df, ["mue_units", "jw_jz_wastage", "icd_dos_validity"],
                               {"ref_dir": _refdir, "mapping": _map})
        _sug = _SF.build_suggestions(_res, std=_df, mapping=_map)
        ok &= _check("v45 suggestions: builds corrections carrying rule, provenance, "
                     "and safe-to-auto-apply",
                     not _sug.empty and set(["fix_rule", "provenance", "safe_to_auto_apply",
                                             "suggested_value"]).issubset(_sug.columns))
        ok &= _check("v45 suggestions: MUE over-cap suggests capping units to the MUE value",
                     ((_sug["issue"] == "units_exceed_mue") &
                      (_sug["suggested_value"].astype(str).str.contains("150"))).any())
        ok &= _check("v45 suggestions: nothing is marked auto-applicable by default "
                     "(coding fixes are review-required)",
                     not _sug["safe_to_auto_apply"].any())
        # apply_safe_suggestions leaves review-required rows untouched
        _cleaned, _applied = _SF.apply_safe_suggestions(_df, _sug, _map)
        ok &= _check("v45 suggestions: apply-safe applies nothing when all are "
                     "review-required (data untouched)",
                     len(_applied) == 0 and _cleaned["units"].iloc[0] == _df["units"].iloc[0])

        # ---- issue analysis ----
        _summ, _det = _IA.analyze_all(_res, _df, mapping=_map)
        ok &= _check("v45 issue analysis: sizes each issue with dollars and a "
                     "systematic verdict",
                     not _summ.empty and "systematic_signal" in _summ.columns
                     and "dollar_exposure" in _summ.columns)
        # concentrated issue reads systematic; diffuse reads random
        _rng = _np.random.default_rng(9)
        _diff_std = _pd.DataFrame({
            "billing_npi": [f"1{_rng.integers(100000000,999999999)}" for _ in range(400)],
            "hcpcs": [f"J{_rng.integers(1000,9999)}" for _ in range(400)],
            "allowed_amt": _rng.uniform(100, 1000, 400),
        })
        _diff_flag = _pd.DataFrame({"row": _rng.choice(_diff_std.index, 40, replace=False)})
        _a_diff = _IA.analyze_issue(_diff_flag, _diff_std, "diffuse",
                                    mapping={c: c for c in _diff_std.columns})
        ok &= _check("v45 issue analysis: a diffuse issue reads as random error",
                     _a_diff["systematic_signal"].startswith("diffuse"))
        _conc_flag = _pd.DataFrame({"row": list(_df.index[:40])})  # all J1745, one provider-ish
        _a_conc = _IA.analyze_issue(_conc_flag, _df, "concentrated", mapping=_map)
        ok &= _check("v45 issue analysis: a concentrated issue reads as systematic",
                     _a_conc["systematic_signal"].startswith("systematic"))

        # ---- orchestrator ----
        _out = _CO.clean_all(_df, ref_dir=_refdir, mapping=_map, run_deep_clean=False)
        ok &= _check("v45 orchestrator: returns cleaned data, screens, suggestions, "
                     "issue summary, and a scorecard",
                     set(["cleaned", "screens", "suggestions", "issue_summary",
                          "scorecard"]).issubset(_out.keys()))
        ok &= _check("v45 orchestrator: scorecard has an auto-fixed row and "
                     "suggested-review rows",
                     (_out["scorecard"]["category"] == "auto_fixed_deterministic").any()
                     and (_out["scorecard"]["category"] == "suggested_review_required").any())
        ok &= _check("v45 orchestrator: folds cross-field consistency issues into "
                     "the suggestions",
                     (_out["suggestions"]["issue"] == "money_ordering").any())
        # registry now exposes the consistency screens as selectable fixes
        ok &= _check("v45 registry: consistency screens are registered as fixes",
                     all(k in _R.REGISTRY_BY_KEY for k in
                         ("money_ordering", "date_ordering", "npi_role_coherence",
                          "units_days_supply")))
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v45 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v46 ----
    # DuckDB engine (SQL==pandas), review loop, profiling, multi-format export.
    try:
        import pandas as _pd
        import numpy as _np
        import os as _os
        import tempfile as _tf
        from npi_recovery import review as _RV
        from npi_recovery import profiling as _PR
        from npi_recovery import export as _EX
        from npi_recovery import clean_orchestrator as _CO
        from npi_recovery import coding_edits as _CE2
        from npi_recovery import registry as _R
        _refdir = _os.path.join(_os.path.dirname(_R.__file__), "reference")

        _df = _pd.DataFrame({
            "billing_npi": ["1003914151"] * 100 + ["1999999984"] * 60 + ["1184649626"] * 40,
            "hcpcs": ["J1745"] * 80 + ["J9312"] * 60 + ["J3357"] * 60,
            "units": [200] + [10] * 79 + [1] * 60 + [5] * 60,
            "modifiers": [""] * 120 + ["JZ"] * 80,
            "allowed_amt": [1000.0] * 200,
        })
        _map = {c: c for c in _df.columns}

        # ---- engine: SQL screens must match pandas screens ----
        from npi_recovery import engine as _ENG
        if _ENG.duckdb_available():
            _csv = _os.path.join(_tf.gettempdir(), "_selftest_engine.csv")
            _df.to_csv(_csv, index=False)
            ok &= _check("v46 engine: SQL rowcount matches the frame",
                         _ENG.rowcount(_csv) == len(_df))
            _sql_mue = _ENG.mue_screen_sql(_csv, ref_dir=_refdir, mapping=_map)
            _pd_mue = _CE2.mue_screen(_df, ref_dir=_refdir, mapping=_map)
            ok &= _check("v46 engine: SQL MUE screen flags the same count as pandas",
                         len(_sql_mue) == len(_pd_mue))
            _sql_jw = _ENG.jw_jz_screen_sql(_csv, ref_dir=_refdir, mapping=_map)
            _pd_jw = _CE2.jw_jz_wastage(_df, ref_dir=_refdir, mapping=_map)
            ok &= _check("v46 engine: SQL JW/JZ screen matches pandas (NULL-safe "
                         "modifier handling)",
                         len(_sql_jw) == len(_pd_jw))
            _sql_de = _ENG.deactivated_screen_sql(_csv, ref_dir=_refdir, mapping=_map)
            _pd_de = _CE2.deactivated_npi_screen(_df, ref_dir=_refdir, mapping=_map)
            ok &= _check("v46 engine: SQL deactivated-NPI screen matches pandas",
                         len(_sql_de) == len(_pd_de))
            _mix = _ENG.drug_mix_sql(_csv, mapping=_map)
            ok &= _check("v46 engine: SQL drug mix returns dollar shares summing to ~100",
                         abs(float(_mix["share_pct"].sum()) - 100.0) < 1.0)
            _prof = _ENG.profile(_csv, _map)
            ok &= _check("v46 engine: SQL profile flags the 60%-null modifier column",
                         float(_prof["columns"].set_index("column").loc["modifiers", "null_pct"]) == 60.0)
            try:
                _os.remove(_csv)
            except Exception:
                pass
        else:
            ok &= _check("v46 engine: duckdb not installed, SQL path skipped (pandas "
                         "fallback active)", True)

        # ---- review loop ----
        _out = _CO.clean_all(_df, ref_dir=_refdir, mapping=_map, run_deep_clean=False)
        _sug = _out["suggestions"]
        _dec = _RV.blank_decisions(_sug)
        ok &= _check("v46 review: blank decisions start every suggestion on hold",
                     (_dec["decision"] == _RV.HOLD).all())
        _dec = _RV.decide_by_rule(_dec, _sug, accept_issues=["units_exceed_mue"])
        ok &= _check("v46 review: a rule accepts the targeted issue and leaves others on hold",
                     (_dec["decision"] == _RV.ACCEPT).sum() >= 1
                     and (_dec["decision"] == _RV.HOLD).any())
        _final, _led = _RV.apply_decisions(_df, _sug, _dec)
        ok &= _check("v46 review: applying an accepted MUE cap changes the value and "
                     "preserves the original in a shadow column",
                     "units__original" in _final.columns
                     and _final.loc[0, "units"] == 150 and _final.loc[0, "units__original"] == 200)
        ok &= _check("v46 review: the decision ledger records before, after, and applied",
                     not _led.empty and set(["before", "after", "applied"]).issubset(_led.columns))
        # a descriptive (non-concrete) suggestion is accepted but marked manual
        _dec2 = _RV.blank_decisions(_sug)
        _dec2 = _RV.decide_by_rule(_dec2, _sug, accept_issues=["missing_wastage_modifier"])
        _f2, _l2 = _RV.apply_decisions(_df, _sug, _dec2)
        ok &= _check("v46 review: a descriptive fix is accepted but flagged manual, "
                     "not applied blindly",
                     (not _l2.empty) and (_l2["status"] == "accepted_manual_action_required").any())

        # ---- profiling ----
        _pf = _PR.profile_frame(_df)
        ok &= _check("v46 profiling: infers npi, hcpcs, and numeric semantic types",
                     set(["npi", "hcpcs", "numeric"]).issubset(set(_pf["semantic_type"])))
        _const = _pd.DataFrame({"x": [1] * 50, "y": range(50), "z": [None] * 50})
        _pf2 = _PR.profile_frame(_const)
        _flags = dict(zip(_pf2["column"], _pf2["quality_flag"]))
        ok &= _check("v46 profiling: flags constant, key-like, and empty columns",
                     _flags.get("x") == "constant" and _flags.get("z") == "empty"
                     and "key" in _flags.get("y", ""))

        # ---- export ----
        _exp = _os.path.join(_tf.gettempdir(), "_selftest_export")
        _man = _EX.export_result({"cleaned": _out["cleaned"], "suggestions": _sug}, _exp, fmt="parquet")
        ok &= _check("v46 export: writes parquet frames with non-zero size",
                     not _man.empty and (_man["bytes"] > 0).all())
        _rt = _pd.read_parquet(_os.path.join(_exp, "suggestions.parquet"))
        ok &= _check("v46 export: parquet round-trips a mixed-type corrections frame",
                     len(_rt) == len(_sug))
        _csvp = _os.path.join(_tf.gettempdir(), "_selftest_export.json")
        _EX.export_frame(_out["cleaned"], _csvp, fmt="json")
        ok &= _check("v46 export: json export writes a readable records file",
                     _os.path.exists(_csvp) and _os.path.getsize(_csvp) > 0)
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v46 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v47 ----
    # RIF/VRDC ingestion, CMS cell suppression, FFS-specific recovery.
    try:
        import pandas as _pd
        import numpy as _np
        import os as _os
        from npi_recovery import rif_schema as _RIF
        from npi_recovery import rif_cleaning as _RC
        from npi_recovery import coding_edits as _CE3
        from npi_recovery import schema as _SC
        from npi_recovery import registry as _R
        _refdir = _os.path.join(_os.path.dirname(_R.__file__), "reference")

        # ---- RIF carrier (real CCW variable names) ----
        _n = 100
        _carrier = _pd.DataFrame({
            "CLM_ID": [f"C{i:08d}" for i in range(_n)],
            "BENE_ID": [f"B{i%40:06d}" for i in range(_n)],
            "CLM_FROM_DT": ["2025-06-15"] * _n,
            "CARR_CLM_HCPCS_CD": ["J1745"] * 60 + ["J9312"] * 40,
            "HCPCS_CD": ["J1745"] * 60 + ["J9312"] * 40,
            "HCPCS_1ST_MDFR_CD": [""] * 70 + ["JZ"] * 30,
            "PRF_PHYSN_NPI": ["1003914151"] * 50 + ["1999999984"] * 50,
            "ORG_NPI_NUM": ["1184649626"] * 80 + [""] * 20,
            "RFR_PHYSN_NPI": ["1720082779"] * _n,
            "LINE_ALOWD_CHRG_AMT": [1000.0] * _n,
            "LINE_NCH_PMT_AMT": [800.0] * _n,
            "LINE_SBMTD_CHRG_AMT": [1200.0] * _n,
            "LINE_SRVC_CNT": [200] + [10] * 99,
            "LINE_ICD_DGNS_CD": ["G7000"] * _n,
            "CARR_NUM": ["06302"] * _n,
        })
        _rt, _score, _ev = _RIF.detect_rif_type(_carrier)
        ok &= _check("v47 RIF: carrier file detected from CCW signature columns",
                     _rt == "carrier" and _score >= 3)

        _pde = _pd.DataFrame({
            "PDE_ID": ["P1"], "BENE_ID": ["B1"], "SRVC_DT": ["2025-06-15"],
            "PROD_SRVC_ID": ["00074433902"], "QTY_DSPNSD_NUM": [30],
            "DAYS_SUPLY_NUM": [30], "PRSCRBR_ID": ["1003914151"],
            "SRVC_PRVDR_ID": ["1999999984"], "TOT_RX_CST_AMT": [5000.0],
            "CVRD_D_PLAN_PD_AMT": [4500.0],
        })
        ok &= _check("v47 RIF: Part D PDE file detected",
                     _RIF.detect_rif_type(_pde)[0] == "pde")
        _inp = _pd.DataFrame({
            "CLM_ID": ["I1"], "BENE_ID": ["B1"], "CLM_ADMSN_DT": ["2025-06-01"],
            "NCH_BENE_DSCHRG_DT": ["2025-06-05"], "ORG_NPI_NUM": ["1184649626"],
            "AT_PHYSN_NPI": ["1003914151"], "PRNCPAL_DGNS_CD": ["I509"],
            "CLM_TOT_CHRG_AMT": [50000.0], "CLM_PMT_AMT": [12000.0],
            "CLM_PPS_CPTL_FSP_AMT": [800.0],
        })
        ok &= _check("v47 RIF: inpatient file detected",
                     _RIF.detect_rif_type(_inp)[0] == "inpatient")
        # non-RIF must NOT be detected as RIF
        _komodo = _pd.DataFrame({"BillingProviderNPI": ["1003914151"],
                                 "ProcedureCode": ["J1745"], "AllowedAmt": [100.0]})
        ok &= _check("v47 RIF: a commercial extract is not misdetected as RIF",
                     not _RIF.is_rif(_komodo))

        # ---- standardization maps RIF -> canonical ----
        _std, _rtype, _rep = _RIF.standardize_rif(_carrier)
        ok &= _check("v47 RIF: carrier standardizes billing/rendering/referring NPI, "
                     "hcpcs, and money into canonical fields",
                     all(c in _std.columns for c in
                         ("billing_npi", "rendering_npi", "referring_npi", "hcpcs",
                          "allowed_amt", "paid_amt", "units")))
        ok &= _check("v47 RIF: ORG_NPI_NUM maps to billing_npi, PRF_PHYSN_NPI to "
                     "rendering_npi",
                     _std["billing_npi"].dropna().iloc[0] == "1184649626"
                     and _std["rendering_npi"].iloc[0] == "1003914151")

        # ---- the payoff: existing screens run unchanged on RIF data ----
        _m = {c: c for c in _std.columns}
        _mue_rif = _CE3.mue_screen(_std, ref_dir=_refdir, mapping=_m)
        ok &= _check("v47 RIF: the MUE screen runs unchanged on standardized RIF data",
                     "row" in _mue_rif.columns and len(_mue_rif) == 1)
        _jw_rif = _CE3.jw_jz_wastage(_std, ref_dir=_refdir, mapping=_m)
        ok &= _check("v47 RIF: the JW/JZ screen runs unchanged on standardized RIF data",
                     "row" in _jw_rif.columns and len(_jw_rif) == 70)

        # ---- standardize_any auto-routing ----
        _sa_std, _sa_map, _sa_rep = _SC.standardize_any(_carrier)
        ok &= _check("v47 RIF: standardize_any routes RIF input to the RIF path",
                     str(_sa_rep.get("source_format", "")).startswith("rif:"))
        _sa_std2, _, _sa_rep2 = _SC.standardize_any(_komodo)
        ok &= _check("v47 RIF: standardize_any routes a commercial extract to the "
                     "commercial path",
                     _sa_rep2.get("source_format") == "commercial")

        # ---- CMS cell suppression ----
        _agg = _pd.DataFrame({
            "drug": ["IVIG", "Tepezza", "Stelara", "Ocrevus", "RareDrug"],
            "beneficiaries": [500, 250, 120, 45, 7],
            "allowed": [5e6, 3e6, 1.2e6, 450000, 70000],
        })
        _supp = _RC.suppress_small_cells(_agg, count_col="beneficiaries",
                                         value_cols=["allowed"])
        ok &= _check("v47 suppression: a cell below 11 beneficiaries is suppressed "
                     "and its value blanked",
                     bool(_supp[_supp["drug"] == "RareDrug"]["suppressed"].iloc[0])
                     and _pd.isna(_supp[_supp["drug"] == "RareDrug"]["allowed"].iloc[0]))
        ok &= _check("v47 suppression: cells at or above 11 beneficiaries are kept",
                     not bool(_supp[_supp["drug"] == "Ocrevus"]["suppressed"].iloc[0]))
        _, _clean = _RC.export_safe(_agg, count_col="beneficiaries", value_cols=["allowed"])
        ok &= _check("v47 suppression: export gate reports clean once small cells "
                     "are suppressed",
                     _clean is True)
        # complementary suppression: a lone small cell in a group forces a second
        _agg2 = _pd.DataFrame({
            "grp": ["A", "A", "A", "B", "B"],
            "cell": ["a1", "a2", "a3", "b1", "b2"],
            "beneficiaries": [8, 500, 400, 300, 250],  # one small cell in group A
            "allowed": [8000, 5e5, 4e5, 3e5, 25e4],
        })
        _supp2 = _RC.suppress_small_cells(_agg2, count_col="beneficiaries",
                                          value_cols=["allowed"], group_col="grp")
        ok &= _check("v47 suppression: a lone small cell triggers complementary "
                     "suppression of a second cell in the group",
                     int(_supp2[_supp2["grp"] == "A"]["suppressed"].sum()) >= 2)

        # ---- FFS solo-biller recovery ----
        _solo = _RC.flag_solo_biller_gap(_std)
        ok &= _check("v47 RIF: solo-biller gap recovers blank-org lines by promoting "
                     "the rendering NPI",
                     len(_solo) == 20 and "suggested_billing_npi" in _solo.columns
                     and (_solo["suggested_billing_npi"] == _solo["rendering_npi"]).all())
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v47 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v48 ----
    # Specialty-by-drug utilization model replacing the hand-coded taxonomy gate.
    try:
        import pandas as _pd
        import numpy as _np
        from npi_recovery import specialty_drug as _SD

        # ---- prior model ----
        _m = _SD.default_model()
        ok &= _check("v48 specialty-drug: prior model fits and scores a plausible "
                     "biller high, an implausible specialty at zero",
                     _m.plausibility("Rheumatology", "J1745") > 0.5
                     and _m.plausibility("Podiatry", "J1745") == 0.0)
        ok &= _check("v48 specialty-drug: an unknown drug returns neutral (no opinion)",
                     _m.plausibility("Neurology", "J9999") == 0.5)

        # ---- fitted from utilization: frequencies drive graded plausibility ----
        _rng = _np.random.default_rng(3)
        _rows = []
        for _spec, _nn in [("Rheumatology", 400), ("Gastroenterology", 300),
                           ("Neurology", 120), ("Hematology-Oncology", 80),
                           ("Dermatology", 5)]:
            _rows += [{"specialty": _spec, "hcpcs": "J1745",
                       "services": int(_rng.integers(1, 50))} for _ in range(_nn)]
        _util = _pd.DataFrame(_rows)
        _m2 = _SD.SpecialtyDrugModel().fit(_util, weight_col="services", source="synthetic_puf")
        ok &= _check("v48 specialty-drug: fitted model ranks the top biller near 1 and "
                     "a rare biller low",
                     _m2.plausibility("Rheumatology", "J1745") == 1.0
                     and _m2.plausibility("Dermatology", "J1745") < 0.1)
        _dist = _m2.specialty_distribution("J1745")
        ok &= _check("v48 specialty-drug: specialty distribution sums to ~1 and leads "
                     "with the top biller",
                     abs(float(_dist["share_of_drug"].sum()) - 1.0) < 0.05
                     and _dist.iloc[0]["specialty"] == "Rheumatology")

        # ---- plausibility screen flags only the implausible recovery ----
        _pred = _pd.DataFrame({"recovered_npi": ["1", "2", "3"],
                               "recovered_specialty": ["Rheumatology", "Dermatology", "Neurology"]})
        _std = _pd.DataFrame({"hcpcs": ["J1745", "J1745", "J1745"]})
        _fl = _SD.plausibility_screen(_pred, _std, model=_m2, mapping={"hcpcs": "hcpcs"}, threshold=0.1)
        ok &= _check("v48 specialty-drug: screen flags only the rare-biller recovery",
                     len(_fl) == 1 and _fl["recovered_npi"].iloc[0] == "2")

        # ---- persistence round-trip ----
        import tempfile as _tf, os as _os2
        _mp = _os2.path.join(_tf.gettempdir(), "_sd_model.json")
        _m2.to_json(_mp)
        _m3 = _SD.SpecialtyDrugModel.from_json(_mp)
        ok &= _check("v48 specialty-drug: model round-trips through JSON with identical "
                     "scoring",
                     _m3.plausibility("Rheumatology", "J1745") == _m2.plausibility("Rheumatology", "J1745"))

        # ---- recovery model adopts the graded plausibility as its coherence feature ----
        from npi_recovery import recovery_model as _RM3
        _pred2 = _pd.DataFrame({
            "recovered_npi": ["1", "2"], "recovered_specialty": ["Rheumatology", "Podiatry"],
            "confidence": [0.9, 0.8], "score": [0.9, 0.8], "margin": [0.5, 0.4],
            "support": [100, 80], "tier": ["T1_full_key"] * 2, "tier_source": ["in_panel"] * 2,
        })
        _std2 = _pd.DataFrame({"hcpcs": ["J1745", "J1745"]})
        _feats = _RM3.build_features(_pred2, std=_std2, mapping={"hcpcs": "hcpcs"})
        ok &= _check("v48 specialty-drug: recovery model uses the graded plausibility so "
                     "an implausible recovery scores 0 on coherence",
                     _feats["tax_coherent"].iloc[0] > 0.5 and _feats["tax_coherent"].iloc[1] == 0.0)

        # ---- build utilization from RIF carrier data ----
        _rif = _pd.DataFrame({
            "hcpcs": ["J1745"] * 4, "rendering_npi": ["1", "2", "3", "4"],
            "rendering_specialty": ["Rheumatology", "Rheumatology", "Neurology", "Dermatology"],
            "units": [10, 20, 5, 1],
        })
        _rutil = _SD.build_utilization_from_rif(_rif)
        ok &= _check("v48 specialty-drug: utilization can be built from RIF carrier "
                     "specialty-by-drug for a real refit",
                     len(_rutil) == 4 and set(_rutil.columns) >= {"specialty", "hcpcs", "w"})
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v48 checks run ({type(_e).__name__}: {_e})", False)

    # --------------------------------------------------------------- v49 ----
    # PECOS reassignment graph: structural recovery, validation, three-method.
    try:
        import pandas as _pd
        from npi_recovery import reassignment as _RA

        _g = _RA.demo_graph()
        # ---- lookups ----
        ok &= _check("v49 reassignment: individual-to-org lookup returns the "
                     "organization the physician reassigns to",
                     _g.orgs_for_individual("1003914151") == ["1184649626"])
        ok &= _check("v49 reassignment: reassigns_to validates a real link and "
                     "rejects a fake one",
                     _g.reassigns_to("1003914151", "1184649626")
                     and not _g.reassigns_to("1003914151", "9999999999"))

        # ---- structural recovery ----
        _r1 = _g.recover_billing_npi("1003914151")
        ok &= _check("v49 reassignment: a single reassignment recovers the org at "
                     "high confidence",
                     _r1["recovered_npi"] == "1184649626" and _r1["confidence"] >= 0.9
                     and _r1["basis"] == "single_reassignment")
        _r2 = _g.recover_billing_npi("1184649620")
        ok &= _check("v49 reassignment: multiple reassignments narrow to a structural "
                     "candidate set",
                     _r2["basis"] == "multiple_reassignment" and len(_r2["candidates"]) == 2)
        _r3 = _g.recover_billing_npi("5555555555")
        ok &= _check("v49 reassignment: an unknown individual yields no recovery "
                     "(no fabrication)",
                     _r3["recovered_npi"] == "" and _r3["basis"] == "no_reassignment")

        # ---- recover blanks from a claims frame ----
        _std = _pd.DataFrame({
            "billing_npi": [None, None, None, "1184649626"],
            "rendering_npi": ["1003914151", "1720082779", "5555555555", "1999999984"],
            "hcpcs": ["J1745"] * 4,
        })
        _m = {c: c for c in _std.columns}
        _rec = _RA.recover_from_reassignment(_std, _g, mapping=_m)
        ok &= _check("v49 reassignment: recovers exactly the blank rows whose rendering "
                     "NPI is in the graph (2 of 3 blanks)",
                     len(_rec) == 2 and set(_rec["basis"]) == {"single_reassignment"})

        # ---- validate recovered billers ----
        _pred = _pd.DataFrame({"orig_row": [0, 1],
                               "recovered_npi": ["1184649626", "1770009988"]})
        _val = _RA.validate_billers(_pred, _std, _g, mapping=_m)
        ok &= _check("v49 reassignment: validation flags a recovered biller the "
                     "rendering physician does not reassign to",
                     len(_val) == 1 and _val["recovered_npi"].iloc[0] == "1770009988")

        # ---- three-method agreement ----
        _agree = _pd.DataFrame({
            "row": [0, 1, 2],
            "method_a_npi": ["1184649626", "1558887711", "9999999999"],
            "method_b_npi": ["1184649626", "1770009988", "8888888888"],
            "agreement": ["agree", "disagree", "disagree"],
            "agreement_boost": [1.15, 0.6, 0.6],
        })
        _std3 = _pd.DataFrame({"rendering_npi": ["1003914151", "1720082779", "1003914151"]})
        _three = _RA.add_to_agreement(_agree, _std3, _g, mapping={"rendering_npi": "rendering_npi"})
        _states = dict(zip(_three["row"], _three["three_method_state"]))
        ok &= _check("v49 reassignment: all-three-agree gets the highest boost",
                     _states[0] == "all_three_agree"
                     and _three[_three["row"] == 0]["agreement_boost"].iloc[0] == 1.25)
        ok &= _check("v49 reassignment: structure confirms one of two disagreeing "
                     "statistical guesses, lifting the boost",
                     _states[1] == "structural_confirms_one"
                     and _three[_three["row"] == 1]["agreement_boost"].iloc[0] > 1.0)
        ok &= _check("v49 reassignment: structure overrides two bad statistical guesses",
                     _states[2] == "structural_overrides")

        # ---- fit from a PECOS-style table ----
        _pecos = _pd.DataFrame({
            "individual_npi": ["1111111111", "1111111111", "2222222222"],
            "org_npi": ["3333333333", "4444444444", "3333333333"],
            "org_name": ["Group A", "Group B", "Group A"],
        })
        _g2 = _RA.ReassignmentGraph().fit(_pecos, name_col="org_name", source="test_pecos")
        ok &= _check("v49 reassignment: fits from a PECOS reassignment table with "
                     "org names",
                     _g2.orgs_for_individual("1111111111") == ["3333333333", "4444444444"]
                     and _g2.org_names.get("3333333333") == "Group A")
    except Exception as _e:
        import traceback; traceback.print_exc()
        ok &= _check(f"v49 checks run ({type(_e).__name__}: {_e})", False)

    print(f"\n{'ALL CHECKS PASSED' if ok else 'SOME CHECKS FAILED'}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
