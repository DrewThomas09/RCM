"""Multi-sheet Excel export for a :class:`DealAnalysisPacket`.

Partners have been asking for Excel because their downstream workflows
(client decks, LP updates, IC memos) all live in Excel. The CSV export
works but loses structure — one flat sheet, no conditional formatting,
no hyperlinks between tabs, no per-sheet headers. ``render_deal_xlsx``
produces a six-sheet workbook that:

1. RCM Profile — one row per metric, conditional-format coloured vs
   the benchmark P50 (green = better, amber = near-benchmark, red =
   worse), plus current/P25/P50/P75/source/confidence columns.
2. EBITDA Bridge — per-lever impact table + a Waterfall chart so the
   numbers render in Excel exactly as the workbench HTML waterfall.
3. Monte Carlo — P5/P10/P25/P50/P75/P90/P95 summary + histogram data
   the analyst can re-chart in whatever format they like.
4. Risk Flags — severity-sorted table, one row per flag.
5. Raw Data — same columns as the CSV export (kept for compatibility
   with scripts that parse the CSV).
6. Audit — packet hash, run_id, generated_at, model version, analyst
   overrides, completeness grade. Same shape as the AuditFooter used
   in the HTML/PPTX renderers so partners chase numbers back to the
   same run id.

Optional dep: ``openpyxl`` (pinned in ``pyproject.toml``). When it's
missing we raise ``ImportError`` — callers handle the fallback to
:meth:`PacketRenderer.render_raw_data_csv`.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..analysis.packet import (
    PACKET_SCHEMA_VERSION,
    DealAnalysisPacket,
)

logger = logging.getLogger(__name__)


# ── Conditional-format colors (Excel hex without ``#``) ────────────
# Match the workbench palette so the xlsx visually matches the UI.
_FILL_GOOD = "D1FAE5"      # pale green — better than P50
_FILL_NEUTRAL = "FEF3C7"   # pale amber — within ±5% of P50
_FILL_BAD = "FEE2E2"       # pale red — worse than P50

_HEADER_FILL = "1F4E78"    # Chartis blue
_HEADER_FONT = "FFFFFF"

_SEV_COLORS = {
    "CRITICAL": "DC2626",
    "HIGH":     "F59E0B",
    "MEDIUM":   "EAB308",
    "LOW":      "64748B",
}

_SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}


# ── Helpers ────────────────────────────────────────────────────────

def _openpyxl_or_raise():
    """Import openpyxl modules; raise a partner-friendly ImportError.

    Kept in a helper so the ``xlsx_renderer`` module itself can still
    be imported on environments without openpyxl — the error only
    surfaces when the caller actually asks for an xlsx build.
    """
    try:
        from openpyxl import Workbook        # noqa: F401
        from openpyxl.chart import BarChart, Reference   # noqa: F401
        from openpyxl.styles import Font, PatternFill, Alignment   # noqa: F401
        from openpyxl.utils import get_column_letter   # noqa: F401
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for xlsx export. "
            "Install with `pip install openpyxl` or "
            "`pip install rcm-mc[exports]`."
        ) from exc


def _direction_is_better(metric_key: str) -> str:
    """Return ``"lower"`` or ``"higher"`` for the conditional-format
    logic. We keep the list local so xlsx doesn't pull a heavier
    dependency just to read the registry.
    """
    lower_is_better = {
        "denial_rate", "initial_denial_rate", "final_denial_rate",
        "days_in_ar", "ar_over_90_pct", "cost_to_collect",
        "discharged_not_final_billed_days", "bad_debt",
    }
    return "lower" if metric_key in lower_is_better else "higher"


def _benchmark_fill(
    metric_key: str, current: Optional[float], p50: Optional[float],
) -> Optional[str]:
    """Pick green / amber / red based on current vs P50."""
    if current is None or p50 is None or p50 == 0:
        return None
    direction = _direction_is_better(metric_key)
    if direction == "lower":
        if current <= p50 * 0.95:
            return _FILL_GOOD
        if current <= p50 * 1.05:
            return _FILL_NEUTRAL
        return _FILL_BAD
    if current >= p50 * 1.05:
        return _FILL_GOOD
    if current >= p50 * 0.95:
        return _FILL_NEUTRAL
    return _FILL_BAD


def _apply_header_row(ws, row_idx: int, cols: List[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.fill = PatternFill(
            start_color=_HEADER_FILL, end_color=_HEADER_FILL,
            fill_type="solid",
        )
        cell.font = Font(color=_HEADER_FONT, bold=True, size=11)
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws, cols: List[str], max_width: int = 28) -> None:
    from openpyxl.utils import get_column_letter
    for i, _ in enumerate(cols, start=1):
        # Walk the column's body cells and pick the widest string.
        width = len(cols[i - 1]) + 2
        for row in ws.iter_rows(min_col=i, max_col=i, min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                cand = len(str(cell.value))
                if cand > width:
                    width = cand
        ws.column_dimensions[get_column_letter(i)].width = min(
            max_width, width + 1,
        )


# ── Sheet builders ─────────────────────────────────────────────────

def _sheet_rcm_profile(wb, packet: DealAnalysisPacket) -> None:
    """Sheet 1 — one row per metric, coloured vs benchmark P50."""
    from openpyxl.styles import PatternFill
    from ..analysis.completeness import RCM_METRIC_REGISTRY

    ws = wb.create_sheet("RCM Profile")
    cols = [
        "metric_key", "display_name", "category",
        "current_value", "source", "quality",
        "benchmark_p25", "benchmark_p50", "benchmark_p75",
        "delta_vs_p50", "delta_pct",
    ]
    _apply_header_row(ws, 1, cols)
    row_idx = 2
    for key, pm in (packet.rcm_profile or {}).items():
        meta = RCM_METRIC_REGISTRY.get(key) or {}
        current = float(pm.value) if pm is not None else None
        p25 = meta.get("benchmark_p25")
        p50 = meta.get("benchmark_p50")
        p75 = meta.get("benchmark_p75")
        delta = None
        delta_pct = None
        if current is not None and p50 is not None:
            delta = current - float(p50)
            delta_pct = (delta / float(p50)) if p50 else None
        source = pm.source.value if pm is not None and hasattr(pm.source, "value") else ""
        # ``ProfileMetric`` doesn't carry a ``confidence`` field — the
        # closest analogue is ``quality`` (high/medium/low). We surface
        # that string so partners can sort on it without pretending
        # we've got a numeric confidence score.
        conf = getattr(pm, "quality", None) if pm is not None else None
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2,
                value=meta.get("display_name") or key)
        ws.cell(row=row_idx, column=3, value=meta.get("category") or "")
        ws.cell(row=row_idx, column=4, value=current)
        ws.cell(row=row_idx, column=5, value=source)
        ws.cell(row=row_idx, column=6, value=conf)
        ws.cell(row=row_idx, column=7, value=p25)
        ws.cell(row=row_idx, column=8, value=p50)
        ws.cell(row=row_idx, column=9, value=p75)
        ws.cell(row=row_idx, column=10, value=delta)
        ws.cell(row=row_idx, column=11, value=delta_pct)

        fill_color = _benchmark_fill(key, current, p50)
        if fill_color:
            fill = PatternFill(
                start_color=fill_color, end_color=fill_color,
                fill_type="solid",
            )
            for col in (4, 10, 11):  # current, delta, delta_pct
                ws.cell(row=row_idx, column=col).fill = fill
        row_idx += 1
    ws.freeze_panes = "A2"
    _autosize(ws, cols)


def _sheet_bridge(wb, packet: DealAnalysisPacket) -> None:
    """Sheet 2 — per-lever impact table + a BarChart approximating the
    workbench waterfall."""
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("EBITDA Bridge")
    cols = [
        "metric_key", "current_value", "target_value",
        "revenue_impact", "cost_impact", "ebitda_impact",
        "margin_impact_bps", "working_capital_impact",
    ]
    _apply_header_row(ws, 1, cols)

    impacts = (packet.ebitda_bridge.per_metric_impacts or [])
    row_idx = 2
    for imp in impacts:
        ws.cell(row=row_idx, column=1, value=imp.metric_key)
        ws.cell(row=row_idx, column=2, value=float(imp.current_value))
        ws.cell(row=row_idx, column=3, value=float(imp.target_value))
        ws.cell(row=row_idx, column=4, value=float(imp.revenue_impact or 0))
        ws.cell(row=row_idx, column=5, value=float(imp.cost_impact or 0))
        ws.cell(row=row_idx, column=6, value=float(imp.ebitda_impact or 0))
        ws.cell(row=row_idx, column=7,
                value=float(imp.margin_impact_bps or 0))
        ws.cell(row=row_idx, column=8,
                value=float(imp.working_capital_impact or 0))
        row_idx += 1

    # Totals row.
    totals_row = row_idx
    if impacts:
        ws.cell(row=totals_row, column=1, value="TOTAL").font = __import__(
            "openpyxl.styles", fromlist=["Font"],
        ).Font(bold=True)
        for col in (4, 5, 6, 8):
            letter = __import__(
                "openpyxl.utils", fromlist=["get_column_letter"],
            ).get_column_letter(col)
            ws.cell(
                row=totals_row, column=col,
                value=f"=SUM({letter}2:{letter}{totals_row - 1})",
            )
        row_idx += 1

    # Bar chart of ebitda_impact by lever.
    if impacts:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = "EBITDA impact by lever"
        chart.y_axis.title = "Lever"
        chart.x_axis.title = "EBITDA impact ($)"
        data_ref = Reference(
            ws, min_col=6, min_row=1,
            max_row=len(impacts) + 1, max_col=6,
        )
        labels_ref = Reference(
            ws, min_col=1, min_row=2, max_row=len(impacts) + 1,
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, f"J2")

    # Summary below totals.
    summary_start = row_idx + 1
    ws.cell(row=summary_start, column=1, value="Current EBITDA").font = (
        __import__("openpyxl.styles", fromlist=["Font"]).Font(bold=True)
    )
    ws.cell(row=summary_start, column=2,
            value=float(packet.ebitda_bridge.current_ebitda or 0))
    ws.cell(row=summary_start + 1, column=1, value="Target EBITDA")
    ws.cell(row=summary_start + 1, column=2,
            value=float(packet.ebitda_bridge.target_ebitda or 0))
    ws.cell(row=summary_start + 2, column=1, value="New margin")
    ws.cell(row=summary_start + 2, column=2,
            value=float(packet.ebitda_bridge.new_ebitda_margin or 0))
    _autosize(ws, cols)


def _sheet_monte_carlo(wb, packet: DealAnalysisPacket) -> None:
    """Sheet 3 — summary stats + histogram bins if available."""
    ws = wb.create_sheet("Monte Carlo")
    cols = ["percentile", "ebitda_impact", "moic", "irr"]
    _apply_header_row(ws, 1, cols)

    sim = packet.simulation
    if sim is None:
        ws.cell(row=2, column=1, value="(no simulation on packet)")
        _autosize(ws, cols)
        return

    ps_ebitda = sim.ebitda_uplift
    ps_moic = sim.moic
    ps_irr = sim.irr
    percentiles = ["P10", "P25", "P50", "P75", "P90"]
    values = [
        (ps_ebitda.p10, ps_moic.p10, ps_irr.p10),
        (ps_ebitda.p25, ps_moic.p25, ps_irr.p25),
        (ps_ebitda.p50, ps_moic.p50, ps_irr.p50),
        (ps_ebitda.p75, ps_moic.p75, ps_irr.p75),
        (ps_ebitda.p90, ps_moic.p90, ps_irr.p90),
    ]
    for i, (label, (eb, moic, irr)) in enumerate(
        zip(percentiles, values), start=2,
    ):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=float(eb or 0))
        ws.cell(row=i, column=3, value=float(moic or 0))
        ws.cell(row=i, column=4, value=float(irr or 0))

    # Metadata block.
    start = 9
    ws.cell(row=start, column=1, value="n_simulations")
    ws.cell(row=start, column=2, value=int(sim.n_sims or 0))
    ws.cell(row=start + 1, column=1, value="seed")
    ws.cell(row=start + 1, column=2, value=int(sim.seed or 0))
    ws.cell(row=start + 2, column=1, value="p_covenant_breach")
    ws.cell(row=start + 2, column=2,
            value=float(sim.probability_of_covenant_breach or 0))
    ws.cell(row=start + 3, column=1, value="converged")
    ws.cell(
        row=start + 3, column=2,
        value=bool(
            (sim.convergence_check or {}).get("converged") or False,
        ),
    )

    # v2 simulation histogram data — included when present.
    v2 = packet.v2_simulation or {}
    bins = v2.get("histogram_data") or []
    if bins:
        hist_start = start + 6
        _apply_header_row(
            ws, hist_start, ["bin_edge_low", "bin_edge_high", "count"],
        )
        for i, b in enumerate(bins, start=hist_start + 1):
            ws.cell(row=i, column=1, value=float(b.get("bin_edge_low") or 0))
            ws.cell(row=i, column=2, value=float(b.get("bin_edge_high") or 0))
            ws.cell(row=i, column=3, value=int(b.get("count") or 0))
    _autosize(ws, cols)


def _sheet_risk_flags(wb, packet: DealAnalysisPacket) -> None:
    """Sheet 4 — severity-sorted risk flags with per-row tinting."""
    from openpyxl.styles import PatternFill

    ws = wb.create_sheet("Risk Flags")
    cols = [
        "severity", "category", "title", "detail",
        "trigger_metric", "ebitda_at_risk",
    ]
    _apply_header_row(ws, 1, cols)
    flags = list(packet.risk_flags or [])
    def _sev_value(rf):
        v = rf.severity.value if hasattr(rf.severity, "value") else str(rf.severity)
        return _SEV_ORDER.get(v, 99)
    flags.sort(key=_sev_value)

    for i, rf in enumerate(flags, start=2):
        sev = rf.severity.value if hasattr(rf.severity, "value") else str(rf.severity)
        ws.cell(row=i, column=1, value=sev)
        ws.cell(row=i, column=2, value=rf.category or "")
        ws.cell(row=i, column=3, value=rf.title or "")
        ws.cell(row=i, column=4, value=rf.detail or rf.explanation or "")
        ws.cell(row=i, column=5, value=rf.trigger_metric or "")
        ws.cell(row=i, column=6, value=rf.ebitda_at_risk)

        color = _SEV_COLORS.get(sev)
        if color:
            ws.cell(row=i, column=1).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid",
            )
            ws.cell(row=i, column=1).font = __import__(
                "openpyxl.styles", fromlist=["Font"],
            ).Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    _autosize(ws, cols, max_width=60)


def _sheet_raw_data(wb, packet: DealAnalysisPacket) -> None:
    """Sheet 5 — same columns as the CSV export so anyone parsing the
    CSV path can migrate to xlsx without schema drift."""
    from ..analysis.completeness import RCM_METRIC_REGISTRY

    ws = wb.create_sheet("Raw Data")
    cols = [
        "metric_key", "display_name", "current_value", "source",
        "benchmark_p50", "predicted_value", "ci_low", "ci_high",
        "ebitda_impact", "risk_flags",
    ]
    _apply_header_row(ws, 1, cols)

    impact_by_metric: Dict[str, float] = {
        imp.metric_key: float(imp.ebitda_impact or 0)
        for imp in (packet.ebitda_bridge.per_metric_impacts or [])
    }
    risks_by_metric: Dict[str, List[str]] = {}
    for rf in (packet.risk_flags or []):
        for tm in (rf.trigger_metrics or [rf.trigger_metric] or []):
            if not tm:
                continue
            sev = rf.severity.value if hasattr(rf.severity, "value") else str(rf.severity)
            risks_by_metric.setdefault(tm, []).append(
                f"{sev}: {rf.title or ''}",
            )

    seen: set = set()
    ordered: List[str] = []
    for k in packet.rcm_profile or {}:
        if k not in seen:
            seen.add(k); ordered.append(k)
    for k in packet.predicted_metrics or {}:
        if k not in seen:
            seen.add(k); ordered.append(k)

    for i, k in enumerate(ordered, start=2):
        pm = packet.rcm_profile.get(k)
        pred = packet.predicted_metrics.get(k)
        meta = RCM_METRIC_REGISTRY.get(k) or {}
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=meta.get("display_name") or k)
        ws.cell(row=i, column=3,
                value=float(pm.value) if pm is not None else None)
        ws.cell(row=i, column=4,
                value=(pm.source.value if pm is not None
                       and hasattr(pm.source, "value") else ""))
        ws.cell(row=i, column=5, value=meta.get("benchmark_p50"))
        ws.cell(row=i, column=6,
                value=float(pred.value) if pred is not None else None)
        ws.cell(row=i, column=7,
                value=float(pred.ci_low) if pred is not None
                and pred.ci_low is not None else None)
        ws.cell(row=i, column=8,
                value=float(pred.ci_high) if pred is not None
                and pred.ci_high is not None else None)
        ws.cell(row=i, column=9, value=impact_by_metric.get(k))
        ws.cell(row=i, column=10, value="; ".join(risks_by_metric.get(k, [])))
    ws.freeze_panes = "A2"
    _autosize(ws, cols)


def _sheet_audit(
    wb, packet: DealAnalysisPacket, *, inputs_hash: str,
) -> None:
    """Sheet 6 — packet + run identification for traceability."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Audit")
    cols = ["field", "value"]
    _apply_header_row(ws, 1, cols)

    rows = [
        ("deal_id", packet.deal_id or ""),
        ("deal_name", packet.deal_name or ""),
        ("run_id", packet.run_id or ""),
        ("generated_at",
            packet.generated_at.isoformat() if packet.generated_at else ""),
        ("model_version", packet.model_version or PACKET_SCHEMA_VERSION),
        ("scenario_id", packet.scenario_id or ""),
        ("as_of", packet.as_of.isoformat() if packet.as_of else ""),
        ("input_hash", inputs_hash or "n/a"),
        ("completeness_grade",
            str(getattr(packet.completeness, "grade", "") or "")),
        ("coverage_pct",
            float(getattr(packet.completeness, "coverage_pct", 0) or 0)),
        ("observed_metric_count", len(packet.observed_metrics or {})),
        ("predicted_metric_count", len(packet.predicted_metrics or {})),
        ("analyst_override_count",
            len(packet.analyst_overrides or {})),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Override detail block, one row per override key.
    if packet.analyst_overrides:
        header_row = len(rows) + 3
        _apply_header_row(
            ws, header_row, ["override_key", "override_value"],
        )
        for j, (k, v) in enumerate(
            sorted(packet.analyst_overrides.items()), start=header_row + 1,
        ):
            ws.cell(row=j, column=1, value=str(k))
            ws.cell(row=j, column=2, value=str(v))
    _autosize(ws, cols, max_width=40)


# ── Public entry point ────────────────────────────────────────────

def render_deal_xlsx(
    packet: DealAnalysisPacket,
    out_dir: Path,
    *,
    inputs_hash: str = "",
) -> Path:
    """Render the full 6-sheet workbook and return the output path.

    Raises :class:`ImportError` when ``openpyxl`` is not installed —
    callers (e.g. :class:`PacketRenderer`) should catch and fall back
    to the CSV export.
    """
    _openpyxl_or_raise()
    from openpyxl import Workbook

    wb = Workbook()
    # Drop the default empty sheet — we create named ones below.
    default = wb.active
    wb.remove(default)

    _sheet_rcm_profile(wb, packet)
    _sheet_bridge(wb, packet)
    _sheet_monte_carlo(wb, packet)
    _sheet_risk_flags(wb, packet)
    _sheet_raw_data(wb, packet)
    _sheet_audit(wb, packet, inputs_hash=inputs_hash)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    import re as _re
    stem = _re.sub(
        r"[^A-Za-z0-9_.-]+", "_",
        str(packet.run_id or packet.deal_id or "packet"),
    )
    path = out_dir / f"{stem}.xlsx"
    wb.save(str(path))
    return path
