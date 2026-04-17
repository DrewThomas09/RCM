"""Exit-package generator (Prompt 51).

Produces a zip archive for the exit / sell-side process containing:
- 01_Exit_Memo.html — narrative referencing value creation achieved
- 02_Value_Creation_Summary.xlsx — original targets vs actuals per initiative
- 03_Buyer_Data_Room_Checklist.md — auto-generated document list
- manifest.json

The package pulls from the value creation plan (Prompt 41) and the
predicted-vs-actual comparison (Prompt 43) to build the narrative
automatically. Partners click "Export → Exit Package" on the hold
dashboard and get a download-ready bundle.
"""
from __future__ import annotations

import json
import logging
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _today_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


# ── Sub-renderers ─────────────────────────────────────────────────


def _exit_memo_html(
    deal_id: str,
    plan: Any,
    pva_results: List[Any],
) -> str:
    """Build an HTML narrative memo summarising value creation achieved."""
    title = plan.plan_name if plan else f"Exit Memo — {deal_id}"
    total_target = plan.total_target_ebitda if plan else 0.0

    # Compute achieved total from predicted-vs-actual results.
    achieved_total = 0.0
    achieved_items: List[str] = []
    for pva in pva_results:
        achieved_total += float(pva.actual_now or 0)
        var_pct = float(pva.variance_pct or 0) * 100
        sign = "+" if var_pct >= 0 else ""
        achieved_items.append(
            f"<li><strong>{pva.metric_key}</strong>: "
            f"predicted {pva.predicted_at_diligence:.2f}, "
            f"actual {pva.actual_now:.2f} "
            f"({sign}{var_pct:.1f}%)</li>"
        )

    # Initiative summary from plan.
    init_lines: List[str] = []
    if plan:
        for init in plan.initiatives:
            status = init.status or "not_started"
            init_lines.append(
                f"<li>{init.name} — target EBITDA impact "
                f"${init.target_ebitda_impact:,.2f}, status: {status}</li>"
            )

    return (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<title>{title}</title>"
        "<style>"
        "body{font-family:sans-serif;max-width:900px;margin:0 auto;padding:24px;}"
        "h1{color:#1f4e78;} .num{font-family:monospace;font-variant-numeric:tabular-nums;}"
        "table{border-collapse:collapse;width:100%;margin:16px 0;}"
        "th,td{border:1px solid #ccc;padding:6px 10px;text-align:left;}"
        "th{background:#f0f4f8;}"
        "</style></head>"
        f"<body><h1>{title}</h1>"
        f"<p><em>Generated {_today_str()}</em></p>"
        f"<h2>Value Creation Summary</h2>"
        f"<p>Total target EBITDA at plan inception: "
        f"<span class='num'>${total_target:,.2f}</span></p>"
        f"<h3>Initiative Progress</h3>"
        f"<ul>{''.join(init_lines) if init_lines else '<li>No initiatives on plan.</li>'}</ul>"
        f"<h3>Predicted vs Actual Performance</h3>"
        f"<ul>{''.join(achieved_items) if achieved_items else '<li>No actuals available.</li>'}</ul>"
        f"<h3>Conclusion</h3>"
        f"<p>The value creation plan targeted ${total_target:,.2f} in EBITDA improvements "
        f"across {len(plan.initiatives) if plan else 0} initiatives. "
        f"Actual performance data covers {len(pva_results)} metric(s).</p>"
        "</body></html>"
    )


def _value_creation_xlsx(
    plan: Any,
    pva_results: List[Any],
) -> bytes:
    """Build an xlsx workbook: original targets vs actuals per initiative.

    Returns raw bytes suitable for writing into a zip.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Value Creation Summary"

    # Header row.
    headers = [
        "Initiative", "Lever", "Target Value", "Current Value",
        "Target EBITDA Impact", "Status", "Actual (PvA)", "Variance %",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill

    # Build a lookup of metric_key -> PvA for matching.
    pva_map: Dict[str, Any] = {}
    for pva in pva_results:
        pva_map[pva.metric_key] = pva

    row = 2
    initiatives = plan.initiatives if plan else []
    for init in initiatives:
        pva = pva_map.get(init.lever_key)
        actual_val = pva.actual_now if pva else None
        var_pct = (pva.variance_pct * 100) if pva else None

        ws.cell(row=row, column=1, value=init.name)
        ws.cell(row=row, column=2, value=init.lever_key)
        ws.cell(row=row, column=3, value=init.target_value)
        ws.cell(row=row, column=4, value=init.current_value)
        ws.cell(row=row, column=5, value=init.target_ebitda_impact)
        ws.cell(row=row, column=6, value=init.status)
        ws.cell(row=row, column=7, value=actual_val)
        ws.cell(row=row, column=8, value=var_pct)
        row += 1

    # Auto-width columns.
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, row)
        ) if row > 1 else 10
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            max_len + 2, 12,
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _buyer_data_room_checklist_md(
    deal_id: str,
    plan: Any,
) -> str:
    """Auto-generated checklist of documents a buyer would expect."""
    lines = [
        f"# Buyer Data Room Checklist — {deal_id}",
        "",
        f"Generated {_today_str()}",
        "",
        "## Financial Documents",
        "- [ ] Audited financial statements (last 3 years)",
        "- [ ] Monthly P&L trailing 12 months",
        "- [ ] Revenue cycle KPI dashboard",
        "- [ ] Accounts receivable aging schedule",
        "- [ ] Payer contract summaries",
        "",
        "## Operational Documents",
        "- [ ] Organizational chart",
        "- [ ] Key employee agreements",
        "- [ ] Technology systems inventory",
        "- [ ] Compliance program documentation",
        "",
        "## Value Creation Evidence",
    ]
    if plan:
        for init in plan.initiatives:
            lines.append(f"- [ ] {init.name} — supporting data / evidence")
    else:
        lines.append("- [ ] (No value creation plan available)")

    lines.extend([
        "",
        "## Legal & Regulatory",
        "- [ ] Corporate governance documents",
        "- [ ] Material contracts list",
        "- [ ] Litigation summary",
        "- [ ] Regulatory licenses and certifications",
        "",
        "## Diligence Reports",
        "- [ ] Quality of Earnings report",
        "- [ ] Exit memo (this package)",
        "- [ ] Value creation summary workbook",
        "",
    ])
    return "\n".join(lines)


def _build_manifest(
    deal_id: str,
    files: Dict[str, str],
) -> str:
    """JSON manifest listing all files in the package."""
    return json.dumps(
        {
            "deal_id": deal_id,
            "generated_at": _utcnow_iso(),
            "package_type": "exit",
            "files": files,
        },
        indent=2,
    )


# ── Public entry ──────────────────────────────────────────────────


def generate_exit_package(
    store: Any,
    deal_id: str,
    *,
    out_dir: Optional[Path] = None,
) -> Path:
    """Produce a zip with exit-package documents + manifest.

    Returns the zip path. Callers serve it as a download or persist
    it to disk.
    """
    from ..pe.value_creation_plan import load_latest_plan
    from ..pe.predicted_vs_actual import compute_predicted_vs_actual

    plan = load_latest_plan(store, deal_id)
    pva_results = compute_predicted_vs_actual(store, deal_id)

    if out_dir is None:
        out_dir = Path(".")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    zip_name = f"{deal_id}_exit_package.zip"
    zip_path = out_dir / zip_name
    manifest_entries: Dict[str, str] = {}

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        # 01: Exit Memo.
        memo_html = _exit_memo_html(deal_id, plan, pva_results)
        z.writestr("01_Exit_Memo.html", memo_html)
        manifest_entries["01_Exit_Memo.html"] = "Exit memo with value creation narrative"

        # 02: Value Creation Summary xlsx.
        xlsx_bytes = _value_creation_xlsx(plan, pva_results)
        z.writestr("02_Value_Creation_Summary.xlsx", xlsx_bytes)
        manifest_entries["02_Value_Creation_Summary.xlsx"] = (
            "Original targets vs actuals per initiative"
        )

        # 03: Buyer Data Room Checklist.
        checklist_md = _buyer_data_room_checklist_md(deal_id, plan)
        z.writestr("03_Buyer_Data_Room_Checklist.md", checklist_md)
        manifest_entries["03_Buyer_Data_Room_Checklist.md"] = (
            "Auto-generated buyer data room document checklist"
        )

        # manifest.json
        manifest = _build_manifest(deal_id, manifest_entries)
        z.writestr("manifest.json", manifest)

    logger.info("Exit package written to %s", zip_path)
    return zip_path
