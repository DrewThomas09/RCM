"""EBITDA Bridge Excel Export — structured workbook for IC workflow.

Creates a multi-sheet Excel workbook from the EBITDA bridge:
  Sheet 1: Bridge Summary (levers, current→target, impact, provenance)
  Sheet 2: Returns Sensitivity (entry × exit multiple grid)
  Sheet 3: Achievement Sensitivity (50%/75%/100%/120%)
  Sheet 4: Timing Curve (monthly ramp per lever)
  Sheet 5: Peer Context (hospital vs P25/P50/P75)

Partners paste these directly into IC decks.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

import numpy as np


def export_bridge_xlsx(
    bridge: Dict[str, Any],
    hospital_name: str = "",
    ccn: str = "",
    returns_grid: Optional[List[Dict]] = None,
    peer_context: Optional[List[Dict]] = None,
) -> bytes:
    """Export the EBITDA bridge as an Excel workbook. Returns bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    money_fmt = '#,##0'
    pct_fmt = '0.0%'
    num_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    green_font = Font(name="Calibri", size=10, bold=True, color="2ECC71")
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def _style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Bridge Summary ──
    ws1 = wb.active
    ws1.title = "Bridge Summary"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 16
    ws1.column_dimensions["G"].width = 12
    ws1.column_dimensions["H"].width = 14

    # Title
    ws1["A1"] = f"EBITDA Bridge — {hospital_name}"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws1["A2"] = f"CCN {ccn} | Net Revenue: ${bridge.get('net_revenue', 0)/1e6:.1f}M"
    ws1["A2"].font = Font(name="Calibri", size=10, color="666666")

    # Headers
    headers = ["Lever", "Current", "Target", "Revenue Impact", "Cost Impact",
               "EBITDA Impact", "Margin (bps)", "Ramp (mo)"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=4, column=col, value=h)
    _style_header(ws1, 4, len(headers))

    # Lever rows
    row = 5
    for lev in bridge.get("levers", []):
        if lev.get("ebitda_impact", 0) == 0:
            continue
        ws1.cell(row=row, column=1, value=lev["name"]).font = num_font
        ws1.cell(row=row, column=2, value=lev["current"]).font = num_font
        ws1.cell(row=row, column=2).number_format = pct_fmt if lev["current"] < 2 else money_fmt
        ws1.cell(row=row, column=3, value=lev["target"]).font = num_font
        ws1.cell(row=row, column=3).number_format = pct_fmt if lev["target"] < 2 else money_fmt
        ws1.cell(row=row, column=4, value=lev["revenue_impact"]).font = num_font
        ws1.cell(row=row, column=4).number_format = money_fmt
        ws1.cell(row=row, column=5, value=lev["cost_impact"]).font = num_font
        ws1.cell(row=row, column=5).number_format = money_fmt
        ws1.cell(row=row, column=6, value=lev["ebitda_impact"]).font = green_font
        ws1.cell(row=row, column=6).number_format = money_fmt
        ws1.cell(row=row, column=7, value=lev["margin_bps"]).font = num_font
        ws1.cell(row=row, column=8, value=lev["ramp_months"]).font = num_font
        for c in range(1, 9):
            ws1.cell(row=row, column=c).border = thin_border
        row += 1

    # Total row
    ws1.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws1.cell(row=row, column=4, value=bridge.get("total_revenue_impact", 0)).font = bold_font
    ws1.cell(row=row, column=4).number_format = money_fmt
    ws1.cell(row=row, column=5, value=bridge.get("total_cost_impact", 0)).font = bold_font
    ws1.cell(row=row, column=5).number_format = money_fmt
    ws1.cell(row=row, column=6, value=bridge.get("total_ebitda_impact", 0)).font = Font(name="Calibri", size=11, bold=True, color="2ECC71")
    ws1.cell(row=row, column=6).number_format = money_fmt
    ws1.cell(row=row, column=7, value=bridge.get("margin_improvement_bps", 0)).font = bold_font

    # Summary below
    row += 2
    summaries = [
        ("Current EBITDA", bridge.get("current_ebitda", 0)),
        ("RCM Uplift", bridge.get("total_ebitda_impact", 0)),
        ("Pro Forma EBITDA", bridge.get("new_ebitda", 0)),
        ("Current Margin", bridge.get("current_margin", 0)),
        ("Pro Forma Margin", bridge.get("new_margin", 0)),
        ("Working Capital Released", bridge.get("total_wc_released", 0)),
    ]
    for label, val in summaries:
        ws1.cell(row=row, column=1, value=label).font = bold_font
        cell = ws1.cell(row=row, column=2, value=val)
        cell.font = bold_font
        if "Margin" in label:
            cell.number_format = pct_fmt
        else:
            cell.number_format = money_fmt
        row += 1

    # ── Sheet 2: Returns Grid ──
    if returns_grid:
        ws2 = wb.create_sheet("Returns Sensitivity")
        ws2.column_dimensions["A"].width = 16
        ws2["A1"] = "Returns Sensitivity (IRR / MOIC)"
        ws2["A1"].font = Font(name="Calibri", size=12, bold=True)
        ws2["A2"] = "5-year hold, 5.5x leverage, 3% organic growth"
        ws2["A2"].font = Font(name="Calibri", size=9, color="666666")

        # Extract unique entry/exit multiples
        entry_ms = sorted(set(g["entry_multiple"] for g in returns_grid))
        exit_ms = sorted(set(g["exit_multiple"] for g in returns_grid))

        # Headers
        ws2.cell(row=4, column=1, value="Entry \\ Exit").font = header_font
        ws2.cell(row=4, column=1).fill = header_fill
        for j, xm in enumerate(exit_ms, 2):
            ws2.cell(row=4, column=j, value=f"{xm:.1f}x").font = header_font
            ws2.cell(row=4, column=j).fill = header_fill
            ws2.column_dimensions[chr(64 + j)].width = 14

        for i, em in enumerate(entry_ms, 5):
            ws2.cell(row=i, column=1, value=f"{em:.1f}x").font = bold_font
            for j, xm in enumerate(exit_ms, 2):
                cell_data = next((g for g in returns_grid
                                   if g["entry_multiple"] == em and g["exit_multiple"] == xm), None)
                if cell_data:
                    val = f'{cell_data["irr"]:.1%} / {cell_data["moic"]:.2f}x'
                    cell = ws2.cell(row=i, column=j, value=val)
                    cell.font = num_font
                    cell.alignment = Alignment(horizontal="center")

    # ── Sheet 3: Achievement Sensitivity ──
    ws3 = wb.create_sheet("Achievement")
    ws3.column_dimensions["A"].width = 28
    ws3["A1"] = "Achievement Sensitivity"
    ws3["A1"].font = Font(name="Calibri", size=12, bold=True)

    ach_headers = ["Lever", "50%", "75%", "100%", "120%"]
    for col, h in enumerate(ach_headers, 1):
        ws3.cell(row=3, column=col, value=h)
        ws3.column_dimensions[chr(64 + col)].width = 16
    _style_header(ws3, 3, len(ach_headers))

    row = 4
    for lev in bridge.get("levers", []):
        if lev.get("ebitda_impact", 0) == 0:
            continue
        ws3.cell(row=row, column=1, value=lev["name"]).font = num_font
        for j, pct in enumerate([50, 75, 100, 120], 2):
            cell = ws3.cell(row=row, column=j, value=lev["ebitda_impact"] * pct / 100)
            cell.font = num_font
            cell.number_format = money_fmt
        row += 1

    # Totals
    ws3.cell(row=row, column=1, value="TOTAL").font = bold_font
    total_impact = bridge.get("total_ebitda_impact", 0)
    for j, pct in enumerate([50, 75, 100, 120], 2):
        cell = ws3.cell(row=row, column=j, value=total_impact * pct / 100)
        cell.font = bold_font
        cell.number_format = money_fmt

    # ── Sheet 4: Timing ──
    ws4 = wb.create_sheet("Timing Curve")
    ws4.column_dimensions["A"].width = 28
    ws4["A1"] = "Implementation Timing"
    ws4["A1"].font = Font(name="Calibri", size=12, bold=True)

    months = [0, 3, 6, 9, 12, 18, 24, 36]
    t_headers = ["Lever"] + [f"M{m}" for m in months]
    for col, h in enumerate(t_headers, 1):
        ws4.cell(row=3, column=col, value=h)
        ws4.column_dimensions[chr(64 + col)].width = 12
    _style_header(ws4, 3, len(t_headers))

    row = 4
    for lev in bridge.get("levers", []):
        if lev.get("ebitda_impact", 0) == 0:
            continue
        ws4.cell(row=row, column=1, value=lev["name"]).font = num_font
        for j, m in enumerate(months, 2):
            ramp = lev.get("ramp_months", 12)
            pct = min(1.0, m / ramp) if ramp > 0 else 1.0
            cell = ws4.cell(row=row, column=j, value=lev["ebitda_impact"] * pct)
            cell.font = num_font
            cell.number_format = money_fmt
        row += 1

    # ── Save to bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
