"""Polish helpers for ``diligence_workbook.xlsx``.

Turns a pandas-dumped workbook into something a partner opens in Excel
without wincing. Handles:

- Header styling (white-on-blue, bold, centered, frozen)
- Column-specific number formats (currency, percent, integer)
- Source-tag coloring in the Assumptions tab
- Percentile highlighting in the Peer Percentiles tab
- Cover sheet with target metadata and table of contents
- Reasonable column widths

Kept as a sibling module so ``_bundle.py`` stays focused on orchestration.
The styling surface is testable (every helper takes a worksheet + columns
and mutates it in place) and reversible — nothing here changes the
underlying data, only its presentation.
"""
from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ── Color palette ───────────────────────────────────────────────────────────

# Dark blue from Chartis-ish palette; muted enough to print well in B&W.
_HEADER_BG_HEX = "1F4E78"
_HEADER_FG_HEX = "FFFFFF"

# Semantic fills for source tags (Assumptions tab)
_OBSERVED_HEX = "C6EFCE"   # pastel green
_PRIOR_HEX = "FFEB9C"      # pastel amber
_ASSUMED_HEX = "FFC7CE"    # pastel red

# Percentile highlights (Peer Percentiles tab): extreme outliers only
_PCT_HIGH_HEX = "C6EFCE"   # ≥75th percentile: high (usually "bigger/stronger")
_PCT_LOW_HEX = "FFC7CE"    # ≤25th percentile: low


HEADER_FONT = Font(color=_HEADER_FG_HEX, bold=True, size=11, name="Calibri")
HEADER_FILL = PatternFill("solid", fgColor=_HEADER_BG_HEX)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

COVER_TITLE_FONT = Font(bold=True, size=20, color=_HEADER_BG_HEX, name="Calibri")
COVER_LABEL_FONT = Font(bold=True, size=11, color="595959")
COVER_TOC_HEADER = Font(bold=True, size=13, color=_HEADER_BG_HEX)
COVER_TOC_TAB = Font(bold=True, size=11)

_SOURCE_FILLS: Dict[str, PatternFill] = {
    "observed": PatternFill("solid", fgColor=_OBSERVED_HEX),
    "prior":    PatternFill("solid", fgColor=_PRIOR_HEX),
    "assumed":  PatternFill("solid", fgColor=_ASSUMED_HEX),
}

_PCT_HIGH_FILL = PatternFill("solid", fgColor=_PCT_HIGH_HEX)
_PCT_LOW_FILL = PatternFill("solid", fgColor=_PCT_LOW_HEX)


# ── Column name → Excel number format ─────────────────────────────────────

# Straight-match formats. Currency uses "$#,##0" with parens/red for negatives
# on net_income specifically (the one KPI where sign matters for partners).
_MONEY_FMT = '"$"#,##0;[Red]-"$"#,##0'
_MONEY_WITH_SIGN_FMT = '"$"#,##0;[Red]-"$"#,##0'  # same; explicit for readability
_INT_FMT = "#,##0"
_PERCENT_0_1_FMT = "0.0%"   # value is 0.0-1.0 (store ×1)
_PERCENT_0_100_FMT = '0.0"%"'  # value is already 0-100 (store ×100)
_SCORE_FMT = "0.000"

NUMBER_FORMATS: Dict[str, str] = {
    # ── Money columns ────────
    "net_patient_revenue":       _MONEY_FMT,
    "gross_patient_revenue":     _MONEY_FMT,
    "operating_expenses":        _MONEY_FMT,
    "net_income":                _MONEY_WITH_SIGN_FMT,
    "contractual_allowances":    _MONEY_FMT,
    "annual_revenue":            _MONEY_FMT,
    "avg_claim_dollars":         _MONEY_FMT,
    "mean":                      _MONEY_FMT,
    "median":                    _MONEY_FMT,
    "p10":                       _MONEY_FMT,
    "p90":                       _MONEY_FMT,
    "target":                    _MONEY_FMT,
    "peer_p10":                  _MONEY_FMT,
    "peer_median":               _MONEY_FMT,
    "peer_p90":                  _MONEY_FMT,
    "uplift_oat":                _MONEY_FMT,
    "ebitda_drag_mean":          _MONEY_FMT,
    "ebitda_drag_p10":           _MONEY_FMT,
    "ebitda_drag_p90":           _MONEY_FMT,
    "remaining_drag":            _MONEY_FMT,
    "target_value":              _MONEY_FMT,

    # ── Integer counts ────────
    "beds":                      _INT_FMT,
    "bed_days_available":        _INT_FMT,
    "medicare_days":             _INT_FMT,
    "medicaid_days":             _INT_FMT,
    "total_patient_days":        _INT_FMT,

    # ── Percent (stored 0.0-1.0) ────────
    "medicare_day_pct":          _PERCENT_0_1_FMT,
    "medicaid_day_pct":          _PERCENT_0_1_FMT,
    "revenue_share":             _PERCENT_0_1_FMT,
    "idr_mean":                  _PERCENT_0_1_FMT,
    "fwr_mean":                  _PERCENT_0_1_FMT,
    "ebitda_margin":             _PERCENT_0_1_FMT,
    "achievement":               _PERCENT_0_1_FMT,
    "actual_blended":            _PERCENT_0_1_FMT,
    "benchmark_blended":         _PERCENT_0_1_FMT,

    # ── Percent stored 0-100 ────────
    "target_percentile":         _PERCENT_0_100_FMT,

    # ── Unitless scores ────────
    "similarity_score":          _SCORE_FMT,
    "progress_ratio":            _SCORE_FMT,
    "median_ramp_months":        _INT_FMT,
}


# ── Core polish helpers (per-sheet) ─────────────────────────────────────────

def _style_header_row(ws: Worksheet) -> None:
    """Bold white-on-dark-blue fill on row 1."""
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    # Header row a bit taller for visual weight
    ws.row_dimensions[1].height = 22


def _freeze_header(ws: Worksheet) -> None:
    """Freeze row 1 so it stays visible when scrolling a large data sheet."""
    ws.freeze_panes = "A2"


def _auto_size_columns(
    ws: Worksheet,
    min_width: int = 10,
    max_width: int = 45,
    header_weight: int = 3,
) -> None:
    """Heuristic column widths. Header length is weighted more than typical data
    lengths so the header never gets truncated.
    """
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        longest = 0
        for i, cell in enumerate(col_cells):
            val = "" if cell.value is None else str(cell.value)
            weight = header_weight if i == 0 else 1
            longest = max(longest, len(val) * weight if i == 0 else len(val))
        ws.column_dimensions[col_letter].width = min(max(min_width, longest + 2), max_width)


def _apply_number_formats(ws: Worksheet, headers: List[str]) -> None:
    """Write per-column Excel number formats based on header name."""
    for idx, col_name in enumerate(headers, start=1):
        fmt = NUMBER_FORMATS.get(col_name)
        if not fmt:
            continue
        col_letter = get_column_letter(idx)
        for row in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = fmt


def apply_sheet_polish(ws: Worksheet) -> None:
    """One-shot: style header, freeze, apply number formats, size columns."""
    if ws.max_row < 1:
        return
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    _style_header_row(ws)
    _freeze_header(ws)
    _apply_number_formats(ws, headers)
    _auto_size_columns(ws)


# ── Sheet-specific conditional fills ────────────────────────────────────────

def apply_source_coloring(ws: Worksheet, source_col_name: str = "source") -> int:
    """Color cells in a ``source`` column by observed/prior/assumed. Returns
    the number of cells that received a fill (for tests / metrics)."""
    if ws.max_row < 2:
        return 0
    source_col_idx: Optional[int] = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").lower() == source_col_name.lower():
            source_col_idx = idx
            break
    if source_col_idx is None:
        return 0
    col_letter = get_column_letter(source_col_idx)
    colored = 0
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        fill = _SOURCE_FILLS.get(str(cell.value or "").lower())
        if fill:
            cell.fill = fill
            colored += 1
    return colored


def apply_peer_percentile_row_formats(
    ws: Worksheet,
    kpi_col_name: str = "kpi",
    value_cols: Tuple[str, ...] = ("target", "peer_p10", "peer_median", "peer_p90"),
) -> int:
    """Per-row number formats on the Peer Percentiles tab.

    The `target` and `peer_*` columns hold heterogeneous value types by row
    (dollars, integers, ratios) once derived KPIs are included. A column-wide
    format would mis-render rows like ``operating_margin = -0.177`` as ``"$0"``.
    This helper writes per-cell ``number_format`` based on each row's KPI.

    Returns the number of cells formatted.
    """
    if ws.max_row < 2:
        return 0

    header = {str(c.value or ""): idx for idx, c in enumerate(ws[1], start=1)}
    if kpi_col_name not in header:
        return 0
    kpi_col_idx = header[kpi_col_name]
    value_col_idxs = [header[name] for name in value_cols if name in header]
    if not value_col_idxs:
        return 0

    formatted = 0
    for row in range(2, ws.max_row + 1):
        kpi = ws.cell(row=row, column=kpi_col_idx).value
        if not isinstance(kpi, str):
            continue
        fmt = NUMBER_FORMATS.get(kpi)
        # Derived ratios aren't shipped per-field in NUMBER_FORMATS — map here
        if fmt is None:
            if kpi == "operating_margin":
                fmt = _PERCENT_0_1_FMT
            elif kpi in ("cost_per_patient_day", "npsr_per_bed"):
                fmt = _MONEY_FMT
            elif kpi == "payer_mix_hhi":
                fmt = _SCORE_FMT
        if fmt is None:
            continue
        for col_idx in value_col_idxs:
            ws.cell(row=row, column=col_idx).number_format = fmt
            formatted += 1
    return formatted


def apply_data_bar(
    ws: Worksheet,
    column_name: str,
    color: str = "4F81BD",
) -> int:
    """Add a solid blue data bar to ``column_name`` spanning the full column.

    Data bars give an at-a-glance visual sense of relative magnitudes in
    numeric columns — used on variance, IRR, MOIC columns where partners
    skim for outliers. Returns 1 if the rule was added, 0 if the column
    wasn't found or the sheet has no data rows.
    """
    from openpyxl.formatting.rule import DataBarRule

    if ws.max_row < 2:
        return 0
    col_idx: Optional[int] = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "") == column_name:
            col_idx = idx
            break
    if col_idx is None:
        return 0

    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    rule = DataBarRule(
        start_type="min", end_type="max",
        color=color, showValue=True,
    )
    ws.conditional_formatting.add(rng, rule)
    return 1


def apply_color_scale(
    ws: Worksheet,
    column_name: str,
    *,
    higher_is_better: bool = True,
) -> int:
    """Red/yellow/green 3-color scale across a column's values.

    - ``higher_is_better=True``: low=red, high=green (e.g., MOIC, IRR)
    - ``higher_is_better=False``: low=green, high=red (e.g., variance,
      leverage where higher is worse)
    """
    from openpyxl.formatting.rule import ColorScaleRule

    if ws.max_row < 2:
        return 0
    col_idx: Optional[int] = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "") == column_name:
            col_idx = idx
            break
    if col_idx is None:
        return 0

    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    if higher_is_better:
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",       # red
            mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
            end_type="max", end_color="63BE7B",           # green
        )
    else:
        rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    ws.conditional_formatting.add(rng, rule)
    return 1


def add_tab_note(ws: Worksheet, note: str) -> None:
    """Insert a muted description row above the header (row 1 becomes row 2).

    Turns opaque tab names into self-explanatory artifacts. ``note`` is
    expected to be short (one line) — wider text wraps within the merged
    cell. Idempotent only in the sense that re-calling inserts *another*
    row, so callers must apply exactly once per sheet, AND apply after
    any other helpers that rely on row 1 containing the header.

    Also shifts the frozen pane one row down so the note row stays
    visible when scrolling.
    """
    ws.insert_rows(1)
    note_cell = ws.cell(row=1, column=1)
    note_cell.value = note
    note_cell.font = Font(italic=True, color="6B7280", size=10)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Merge across the original header span so the note reads clean
    if ws.max_column >= 2:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=ws.max_column,
        )
    ws.row_dimensions[1].height = 22
    # Freeze pane was originally A2 (header row only); after insertion
    # the header is now row 2, data starts at row 3 — adjust.
    ws.freeze_panes = "A3"


def apply_percentile_coloring(
    ws: Worksheet,
    percentile_col_name: str = "target_percentile",
    high_cut: float = 75.0,
    low_cut: float = 25.0,
) -> int:
    """Highlight extremes in a percentile column: ≥``high_cut`` green,
    ≤``low_cut`` amber. Returns cells colored.

    Neutral on direction — the analyst reads meaning from KPI context.
    (NPSR high = bigger hospital; Medicare day% high = different risk profile.)
    """
    if ws.max_row < 2:
        return 0
    pct_col_idx: Optional[int] = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "") == percentile_col_name:
            pct_col_idx = idx
            break
    if pct_col_idx is None:
        return 0
    col_letter = get_column_letter(pct_col_idx)
    colored = 0
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        try:
            v = float(cell.value)
        except (TypeError, ValueError):
            continue
        if v >= high_cut:
            cell.fill = _PCT_HIGH_FILL
            colored += 1
        elif v <= low_cut:
            cell.fill = _PCT_LOW_FILL
            colored += 1
    return colored


# ── Cover sheet ─────────────────────────────────────────────────────────────

_TAB_DESCRIPTIONS: Dict[str, str] = {
    "Summary":             "Headline metrics — mean, median, P10, P90 per metric",
    "Payers":              "Per-payer KPIs (revenue share, IDR, FWR, DAR) with source tags",
    "Assumptions":         "Full source map — every input labeled observed / prior / assumed",
    "Value Drivers":       "OAT attribution uplifts (or correlation sensitivity as fallback)",
    "Stress Tests":        "Adverse-scenario runs and their modeled impact",
    "Action Plan":         "100-day workstreams ranked by EBITDA impact",
    "Plan Pressure Test":  "Classification of management plan targets (conservative/stretch/aggressive/aspirational)",
    "Plan Miss Scenarios": "EBITDA drag at 100/75/50/0% of plan achievement",
    "Peer Percentiles":    "Target's rank against matched CMS peers on each KPI",
    "Peer Set":            "The 15 matched HCRIS peer hospitals with full metrics",
    "Trend Signals":       "Year-over-year deltas on target's HCRIS metrics (first → last fiscal year)",
    "PE Bridge":           "Value-creation waterfall: entry EV → organic / RCM / multiple-expansion → exit EV",
    "PE Returns":          "Base-case MOIC + IRR at deal's hold + exit-multiple assumptions",
    "PE Hold Grid":        "Hold-years × exit-multiple sensitivity: IRR / MOIC across scenarios",
    "PE Covenant":         "Leverage check: actual vs covenant, EBITDA cushion before trip",
    "Lineage":             "Per-metric formula + config-key inputs + source-grade rollup (IC defensibility)",
    "Challenge":           "Reverse solver: assumption moves needed to reach a target EBITDA drag",
}


def _load_grade_info(outdir: str) -> Optional[Dict[str, Any]]:
    """Pull source classification counts from provenance.json if present."""
    path = os.path.join(outdir, "provenance.json")
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            doc = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    sources = doc.get("sources") or {}
    counts = sources.get("counts") or {}
    total = int(counts.get("total") or 0)
    observed = int(counts.get("observed") or 0)
    if total == 0:
        return None
    return {
        "grade": sources.get("grade", "?"),
        "observed": observed,
        "total": total,
        "pct": (observed / total * 100.0) if total else 0.0,
    }


def build_cover_sheet(
    wb: Workbook,
    *,
    hospital_name: Optional[str],
    ccn: Optional[str],
    outdir: str,
    tab_order: Optional[List[str]] = None,
) -> Worksheet:
    """Insert a branded cover sheet at position 0 with target metadata + TOC.

    ``tab_order`` is the ordered list of the workbook's existing data sheets
    (by title). If omitted we use the current workbook order.
    """
    if "Cover" in wb.sheetnames:
        del wb["Cover"]
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72

    # Title row
    ws["A1"] = "RCM Due Diligence Workbook"
    ws["A1"].font = COVER_TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 32

    # Metadata
    gen_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    grade = _load_grade_info(outdir)

    rows: List[Tuple[str, str]] = [
        ("Target hospital:",  hospital_name or "—"),
        ("Medicare CCN:",     ccn or "—"),
        ("Generated:",        gen_at),
    ]
    if grade:
        rows.append((
            "Evidence grade:",
            f"{grade['grade']}  —  {grade['observed']} of {grade['total']} inputs observed "
            f"({grade['pct']:.0f}%)",
        ))

    start = 3
    for i, (label, value) in enumerate(rows):
        r = start + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = COVER_LABEL_FONT
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws[f"B{r}"] = value
        ws[f"B{r}"].alignment = Alignment(vertical="center")

    # Table of contents
    toc_start = start + len(rows) + 2
    ws[f"A{toc_start}"] = "Tabs in this workbook"
    ws[f"A{toc_start}"].font = COVER_TOC_HEADER
    ws.merge_cells(f"A{toc_start}:B{toc_start}")
    ws.row_dimensions[toc_start].height = 22

    tabs = tab_order or [ws_.title for ws_ in wb.worksheets if ws_.title != "Cover"]
    for i, tab in enumerate(tabs):
        r = toc_start + 1 + i
        ws[f"A{r}"] = tab
        ws[f"A{r}"].font = COVER_TOC_TAB
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws[f"B{r}"] = _TAB_DESCRIPTIONS.get(tab, "")
        ws[f"B{r}"].alignment = Alignment(vertical="center", wrap_text=True)

    # Footer note
    footer_row = toc_start + len(tabs) + 3
    ws[f"A{footer_row}"] = "Confidential — prepared for diligence use only."
    ws[f"A{footer_row}"].font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(f"A{footer_row}:B{footer_row}")

    return ws
