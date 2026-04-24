"""Source-format readers.

Each reader is a thin function that returns a list of row dicts +
per-row source metadata. The ingester consumes these uniformly; the
normalization + CCD construction is format-agnostic.

Design note: we deliberately keep readers stdlib-heavy where possible
(no pandas for CSV; use stdlib ``csv``). Excel requires openpyxl
(already a base dependency per ``pyproject.toml``). 837/835 parsers
are a small hand-rolled subset sufficient for the diligence fixtures
— a full X12 parser is out of scope.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

from .ccd import SourceFormat


@dataclass
class RawRow:
    row_number: int
    source_file: str
    source_format: SourceFormat
    values: Dict[str, Any] = field(default_factory=dict)
    note: str = ""              # populated when the reader hits a skip


@dataclass
class ReaderResult:
    rows: List[RawRow] = field(default_factory=list)
    encoding: str = "utf-8"
    header_dropped_lines: int = 0
    footer_dropped_lines: int = 0
    malformed: bool = False
    note: str = ""


# ── Dispatcher ──────────────────────────────────────────────────────

def read_file(path: Path | str) -> ReaderResult:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in (".csv", ".tsv"):
        return read_csv(p)
    if suffix == ".parquet":
        return read_parquet(p)
    if suffix in (".xlsx", ".xlsm"):
        return read_excel(p)
    if suffix == ".edi":
        return read_edi(p)
    return ReaderResult(
        rows=[], note=f"unsupported file type: {suffix}", malformed=True,
    )


# ── CSV / TSV ───────────────────────────────────────────────────────

def read_csv(path: Path) -> ReaderResult:
    """Handles encoding chaos, trailing total rows, and empty headers.

    Encoding strategy: try utf-8, then windows-1252 (the common
    mis-encoding for Excel CSV exports), then latin-1 as a last
    resort. A file with mixed encodings across lines goes line-by-line
    with a best-effort decode.
    """
    raw_bytes = path.read_bytes()
    text, encoding = _decode_with_fallback(raw_bytes)
    if text is None:
        # Mixed-encoding hellscape — decode line by line.
        lines: List[str] = []
        encoding = "mixed"
        for line in raw_bytes.split(b"\n"):
            for enc in ("utf-8", "windows-1252", "latin-1"):
                try:
                    lines.append(line.decode(enc))
                    break
                except UnicodeDecodeError:
                    continue
            else:
                lines.append(line.decode("latin-1", errors="replace"))
        text = "\n".join(lines)

    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    # Pre-scan: drop leading junk lines that aren't CSV (e.g., a
    # "Report: Q4 Claims Extract" banner row).
    all_lines = text.splitlines()
    header_idx = _find_header_line(all_lines, delimiter)
    header_dropped = header_idx
    usable = "\n".join(all_lines[header_idx:])

    reader = csv.DictReader(io.StringIO(usable), delimiter=delimiter)
    rows: List[RawRow] = []
    fmt = SourceFormat.CSV if delimiter == "," else SourceFormat.TSV
    for i, raw in enumerate(reader, start=2):  # +2 for 1-indexed + header
        # Drop rows where every value is empty (trailing blank lines).
        if not any((v or "").strip() for v in raw.values() if isinstance(v, str)):
            continue
        # Drop trailing-total rows — heuristic: first column contains
        # "Total" / "Grand Total" / empty and the rest are numbers or
        # blank. We record the drop in ReaderResult.footer_dropped_lines.
        first_key = next(iter(raw.keys()), "")
        first_val = (raw.get(first_key) or "").strip().lower()
        if first_val in ("total", "grand total", "totals", "subtotal"):
            continue
        rows.append(RawRow(
            row_number=i + header_dropped,
            source_file=str(path.name),
            source_format=fmt,
            values={k: v for k, v in raw.items() if k is not None},
        ))

    footer_dropped = max(0, (len(all_lines) - header_idx - 1) - len(rows))
    return ReaderResult(
        rows=rows, encoding=encoding,
        header_dropped_lines=header_dropped,
        footer_dropped_lines=footer_dropped,
    )


def _decode_with_fallback(raw: bytes) -> Tuple[Optional[str], str]:
    for enc in ("utf-8", "windows-1252", "latin-1"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return None, "mixed"


def _find_header_line(lines: List[str], delimiter: str) -> int:
    """Return the index of the line that looks like a header row.

    A header has multiple delimited fields, most of which are alpha-
    dominant (column names, not numbers). We scan the first 20 lines
    and pick the first one that has ≥3 fields with ≥60% alpha.
    """
    for i, line in enumerate(lines[:20]):
        parts = [p.strip().strip('"') for p in line.split(delimiter)]
        if len(parts) < 3:
            continue
        alpha_frac = sum(1 for p in parts if p and any(c.isalpha() for c in p)) / len(parts)
        if alpha_frac >= 0.6:
            return i
    return 0


# ── Parquet ─────────────────────────────────────────────────────────

def read_parquet(path: Path) -> ReaderResult:
    import pyarrow.parquet as pq
    tbl = pq.read_table(path)
    rows: List[RawRow] = []
    for i, r in enumerate(tbl.to_pylist(), start=1):
        rows.append(RawRow(
            row_number=i,
            source_file=str(path.name),
            source_format=SourceFormat.PARQUET,
            values=dict(r),
        ))
    return ReaderResult(rows=rows, encoding="binary")


# ── Excel ──────────────────────────────────────────────────────────

def read_excel(path: Path) -> ReaderResult:
    """Reads the first worksheet. Handles merged cells by un-merging
    (propagating the top-left value across the span), drops junk
    header rows (non-delimited banner lines), and drops trailing
    total rows."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return ReaderResult(rows=[], malformed=True,
                            note="no active worksheet")

    # Un-merge: copy the top-left value into every merged cell in the
    # range, then remove the merge so iter_rows sees the full grid.
    for mr in list(ws.merged_cells.ranges):
        top = ws.cell(row=mr.min_row, column=mr.min_col).value
        ws.unmerge_cells(str(mr))
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                ws.cell(row=row, column=col).value = top

    # Read all rows, find header.
    all_rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    header_idx = _find_excel_header(all_rows)
    header_dropped = header_idx
    header = [_sanitise_key(v) for v in all_rows[header_idx]] if all_rows else []

    rows: List[RawRow] = []
    for i, r in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not any((v is not None and str(v).strip() != "") for v in r):
            continue
        first_val = str(r[0] or "").strip().lower() if r else ""
        if first_val in ("total", "grand total", "totals", "subtotal"):
            continue
        # Pad / trim to header width.
        r = (r + [None] * len(header))[: len(header)]
        values = {k: v for k, v in zip(header, r) if k}
        rows.append(RawRow(
            row_number=i,
            source_file=str(path.name),
            source_format=SourceFormat.EXCEL,
            values=values,
        ))

    return ReaderResult(
        rows=rows, encoding="binary",
        header_dropped_lines=header_dropped,
    )


def _find_excel_header(all_rows: List[List[Any]]) -> int:
    for i, row in enumerate(all_rows[:20]):
        cells = [str(v).strip() if v is not None else "" for v in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 3:
            continue
        alpha_frac = sum(
            1 for c in non_empty if any(ch.isalpha() for ch in c)
        ) / len(non_empty)
        if alpha_frac >= 0.6:
            return i
    return 0


def _sanitise_key(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ── EDI 837 / 835 ───────────────────────────────────────────────────

# Minimal subset of X12 — pulls just enough per-claim data for the
# fixtures. A production parser (X12 v5010) is out of scope.
_EDI_CLAIM_BEGIN = "CLM"
_EDI_SEGMENT_RE = re.compile(r"([A-Z0-9]+)\*([^~]*)")


@dataclass
class _EdiContext:
    """Per-claim accumulator."""
    claim_id: Optional[str] = None
    claim_amount: Optional[float] = None
    service_date: Optional[str] = None
    service_date_to: Optional[str] = None
    cpt_code: Optional[str] = None
    icd10_primary: Optional[str] = None
    patient_id: Optional[str] = None
    payer: Optional[str] = None
    paid_amount: Optional[float] = None
    status_code: Optional[str] = None  # from CLP in 835


def read_edi(path: Path) -> ReaderResult:
    """Parses a subset of X12 837 (submit) or 835 (remittance).

    We look at CLM (claim) segments for 837 and CLP (claim payment)
    for 835, pulling the common fields needed for Phase 1. A
    truncated file produces a partial read with ``malformed=True``
    so the ingester can surface the risk without dropping what did
    parse.
    """
    text = path.read_text(encoding="latin-1")
    # Segment terminator is usually `~`; element is `*`; sub-element
    # is `:`. Production parsers read ISA to confirm; we assume
    # standard delimiters to keep this compact.
    segments = [s.strip() for s in text.split("~") if s.strip()]

    # Detect 837 vs 835 by looking at ST segment.
    kind = SourceFormat.EDI_837
    for seg in segments[:10]:
        elems = seg.split("*")
        if elems[0] == "ST" and len(elems) > 1:
            if elems[1] == "835":
                kind = SourceFormat.EDI_835
            break

    rows: List[RawRow] = []
    ctx: Optional[_EdiContext] = None
    service_date_hits = 0
    malformed = False
    last_segment_ok = False
    # In X12, NM1*PR (payer) and NM1*QC (patient) commonly precede the
    # CLM segment within a 2000B loop. We buffer them here and apply
    # to the next CLM when it arrives.
    pending_payer: Optional[str] = None
    pending_patient_id: Optional[str] = None

    for row_idx, seg in enumerate(segments, start=1):
        elems = seg.split("*")
        tag = elems[0]
        last_segment_ok = False

        if kind == SourceFormat.EDI_837:
            if tag == "CLM":
                # Flush previous claim.
                if ctx and ctx.claim_id:
                    rows.append(_edi_ctx_to_row(ctx, row_idx, path.name, kind))
                ctx = _EdiContext()
                ctx.claim_id = elems[1] if len(elems) > 1 else None
                try:
                    ctx.claim_amount = float(elems[2]) if len(elems) > 2 else None
                except ValueError:
                    ctx.claim_amount = None
                # Attach any buffered payer / patient from the 2000B
                # loop header that preceded this CLM.
                if pending_payer is not None:
                    ctx.payer = pending_payer
                    pending_payer = None
                if pending_patient_id is not None:
                    ctx.patient_id = pending_patient_id
                    pending_patient_id = None
                last_segment_ok = True
            elif tag == "NM1" and len(elems) > 3:
                # In X12, NM1*PR (payer) and NM1*QC (patient) are loop-
                # header segments that precede their CLM. When an NM1
                # shows up inside an in-flight ctx, it's signalling a
                # NEW 2000B loop — flush the current ctx first, then
                # buffer for the next CLM.
                if elems[1] in ("PR", "QC") and ctx is not None:
                    rows.append(_edi_ctx_to_row(ctx, row_idx, path.name, kind))
                    ctx = None
                if elems[1] == "PR":
                    pending_payer = elems[3]
                elif elems[1] == "QC" and len(elems) > 9:
                    pending_patient_id = elems[9]
                last_segment_ok = True
            elif ctx is not None:
                if tag == "SV1" and len(elems) > 1:
                    # SV1*HC:99213*…
                    parts = elems[1].split(":")
                    if len(parts) > 1:
                        ctx.cpt_code = parts[1]
                    last_segment_ok = True
                elif tag == "DTP" and len(elems) > 3:
                    # DTP*472*D8*20240115  or DTP*472*RD8*20240115-20240117
                    qualifier = elems[1]
                    date_raw = elems[3]
                    if qualifier == "472":  # service date
                        if "-" in date_raw and len(date_raw) >= 17:
                            ctx.service_date = date_raw[:8]
                            ctx.service_date_to = date_raw[9:17]
                        else:
                            ctx.service_date = date_raw
                        service_date_hits += 1
                    last_segment_ok = True
                elif tag == "HI" and len(elems) > 1:
                    # HI*ABK:E11.9 or HI*BK:25000
                    for e in elems[1:]:
                        parts = e.split(":")
                        if len(parts) > 1 and not ctx.icd10_primary:
                            ctx.icd10_primary = parts[1]
                            break
                    last_segment_ok = True
                # NM1 handling moved up — loop-header segments flush
                # the current ctx rather than updating it in place.

        elif kind == SourceFormat.EDI_835:
            if tag == "CLP":
                if ctx and ctx.claim_id:
                    rows.append(_edi_ctx_to_row(ctx, row_idx, path.name, kind))
                ctx = _EdiContext()
                ctx.claim_id = elems[1] if len(elems) > 1 else None
                ctx.status_code = elems[2] if len(elems) > 2 else None
                try:
                    ctx.claim_amount = float(elems[3]) if len(elems) > 3 else None
                    ctx.paid_amount = float(elems[4]) if len(elems) > 4 else None
                except ValueError:
                    pass
                last_segment_ok = True
            elif ctx is not None and tag == "NM1" and len(elems) > 3:
                if elems[1] == "QC":
                    ctx.patient_id = elems[9] if len(elems) > 9 else None
                elif elems[1] == "PR":
                    ctx.payer = elems[3]
                last_segment_ok = True
            elif ctx is not None and tag == "DTM" and len(elems) > 2:
                # DTM*405*YYYYMMDD — adjudication date
                ctx.service_date = elems[2]
                last_segment_ok = True

    # Flush final claim.
    if ctx and ctx.claim_id:
        rows.append(_edi_ctx_to_row(row_idx, 999_999, path.name, kind)
                    if False else _edi_ctx_to_row(ctx, len(segments), path.name, kind))

    # Heuristic for truncated EOF: last segment is not a well-known
    # trailer (SE, GE, IEA) — we may have lost a claim mid-segment.
    last_tag = segments[-1].split("*")[0] if segments else ""
    if last_tag not in ("SE", "GE", "IEA") and not last_segment_ok:
        malformed = True

    return ReaderResult(
        rows=rows, encoding="latin-1",
        malformed=malformed,
        note=("truncated EDI — last segment was not a trailer"
              if malformed else ""),
    )


def _edi_ctx_to_row(
    ctx: _EdiContext, row_idx: int, source_file: str, kind: SourceFormat
) -> RawRow:
    values: Dict[str, Any] = {
        "claim_id": ctx.claim_id,
        "patient_id": ctx.patient_id,
        "cpt_code": ctx.cpt_code,
        "icd10_primary": ctx.icd10_primary,
        "payer": ctx.payer,
        "service_date_from": ctx.service_date,
        "service_date_to": ctx.service_date_to,
        "charge_amount": ctx.claim_amount,
        "paid_amount": ctx.paid_amount,
        "status_code": ctx.status_code,
    }
    return RawRow(
        row_number=row_idx,
        source_file=source_file,
        source_format=kind,
        values=values,
    )
