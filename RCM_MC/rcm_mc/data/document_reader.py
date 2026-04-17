"""Drag-drop extraction of RCM metrics from seller files.

Sellers send a mix of Excel workbooks, CSVs, and the occasional TSV.
Each one has the same conceptual contents (denial rate, AR aging,
payer mix, collection reports) but every vendor uses different
column names. Before Prompt 25 an associate spent 4-8 hours per deal
manually rekeying values from these files.

This module does three things:

1. Walks every sheet of every uploaded file.
2. Maps column headers to registered RCM metric keys using a
   hand-curated alias table (the biggest vendors ship ~100 distinct
   header strings for the same underlying metrics) plus a fuzzy
   fallback.
3. Produces an :class:`ExtractionResult` the packet builder can
   accept wholesale — per-metric extractions, multi-period detection,
   and explicit conflicts when two sheets disagree.

No calibration math here; the goal is accurate *extraction* plus
loud provenance so the analyst always knows where a number came
from (source file, sheet, cell). Optional dep is ``openpyxl`` for
.xlsx/.xlsm; CSV/TSV use stdlib ``csv``.
"""
from __future__ import annotations

import csv
import difflib
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ── Dataclasses ────────────────────────────────────────────────────

@dataclass
class MetricExtraction:
    """One extracted value, with enough provenance to cite in the
    diligence memo."""
    period: str
    value: float
    source_sheet: str = ""
    source_cell: str = ""
    confidence: float = 0.8    # 0.0-1.0, from column-match score

    def to_dict(self) -> Dict[str, Any]:
        return {
            "period": self.period,
            "value": float(self.value),
            "source_sheet": self.source_sheet,
            "source_cell": self.source_cell,
            "confidence": float(self.confidence),
        }


@dataclass
class ConflictItem:
    """Same metric extracted multiple times with divergent values.

    ``values`` lists the contenders; ``recommended`` is the
    highest-confidence pick (or None when no single extraction is
    clearly best). Conflict-resolution UI lets the analyst pick.
    """
    metric_key: str
    values: List[Tuple[str, float, str]] = field(default_factory=list)
    recommended: Optional[float] = None
    reason: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "metric_key": self.metric_key,
            "values": [
                {"source": s, "value": float(v), "period": p}
                for s, v, p in self.values
            ],
            "recommended": (float(self.recommended)
                            if self.recommended is not None else None),
            "reason": self.reason,
        }


@dataclass
class ExtractionResult:
    """Output of :func:`read_seller_file`.

    ``metrics`` maps metric_key → list of extractions. Usually length
    one per metric; multi-period sheets produce one per period.
    ``conflicts`` surfaces metrics the reader couldn't pick a single
    value for. ``raw_columns_mapped`` / ``unmapped_columns`` help the
    analyst tune the alias table over time.
    """
    source_file: str = ""
    metrics: Dict[str, List[MetricExtraction]] = field(default_factory=dict)
    conflicts: List[ConflictItem] = field(default_factory=list)
    raw_columns_mapped: Dict[str, str] = field(default_factory=dict)
    unmapped_columns: List[str] = field(default_factory=list)
    extraction_confidence: float = 0.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "source_file": self.source_file,
            "metrics": {
                k: [ex.to_dict() for ex in v]
                for k, v in self.metrics.items()
            },
            "conflicts": [c.to_dict() for c in self.conflicts],
            "raw_columns_mapped": dict(self.raw_columns_mapped),
            "unmapped_columns": list(self.unmapped_columns),
            "extraction_confidence": float(self.extraction_confidence),
        }


# ── Alias table ────────────────────────────────────────────────────

# Partner-curated vendor aliases. Keys are lowercased / stripped /
# punctuation-removed. When a header doesn't hit this table we fall
# through to a fuzzy match against the registry's ``display_name``.
#
# Curated against header samples from Waystar, Experian, R1 RCM,
# nThrive, Parallon, Epic, Cerner exports partners commonly send.
# Not exhaustive — the fuzzy fallback catches the long tail.
COLUMN_ALIAS_TABLE: Dict[str, str] = {
    # Denials
    "initial denial rate":              "denial_rate",
    "init denial":                      "denial_rate",
    "init denial pct":                  "denial_rate",
    "init denial %":                    "denial_rate",
    "first pass denial rate":           "denial_rate",
    "first pass denial pct":            "denial_rate",
    "front end denial rate":            "denial_rate",
    "initial denials":                  "denial_rate",
    "idr":                              "denial_rate",
    "final denial rate":                "final_denial_rate",
    "final write off rate":             "final_denial_rate",
    "fwr":                              "final_denial_rate",
    "final denial write off":           "final_denial_rate",
    "net denial rate":                  "final_denial_rate",
    "appeal win rate":                  "appeals_overturn_rate",
    "denial overturn rate":             "appeals_overturn_rate",
    "appeals overturn":                 "appeals_overturn_rate",
    "overturn pct":                     "appeals_overturn_rate",
    "coding denial rate":               "coding_denial_rate",
    "auth denial rate":                 "auth_denial_rate",
    "authorization denial rate":        "auth_denial_rate",
    "pre auth denial":                  "auth_denial_rate",
    "eligibility denial rate":          "eligibility_denial_rate",
    "elig denial":                      "eligibility_denial_rate",
    "medical necessity denial rate":    "medical_necessity_denial_rate",
    "timely filing denial rate":        "timely_filing_denial_rate",
    "avoidable denial rate":            "avoidable_denial_pct",
    "avoidable denials":                "avoidable_denial_pct",

    # AR & collections
    "days in ar":                       "days_in_ar",
    "ar days":                          "days_in_ar",
    "a r days":                         "days_in_ar",
    "dar":                              "days_in_ar",
    "clean dar":                        "days_in_ar",
    "discharged not final billed":      "discharged_not_final_billed_days",
    "dnfb days":                        "discharged_not_final_billed_days",
    "dnfb":                             "discharged_not_final_billed_days",
    "ar over 90":                       "ar_over_90_pct",
    "ar 90 plus":                       "ar_over_90_pct",
    "ar over 90 pct":                   "ar_over_90_pct",
    "aged ar 90 plus":                  "ar_over_90_pct",
    "ar 90 percent":                    "ar_over_90_pct",

    # Clean claim / FPR
    "clean claim rate":                 "clean_claim_rate",
    "ccr":                              "clean_claim_rate",
    "first pass resolution":            "first_pass_resolution_rate",
    "first pass resolution rate":       "first_pass_resolution_rate",
    "fpr":                              "first_pass_resolution_rate",
    "first pass yield":                 "first_pass_resolution_rate",

    # Collections / ncr / cost-to-collect
    "net collection rate":              "net_collection_rate",
    "net coll rate":                    "net_collection_rate",
    "ncr":                              "net_collection_rate",
    "gross collection rate":            "gross_collection_rate",
    "gcr":                              "gross_collection_rate",
    "cost to collect":                  "cost_to_collect",
    "ctc":                              "cost_to_collect",
    "cost to collect pct":              "cost_to_collect",

    # Coding / CDI
    "case mix index":                   "case_mix_index",
    "cmi":                              "case_mix_index",
    "coding accuracy":                  "coding_accuracy_rate",
    "coding accuracy rate":             "coding_accuracy_rate",

    # Bad debt
    "bad debt":                         "bad_debt",
    "bad debt pct":                     "bad_debt",
    "bad debt rate":                    "bad_debt",
    "bad debt percent":                 "bad_debt",
}


# Period column name aliases.
_PERIOD_HEADERS = frozenset({
    "period", "month", "date", "quarter", "reporting period",
    "as of", "fiscal period", "fiscal month", "reporting month",
    "time period", "month end",
})


# ── Helpers ────────────────────────────────────────────────────────

def _normalize_header(s: str) -> str:
    """Lowercase, strip, drop non-letter/digit so ``"Init Denial %"``
    and ``"init_denial_pct"`` both collapse to ``"init denial pct"``.
    """
    if not isinstance(s, str):
        return ""
    # Replace non-alphanum with space, collapse whitespace.
    out = re.sub(r"[^a-z0-9]+", " ", s.strip().lower())
    out = re.sub(r"\s+", " ", out).strip()
    # Replace common percent-sign words.
    out = out.replace(" pct", "").replace(" percent", "").strip()
    return out


def _fuzzy_column_match(
    header: str,
    known_metrics: Dict[str, Dict[str, Any]],
    threshold: float = 0.72,
) -> Tuple[Optional[str], float]:
    """Return ``(metric_key, confidence)`` for a column header.

    Tries the alias table first (confidence 0.95); falls back to
    SequenceMatcher on the registry's ``display_name`` with a 0.72
    threshold. Below that we don't guess — unmapped columns land in
    ``ExtractionResult.unmapped_columns`` for the analyst to review.
    """
    if not header or not isinstance(header, str):
        return None, 0.0
    norm = _normalize_header(header)
    if not norm:
        return None, 0.0

    # Alias table is authoritative — hand-curated by partners.
    if norm in COLUMN_ALIAS_TABLE:
        return COLUMN_ALIAS_TABLE[norm], 0.95
    # Also try with underscores stripped further (``"net_coll_rate"``).
    collapsed = norm.replace(" ", "")
    for alias, metric in COLUMN_ALIAS_TABLE.items():
        if alias.replace(" ", "") == collapsed:
            return metric, 0.95

    # Fuzzy fallback: match against the registry's display names.
    best_key: Optional[str] = None
    best_score = 0.0
    for key, meta in known_metrics.items():
        display = _normalize_header(str(meta.get("display_name") or key))
        key_norm = _normalize_header(key.replace("_", " "))
        for candidate in (display, key_norm):
            score = difflib.SequenceMatcher(None, norm, candidate).ratio()
            if score > best_score:
                best_score = score
                best_key = key
    if best_score >= threshold and best_key:
        return best_key, best_score
    return None, best_score


def _parse_value(raw: Any) -> Optional[float]:
    """Coerce a cell into a float or ``None``. Accepts ``"12.5%"``,
    ``"$ 1,234"``, ``"45 days"``. Blank/dash returns ``None`` — the
    extractor drops rows where the metric column is empty."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return float(raw)
    if isinstance(raw, (int, float)):
        try:
            f = float(raw)
        except (TypeError, ValueError):
            return None
        if f != f:   # NaN
            return None
        return f
    if not isinstance(raw, str):
        return None
    s = raw.strip()
    if not s or s in ("-", "—", "n/a", "N/A", "NA", "null"):
        return None
    # Strip percent signs, currency, commas, whitespace, trailing units.
    cleaned = s.replace("$", "").replace(",", "").strip()
    cleaned = re.sub(r"\s*(days?|%)\s*$", "", cleaned, flags=re.I).strip()
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1].strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _detect_period(row_values: Dict[str, Any]) -> str:
    """If the row has a column whose *normalized* name is in
    :data:`_PERIOD_HEADERS`, return its value as a string. Otherwise
    return ``""`` (single-period extraction)."""
    for col, v in row_values.items():
        if _normalize_header(col) in _PERIOD_HEADERS and v not in (None, ""):
            return str(v)
    return ""


def _metric_registry() -> Dict[str, Dict[str, Any]]:
    try:
        from ..analysis.completeness import RCM_METRIC_REGISTRY
        return RCM_METRIC_REGISTRY
    except Exception:  # noqa: BLE001
        return {}


# ── CSV/TSV reader ─────────────────────────────────────────────────

def _read_csv(path: Path) -> ExtractionResult:
    """Read a .csv or .tsv. Auto-detect delimiter."""
    registry = _metric_registry()
    result = ExtractionResult(source_file=str(path))
    try:
        # Sniff the delimiter on the first ~4KB so mixed-format files
        # don't break. Fall back to comma on sniffer failure.
        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(fh, dialect=dialect)
            rows = list(reader)
            headers = list(reader.fieldnames or [])
    except FileNotFoundError:
        return result
    except Exception as exc:  # noqa: BLE001
        logger.debug("csv read failed for %s: %s", path, exc)
        return result

    return _process_rows(
        result,
        headers,
        rows,
        registry,
        sheet_name="csv",
        row_number_start=2,
    )


# ── Excel reader ───────────────────────────────────────────────────

def _read_excel(path: Path) -> ExtractionResult:
    """Read every sheet of a .xlsx/.xlsm workbook. Requires openpyxl."""
    registry = _metric_registry()
    result = ExtractionResult(source_file=str(path))
    try:
        from openpyxl import load_workbook   # type: ignore
    except ImportError:
        logger.debug("openpyxl missing; xlsx extraction disabled")
        result.unmapped_columns = ["<openpyxl not installed>"]
        return result
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.debug("openpyxl load failed for %s: %s", path, exc)
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not all_rows:
            continue
        header_idx = _detect_header_row(all_rows, registry)
        if header_idx is None:
            continue
        headers = [
            str(c).strip() if c is not None else ""
            for c in all_rows[header_idx]
        ]
        body = all_rows[header_idx + 1:]
        rows = [
            {h: (row[i] if i < len(row) else None)
             for i, h in enumerate(headers) if h}
            for row in body
        ]
        _process_rows(
            result,
            headers,
            rows,
            registry,
            sheet_name=sheet_name,
            row_number_start=header_idx + 2,
        )
    return result


def _detect_header_row(
    rows: List[List[Any]], registry: Dict[str, Dict[str, Any]],
) -> Optional[int]:
    """Heuristic: the first row where at least two cells map to a
    known metric OR a period column. Keeps us from treating the
    company-logo block at the top of seller reports as headers.

    Threshold 2 rather than 3 so sparse two-column reports (a
    common seller format for ad-hoc extracts) still get picked up;
    logo / title blocks rarely carry two RCM-jargon tokens.

    Returns the zero-based row index, or ``None`` if no row in the
    first 20 meets the bar (skip the sheet in that case).
    """
    for i, row in enumerate(rows[:20]):
        hits = 0
        for cell in row:
            if not isinstance(cell, str):
                continue
            norm = _normalize_header(cell)
            if norm in COLUMN_ALIAS_TABLE or norm in _PERIOD_HEADERS:
                hits += 1
                continue
            # Try a looser match for counting purposes.
            key, score = _fuzzy_column_match(cell, registry)
            if key is not None and score >= 0.75:
                hits += 1
        if hits >= 2:
            return i
    return None


# ── Shared row processing ─────────────────────────────────────────

def _process_rows(
    result: ExtractionResult,
    headers: List[str],
    rows: List[Dict[str, Any]],
    registry: Dict[str, Dict[str, Any]],
    *,
    sheet_name: str,
    row_number_start: int,
) -> ExtractionResult:
    """Map headers → metric keys, then walk the body rows pulling
    values.

    ``row_number_start`` keeps provenance aligned with the original
    file when the header row doesn't start at line 1 (common in
    seller workbooks that begin with a logo/title block).

    Updates ``result`` in place and returns it.
    """
    # Header → metric_key map with confidence.
    col_map: Dict[str, Tuple[str, float]] = {}
    for h in headers:
        if not h:
            continue
        key, conf = _fuzzy_column_match(h, registry)
        if key is not None:
            col_map[h] = (key, conf)
            result.raw_columns_mapped[h] = key
        else:
            if _normalize_header(h) not in _PERIOD_HEADERS:
                if h not in result.unmapped_columns:
                    result.unmapped_columns.append(h)

    if not col_map:
        return result

    # Row walk.
    for row_idx, row in enumerate(rows, start=row_number_start):
        period = _detect_period(row)
        for header, (metric_key, conf) in col_map.items():
            val = _parse_value(row.get(header))
            if val is None:
                continue
            ex = MetricExtraction(
                period=period,
                value=val,
                source_sheet=sheet_name,
                source_cell=f"{header} / row {row_idx}",
                confidence=float(conf),
            )
            result.metrics.setdefault(metric_key, []).append(ex)

    # Update overall confidence: weighted mean of the mapping confidences
    # we hit. Empty → 0.
    if col_map:
        result.extraction_confidence = sum(c for _, c in col_map.values()) / len(col_map)
    return result


# ── Conflict detection ────────────────────────────────────────────

def _detect_conflicts(result: ExtractionResult) -> None:
    """Populate ``result.conflicts`` for metrics where two extractions
    (different sheets, same period) disagree by more than 5%."""
    for metric_key, extractions in result.metrics.items():
        if len(extractions) < 2:
            continue
        # Group by period so multi-period series don't look like
        # conflicts.
        by_period: Dict[str, List[MetricExtraction]] = {}
        for ex in extractions:
            by_period.setdefault(ex.period, []).append(ex)
        for period, items in by_period.items():
            if len(items) < 2:
                continue
            values = [ex.value for ex in items]
            mn, mx = min(values), max(values)
            if mx == 0:
                continue
            spread = (mx - mn) / max(abs(mx), 1.0)
            if spread <= 0.05:
                continue
            # Pick the highest-confidence winner; tie → first one.
            winner = max(items, key=lambda ex: ex.confidence)
            result.conflicts.append(ConflictItem(
                metric_key=metric_key,
                values=[
                    (ex.source_sheet, ex.value, ex.period) for ex in items
                ],
                recommended=winner.value,
                reason=(
                    f"spread {spread * 100:.1f}% across "
                    f"{len(items)} extractions"
                ),
            ))


# ── Public entry ──────────────────────────────────────────────────

def read_seller_file(filepath: Path) -> ExtractionResult:
    """Extract RCM metrics from one Excel / CSV / TSV file.

    Returns an :class:`ExtractionResult` — empty when the file is
    unreadable or carries no recognizable columns. Never raises on
    bad input; the diligence-upload flow needs this to stay resilient
    so one malformed sheet doesn't kill the batch.
    """
    path = Path(filepath)
    if not path.exists():
        return ExtractionResult(source_file=str(path))
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        result = _read_excel(path)
    elif suffix in (".csv", ".tsv", ".txt"):
        result = _read_csv(path)
    else:
        # Try CSV as the fallback for unknown extensions — partners
        # sometimes send ".data" or ".report" files.
        result = _read_csv(path)
    _detect_conflicts(result)
    return result


def read_data_room(directory: Path) -> ExtractionResult:
    """Walk a directory, merge every readable file into a single
    :class:`ExtractionResult`. Used by the ``rcm-mc ingest`` CLI and
    the wizard's batch-upload handler.

    Duplicates across files surface as conflicts rather than being
    silently overwritten.
    """
    directory = Path(directory)
    combined = ExtractionResult(source_file=str(directory))
    if not directory.is_dir():
        return combined
    for path in sorted(directory.iterdir()):
        if path.is_dir():
            continue
        if path.suffix.lower() not in (
            ".xlsx", ".xlsm", ".xls", ".csv", ".tsv", ".txt",
        ):
            continue
        part = read_seller_file(path)
        # Merge metrics.
        for metric_key, extractions in part.metrics.items():
            combined.metrics.setdefault(metric_key, []).extend(extractions)
        # Merge column-map notes.
        for col, key in part.raw_columns_mapped.items():
            combined.raw_columns_mapped[f"{path.name}:{col}"] = key
        for col in part.unmapped_columns:
            combined.unmapped_columns.append(f"{path.name}:{col}")
    _detect_conflicts(combined)
    return combined
