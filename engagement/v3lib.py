"""House-style sheet builder for IFT Sourced Evidence Master v3.

Renders new v3 tabs in the exact v2.7 visual language:
  - title row: Arial 15 bold navy #00294C
  - subtitle row: Arial 9 grey #555555, wrapped
  - column headers: Arial 9 bold white on navy fill, medium bottom border
  - section banners: Arial 10 bold white on teal #1F6F8B
  - data cells: Arial 9; BLUE #0000FF = hardcoded from a source document,
    BLACK = Excel formula, GREEN #1F7A33 = link to another tab
  - thin bottom borders on data rows; landscape / fit-to-width print setup
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = 'FF00294C'
TEAL = 'FF1F6F8B'
GREY = 'FF555555'
BLUE = 'FF0000FF'   # hardcoded from source
BLACK = 'FF000000'  # formula
GREEN = 'FF1F7A33'  # cross-tab link
PURPLE = 'FF7A5195'

FMT_INT = '#,##0;\\(#,##0\\);\\-'
FMT_DEC1 = '#,##0.0;\\(#,##0.0\\);\\-'
FMT_DEC2 = '#,##0.00;\\(#,##0.00\\);\\-'
FMT_USD = '$#,##0;\\($#,##0\\);\\-'
FMT_USD2 = '$#,##0.00;\\($#,##0.00\\);\\-'
FMT_PCT1 = '0.0%;\\(0.0%\\);\\-'
FMT_PCT2 = '0.00%;\\(0.00%\\);\\-'
FMT_X = '0.00"x"'

_thin = Border(bottom=Side(style='thin', color='FFD9D9D9'))
_medium = Border(bottom=Side(style='medium', color=NAVY))

F_TITLE = Font(name='Arial', size=15, bold=True, color=NAVY)
F_SUB = Font(name='Arial', size=9, color=GREY)
F_HDR = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
F_BANNER = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
F_SRC = Font(name='Arial', size=9, color=BLUE)
F_FML = Font(name='Arial', size=9, color=BLACK)
F_LINK = Font(name='Arial', size=9, color=GREEN)
F_TXT = Font(name='Arial', size=9, color=BLACK)
F_LABEL = Font(name='Arial', size=9, bold=True, color=BLACK)
F_NOTE = Font(name='Arial', size=8, italic=True, color=GREY)

FILL_HDR = PatternFill('solid', fgColor=NAVY)
FILL_BANNER = PatternFill('solid', fgColor=TEAL)

AL_HDR = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_WRAP = Alignment(vertical='top', wrap_text=True)
AL_TOP = Alignment(vertical='top')


class SheetBuilder:
    """Row-cursor builder that writes v2.7-house-style content."""

    def __init__(self, ws, n_cols, col_widths=None, tab_color=None):
        self.ws = ws
        self.n = n_cols
        self.r = 0
        ws.sheet_view.showGridLines = False
        if col_widths:
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = 5

    def _row(self):
        self.r += 1
        return self.r

    def title(self, text):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_TITLE
        c.alignment = AL_TOP
        self.ws.row_dimensions[r].height = 24
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def subtitle(self, text, height=30):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_SUB
        c.alignment = AL_WRAP
        self.ws.row_dimensions[r].height = height
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def blank(self):
        r = self._row()
        self.ws.row_dimensions[r].height = 13.5
        return self

    def banner(self, text):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_BANNER
        c.fill = FILL_BANNER
        c.alignment = AL_TOP
        self.ws.row_dimensions[r].height = 16
        for i in range(2, self.n + 1):
            self.ws.cell(row=r, column=i).fill = FILL_BANNER
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def headers(self, cols, freeze=True, height=None):
        r = self._row()
        if height is None:
            # size to the tallest wrapped header given its column width
            lines = 1
            for i, h in enumerate(cols, start=1):
                if not h:
                    continue
                dim = self.ws.column_dimensions.get(get_column_letter(i))
                w = (dim.width if dim is not None and dim.width else 8.43)
                lines = max(lines, -(-len(str(h)) // max(6, int(w * 0.95))))
            height = max(15, min(46, lines * 10.5 + 5))
        for i, h in enumerate(cols, start=1):
            c = self.ws.cell(row=r, column=i, value=h)
            c.font = F_HDR
            c.fill = FILL_HDR
            c.alignment = AL_HDR
            c.border = _medium
        for i in range(len(cols) + 1, self.n + 1):
            c = self.ws.cell(row=r, column=i)
            c.fill = FILL_HDR
            c.border = _medium
        self.ws.row_dimensions[r].height = height
        if freeze and self.ws.freeze_panes is None:
            self.ws.freeze_panes = f'A{r + 1}'
        return self

    def row(self, cells, height=None, wrap=False):
        """cells: list of (value, kind[, numfmt]) or plain values (kind='text').

        kinds: 'src' blue source-hardcoded | 'fml' black formula |
               'link' green cross-tab | 'text' | 'label' bold | 'note'
        """
        r = self._row()
        for i, spec in enumerate(cells, start=1):
            if spec is None:
                continue
            if not isinstance(spec, tuple):
                spec = (spec, 'text')
            value, kind = spec[0], spec[1]
            numfmt = spec[2] if len(spec) > 2 else None
            c = self.ws.cell(row=r, column=i, value=value)
            c.font = {'src': F_SRC, 'fml': F_FML, 'link': F_LINK,
                      'label': F_LABEL, 'note': F_NOTE}.get(kind, F_TXT)
            c.border = _thin
            c.alignment = AL_WRAP if wrap else AL_TOP
            if numfmt:
                c.number_format = numfmt
        if height:
            self.ws.row_dimensions[r].height = height
        elif wrap:
            self.ws.row_dimensions[r].height = 30
        return self

    def note(self, text, height=24):
        r = self._row()
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = F_NOTE
        c.alignment = AL_WRAP
        self.ws.row_dimensions[r].height = height
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=self.n)
        return self

    def prose(self, text, kind='text', lead=None):
        """A merged, wrapped body-text row with height sized to the text.

        lead: optional bold run rendered as its own first column instead of
        merging, for 'Term. Definition' rows.
        """
        r = self._row()
        width_chars = sum(
            (self.ws.column_dimensions[get_column_letter(i)].width or 8.43)
            for i in range(1, self.n + 1)) - 6
        font = {'src': F_SRC, 'fml': F_FML, 'link': F_LINK,
                'label': F_LABEL, 'note': F_NOTE}.get(kind, F_TXT)
        c = self.ws.cell(row=r, column=1, value=text)
        c.font = font
        c.alignment = AL_WRAP
        lines = max(1, -(-len(text) // max(60, int(width_chars * 0.92))))
        self.ws.row_dimensions[r].height = lines * 11.5 + 3.5
        if self.n > 1:
            self.ws.merge_cells(start_row=r, start_column=1, end_row=r,
                                end_column=self.n)
        return self


import re as _re

_REF = _re.compile(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def _mkref(wb, f):
    from openpyxl.chart import Reference
    from openpyxl.utils import column_index_from_string
    m = _REF.match(f)
    if not m:
        raise ValueError(f'bad ref {f!r}')
    sheet, c1, r1, c2, r2 = m.groups()
    return Reference(wb[sheet], min_col=column_index_from_string(c1), min_row=int(r1),
                     max_col=column_index_from_string(c2) if c2 else column_index_from_string(c1),
                     max_row=int(r2) if r2 else int(r1))


# ---------------------------------------------------------------- charts ---
# House chart style. Every chart in the workbook goes through add_chart()
# and/or normalize_all_charts(); both enforce the same rules so v3-built and
# v2.7-carried charts are indistinguishable.

CHART_COLORS = ['00294C', '1F6F8B', 'C58F00', '8C1D40', '4C7C2C', '6B7C93',
                '7A5195', 'B75D2A']
GRID_COLOR = 'E3E7EB'
TICK_COLOR = '595959'


def _rich_font(sz, color, bold=False):
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (CharacterProperties, Paragraph,
                                       ParagraphProperties, RichTextProperties)
    cp = CharacterProperties(sz=sz, b=bold, solidFill=color)
    return RichText(bodyPr=RichTextProperties(vert='horz'),
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def _mk_title(text, sz=1100):
    from openpyxl.chart.text import RichText, Text
    from openpyxl.chart.title import Title
    from openpyxl.drawing.text import (CharacterProperties, Paragraph,
                                       ParagraphProperties, RegularTextRun)
    cp = CharacterProperties(sz=sz, b=True, solidFill='00294C')
    p = Paragraph(pPr=ParagraphProperties(defRPr=cp),
                  r=[RegularTextRun(rPr=cp, t=text)])
    return Title(tx=Text(rich=RichText(p=[p])), overlay=False)


def _style_series(ch, kind):
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    for i, s in enumerate(ch.series):
        color = CHART_COLORS[i % len(CHART_COLORS)]
        if s.spPr is None:
            s.spPr = GraphicalProperties()
        if kind == 'line':
            s.graphicalProperties.line = LineProperties(solidFill=color, w=23000)
            s.marker = Marker(symbol='none')
            s.smooth = False
        else:
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line = LineProperties(noFill=True)


def _style_axes(ch, y_fmt=None, horizontal=False):
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    x, y = ch.x_axis, ch.y_axis
    x.delete = False
    y.delete = False
    if horizontal:      # horizontal bars: categories left, values bottom
        x.axPos = 'l'
        y.axPos = 'b'
    else:
        x.axPos = 'b'
        y.axPos = 'l'
    x.majorTickMark = 'out'
    x.minorTickMark = 'none'
    y.majorTickMark = 'out'
    y.minorTickMark = 'none'
    x.majorGridlines = None
    y.majorGridlines = ChartLines(spPr=GraphicalProperties(
        ln=LineProperties(solidFill=GRID_COLOR, w=9525)))
    x.spPr = GraphicalProperties(ln=LineProperties(solidFill=TICK_COLOR, w=9525))
    y.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
    x.txPr = _rich_font(800, TICK_COLOR)
    y.txPr = _rich_font(800, TICK_COLOR)
    if y_fmt:
        y.number_format = y_fmt
    elif not y.number_format or y.number_format == 'General':
        y.number_format = '#,##0'


def _style_legend(ch):
    from openpyxl.chart.legend import Legend
    n = len(ch.series)
    if n <= 1:
        ch.legend = None
    else:
        if ch.legend is None:
            ch.legend = Legend()
        ch.legend.position = 'b'
        ch.legend.overlay = False
        ch.legend.txPr = _rich_font(800, TICK_COLOR)


def style_chart(ch, kind, y_fmt=None):
    """Apply the full house style to a Bar/LineChart object in place."""
    if ch.title is not None:
        txt = None
        try:
            txt = ''.join(r.t or '' for p in ch.title.tx.rich.p for r in (p.r or []))
        except AttributeError:
            pass
        if txt:
            ch.title = _mk_title(txt)
    _style_series(ch, kind)
    _style_axes(ch, y_fmt=y_fmt,
                horizontal=(kind == 'bar' and getattr(ch, 'type', 'col') == 'bar'))
    _style_legend(ch)
    ch.varyColors = False
    if kind == 'bar':
        ch.gapWidth = 60
        if ch.grouping == 'stacked':
            ch.overlap = 100


def add_chart(ws, anchor, title, cat_ref, series, kind='line', width=15.2,
              height=6.8, y_title=None, x_title=None, y_fmt=None, stacked=False):
    """Add a native house-style chart whose series are live range references.

    series: list of (series_name:str, values_ref:'Sheet!$B$5:$B$20').
    One chart = one value axis. Anything needing a second axis is built as
    two separate charts at the call site.
    """
    from openpyxl.chart import BarChart, LineChart
    from openpyxl.chart.data_source import StrRef
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.series_factory import SeriesFactory as Series

    wb = ws.parent
    if kind == 'bar':
        main = BarChart()
        main.type = 'col'
        main.grouping = 'stacked' if stacked else 'clustered'
    else:
        main = LineChart()
    for name, ref in series:
        s = Series(_mkref(wb, ref), title_from_data=False)
        s.tx = SeriesLabel(strRef=StrRef(name)) if _REF.match(name) else SeriesLabel(v=name)
        main.series.append(s)
    if cat_ref:
        main.set_categories(_mkref(wb, cat_ref))
        main.x_axis.number_format = 'General'
    main.title = title
    style_chart(main, kind, y_fmt=y_fmt)
    if y_title:
        main.y_axis.title = y_title
    if x_title:
        main.x_axis.title = x_title
    main.width = width
    main.height = height
    ws.add_chart(main, anchor)
    return main


def _col_x_cm(ws, col_idx):
    """x offset (cm) of the left edge of 1-based column col_idx."""
    total = 0.0
    for i in range(1, col_idx):
        dim = ws.column_dimensions.get(get_column_letter(i))
        w = dim.width if dim is not None and dim.width else 8.43
        total += (int(round(w * 7)) + 5) * 2.54 / 96
    return total


def _row_y_cm(ws, row_idx):
    """y offset (cm) of the top edge of 1-based row row_idx."""
    total = 0.0
    for i in range(1, row_idx):
        dim = ws.row_dimensions.get(i)
        h = dim.height if dim is not None and dim.height else 15.0
        total += h * 2.54 / 72
    return total


def normalize_all_charts(wb, log=None):
    """Workbook-wide chart pass: house style + collision-free geometry.

    Restyles every chart (including v2.7-carried rebuilds) and shrinks
    width/height where a chart would otherwise overlap the next chart
    below or beside it.
    """
    from openpyxl.chart import BarChart, LineChart
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    n_styled = n_clipped = 0
    for name in wb.sheetnames:
        ws = wb[name]
        charts = list(ws._charts)
        if not charts:
            continue
        geo = []
        for ch in charts:
            kind = 'bar' if isinstance(ch, BarChart) else (
                'line' if isinstance(ch, LineChart) else None)
            if kind:
                fmt = ch.y_axis.number_format
                style_chart(ch, kind, y_fmt=fmt)
                n_styled += 1
            anchor = ch.anchor
            if not isinstance(anchor, str):
                continue
            col_s, row = coordinate_from_string(anchor)
            col = column_index_from_string(col_s)
            geo.append({'ch': ch, 'col': col, 'row': row,
                        'x': _col_x_cm(ws, col), 'y': _row_y_cm(ws, row)})
        # vertical collisions: same column band, chart below starts inside me
        for g in geo:
            below = [o for o in geo if o is not g and o['row'] > g['row']
                     and o['x'] < g['x'] + g['ch'].width - 0.3
                     and o['x'] + o['ch'].width > g['x'] + 0.3]
            if below:
                limit = min(o['y'] for o in below) - g['y'] - 0.25
                if g['ch'].height > limit > 3.0:
                    g['ch'].height = round(limit, 2)
                    n_clipped += 1
        # horizontal collisions: chart to the right in my row band
        for g in geo:
            right = [o for o in geo if o is not g and o['col'] > g['col']
                     and o['y'] < g['y'] + g['ch'].height - 0.3
                     and o['y'] + o['ch'].height > g['y'] + 0.3]
            if right:
                limit = min(o['x'] for o in right) - g['x'] - 0.25
                if g['ch'].width > limit > 5.0:
                    g['ch'].width = round(limit, 2)
                    n_clipped += 1
        # floor sizes so nothing renders as a stamp
        for g in geo:
            if g['ch'].width < 8.0:
                g['ch'].width = 8.0
            if g['ch'].height < 5.2:
                g['ch'].height = 5.2
    if log:
        log(f'chart normalize: {n_styled} styled, {n_clipped} resized for collisions')
    return n_styled, n_clipped


def format_sweep(wb, log=None, cap=42):
    """Workbook-wide format pass.

    1) Print setup: every sheet landscape / fit-to-width / legal, including
       carried v2.7 tabs that never got the normalization.
    2) Column truncation: widen a column when its longest unwrapped text cell
       is clipped against a non-empty right neighbour (e.g. state names cut
       by a FIPS column). Never narrows, capped so nothing balloons.
    """
    widened = printed = 0
    for name in wb.sheetnames:
        ws = wb[name]
        ps = ws.page_setup
        if ps.orientation != 'landscape' or not ps.fitToWidth:
            ps.orientation = 'landscape'
            ps.fitToWidth = 1
            ps.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ps.paperSize = 5
            printed += 1
        merged = [r.bounds for r in ws.merged_cells.ranges]
        cand = {}   # col -> up to 3 longest unmerged (len, row)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v or v.startswith('='):
                    continue
                al = c.alignment
                if al is not None and al.wrap_text:
                    continue
                if any(c1 <= c.column <= c2 and r1 <= c.row <= r2
                       for (c1, r1, c2, r2) in merged):
                    continue
                lst = cand.setdefault(c.column, [])
                lst.append((len(v), c.row))
                if len(lst) > 3:
                    lst.sort(reverse=True)
                    del lst[3:]
        for col, lst in cand.items():
            letter = get_column_letter(col)
            cur = ws.column_dimensions[letter].width or 8.43
            for L, r in sorted(lst, reverse=True):
                if L * 1.05 + 1 <= cur:
                    break
                if ws.cell(row=r, column=col + 1).value is None:
                    continue
                ws.column_dimensions[letter].width = min(cap, L * 1.02 + 1.5)
                widened += 1
                break
    if log:
        log(f'format sweep: {printed} sheets print-normalized, '
            f'{widened} columns widened for truncation')
    return printed, widened


def cagr_formula(end_ref, start_ref, years):
    """CAGR as a live Excel formula string."""
    return f'=({end_ref}/{start_ref})^(1/{years})-1'


def load_cache(cache_dir, key):
    """Load a pull artifact; accepts plain .json (scratchpad) or .json.gz (repo)."""
    import gzip
    import json as _json
    import os as _os
    p = _os.path.join(cache_dir, key + '.json')
    if _os.path.exists(p):
        return _json.load(open(p))
    with gzip.open(p + '.gz', 'rt') as f:
        return _json.load(f)


def estimate_print_pages(ws, rows_per_page=42.0):
    """Printed-page estimate under landscape fit-to-width-1 setup.

    Counts weighted rows (taller rows consume more of a page) so dense wrapped
    sheets estimate honestly. A page at 100% zoom fits ~42 rows of 13.5pt;
    fit-to-width shrink adds rows per page for wide sheets, so this is a
    conservative (low) page estimate for narrow sheets and roughly right for
    wide ones.
    """
    used = 0.0
    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        h = ws.row_dimensions[r].height if r in ws.row_dimensions and ws.row_dimensions[r].height else 13.5
        used += h / 13.5
    if used == 0:
        return 0
    return max(1, round(used / rows_per_page + 0.499))
