"""Firewall leak check: confirm ZERO Komodo-derived values in the master.

The Komodo extract is licensed engagement data. This scans the public master
workbook AND the master pipeline tree (RCM_MC/, excluding the engagement/
directory) for the extract's signature figures. Any hit is a firewall breach.
Writes leak_check_master.json. Exits non-zero on any hit.
"""
import json
import os
import re
import sys

from openpyxl import load_workbook

MASTER_XLSX = ('/home/user/RCM/RCM_MC/deliverables/'
               'IFT_Sourced_Evidence_Master_v4_2.xlsx')
REPO = '/home/user/RCM'
SCAN_DIRS = ['RCM_MC']            # master tree only; engagement/ is excluded
EXCLUDE = ('engagement',)

# Komodo-EXTRACT-SPECIFIC signatures (totals and Komodo claim/dollar counts).
# These do NOT include any public benchmark (e.g. 10,637,766 / 11,300,000 /
# 581,532), which legitimately live in the master.
SIG = [47061988, 14320080327, 9435317565, 12536776941, 7885341451,
       4019680332, 2415058544, 28660137, 8044685, 10357334,
       5409977, 1826100921, 632029, 259868460, 8919408, 5250804433,
       262139, 329191232, 24881351, 14723330, 37710269, 9351749]
SIG_STR = [str(s) for s in SIG] + ['{:,}'.format(s) for s in SIG]
# Digit-boundary regex per signature: a plain or comma-grouped signature not
# embedded in a longer digit run (avoids matching inside e.g. 0.0092621398).
SIG_RX = [re.compile(r'(?<!\d)' + re.escape(s) + r'(?!\d)') for s in SIG_STR]


def scan_workbook():
    wb = load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    hits = []
    sigset = set(SIG)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, (int, float)) and int(v) in sigset:
                    hits.append(f'{ws.title}!{c.coordinate}={v}')
                elif isinstance(v, str):
                    for rx in SIG_RX:
                        if rx.search(v):
                            hits.append(f'{ws.title}!{c.coordinate}~{rx.pattern}')
                            break
    return hits


def scan_repo():
    hits = []
    for d in SCAN_DIRS:
        for root, dirs, files in os.walk(os.path.join(REPO, d)):
            dirs[:] = [x for x in dirs if x not in EXCLUDE]
            for fn in files:
                if not fn.endswith(('.py', '.json', '.md', '.txt', '.csv')):
                    continue
                path = os.path.join(root, fn)
                try:
                    txt = open(path, encoding='utf-8', errors='ignore').read()
                except OSError:
                    continue
                for rx in SIG_RX:
                    if rx.search(txt):
                        hits.append(f'{os.path.relpath(path, REPO)}~{rx.pattern}')
                        break
    return hits


def main():
    wb_hits = scan_workbook()
    repo_hits = scan_repo()
    result = {
        'master_workbook': os.path.basename(MASTER_XLSX),
        'signatures_scanned': len(SIG),
        'workbook_hits': wb_hits,
        'repo_hits': repo_hits,
        'clean': not (wb_hits or repo_hits),
        'statement': ('Zero Komodo-derived values in the master or its RCM_MC '
                      'tree; the engagement extract stays engagement-side.'),
    }
    json.dump(result, open(os.path.join(os.path.dirname(os.path.abspath(
        __file__)), 'leak_check_master.json'), 'w'), indent=1)
    print(f'master leak check: {len(SIG)} signatures; '
          f'{len(wb_hits)} workbook hits, {len(repo_hits)} repo hits')
    for h in (wb_hits + repo_hits)[:20]:
        print('  HIT', h)
    if not result['clean']:
        sys.exit(1)
    print('MASTER LEAK CHECK CLEAN - master untouched')


if __name__ == '__main__':
    main()
