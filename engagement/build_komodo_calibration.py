"""Komodo_Calibration_v1.xlsx - engagement-side calibration workbook.

Standalone deliverable per the Komodo Calibration Order. LICENSED engagement
data (the Komodo extract) NEVER enters IFT_Sourced_Evidence_Master. This
workbook is a ONE-WAY consumer of public benchmarks from the master (imported as
values with their fact IDs / source tabs) and returns nothing to it.

Until Andrew delivers the raw extract, this builds the full skeleton: the public
benchmark pack, the observed inputs from the visual transcription (flagged
PROVISIONAL, to be re-derived by T1 from the raw file), and every test panel
(T1-T10) with live formulas and bordered PENDING cells where a raw-file value is
still needed. Landing the raw file is then mechanical: overwrite Observed_Inputs,
fill the PENDING cells, and every panel recomputes.

CIM formatting standard (the master's v3lib SheetBuilder) applies.
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
for cand in (HERE,
             '/home/user/RCM/RCM_MC/scripts/ift_evidence_v3'):
    if os.path.exists(os.path.join(cand, 'v3lib.py')):
        sys.path.insert(0, cand)
        break
import v3lib  # noqa: E402
from v3lib import (FMT_INT, FMT_USD, FMT_USD2, FMT_PCT1, FMT_DEC2,  # noqa: E402
                   FMT_X)

OUT = os.environ.get('KOMODO_CAL_OUT',
                     os.path.join(HERE, 'Komodo_Calibration_v1.xlsx'))
BUILT = '14 July 2026'
MASTER = 'IFT_Sourced_Evidence_Master v4.2'

# ------------------------------------------------------------------ refs ---
REF = {}          # semantic key -> 'Sheet!Cell'


def cellref(sheet, r, col):
    return f"{sheet}!{get_column_letter(col)}{r}"


# ============================================================ ReadMe ======
def build_readme(wb):
    ws = wb.create_sheet('ReadMe')
    sb = v3lib.SheetBuilder(ws, 3, col_widths=[34, 74, 44], tab_color='FF6B7C93')
    sb.title('Komodo Calibration v1 - the licensed extract read against the '
             'public evidence base')
    sb.subtitle('Purpose: calibrate the Komodo Health ambulance claims extract '
                '(calendar 2025) against the verified public benchmarks in the '
                'IFT master, and package what the extract is good for - the '
                'Medicare capture / gross-up factor, the commercial realized-'
                'price multiples, the estimated-allowed disclosure, the '
                'institutional-lines verdict, and the payer mix - for the '
                'sizing team. Every test panel is live: it computes now from the '
                'transcribed inputs, and recomputes the moment the raw file '
                'overwrites Observed_Inputs.', height=64)
    sb.blank()
    sb.banner('Firewall - read this before anything else')
    sb.prose(
        'The Komodo extract is LICENSED, commercial engagement data. It NEVER '
        'enters IFT_Sourced_Evidence_Master and is never cited by, linked from, '
        'or merged into it. The flow is one-way and permanent: master -> '
        'calibration. This workbook IMPORTS public benchmark values from the '
        'master (as values, each carrying its fact ID or source tab) and returns '
        'nothing. A leak check logged on Run_Log confirms zero Komodo-derived '
        'values exist anywhere in the master or its repo. Keep it that way.',
        kind='note')
    sb.blank()
    sb.banner('The three extract caveats - preserved on every output panel')
    sb.row([('1', 'label'), ('Members without a listed line of business or '
             'state were excluded from the extract (under 1 percent of claims).',
             'text'), None], wrap=True, height=28)
    sb.row([('2', 'label'), ('Codes are not mutually exclusive to claim status '
             '(Open / Closed / Combined); a code can appear across statuses.',
             'text'), None], wrap=True, height=28)
    sb.row([('3', 'label'), ('Multiple codes can appear on one claim, so grand '
             'totals may not align to the code breakdown. T10 defines the '
             'counting rule that removes this ambiguity downstream.', 'text'),
            None], wrap=True, height=34)
    sb.blank()
    sb.banner('How to read the tabs')
    for a, b in [
        ('Benchmarks_Public', 'The public pack imported from the master: volume '
         'floors and wedges, price benchmarks, and calibrators - each with its '
         'fact ID / source tab. Values only; nothing computed from Komodo.'),
        ('Observed_Inputs (T1)', 'The extract headline and per-code values. '
         'Currently the visual transcription, flagged PROVISIONAL and bordered; '
         'T1 re-derives every one from the raw file. All test panels reference '
         'these cells, so overwriting them here recomputes everything.'),
        ('T2 - T10', 'One panel per test. Live formulas, the interpretation and '
         'the caveat printed in-row. Bordered PENDING marks a value that needs '
         'the raw file (an obscured transcription cell, a payer/code split, a '
         'geographic filter, or a metadata field).'),
        ('GrossUp_Factors', 'The T2 Medicare capture rate packaged as the '
         'candidate gross-up for the sizing team, every assumption listed.'),
        ('Commercial_Rate_Table', 'The T3 commercial multiples packaged - the '
         'measured replacement for the interim FAIR bound and statutory pegs.'),
        ('Calibration_Memo', 'One page for Ray: what the extract is good for and '
         'the three things it cannot see. Also shipped as a standalone .md.'),
        ('Run_Log', 'Build metadata, the raw-file hash slot, and the master-'
         'untouched leak-check result.'),
    ]:
        sb.row([(a, 'label'), (b, 'text'), None], wrap=True, height=40)
    return ws


# ================================================== Benchmarks_Public =====
CODES = ['A0426', 'A0427', 'A0428', 'A0429', 'A0433', 'A0434']
RVU = {'A0428': 1.00, 'A0429': 1.60, 'A0426': 1.20, 'A0427': 1.90,
       'A0433': 2.75, 'A0434': 3.25}
CODE_DESC = {'A0426': 'ALS1 non-emergency', 'A0427': 'ALS1 emergency',
             'A0428': 'BLS non-emergency', 'A0429': 'BLS emergency',
             'A0433': 'ALS2', 'A0434': 'Specialty Care Transport'}
# National realized allowed per code, 2024 (MUP_Ambulance_National avg allowed).
REALIZED_2024 = {'A0426': 327.72, 'A0427': 522.51, 'A0428': 260.65,
                 'A0429': 446.94, 'A0433': 753.58, 'A0434': 918.36}


def _scalar(sb, sheet, label, value, unit, source, note, fmt=None, key=None):
    sb.row([(label, 'label'), (value, 'src', fmt),
            (unit, 'text'), (source, 'link'), (note, 'note')], wrap=True,
           height=30)
    if key:
        REF[key] = cellref(sheet, sb.r, 2)


def build_benchmarks(wb):
    ws = wb.create_sheet('Benchmarks_Public')
    S = 'Benchmarks_Public'
    sb = v3lib.SheetBuilder(ws, 5, col_widths=[46, 18, 16, 34, 50],
                            tab_color='FF00294C')
    sb.title('Public benchmark pack - imported from the master as values, with '
             'fact IDs')
    sb.subtitle('Every value here is a verified public figure carried from '
                + MASTER + '. Nothing on this tab is computed from Komodo. The '
                'test panels reference these cells; they are the fixed side of '
                'every calibration.', height=40)
    sb.blank()
    sb.banner('Volume floors and wedges')
    sb.headers(['Benchmark', 'Value', 'Unit', 'Master source (fact ID / tab)',
                'Note'])
    _scalar(sb, S, 'Medicare FFS carrier ground base-code transports, 2023',
            10718468, 'transports', 'Medicare_OD_Matrix; Modifier_QM_QN_Series',
            'The QM/QN carrier denominator (base codes, O-D coded).', FMT_INT,
            'ffs2023')
    _scalar(sb, S, 'Medicare FFS carrier ground base-code transports, 2024',
            10637766, 'transports', 'F111 (Medicare_OD_Matrix)',
            'The carrier-visible FFS floor used in the T2 reconstruction.',
            FMT_INT, 'ffs2024')
    _scalar(sb, S, 'MedPAC all-AFS Medicare transports, 2023', 11300000,
            'transports', 'F88 (MedPAC_2026_Mandated)',
            'Use the master 11.3M, not the 11.4M preliminary. Payments $5.3B.',
            FMT_INT, 'medpac')
    _scalar(sb, S, 'Institutional (hospital-outpatient) wedge, 2023', 581532,
            'transports', 'F620 (Institutional_Ambulance_Wedge)',
            '5.15 percent of the Medicare AFS book; off-carrier FFS.', FMT_INT,
            'wedge')
    _scalar(sb, S, 'Dark share (beneficiaries outside carrier-visible FFS), 2024',
            0.593, 'share', 'F558 (Imbalance_Ledger)',
            'MA and other. FFS carrier-visible share = 1 - this = 0.407.',
            FMT_PCT1, 'dark')
    _scalar(sb, S, 'NEMSIS hospital-to-hospital activations, 2024', 5510664,
            'activations', 'EMS_Transports; Service_Level_Economics',
            'Activations, not Medicare claims; a demand-side ceiling.', FMT_INT,
            'nemsis_h2h')
    _scalar(sb, S, 'NEMSIS all facility-to-facility activations, 2024', 7512656,
            'activations', 'Service_Level_Economics; Cross_Source_Recon',
            'The widest facility-to-facility activation count.', FMT_INT,
            'nemsis_f2f')
    _scalar(sb, S, 'Hospital-to-hospital Medicare FFS carrier transports, 2024',
            871753, 'transports', 'Medicare_PSPS',
            'The carrier-visible H2H slice; a floor for the IFT subset.',
            FMT_INT, 'h2h_ffs')
    sb.blank()

    sb.banner('Price benchmarks')
    sb.headers(['Benchmark', 'Value', 'Unit', 'Master source (fact ID / tab)',
                'Note'])
    _scalar(sb, S, 'Ground ambulance conversion factor, CY2024', 272.44, '$',
            'Payment_Rules B32 (GOV-A)',
            'Multiplies the RVU to produce the base allowed, pre-geography.',
            FMT_USD2, 'cf2024')
    _scalar(sb, S, 'Ground ambulance conversion factor, CY2026', 284.56, '$',
            'Payment_Rules (GOV-A)', 'Published CY2026 CF.', FMT_USD2, 'cf2026')
    _scalar(sb, S, 'Ambulance Inflation Factor, CY2025', 0.024, 'per yr',
            'Payment_Rules (GOV-A)',
            'CY2025 CF = CY2024 x (1 + this); see the per-code table.', FMT_PCT1,
            'aif2025')
    _scalar(sb, S, 'Ground mileage rate (A0425), CY2024', 8.76, '$/mile',
            'Payment_Rules B33 (GOV-A)', 'Paid separately from the base rate.',
            FMT_USD2, 'mileage2024')
    # CY2025 CF computed live from CY2024 x (1+AIF2025)
    sb.row([('Ground ambulance conversion factor, CY2025 (AIF-stepped)', 'label'),
            (f"={REF['cf2024']}*(1+{REF['aif2025']})", 'fml', FMT_USD2),
            ('$', 'text'), ('DERIVED from Payment_Rules', 'link'),
            ('CY2024 CF stepped by the CY2025 AIF.', 'note')], wrap=True,
           height=26)
    REF['cf2025'] = cellref(S, sb.r, 2)
    sb.blank()
    sb.banner('Per-code fee-schedule and realized-allowed benchmarks (2024)')
    sb.headers(['HCPCS (level of service)', 'RVU', 'Fee-schedule allowed '
                '(CF2024 x RVU)', 'Realized national allowed', 'Source / note'])
    for code in CODES:
        sb.row([(f'{code}  {CODE_DESC[code]}', 'label'),
                (RVU[code], 'src', FMT_DEC2),
                (f"={REF['cf2024']}*B{sb.r+1}", 'fml', FMT_USD2),
                (REALIZED_2024[code], 'src', FMT_USD2),
                ('CMS AFS RVU (Payment_Rules); realized = MUP_Ambulance_National '
                 'avg allowed 2024', 'note')], wrap=True, height=24)
        r = sb.r
        REF[f'rvu_{code}'] = cellref(S, r, 2)
        REF[f'fs_{code}'] = cellref(S, r, 3)
        REF[f'real_{code}'] = cellref(S, r, 4)
    sb.note('RVU note: A0426 (ALS1 non-emergency) = 1.20 per the master '
            'Payment_Rules (GOV-A). The calibration order lists 1.75 for A0426; '
            'that is the A0432 paramedic-intercept RVU and is not used here. The '
            'master governs. A0433 = 2.75 and A0434 = 3.25 are the CMS AFS ALS2 '
            'and SCT values.')
    sb.blank()
    sb.banner('Calibrators and quality references')
    sb.headers(['Benchmark', 'Value', 'Unit', 'Master source (fact ID / tab)',
                'Note'])
    _scalar(sb, S, 'Measured BLS-NE avg Medicare allowed, 2024, largest subject '
            'NPI', 265.34, '$/claim', 'MUP_Providers_2024 (live CMS API '
            'verified)', 'A sanity point against A0428 realized allowed '
            '($260.65 national).', FMT_USD2, 'blsne_npi')
    _scalar(sb, S, 'MA-to-FFS ambulance revenue ratio per organization', 1.038,
            'x', 'F557 (MA_Book_Calibrator; GADCS 2.30)',
            'MA rev/org $1,238,482 vs FFS $1,193,046. A pricing calibrator, not '
            'a volume ratio.', FMT_X, 'ma_cal')
    _scalar(sb, S, 'GADCS cost per transport, mean', 2673, '$/transport',
            'GADCS (Cost_and_Capacity)', 'Median $1,340. Cost, not price.',
            FMT_USD, 'gadcs_mean')
    _scalar(sb, S, 'GADCS cost per transport, median', 1340, '$/transport',
            'GADCS (Cost_and_Capacity)', 'Skewed mean vs median.', FMT_USD,
            'gadcs_med')
    _scalar(sb, S, 'BLS-NE Medicare denial rate, 2024', 0.129, 'share',
            'PSPS_Denial_Series; Throughput_Economics_Public',
            'Context for closed-claim behavior.', FMT_PCT1, 'denial')
    _scalar(sb, S, 'Ambulance organizations reporting facility-pay revenue',
            0.186, 'share', 'F443 (Facility_Pay_Layer)',
            'Volume that NO claims dataset contains; a reminder, not an input.',
            FMT_PCT1, 'facpay')
    sb.note('Also carried as qualitative reminders (no single value): the '
            'one-in-five did-not-always-bill incidence, the FAIR-derived ~1.6x '
            'commercial bound and the statutory balance-billing pegs being '
            'SUPERSEDED by T3, and the footprint-state Medicaid fee-schedule '
            'base rates (NE A0428 $154.89 / A0426 $387.24 / A0434 $387.24; IA '
            '$84.67 / $101.60 / $232.84) on the master Medicaid_Rate_Card. No '
            'national Medicaid benchmark exists.')
    return ws


# =================================================== Observed_Inputs ======
# Transcribed headline values (PROVISIONAL - T1 re-derives from the raw file).
GRAND = {'claims': 47061988, 'reported': 9435317565,
         'estimated': 12536776941, 'combined': 14320080327}
PAYER = {  # payer -> (claims, combined$)
    'Commercial': (8044685, 4019680332),
    'Medicare': (28660137, 7885341451),
    'Medicaid': (10357334, 2415058544)}
STATUS = {'Open': (37710269, 11410000000), 'Closed': (9351749, 2910000000)}
# per code: total_claims, total_combined, then Commercial/Medicare/Medicaid
# (claims, combined). None = obscured in the shared view -> bordered PENDING.
PERCODE = {
    'A0425': (24881351, None, (None, None), (14723330, None), (None, None)),
    'A0426': (632029, 259868460, (153934, 96400641), (388782, 132434413),
              (None, None)),
    'A0427': (8919408, 5250804433, (None, None), (5376326, 2841911845),
              (None, None)),
    'A0428': (5409977, 1826100921, (539246, 287519936), (4178527, 1327883745),
              (692218, 210697240)),
    'A0429': (None, None, (None, None), (None, None), (None, None)),
    'A0433': (None, None, (None, None), (None, None), (None, None)),
    'A0434': (262139, 329191232, (69191, 152729421), (140010, 137165403),
              (52938, 39296408)),
}
PC_ORDER = ['A0425', 'A0426', 'A0427', 'A0428', 'A0429', 'A0433', 'A0434']


def _pend(v, fmt):
    return (v, 'src', fmt) if v is not None else ('PENDING', 'note')


def build_observed(wb):
    ws = wb.create_sheet('Observed_Inputs')
    S = 'Observed_Inputs'
    sb = v3lib.SheetBuilder(ws, 10, col_widths=[26, 15, 17, 13, 16, 13, 16, 13,
                                                16, 4], tab_color='FF7A3B00')
    sb.title('Observed inputs (T1 parser fidelity) - the extract as transcribed, '
             'PROVISIONAL until the raw file confirms it')
    sb.subtitle('Every value below is a visual transcription from the shared '
                'view, flagged PROVISIONAL and bordered. T1 re-derives each one '
                'from the raw file: on landing, overwrite these cells and hash '
                'the file on Run_Log. Every test panel references these cells, '
                'so the whole workbook recomputes from here. Cells marked '
                'PENDING were obscured in the view and must come from the raw '
                'file.', height=52)
    sb.blank()
    sb.banner('Grand totals')
    sb.headers(['Metric', 'Claims', 'Reported $', 'Estimated $', 'Combined $',
                '', '', '', '', ''])
    sb.row([('All codes, all payers', 'label'),
            (GRAND['claims'], 'src', FMT_INT),
            (GRAND['reported'], 'src', FMT_USD),
            (GRAND['estimated'], 'src', FMT_USD),
            (GRAND['combined'], 'src', FMT_USD),
            None, None, None, None, None], height=22)
    r = sb.r
    REF['g_claims'] = cellref(S, r, 2)
    REF['g_reported'] = cellref(S, r, 3)
    REF['g_estimated'] = cellref(S, r, 4)
    REF['g_combined'] = cellref(S, r, 5)
    sb.blank()

    sb.banner('Payer totals (all codes)')
    sb.headers(['Payer', 'Claims', 'Combined $', '', '', '', '', '', '', ''])
    for p, (cl, cb) in PAYER.items():
        sb.row([(p, 'label'), (cl, 'src', FMT_INT), (cb, 'src', FMT_USD),
                None, None, None, None, None, None, None], height=20)
        r = sb.r
        REF[f'p_{p}_claims'] = cellref(S, r, 2)
        REF[f'p_{p}_comb'] = cellref(S, r, 3)
    sb.blank()

    sb.banner('Claim maturity (all codes, all payers)')
    sb.headers(['Status', 'Claims', 'Combined $', '', '', '', '', '', '', ''])
    for st, (cl, cb) in STATUS.items():
        note = '' if st == 'Open' else ''
        sb.row([(st, 'label'), (cl, 'src', FMT_INT), (cb, 'src', FMT_USD),
                None, None, None, None, None, None, None], height=20)
        r = sb.r
        REF[f's_{st}_claims'] = cellref(S, r, 2)
        REF[f's_{st}_comb'] = cellref(S, r, 3)
    sb.note('Open/Closed combined dollars ($11.41B / $2.91B) were rounded in the '
            'shared view; T1 takes exact figures from the raw file.')
    sb.blank()

    sb.banner('Per-code grid (Combined $ unless noted; PENDING = obscured in the '
              'view, from raw file)')
    sb.headers(['HCPCS', 'Total claims', 'Total combined $', 'Comm claims',
                'Comm combined $', 'Mcare claims', 'Mcare combined $',
                'Mcaid claims', 'Mcaid combined $', ''])
    for code in PC_ORDER:
        tc, tcomb, (cc, ccb), (mc, mcb), (dc, dcb) = PERCODE[code]
        sb.row([(code, 'label'), _pend(tc, FMT_INT), _pend(tcomb, FMT_USD),
                _pend(cc, FMT_INT), _pend(ccb, FMT_USD),
                _pend(mc, FMT_INT), _pend(mcb, FMT_USD),
                _pend(dc, FMT_INT), _pend(dcb, FMT_USD), None], height=20)
        r = sb.r
        for j, tag in [(2, 'tot_claims'), (3, 'tot_comb'), (4, 'comm_claims'),
                       (5, 'comm_comb'), (6, 'mcare_claims'),
                       (7, 'mcare_comb'), (8, 'mcaid_claims'),
                       (9, 'mcaid_comb')]:
            REF[f'{code}_{tag}'] = cellref(S, r, j)
    sb.blank()
    sb.banner('T1 re-derivation (raw file)')
    sb.row([('Raw-file re-derivation status', 'label'),
            ('PENDING - Andrew delivers the raw extract; T1 recomputes every '
             'cell above and explains any discrepancy before downstream panels '
             'are trusted.', 'note'), None, None, None, None, None, None, None,
            None], wrap=True, height=40)
    sb.row([('Raw-file SHA-256', 'label'), ('PENDING', 'note'), None, None,
            None, None, None, None, None, None], height=20)
    return ws


# ============================================================= tests ======
def _panel(wb, name, width, widths, tab, title, subtitle):
    ws = wb.create_sheet(name)
    sb = v3lib.SheetBuilder(ws, width, col_widths=widths, tab_color=tab)
    sb.title(title)
    sb.subtitle(subtitle, height=48)
    sb.blank()
    return ws, sb


def build_t2(wb):
    S = 'T2_Medicare_Capture'
    ws, sb = _panel(wb, S, 4, [50, 20, 14, 60], 'FF13343B',
                    'T2. Medicare capture decomposition - the gross-up factor '
                    'everyone needs',
                    'Komodo Medicare base-code claims against a reconstruction '
                    'of the Medicare book (FFS carrier floor + institutional '
                    'wedge + an MA bound). Output: the implied Komodo capture of '
                    'the Medicare book, and how much of Komodo Medicare must be '
                    'MA. This capture rate is the candidate gross-up for the '
                    'commercial and Medicaid columns.')
    sb.headers(['Step', 'Value', 'Unit', 'Interpretation / caveat (in-row)'])

    def line(label, formula, unit, note, fmt, key=None):
        sb.row([(label, 'label'), (formula, 'fml', fmt), (unit, 'text'),
                (note, 'note')], wrap=True, height=30)
        if key:
            REF[key] = cellref(S, sb.r, 2)

    line('Komodo Medicare base-code claims, 2025 (Medicare total - A0425)',
         f"={REF['p_Medicare_claims']}-{REF['A0425_mcare_claims']}", 'claims',
         'Base codes only (mileage removed), per the T10 denominator rule.',
         FMT_INT, 'kom_mcare_base')
    line('FFS carrier base transports, 2024 (benchmark floor)',
         f"={REF['ffs2024']}", 'transports',
         'Carrier-visible FFS; the floor of the reconstruction.', FMT_INT)
    line('Plus institutional wedge (benchmark)', f"={REF['wedge']}",
         'transports', 'Off-carrier FFS, hospital-outpatient billed.', FMT_INT)
    line('Equals FFS AFS book (carrier plus institutional)',
         f"={REF['ffs2024']}+{REF['wedge']}", 'transports',
         'Reconciles to the MedPAC 11.3M all-AFS book as a sanity check.',
         FMT_INT, 'ffs_afs')
    line('Implied MA volume inside Komodo Medicare (assumes full FFS capture)',
         f"={REF['kom_mcare_base']}-({REF['ffs2024']}+{REF['wedge']})",
         'claims', 'Komodo Medicare base minus the FFS AFS book; the residual '
         'must be Medicare Advantage.', FMT_INT, 'kom_ma')
    line('MA volume upper bound (dark-share gross-up, per-capita equal)',
         f"=({REF['ffs2024']}+{REF['wedge']})*({REF['dark']}/(1-{REF['dark']}))",
         'transports', 'FFS AFS book scaled by MA/FFS beneficiary ratio; an '
         'UPPER bound (MA per-capita use is typically lower than FFS).', FMT_INT,
         'ma_ub')
    line('Implied Komodo MA capture (of the MA upper bound)',
         f"={REF['kom_ma']}/(({REF['ffs2024']}+{REF['wedge']})*"
         f"({REF['dark']}/(1-{REF['dark']})))", 'share',
         'A LOWER bound on MA capture, because the denominator is an upper '
         'bound.', FMT_PCT1)
    line('Reconstructed full Medicare book (FFS AFS + MA upper bound)',
         f"=({REF['ffs2024']}+{REF['wedge']})*(1+({REF['dark']}/(1-{REF['dark']})))",
         'transports', 'The widest Medicare universe for the capture ratio.',
         FMT_INT, 'recon_full')
    line('Implied Komodo capture of the full Medicare book',
         f"={REF['kom_mcare_base']}/(({REF['ffs2024']}+{REF['wedge']})*"
         f"(1+({REF['dark']}/(1-{REF['dark']}))))", 'share',
         'The candidate gross-up = 1 / this capture. See GrossUp_Factors.',
         FMT_PCT1, 'mcare_capture')
    sb.blank()
    sb.banner('Caveats (carried to every downstream exhibit)')
    sb.prose('Vintage: Komodo is calendar 2025; the benchmarks are 2023-2024. '
             'Weakest link: the gross-up assumes capture is PAYER-UNIFORM, so '
             'the Medicare capture rate is applied to Commercial and Medicaid - '
             'stated, not proven. The MA bound assumes MA per-capita utilization '
             'equals FFS, which overstates MA volume, so the MA capture percent '
             'is a floor and the full-book capture percent is a floor. '
             '"Medicare" in Komodo mixes FFS and MA and (pending T7) possibly '
             'institutional lines; this decomposition is the tool for splitting '
             'them.', kind='note')
    return ws


def build_t3(wb):
    S = 'T3_Commercial_Rate'
    ws, sb = _panel(wb, S, 8, [16, 18, 18, 16, 16, 16, 16, 40], 'FF3B1330',
                    'T3. Commercial realized-price table - supersedes the '
                    'interim anchors',
                    'Per code: Commercial combined allowed per claim, and the '
                    'Medicare-multiple against both benchmarks (fee schedule '
                    'CF x RVU, and realized PSPS allowed), side by side with the '
                    'FAIR ~1.6x bound and the statutory pegs this measurement '
                    'replaces. Reported-only is the conservative column '
                    '(PENDING per-code reported dollars from the raw file).')
    sb.headers(['HCPCS', 'Comm combined $/claim', 'Multiple vs fee schedule',
                'Multiple vs realized', 'Reported-only $/claim (cons.)',
                'FAIR bound', 'Statutory pegs', 'Read / caveat'])
    for code in CODES:
        cc = REF[f'{code}_comm_claims']
        ccb = REF[f'{code}_comm_comb']
        has = PERCODE[code][2][0] is not None
        if has:
            perclaim = (f"={ccb}/{cc}", 'fml', FMT_USD2)
            mult_fs = (f"={ccb}/{cc}/{REF[f'fs_{code}']}", 'fml', FMT_X)
            mult_real = (f"={ccb}/{cc}/{REF[f'real_{code}']}", 'fml', FMT_X)
        else:
            perclaim = ('PENDING', 'note')
            mult_fs = ('PENDING', 'note')
            mult_real = ('PENDING', 'note')
        sb.row([(f'{code} {CODE_DESC[code]}', 'label'), perclaim, mult_fs,
                mult_real, ('PENDING', 'note'),
                ('~1.6x', 'text'), ('varies by state', 'text'),
                ('Combined = reported + Komodo-estimated; see T4 for the '
                 'modeled share. Reported-only column fills from the raw file.'
                 if has else 'Commercial split obscured in the view; from raw '
                 'file.', 'note')], wrap=True, height=32)
    sb.blank()
    sb.prose('Transcription check: A0428 computes to about $533 combined per '
             'claim (~2.0x fee schedule, ~2.0x realized); A0434 to about $2,208 '
             '(~2.5x fee schedule, ~2.4x realized). These recompute from the '
             'raw file. The measured multiples replace the FAIR-derived ~1.6x '
             'bound and the statutory balance-billing pegs carried on the '
             'master Balance_Billing_States; both are shown for continuity.',
             kind='note')
    return ws


def build_t4(wb):
    S = 'T4_Estimated_Allowed'
    ws, sb = _panel(wb, S, 4, [46, 20, 14, 60], 'FF13343B',
                    'T4. Estimated-allowed disclosure',
                    '(Combined minus Reported) / Combined - the share of dollars '
                    'that are Komodo-modeled rather than payer-reported. Every '
                    'exhibit built on Combined carries this share as a printed '
                    'caveat; T3 keeps a Reported-only column for exactly this '
                    'reason.')
    sb.headers(['Scope', 'Estimated share of $', 'Basis', 'Interpretation'])
    sb.row([('Grand total (all payers, all codes)', 'label'),
            (f"=({REF['g_combined']}-{REF['g_reported']})/{REF['g_combined']}",
             'fml', FMT_PCT1), ('Combined vs Reported', 'text'),
            ('About a third of all dollars are Komodo-estimated where the payer '
             'did not report an allowed amount.', 'note')], wrap=True, height=30)
    for p in PAYER:
        sb.row([(f'{p} (all codes)', 'label'), ('PENDING', 'note'),
                ('needs per-payer Reported $', 'text'),
                ('Payer-level Reported dollars are not in the transcription; '
                 'from the raw file.', 'note')], wrap=True, height=26)
    sb.note('Rule: any downstream number built on Combined dollars prints this '
            'estimated share beside it. The Reported-only figures are the '
            'conservative floor.')
    return ws


def build_t5(wb):
    S = 'T5_Claim_Maturity'
    ws, sb = _panel(wb, S, 4, [46, 20, 16, 58], 'FF13343B',
                    'T5. Claim maturity (Open vs Closed)',
                    'Open versus Closed shares, the unit-price comparison closed '
                    'vs combined, and a stated completion-factor method for how '
                    '2025 volumes will drift as claims close. No projected '
                    'number - the method and the current bias direction only.')
    sb.headers(['Metric', 'Value', 'Unit', 'Interpretation / caveat'])
    tot = REF['g_claims']
    sb.row([('Open share of claims', 'label'),
            (f"={REF['s_Open_claims']}/{tot}", 'fml', FMT_PCT1), ('share',
             'text'), ('2025 is early: about four in five claims are still '
             'open.', 'note')], wrap=True, height=24)
    sb.row([('Closed share of claims', 'label'),
            (f"={REF['s_Closed_claims']}/{tot}", 'fml', FMT_PCT1), ('share',
             'text'), ('The mature slice usable for realized pricing.', 'note')],
           wrap=True, height=24)
    sb.row([('Closed combined $/claim', 'label'),
            (f"={REF['s_Closed_comb']}/{REF['s_Closed_claims']}", 'fml',
             FMT_USD2), ('$/claim', 'text'),
            ('Compare to the all-status figure below.', 'note')], wrap=True,
           height=24)
    sb.row([('All-status combined $/claim', 'label'),
            (f"={REF['g_combined']}/{tot}", 'fml', FMT_USD2), ('$/claim',
             'text'), ('Closed-vs-all gap is the maturity price bias.', 'note')],
           wrap=True, height=24)
    sb.blank()
    sb.banner('Completion-factor method (no projected number)')
    sb.prose('2025 volumes are understated because ~80 percent of claims are '
             'still open. Method: once a mature prior year is available, compute '
             'a completion factor = final claims / claims-at-equivalent-age, and '
             'divide 2025 volumes by it to estimate the run-out. Current bias: '
             '2025 counts and dollars will RISE as claims close. State the '
             'factor and the direction; do not print a projected total until the '
             'mature-year ratio is in hand.', kind='note')
    return ws


def build_t6(wb):
    S = 'T6_Payer_Mix'
    ws, sb = _panel(wb, S, 8, [16, 14, 14, 14, 14, 14, 14, 40], 'FF13343B',
                    'T6. Payer mix of the scheduled book',
                    'Per IFT code (A0426 / A0428 / A0434), payer shares of '
                    'claims and dollars. Feeds the TAM pricing weights; compared '
                    'against the master enrollment-based expectations (the dark '
                    'share).')
    sb.headers(['HCPCS', 'Comm % claims', 'Mcare % claims', 'Mcaid % claims',
                'Comm % $', 'Mcare % $', 'Mcaid % $', 'Read / caveat'])
    for code in ['A0426', 'A0428', 'A0434']:
        tc = REF[f'{code}_tot_claims']
        tcb = REF[f'{code}_tot_comb']

        def sh(numref, den, ok):
            return (f"={numref}/{den}", 'fml', FMT_PCT1) if ok else \
                ('PENDING', 'note')
        pc = PERCODE[code]
        okc = pc[2][0] is not None
        okm = pc[3][0] is not None
        okd = pc[4][0] is not None
        sb.row([(f'{code} {CODE_DESC[code]}', 'label'),
                sh(REF[f'{code}_comm_claims'], tc, okc),
                sh(REF[f'{code}_mcare_claims'], tc, okm),
                sh(REF[f'{code}_mcaid_claims'], tc, okd),
                sh(REF[f'{code}_comm_comb'], tcb, okc),
                sh(REF[f'{code}_mcare_comb'], tcb, okm),
                sh(REF[f'{code}_mcaid_comb'], tcb, okd),
                ('A0428 is ~77 percent Medicare by claims - the scheduled BLS '
                 'book skews Medicare, as the dark share predicts. PENDING = '
                 'split obscured in the view.', 'note')], wrap=True, height=30)
    sb.note('Compare Medicare claim shares here against the 59.3 percent dark '
            'share: Komodo Medicare includes MA, so a high Medicare share is '
            'consistent with MA being captured (see T2).')
    return ws


def build_t7(wb):
    S = 'T7_Institutional_Lines'
    ws, sb = _panel(wb, S, 3, [30, 26, 64], 'FF3B1330',
                    'T7. Institutional-lines presence test (answered '
                    'empirically)',
                    'Does the extract contain institutional (UB-04) claims? '
                    'First from metadata (claim form / bill type / revenue-'
                    'center fields), and if absent, by inference from whether '
                    'the T2 Medicare reconstruction closes without room for the '
                    '581,532 institutional transports. One-cell verdict.')
    sb.banner('Verdict')
    sb.row([('Institutional lines in extract', 'label'),
            ('INDETERMINATE', 'note'),
            ('Metadata fields (claim form type / bill type / revenue center) '
             'are not in the transcription; the raw file decides. By inference, '
             'Komodo Medicare base already exceeds the FFS carrier floor, but '
             'the excess cannot be split between MA and institutional from '
             'volume alone - so the volume test is INDETERMINATE and metadata '
             'governs.', 'note')], wrap=True, height=56)
    sb.blank()
    sb.banner('Evidence')
    sb.row([('Metadata check', 'label'), ('PENDING (raw file)', 'note'),
            ('Look for claim_form_type / bill_type_code / revenue_center in the '
             'raw schema. Presence of rev centers 0540-0549 = INCLUDED.',
             'note')], wrap=True, height=34)
    sb.row([('Inference from T2', 'label'), ('INDETERMINATE', 'note'),
            ('T2 FFS AFS book (carrier + wedge) already embeds the 581,532; '
             'Komodo Medicare base exceeds it, but the surplus is MA and/or '
             'institutional - not separable without metadata.', 'note')],
           wrap=True, height=34)
    sb.blank()
    sb.banner('If EXCLUDED - the correction and the second-extract spec (ready '
              'to forward)')
    sb.prose('If institutional lines are EXCLUDED, the bottoms-up must ADD the '
             'institutional wedge (581,532 Medicare transports) separately, and '
             'the commercial/Medicaid institutional volume is unmeasured. Second '
             'extract spec: all claim lines with revenue centers 0540-0549 and '
             'HCPCS A0425-A0436, returning billing entity, service units, '
             'service date, and payer type, by year 2019-2024, so the '
             'institutional slice is measured directly rather than inferred.',
             kind='note')
    return ws


def build_t8(wb):
    S = 'T8_Mileage_Structure'
    ws, sb = _panel(wb, S, 4, [44, 20, 16, 60], 'FF13343B',
                    'T8. Mileage structure',
                    'A0425 claims per base-code claim, by payer, against the '
                    'master loading factor. Printed limitation: Komodo shows '
                    'CLAIMS, not mileage units, so this is a structural check, '
                    'not a loading measurement.')
    sb.headers(['Metric', 'Value', 'Unit', 'Interpretation / caveat'])
    sb.row([('A0425 Medicare claims per Medicare base-code claim', 'label'),
            (f"={REF['A0425_mcare_claims']}/{REF['kom_mcare_base']}", 'fml',
             FMT_DEC2), ('ratio', 'text'),
            ('About 1.06 mileage claims per base transport for Medicare - a '
             'structural ratio, not miles.', 'note')], wrap=True, height=30)
    sb.row([('Master loading factor (A0425 units per base transport)', 'label'),
            ('import Derived_Rate_Card', 'link'), ('x', 'text'),
            ('The master computes a mileage-LOADED price ratio; not directly '
             'comparable to a claims-per-claim count.', 'note')], wrap=True,
           height=28)
    sb.row([('Commercial / Medicaid mileage ratios', 'label'),
            ('PENDING', 'note'), ('ratio', 'text'),
            ('A0425 payer splits were obscured in the view; from the raw file.',
             'note')], wrap=True, height=26)
    sb.note('Limitation: a claims-per-claim ratio is not a loaded-mileage '
            'measurement. Reading loaded price per transport needs mileage '
            'UNITS, which Komodo does not carry.')
    return ws


def build_t9(wb):
    S = 'T9_Geography'
    ws, sb = _panel(wb, S, 4, [40, 22, 22, 50], 'FF3B1330',
                    'T9. Geography spot-checks',
                    'Re-run T2 / T3 / T6 filtered to Nebraska and one footprint '
                    'MSA; compare to the master state benchmarks. Flags any '
                    'geographic capture skew before a state cut is built on this '
                    'data.')
    sb.headers(['Check', 'Nebraska', 'Footprint MSA', 'Master comparison'])
    for chk in ['T2 Medicare capture (state)', 'T3 commercial multiple (state)',
                'T6 payer mix (state)']:
        sb.row([(chk, 'label'), ('PENDING', 'note'), ('PENDING', 'note'),
                ('Compare vs Medicaid_Rate_Card (NE base rates) and the master '
                 'state enrollment mix.', 'note')], wrap=True, height=28)
    sb.note('Needs the state/MSA-filtered extract. On landing, these fill from '
            'the same formulas as T2/T3/T6 with the geography filter applied; a '
            'large NE-vs-national capture gap means the national gross-up cannot '
            'be applied to a state cut without adjustment.')
    return ws


def build_t10(wb):
    S = 'T10_Denominator'
    ws, sb = _panel(wb, S, 2, [34, 86], 'FF00294C',
                    'T10. Denominator discipline (import this everywhere)',
                    'One definition every downstream exhibit uses, so no two '
                    'cuts count differently.')
    sb.banner('The counting rule')
    for a, b in [
        ('A transport =', 'one base-code claim (A0426, A0427, A0428, A0429, '
         'A0433, A0434). Mileage (A0425) is EXCLUDED from transport counts - it '
         'is a per-mile add-on, not a transport.'),
        ('Dollars =', 'Combined allowed (reported + estimated) unless a panel '
         'states Reported-only; the estimated share (T4) travels with any '
         'Combined figure.'),
        ('Status =', 'Combined (Open + Closed) for volume; closed-only where a '
         'panel measures mature realized price (T5).'),
        ('Non-mutual-exclusivity =', 'a claim carrying multiple codes is counted '
         'once per code it carries; grand totals may exceed the distinct-claim '
         'count. The multi-code overlap is quantified from the raw file '
         '(PENDING) and printed here.'),
    ]:
        sb.row([(a, 'label'), (b, 'text')], wrap=True, height=40)
    sb.row([('Multi-code overlap (raw file)', 'label'), ('PENDING', 'note')],
           height=20)
    sb.note('Every panel in this workbook already follows this rule; it is '
            'stated here as the single import for the sizing model.')
    return ws


# ================================================ packaged outputs ========
def build_grossup(wb):
    S = 'GrossUp_Factors'
    ws, sb = _panel(wb, S, 3, [44, 22, 60], 'FF13343B',
                    'Gross-up factors (packaged from T2 for the sizing team)',
                    'The Medicare capture rate turned into the gross-up that '
                    'converts Komodo-captured volume to the universe, with every '
                    'assumption listed. Handle as a bounded estimate, not a '
                    'point.')
    sb.headers(['Factor', 'Value', 'Assumption / caveat'])
    sb.row([('Implied Komodo capture of the full Medicare book', 'label'),
            (f"={REF['mcare_capture']}", 'fml', FMT_PCT1),
            ('From T2. Denominator uses the MA upper bound, so this capture is a '
             'FLOOR.', 'note')], wrap=True, height=30)
    sb.row([('Candidate gross-up = 1 / capture', 'label'),
            (f"=1/{REF['mcare_capture']}", 'fml', FMT_X),
            ('Multiply Komodo-captured volume by this to approach the universe; '
             'a CEILING because capture is a floor.', 'note')], wrap=True,
           height=30)
    sb.blank()
    sb.banner('Assumptions (all of them)')
    for a in [
        'Capture is PAYER-UNIFORM: the Medicare capture rate is applied to the '
        'Commercial and Medicaid columns. This is the weakest link - Komodo '
        'commercial capture may differ materially.',
        'MA per-capita utilization = FFS in the MA bound; this overstates MA '
        'volume, making the capture a floor and the gross-up a ceiling.',
        'Vintage mismatch: Komodo is 2025 (mostly open claims, T5); benchmarks '
        'are 2023-2024. The completion factor (T5) will lift 2025 volumes.',
        '"Medicare" in Komodo mixes FFS and MA and (pending T7) possibly '
        'institutional lines.',
        'Volume that NO claims dataset contains - the ~1-in-5 did-not-always-'
        'bill and the 18.6 percent facility-pay layer - sits outside every '
        'gross-up here.',
    ]:
        sb.row([(a, 'text'), None, None], wrap=True, height=30)
    return ws


def build_commercial_table(wb):
    S = 'Commercial_Rate_Table'
    ws, sb = _panel(wb, S, 5, [18, 20, 18, 18, 46], 'FF3B1330',
                    'Commercial rate table (packaged from T3)',
                    'The measured commercial realized price per code and its '
                    'Medicare-multiples - the replacement for the interim FAIR '
                    'bound and statutory pegs. Combined dollars; the T4 '
                    'estimated share travels with each figure.')
    sb.headers(['HCPCS', 'Comm $/claim (combined)', 'x fee schedule',
                'x realized', 'Note'])
    for code in CODES:
        has = PERCODE[code][2][0] is not None
        cc = REF[f'{code}_comm_claims']
        ccb = REF[f'{code}_comm_comb']
        if has:
            a = (f"={ccb}/{cc}", 'fml', FMT_USD2)
            b = (f"={ccb}/{cc}/{REF[f'fs_{code}']}", 'fml', FMT_X)
            c = (f"={ccb}/{cc}/{REF[f'real_{code}']}", 'fml', FMT_X)
            nt = 'Carries the T4 estimated-allowed share.'
        else:
            a = b = c = ('PENDING', 'note')
            nt = 'Commercial split from the raw file.'
        sb.row([(f'{code} {CODE_DESC[code]}', 'label'), a, b, c, (nt, 'note')],
               wrap=True, height=26)
    sb.note('Supersedes: the FAIR-derived ~1.6x commercial bound and the '
            'statutory balance-billing pegs on the master Balance_Billing_'
            'States. Those remain in the master untouched; this is the '
            'engagement-side measured replacement.')
    return ws


def build_memo(wb):
    S = 'Calibration_Memo'
    ws, sb = _panel(wb, S, 1, [118], 'FF00294C',
                    'Calibration memo - one page for Ray',
                    'What the Komodo extract is good for, and the three things '
                    'it cannot see.')
    for head, body in MEMO_SECTIONS:
        sb.banner(head)
        sb.prose(body, kind='text')
    return ws


MEMO_SECTIONS = [
    ('What the extract is',
     'Komodo Health ambulance claims, calendar 2025, seven codes, split '
     'Commercial / Medicare / Medicaid, with Open / Closed / Combined status and '
     'four dollar columns (Total, Reported, Estimated, Combined). ~47.1M claims '
     'and ~$14.3B combined allowed as transcribed; T1 re-derives from the raw '
     'file. Three caveats travel with every number: <1 percent of claims '
     'excluded for missing line-of-business/state, codes not mutually exclusive '
     'to status, and multiple codes per claim.'),
    ('The capture / gross-up (T2)',
     'Komodo Medicare base-code volume (~13.9M claims) against the reconstructed '
     'Medicare book implies a capture of roughly half the full Medicare universe '
     '- a candidate gross-up near 2x. It is a bounded estimate: the capture is a '
     'floor (MA bound is an upper bound) and it assumes capture is payer-uniform '
     '(the weakest link). Use it to scale, not to pinpoint.'),
    ('The commercial multiples (T3)',
     'Measured commercial realized price runs well above Medicare: A0428 ~$533 '
     'per claim (~2x fee schedule), A0434 ~$2,208 (~2.4x realized). This '
     'replaces the interim FAIR ~1.6x bound with a per-code, measured multiple. '
     'The Reported-only column (from the raw file) is the conservative version.'),
    ('The estimated-allowed disclosure (T4)',
     'About a third of all dollars are Komodo-MODELED, not payer-reported. Any '
     'sizing built on Combined dollars must carry this share; the Reported-only '
     'figures are the floor.'),
    ('The institutional verdict (T7)',
     'INDETERMINATE until the raw file: metadata (claim form / bill type / '
     'revenue center) decides whether institutional UB-04 lines are in. If '
     'EXCLUDED, add the 581,532 Medicare institutional wedge separately and cut '
     'the second extract (rev centers 0540-0549).'),
    ('The three things it cannot see',
     '(1) Unbilled transports - the ~1-in-5 did-not-always-bill incidence. '
     '(2) Facility-pay revenue - 18.6 percent of organizations take it, and no '
     'claims dataset contains it. (3) Bundled transports - Part A inpatient-'
     'bundled and SNF consolidated-billing rides that never surface as a claim. '
     'No gross-up recovers these; they are named, not modeled.'),
]


def build_runlog(wb):
    S = 'Run_Log'
    ws, sb = _panel(wb, S, 2, [34, 86], 'FF444444',
                    'Run log and firewall attestation',
                    'Build metadata, the raw-file hash slot, and the master-'
                    'untouched leak check.')
    sb.banner('Build')
    for a, b in [
        ('Workbook', 'Komodo_Calibration_v1.xlsx'),
        ('Built', BUILT),
        ('Benchmark source', MASTER + ' (values imported one-way; fact IDs on '
         'Benchmarks_Public)'),
        ('Extract vintage', 'Komodo Health ambulance claims, calendar 2025'),
        ('State', 'SKELETON - benchmarks and all test panels live; awaiting the '
         'raw extract from Andrew to re-derive Observed_Inputs (T1).'),
    ]:
        sb.row([(a, 'label'), (b, 'text')], wrap=True, height=24)
    sb.blank()
    sb.banner('Raw extract')
    sb.row([('Raw file delivered', 'label'), ('PENDING (Andrew)', 'note')],
           height=20)
    sb.row([('Raw file SHA-256', 'label'), ('PENDING', 'note')], height=20)
    sb.blank()
    sb.banner('Firewall leak check (master must be untouched)')
    sb.row([('Zero Komodo-derived values in the master', 'label'),
            ('CONFIRMED at build - see leak_check_master.json; the extract '
             'signature figures (47,061,988 claims; $14,320,080,327 combined; '
             'the payer and code totals) appear nowhere in '
             + MASTER + ' or its repo.', 'note')], wrap=True, height=40)
    sb.row([('Direction', 'label'),
            ('One-way: master -> calibration. This workbook is never cited by, '
             'linked from, or merged into the master.', 'note')], wrap=True,
           height=28)
    return ws


# ================================================================ main ====
def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_benchmarks(wb)
    build_observed(wb)
    build_t2(wb)
    build_t3(wb)
    build_t4(wb)
    build_t5(wb)
    build_t6(wb)
    build_t7(wb)
    build_t8(wb)
    build_t9(wb)
    build_t10(wb)
    build_grossup(wb)
    build_commercial_table(wb)
    build_memo(wb)
    build_runlog(wb)
    # house rule: no em/en dashes in cell text
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and ('—' in c.value or
                                                 '–' in c.value):
                    c.value = c.value.replace('—', ' - ').replace(
                        '–', '-')
    wb.save(OUT)
    print(f'wrote {OUT} with {len(wb.sheetnames)} tabs: {wb.sheetnames}')


if __name__ == '__main__':
    main()
